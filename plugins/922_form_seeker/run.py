"""
922 FormSeeker
==============
Discovers formed plate PDFs in a 922 QTDR batch using three methods:
  1. Filename scan for "PLT F"
  2. PO workbook NOTES column containing "bend"
  3. PDF content / spatial analysis for unlabeled formed plates

Copies all discovered PDFs into Batch {n} - Documentation/Forming {n}/, merges
them into Forming {n}.pdf, and writes one row per part to the Bent Plates
sheet of the Pallet & Rod Organizer workbook (sorted by SOURCE MATERIAL).
"""
from __future__ import annotations

import os
import re
import shutil
import threading
from pathlib import Path
from typing import Optional

import fitz
import openpyxl
from openpyxl.styles import Font

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk


# Regex patterns ---------------------------------------------------------------

_UP_DEG_R_RE = re.compile(r'UP\s*\d+(?:\.\d+)?\s*°\s*R', re.IGNORECASE)
_DEG_WORD_RE = re.compile(r'\d+\s*°')
_NUMERIC_WORD_RE = re.compile(r'^\d+\.?\d*$')

PURPLE_HEX = "FF7030A0"


# DYPN helpers -----------------------------------------------------------------

def _dypn_from_filename(filename: str) -> str:
    """Extract DYPN from filename: the portion before ' PLT'."""
    idx = filename.find(" PLT")
    if idx == -1:
        return Path(filename).stem
    return filename[:idx].strip()


def _bent_suffix(dypn: str) -> str:
    """E.g. 'E6492162-H22-3' -> 'H22-3 BEND' (suffix is everything from first -H)."""
    idx = dypn.find("-H")
    if idx == -1:
        return f"{dypn} BEND"
    return f"{dypn[idx + 1:]} BEND"


def _is_skipped(pdf: Path, batch_path: Path, batch_no: str) -> bool:
    try:
        rel = pdf.relative_to(batch_path)
    except ValueError:
        return False
    parts_lower = [seg.casefold() for seg in rel.parts]
    if "repeat batches" in parts_lower:
        return True
    doc_name = f"batch {batch_no} - documentation".casefold()
    if doc_name in parts_lower:
        return True
    return False


# PO workbook ------------------------------------------------------------------

def _find_po_workbook(doc_folder: Path, batch_no: str) -> Optional[Path]:
    matches = [
        p for p in doc_folder.glob(f"PO H{batch_no} QF-QU-09 REV C*.xlsx")
        if not p.name.startswith("~")
    ]
    return sorted(matches)[0] if matches else None


def _find_organizer_workbook(doc_folder: Path, batch_no: str) -> Optional[Path]:
    matches = [
        p for p in doc_folder.glob(f"PO H{batch_no} Pallet & Rod Organizer*.xlsx")
        if not p.name.startswith("~")
    ]
    return sorted(matches)[0] if matches else None


def _load_po_lookup(po_path: Path, log) -> dict:
    """Return {dypn_casefold: row metadata}. Bend rows preferred over non-bend
    when the same DYPN appears on multiple rows."""
    wb = openpyxl.load_workbook(po_path, data_only=True)
    if 'PO' not in wb.sheetnames:
        log("PO workbook: 'PO' sheet not found.")
        wb.close()
        return {}
    ws = wb['PO']

    # Locate header row by scanning for cells named ORDER + DYPN
    header_row = None
    headers: dict[str, int] = {}
    for r in range(1, min(ws.max_row, 30) + 1):
        row_map: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                row_map[v.strip().upper()] = c
        if 'DYPN' in row_map and 'ORDER' in row_map:
            header_row = r
            headers = row_map
            break

    if header_row is None:
        log("PO workbook: could not locate header row.")
        wb.close()
        return {}

    required = ['ORDER', 'PPN', 'DYPN', 'NOTES', 'SOURCE MATERIAL']
    missing = [h for h in required if h not in headers]
    if missing:
        log(f"PO workbook: missing header columns: {missing}")

    def cell_val(r: int, name: str):
        col = headers.get(name)
        if col is None:
            return None
        return ws.cell(row=r, column=col).value

    bend_keys: set[str] = set()
    lookup: dict[str, dict] = {}

    # Pass 1: bend rows take priority
    for r in range(header_row + 1, ws.max_row + 1):
        dypn = cell_val(r, 'DYPN')
        if not dypn:
            continue
        notes = cell_val(r, 'NOTES')
        if not notes or 'bend' not in str(notes).lower():
            continue
        key = str(dypn).strip().casefold()
        if key in lookup:
            continue
        lookup[key] = {
            'ORDER': cell_val(r, 'ORDER'),
            'PPN': cell_val(r, 'PPN'),
            'DYPN': str(dypn).strip(),
            'NOTES': notes,
            'SOURCE MATERIAL': cell_val(r, 'SOURCE MATERIAL'),
        }
        bend_keys.add(key)

    # Pass 2: fill in non-bend rows for DYPNs not already covered
    for r in range(header_row + 1, ws.max_row + 1):
        dypn = cell_val(r, 'DYPN')
        if not dypn:
            continue
        key = str(dypn).strip().casefold()
        if key in lookup:
            continue
        lookup[key] = {
            'ORDER': cell_val(r, 'ORDER'),
            'PPN': cell_val(r, 'PPN'),
            'DYPN': str(dypn).strip(),
            'NOTES': cell_val(r, 'NOTES'),
            'SOURCE MATERIAL': cell_val(r, 'SOURCE MATERIAL'),
        }

    wb.close()
    return lookup


# Method 1 ---------------------------------------------------------------------

def _method1(batch_path: Path, batch_no: str) -> dict[str, dict]:
    """Recursively scan for PDFs with 'PLT F' in the filename (case-sensitive)."""
    found: dict[str, dict] = {}
    for pdf in batch_path.rglob("*.pdf"):
        if "PLT F" not in pdf.name:
            continue
        if _is_skipped(pdf, batch_path, batch_no):
            continue
        dypn = _dypn_from_filename(pdf.name)
        if not dypn:
            continue
        key = dypn.casefold()
        if key not in found:
            found[key] = {'dypn': dypn, 'pdf_path': pdf, 'methods': {1}}
    return found


# Method 2 ---------------------------------------------------------------------

def _resolve_method2_pdf(batch_path: Path, order: str, dypn: str) -> Optional[Path]:
    """Find a PDF for DYPN under {batch_root}/{order folder}/CAD-AND-SHOP-PRINTS/.
    Prefer 'PLT F' over 'PLT' if both exist."""
    order_folder = None
    for child in batch_path.iterdir():
        if not child.is_dir():
            continue
        if child.name == order or child.name.startswith(f"{order}-"):
            order_folder = child
            break
    if not order_folder:
        return None

    # CAD-AND-SHOP-PRINTS case-insensitive
    cad_root = None
    for child in order_folder.iterdir():
        if child.is_dir() and child.name.lower() == "cad-and-shop-prints":
            cad_root = child
            break
    if not cad_root:
        return None

    plt_f_match: Optional[Path] = None
    plt_match: Optional[Path] = None
    needle = f"{dypn} "
    for pdf in cad_root.rglob("*.pdf"):
        name = pdf.name
        if needle not in name or "PLT" not in name:
            continue
        if "PLT F" in name:
            if plt_f_match is None:
                plt_f_match = pdf
        else:
            if plt_match is None:
                plt_match = pdf
    return plt_f_match or plt_match


def _method2(batch_path: Path, po_lookup: dict, log) -> dict[str, dict]:
    """Find DYPNs where PO NOTES contain 'bend'; resolve each to a PDF on disk."""
    found: dict[str, dict] = {}
    for key, meta in po_lookup.items():
        notes = meta.get('NOTES')
        if not notes or 'bend' not in str(notes).lower():
            continue
        order = meta.get('ORDER')
        dypn = meta['DYPN']
        if not order:
            log(f"  Method 2: row for {dypn} has no ORDER value; cannot resolve PDF")
            continue
        pdf = _resolve_method2_pdf(batch_path, str(order).strip(), dypn)
        if not pdf:
            log(f"  Method 2: PDF not found on disk for {dypn} (order {order})")
            continue
        found[key] = {'dypn': dypn, 'pdf_path': pdf, 'methods': {2}}
    return found


# Method 3 ---------------------------------------------------------------------

def _is_formed_pdf(pdf: Path) -> bool:
    """Open PDF; return True if Check A (UP X deg R) or Check B (spatial) passes."""
    try:
        doc = fitz.open(pdf)
    except Exception:
        return False
    try:
        page = doc[0]
        text = page.get_text("text") or ""

        if _UP_DEG_R_RE.search(text):
            return True

        words = page.get_text("words") or []
        tb_top: Optional[float] = None
        for w in words:
            if 'WELDMENT' in w[4].upper() and 'FAB' in w[4].upper():
                tb_top = w[1]
                break
        if tb_top is None:
            for w in words:
                if 'WELDMENT' in w[4].upper():
                    tb_top = w[1]
                    break
        if tb_top is None:
            return False

        x_min = page.rect.width / 2.0
        x_max = page.rect.width
        y_min = tb_top - (tb_top * 0.45)
        y_max = tb_top
        x_span = x_max - x_min
        bevel_cutoff = x_min + x_span * 0.15
        gap_cutoff = page.rect.width * 0.15

        def _in_region(w):
            cx = (w[0] + w[2]) / 2.0
            cy = (w[1] + w[3]) / 2.0
            return x_min <= cx <= x_max and y_min <= cy <= y_max

        sv_words = [w for w in words if _in_region(w)]
        deg_words = [w for w in sv_words if _DEG_WORD_RE.search(w[4])]

        for dw in deg_words:
            if dw[0] < bevel_cutoff:
                continue
            dw_yc = (dw[1] + dw[3]) / 2.0
            for nw in sv_words:
                if not _NUMERIC_WORD_RE.fullmatch(nw[4]):
                    continue
                if nw[2] >= dw[0]:
                    continue
                if dw[0] - nw[2] >= gap_cutoff:
                    continue
                nw_yc = (nw[1] + nw[3]) / 2.0
                if abs(dw_yc - nw_yc) > 8:
                    continue
                return True
        return False
    finally:
        doc.close()


def _method3(
    batch_path: Path,
    batch_no: str,
    already_found: set[str],
    log,
    progress_callback,
    cancel_event: threading.Event,
    progress_start: int,
    progress_end: int,
) -> dict[str, dict]:
    found: dict[str, dict] = {}
    candidates: list[Path] = []
    for pdf in batch_path.rglob("*.pdf"):
        if "PLT" not in pdf.name:
            continue
        if "PLT F" in pdf.name:
            continue
        if _is_skipped(pdf, batch_path, batch_no):
            continue
        candidates.append(pdf)

    log(f"  Method 3 candidate PDFs: {len(candidates)}")
    total = max(1, len(candidates))
    for i, pdf in enumerate(candidates):
        if cancel_event.is_set():
            return found
        progress_callback(progress_start + int((i / total) * (progress_end - progress_start)))
        dypn = _dypn_from_filename(pdf.name)
        if not dypn:
            continue
        key = dypn.casefold()
        if key in already_found:
            continue
        if _is_formed_pdf(pdf):
            found[key] = {'dypn': dypn, 'pdf_path': pdf, 'methods': {3}}
            log(f"  Method 3: formed plate detected -- {dypn}")
    return found


# Phase 2 ----------------------------------------------------------------------

def _update_bent_plates(
    organizer_path: Path,
    batch_no: str,
    rows_data: list[dict],
    log,
) -> None:
    wb = openpyxl.load_workbook(organizer_path)
    if 'Bent Plates' not in wb.sheetnames:
        wb.close()
        raise RuntimeError("'Bent Plates' sheet not found in organizer workbook")
    ws = wb['Bent Plates']

    header_row = 2
    headers: dict[str, int] = {}
    scan_cols = max(ws.max_column, 8)
    for c in range(1, scan_cols + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str):
            headers[v.strip().upper()] = c

    required = ['ORDER', 'PPN', 'DYPN', 'SOURCE MATERIAL', 'NOTES']
    missing = [h for h in required if h not in headers]
    if missing:
        wb.close()
        raise RuntimeError(f"Bent Plates header row missing columns: {missing}")

    # Clear existing data rows from row 3 down across all known header columns
    last_col = max(headers.values())
    clear_to = max(ws.max_row, 3)
    for r in range(3, clear_to + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.value = None

    purple_font = Font(color=PURPLE_HEX)
    for i, row in enumerate(rows_data):
        r = 3 + i
        ws.cell(row=r, column=headers['ORDER'], value=row.get('ORDER'))
        ws.cell(row=r, column=headers['PPN'], value=row.get('PPN'))
        ws.cell(row=r, column=headers['DYPN'], value=row.get('DYPN'))
        ws.cell(row=r, column=headers['SOURCE MATERIAL'], value=row.get('SOURCE MATERIAL'))
        notes_cell = ws.cell(row=r, column=headers['NOTES'], value=row.get('NOTES'))
        notes_cell.font = purple_font

    # Row 1 title (merged cell; write to top-left)
    ws.cell(row=1, column=1, value=f"BATCH {batch_no}: FORMED PLATES")

    wb.save(organizer_path)
    wb.close()


# Plugin entry -----------------------------------------------------------------

def run(params: dict, progress_callback, cancel_event: threading.Event) -> None:
    log = params.get('log', print)
    console = params.get('console')

    log("Starting 922 FormSeeker...")
    progress_callback(0)

    if console and hasattr(console, 'request_input'):
        raw = console.request_input(
            'Enter batch number (e.g. "473", "Batch 473", "PO #473"):'
        )
    else:
        raw = input('Enter batch number: ')

    batch_no = sdk.parse_922_batch(raw or '')
    if not batch_no:
        raise ValueError(f"Unrecognised batch input: {raw!r}")
    log(f"Batch: {batch_no}")

    root = sdk.resolve_922_root()
    if not root:
        raise RuntimeError(
            "Could not locate '922 QTDR Production Packages'. Verify OneDrive sync."
        )
    batch_path = sdk.find_922_batch_path(root, batch_no)
    if not batch_path:
        raise RuntimeError(
            f"Batch {batch_no} not found under {root} (also checked '1 - Completed')."
        )
    log(f"Batch path: {batch_path}")
    doc_folder = batch_path / f"Batch {batch_no} - Documentation"
    progress_callback(5)

    # Existing-forming check: prompt before doing any work.
    forming_dir = doc_folder / f"Forming {batch_no}"
    if forming_dir.exists() and any(forming_dir.iterdir()):
        prompt = "Looks like this batch already has forming. Overwrite existing? Y/N"
        if console and hasattr(console, 'request_input'):
            ans = console.request_input(prompt)
        else:
            ans = input(prompt + ": ")
        if (ans or "").strip().lower() not in {"y", "yes"}:
            log("Aborted -- existing forming preserved.")
            return
        log("Overwriting existing forming.")

    # PO workbook lookup (drives Method 2 and Bent Plates metadata)
    po_path = _find_po_workbook(doc_folder, batch_no)
    if po_path:
        log(f"PO workbook: {po_path.name}")
        po_lookup = _load_po_lookup(po_path, log)
        log(f"  PO rows indexed: {len(po_lookup)}")
    else:
        log("WARNING: PO workbook (QF-QU-09) not found. Skipping Method 2 and PO metadata lookup.")
        po_lookup = {}
    progress_callback(10)
    if cancel_event.is_set():
        log("Cancelled.")
        return

    # Method 1
    log("Method 1: scanning filenames for 'PLT F'...")
    m1 = _method1(batch_path, batch_no)
    log(f"  Method 1 found: {len(m1)} DYPN(s)")
    progress_callback(25)
    if cancel_event.is_set():
        log("Cancelled.")
        return

    # Method 2
    if po_lookup:
        log("Method 2: scanning PO NOTES for 'bend'...")
        m2 = _method2(batch_path, po_lookup, log)
        log(f"  Method 2 found: {len(m2)} DYPN(s)")
    else:
        m2 = {}
    progress_callback(40)
    if cancel_event.is_set():
        log("Cancelled.")
        return

    # Combine Methods 1 and 2 results before running Method 3 (which skips duplicates)
    combined: dict[str, dict] = {}
    for src in (m1, m2):
        for key, entry in src.items():
            if key in combined:
                combined[key]['methods'].update(entry['methods'])
            else:
                combined[key] = {
                    'dypn': entry['dypn'],
                    'pdf_path': entry['pdf_path'],
                    'methods': set(entry['methods']),
                }

    # Method 3
    log("Method 3: scanning PDFs for formed-plate visual signature...")
    m3 = _method3(
        batch_path, batch_no, set(combined.keys()),
        log, progress_callback, cancel_event,
        progress_start=40, progress_end=70,
    )
    log(f"  Method 3 found: {len(m3)} additional DYPN(s)")
    for key, entry in m3.items():
        combined[key] = entry
    progress_callback(70)
    if cancel_event.is_set():
        log("Cancelled.")
        return

    if not combined:
        log("No formed plates discovered by any method. Nothing to write.")
        progress_callback(100)
        return

    log(f"Total unique formed plates: {len(combined)}")

    # Phase 2: copy PDFs into Forming subfolder
    forming_dir.mkdir(parents=True, exist_ok=True)
    log(f"Output folder: {forming_dir}")

    sorted_for_copy = sorted(combined.values(), key=lambda e: e['dypn'].casefold())
    copied_paths: list[Path] = []
    for f in sorted_for_copy:
        if cancel_event.is_set():
            log("Cancelled.")
            return
        src = f['pdf_path']
        dest = forming_dir / src.name
        shutil.copy2(src, dest)
        methods_str = ",".join(str(m) for m in sorted(f['methods']))
        log(f"  copied {dest.name}  (methods: {methods_str})")
        copied_paths.append(dest)
    progress_callback(85)

    # Merge
    merged_path = forming_dir / f"Forming {batch_no}.pdf"
    log(f"Merging {len(copied_paths)} PDF(s) into {merged_path.name}")
    sdk.merge_pdfs(copied_paths, merged_path)
    progress_callback(90)

    # Build Bent Plates row data (sorted by SOURCE MATERIAL)
    rows_data: list[dict] = []
    for f in sorted_for_copy:
        meta = po_lookup.get(f['dypn'].casefold(), {})
        rows_data.append({
            'ORDER': meta.get('ORDER'),
            'PPN': meta.get('PPN'),
            'DYPN': f['dypn'],
            'SOURCE MATERIAL': meta.get('SOURCE MATERIAL'),
            'NOTES': _bent_suffix(f['dypn']),
        })

    def _sort_key(row):
        sm = row.get('SOURCE MATERIAL')
        if sm is None or sm == '':
            return (1, '')
        return (0, str(sm))
    rows_data.sort(key=_sort_key)

    organizer_path = _find_organizer_workbook(doc_folder, batch_no)
    if not organizer_path:
        raise RuntimeError(
            f"Pallet & Rod Organizer workbook not found in {doc_folder}. "
            "Halted Phase 2 write."
        )
    log(f"Updating Bent Plates in: {organizer_path.name}")
    _update_bent_plates(organizer_path, batch_no, rows_data, log)
    progress_callback(100)

    log("=" * 50)
    log(f"922 FormSeeker -- Batch {batch_no}")
    log(f"  Total formed plates : {len(rows_data)}")
    log("=" * 50)


# Standalone test harness ------------------------------------------------------
if __name__ == "__main__":
    cancel = threading.Event()
    run(
        params={'log': print, 'settings': {}, 'console': None},
        progress_callback=lambda p: print(f"[{p}%]"),
        cancel_event=cancel,
    )
