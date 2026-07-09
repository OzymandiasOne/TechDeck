"""
911 Linear Inch CutTime plugin for TechDeck.

Closed-loop cut-length + cut-time report for a 911 QTDR batch.

The user picks a batch's ``IGES FILES`` folder (or the batch folder). Each IGES subfolder
is named by the nest package number (``N<nest>S`` -> nest ``<nest>``) and holds that
nest's part DXFs. Cut length per part is computed EXACTLY from the DXF vector geometry;
the per-part quantity is read from the nest package PDF's "SUMMARY OF NEST" table
(WORK ORDER -> qty), matched to the DXF by the work order embedded in its filename.

Cut time = total linear inches / feed rate (IPM), where the feed rate is looked up by the
part's thickness (pulled from its PART SKETCH page) in a single feed-rate chart. A plate
whose thickness falls BETWEEN two listed thicknesses rounds UP to the next (thicker) row's
feed rate - the conservative choice (thicker = slower = more estimated time), so oddball
outlier thicknesses stay protected without editing the chart.

Output: one Excel report - per-part rows, per-nest subtotals, batch total.

Worker-thread plugin: uses the SDK console helpers (native folder picker + modal warning).

The DXF parser is vendored from the ``customer_dxf_quoting`` plugin (stdlib only, ASCII
DXF, LINE/ARC/CIRCLE/LWPOLYLINE/POLYLINE + bulge). It should later be centralized into
``plugin_sdk`` so both plugins share one implementation.
"""

import os
import re
import glob
import math
from pathlib import Path

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Work order token in DXF filenames + the SUMMARY OF NEST table. Two SigmaNest
# formats seen: current "X4549775"/"BK514922" (1-2 letters + 5-8 digits, GX029+)
# and legacy "3X24-398" (digit-letter-digits-dash-digits, batch 3X010). Matches the
# work-order form 911_setup's summary parser uses (`^[A-Z]{1,2}\d{5,8}$`).
WO_RE = re.compile(r"(?:[A-Z]{1,2}\d{5,8}|\d[A-Z]\d+-\d+)", re.I)
# Part number label on a PART SKETCH page, e.g. "PART: R4138947-8M".
PART_RE = re.compile(r"PART:\s*([A-Z0-9][A-Z0-9./\-]{3,})", re.I)
# Part thickness on a PART SKETCH page, e.g. "0.250 THK" or "0.500 THICK".
THK_RE = re.compile(r"([\d.]+)\s*(?:THK|THICK)", re.I)
# Layers that are reference-only and never count as cut (matches customer_dxf_quoting).
EXCLUDE_LAYERS = {"BOUNDING_BOX", "IGNORE"}
# Entity types we do NOT measure; their presence means the length may be under-counted.
UNSUPPORTED_FLAGGED = {"SPLINE", "ELLIPSE"}

# ---------------------------------------------------------------------------
# Feed rates (cut speed, inches/min) by thickness - cut-time estimate
# ---------------------------------------------------------------------------
# One combined chart (no more separate Plasma vs Oxy sets). Speed is the DIVISOR:
#   cut minutes = total linear inches / speed (IPM).
# Rows are ASCENDING by thickness - required by the round-up lookup below.
# TO UPDATE: replace the (thickness_in, speed_ipm) pairs from the reference sheet's
# THICKNESS + FEEDRATE (IPM) columns, keeping them sorted thin -> thick.
FEED_RATES_SOURCE = "Plasma & Oxy Feed Rates (single combined chart, rev 2026-07-09)"

FEED_RATES = [  # (thickness in, speed IPM) - ascending
    (0.125, 180), (0.188, 140), (0.25, 100), (0.313, 80), (0.375, 100),
    (0.438, 90), (0.5, 75), (0.625, 55), (0.75, 80), (0.875, 70),
    (1.0, 50), (1.125, 45), (1.25, 40), (1.375, 35), (1.5, 30),
    (1.625, 45), (1.75, 40), (2.0, 30), (2.25, 25), (2.5, 20),
    (2.75, 9.25), (3.0, 9.25), (3.25, 9.0), (3.5, 8.5), (3.75, 8.5),
    (4.0, 8.5), (4.25, 8.0), (4.5, 8.0), (5.0, 7.25), (5.25, 6.5),
    (5.5, 6.0), (5.75, 6.0), (6.0, 6.0), (6.25, 6.0), (6.5, 6.0),
    (6.75, 6.0), (7.0, 6.0),
]


def lookup_feed(thickness):
    """(ipm, note) for a thickness, rounding UP to the next listed thickness.

    A plate whose thickness sits between two rows uses the next-THICKER row's feed
    rate (thicker = slower = more time = conservative). Below the thinnest row it
    clamps to the thinnest; above the thickest it clamps to the thickest. ``note`` is
    None on an exact hit, else a short string describing the round-up (for flagging).
    Returns None only when ``thickness`` is None.
    """
    if thickness is None:
        return None
    for t, ipm in FEED_RATES:
        if thickness <= t + 1e-9:
            note = None if abs(t - thickness) < 1e-6 else f"rounded up to {t}\" row"
            return ipm, note
    t, ipm = FEED_RATES[-1]  # thicker than the whole chart - clamp to the last row
    return ipm, f"above chart max {t}\" - used {t}\" row"


def fmt_hm(minutes):
    """Minutes -> 'Hh MMm' (blank for None)."""
    if minutes is None:
        return ""
    total = int(round(minutes))
    return f"{total // 60}h {total % 60:02d}m"


def part_cut_times(total_inches, thickness):
    """Feed rate + cut minutes for a part's total inches (single combined chart).
    ipm/min are None only when the thickness is unknown; ``note`` describes any
    round-up so callers can flag it."""
    out = {"ipm": None, "min": None, "note": None}
    r = lookup_feed(thickness)
    if r:
        out["ipm"], out["note"] = r
        out["min"] = total_inches / r[0] if r[0] else None
    return out


# ---------------------------------------------------------------------------
# DXF parsing (vendored from customer_dxf_quoting; TODO: centralize in plugin_sdk)
# ---------------------------------------------------------------------------

def _decode_dxf(path):
    data = Path(path).read_bytes()
    if data.startswith(b"AutoCAD Binary DXF"):
        raise ValueError("Binary DXF is not supported - re-export as ASCII DXF.")
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not decode file - is this a DXF?")


def _read_pairs(path):
    lines = _decode_dxf(path).splitlines()
    pairs = []
    it = iter(lines)
    for code_line in it:
        try:
            value = next(it)
        except StopIteration:
            break
        try:
            code = int(code_line.strip())
        except ValueError:
            raise ValueError(f"Malformed DXF (expected numeric group code, got {code_line.strip()!r}).")
        pairs.append((code, value.strip()))
    if not pairs:
        raise ValueError("File is empty.")
    return pairs


def _bulge_length(p1, p2, bulge):
    chord = math.hypot(p2[0] - p1[0], p2[1] - p1[1])
    if abs(bulge) < 1e-12 or chord < 1e-12:
        return chord
    theta = 4.0 * math.atan(bulge)
    radius = chord * (1.0 + bulge * bulge) / (4.0 * abs(bulge))
    return radius * abs(theta)


def _collect_until_next_entity(pairs, j):
    groups = []
    n = len(pairs)
    while j < n and pairs[j][0] != 0:
        groups.append(pairs[j])
        j += 1
    return groups, j


def _build_line(groups):
    g = {}
    for code, val in groups:
        if code in (8, 10, 20, 11, 21):
            g[code] = val
    try:
        p1 = (float(g[10]), float(g[20]))
        p2 = (float(g[11]), float(g[21]))
    except KeyError:
        return None
    return {"layer": g.get(8, "0"), "points": [p1, p2], "length": math.dist(p1, p2)}


def _build_arc(groups, full_circle=False):
    g = {}
    for code, val in groups:
        if code in (8, 10, 20, 40, 50, 51):
            g[code] = val
    try:
        cx, cy, r = float(g[10]), float(g[20]), float(g[40])
    except KeyError:
        return None
    if full_circle:
        a0, sweep = 0.0, math.tau
    else:
        a0 = math.radians(float(g.get(50, 0.0)))
        a1 = math.radians(float(g.get(51, 360.0)))
        sweep = (a1 - a0) % math.tau or math.tau
    n = max(8, int(abs(sweep) / math.tau * 64) + 1)
    pts = [(cx + r * math.cos(a0 + sweep * t / n), cy + r * math.sin(a0 + sweep * t / n))
           for t in range(n + 1)]
    return {"layer": g.get(8, "0"), "points": pts, "length": r * sweep}


def _polyline_entity(layer, verts, closed):
    verts = [v for v in verts if v[1] is not None]
    if len(verts) < 2:
        return None
    total = 0.0
    seg_pairs = list(zip(verts, verts[1:]))
    if closed:
        seg_pairs.append((verts[-1], verts[0]))
    for v1, v2 in seg_pairs:
        total += _bulge_length((v1[0], v1[1]), (v2[0], v2[1]), v1[2])
    points = [(v[0], v[1]) for v in verts]
    return {"layer": layer, "points": points, "length": total}


def _build_lwpolyline(groups):
    layer, flags = "0", 0
    verts = []  # [x, y, bulge]
    for code, val in groups:
        if code == 8:
            layer = val
        elif code == 70:
            flags = int(val)
        elif code == 10:
            verts.append([float(val), None, 0.0])
        elif code == 20 and verts:
            verts[-1][1] = float(val)
        elif code == 42 and verts:
            verts[-1][2] = float(val)
    return _polyline_entity(layer, verts, closed=bool(flags & 1))


def _build_polyline(pairs, j):
    groups, j = _collect_until_next_entity(pairs, j)
    layer, flags = "0", 0
    for code, val in groups:
        if code == 8:
            layer = val
        elif code == 70:
            flags = int(val)
    verts = []
    n = len(pairs)
    while j < n:
        code, val = pairs[j]
        if code == 0 and val == "VERTEX":
            vgroups, j = _collect_until_next_entity(pairs, j + 1)
            v = [None, None, 0.0]
            vflags = 0
            for c, vv in vgroups:
                if c == 10:
                    v[0] = float(vv)
                elif c == 20:
                    v[1] = float(vv)
                elif c == 42:
                    v[2] = float(vv)
                elif c == 70:
                    vflags = int(vv)
            if v[0] is not None and v[1] is not None and not (vflags & 16):
                verts.append(v)
        elif code == 0 and val == "SEQEND":
            _, j = _collect_until_next_entity(pairs, j + 1)
            break
        else:
            j += 1
    return _polyline_entity(layer, verts, closed=bool(flags & 1)), j


def parse_dxf(path):
    """Return (entities, skipped, insunits). entity = {layer, points, length}."""
    pairs = _read_pairs(path)
    n = len(pairs)
    entities, skipped = [], {}
    insunits = None
    section = None
    i = 0
    while i < n:
        code, val = pairs[i]
        if code == 0 and val == "SECTION" and i + 1 < n and pairs[i + 1][0] == 2:
            section = pairs[i + 1][1]
            i += 2
            continue
        if code == 0 and val == "ENDSEC":
            section = None
            i += 1
            continue
        if section == "HEADER" and code == 9 and val == "$INSUNITS":
            if i + 1 < n and pairs[i + 1][0] == 70:
                insunits = int(pairs[i + 1][1])
            i += 1
            continue
        if section == "ENTITIES" and code == 0:
            if val == "LINE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_line(groups)
            elif val == "ARC":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_arc(groups)
            elif val == "CIRCLE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_arc(groups, full_circle=True)
            elif val == "LWPOLYLINE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_lwpolyline(groups)
            elif val == "POLYLINE":
                ent, i = _build_polyline(pairs, i + 1)
            else:
                skipped[val] = skipped.get(val, 0) + 1
                i += 1
                continue
            if ent is not None:
                entities.append(ent)
            continue
        i += 1
    return entities, skipped, insunits


def dxf_metrics(path):
    """Cut length (in) + anomaly info for one part DXF."""
    entities, skipped, insunits = parse_dxf(path)
    scale, unit_note = 1.0, None
    if insunits == 4:            # millimetres
        scale, unit_note = 1.0 / 25.4, "units=mm (converted to in)"
    elif insunits not in (1, 0, None):
        unit_note = f"unexpected $INSUNITS={insunits}"
    total = 0.0
    layers = set()
    for e in entities:
        layers.add(e["layer"])
        if e["layer"].strip().upper() in EXCLUDE_LAYERS:
            continue
        total += e["length"]
    return {
        "linear_inches": total * scale,
        "skipped": skipped,
        "layers": layers,
        "unit_note": unit_note,
        "n_entities": len(entities),
    }


# ---------------------------------------------------------------------------
# Folder resolution + nest package lookup
# ---------------------------------------------------------------------------

def _find_child_dir(parent, name):
    try:
        for d in os.listdir(parent):
            if d.strip().upper() == name.upper() and os.path.isdir(os.path.join(parent, d)):
                return os.path.join(parent, d)
    except OSError:
        pass
    return None


def resolve_roots(selected):
    """Return (iges_root, batch_dir) from whatever the user picked."""
    selected = os.path.abspath(selected)
    if os.path.basename(selected).strip().upper() == "IGES FILES":
        return selected, os.path.dirname(selected)
    child = _find_child_dir(selected, "IGES FILES")
    if child:
        return child, selected
    return selected, os.path.dirname(selected)


def derive_nest(subfolder_name):
    """N5CDARQS -> 5CDARQ; anything else returned unchanged."""
    s = subfolder_name
    if len(s) > 2 and s[0].upper() == "N" and s[-1].upper() == "S":
        return s[1:-1]
    return s


def match_work_order(filename, summary):
    """Resolve a DXF filename to its work order.

    DXF names embed PART.REV.WORKORDER (e.g. "H4533219-16M.01.X4549775.dxf"), so a
    blind regex search grabs the part number's digits first. Instead we match against
    the KNOWN summary work orders: the filename's dot-separated component that equals a
    summary WO, else a summary WO that's a substring of the name. Falls back to a
    WO_RE search only when there is no summary (legacy filenames). Returns "?" if none.
    """
    stem = os.path.splitext(os.path.basename(filename))[0].upper()
    comps = re.split(r"[.\s_]+", stem)
    for wo in summary:                       # exact dot-separated component (preferred)
        if wo in comps:
            return wo
    for wo in summary:                       # substring fallback
        if wo in stem:
            return wo
    # No summary match: the work order trails the part in the name (PART.REV.WO), so
    # take the last WO-shaped component - NOT a WO_RE search, which grabs the part's
    # leading digits first.
    for comp in reversed(comps):
        if WO_RE.fullmatch(comp):
            return comp
    return "?"


def parse_package_pdf(pdf, log):
    """Parse ONE nest-package PDF -> (summary {WO: (PART, qty)}, thk_by_wo {WO: in},
    thk_by_part {PART: in}, first_thickness). Keys are UPPERCASED so DXF filename
    matching is case-insensitive.

    The SUMMARY OF NEST table renders as a flat line sequence
    ``REF / PART NUMBER / QTY / WORK ORDER / SK`` then one row of those 5 fields per
    part; we scan for each WORK ORDER token and read qty (1 line back) + part (2 back).
    Thickness comes off each PART SKETCH page (SIZE: "1.750 THK"), keyed by BOTH the
    page's work order (the line just before "PART SKETCH") and its ``PART:`` number, so
    a DXF resolves its thickness by whichever it can match.
    """
    summary, thk_by_wo, thk_by_part, first_thk = {}, {}, {}, None
    try:
        sdk.ensure_local(pdf, log=log)
        doc = fitz.open(pdf)
    except Exception as exc:
        log(f"  ! could not open package PDF {os.path.basename(pdf)}: {exc}")
        return summary, thk_by_wo, thk_by_part, first_thk
    try:
        for page in doc:
            text = page.get_text()
            if not summary and "SUMMARY OF NEST" in text:
                toks = [t.strip() for t in text.splitlines() if t.strip()]
                for i, tk in enumerate(toks):
                    if WO_RE.fullmatch(tk) and i >= 2:
                        try:
                            qty = int(toks[i - 1])
                        except ValueError:
                            continue
                        summary[tk.upper()] = (toks[i - 2].upper(), qty)
            if "PART SKETCH" in text:
                lines = [t.strip() for t in text.splitlines() if t.strip()]
                tk = THK_RE.search(text)
                try:
                    val = float(tk.group(1)) if tk else None
                except ValueError:
                    val = None
                if val is not None:
                    pm = PART_RE.search(text)
                    if pm:
                        thk_by_part[pm.group(1).upper()] = val
                    # The work order is the line right before "PART SKETCH".
                    if "PART SKETCH" in lines:
                        idx = lines.index("PART SKETCH")
                        wo_line = lines[idx - 1] if idx > 0 else ""
                        if WO_RE.fullmatch(wo_line):
                            thk_by_wo[wo_line.upper()] = val
                    if first_thk is None:
                        first_thk = val
    finally:
        doc.close()
    return summary, thk_by_wo, thk_by_part, first_thk


def _locate_nest_pdf(nest, nest_pkg_dir, batch_dir):
    """The PDF for one nest: NEST PACKAGES\\<nest>.pdf, else the nest order folder's
    MOVE TICKET OMIT pdf. None if neither exists."""
    if nest_pkg_dir:
        cands = glob.glob(os.path.join(nest_pkg_dir, f"{nest}.pdf")) + \
            glob.glob(os.path.join(nest_pkg_dir, f"{nest} *.pdf"))
        if cands:
            return cands[0]
    nest_folder = _find_child_dir(batch_dir, nest)
    if nest_folder:
        cands = glob.glob(os.path.join(nest_folder, "*MOVE TICKET OMIT*.pdf"))
        if cands:
            return cands[0]
    return None


def read_nest_package(nest, nest_pkg_dir, batch_dir, log):
    """(summary, thk_by_wo, thk_by_part, nest_thickness) for one nest, or empty when it
    has no package of its own (a foreign nest — the batch-wide index covers those)."""
    pdf = _locate_nest_pdf(nest, nest_pkg_dir, batch_dir)
    if pdf is None:
        return {}, {}, {}, None
    return parse_package_pdf(pdf, log)


def build_batch_index(nest_pkg_dir, batch_dir, log):
    """Batch-wide fallback: merge the summaries + thickness of EVERY nest-package PDF so
    a foreign nest (its parts nested under a different order, no package of its own —
    e.g. GX029's 5CDAWH resolves via 5CDAYP.pdf) still gets qty + thickness. Returns
    (summary, thk_by_wo, thk_by_part)."""
    all_summary, all_thk_wo, all_thk_part = {}, {}, {}
    pdfs = glob.glob(os.path.join(nest_pkg_dir, "*.pdf")) if nest_pkg_dir else []
    for pdf in sorted(pdfs):
        s, tw, tp, _ = parse_package_pdf(pdf, log)
        all_summary.update(s)
        all_thk_wo.update(tw)
        all_thk_part.update(tp)
    if all_summary:
        log(f"Batch index: {len(all_summary)} work order(s) across {len(pdfs)} package PDF(s).")
    return all_summary, all_thk_wo, all_thk_part


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def write_report(out_path, batch, nests, log):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Linear Inches"
    header = ["Batch", "Nest", "Part Number", "Work Order", "Qty", "Thickness (in)",
              "Individual Linear Inches", "Total Linear Inches (Part x Qty)",
              "Feedrate (IPM)", "Cut Time"]
    NCOL = len(header)
    LI_COL = 8            # "Total Linear Inches" column index (1-based)
    hf, hb = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="305496")
    nf, nb = Font(bold=True), PatternFill("solid", fgColor="D9E1F2")
    bf, bb = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F3864")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(thin, thin, thin, thin)

    def blank_row():
        return [""] * NCOL

    ws.append(header)
    for c in ws[1]:
        c.font, c.fill = hf, hb
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border

    CT_COL = 10          # "Cut Time" column index (1-based)

    r = 2
    batch_total = 0.0
    batch_min = 0.0
    for nest, parts in nests:
        nest_total = sum(p["total"] for p in parts)
        nest_min = sum(p.get("min") or 0.0 for p in parts)
        has_time = any(p.get("min") is not None for p in parts)
        for p in parts:
            ws.append([
                batch, nest, p["part"], p["wo"], p["qty"],
                p.get("thickness"), round(p["indiv"], 2), round(p["total"], 2),
                p.get("ipm") if p.get("ipm") is not None else "",
                fmt_hm(p.get("min")),
            ])
            for c in ws[r]:
                c.border = border
            r += 1
        sub = blank_row()
        sub[4] = f"NEST {nest} TOTAL"
        sub[LI_COL - 1] = round(nest_total, 2)
        sub[CT_COL - 1] = fmt_hm(nest_min) if has_time else ""
        ws.append(sub)
        for c in ws[r]:
            c.font, c.fill, c.border = nf, nb, border
        r += 1
        batch_total += nest_total
        batch_min += nest_min

    tot = blank_row()
    tot[4] = f"BATCH {batch} TOTAL"
    tot[LI_COL - 1] = round(batch_total, 2)
    tot[CT_COL - 1] = fmt_hm(batch_min) if batch_min else ""
    ws.append(tot)
    for c in ws[r]:
        c.font, c.fill, c.border = bf, bb, border
    r += 1

    note = blank_row()
    note[0] = (f"Feed rates: {FEED_RATES_SOURCE}. Cut time = total linear inches / feedrate. "
               "A thickness between two chart rows uses the next-thicker row's feedrate "
               "(conservative), so outlier thicknesses stay covered without editing the chart.")
    ws.append(note)
    ws[r][0].font = Font(italic=True, color="808080")

    for i, w in enumerate([10, 12, 22, 13, 6, 12, 16, 20, 14, 13], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    try:
        wb.save(out_path)
    except PermissionError:
        alt = out_path[:-5] + " (new).xlsx"
        wb.save(alt)
        log(f"  ! '{os.path.basename(out_path)}' was open/locked - saved as '{os.path.basename(alt)}' instead.")
        return alt, batch_total
    return out_path, batch_total


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run(params, progress_callback, cancel_event):
    log = params.get("log", print)
    settings = params.get("settings", {}) or {}

    selected = sdk.request_directory(
        params,
        "Select the batch's 'IGES FILES' folder (or the batch folder)",
        start_dir=(settings.get("default_root") or "").strip(),
    )
    if not selected:
        log("Cancelled - no folder selected.")
        return

    iges_root, batch_dir = resolve_roots(selected)
    batch = os.path.basename(batch_dir.rstrip("\\/")) or batch_dir
    nest_pkg_dir = _find_child_dir(batch_dir, "NEST PACKAGES")
    log(f"Batch: {batch}")
    log(f"IGES root: {iges_root}")
    log(f"Nest packages: {nest_pkg_dir or '(not found - will try MOVE TICKET OMIT pdfs)'}")

    # Batch-wide fallback so foreign nests (parts nested under another order) still
    # resolve qty + thickness from whichever package PDF actually lists them.
    batch_summary, batch_thk_wo, batch_thk_part = build_batch_index(
        nest_pkg_dir, batch_dir, log)

    try:
        subs = sorted(d for d in os.listdir(iges_root)
                      if os.path.isdir(os.path.join(iges_root, d)))
    except OSError as exc:
        sdk.show_warning(params, "Folder error", f"Could not read:\n{iges_root}\n\n{exc}")
        return
    if not subs:
        sdk.show_warning(params, "Nothing to process",
                         f"No nest subfolders found in:\n{iges_root}")
        return

    nests = []       # (nest, [part dicts])
    flags = []
    for idx, sub in enumerate(subs):
        if cancel_event is not None and cancel_event.is_set():
            log("Cancelled.")
            return
        progress_callback(int(5 + 85 * idx / len(subs)))
        nest = derive_nest(sub)
        if nest == sub:
            flags.append(f"{sub}: folder name is not the expected N<nest>S format")
        summary, thk_by_wo, thk_by_part, nest_thk = read_nest_package(
            nest, nest_pkg_dir, batch_dir, log)
        # Nest's own package wins; fall back to the batch-wide index for foreign nests.
        eff_summary = {**batch_summary, **summary}
        eff_thk_wo = {**batch_thk_wo, **thk_by_wo}
        eff_thk_part = {**batch_thk_part, **thk_by_part}
        if not eff_summary:
            flags.append(f"{nest}: no SUMMARY OF NEST found - quantities unknown")

        dxfs = sorted(glob.glob(os.path.join(iges_root, sub, "*.dxf")))
        if not dxfs:
            flags.append(f"{nest}: subfolder has no DXF files")
        parts = []
        for dxf in dxfs:
            if cancel_event is not None and cancel_event.is_set():
                log("Cancelled.")
                return
            fn = os.path.basename(dxf)
            try:
                sdk.ensure_local(dxf, log=log)
                m = dxf_metrics(dxf)
            except Exception as exc:
                flags.append(f"{nest}/{fn}: could not read DXF ({exc})")
                continue
            wo = match_work_order(fn, eff_summary)
            part, qty = eff_summary.get(wo, ("(unmatched)", 0))
            if qty == 0:
                flags.append(f"{nest}/{fn}: no quantity match for work order {wo}")
            if m["skipped"]:
                bad = ", ".join(f"{k}x{v}" for k, v in m["skipped"].items()
                                if k in UNSUPPORTED_FLAGGED)
                if bad:
                    flags.append(f"{nest}/{fn}: unmeasured geometry ({bad}) - length may be low")
            if len(m["layers"]) > 1:
                flags.append(f"{nest}/{fn}: {len(m['layers'])} layers - verify all count as cut")
            if m["unit_note"]:
                flags.append(f"{nest}/{fn}: {m['unit_note']}")
            indiv = m["linear_inches"]
            total = indiv * qty
            thickness = eff_thk_wo.get(wo) or eff_thk_part.get(part) or nest_thk
            times = part_cut_times(total, thickness)
            if thickness is None:
                flags.append(f"{nest}/{fn}: thickness not found - no cut-time estimate")
            elif times["note"]:
                flags.append(f"{nest}/{fn}: thickness {thickness}\" {times['note']} "
                             f"({times['ipm']} IPM)")
            rec = {"part": part, "wo": wo, "qty": qty, "indiv": indiv,
                   "total": total, "thickness": thickness}
            rec.update(times)
            parts.append(rec)
            log(f"  {nest}  {part}  WO {wo}  {indiv:.2f} in x {qty} = {total:.2f} in"
                f"  thk {thickness}  {fmt_hm(times['min']) or '-'}"
                f"{' @ ' + str(times['ipm']) + ' IPM' if times['ipm'] is not None else ''}")
        if parts:
            nests.append((nest, parts))

    if not nests:
        sdk.show_warning(params, "No results",
                         "No part DXFs could be measured. Check the folder selection.")
        return

    progress_callback(94)
    out_path = os.path.join(batch_dir, f"{batch} Linear Inches.xlsx")
    saved_path, batch_total = write_report(out_path, batch, nests, log)

    log("")
    log(f"BATCH {batch} TOTAL = {batch_total:.2f} in = {batch_total / 12:.2f} ft")
    log(f"Report: {saved_path}")
    console = params.get("console")
    if console is not None and hasattr(console, "append_link"):
        try:
            console.append_link(os.path.basename(saved_path), saved_path,
                                prefix="REPORT", at_run_end=True)
        except TypeError:
            console.append_link(os.path.basename(saved_path), saved_path)

    if flags:
        log("")
        log(f"{len(flags)} item(s) flagged for review:")
        for f in flags:
            log(f"  ! {f}")
        sdk.show_warning(params, "911 Linear Inch CutTime - review needed",
                         "Report saved, but some items need a look:\n\n" +
                         "\n".join(f"- {f}" for f in flags[:20]) +
                         ("\n..." if len(flags) > 20 else ""))
    progress_callback(100)
