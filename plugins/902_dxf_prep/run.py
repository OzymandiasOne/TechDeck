"""
902 DXF Prep - TechDeck plugin.

Automates the SharePoint-side steps of the "Batch DXF Cleanup and Preparation
for Boost 902-Part Files" SOP. The SigmaNest VM import/export and the AutoCAD
Layer-12 macro stay manual; this plugin covers the bookends as three processes
picked in a 911-Setup-style checkbox dialog (root "All processes" toggles the
set; everything is selected by default, so an out-of-the-gate run walks the
whole workflow back-to-back). Already-done processes are flagged in the picker;
all are idempotent. Selected processes run in order, stopping if one fails:

  1. SETUP            - copy any .igs files sitting in the selected folder
                        into "BATCH# - IGES CONVERT" (skipped with a note if
                        there are none), and build a standalone
                        "BATCH# QTY.xlsx" (DYPN + QTY + NOTES, sorted, repeat
                        groups boxed, multi-qty rows highlighted) from the PO
                        spreadsheet.
  2. RENAME + SORT    - after the DXF export returns from the VM: strip
                        _FLAT-PATTERN#n / revision / trailing-number suffixes
                        from the DXF filenames, then split them into numbered
                        subfolders of 25 for the AutoCAD review.
  3. RECOMBINE+VERIFY - merge the numbered subfolders back, reconcile files
                        against the PO spreadsheet DYPN list both ways
                        (missing parts -> "BATCH# - MISSING PARTS.txt",
                        extra files -> EXTRA subfolder), and prefix "(Nx) "
                        onto filenames of parts ordered more than once.

Runtime inputs: the folder holding the part files - the SigmaNest DXF export,
or the mixed EB folder with IGES + uncleaned DXFs side by side (native
pick-a-folder dialog via sdk.request_directory; headless fallback pastes a
path) - then the process picker (headless fallback: a text prompt, blank =
all). The PO spreadsheet (the *PRICING* workbook with the PO-#### sheet) is
discovered in the selected folder or its parent; its folder is the BATCH
folder where the outputs (IGES CONVERT, QTY workbook, MISSING PARTS report)
are written. The batch number derives from the PO sheet name.
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path

# SDK bootstrap - works both in-process (TechDeck/frozen exe) and for standalone
# CLI testing (python plugins/902_dxf_prep/run.py).
try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

VERSION = "1.4.0"

CHUNK_SIZE = 25  # AutoCAD review batches (Windows 25-file open limit per SOP)

# A valid 902 part number (DYPN): dash-separated alphanumeric segments that
# contain at least one digit overall. Shapes seen on real POs: H4130810-484 /
# R4533683-5-3 (letter prefix + digits), 263500220-5001 (all-numeric head),
# H5730805A230-1 / H5731304C49-13 (letters embedded mid-number), LTG-2213AE
# (letter-only head, digits in a later segment). The digit requirement plus
# no-spaces/no-underscores is what rejects footer/junk rows like
# "NAME OF VENDOR REPRESENTATIVE".
_PART_RE = re.compile(r"^(?=[A-Z0-9-]*\d)[A-Z0-9]{2,16}(?:-[A-Z0-9]+)+$",
                      re.IGNORECASE)

# Suffix tokens SigmaNest's export appends to a DYPN, stripped right-to-left:
# _FLAT-PATTERN#1, a lone revision letter (_A), or a trailing number (_2).
# Part numbers use dashes, never underscores, so underscore tokens are safe
# to strip.
_SUFFIX_TOKEN_RE = re.compile(r"_(?:FLAT-PATTERN#\d+|[A-Z]|\d+)$", re.IGNORECASE)

# "(18x) " quantity prefix (mode 3 output; stripped when re-deriving the DYPN).
_QTY_PREFIX_RE = re.compile(r"^\(\d+x\)\s*", re.IGNORECASE)


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def clean_stem(stem: str) -> str:
    """Reduce a DXF filename stem to its DYPN: drop a (Nx) prefix, then strip
    export suffix tokens right-to-left. Returns the cleaned stem unchanged if
    it never resolves to a valid part number (caller decides what to do)."""
    s = _QTY_PREFIX_RE.sub("", stem).strip()
    while True:
        if _PART_RE.match(s):
            return s
        m = _SUFFIX_TOKEN_RE.search(s)
        if not m:
            return s
        s = s[: m.start()]


def _find_pricing_workbook(batch_folder: Path, log) -> Path | None:
    """The PO spreadsheet inside the batch folder: prefer a *PRICING* workbook,
    else any .xlsm/.xlsx (skipping Excel ~$ locks and our own QTY output)."""
    candidates = [
        p for p in batch_folder.glob("*.xls[mx]")
        if not p.name.startswith("~$")
        and not p.stem.upper().endswith("QTY")
    ]
    if not candidates:
        return None
    pricing = [p for p in candidates if "PRICING" in p.name.upper()]
    pick = sorted(pricing or candidates)[0]
    if len(candidates) > 1:
        log(f"  ({len(candidates)} workbooks in the batch folder; using {pick.name})")
    return pick


def _load_po_rows(pricing_path: Path, log):
    """Read the PO sheet: returns (batch_number|None, rows) where rows are
    (dypn_upper, qty) in sheet order. Prefers a sheet named PO-#### ; falls
    back to the first sheet containing DYPN + QTY ORDERED headers."""
    wb = sdk.load_workbook_resilient(pricing_path, log=log,
                                     read_only=True, data_only=True)
    try:
        # Scan VISIBLE sheets only: pricing workbooks carry hidden staging
        # sheets (FORMING, PRICING, ...) whose headers also match, and the
        # fallback below must never read one of those. [-_ ]* tolerates any
        # separator run in the tab name ('PO-4130', 'PO- 4130', 'PO 4130').
        visible = sdk.visible_sheetnames(wb)
        po_sheets = [n for n in visible
                     if re.match(r"^PO[-_ ]*\d+", n.strip(), re.IGNORECASE)]
        sheet_names = po_sheets or visible

        for name in sheet_names:
            ws = wb[name]
            header_row, cols = sdk.find_header_row(ws, ["DYPN"])
            if header_row is None:
                continue
            # Headers can contain line breaks ("QTY \nORDERED") - normalize
            # whitespace before matching names (Hard Rules 1-2: by name only).
            norm = {" ".join(k.split()): v for k, v in cols.items()}
            dypn_col = norm.get("DYPN")
            qty_col = norm.get("QTY ORDERED")
            if dypn_col is None or qty_col is None:
                continue

            m = re.search(r"(\d+)", name)
            batch = m.group(1) if m and po_sheets else None

            rows: list[tuple[str, float]] = []
            skipped = []
            for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
                raw = r[dypn_col - 1]
                if raw is None or not str(raw).strip():
                    continue
                dypn = str(raw).strip().upper()
                if not _PART_RE.match(dypn):
                    skipped.append(dypn)
                    continue
                q = r[qty_col - 1]
                qty = float(q) if isinstance(q, (int, float)) else 1.0
                rows.append((dypn, qty))
            if skipped:
                log(f"  (skipped {len(skipped)} non-part row(s) on '{name}', "
                    f"e.g. {skipped[0][:40]!r})")
            if not po_sheets:
                log(f"  WARNING: no PO-#### sheet in {pricing_path.name} - "
                    f"using '{name}' (first visible sheet with DYPN + QTY "
                    f"ORDERED headers). Verify this is the right sheet!")
            log(f"  PO sheet '{name}': {len(rows)} part rows, "
                f"{len({d for d, _ in rows})} distinct DYPNs")
            return batch, rows

        log(f"ERROR: no sheet in {pricing_path.name} has DYPN + QTY ORDERED headers.")
        return None, []
    finally:
        wb.close()


def _resolve_batch_context(params, cancel_event, log):
    """Pick the batch folder (native folder dialog; pasted path when headless),
    find the PO spreadsheet, derive the batch number. Returns (batch_folder,
    pricing_path, batch, po_rows) - pricing may be None (some modes can
    proceed without it)."""
    start_dir = ""
    try:
        roots = sdk.pilot_program_roots()
        if roots:
            start_dir = str(roots[0])
    except Exception:
        pass
    raw = sdk.request_directory(
        params, "Select the folder with the part files (DXF / IGES)", start_dir)
    if cancel_event.is_set():
        return None, None, None, None, []
    if not raw:
        log("Folder selection cancelled - nothing was run.")
        cancel_event.set()  # user cancel: not a successful (ticket-earning) run
        return None, None, None, None, []
    work_folder = Path(raw.strip().strip('"'))
    if not sdk.is_dir(work_folder):
        log(f"ERROR: folder not found: {work_folder}")
        return None, None, None, None, []
    log(f"Working folder: {work_folder}")

    # The PO/pricing workbook anchors the BATCH folder: usually the selected
    # folder's parent (the user selects the part-files folder inside the batch
    # folder), sometimes the selected folder itself.
    pricing = _find_pricing_workbook(work_folder, log)
    if pricing is None and work_folder.parent != work_folder:
        pricing = _find_pricing_workbook(work_folder.parent, log)
    batch, po_rows = (None, [])
    if pricing is None:
        log("WARNING: no PO/pricing workbook (.xlsm/.xlsx) found in the selected "
            "folder or its parent.")
    else:
        log(f"PO spreadsheet: {pricing.name}")
        batch, po_rows = _load_po_rows(pricing, log)
    batch_root = pricing.parent if pricing is not None else work_folder
    if batch_root != work_folder:
        log(f"Batch folder (outputs go here): {batch_root}")

    if not batch:
        # Try the folder names, then ask (a real batch number -> shared cache).
        m = (re.search(r"\b(\d{3,5})\b", work_folder.name)
             or re.search(r"\b(\d{3,5})\b", batch_root.name))
        batch = m.group(1) if m else None
    if not batch:
        batch = sdk.request_batch_number(params, "Enter the 902 batch number:").strip()
        if cancel_event.is_set():
            return None, None, None, None, []
    log(f"Batch number: {batch}")
    return work_folder, batch_root, pricing, batch, po_rows


def _find_dxf_folder(batch_folder: Path, batch: str, log) -> Path | None:
    """The exported-DXF folder: "BATCH# - DXF" per the SOP, located loosely
    (Hard Rule: resolve folders defensively, layouts vary). If the user pasted
    the DXF folder itself, batch_folder already contains loose .dxf files."""
    if any(True for _ in batch_folder.glob("*.dxf")):
        return batch_folder
    # Pasted the DXF folder itself while its files sit in numbered subfolders?
    if "DXF" in batch_folder.name.upper() or any(
            sdk.is_dir(d) and d.name.isdigit() and any(True for _ in d.glob("*.dxf"))
            for d in batch_folder.iterdir()):
        return batch_folder
    candidates = []
    for d in batch_folder.iterdir():
        if not sdk.is_dir(d):
            continue
        name = d.name.upper()
        if "DXF" in name and "CONVERT" not in name:
            candidates.append(d)
    if not candidates:
        log(f"ERROR: no .dxf files in the selected folder (and no DXF subfolder "
            f'like "{batch} - DXF" under it). Select the folder that holds the '
            f"DXF part files.")
        return None
    exact = [d for d in candidates if d.name.upper().startswith(str(batch))]
    pick = sorted(exact or candidates)[0]
    if len(candidates) > 1:
        log(f"  ({len(candidates)} DXF-named folders; using {pick.name})")
    return pick


def _loose_dxfs(dxf_folder: Path) -> list[Path]:
    """The .dxf files sitting directly in the DXF folder (not in numbered
    subfolders / EXTRA), sorted by name."""
    return sorted(p for p in dxf_folder.glob("*.dxf") if sdk.is_file(p))


# ─────────────────────────────────────────────────────────────────────────────
# Mode 1 - Setup: IGES CONVERT folder + QTY workbook
# ─────────────────────────────────────────────────────────────────────────────

def _mode_setup(work_folder: Path, batch_root: Path, batch: str,
                po_rows, log, progress_callback, cancel_event) -> bool:
    # 1) BATCH# - IGES CONVERT: copy any .igs sitting in the selected folder
    #    (mixed EB folders hold IGES + uncleaned DXFs side by side). No IGES
    #    present is fine - a SigmaNest DXF export folder has none, and the QTY
    #    sheet below is still worth building.
    iges = sorted({p for p in list(work_folder.glob("*.igs"))
                   + list(work_folder.glob("*.iges"))})
    if not iges:
        log("No .igs/.iges files in the selected folder - skipping the IGES copy.")
    else:
        convert_dir = batch_root / f"{batch} - IGES CONVERT"
        sdk.ensure_dir(convert_dir)
        log(f"IGES working folder: {convert_dir.name} ({len(iges)} .igs found)")
        copied = skipped = 0
        for i, p in enumerate(iges):
            if cancel_event.is_set():
                log("Cancelled.")
                return False
            dest = convert_dir / p.name
            if sdk.exists(dest):
                skipped += 1
                continue
            sdk.ensure_local(p, log=log)  # OneDrive placeholder-safe (Hard Rule 13)
            shutil.copy2(sdk.long_path(p), sdk.long_path(dest))
            copied += 1
            if copied % 25 == 0:
                log(f"  copied {copied}/{len(iges)}...")
            progress_callback(5 + int(55 * (i + 1) / len(iges)))
        log(f"Copied {copied} .igs file(s)" +
            (f" ({skipped} already present, skipped)" if skipped else "") + ".")

    # 2) QTY workbook (written to the batch folder, next to the PO spreadsheet)
    if not po_rows:
        log("WARNING: no PO rows read - QTY workbook not built.")
        return True
    qty_path = batch_root / f"{batch} QTY.xlsx"
    if sdk.exists(qty_path):
        log(f"NOTE: {qty_path.name} already exists - leaving it untouched "
            f"(it may hold review notes). Delete it and rerun Setup to rebuild.")
        return True
    _write_qty_workbook(qty_path, po_rows, log)
    progress_callback(95)
    return True


def _write_qty_workbook(qty_path: Path, po_rows, log) -> None:
    """Standalone "BATCH# QTY.xlsx": DYPN | QTY | NOTES sorted by part number,
    an outline border boxing each run of repeated DYPNs, and a highlight on
    every row whose part is ordered more than once (per SOP)."""
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = sorted(po_rows, key=lambda t: t[0])
    totals: dict[str, float] = {}
    for dypn, qty in rows:
        totals[dypn] = totals.get(dypn, 0) + qty

    wb = Workbook()
    ws = wb.active
    ws.title = "QTY"
    ws.append(["DYPN", "QTY", "NOTES"])
    for c in range(1, 4):
        ws.cell(1, c).font = Font(bold=True)
    for dypn, qty in rows:
        ws.append([dypn, int(qty) if float(qty).is_integer() else qty, None])

    thin = Side(style="thin")
    thick = Side(style="medium")
    fill = PatternFill("solid", start_color="FFF3B0")  # multi-qty highlight
    r = 2
    while r <= ws.max_row:
        dypn = ws.cell(r, 1).value
        end = r
        while end + 1 <= ws.max_row and ws.cell(end + 1, 1).value == dypn:
            end += 1
        multi_qty = totals.get(dypn, 0) > 1
        repeat_group = end > r
        for rr in range(r, end + 1):
            for cc in range(1, 4):
                cell = ws.cell(rr, cc)
                if repeat_group:
                    cell.border = Border(
                        left=thick if cc == 1 else thin,
                        right=thick if cc == 3 else thin,
                        top=thick if rr == r else thin,
                        bottom=thick if rr == end else thin,
                    )
                else:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if multi_qty:
                    cell.fill = fill
        r = end + 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 50
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.freeze_panes = "A2"
    sdk.save_workbook(wb, qty_path)
    n_multi = sum(1 for v in totals.values() if v > 1)
    log(f"QTY workbook written: {qty_path.name} "
        f"({len(rows)} rows, {len(totals)} distinct parts, {n_multi} multi-qty).")


# ─────────────────────────────────────────────────────────────────────────────
# Mode 2 - Rename + Sort
# ─────────────────────────────────────────────────────────────────────────────

def _mode_rename_sort(dxf_folder: Path, log, progress_callback, cancel_event) -> bool:
    existing_numbered = [d for d in dxf_folder.iterdir()
                         if sdk.is_dir(d) and d.name.isdigit()]
    if existing_numbered:
        log(f"ERROR: {dxf_folder.name} already has numbered subfolders "
            f"({len(existing_numbered)}) - it looks sorted already. "
            f"Run step 3 (Recombine + Verify) first if you need to redo this.")
        return False

    files = _loose_dxfs(dxf_folder)
    if not files:
        log(f"ERROR: no .dxf files found in {dxf_folder}.")
        return False
    log(f"Renaming {len(files)} DXF file(s) in {dxf_folder.name}...")

    renamed = unchanged = 0
    problems: list[str] = []
    for i, p in enumerate(files):
        if cancel_event.is_set():
            log("Cancelled.")
            return False
        cleaned = clean_stem(p.stem)
        if not _PART_RE.match(cleaned):
            problems.append(f"UNRECOGNIZED NAME (left as-is): {p.name}")
            continue
        if cleaned == p.stem:
            unchanged += 1
            continue
        target = p.with_name(cleaned + p.suffix.lower())
        if sdk.exists(target):
            problems.append(f"DUPLICATE after cleanup (left as-is): {p.name} -> {target.name}")
            continue
        p.rename(target)
        renamed += 1
        progress_callback(5 + int(45 * (i + 1) / len(files)))

    log(f"Renamed {renamed}, already clean {unchanged}.")
    for msg in problems:
        log(f"  WARNING: {msg}")

    # Split into numbered folders of 25 for the AutoCAD pass.
    files = _loose_dxfs(dxf_folder)
    n_folders = (len(files) + CHUNK_SIZE - 1) // CHUNK_SIZE
    for idx, p in enumerate(files):
        if cancel_event.is_set():
            log("Cancelled mid-sort - rerun step 2 to finish.")
            return False
        sub = dxf_folder / str(idx // CHUNK_SIZE + 1)
        sdk.ensure_dir(sub)
        p.rename(sub / p.name)
        progress_callback(50 + int(48 * (idx + 1) / len(files)))
    log(f"Sorted {len(files)} file(s) into {n_folders} folder(s) of {CHUNK_SIZE}.")
    if problems:
        log(f"NOTE: {len(problems)} file(s) need a manual look (see warnings above) - "
            f"document them on the QTY spreadsheet.")
    return True


# ─────────────────────────────────────────────────────────────────────────────
# Mode 3 - Recombine + Verify
# ─────────────────────────────────────────────────────────────────────────────

def _mode_recombine_verify(batch_root: Path, dxf_folder: Path, batch: str,
                           po_rows, log, progress_callback, cancel_event) -> bool:
    # 1) Recombine: pull files out of numbered subfolders, drop empty ones.
    numbered = sorted((d for d in dxf_folder.iterdir()
                       if sdk.is_dir(d) and d.name.isdigit()),
                      key=lambda d: int(d.name))
    moved = 0
    for d in numbered:
        for p in sorted(d.glob("*.dxf")):
            if cancel_event.is_set():
                log("Cancelled mid-recombine - rerun step 3 to finish.")
                return False
            dest = dxf_folder / p.name
            if sdk.exists(dest):
                log(f"  WARNING: {p.name} exists in both {d.name}\\ and the main "
                    f"folder - left in {d.name}\\.")
                continue
            p.rename(dest)
            moved += 1
        try:
            d.rmdir()
        except OSError:
            log(f"  WARNING: folder {d.name}\\ not empty after merge - left in place.")
    if numbered:
        log(f"Recombined {moved} file(s) from {len(numbered)} numbered folder(s).")
    else:
        log("No numbered subfolders - folder already recombined.")
    progress_callback(30)

    if not po_rows:
        log("ERROR: no PO rows available - cannot reconcile. Is the PO/pricing "
            "workbook in the batch folder?")
        return False

    # 2) Reconcile both ways (the flowchart): spreadsheet->folder = missing,
    #    folder->spreadsheet = extra.
    totals: dict[str, float] = {}
    for dypn, qty in po_rows:
        totals[dypn] = totals.get(dypn, 0) + qty

    files = _loose_dxfs(dxf_folder)
    by_part: dict[str, Path] = {}
    unrecognized: list[Path] = []
    for p in files:
        part = clean_stem(p.stem).upper()
        if _PART_RE.match(part):
            if part in by_part:
                log(f"  WARNING: duplicate files for {part}: "
                    f"{by_part[part].name} and {p.name}")
            by_part.setdefault(part, p)
        else:
            unrecognized.append(p)

    missing = sorted(set(totals) - set(by_part))
    extra_parts = sorted(set(by_part) - set(totals))
    matched = len(by_part) - len(extra_parts)
    progress_callback(50)

    # 3) Extras -> EXTRA subfolder (per request: relocate, don't delete).
    extras_moved = []
    if extra_parts or unrecognized:
        extra_dir = dxf_folder / "EXTRA"
        sdk.ensure_dir(extra_dir)
        for part in extra_parts:
            p = by_part[part]
            dest = extra_dir / p.name
            if not sdk.exists(dest):
                p.rename(dest)
            extras_moved.append(p.name)
        for p in unrecognized:
            dest = extra_dir / p.name
            if not sdk.exists(dest):
                p.rename(dest)
            extras_moved.append(p.name + "  (unrecognized filename)")
        log(f"Moved {len(extras_moved)} extra file(s) to {extra_dir.name}\\.")
    progress_callback(65)

    # 4) (Nx) quantity prefixes on multi-qty parts still in the main folder.
    prefixed = 0
    for part, total in sorted(totals.items()):
        if cancel_event.is_set():
            return False
        if total <= 1 or part not in by_part or part in extra_parts:
            continue
        p = by_part[part]
        if not sdk.exists(p):
            continue
        want = f"({int(total)}x) {clean_stem(p.stem)}{p.suffix}"
        if p.name == want:
            continue
        target = p.with_name(want)
        if sdk.exists(target):
            log(f"  WARNING: {want} already exists - left {p.name} alone.")
            continue
        p.rename(target)
        prefixed += 1
    log(f"Quantity-prefixed {prefixed} multi-qty file(s).")
    progress_callback(80)

    # 5) Missing-parts text file + console summary.
    report = batch_root / f"{batch} - MISSING PARTS.txt"
    lines = [f"BATCH {batch} - DXF RECONCILIATION",
             f"Matched: {matched}   Missing: {len(missing)}   Extra: {len(extras_moved)}",
             "",
             f"MISSING FROM FOLDER ({len(missing)}) - on the PO spreadsheet, no DXF file:"]
    lines += [f"  {part}" for part in missing] if missing else ["  (none)"]
    if extras_moved:
        lines += ["", f"EXTRA FILES ({len(extras_moved)}) - moved to EXTRA\\, "
                      "not on the PO spreadsheet:"]
        lines += [f"  {name}" for name in extras_moved]
    report.write_text("\n".join(lines) + "\n", encoding="utf-8")

    log("")
    log("=" * 50)
    log(f"RECONCILIATION - batch {batch}")
    log(f"  Parts on PO spreadsheet : {len(totals)}")
    log(f"  DXF files matched       : {matched}")
    log(f"  MISSING (no file)       : {len(missing)}")
    for part in missing[:30]:
        log(f"      {part}")
    if len(missing) > 30:
        log(f"      ... and {len(missing) - 30} more (see the text file)")
    log(f"  EXTRA (moved to EXTRA\\) : {len(extras_moved)}")
    log(f"Report: {report.name}")
    log("=" * 50)
    progress_callback(98)
    return True


# ─────────────────────────────────────────────────────────────────────────────
# Process selection (911-Setup-style checkbox dialog) + entry point
# ─────────────────────────────────────────────────────────────────────────────

STEP_LABELS = {
    1: "1. Setup  -  IGES CONVERT folder + QTY sheet",
    2: "2. Rename + Sort  -  clean DXF names, split into folders of 25",
    3: "3. Recombine + Verify  -  reconcile vs PO, EXTRA folder + missing list",
}


def _detect_done_steps(work_folder: Path, batch_root: Path, batch: str) -> set[int]:
    """Which steps show an "already done" flag in the picker. Heuristic only -
    every step is idempotent, so re-running a flagged one is always safe."""
    done: set[int] = set()
    if sdk.is_dir(batch_root / f"{batch} - IGES CONVERT"):
        done.add(1)
    dxf = _find_dxf_folder(work_folder, batch, lambda *_: None)
    if dxf is not None and any(sdk.is_dir(d) and d.name.isdigit()
                               for d in dxf.iterdir()):
        done.add(2)
    if sdk.exists(batch_root / f"{batch} - MISSING PARTS.txt"):
        done.add(3)
    return done


def _choose_steps(params, work_folder: Path, batch_root: Path, batch: str,
                  cancel_event, log) -> list[int] | None:
    """Pop the process picker (all checked by default) and return the chosen
    step numbers in run order; None = user cancelled. Headless fallback: a
    text prompt where blank means all."""
    console = params.get("console")
    if console is not None and hasattr(console, "request_selection"):
        labels = [STEP_LABELS[i] for i in (1, 2, 3)]
        done = {STEP_LABELS[i]
                for i in _detect_done_steps(work_folder, batch_root, batch)}
        selection = sdk.request_selection(
            params,
            labels,
            done,
            window_title=f"902 DXF Prep - Batch {batch}",
            header="Select Processes to Run",
            root_label="All processes",
            noun="process",
            done_label="already done",
            overwrite_note="all steps are safe to re-run",
            prompt_note="Selected processes run in order, no breaks.",
            subhead_prefix=f"Batch {batch} - ",
            run_button_text="Run Selected",
        )
        if selection is None:
            log("Process selection cancelled - nothing was run.")
            return None   # sdk.request_selection already flagged the cancel
        return sorted(int(label[0]) for label in selection)

    raw = sdk.request_text(
        params, "Steps to run (e.g. '1 2 3' or '2 3'; blank = all):").strip()
    if cancel_event.is_set():
        return None
    picks = sorted({int(tok) for tok in re.findall(r"[123]", raw)}) or [1, 2, 3]
    return picks


def run(params: dict, progress_callback, cancel_event):
    log = params.get("log", print)

    log("=" * 60)
    log(f"902 DXF Prep v{VERSION}")
    log("=" * 60)

    work_folder, batch_root, pricing, batch, po_rows = _resolve_batch_context(
        params, cancel_event, log)
    if work_folder is None:
        return

    steps = _choose_steps(params, work_folder, batch_root, batch,
                          cancel_event, log)
    if not steps:
        return
    log("Running: " + ",  ".join(STEP_LABELS[s] for s in steps))
    progress_callback(3)

    # Each selected step gets an equal slice of the progress bar; the step
    # implementations report 0-100 internally and get rescaled here.
    span = 95 // len(steps)
    for n, step in enumerate(steps):
        base = 3 + n * span
        scaled = lambda p, b=base: progress_callback(b + p * span // 100)
        log("")
        log(f"--- {STEP_LABELS[step]} ---")

        if step == 1:
            ok = _mode_setup(work_folder, batch_root, batch, po_rows,
                             log, scaled, cancel_event)
        else:
            dxf_folder = _find_dxf_folder(work_folder, batch, log)
            if dxf_folder is None:
                ok = False
            elif step == 2:
                if dxf_folder != work_folder:
                    log(f"DXF folder: {dxf_folder}")
                ok = _mode_rename_sort(dxf_folder, log, scaled, cancel_event)
            else:
                if dxf_folder != work_folder:
                    log(f"DXF folder: {dxf_folder}")
                ok = _mode_recombine_verify(batch_root, dxf_folder, batch,
                                            po_rows, log, scaled, cancel_event)

        if cancel_event.is_set():
            return
        if not ok:
            remaining = [s for s in steps if s > step]
            if remaining:
                log(f"Stopping - step {step} did not complete, skipping "
                    f"step(s) {', '.join(map(str, remaining))}.")
            return

    progress_callback(100)
    log("\nDONE.")


if __name__ == "__main__":
    # Headless smoke test: python plugins/902_dxf_prep/run.py
    import threading
    run({"log": print, "settings": {}, "console": None},
        lambda p: print(f"[{p}%]"), threading.Event())
