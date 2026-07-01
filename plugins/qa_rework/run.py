"""
Gemba Analysis  (plugin id: qa_rework, family: QA)
==================================================
A GUI plugin for the QA team. Two jobs:

  1. Log Entry  - a quick form that appends one rework event as a row to a single
                  shared workbook (one flat sheet, no formulas/charts in the file).
  2. Charts     - reads that workbook back and renders pie / column / stacked-column
                  views with toggles (time window, group-by dimension, %-threshold,
                  best-fit line), plus a one-click "Gemba Pack" PDF of the four
                  standard meeting charts.

Design notes:
  - Charts use the bundled PySide6.QtCharts (NOT matplotlib) so nothing new has to be
    installed on locked-down laptops and the build stays small.
  - PDF export uses Qt's native QPdfWriter/QPainter - also zero new dependencies.
  - The FAILURE MODE cell keeps the team's existing "Category - Subcategory" string
    (e.g. "Laser - Angle"); we split it on the first " - " for category vs. detailed
    grouping, so the file stays compatible with their Master Rework Log.
"""

import datetime
import json
import os
import random
import re
import time
from collections import Counter, OrderedDict
from pathlib import Path

# --- SDK bootstrap (works under the app and for headless CLI testing) -------------
try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys
    import pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

from techdeck.core.plugin_window import PluginWindow

from PySide6.QtWidgets import (
    QWidget, QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QComboBox, QLineEdit, QPushButton, QLabel, QDateEdit, QDoubleSpinBox,
    QCheckBox, QFileDialog, QFrame, QSizePolicy, QStackedWidget, QToolButton,
    QButtonGroup,
)
from PySide6.QtCore import Qt, QDate, QTimer, QPointF, QSize
from PySide6.QtGui import (
    QColor, QBrush, QPainter, QPdfWriter, QPageSize, QPageLayout, QFont,
    QPixmap, QIcon, QPen,
)
from PySide6.QtCharts import (
    QChart, QChartView, QPieSeries, QBarSeries, QStackedBarSeries, QBarSet,
    QBarCategoryAxis, QValueAxis, QLineSeries,
)

# Module-level reference so Qt does not garbage-collect the window (Hard Rule).
_window = None

# ---------------------------------------------------------------------------------
# Controlled vocabulary (mirrors the team's "Tables & Formulae" dropdowns)
# ---------------------------------------------------------------------------------
DATA_FILENAME = "QA Rework Log.xlsx"
SHEET_NAME = "Master Rework Log"
HEADERS = [
    "BATCH (PO)", "DATE", "ITEM NUMBER (DYPN)", "SOURCE MATERIAL",
    "RECUT? (Y/N)", "FAILURE MODE", "COMMENTS",
]

# Category -> subcategories. "" = the bare "Other" with no subcategory.
FAILURE_CATEGORIES = OrderedDict([
    ("Form",   ["Angle", "Location"]),
    ("Grind",  ["Angle", "Surface Finish"]),
    ("Laser",  ["Angle", "Blowout", "Length"]),
    ("Saw",    ["Angle", "Length"]),
    ("Scribe", ["Illegible", "Incorrect Information"]),
    ("Other",  ["", "Material Not Provided", "Product Damaged",
                "Product Not Inspected", "Product Not Processed / Step(s) Skipped"]),
])
MATERIALS = ["CLEVIS", "LUG", "MISC.", "PLATE", "ROD", "TUBE", "OTHER"]
RECUT_OPTIONS = ["NO", "YES"]

GROUP_BY = ["Failure category", "Failure mode (detailed)", "Material", "Recut"]
WINDOWS = [
    "Year-to-date", "Last 7 days", "Last 10 days",
    "Last business week (Mon-Fri)", "All time",
]
DAY_WINDOWS = {"Last 7 days", "Last 10 days", "Last business week (Mon-Fri)"}

CHART_TYPES = ["Pie", "Column", "Stacked column"]
XAXIS_MODES = ["By time (month/day)", "By group"]

# One config per graph. Defaults mirror the Gemba Pack so the 2x2 grid opens with a
# useful board; each config is fully editable from the shared control row.
def _default_graph_configs():
    return [
        {"type": "Column", "window": "Year-to-date", "group": "Failure category",
         "xaxis": "By time (month/day)", "threshold": 5.0, "bestfit": True},
        {"type": "Pie", "window": "Year-to-date", "group": "Failure category",
         "xaxis": "By time (month/day)", "threshold": 5.0, "bestfit": True},
        {"type": "Stacked column", "window": "Last business week (Mon-Fri)",
         "group": "Failure mode (detailed)",
         "xaxis": "By time (month/day)", "threshold": 5.0, "bestfit": True},
        {"type": "Pie", "window": "Last business week (Mon-Fri)",
         "group": "Failure mode (detailed)",
         "xaxis": "By time (month/day)", "threshold": 5.0, "bestfit": True},
    ]

# Print palette for PDFs - white background + dark ink so Gemba printouts read well.
PRINT_BG = "#FFFFFF"
PRINT_TEXT = "#1A1A1A"
PRINT_CYCLE = [
    "#2E5EAA", "#D7791F", "#3E8E41", "#B23A48", "#6A4C93",
    "#1C8C9C", "#C9A227", "#8C5E3C", "#5A5A5A", "#A03E78",
]


# ---------------------------------------------------------------------------------
# Failure-mode parsing
# ---------------------------------------------------------------------------------
def split_mode(mode):
    """'Laser - Angle' -> ('Laser', 'Angle'); 'Other' -> ('Other', '')."""
    mode = (mode or "").strip()
    if " - " in mode:
        cat, sub = mode.split(" - ", 1)
        return cat.strip(), sub.strip()
    return mode, ""


def compose_mode(category, subcategory):
    category = (category or "").strip()
    subcategory = (subcategory or "").strip()
    return f"{category} - {subcategory}" if subcategory else category


# ---------------------------------------------------------------------------------
# Data layer
# ---------------------------------------------------------------------------------
def resolve_data_path(settings, log=print):
    """Return the workbook path, creating its folder. Defaults to a 'QA Rework Log'
    folder under the Pilot Program library so every laptop syncs to one file."""
    configured = ((settings or {}).get("data_dir") or "").strip()
    if configured:
        folder = Path(configured)
    else:
        roots = []
        try:
            roots = sdk.pilot_program_roots()
        except Exception:
            roots = []
        if roots:
            folder = Path(roots[0]) / "QA Rework Log"
        else:
            folder = Path.home() / "TechDeck QA Rework"
            log(f"Pilot Program library not found; using local folder {folder}")
    folder.mkdir(parents=True, exist_ok=True)
    return folder / DATA_FILENAME


def _norm(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip()).upper()


def load_records(path, log=print):
    """Read the log into a list of dicts. Empty list if the file does not exist yet."""
    path = Path(path)
    if not path.exists():
        return []
    wb = sdk.load_workbook_resilient(path, log=log, data_only=True, read_only=True)
    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []

    # Locate the header row + map our known columns by normalized name.
    wanted = {_norm(h) for h in HEADERS}
    header_idx, col = 0, {}
    for i, row in enumerate(rows[:10]):
        hits = {_norm(v): j for j, v in enumerate(row) if v is not None}
        if len(wanted & set(hits)) >= 4:
            header_idx, col = i, hits
            break
    else:
        col = {_norm(v): j for j, v in enumerate(rows[0]) if v is not None}

    def cell(row, header):
        j = col.get(_norm(header))
        return row[j] if j is not None and j < len(row) else None

    records = []
    for row in rows[header_idx + 1:]:
        if row is None or all(v is None for v in row):
            continue
        mode = str(cell(row, "FAILURE MODE") or "").strip()
        cat, sub = split_mode(mode)
        if not mode and not cell(row, "ITEM NUMBER (DYPN)"):
            continue
        records.append({
            "batch": _as_str(cell(row, "BATCH (PO)")),
            "date": _as_date(cell(row, "DATE")),
            "item": _as_str(cell(row, "ITEM NUMBER (DYPN)")),
            "material": _as_str(cell(row, "SOURCE MATERIAL")).upper(),
            "recut": _as_str(cell(row, "RECUT? (Y/N)")).upper(),
            "mode": mode,
            "category": cat,
            "subcategory": sub,
            "comments": _as_str(cell(row, "COMMENTS")),
        })
    return records


def _as_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _as_date(v):
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, datetime.date):
        return datetime.datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"):
            try:
                return datetime.datetime.strptime(v.strip(), fmt)
            except ValueError:
                continue
    return None


def append_record(path, record, log=print, attempts=12):
    """Append one row. Open->append->save->close fast; retry if the file is briefly
    locked (someone has it open in Excel, OneDrive mid-sync, or a concurrent write)."""
    import openpyxl
    path = Path(path)
    ordered = [
        record["batch"], record["date"], record["item"], record["material"],
        record["recut"], record["mode"], record["comments"],
    ]
    last_err = None
    for attempt in range(attempts):
        try:
            if path.exists():
                wb = sdk.load_workbook_resilient(path, log=log)
                ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = SHEET_NAME
                ws.append(HEADERS)
            ws.append(ordered)
            wb.save(path)
            try:
                wb.close()
            except Exception:
                pass
            return
        except (PermissionError, OSError) as e:
            last_err = e
            time.sleep(0.35 + random.random() * 0.5)
    raise last_err if last_err else RuntimeError("append failed")


def ensure_seeded(path, log=print):
    """First run only: if the shared workbook does not exist yet, create it from the
    bundled seed data (the team's starting rework dataset) so the app opens with the
    charts already populated. Never overwrites an existing log."""
    path = Path(path)
    if path.exists():
        return
    seed_file = Path(__file__).resolve().parent / "seed_data.json"
    if not seed_file.exists():
        return
    try:
        rows = json.loads(seed_file.read_text(encoding="utf-8"))
    except Exception as e:
        log(f"Could not read seed data: {e}")
        return
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    for r in rows:
        batch, date_s, item, mat, recut, mode, comments = (list(r) + [""] * 7)[:7]
        d = _as_date(date_s)
        ws.append([batch, d, item, mat, recut, mode, comments])
    wb.save(path)
    log(f"Seeded {len(rows)} starting rows into {path.name}")


# ---------------------------------------------------------------------------------
# Time windows + aggregation
# ---------------------------------------------------------------------------------
def filter_window(records, window, today=None):
    today = today or datetime.date.today()
    def d(r):
        return r["date"].date() if r["date"] else None

    if window == "All time":
        return [r for r in records if d(r)]
    if window == "Year-to-date":
        start = datetime.date(today.year, 1, 1)
        return [r for r in records if d(r) and d(r) >= start]
    if window == "Last 7 days":
        start = today - datetime.timedelta(days=6)
        return [r for r in records if d(r) and start <= d(r) <= today]
    if window == "Last 10 days":
        start = today - datetime.timedelta(days=9)
        return [r for r in records if d(r) and start <= d(r) <= today]
    if window == "Last business week (Mon-Fri)":
        this_monday = today - datetime.timedelta(days=today.weekday())
        last_monday = this_monday - datetime.timedelta(days=7)
        last_friday = last_monday + datetime.timedelta(days=4)
        return [r for r in records if d(r) and last_monday <= d(r) <= last_friday]
    return [r for r in records if d(r)]


def dimension_key(rec, dim):
    if dim == "Failure category":
        return rec["category"] or "(blank)"
    if dim == "Failure mode (detailed)":
        return rec["mode"] or "(blank)"
    if dim == "Material":
        return rec["material"] or "(blank)"
    if dim == "Recut":
        return rec["recut"] or "(blank)"
    return "(blank)"


def counts_by(records, dim):
    c = Counter()
    for r in records:
        c[dimension_key(r, dim)] += 1
    return c


def pie_slices(counts, threshold_pct):
    """Sorted (label, value) list; anything below threshold_pct rolls into 'Other'."""
    total = sum(counts.values()) or 1
    big, other = {}, 0
    for k, v in counts.items():
        if (v / total * 100.0) < threshold_pct:
            other += v
        else:
            big[k] = v
    items = sorted(big.items(), key=lambda kv: -kv[1])
    if other:
        # Avoid a literal '<' — QtCharts renders labels as rich text and treats it
        # as an unclosed HTML tag, swallowing the rest of the label ("Other (").
        items.append((f"Other (under {threshold_pct:g}%)", other))
    return items, total


def _month_label(key, all_keys):
    multi_year = len({k[0] for k in all_keys}) > 1
    y, m = key
    base = datetime.date(y, m, 1)
    return base.strftime("%b '%y") if multi_year else base.strftime("%b")


def time_buckets(records, window):
    """Return (labels, key_of_record_fn) bucketing by day for short windows else month."""
    if window in DAY_WINDOWS:
        days = sorted({r["date"].date() for r in records if r["date"]})
        labels = [d.strftime("%a %m/%d") for d in days]
        idx = {d: i for i, d in enumerate(days)}
        return labels, (lambda r: idx.get(r["date"].date()) if r["date"] else None)
    months = sorted({(r["date"].year, r["date"].month) for r in records if r["date"]})
    labels = [_month_label(m, months) for m in months]
    idx = {m: i for i, m in enumerate(months)}
    return labels, (lambda r: idx.get((r["date"].year, r["date"].month))
                    if r["date"] else None)


def bucket_totals(records, window):
    labels, key_fn = time_buckets(records, window)
    totals = [0] * len(labels)
    for r in records:
        i = key_fn(r)
        if i is not None:
            totals[i] += 1
    return labels, totals


def bucket_matrix(records, window, dim):
    """Return (labels, OrderedDict[dim_value -> [counts per bucket]]) for stacking."""
    labels, key_fn = time_buckets(records, window)
    dims = sorted({dimension_key(r, dim) for r in records})
    matrix = OrderedDict((d, [0] * len(labels)) for d in dims)
    for r in records:
        i = key_fn(r)
        if i is not None:
            matrix[dimension_key(r, dim)][i] += 1
    return labels, matrix


def linregress(ys):
    """Least-squares slope/intercept over x = 0..n-1. Flat line if <2 points."""
    n = len(ys)
    if n < 2:
        return 0.0, (ys[0] if ys else 0.0)
    xs = list(range(n))
    mx = sum(xs) / n
    my = sum(ys) / n
    denom = sum((x - mx) ** 2 for x in xs) or 1.0
    slope = sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / denom
    return slope, my - slope * mx


# ---------------------------------------------------------------------------------
# Chart builders (return a QChart; colors/bg/text are caller-supplied so the same
# code drives both the on-screen themed view and the white-background PDF)
# ---------------------------------------------------------------------------------
def _style_chart(chart, title, bg, text):
    chart.setBackgroundBrush(QBrush(QColor(bg)))
    chart.setPlotAreaBackgroundVisible(False)
    if title:
        chart.setTitle(title)
        chart.setTitleBrush(QBrush(QColor(text)))
        f = chart.titleFont()
        f.setBold(True)
        f.setPointSize(max(f.pointSize() + 1, 11))
        chart.setTitleFont(f)
    lg = chart.legend()
    lg.setLabelColor(QColor(text))
    lg.setAlignment(Qt.AlignmentFlag.AlignBottom)


def _style_axes(axes, text):
    for ax in axes:
        ax.setLabelsColor(QColor(text))
        ax.setTitleBrush(QBrush(QColor(text)))
        ax.setGridLineColor(QColor(text).lighter(180) if QColor(text).lightness() < 128
                            else QColor(text).darker(140))
        ax.setLinePenColor(QColor(text))


def build_pie(items, total, title, cycle, bg, text):
    series = QPieSeries()
    for i, (label, value) in enumerate(items):
        pct = value / (total or 1) * 100.0
        sl = series.append(f"{label}  {pct:.0f}%", float(value))
        sl.setColor(QColor(cycle[i % len(cycle)]))
        sl.setLabelColor(QColor(text))
    series.setLabelsVisible(True)
    chart = QChart()
    chart.addSeries(series)
    _style_chart(chart, title, bg, text)
    chart.legend().setVisible(True)
    return chart


def build_column(labels, totals, title, cycle, bg, text, bestfit=False):
    bar_set = QBarSet("Reworks")
    for v in totals:
        bar_set.append(float(v))
    bar_set.setColor(QColor(cycle[0]))
    series = QBarSeries()
    series.append(bar_set)

    chart = QChart()
    chart.addSeries(series)
    _style_chart(chart, title, bg, text)

    cat_axis = QBarCategoryAxis()
    cat_axis.append(labels or [""])
    chart.addAxis(cat_axis, Qt.AlignmentFlag.AlignBottom)
    series.attachAxis(cat_axis)

    ymax = max(totals, default=0) or 1
    val_axis = QValueAxis()
    val_axis.setRange(0, ymax * 1.15)
    val_axis.setLabelFormat("%d")
    chart.addAxis(val_axis, Qt.AlignmentFlag.AlignLeft)
    series.attachAxis(val_axis)

    chart.legend().setVisible(False)

    if bestfit and len(totals) >= 2:
        slope, intercept = linregress([float(t) for t in totals])
        line = QLineSeries()
        line.setName("Best fit")
        for i in range(len(totals)):
            line.append(i + 0.5, slope * i + intercept)
        pen = line.pen()
        pen.setWidth(3)
        pen.setColor(QColor("#B23A48"))
        line.setPen(pen)
        chart.addSeries(line)
        hidden_x = QValueAxis()
        hidden_x.setRange(0, len(totals))
        hidden_x.setVisible(False)
        chart.addAxis(hidden_x, Qt.AlignmentFlag.AlignBottom)
        line.attachAxis(hidden_x)
        line.attachAxis(val_axis)
        chart.legend().setVisible(True)

    _style_axes([cat_axis, val_axis], text)
    return chart


def build_stacked(labels, matrix, title, cycle, bg, text):
    series = QStackedBarSeries()
    seg_colors = []
    for i, (name, values) in enumerate(matrix.items()):
        bs = QBarSet(str(name))
        for v in values:
            bs.append(float(v))
        col = QColor(cycle[i % len(cycle)])
        bs.setColor(col)
        seg_colors.append(col)
        series.append(bs)

    chart = QChart()
    chart.addSeries(series)
    _style_chart(chart, title, bg, text)

    cat_axis = QBarCategoryAxis()
    cat_axis.append(labels or [""])
    chart.addAxis(cat_axis, Qt.AlignmentFlag.AlignBottom)
    series.attachAxis(cat_axis)

    stacked_max = max((sum(col) for col in zip(*matrix.values())), default=0) \
        if matrix else 0
    val_axis = QValueAxis()
    val_axis.setRange(0, (stacked_max or 1) * 1.15)
    val_axis.setLabelFormat("%d")
    chart.addAxis(val_axis, Qt.AlignmentFlag.AlignLeft)
    series.attachAxis(val_axis)

    chart.legend().setVisible(False)  # names are written on the segments instead
    _style_axes([cat_axis, val_axis], text)

    # QtCharts only stamps @value on a bar, so to write the GROUP NAME on each
    # segment we draw our own text in LabeledChartView.drawForeground, positioned
    # via chart.mapToPosition at each segment's vertical midpoint. Stash the metadata
    # on the chart's python wrapper for the view to read (we keep a ref so it lives).
    chart._stacked_series = series
    chart._stacked_names = list(matrix.keys())
    chart._stacked_matrix = matrix
    chart._stacked_colors = seg_colors
    return chart


def _contrast_text(bg_color):
    """Black or white, whichever reads better on bg_color (segment label color)."""
    return QColor("#FFFFFF") if bg_color.lightness() < 140 else QColor("#111111")


def _short(name, limit=20):
    name = str(name)
    return name if len(name) <= limit else name[:limit - 1] + "…"


class LabeledChartView(QChartView):
    """QChartView that writes the group name + count centered on each stacked
    segment (drawn in drawForeground so it reflows on every resize / PDF render,
    and works without a color-key legend). A no-op for non-stacked charts."""

    def drawForeground(self, painter, rect):
        super().drawForeground(painter, rect)
        chart = self.chart()
        names = getattr(chart, "_stacked_names", None)
        matrix = getattr(chart, "_stacked_matrix", None)
        series = getattr(chart, "_stacked_series", None)
        colors = getattr(chart, "_stacked_colors", None)
        if not names or not matrix or series is None:
            return
        n_buckets = len(next(iter(matrix.values()))) if matrix else 0
        painter.save()
        font = painter.font()
        font.setPointSizeF(max(7.5, font.pointSizeF()))
        painter.setFont(font)
        fm = painter.fontMetrics()
        for i in range(n_buckets):
            cum = 0.0
            for name, color in zip(names, colors):
                val = matrix[name][i]
                if val <= 0:
                    cum += val
                    continue
                top = chart.mapToPosition(QPointF(i, cum + val), series)
                bot = chart.mapToPosition(QPointF(i, cum), series)
                cum += val
                seg_px = abs(bot.y() - top.y())
                if seg_px < fm.height() + 2:  # too thin to label legibly
                    continue
                cx = (top.x() + bot.x()) / 2.0
                cy = (top.y() + bot.y()) / 2.0
                txt = f"{_short(name)} ({int(val)})"
                w = fm.horizontalAdvance(txt)
                painter.setPen(_contrast_text(QColor(color)))
                painter.drawText(QPointF(cx - w / 2.0, cy + fm.ascent() / 2.0), txt)
        painter.restore()


def empty_chart(message, bg, text):
    chart = QChart()
    _style_chart(chart, message, bg, text)
    chart.legend().setVisible(False)
    return chart


# ---------------------------------------------------------------------------------
# Log Rework dialog (the quick entry form; closes on submit)
# ---------------------------------------------------------------------------------
class LogReworkDialog(QDialog):
    def __init__(self, data_path, log=print, parent=None):
        super().__init__(parent)
        self.data_path = Path(data_path)
        self.log = log
        self.setWindowTitle("Log Rework")
        self.setMinimumWidth(440)
        if parent is not None and parent.window() is not None:
            self.setStyleSheet(parent.window().styleSheet())  # match the app theme

        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self.in_batch = QLineEdit()
        self.in_batch.setPlaceholderText("e.g. 100")
        self.in_date = QDateEdit()
        self.in_date.setCalendarPopup(True)
        self.in_date.setDisplayFormat("MM/dd/yyyy")
        self.in_date.setDate(QDate.currentDate())
        self.in_item = QLineEdit()
        self.in_item.setPlaceholderText("e.g. DYPN404")
        self.in_material = QComboBox()
        self.in_material.addItems(MATERIALS)
        self.in_recut = QComboBox()
        self.in_recut.addItems(RECUT_OPTIONS)
        self.in_category = QComboBox()
        self.in_category.addItems(list(FAILURE_CATEGORIES.keys()))
        self.in_subcategory = QComboBox()
        self.in_category.currentTextChanged.connect(self._refresh_subcategories)
        self._refresh_subcategories(self.in_category.currentText())
        self.in_comments = QLineEdit()
        self.in_comments.setPlaceholderText("Optional")

        form.addRow("Batch (PO):", self.in_batch)
        form.addRow("Date:", self.in_date)
        form.addRow("Item Number (DYPN):", self.in_item)
        form.addRow("Source Material:", self.in_material)
        form.addRow("Recut?:", self.in_recut)
        form.addRow("Failure Category:", self.in_category)
        form.addRow("Failure Subcategory:", self.in_subcategory)
        form.addRow("Comments:", self.in_comments)
        layout.addLayout(form)

        self.status = QLabel("")
        self.status.setWordWrap(True)
        layout.addWidget(self.status)

        btn_row = QHBoxLayout()
        cancel = QPushButton("Cancel")
        cancel.clicked.connect(self.reject)
        submit = QPushButton("Submit")
        submit.setDefault(True)
        submit.clicked.connect(self._submit)
        btn_row.addStretch(1)
        btn_row.addWidget(cancel)
        btn_row.addWidget(submit)
        layout.addLayout(btn_row)
        self.in_batch.setFocus()

    def _refresh_subcategories(self, category):
        self.in_subcategory.clear()
        for sub in FAILURE_CATEGORIES.get(category, []):
            self.in_subcategory.addItem(sub if sub else "(none)", sub)

    def _submit(self):
        batch = self.in_batch.text().strip()
        item = self.in_item.text().strip()
        if not batch or not item:
            self.status.setStyleSheet("color: #B23A48; font-weight: bold;")
            self.status.setText("Batch (PO) and Item Number are required.")
            return
        category = self.in_category.currentText()
        sub = self.in_subcategory.currentData()
        mode = compose_mode(category, sub if sub is not None else "")
        qd = self.in_date.date()
        record = {
            "batch": batch,
            "date": datetime.datetime(qd.year(), qd.month(), qd.day()),
            "item": item,
            "material": self.in_material.currentText(),
            "recut": self.in_recut.currentText(),
            "mode": mode,
            "comments": self.in_comments.text().strip(),
        }
        try:
            append_record(self.data_path, record, log=self.log)
        except Exception as e:
            self.status.setStyleSheet("color: #B23A48; font-weight: bold;")
            self.status.setText(
                f"Could not write (the log may be open in Excel): {e}")
            return
        self.log(f"Logged {item}: {mode} (batch {batch})")
        self.accept()   # closes the dialog; parent reloads + re-renders


# ---------------------------------------------------------------------------------
# The window content (charts dashboard — the only window)
# ---------------------------------------------------------------------------------
class QAReworkContent(QWidget):
    def __init__(self, data_path, log=print, parent=None):
        super().__init__(parent)
        self.data_path = Path(data_path)
        self.log = log
        self.records = []
        self._last_mtime = None
        self._palette = self._theme_palette()

        # Four independent graph configs; the shared control row edits the active one.
        self.graph_configs = _default_graph_configs()
        self.active_graph = 0
        self._loading = False        # guard: True while syncing controls from a config
        self._live_charts = []       # keep python refs so stacked metadata survives

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        self._build_charts_ui(layout)

        self.reload_records()

        # "Live" feel: poll the shared file's mtime and auto-refresh when it changes,
        # so entries from this laptop OR another writer flow into the charts on their
        # own. The file is a local OneDrive cache, so the stat()+reload is cheap.
        self._poll_timer = QTimer(self)
        self._poll_timer.setInterval(8000)
        self._poll_timer.timeout.connect(self._poll_file)
        self._poll_timer.start()

    # ---- theme -----------------------------------------------------------------
    def _theme_palette(self):
        try:
            from techdeck.ui.theme_manager import get_theme_manager
            return get_theme_manager().get_current_palette()
        except Exception:
            return None

    def _screen_colors(self):
        p = self._palette
        if p is None:
            return (PRINT_CYCLE, "#202020", "#EAEAEA")
        cycle = [getattr(p, f) for f in
                 ("accent", "accent_two", "success", "info", "warning", "error")
                 if hasattr(p, f)] or PRINT_CYCLE
        return (cycle, getattr(p, "surface", "#202020"),
                getattr(p, "text", "#EAEAEA"))

    # ---- view-toggle icons (drawn in code so no asset files / .spec changes) ----
    def _view_icon(self, kind, color):
        pm = QPixmap(20, 20)
        pm.fill(Qt.GlobalColor.transparent)
        p = QPainter(pm)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        pen = QPen(QColor(color))
        pen.setWidth(2)
        p.setPen(pen)
        if kind == "single":            # one monitor on a stand
            p.drawRoundedRect(3, 3, 14, 10, 2, 2)
            p.drawLine(10, 13, 10, 16)
            p.drawLine(6, 17, 14, 17)
        else:                            # 2x2 grid of screens
            for rx, ry in ((3, 3), (11, 3), (3, 11), (11, 11)):
                p.drawRoundedRect(rx, ry, 6, 6, 1, 1)
        p.end()
        return QIcon(pm)

    # ---- charts UI (the only window) -------------------------------------------
    def _build_charts_ui(self, outer):
        _, _, icon_color = self._screen_colors()

        # Top action row: log an entry / open the raw log, then the view toggle
        # (single screen | multi screen) and the "which graph am I editing" picker.
        action_row = QHBoxLayout()
        log_btn = QPushButton("➕  Log Rework")
        log_btn.clicked.connect(self._open_log_dialog)
        open_btn = QPushButton("📂  Open Rework Log")
        open_btn.clicked.connect(self._open_rework_log)
        action_row.addWidget(log_btn)
        action_row.addWidget(open_btn)
        action_row.addStretch(1)

        self.btn_single = QToolButton()
        self.btn_single.setCheckable(True)
        self.btn_single.setChecked(True)
        self.btn_single.setIcon(self._view_icon("single", icon_color))
        self.btn_single.setIconSize(QSize(20, 20))
        self.btn_single.setToolTip("Single graph")
        self.btn_multi = QToolButton()
        self.btn_multi.setCheckable(True)
        self.btn_multi.setIcon(self._view_icon("multi", icon_color))
        self.btn_multi.setIconSize(QSize(20, 20))
        self.btn_multi.setToolTip("Four graphs (2x2)")
        self._view_group = QButtonGroup(self)
        self._view_group.setExclusive(True)
        self._view_group.addButton(self.btn_single, 0)
        self._view_group.addButton(self.btn_multi, 1)
        self._view_group.idToggled.connect(self._on_view_changed)
        sep = QLabel("|")
        sep.setEnabled(False)
        action_row.addWidget(self.btn_single)
        action_row.addWidget(sep)
        action_row.addWidget(self.btn_multi)

        action_row.addSpacing(16)
        action_row.addWidget(QLabel("Graph:"))
        self.c_graph = QComboBox()
        self.c_graph.addItems(["Graph 1", "Graph 2", "Graph 3", "Graph 4"])
        self.c_graph.currentIndexChanged.connect(self._on_graph_changed)
        action_row.addWidget(self.c_graph)
        outer.addLayout(action_row)

        controls = QGridLayout()
        self.c_type = QComboBox()
        self.c_type.addItems(CHART_TYPES)
        self.c_window = QComboBox()
        self.c_window.addItems(WINDOWS)
        self.c_group = QComboBox()
        self.c_group.addItems(GROUP_BY)
        self.c_xaxis = QComboBox()
        self.c_xaxis.addItems(XAXIS_MODES)
        self.c_threshold = QDoubleSpinBox()
        self.c_threshold.setRange(0.0, 50.0)
        self.c_threshold.setValue(5.0)
        self.c_threshold.setSuffix(" %")
        self.c_bestfit = QCheckBox("Best-fit line")
        self.c_bestfit.setChecked(True)

        # Every control edits the ACTIVE graph's config, then re-renders.
        for w in (self.c_type, self.c_window, self.c_group, self.c_xaxis):
            w.currentIndexChanged.connect(self._on_controls_changed)
        self.c_threshold.valueChanged.connect(self._on_controls_changed)
        self.c_bestfit.stateChanged.connect(self._on_controls_changed)

        self.lbl_xaxis = QLabel("Column X:")
        self.lbl_threshold = QLabel("Pie hide <")
        controls.addWidget(QLabel("Chart:"), 0, 0)
        controls.addWidget(self.c_type, 0, 1)
        controls.addWidget(QLabel("Time:"), 0, 2)
        controls.addWidget(self.c_window, 0, 3)
        controls.addWidget(QLabel("Group by:"), 0, 4)
        controls.addWidget(self.c_group, 0, 5)
        controls.addWidget(self.lbl_xaxis, 1, 0)
        controls.addWidget(self.c_xaxis, 1, 1)
        controls.addWidget(self.lbl_threshold, 1, 2)
        controls.addWidget(self.c_threshold, 1, 3)
        controls.addWidget(self.c_bestfit, 1, 4, 1, 2)
        outer.addLayout(controls)

        # Page 0: the single active graph.  Page 1: a 2x2 grid of all four.
        self.view_stack = QStackedWidget()

        self.single_view = LabeledChartView()
        self.single_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.single_view.setSizePolicy(QSizePolicy.Policy.Expanding,
                                       QSizePolicy.Policy.Expanding)
        self.view_stack.addWidget(self.single_view)

        grid_page = QWidget()
        grid = QGridLayout(grid_page)
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setSpacing(8)
        self.grid_views = []
        for i in range(4):
            v = LabeledChartView()
            v.setRenderHint(QPainter.RenderHint.Antialiasing)
            v.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
            grid.addWidget(v, i // 2, i % 2)
            self.grid_views.append(v)
        self.view_stack.addWidget(grid_page)

        self.view_stack.setMinimumHeight(440)
        outer.addWidget(self.view_stack, 1)

        btn_row = QHBoxLayout()
        refresh = QPushButton("↻ Reload data")
        refresh.clicked.connect(self.reload_records)
        self.export_view_btn = QPushButton("Export view (PDF)")
        self.export_view_btn.clicked.connect(self._export_view)
        export_pack = QPushButton("Export Gemba Pack (PDF)")
        export_pack.clicked.connect(self._export_gemba_pack)
        btn_row.addWidget(refresh)
        btn_row.addStretch(1)
        btn_row.addWidget(self.export_view_btn)
        btn_row.addWidget(export_pack)
        outer.addLayout(btn_row)

        self.chart_status = QLabel("")
        outer.addWidget(self.chart_status)

        self._config_to_controls()   # load Graph 1 into the controls + set enablement

    # ---- config <-> controls sync ----------------------------------------------
    def _current_config(self):
        return self.graph_configs[self.active_graph]

    def _config_to_controls(self):
        """Load the active graph's config into the control row (no re-save churn)."""
        self._loading = True
        cfg = self._current_config()
        self.c_type.setCurrentText(cfg["type"])
        self.c_window.setCurrentText(cfg["window"])
        self.c_group.setCurrentText(cfg["group"])
        self.c_xaxis.setCurrentText(cfg["xaxis"])
        self.c_threshold.setValue(cfg["threshold"])
        self.c_bestfit.setChecked(cfg["bestfit"])
        self._loading = False
        self._update_control_enablement()

    def _controls_to_config(self):
        cfg = self._current_config()
        cfg["type"] = self.c_type.currentText()
        cfg["window"] = self.c_window.currentText()
        cfg["group"] = self.c_group.currentText()
        cfg["xaxis"] = self.c_xaxis.currentText()
        cfg["threshold"] = self.c_threshold.value()
        cfg["bestfit"] = self.c_bestfit.isChecked()

    def _on_controls_changed(self, *_):
        if self._loading:
            return
        self._controls_to_config()
        self._update_control_enablement()
        self.render_all()

    def _on_graph_changed(self, idx):
        self.active_graph = max(0, min(idx, len(self.graph_configs) - 1))
        self._config_to_controls()
        self.render_all()

    def _on_view_changed(self, _id=None, _checked=None):
        self.view_stack.setCurrentIndex(1 if self.btn_multi.isChecked() else 0)
        self.export_view_btn.setText(
            "Export 2x2 (PDF)" if self.btn_multi.isChecked() else "Export graph (PDF)")
        self.render_all()

    def _update_control_enablement(self):
        """Enable only the controls that apply to the active graph's chart type, so
        it's obvious which toggles do something (By time/By group + best-fit only
        affect Column; the %-threshold only affects Pie)."""
        ctype = self.c_type.currentText()
        is_pie = ctype == "Pie"
        is_col = ctype == "Column"
        for w in (self.lbl_xaxis, self.c_xaxis):
            w.setEnabled(is_col)
        self.c_bestfit.setEnabled(is_col and
                                  self.c_xaxis.currentText() != "By group")
        for w in (self.lbl_threshold, self.c_threshold):
            w.setEnabled(is_pie)

    def _open_log_dialog(self):
        dlg = LogReworkDialog(self.data_path, self.log, parent=self)
        if dlg.exec():               # modal; True when an entry was submitted
            self.reload_records()    # charts update immediately

    def _open_rework_log(self):
        try:
            ensure_seeded(self.data_path, self.log)  # create if somehow missing
            os.startfile(str(self.data_path))        # opens in Excel on Windows
            self.chart_status.setText(
                "Opened the rework log in Excel. Edits you save there refresh "
                "the charts automatically.")
        except Exception as e:
            self.chart_status.setText(f"Could not open the log: {e}")

    def reload_records(self):
        try:
            self.records = load_records(self.data_path, log=self.log)
            self._last_mtime = self._current_mtime()
        except Exception as e:
            self.records = []
            self.log(f"Could not read the log: {e}")
            if hasattr(self, "chart_status"):
                self.chart_status.setText(f"Could not read the log: {e}")
        if hasattr(self, "view_stack"):
            self.render_all()

    def _current_mtime(self):
        try:
            return self.data_path.stat().st_mtime
        except OSError:
            return None

    def _poll_file(self):
        """Reload only when the shared workbook actually changed on disk."""
        mtime = self._current_mtime()
        if mtime is not None and mtime != self._last_mtime:
            self.log("QA rework log changed on disk - refreshing charts.")
            self.reload_records()

    # ---- rendering -------------------------------------------------------------
    def render_all(self):
        """Render the active view: one chart in single mode, all four in 2x2 mode.
        We hold python refs to the live charts so their stacked-label metadata is not
        garbage-collected out from under the view."""
        if not hasattr(self, "view_stack"):
            return
        cycle, bg, text = self._screen_colors()
        live = []
        if self.btn_multi.isChecked():
            for cfg, view in zip(self.graph_configs, self.grid_views):
                chart = self._build_chart(cfg, cycle, bg, text)
                self._swap_chart(view, chart)
                live.append(chart)
        else:
            chart = self._build_chart(self._current_config(), cycle, bg, text)
            self._swap_chart(self.single_view, chart)
            live.append(chart)
        self._live_charts = live

        cfg = self._current_config()
        n = len(filter_window(self.records, cfg["window"]))
        scope = "4 graphs" if self.btn_multi.isChecked() else f"Graph {self.active_graph + 1}"
        self.chart_status.setText(
            f"{scope} · {n} rework events in '{cfg['window']}' "
            f"({len(self.records)} total logged).")

    def _swap_chart(self, view, chart):
        old = view.chart()
        view.setChart(chart)
        if old is not None:
            old.deleteLater()

    def _build_chart(self, cfg, cycle, bg, text):
        """Build one QChart from a graph config dict (no widget reads, so it drives
        the on-screen single/grid views AND the PDF exports identically)."""
        window = cfg["window"]
        group = cfg["group"]
        recs = filter_window(self.records, window)
        if not recs:
            return empty_chart("No data in this time window yet.", bg, text)

        ctype = cfg["type"]
        if ctype == "Pie":
            items, total = pie_slices(counts_by(recs, group), cfg["threshold"])
            return build_pie(items, total, f"{group} — {window}", cycle, bg, text)

        if ctype == "Stacked column":
            labels, matrix = bucket_matrix(recs, window, group)
            return build_stacked(labels, matrix,
                                 f"{group} by {'day' if window in DAY_WINDOWS else 'month'} — {window}",
                                 cycle, bg, text)

        # Column
        if cfg["xaxis"] == "By group":
            counts = counts_by(recs, group)
            ordered = sorted(counts.items(), key=lambda kv: -kv[1])
            labels = [k for k, _ in ordered]
            totals = [v for _, v in ordered]
            return build_column(labels, totals, f"{group} — {window}", cycle, bg, text,
                                bestfit=False)
        labels, totals = bucket_totals(recs, window)
        return build_column(labels, totals, f"Reworks by "
                            f"{'day' if window in DAY_WINDOWS else 'month'} — {window}",
                            cycle, bg, text, bestfit=cfg["bestfit"])

    # ---- PDF export ------------------------------------------------------------
    def _export_view(self):
        """Export whatever is on screen: the single active graph, or the 2x2 board
        as one landscape page. Charts are rebuilt with the white print palette."""
        multi = self.btn_multi.isChecked()
        default = "QA Rework 2x2.pdf" if multi else "QA Rework Chart.pdf"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save view as PDF", str(Path.home() / default), "PDF files (*.pdf)")
        if not path:
            return
        cy, bg, tx = PRINT_CYCLE, PRINT_BG, PRINT_TEXT
        try:
            if multi:
                charts = [self._build_chart(c, cy, bg, tx) for c in self.graph_configs]
                self._write_pdf_grid(path, "QA Rework — Gemba Board", charts, cols=2)
            else:
                chart = self._build_chart(self._current_config(), cy, bg, tx)
                self._write_pdf(path, [("", [chart])])
            self.chart_status.setText(f"Saved {path}")
        except Exception as e:
            self.chart_status.setText(f"Export failed: {e}")

    def _export_gemba_pack(self):
        if not self.records:
            self.chart_status.setText("Nothing to export yet — log some data first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Gemba Pack PDF",
            str(Path.home() / "QA Gemba Pack.pdf"), "PDF files (*.pdf)")
        if not path:
            return
        try:
            self._write_pdf(path, self._gemba_pages())
            self.chart_status.setText(f"Saved Gemba Pack: {path}")
            self.log(f"Exported Gemba Pack to {path}")
        except Exception as e:
            self.chart_status.setText(f"Export failed: {e}")

    def _gemba_pages(self):
        """The four standard meeting charts, two per page."""
        cy, bg, tx = PRINT_CYCLE, PRINT_BG, PRINT_TEXT
        ytd = filter_window(self.records, "Year-to-date")
        lastwk = filter_window(self.records, "Last business week (Mon-Fri)")

        # Page 1 - Year to date
        ytd_labels, ytd_totals = bucket_totals(ytd, "Year-to-date")
        col = build_column(ytd_labels, ytd_totals, "YTD Reworks by Month",
                           cy, bg, tx, bestfit=True) if ytd \
            else empty_chart("No YTD data", bg, tx)
        items, total = pie_slices(counts_by(ytd, "Failure category"), 5.0)
        ytd_pie = build_pie(items, total, "YTD Failure Category %", cy, bg, tx) if ytd \
            else empty_chart("No YTD data", bg, tx)

        # Page 2 - Last business week
        if lastwk:
            labels, matrix = bucket_matrix(lastwk, "Last business week (Mon-Fri)",
                                           "Failure mode (detailed)")
            wk_col = build_stacked(labels, matrix, "Last Week Reworks by Failure Mode",
                                   cy, bg, tx)
            it2, tot2 = pie_slices(counts_by(lastwk, "Failure mode (detailed)"), 5.0)
            wk_pie = build_pie(it2, tot2, "Last Week Failure Mode %", cy, bg, tx)
        else:
            wk_col = empty_chart("No data for last business week", bg, tx)
            wk_pie = empty_chart("No data for last business week", bg, tx)

        return [("Year-to-Date", [col, ytd_pie]),
                ("Last Week", [wk_col, wk_pie])]

    def _write_pdf(self, path, pages):
        """pages = list of (page_title, [charts]); charts laid out stacked per page."""
        writer = QPdfWriter(str(path))
        writer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
        writer.setPageOrientation(QPageLayout.Orientation.Landscape)
        writer.setResolution(150)
        painter = QPainter(writer)
        try:
            for pi, (title, charts) in enumerate(pages):
                if pi > 0:
                    writer.newPage()
                self._paint_page(painter, writer, title, charts)
        finally:
            painter.end()

    def _write_pdf_grid(self, path, title, charts, cols=2):
        """Write all charts onto ONE landscape page in a `cols`-wide grid (used for
        the 2x2 view export)."""
        writer = QPdfWriter(str(path))
        writer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
        writer.setPageOrientation(QPageLayout.Orientation.Landscape)
        writer.setResolution(150)
        painter = QPainter(writer)
        try:
            self._paint_grid_page(painter, title, charts, cols)
        finally:
            painter.end()

    def _paint_grid_page(self, painter, title, charts, cols):
        page = painter.viewport()
        margin = int(min(page.width(), page.height()) * 0.04)
        x = page.left() + margin
        y = page.top() + margin
        w = page.width() - 2 * margin
        h = page.height() - 2 * margin

        title_h = 0
        if title:
            font = QFont()
            font.setBold(True)
            font.setPointSize(16)
            painter.setFont(font)
            painter.setPen(QColor(PRINT_TEXT))
            title_h = int(h * 0.06)
            painter.drawText(x, y, w, title_h,
                             int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
                             title)

        charts = [c for c in charts if c is not None]
        if not charts:
            return
        rows = (len(charts) + cols - 1) // cols
        gap = margin // 2
        cell_w = (w - gap * (cols - 1)) // cols
        cell_h = (h - title_h - gap * (rows - 1)) // rows
        for idx, chart in enumerate(charts):
            r, c = divmod(idx, cols)
            cx = x + c * (cell_w + gap)
            cy = y + title_h + r * (cell_h + gap)
            self._render_chart_into(painter, chart, cx, cy, cell_w, cell_h)

    def _paint_page(self, painter, writer, title, charts):
        page = painter.viewport()
        margin = int(min(page.width(), page.height()) * 0.04)
        x = page.left() + margin
        y = page.top() + margin
        w = page.width() - 2 * margin
        h = page.height() - 2 * margin

        title_h = 0
        if title:
            font = QFont()
            font.setBold(True)
            font.setPointSize(16)
            painter.setFont(font)
            painter.setPen(QColor(PRINT_TEXT))
            title_h = int(h * 0.07)
            painter.drawText(x, y, w, title_h,
                             int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
                             title)

        charts = [c for c in charts if c is not None]
        if not charts:
            return
        avail_h = h - title_h
        gap = margin // 2
        each_h = (avail_h - gap * (len(charts) - 1)) // len(charts)
        cy = y + title_h
        for chart in charts:
            self._render_chart_into(painter, chart, x, cy, w, each_h)
            cy += each_h + gap

    def _render_chart_into(self, painter, chart, x, y, w, h):
        view = LabeledChartView()
        view.setChart(chart)
        view.setRenderHint(QPainter.RenderHint.Antialiasing)
        view.setStyleSheet("background: white; border: none;")
        view.resize(w, h)
        view.setAttribute(Qt.WidgetAttribute.WA_DontShowOnScreen, True)
        view.show()
        painter.save()
        painter.translate(x, y)
        view.render(painter)
        painter.restore()
        view.hide()
        view.setChart(QChart())  # detach so deleteLater doesn't take our chart
        view.deleteLater()


# ---------------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------------
def run(params, progress_callback, cancel_event):
    global _window
    log = params.get("log", print)
    settings = params.get("settings", {})
    try:
        data_path = resolve_data_path(settings, log)
    except Exception as e:
        log(f"Could not resolve the QA rework data folder: {e}")
        return
    log(f"QA Rework log file: {data_path}")
    try:
        ensure_seeded(data_path, log)
    except Exception as e:
        log(f"Seeding skipped: {e}")

    _window = PluginWindow("qa_rework", "Gemba Analysis")
    content = QAReworkContent(data_path, log)
    _window.set_content(content)
    _window.resize(1120, 780)
    _window.show()
    progress_callback(100)
