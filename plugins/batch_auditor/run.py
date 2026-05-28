"""
Batch Auditor
=============
Read-only "is this batch ready for the floor?" check for a 911 or 922 batch.
The user chooses which line to audit; it never audits both in one run.

Results are presented two ways:
  - a dashboard (KPI cards + pie/bar charts) on the console's Dashboard tab, and
  - a concise text summary on the Console tab.

Nothing is written to the batch — this plugin only reads.
"""
from __future__ import annotations

import re
import threading
from pathlib import Path

import openpyxl

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk


_NEST_RE = re.compile(r'^[PS]?\d{3,}$', re.IGNORECASE)


# ── status helpers ──────────────────────────────────────────────────────────

def _ratio_status(done: int, total: int) -> str:
    if total == 0:
        return "info"
    if done >= total:
        return "ok"
    if done >= total * 0.5:
        return "warning"
    return "error"


def _bool_status(ok: bool) -> str:
    return "ok" if ok else "error"


# ── small workbook readers (read-only) ──────────────────────────────────────

def _count_bend_notes(po_path: Path) -> int:
    wb = openpyxl.load_workbook(po_path, data_only=True)
    try:
        ws = wb['PO'] if 'PO' in wb.sheetnames else wb.active
        hr, cols = sdk.find_header_row(ws, ['DYPN', 'NOTES'])
        if hr is None or 'NOTES' not in cols:
            return 0
        c = cols['NOTES']
        n = 0
        for r in range(hr + 1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if v and 'bend' in str(v).lower():
                n += 1
        return n
    finally:
        wb.close()


def _bent_plates_rows(organizer_path: Path) -> int:
    wb = openpyxl.load_workbook(organizer_path, data_only=True)
    try:
        if 'Bent Plates' not in wb.sheetnames:
            return 0
        ws = wb['Bent Plates']
        hr, cols = sdk.find_header_row(ws, ['DYPN'])
        if hr is None:
            return 0
        c = cols['DYPN']
        return sum(1 for r in range(hr + 1, ws.max_row + 1)
                   if ws.cell(r, c).value not in (None, ''))
    finally:
        wb.close()


def _pallet_assignment_count(organizer_path: Path) -> int:
    """Orders assigned to pallets in the 'Pallet Organizer' sheet. Layout
    mirrors 922_pallet_stamper: orders live in columns B/E/H from row 5 down."""
    wb = openpyxl.load_workbook(organizer_path, data_only=True)
    try:
        if 'Pallet Organizer' not in wb.sheetnames:
            return 0
        ws = wb['Pallet Organizer']
        n = 0
        for r in range(5, 24):
            for c in (2, 5, 8):
                if ws.cell(r, c).value not in (None, ''):
                    n += 1
        return n
    finally:
        wb.close()


def _order_has_7000_prints(order_dir: Path) -> bool:
    """True if the order has a CAD-AND-SHOP-PRINTS/.../7000/ folder with PDFs."""
    for p in order_dir.rglob("7000"):
        if not p.is_dir():
            continue
        if "cad-and-shop-prints" in [s.lower() for s in p.parts]:
            if any(f.suffix.lower() == ".pdf" for f in p.glob("*.pdf")):
                return True
    return False


# ── 922 audit ───────────────────────────────────────────────────────────────

def _audit_922(batch_no: str, batch_path: Path, log):
    doc_folder = batch_path / f"Batch {batch_no} - Documentation"

    po_matches = [p for p in doc_folder.glob(f"PO H{batch_no} QF-QU-09 REV C*.xlsx")
                  if not p.name.startswith("~")]
    organizer_matches = [p for p in doc_folder.glob(f"PO H{batch_no} Pallet & Rod Organizer*.xlsx")
                         if not p.name.startswith("~")]
    po_path = sorted(po_matches)[0] if po_matches else None
    organizer_path = sorted(organizer_matches)[0] if organizer_matches else None

    cards: list[dict] = []
    summary: list[str] = []

    # --- Orders & prints ---
    total_orders = folders_present = laser_orders = prints_present = 0
    missing_folders = missing_prints = 0
    if po_path is None:
        cards.append({"label": "Orders", "value": "No PO", "status": "error"})
        summary.append("ERROR: PO QF-QU-09 workbook not found - cannot audit orders.")
    else:
        order_to_serials, _ = sdk.read_qf_qu_09(po_path)
        total_orders = len(order_to_serials)
        children = [c for c in batch_path.iterdir() if c.is_dir()]

        def _folder_for(order):
            for c in children:
                if c.name == order or c.name.startswith(order + "-"):
                    return c
            return None

        for order, serials in order_to_serials.items():
            folder = _folder_for(order)
            if folder is None:
                missing_folders += 1
                continue
            folders_present += 1
            is_laser = any(sdk.tube_class(s) == "standard" for s in serials)
            if is_laser:
                laser_orders += 1
                if _order_has_7000_prints(folder):
                    prints_present += 1
                else:
                    missing_prints += 1

        cards.append({"label": "Order folders", "value": f"{folders_present}/{total_orders}",
                      "status": _ratio_status(folders_present, total_orders)})
        cards.append({"label": "Laser prints", "value": f"{prints_present}/{laser_orders}",
                      "status": _ratio_status(prints_present, laser_orders)})
        summary.append(f"Order folders: {folders_present}/{total_orders} present "
                       f"({missing_folders} missing).")
        summary.append(f"Laser-tube shop prints: {prints_present}/{laser_orders} present "
                       f"({missing_prints} missing).")

    # --- LST & run time ---
    lst_dir = doc_folder / "LST"
    lst_ok = lst_dir.exists() and any(lst_dir.rglob("*.lst"))
    runtime_ok = (lst_dir / f"Run Time Estimate - Batch {batch_no}.txt").exists()
    cards.append({"label": "LST organized", "value": "Done" if lst_ok else "Missing",
                  "status": _bool_status(lst_ok)})
    cards.append({"label": "Run time est.", "value": "Done" if runtime_ok else "Missing",
                  "status": _bool_status(runtime_ok)})
    summary.append(f"LST organized: {'yes' if lst_ok else 'NO'}.")
    summary.append(f"Run time estimate: {'present' if runtime_ok else 'MISSING'}.")

    # --- Forming ---
    # The deliverable that feeds Kitting is the populated Bent Plates sheet, so
    # that's the primary "done" signal. The Forming {n} folder of copied PDFs is
    # a secondary artifact (often cleaned up after merge), so its absence does
    # NOT by itself mean forming is incomplete.
    bend_count = _count_bend_notes(po_path) if po_path else 0
    bent_rows = _bent_plates_rows(organizer_path) if organizer_path else 0
    forming_issue = 0
    if bend_count == 0:
        cards.append({"label": "Forming", "value": "None needed", "status": "info"})
        summary.append("Forming: no 'bend' parts in PO - none expected.")
    elif bent_rows >= bend_count:
        cards.append({"label": "Forming", "value": "Done", "status": "ok"})
        summary.append(f"Forming: done ({bent_rows} bent plates for {bend_count} bend notes).")
    elif bent_rows > 0:
        forming_issue = 1
        cards.append({"label": "Forming", "value": f"{bent_rows}/{bend_count}", "status": "warning"})
        summary.append(f"Forming: PARTIAL - {bent_rows} bent plates for {bend_count} bend notes.")
    else:
        forming_issue = 1
        cards.append({"label": "Forming", "value": "Incomplete", "status": "error"})
        summary.append(f"Forming: INCOMPLETE - {bend_count} bend notes but Bent Plates sheet empty.")

    # --- Pallets & kitting ---
    pallets = _pallet_assignment_count(organizer_path) if organizer_path else 0
    pallets_ok = pallets > 0
    kitting_ok = (doc_folder / "Kitting" / f"Kitting {batch_no}.pdf").exists()
    cards.append({"label": "Pallets", "value": "Assigned" if pallets_ok else "Not set",
                  "status": "ok" if pallets_ok else "warning"})
    cards.append({"label": "Kitting PDF", "value": "Done" if kitting_ok else "Missing",
                  "status": "ok" if kitting_ok else "warning"})
    summary.append(f"Pallet assignments: {'set' if pallets_ok else 'NOT set'}.")
    summary.append(f"Kitting PDF: {'present' if kitting_ok else 'MISSING'}.")

    # --- charts ---
    ready = max(total_orders - missing_folders - missing_prints, 0)
    pie_slices = [s for s in [
        ("Ready", ready, "ok"),
        ("Missing prints", missing_prints, "warning"),
        ("Missing folder", missing_folders, "error"),
    ] if s[1] > 0]

    bar = {
        "type": "bar", "title": "Issues by category",
        "categories": ["Folders", "Prints", "LST", "Run time", "Forming", "Pallets", "Kitting"],
        "values": [missing_folders, missing_prints,
                   0 if lst_ok else 1, 0 if runtime_ok else 1,
                   forming_issue, 0 if pallets_ok else 1, 0 if kitting_ok else 1],
    }
    charts = []
    if pie_slices:
        charts.append({"type": "pie", "title": "Orders", "slices": pie_slices})
    charts.append(bar)

    return cards, charts, summary


# ── 911 audit (first pass) ───────────────────────────────────────────────────

def _audit_911(batch: str, batch_folder: Path, log):
    cards: list[dict] = []
    summary: list[str] = []

    batch_list = batch_folder / f"{batch} BATCH LIST.xlsx"
    if not batch_list.exists():
        for f in batch_folder.glob("*.xlsx"):
            if "BATCH LIST" in f.name.upper() and not f.name.startswith("~"):
                batch_list = f
                break

    nests: list[str] = []
    if batch_list.exists():
        wb = openpyxl.load_workbook(batch_list, data_only=True)
        try:
            ws = wb["BATCH"] if "BATCH" in wb.sheetnames else wb.active
            col = sdk.find_header_col(ws, "Nest Pkg Nbr", header_row=3)
            seen = set()
            if col:
                for r in range(4, ws.max_row + 1):
                    v = ws.cell(r, col).value
                    if v is not None:
                        s = str(v).strip()
                        if s and _NEST_RE.match(s) and s not in seen:
                            seen.add(s)
                            nests.append(s)
        finally:
            wb.close()

    total = len(nests)
    folders = workbooks = tickets = 0
    for nest in nests:
        nest_dir = batch_folder / nest
        if not nest_dir.is_dir():
            continue
        folders += 1
        if (nest_dir / f"911 BATCH {batch} {nest}.xlsx").exists() or \
           any(p.name.lower().startswith("911 batch") for p in nest_dir.glob("*.xlsx")):
            workbooks += 1
        if (nest_dir / f"{nest} MOVE TICKET OMIT.pdf").exists() or \
           any("move ticket omit" in p.name.lower() for p in nest_dir.glob("*.pdf")):
            tickets += 1

    if total == 0:
        cards.append({"label": "Nests", "value": "None found", "status": "error"})
        summary.append("ERROR: No nests found in BATCH LIST - has 911 Setup run?")
    else:
        cards.append({"label": "Nest folders", "value": f"{folders}/{total}",
                      "status": _ratio_status(folders, total)})
        cards.append({"label": "Nest workbooks", "value": f"{workbooks}/{total}",
                      "status": _ratio_status(workbooks, total)})
        cards.append({"label": "Move tickets", "value": f"{tickets}/{total}",
                      "status": _ratio_status(tickets, total)})
        summary.append(f"Nest folders: {folders}/{total}.")
        summary.append(f"Nest workbooks: {workbooks}/{total}.")
        summary.append(f"Move Ticket Omit PDFs: {tickets}/{total}.")

    charts = []
    if total:
        complete = min(folders, workbooks, tickets)
        incomplete = total - complete
        slices = [s for s in [("Complete", complete, "ok"),
                              ("Incomplete", incomplete, "error")] if s[1] > 0]
        if slices:
            charts.append({"type": "pie", "title": "Nests", "slices": slices})
        charts.append({
            "type": "bar", "title": "Missing by category",
            "categories": ["Folders", "Workbooks", "Move tickets"],
            "values": [total - folders, total - workbooks, total - tickets],
        })
    return cards, charts, summary


# ── entry point ───────────────────────────────────────────────────────────────

def run(params: dict, progress_callback, cancel_event: threading.Event) -> None:
    log = params.get("log", print)
    console = params.get("console")

    log("Starting Batch Auditor...")
    progress_callback(0)

    # Not a batch number prompt — use request_text so the answer doesn't get
    # cached into params['shared_state'][family]['batch_number'] and silently
    # filled in for the real batch prompt below.
    raw_line = sdk.request_text(params, "Which line to audit? Enter 911 or 922:")
    line = "911" if "911" in (raw_line or "") else "922" if "922" in (raw_line or "") else None
    if line is None:
        raise ValueError(f"Unrecognised line: {raw_line!r} (expected 911 or 922)")

    raw_batch = sdk.request_batch_number(params, "Enter batch number:")
    progress_callback(10)

    if line == "922":
        batch_no = sdk.parse_922_batch(raw_batch or "")
        if not batch_no:
            raise ValueError(f"Unrecognised 922 batch: {raw_batch!r}")
        root = sdk.resolve_922_root()
        if not root:
            raise RuntimeError("Could not locate '922 QTDR Production Packages'. Verify OneDrive sync.")
        batch_path = sdk.find_922_batch_path(root, batch_no)
        if not batch_path:
            raise RuntimeError(f"Batch {batch_no} not found under {root} (also checked '1 - Completed').")
        log(f"Auditing 922 Batch {batch_no}: {batch_path}")
        progress_callback(20)
        cards, charts, summary = _audit_922(batch_no, batch_path, log)
        title = f"922 Batch {batch_no} - Readiness"
        label = f"922 Batch {batch_no}"
    else:
        batch = sdk.normalize_911_batch(raw_batch or "")
        if not batch:
            raise ValueError("No 911 batch entered.")
        root = sdk.resolve_911_qtdr_root()
        if not root:
            raise RuntimeError("Could not locate '911 QTDR'. Verify OneDrive sync.")
        batch_folder = sdk.find_911_batch_folder(root, batch)
        if not batch_folder:
            raise RuntimeError(f"911 batch folder '{batch}' not found under {root}.")
        log(f"Auditing 911 Batch {batch}: {batch_folder}")
        progress_callback(20)
        cards, charts, summary = _audit_911(batch, batch_folder, log)
        title = f"911 Batch {batch} - Readiness"
        label = f"911 Batch {batch}"

    if cancel_event.is_set():
        log("Cancelled.")
        return

    progress_callback(90)

    # Dashboard
    if console is not None and hasattr(console, "show_dashboard"):
        console.show_dashboard({"title": title, "cards": cards, "charts": charts})

    # Text summary
    log("=" * 50)
    log(f"Batch Auditor - {label}")
    log("=" * 50)
    for line_text in summary:
        log(line_text)
    issues = sum(1 for c in cards if c.get("status") == "error") + \
        sum(1 for c in cards if c.get("status") == "warning")
    log("-" * 50)
    log("Batch looks ready." if issues == 0 else f"{issues} area(s) need attention (see dashboard).")
    log("=" * 50)
    progress_callback(100)


if __name__ == "__main__":
    import sys
    cancel = threading.Event()
    _line = input("Line (911/922): ").strip()
    _batch = input("Batch number: ").strip()

    class _Console:
        def request_input(self, prompt):
            # feed the two prompts in order
            if "line" in prompt.lower() or "911 or 922" in prompt.lower():
                return _line
            return _batch
        def show_dashboard(self, spec):
            print("[dashboard]", spec.get("title"))
            for c in spec.get("cards", []):
                print("  card:", c)

    run({"log": print, "settings": {}, "console": _Console()},
        lambda p: print(f"[{p}%]"), cancel)
