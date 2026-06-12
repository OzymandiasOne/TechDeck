"""
911 Sketch Extractor Plugin for TechDeck v2.3.0
Extracts PART SKETCH data from PDF packets into Excel workbook.

NEW in v2.3.0:
- Updated to 17-column structure (A-Q) matching new Excel format
- Combined PART_WT and CALC_WT into single WEIGHT column (I)
- Applied 2-decimal rounding to LENGTH (H) and WEIGHT (I) columns
- Added blank placeholder columns: WIDTH (G), CATALOG (M), CATEGORY (N), CUSTOMER/SUPPLIER (O), DIVISION 1 (P), DIVISION 2 (Q)
- Renamed columns: PART -> DYPN, SRCE -> SOURCE, FAB_DIM -> LENGTH, MIL_S -> MIL SPEC, PacketName -> NEST
- Organized output folder structure (creates folder named after Excel file)
- Moved ORDER column from J to A

Previous updates:
- Fixed ORDER column extraction to capture all barcode formats
- Updated regex to match [LETTER][LETTER OR DIGIT][6 DIGITS] pattern
- Strip .pdf extension from PacketName

Prompts via console for:
- Folder path containing PDFs  
- Whether to extract drawings to combined PDF
- Output Excel filename
"""

import os
import re
import errno
import time
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
from collections import defaultdict
import threading

# Third-party imports
import fitz  # PyMuPDF
from pypdf import PdfWriter, PdfReader
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk


# ===== CONSTANTS =====
VERSION = "2.3.0"
FIELD_LABELS = {"SRCE", "MATL", "DESC", "SIZE", "PART", "FAB", "MIL-SPEC", "MIL-S", "PRT WT", "NOUN"}

# NEW 17-column structure (A-Q)
HEADERS = [
    "ORDER",              # A - barcode number (moved from column J)
    "DYPN",               # B - part number (renamed from PART)
    "SOURCE",             # C - source (renamed from SRCE)
    "MATL",               # D - material
    "DESC",               # E - description
    "SIZE",               # F - size
    "WIDTH",              # G - NEW blank placeholder
    "LENGTH",             # H - length (renamed from FAB_DIM, 2 decimals)
    "WEIGHT",             # I - combined weight (was PART_WT + CALC_WT, 2 decimals)
    "MIL SPEC",           # J - mil spec (renamed from MIL_S)
    "NEST",               # K - packet name (renamed from PacketName)
    "PageNumber",         # L - page number
    "CATALOG",            # M - NEW blank placeholder
    "CATEGORY",           # N - NEW blank placeholder
    "CUSTOMER/SUPPLIER",  # O - NEW blank placeholder
    "DIVISION 1",         # P - NEW blank placeholder
    "DIVISION 2"          # Q - NEW blank placeholder
]


# ===== HELPER FUNCTIONS =====
def get_console_input(params: Dict[str, Any], prompt: str) -> str:
    """Get input from user via TechDeck console."""
    console = params.get('console')
    if console and hasattr(console, 'request_input'):
        return console.request_input(prompt)
    else:
        return input(f"{prompt}: ")


def retry_file_op(fn, *args, retries=3, delay=0.1, **kwargs):
    """Retry file operations for Windows file locking issues."""
    for attempt in range(retries):
        try:
            return fn(*args, **kwargs)
        except OSError as e:
            if e.errno in (errno.EACCES, errno.EPERM, errno.EBUSY) and attempt < retries - 1:
                time.sleep(delay)
                continue
            raise


# ===== PARSING LOGIC =====
def is_field_label(line: str) -> bool:
    """Check if a line starts with a known field label."""
    line_upper = line.upper().strip()
    return any(line_upper.startswith(label) for label in FIELD_LABELS)


def extract_field_value(line: str, idx: int, lines: List[str], field_name: str) -> Optional[str]:
    """Extract field value from current line or next line if empty."""
    parts = line.split(":", 1)
    if len(parts) > 1 and parts[1].strip():
        return parts[1].strip()
    
    if idx + 1 < len(lines):
        next_line = lines[idx + 1].strip()
        if next_line and not is_field_label(next_line):
            return next_line
    
    return None


def extract_numeric_from_fab_dim(fab_dim_text: str) -> Optional[str]:
    """
    Extract numeric value from FAB DIM field.
    
    Examples:
    - "130 CUT NEAT AND STAMP PER WORK PACKAGE PART NOMECLATURE." -> "130"
    - "24.5" -> "24.5"
    - "19.13 INCHES" -> "19.13"
    
    Returns the numeric value as a string, or None if no number found.
    """
    if not fab_dim_text:
        return None
    
    # Look for number at the start of the string (handles both int and float)
    match = re.match(r'^(\d+\.?\d*)', fab_dim_text.strip())
    if match:
        return match.group(1)
    
    return None


def extract_order_number(text: str) -> Optional[str]:
    """
    Extract ORDER number (barcode) from PART SKETCH page.
    Looks for patterns like X6385131, BK400428, BL318989, etc.
    Format: 8 characters - [LETTER][LETTER OR DIGIT][6 DIGITS]
    These appear near the top left under a barcode on PART SKETCH pages.
    """
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    
    # Look for pattern: 1 letter + (1 letter OR 1 digit) + 6 digits (e.g., X6385131, BK400428, BL318989)
    # Check first 30 lines (top of page where barcode appears)
    for line in lines[:30]:
        # Try exact match first (most common)
        match = re.match(r'^([A-Z][A-Z0-9]\d{6})$', line)
        if match:
            return match.group(1)
        
        # Also check if it's part of a longer line (fallback)
        match = re.search(r'\b([A-Z][A-Z0-9]\d{6})\b', line)
        if match:
            # Make sure it's not part of "PART SKETCH" heading or similar
            if 'SKETCH' not in line and 'PART:' not in line:
                return match.group(1)
    
    return None


def extract_lb_per_foot(size_text: str) -> float:
    """Extract pounds per foot from SIZE field."""
    if not size_text:
        return 0.0
    
    match = re.search(r'X?([\d.]+)LB\s*$', size_text, re.IGNORECASE)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return 0.0
    return 0.0


def parse_part_sketch_page(text: str, packet_name: str, page_number: int) -> Tuple:
    """
    Parse PART SKETCH page and extract fields in new 17-column structure.
    
    Returns tuple with 17 values (A-Q):
    ORDER, DYPN, SOURCE, MATL, DESC, SIZE, WIDTH, LENGTH, WEIGHT, MIL SPEC,
    NEST, PageNumber, CATALOG, CATEGORY, CUSTOMER/SUPPLIER, DIVISION 1, DIVISION 2
    """
    srce = matl = desc = size = part = fab_dim = part_wt = mils = order_num = None
    noun = None
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    
    # Extract ORDER number first
    order_num = extract_order_number(text)

    for idx, line in enumerate(lines):
        line_upper = line.upper()
        
        if line.startswith("SRCE"):
            srce = extract_field_value(line, idx, lines, "SRCE")
        elif line.startswith("MATL"):
            matl = extract_field_value(line, idx, lines, "MATL")
        elif line.startswith("NOUN"):
            noun = extract_field_value(line, idx, lines, "NOUN")
        elif line.startswith("DESC"):
            desc = extract_field_value(line, idx, lines, "DESC")
        elif line.startswith("SIZE"):
            size = extract_field_value(line, idx, lines, "SIZE")
        elif line.startswith("PART"):
            if "SKETCH" not in line_upper:
                parts = line.split(":", 1)
                if len(parts) > 1 and parts[1].strip():
                    value = parts[1].strip()
                    if " REV" in value.upper():
                        value = value.split(" REV")[0].strip()
                    part = value
        elif line_upper.startswith("PRT WT"):
            if idx + 1 < len(lines):
                next_line = lines[idx + 1].strip()
                if next_line and not is_field_label(next_line):
                    weight_match = re.match(r"(\d+\.?\d*)\s*LB", next_line, re.IGNORECASE)
                    if weight_match:
                        part_wt = weight_match.group(1)
        elif line_upper.startswith("FAB DIM") or line_upper.startswith("FABDIM") or line_upper.startswith("FAB-DIM"):
            raw_fab_dim = extract_field_value(line, idx, lines, "FAB DIM")
            # Extract numeric value from FAB DIM (handles "130 CUT NEAT AND STAMP..." -> "130")
            if raw_fab_dim:
                fab_dim = extract_numeric_from_fab_dim(raw_fab_dim) or raw_fab_dim
        elif line.startswith("MIL-SPEC") or line.startswith("MIL-S"):
            for j in range(1, 6):
                if idx + j < len(lines):
                    cand = lines[idx + j].strip()
                    if re.match(r"[A-Z]+[-][A-Z0-9]+", cand):
                        mils = cand
                        break

    # Only include PART_WT if this is a plate
    if not (noun and ("PLATE" in noun.upper() or "PLT" in noun.upper())):
        part_wt = None

    # === APPLY 2-DECIMAL ROUNDING TO LENGTH (H) ===
    length_rounded = ""
    if fab_dim:
        try:
            length_rounded = f"{float(fab_dim):.2f}"
        except ValueError:
            length_rounded = fab_dim or ""
    
    # === COMBINE WEIGHT LOGIC WITH 2-DECIMAL ROUNDING (I) ===
    weight_combined = ""
    lb_per_foot = extract_lb_per_foot(size or "")
    
    # Priority 1: Calculate weight if we have the data to do so
    if fab_dim and lb_per_foot > 0:
        try:
            fab_dim_val = float(fab_dim)
            calculated = (fab_dim_val / 12.0) * lb_per_foot
            weight_combined = f"{calculated:.2f}"
        except ValueError:
            # If calculation fails, try to use extracted weight
            if part_wt and part_wt.strip():
                try:
                    weight_combined = f"{float(part_wt):.2f}"
                except ValueError:
                    weight_combined = part_wt
    # Priority 2: Use extracted part_wt if calculation not possible
    elif part_wt and part_wt.strip():
        try:
            weight_combined = f"{float(part_wt):.2f}"
        except ValueError:
            weight_combined = part_wt
    # Priority 3: Leave blank if neither available
    else:
        weight_combined = ""

    # Return 17-column tuple matching new structure (A-Q)
    return (
        order_num or "",      # A: ORDER
        part or "",           # B: DYPN
        srce or "",           # C: SOURCE
        matl or "",           # D: MATL
        desc or "",           # E: DESC
        size or "",           # F: SIZE
        "",                   # G: WIDTH (blank placeholder)
        length_rounded,       # H: LENGTH (2 decimals)
        weight_combined,      # I: WEIGHT (combined, 2 decimals)
        mils or "",           # J: MIL SPEC
        packet_name,          # K: NEST
        page_number,          # L: PageNumber
        "",                   # M: CATALOG (blank placeholder)
        "",                   # N: CATEGORY (blank placeholder)
        "",                   # O: CUSTOMER/SUPPLIER (blank placeholder)
        "",                   # P: DIVISION 1 (blank placeholder)
        ""                    # Q: DIVISION 2 (blank placeholder)
    )


def extract_from_pdf(pdf_path: Path, log) -> List[Tuple]:
    """Extract PART SKETCH data from a single PDF."""
    rows = []
    # Strip .pdf extension from packet name
    packet_name = pdf_path.stem  # .stem gives filename without extension

    try:
        sdk.ensure_local(pdf_path)  # OneDrive placeholder -> download first (Hard Rule 13)
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        raise RuntimeError(f"Could not open PDF '{pdf_path}': {e}")

    for i, page in enumerate(doc, start=1):
        text = page.get_text("text") or ""
        if "PART SKETCH" in text:
            # Pass packet_name and page number to parsing function
            row = parse_part_sketch_page(text, packet_name, i)
            rows.append(row)

    return rows


# ===== VALIDATION FUNCTIONS =====
def validate_and_create_reports(output_folder: Path, output_name: str, ws: Worksheet, log) -> None:
    """
    Create validation reports for missing data:
    1. Excel file with missing cell information
    2. Text file with analysis summary
    
    Uses ORDER column for identification (no LINE number validation).
    """
    # Required columns to validate (indices 0-9: ORDER through MIL SPEC)
    required_columns = {
        0: "ORDER",
        1: "DYPN",
        2: "SOURCE",
        3: "MATL",
        4: "DESC",
        5: "SIZE",
        7: "LENGTH",  # Column H (skip G: WIDTH as it's blank placeholder)
        8: "WEIGHT",  # Column I
        9: "MIL SPEC" # Column J
    }
    
    # Collect missing data
    missing_data = []  # List of (row_num, ORDER, column_name, column_letter)
    missing_counts = {col_name: 0 for col_name in required_columns.values()}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        order = row[0].value or ""  # Column A: ORDER
        
        for col_idx, col_name in required_columns.items():
            cell_value = row[col_idx].value
            if cell_value is None or str(cell_value).strip() == "":
                col_letter = chr(65 + col_idx)  # Convert to A, B, C, etc.
                missing_data.append((row_idx, order, col_name, col_letter))
                missing_counts[col_name] += 1
    
    # Create missing items Excel report
    missing_xlsx_path = output_folder / f"{output_name}_missing_items.xlsx"
    create_missing_items_excel(missing_xlsx_path, missing_data, log)
    
    # Create analysis text report
    analysis_txt_path = output_folder / f"{output_name}_analysis.txt"
    create_analysis_text(analysis_txt_path, missing_counts, log)
    
    # Log summary
    total_missing = sum(missing_counts.values())
    if total_missing > 0:
        log(f"WARNING: Found {total_missing} missing cells across {len(missing_data)} locations")
        log(f"Validation reports created:")
        log(f"   - {missing_xlsx_path.name}")
        log(f"   - {analysis_txt_path.name}")
    else:
        log("No missing data found - all required fields populated!")


def create_missing_items_excel(xlsx_path: Path, missing_data: List[Tuple], log) -> None:
    """
    Create Excel file showing missing cell data.
    Format: ORDER BK503650 (A2) | DYPN (B2) | SOURCE (C2)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Data"
    
    # Header
    ws.append(["Missing Cell Data"])
    ws.append(["Format: ORDER <value> (<cell>) | <column> (<cell>) | ..."])
    ws.append([])  # Blank row
    
    if not missing_data:
        ws.append(["No missing data found!"])
    else:
        # Group by row
        by_row = defaultdict(list)
        
        for row_num, order, col_name, col_letter in missing_data:
            by_row[row_num].append((order, col_name, col_letter))
        
        # Write each row's missing data
        for row_num in sorted(by_row.keys()):
            items = by_row[row_num]
            order = items[0][0]  # ORDER is same for all items in row
            
            # Build string: "ORDER BK503650 (A2) | DYPN (B2) | SOURCE (C2)"
            parts = [f"ORDER {order} (A{row_num})"]
            for _, col_name, col_letter in items:
                parts.append(f"{col_name} ({col_letter}{row_num})")
            
            row_text = " | ".join(parts)
            ws.append([row_text])
    
    retry_file_op(wb.save, str(xlsx_path))


def create_analysis_text(txt_path: Path, missing_counts: Dict[str, int], log) -> None:
    """
    Create text file with analysis summary.
    Shows count of missing cells per column.
    """
    lines = []
    lines.append("=" * 50)
    lines.append("MISSING DATA ANALYSIS")
    lines.append("=" * 50)
    lines.append("")
    lines.append("Missing Cell Counts by Column:")
    lines.append("-" * 50)
    
    for col_name, count in missing_counts.items():
        lines.append(f"  Missing {col_name}: {count}")
    
    lines.append("")
    lines.append("=" * 50)
    
    # Write to file
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))



def load_or_create_workbook(xlsx_path: Path) -> Tuple[Workbook, Worksheet]:
    """Load existing workbook or create new one."""
    if xlsx_path.exists():
        wb = retry_file_op(load_workbook, filename=str(xlsx_path))
        ws = wb[wb.sheetnames[0]] if wb.sheetnames else wb.create_sheet("PartSketches")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "PartSketches"
        ws.append(HEADERS)
    
    return wb, ws


def existing_keys(ws: Worksheet) -> set:
    """Get set of existing (NEST, PageNumber) keys to prevent duplicates."""
    keys = set()
    for row in ws.iter_rows(min_row=2):
        nest = row[10].value      # Column K: NEST
        page_num = row[11].value  # Column L: PageNumber
        if nest and page_num:
            try:
                keys.add((nest, int(page_num)))
            except (ValueError, TypeError):
                pass
    return keys


def clean_cut_and_stamp(ws: Worksheet, log) -> int:
    """Remove 'CUT AND STAMP' text from cells in first 10 columns."""
    cleaned_count = 0
    for row in ws.iter_rows(min_row=2):
        # Check first 10 data columns (A-J: ORDER through MIL SPEC)
        for cell in row[0:10]:
            if cell.value and isinstance(cell.value, str) and "CUT AND STAMP" in cell.value.upper():
                cell.value = ""
                cleaned_count += 1
    return cleaned_count


def delete_empty_sequential_rows(ws: Worksheet, log) -> int:
    """Delete rows with empty data and sequential page numbers."""
    deleted_count = 0
    rows_to_delete = []
    all_rows = list(ws.iter_rows(min_row=2))
    
    for idx in range(len(all_rows)):
        current_row = all_rows[idx]
        # Check if first 10 columns are empty (A-J: ORDER through MIL SPEC)
        first_ten_empty = all(
            cell.value is None or str(cell.value).strip() == "" 
            for cell in current_row[0:10]
        )
        
        if first_ten_empty and idx > 0:
            current_page = current_row[11].value  # Column L (PageNumber)
            prev_page = all_rows[idx - 1][11].value
            
            try:
                if current_page and prev_page:
                    if int(current_page) == int(prev_page) + 1:
                        rows_to_delete.append(current_row[0].row)
            except (ValueError, TypeError):
                pass
    
    for row_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_num)
        deleted_count += 1
    
    return deleted_count


def append_rows(xlsx_path: Path, rows: List[Tuple], log) -> Tuple[int, Worksheet]:
    """Append rows to Excel, skipping duplicates. Returns (added_count, worksheet)."""
    wb, ws = load_or_create_workbook(xlsx_path)
    keys = existing_keys(ws)
    added = 0
    
    for r in rows:
        key = (r[10], int(r[11]))  # (NEST at index 10, PageNumber at index 11)
        if key in keys:
            continue
        ws.append(list(r))
        keys.add(key)
        added += 1
    
    cleaned = clean_cut_and_stamp(ws, log)
    if cleaned > 0:
        log(f"   Cleaned {cleaned} 'CUT AND STAMP' cells")
    
    deleted = delete_empty_sequential_rows(ws, log)
    if deleted > 0:
        log(f"   Deleted {deleted} empty sequential rows")
    
    retry_file_op(wb.save, str(xlsx_path))
    return added, ws


# ===== PDF EXTRACTION =====
def extract_drawings_to_pdf(output_folder: Path, pdfs: List[Path], all_rows: List[Tuple], log) -> Optional[Path]:
    """Extract PART SKETCH pages to combined PDF in output folder."""
    pdf_writer = PdfWriter()
    pdf_output = output_folder / "extracted_drawings.pdf"
    
    # Build set of (nest, page_num) that have data
    valid_pages = set()
    for row in all_rows:
        # Has data in first 10 fields (A-J: ORDER through MIL SPEC)
        if any(row[i] for i in range(10)):
            valid_pages.add((row[10], row[11]))  # (NEST at 10, PageNumber at 11)
    
    if not valid_pages:
        return None
    
    try:
        for pdf in pdfs:
            sdk.ensure_local(pdf)
            pdf_reader = PdfReader(str(pdf))
            packet_name = pdf.stem  # Strip .pdf for comparison
            for page_idx, page_num in enumerate(range(1, len(pdf_reader.pages) + 1)):
                if (packet_name, page_num) in valid_pages:
                    pdf_writer.add_page(pdf_reader.pages[page_idx])
        
        if len(pdf_writer.pages) > 0:
            with open(pdf_output, "wb") as f:
                pdf_writer.write(f)
            return pdf_output
    except Exception as e:
        log(f"WARNING: Could not create drawings PDF: {e}")
    
    return None


# ===== MAIN PLUGIN FUNCTION =====
def run(params: Dict[str, Any], progress_callback, cancel_event: threading.Event) -> None:
    """Main plugin execution."""
    log = params.get('log', print)
    
    log("Starting 911 Sketch Extractor v2.3.0...")
    progress_callback(0)
    
    # === PROMPT 1: FOLDER PATH ===
    log("Input required from user...")
    folder_input = get_console_input(params, "Enter folder path containing PDFs")
    
    folder = Path(folder_input.strip()).expanduser().resolve()
    if not folder.exists() or not folder.is_dir():
        log(f"ERROR: Folder not found: {folder}")
        raise ValueError(f"Folder not found: {folder}")
    
    log(f"Folder: {folder}")
    progress_callback(5)
    
    # === PROMPT 2: EXTRACT DRAWINGS ===
    extract_input = get_console_input(params, "Extract drawings to PDF? (yes/no)")
    extract_drawings = extract_input.strip().lower() in ('yes', 'y')
    log(f"Extract drawings: {'Yes' if extract_drawings else 'No'}")
    progress_callback(10)
    
    # === PROMPT 3: OUTPUT FILENAME ===
    output_input = get_console_input(params, "Enter output Excel filename (e.g., 'part_sketches')")
    
    if not output_input.strip():
        log("ERROR: Output filename cannot be empty!")
        raise ValueError("Output filename is required")
    
    # Auto-append .xlsx if not present
    output_filename = output_input.strip()
    if not output_filename.lower().endswith('.xlsx'):
        output_filename += '.xlsx'
    
    # === CREATE ORGANIZED OUTPUT FOLDER ===
    output_name = output_filename.replace('.xlsx', '')
    output_folder = folder / output_name
    output_folder.mkdir(exist_ok=True)
    
    xlsx_path = output_folder / output_filename
    log(f"Output folder: {output_folder.name}/")
    log(f"Output file: {output_filename}")
    log("")
    progress_callback(15)
    
    if cancel_event.is_set():
        log("Operation cancelled")
        return
    
    # === FIND PDFs ===
    log("Finding PDFs...")
    pdfs = sorted([p for p in folder.iterdir() if p.suffix.lower() == ".pdf"])
    
    if not pdfs:
        log("ERROR: No PDFs found in folder")
        raise ValueError("No PDFs found")
    
    log(f"Found {len(pdfs)} PDF(s)")
    progress_callback(20)
    
    # === EXTRACT DATA ===
    log("Extracting PART SKETCH data...")
    log("   This may take a few minutes...")
    log("")
    
    all_rows = []
    total = len(pdfs)
    
    for idx, pdf in enumerate(pdfs):
        if cancel_event.is_set():
            log("WARNING: Cancelled by user")
            return
        
        try:
            log(f"   Processing: {pdf.name}")
            rows = extract_from_pdf(pdf, log)
            all_rows.extend(rows)
            log(f"   Found {len(rows)} PART SKETCH page(s)")
        except Exception as e:
            log(f"   WARNING: Error: {e}")
        
        progress = 20 + int((idx + 1) / total * 60)
        progress_callback(progress)
    
    log("")
    log(f"Total rows extracted: {len(all_rows)}")
    progress_callback(80)
    
    # === APPEND TO EXCEL ===
    log("Writing to Excel...")
    try:
        added, ws = append_rows(xlsx_path, all_rows, log)
        log(f"Added {added} new rows to {output_filename}")
    except Exception as e:
        log(f"ERROR: Failed to write Excel: {e}")
        raise
    
    progress_callback(85)
    
    # === CREATE VALIDATION REPORTS ===
    log("Creating validation reports...")
    try:
        validate_and_create_reports(output_folder, output_name, ws, log)
    except Exception as e:
        log(f"WARNING: Warning: Could not create validation reports: {e}")
    
    progress_callback(90)
    
    # === EXTRACT DRAWINGS ===
    if extract_drawings:
        log("Extracting drawings to PDF...")
        pdf_output = extract_drawings_to_pdf(output_folder, pdfs, all_rows, log)
        if pdf_output:
            log(f"Drawings saved: {pdf_output.name}")
    
    progress_callback(100)
    
    # === SUMMARY ===
    log("")
    log("=" * 50)
    log("EXTRACTION SUMMARY")
    log("=" * 50)
    log(f"Processed: {len(pdfs)} PDFs")
    log(f"Extracted: {len(all_rows)} rows")
    log(f"Added: {added} new rows")
    log(f"Output folder: {output_folder}/")
    log(f"Excel file: {xlsx_path}")
    log(f"Validation reports:")
    log(f"   - {output_name}_missing_items.xlsx")
    log(f"   - {output_name}_analysis.txt")
    if extract_drawings and pdf_output:
        log(f"Drawings: {pdf_output}")
    log("=" * 50)
    log("911 Sketch Extractor completed successfully!")


if __name__ == "__main__":
    import threading
    cancel_event = threading.Event()
    
    def progress(p):
        print(f"Progress: {p}%")
    
    try:
        run(params={'console': None}, progress_callback=progress, cancel_event=cancel_event)
        print("\nDone!")
    except Exception as e:
        print(f"\nERROR: Failed: {e}")
