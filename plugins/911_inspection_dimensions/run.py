"""911 Inspection Dimensions.

Reads the PART SKETCH pages out of a folder of 911 nest-package PDFs and logs every
dimension printed on each drawing - plus the weld preps (the KB codes) - so they can
be dropped onto an inspection sheet instead of being typed off the paper by hand.

The drawings are RASTER images inside the PDF (there is no selectable text on them),
so the dimensions come out of an on-device OCR pass (RapidOCR / ONNX Runtime - no
network, no external exe). Everything above the drawing frame is real PDF text and is
read directly, which is where the part number / work order / FAB DIM come from.

Two things the reader deliberately does NOT count as inspection dimensions:
  * anything labelled REF (reference only) - listed separately,
  * numbers that live inside a drawing NOTE ("0.375 THK MECHANICAL SQUARE STEEL TUBE").

When the nest workbook is sitting next to the packet (the normal case in a batch
folder), the dimensions are also typed onto that part's QF-QU-09 inspection tab -
the one 911 Setup already copied out of the INSPECTION SHEET template. Only the
colour-filled TARGET cells are written; MIN and MAX beside them are array formulas
that derive themselves from the nominal, and a tab that already has numbers on it
is never touched.
"""

import os
import re
import glob
import datetime

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:  # standalone CLI testing
    import sys
    import pathlib

    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk


VERSION = "0.5.2"

# The detector is run twice at different working resolutions and the two results are
# merged: the small pass reliably picks up crowded dimension stacks, the large pass
# picks up small isolated callouts (an R .25 hanging off a corner). Either one alone
# misses roughly one dimension in ten on these drawings.
_DET_SIDES = (736, 2000)
_PAD_PX = 60          # white margin so text flush with the image edge still detects
_RENDER_DPI = 300


# --------------------------------------------------------------------------- OCR
_ENGINES = None


def _engines(log):
    """Lazily build the OCR engines (model load is ~1s and only needed on a real run)."""
    global _ENGINES
    if _ENGINES is not None:
        return _ENGINES
    # BOTH the import AND the construction have to be guarded. RapidOCR loads its
    # three engine classes by NAME out of its own config.yaml, so a frozen build
    # that did not bundle them raises HERE, not at import - and it used to surface
    # as "could not read this PDF", which made a broken reader look like a packet
    # with no sketches in it (FTOURIGNY-LT, 2026-09-01).
    try:
        from rapidocr_onnxruntime import RapidOCR
    except Exception as exc:  # pragma: no cover - dependency guard
        raise _reader_dead(exc)
    log("Starting the drawing reader...")
    try:
        _ENGINES = [
            RapidOCR(
                det_model_path=None,
                det_limit_side_len=side,
                det_box_thresh=0.3,
                cls_model_path=None,
                rec_model_path=None,
                text_score=0.3,
            )
            for side in _DET_SIDES
        ]
    except Exception as exc:  # pragma: no cover - dependency guard
        _ENGINES = None
        raise _reader_dead(exc)
    return _ENGINES


def _reader_dead(exc):
    """The one message for a drawing reader that will not start at all."""
    return sdk.UserFacingError(
        "The drawing reader could not start up (%s)." % exc,
        "Nothing was read and no workbook was changed. This app needs the built-in "
        "drawing reader, which ships inside TechDeck. Update TechDeck to the latest "
        "version, and if it still fails send a Debug Report to a TechDeck admin.",
    )


# --------------------------------------------------------------- text normalisation
# The recogniser is a general-purpose model, so a stroke-font period sometimes comes
# back as a CJK full stop and a hyphen as a CJK dash. Fold those back before parsing.
_TRANSLATE = {
    ord("。"): ".", ord("．"): ".", ord("·"): ".",
    ord("，"): ",", ord("、"): ",",
    ord("一"): "-", ord("－"): "-", ord("–"): "-", ord("—"): "-",
    ord("Ｘ"): "X", ord("×"): "X",
}
for _i, _ch in enumerate("０１２３４５６７８９"):
    _TRANSLATE[ord(_ch)] = str(_i)

# Arrowheads and extension lines touching a dimension come back as leading/trailing
# punctuation. Both a left arrow ("-.46") and a right one ("+.46") happen - missing the
# "+" dropped a real .46 off H5370103-32.
_EDGE_JUNK = re.compile(r"^[\s\-+_=~|,'`\"<>*]+|[\s\-+_=~|,'`\"<>*]+$")


def _norm(text):
    """Fold OCR quirks and strip the arrowhead dashes that glue onto a dimension."""
    t = (text or "").translate(_TRANSLATE)
    t = t.replace("　", " ").strip().upper()
    return _EDGE_JUNK.sub("", t)


# ----------------------------------------------------------------- classification
# Words that may sit next to a number without making it prose.
MODIFIER_WORDS = {
    "TYP", "REF", "THK", "MIN", "MAX", "NOM", "DIA", "RAD", "R", "X",
    "SNIPE", "CUT", "NEAT", "PL", "PLCS", "PLC", "EA", "DEG", "BOTH", "SIDES",
    "NS", "FS", "AND", "OR", "TO", "OF",
}

# --- weld preps -------------------------------------------------------------
# A weld prep is called out as a bevel code on a leader line, usually with the side
# it applies to on the line underneath ("KB114" / "NS & FS" = near side and far
# side). Codes come back from OCR with the leader dash glued on ("-KB114") and
# sometimes with the digits spaced out ("KB 1 1 4"), so match on the space-stripped
# text after trimming leading punctuation.
#
# The bevel book uses four prefixes, not just KB: KB (standard), SB (shell), FB
# (flange) and WB (web). Matching only KB silently dropped every shell and flange
# prep on the drawing - and SB alone is 198 of the 947 sheets.
WELD_PREP_RE = re.compile(r"^(?:KB|SB|FB|WB)[A-Z0-9]{1,6}$")
SIDE_RE = re.compile(r"^(?:[NFB]S(?:&[NFB]S)?|BOTHSIDES?|NEARSIDE|FARSIDE)$")
# Modifiers that ride along on the dimension itself.
TRAILING_MODS = ("TYP", "REF", "THK", "MIN", "MAX", "NOM", "SNIPE")

_NUM = r"\d*\.?\d+"
NUM_RE = re.compile(r"^%s$" % _NUM)
COMPOUND_RE = re.compile(r"^(%s)(?:X(%s))+$" % (_NUM, _NUM))
COUNT_PREFIX_RE = re.compile(r"^\(?(\d{1,2})\)?X(.+)$")
ANGLE_RE = re.compile(r"^(\d+(?:\.\d+)?)\s*(?:°|DEG)$")
FRACTION_RE = re.compile(r"^(\d+)-(\d+)/(\d+)$")
# A chamfer callout, written either way round on these drawings:
#   "1.00 X 45°" (land first) and "45° X .5" (angle first).
CHAMFER_RE = re.compile(r"^(%s)X(\d+(?:\.\d+)?)(?:°|DEG)$" % _NUM)
CHAMFER_REV_RE = re.compile(r"^(\d+(?:\.\d+)?)(?:°|DEG)X(%s)$" % _NUM)


def classify(raw):
    """Turn one OCR token into (kind, value, mods), or None if it isn't a dimension."""
    t = _norm(raw).replace(" ", "")
    if not t:
        return None
    mods = []
    kind = "linear"

    cham = CHAMFER_RE.match(t)
    if cham:
        return "chamfer", "%s X %s deg" % (cham.group(1), cham.group(2)), mods
    cham = CHAMFER_REV_RE.match(t)
    if cham:
        return "chamfer", "%s deg X %s" % (cham.group(1), cham.group(2)), mods

    ang = ANGLE_RE.match(t)
    if ang:
        return "angle", ang.group(1), mods

    if t.startswith("DIA"):
        kind, t = "diameter", t[3:]
    elif t and t[0] in "Ø⌀" and len(t) > 1:
        kind, t = "diameter", t[1:]
    elif t.startswith("R") and len(t) > 1 and (t[1].isdigit() or t[1] == "."):
        kind, t = "radius", t[1:]

    # a leading count multiplier: 2X .50, (4)X .50
    m = COUNT_PREFIX_RE.match(t)
    if m and not m.group(2).startswith("."):
        pass  # 1.00X1.00 style compound, not a count - fall through
    elif m:
        mods.append("%sX" % m.group(1))
        t = m.group(2)

    for word in TRAILING_MODS:
        if t.endswith(word) and len(t) > len(word):
            mods.append(word)
            t = t[: -len(word)]
        if t.startswith(word) and len(t) > len(word):
            mods.append(word)
            t = t[len(word):]
    t = _EDGE_JUNK.sub("", t)

    if COMPOUND_RE.match(t):
        return kind, " X ".join(t.split("X")), mods
    if FRACTION_RE.match(t):
        return kind, t, mods
    if not NUM_RE.match(t):
        return None
    # A bare integer with no decimal point is almost never a dimension on these
    # drawings (they are all printed to two places) - it is a view tag or a stray.
    if "." not in t:
        return None
    return kind, t, mods


# ------------------------------------------------------------------------ geometry
def _bbox(box):
    xs = [p[0] for p in box]
    ys = [p[1] for p in box]
    return min(xs), min(ys), max(xs), max(ys)


def _near_label(item, items, label):
    """True when a standalone `label` box sits against this one on any side."""
    x0, y0, x1, y1 = item["bb"]
    h = max(1.0, y1 - y0)
    w = max(1.0, x1 - x0)
    for other in items:
        if other is item or other["word"] != label:
            continue
        ox0, oy0, ox1, oy1 = other["bb"]
        xover = min(x1, ox1) - max(x0, ox0)
        yover = min(y1, oy1) - max(y0, oy0)
        if xover > 0.3 * min(w, ox1 - ox0):
            gap = max(oy0 - y1, y0 - oy1)          # above or below
            if -0.2 * h <= gap <= 1.4 * h:
                return True
        if yover > 0.4 * min(h, oy1 - oy0):
            gap = max(ox0 - x1, x0 - ox1)          # left or right
            if -0.2 * w <= gap <= 2.5 * h:
                return True
    return False


_PREFIXES = ("KB", "SB", "FB", "WB")

# Every code in the bevel book is a two-letter prefix followed by 3 or 4 DIGITS -
# checked against all 942, none has a letter in the tail. So a letter that comes
# back there is always an OCR slip, and in this stroke font it is always one of
# these: the digit 1 read as a capital I or lower-case l, and 0 read as O.
# Worth fixing rather than rejecting: on the 30-packet real sweep, KB114 came back
# as "KBI14" or "KBII4" three times out of 184 preps, and each one was dropped.
_TAIL_FIX = {"I": "1", "L": "1", "O": "0"}


def _weld_code(word):
    """Normalise one OCR token to a weld-prep code, or None."""
    t = re.sub(r"^[^A-Z0-9]+", "", word).replace(" ", "")
    if not t.startswith(_PREFIXES):
        return None
    # trim a side note that OCR glued onto the code ("KB114NS&FS")
    m = re.match(r"^((?:KB|SB|FB|WB)[A-Z0-9]{1,6}?)((?:[NFB]S(?:&[NFB]S)?)?)$", t)
    if not m:
        return None
    code = m.group(1)
    code = code[:2] + "".join(_TAIL_FIX.get(c, c) for c in code[2:])
    if not WELD_PREP_RE.match(code) or not code[2:].isdigit():
        return None
    return code, _fmt_side(m.group(2))


def nearest_bevel_code(code, log=None):
    """An in-book code one edit away from a misread one, or "" if it is ambiguous.

    Only ever SUGGESTED in the report, never substituted - a bevel is a cut on a
    real part and a plausible-looking guess is worse than saying "look this up".
    The sweep turned up "KB1141" and "KB1147" (a leader line read as an extra
    digit) where dropping one character gives KB114 and nothing else.
    """
    table = bevel_table(log)
    if not code or code in table:
        return ""
    # A leader line touching the end of the code reads as an extra digit, so try
    # the trailing character first - that is the shape actually seen ("KB1141",
    # "KB1147" for KB114). Deleting an INNER character is a much weaker guess and
    # only offered when exactly one candidate survives.
    if len(code) > 5 and code[:-1] in table:
        return code[:-1]
    cands = {code[:i] + code[i + 1:] for i in range(2, len(code))}
    hits = sorted(c for c in cands if c in table)
    return hits[0] if len(hits) == 1 else ""


def _fmt_side(side):
    """Put the spaces back into a side note the glue-strip ran together."""
    return re.sub(r"\s+", " ", (side or "").replace("&", " & ")).strip()


def weld_preps(items):
    """Find the KB weld-prep callouts and the side each applies to.

    Run over the RAW box list, independent of the note/dimension passes - a weld
    prep is neither, and must not be lost to either one.
    """
    found = []
    for it in items:
        parsed = _weld_code(it["word"])
        if not parsed:
            continue
        code, glued_side = parsed
        side = glued_side
        if not side:
            side = _side_note(it, items)
        found.append({"code": code, "side": side, "score": it["score"], "bb": it["bb"]})

    # one callout can be detected twice (both OCR passes, slightly different boxes)
    unique = []
    for w in sorted(found, key=lambda d: (-len(d["side"]), -d["score"])):
        cx, cy = (w["bb"][0] + w["bb"][2]) / 2.0, (w["bb"][1] + w["bb"][3]) / 2.0
        h = max(1.0, w["bb"][3] - w["bb"][1])
        if any(u["code"] == w["code"]
               and abs(cx - (u["bb"][0] + u["bb"][2]) / 2.0) < 3 * h
               and abs(cy - (u["bb"][1] + u["bb"][3]) / 2.0) < 3 * h
               for u in unique):
            continue
        unique.append(w)
    unique.sort(key=lambda d: (round(d["bb"][1] / 60.0), d["bb"][0]))
    return unique


def _side_note(item, items):
    """The NS / FS / NS & FS label sitting under (or beside) a weld-prep code."""
    x0, y0, x1, y1 = item["bb"]
    h = max(1.0, y1 - y0)
    best = ""
    for other in items:
        if other is item:
            continue
        flat = other["word"].replace(" ", "")
        if not SIDE_RE.match(flat):
            continue
        ox0, oy0, ox1, oy1 = other["bb"]
        xover = min(x1, ox1) - max(x0, ox0)
        yover = min(y1, oy1) - max(y0, oy0)
        close = (xover > 0.3 * min(x1 - x0, ox1 - ox0)
                 and -0.2 * h <= max(oy0 - y1, y0 - oy1) <= 1.6 * h)
        beside = (yover > 0.4 * min(h, oy1 - oy0)
                  and -0.2 * h <= max(ox0 - x1, x0 - ox1) <= 2.5 * h)
        if (close or beside) and len(other["word"]) > len(best):
            best = other["word"]
    return _fmt_side(best)


def _note_lines(items):
    """Group boxes into text lines and return the ones that read as prose notes.

    A note is a line (or a stacked paragraph of lines) carrying two or more ordinary
    words - "0.375 THK MECHANICAL SQUARE STEEL TUBE", "CUT FROM / 3.00X3.00X.188 /
    ANGLE IRON". Numbers inside one are stock/material callouts, not dimensions.
    """
    lines = []
    for it in sorted(items, key=lambda d: d["bb"][1]):
        placed = False
        for line in lines:
            ly0, ly1 = line["y0"], line["y1"]
            h = min(it["bb"][3] - it["bb"][1], ly1 - ly0)
            if min(it["bb"][3], ly1) - max(it["bb"][1], ly0) > 0.5 * max(1.0, h):
                line["items"].append(it)
                line["y0"] = min(ly0, it["bb"][1])
                line["y1"] = max(ly1, it["bb"][3])
                placed = True
                break
        if not placed:
            lines.append({"items": [it], "y0": it["bb"][1], "y1": it["bb"][3]})

    # stack lines into paragraphs (tight vertical spacing + overlapping x range)
    paras = []
    for line in sorted(lines, key=lambda l: l["y0"]):
        lx0 = min(i["bb"][0] for i in line["items"])
        lx1 = max(i["bb"][2] for i in line["items"])
        h = line["y1"] - line["y0"]
        for para in paras:
            if (0 <= line["y0"] - para["y1"] <= 0.9 * h
                    and min(lx1, para["x1"]) - max(lx0, para["x0"]) > 0):
                para["lines"].append(line)
                para["y1"] = line["y1"]
                para["x0"] = min(para["x0"], lx0)
                para["x1"] = max(para["x1"], lx1)
                break
        else:
            paras.append({"lines": [line], "y0": line["y0"], "y1": line["y1"],
                          "x0": lx0, "x1": lx1})

    notes = []
    for para in paras:
        members = [i for line in para["lines"] for i in line["items"]]
        words = []
        for it in members:
            for chunk in re.split(r"[^A-Z]+", it["word"]):
                if len(chunk) >= 3 and chunk not in MODIFIER_WORDS:
                    words.append(chunk)
        if len(words) >= 2:
            text = " ".join(
                i["word"] for i in sorted(members, key=lambda d: (round(d["bb"][1] / 20), d["bb"][0]))
            )
            notes.append({"text": text, "members": members})
    return notes


# ---------------------------------------------------------------------- page work
def drawing_clip(page):
    """Rect of the biggest embedded raster on the page - the drawing frame."""
    best, best_px = None, 0
    for img in page.get_images(full=True):
        px = img[2] * img[3]
        if px < 400 * 400:
            continue
        for rect in page.get_image_rects(img[0]):
            if px > best_px:
                best_px, best = px, rect
    return best


def _read_boxes(page, clip, log):
    """Render the drawing and run both OCR passes over it, merged and de-duplicated."""
    import fitz
    import numpy as np

    scale = _RENDER_DPI / 72.0
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), clip=clip)
    arr = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 4:
        arr = arr[:, :, :3]
    elif pix.n == 1:
        arr = np.repeat(arr, 3, axis=2)
    arr = np.ascontiguousarray(arr[:, :, ::-1])           # RGB -> BGR for the model
    padded = np.full((arr.shape[0] + 2 * _PAD_PX, arr.shape[1] + 2 * _PAD_PX, 3), 255, np.uint8)
    padded[_PAD_PX:_PAD_PX + arr.shape[0], _PAD_PX:_PAD_PX + arr.shape[1]] = arr

    items = []
    for engine in _engines(log):
        result, _ = engine(padded)
        for box, txt, score in (result or []):
            word = _norm(txt)
            if not word:
                continue
            items.append({"raw": txt, "word": word, "bb": _bbox(box), "score": float(score)})

    # Merge the two passes. The same text can come back as one wide box from one pass
    # and two narrower boxes from the other, so drop anything whose box mostly sits
    # inside a longer reading, then collapse near-identical centres.
    items.sort(key=lambda d: (-(d["bb"][2] - d["bb"][0]) * (d["bb"][3] - d["bb"][1]), -d["score"]))
    merged = []
    for it in items:
        x0, y0, x1, y1 = it["bb"]
        area = max(1.0, (x1 - x0) * (y1 - y0))
        h = max(1.0, y1 - y0)
        keep = True
        for seen in merged:
            sx0, sy0, sx1, sy1 = seen["bb"]
            overlap = (max(0.0, min(x1, sx1) - max(x0, sx0))
                       * max(0.0, min(y1, sy1) - max(y0, sy0)))
            if overlap > 0.7 * area:
                keep = False
                break
            if (abs((x0 + x1) - (sx0 + sx1)) / 2.0 < h
                    and abs((y0 + y1) - (sy0 + sy1)) / 2.0 < h):
                keep = False
                break
        if keep:
            merged.append(it)
    return merged


_X_TAIL_RE = re.compile(r"^X\s*(%s)$" % _NUM)


def _join_split_chamfers(items):
    """Stitch a chamfer the detector split into two boxes.

    "45° X .5" often comes back as an angle box and a separate "X.5" box. Left
    alone the angle loses its land and the "X.5" parses as nothing, so the whole
    callout collapses to a bare 45 deg (bit H5370103-32).
    """
    out = []
    used = set()
    for i, it in enumerate(items):
        if i in used:
            continue
        if not ANGLE_RE.match(it["word"].replace(" ", "")):
            continue
        x0, y0, x1, y1 = it["bb"]
        h = max(1.0, y1 - y0)
        for j, other in enumerate(items):
            if j == i or j in used:
                continue
            m = _X_TAIL_RE.match(other["word"].replace(" ", ""))
            if not m:
                continue
            ox0, oy0, ox1, oy1 = other["bb"]
            yover = min(y1, oy1) - max(y0, oy0)
            if yover > 0.4 * h and -0.2 * h <= (ox0 - x1) <= 2.0 * h:
                merged = dict(it)
                merged["raw"] = "%s X %s" % (it["word"], m.group(1))
                merged["word"] = merged["raw"]
                merged["bb"] = (min(x0, ox0), min(y0, oy0), max(x1, ox1), max(y1, oy1))
                merged["score"] = min(it["score"], other["score"])
                out.append(merged)
                used.update((i, j))
                break
    return [it for k, it in enumerate(items) if k not in used] + out


def read_drawing(page, clip, log):
    """Return (dimensions, notes, weld preps) for one drawing page."""
    items = _join_split_chamfers(_read_boxes(page, clip, log))
    welds = weld_preps(items)
    notes = _note_lines(items)
    in_note = {id(m) for n in notes for m in n["members"]}

    dims = []
    for it in items:
        if id(it) in in_note:
            continue
        parsed = classify(it["raw"])
        if not parsed:
            continue
        kind, value, mods = parsed
        ref = "REF" in mods or _near_label(it, items, "REF")
        if _near_label(it, items, "TYP") and "TYP" not in mods:
            mods.append("TYP")
        if _near_label(it, items, "SNIPE") and "SNIPE" not in mods:
            mods.append("SNIPE")
        dims.append({
            "kind": kind,
            "value": value,
            "mods": [m for m in mods if m != "REF"],
            "ref": ref,
            "raw": it["raw"],
            "score": it["score"],
            "bb": it["bb"],
        })
    dims.sort(key=lambda d: (round(d["bb"][1] / 60.0), d["bb"][0]))
    return dims, [n["text"] for n in notes], welds


_FIELD_RE = {
    "part": re.compile(r"PART:\s*([A-Z0-9][A-Z0-9\-\.]*)"),
    "rev": re.compile(r"REV/SEQ:\s*([A-Z0-9/]+)"),
    "qty": re.compile(r"QTY:\s*(\d+)"),
    "noun": re.compile(r"NOUN:\s*([A-Z][A-Z ]*?)\s*\n"),
    "size": re.compile(r"SIZE:\s*\n?\s*([^\n]+)"),
    "fab_dim": re.compile(r"FAB DIM:\s*\n?\s*([0-9.]+)"),
    "srce": re.compile(r"SRCE:\s*([A-Z0-9]+)"),
}


# A blank title-block cell makes the next label bleed into the capture ("REV/SEQ: SRCE").
_LABELS = {"SRCE", "QTY", "NOUN", "MATL", "LVL", "FER", "SIZE", "DESC", "PART",
           "REV", "SEQ", "MFG", "NUC", "NON", "PRT", "WT", "FAB", "DIM", "SHIP", "DUE", "DTE"}


def title_block(page):
    """Pull the real-text header fields off a PART SKETCH page."""
    txt = page.get_text()
    out = {}
    for key, rx in _FIELD_RE.items():
        m = rx.search(txt)
        value = m.group(1).strip() if m else ""
        if value.rstrip(":").upper() in _LABELS:
            value = ""
        out[key] = value
    # the first line after the page counter is the work-order barcode label
    lines = [l.strip() for l in txt.splitlines() if l.strip()]
    out["work_order"] = ""
    for line in lines[1:4]:
        if re.fullmatch(r"[A-Z]{1,2}\d{5,}", line):
            out["work_order"] = line
            break
    return out


# ------------------------------------------------------- writing the nominals
# QF-QU-09 grid. Ten two-row "Characteristic" slots down the sheet; each slot row
# carries FIVE "Specification Requirement" groups across. Only the group's TARGET
# cell is ours to fill - the template's own instruction (AV15) reads "Put Nominals
# in Color-Filled Columns (Use degree sign to indicate angles)", and those
# colour-filled cells are exactly these five columns. MIN and MAX next to each one
# are array formulas that derive themselves from the nominal (+/- 0.1 linear from
# the tolerance table at AV19:AY23, +/- 1 for degrees), so writing anything into
# them would destroy the sheet's own arithmetic.
SLOT_ROWS = tuple(range(16, 36, 2))
VALUE_COLS = ("L", "S", "Z", "AG", "AN")
MAX_NOMINALS = len(SLOT_ROWS) * len(VALUE_COLS)

_PART_CELL = "A16"
_NEST_WORKBOOK_GLOBS = ("911 BATCH*.xlsx", "911 PLATE BATCH*.xlsx")
# Sheets that are never a per-part inspection tab.
_NON_PART_SHEETS = {"NEST", "SCRIBE VERIFICATION", "COVER SHEET",
                    "SOURCE MATERIAL INFO", "INSPECTION SHEET"}

# What a usable angle nominal looks like once folded: a bare number, nothing else.
# Shared by _as_nominal (a standalone angle off the drawing) and weld_prep_angles
# (an angle out of the bevel book) - both write the same "NN°" the sheet expects.
ANGLE_VALUE_RE = re.compile(r"^\d+(?:\.\d+)?$")


def _as_nominal(text, kind="linear"):
    """A dimension value as the sheet wants it: a number, or 'NN°' for an angle.

    KIND is not optional information. A chamfer carries "deg" inside its own value
    string ("1.75 X 45 deg"), but a STANDALONE angle's value is the bare number -
    the " deg" only ever existed in the printed report, added from the kind. So
    reading the text alone wrote a standalone angle as a plain number, and the
    sheet's MIN/MAX array formula branches on `ISNUMBER(SEARCH("°", <cell>))`:
    without the sign it fell through to the LINEAR tolerance bands and quietly put
    a +/-.1 tolerance on a 45 degree angle instead of the +/-1 the form intends
    (V094 503891 tabs '67-199' and '-451', 2026-09-01).
    """
    t = str(text).strip()
    if t.lower().endswith("deg"):
        return t[:-3].strip() + "°"
    if kind == "angle" and ANGLE_VALUE_RE.match(t):
        return t + "°"
    try:
        return float(t)
    except ValueError:
        return t


# ------------------------------------------------------------- bevel reference
# bevel_table.csv is the Electric Boat bevel book (Dept 470/459/415 "BEVEL
# TEMPLATE" sheets) transcribed to one row per code: the near/far side angles, the
# max land, and whether the sheet has been VOIDed in favour of another code.
# It ships next to run.py, so it resolves the same way in dev and in the frozen
# build (both load the plugin from its own folder).
_BEVEL_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bevel_table.csv")
_BEVEL_CACHE = None


def bevel_table(log=None):
    """The bevel book as {code: row}, loaded once. Empty dict if it is missing."""
    global _BEVEL_CACHE
    if _BEVEL_CACHE is not None:
        return _BEVEL_CACHE
    import csv

    table = {}
    try:
        with open(_BEVEL_CSV, newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                table[row["code"].upper()] = row
                # Two sheets are filed under a KB name but printed as WB (KB066 /
                # KB076 read WB066 / WB076). A drawing calls the PRINTED number, so
                # index both spellings at the same row.
                printed = (row.get("printed_no") or "").upper()
                if printed and printed not in table:
                    table[printed] = row
    except OSError as exc:
        if log:
            log("   ! bevel table not readable (%s) - weld preps will be reported "
                "but not looked up" % exc)
    _BEVEL_CACHE = table
    return table


def bevel_lookup(code, log=None):
    """One weld-prep code's bevel data, or None when it is not in the book."""
    return bevel_table(log).get(str(code).upper())


def _weld_faces(side):
    """The faces a side note names, as ('NS',), ('FS',), ('NS', 'FS') or ().

    "BS" is the drawings' shorthand for both sides and has to land on two faces
    like "NS & FS" does - reading it as a single unknown face wrote one angle
    where the sheet wants two.
    """
    flat = (side or "").upper().replace(" ", "").replace("&", "").replace("/", "")
    if "BOTH" in flat or "BS" in flat:
        return ("NS", "FS")
    return tuple(f for f in ("NS", "FS") if f in flat)


def weld_prep_angles(weld, log=None):
    """The angle nominals a weld prep contributes, as ('45°', ...).

    ONE ENTRY PER FACE THE DRAWING NAMES - so "KB114  NS & FS" is two 45 deg
    entries, not one. Verified against the hand-filled sheet for H4136024-41 on
    S026, which reads 35.41 / 45 deg / 45 deg: the part is one tube with the same
    bevel cut on the near and far side, and each one is a separate thing to
    inspect. Emitting a single angle there under-filled the sheet by one row.

    KB114 is a SINGLE N/S BEVEL, so the book only carries a near-side angle - when
    the drawing applies it to both faces, that same angle stands for both. A true
    double bevel (KB200, 22.5/22.5) carries an angle per face and each is used.

    Only the ANGLE is written. The land on these sheets is a MAX ("0-1/16 MAX
    LAND"), not a target, and the QF-QU-09 group derives a +/-0.1 band around
    whatever nominal is typed in - so writing 0.06 there would assert a 0.06 +/- 0.1
    land the print never called for. The land is carried into the report instead.
    """
    row = bevel_lookup(weld["code"], log)
    if not row or row.get("status"):
        return ()                                # unknown or VOID - caller reports it
    ns, fs = row["ns_angle"].strip(), row["fs_angle"].strip()
    faces = _weld_faces(weld.get("side"))
    if faces:
        # one per named face, falling back to whichever angle the sheet does carry
        picked = [{"NS": ns, "FS": fs}[f] or ns or fs for f in faces]
    else:
        # no side note - take the sheet at face value: every angle it specifies
        picked = [a for a in (ns, fs) if a]
    out = []
    for angle in picked:
        if not angle:
            continue
        # the table carries the print's own wording ("22 1/2", "45 TYP", "52.0")
        value = _angle_decimal(re.sub(r"\s*TYP$", "", angle).strip())
        # Anything that is not a bare number is a transcription that needs a human,
        # not a nominal. Two sheets used to carry a compound cell ("30 / 30" on
        # FB064, "50 / 25" on SB716 where 50 was the INCLUDED angle) which would
        # have typed the literal string "30 / 30 deg" into a QA form. Drop it here
        # so a bad cell can never reach a sheet, whatever the table says.
        if value and ANGLE_VALUE_RE.match(value):
            out.append("%s°" % value)
    return tuple(out)


def _angle_decimal(angle):
    """'22 1/2' -> '22.5'. A handful of older sheets print the angle as a mixed
    fraction; the inspection form and every other angle on it are decimal, so fold
    them rather than leaving one odd '22 1/2°' among the '22.5°'s."""
    m = re.match(r"^(\d+)\s+(\d+)/(\d+)$", angle)
    if not m:
        return angle
    whole, num, den = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not den:
        return angle
    value = whole + num / den
    return ("%.10f" % value).rstrip("0").rstrip(".")


def nominals_for(dims, welds=(), log=None):
    """Flatten one part's dimensions into the ordered values to type into the form.

    Compound callouts become separate entries because that is how they are
    inspected and how the sheets are filled by hand: a chamfer "45 deg X .5" is an
    angle plus a land, and a snipe ".50 X .50" is two lengths.

    Weld-prep angles go in after the printed dimensions and BEFORE the TYP block.
    The completed sheets carry them (the 45 deg entries on 503577 '24-41' come from
    the weld prep, not the drawing) - the drawing only names the bevel code, and
    the angle lives in the bevel book. They must not land after the TYP block or
    they break up the run of repeats the reviewer copies out.

    TYP dimensions go LAST (the user's call). A TYP callout is printed once but the
    feature repeats, and the drawing never says how many times - so the reader
    cannot know the count. Putting them at the end of the run means the reviewer
    can copy the final entries out as many times as the part actually needs,
    without having to unpick them from the middle of the list.
    """
    plain, typ = [], []
    for d in dims:
        if d["ref"]:
            continue
        parts = [_as_nominal(p, d["kind"]) for p in str(d["value"]).split(" X ")]
        (typ if "TYP" in d["mods"] else plain).append(parts)

    # Not de-duplicated: a repeated angle is a repeated FEATURE. "KB114 NS & FS" is
    # two 45 deg bevels on one part and the hand-filled sheets carry both.
    # process_pdf already collapses the same (code, side) seen on more than one view
    # page, so nothing double-counts here.
    bevel = []
    for weld in welds or ():
        bevel.extend(weld_prep_angles(weld, log))

    out = []
    for group in plain:
        out.extend(group)
    out.extend(bevel)
    for group in typ:
        out.extend(group)
    return out


def sheet_nominals(ws):
    """Every nominal already on an inspection tab (blank-stripped)."""
    found = []
    for row in SLOT_ROWS:
        for col in VALUE_COLS:
            v = ws["%s%d" % (col, row)].value
            if v not in (None, ""):
                found.append(v)
    return found


def writable_slots(ws):
    """The nominal cells that can actually take a value, in fill order.

    Not every sheet still has all fifty. On some tabs a whole group has been
    merged into one free-text box (K18:P18 on V092 503838's -244 and -238), which
    leaves its nominal cell a read-only MergedCell - assigning to it raises
    "'MergedCell' object attribute 'value' is read-only". Skip those and carry on
    into the next slot rather than writing into somebody's note.
    """
    from openpyxl.cell.cell import MergedCell

    slots = []
    for row in SLOT_ROWS:
        for col in VALUE_COLS:
            cell = ws["%s%d" % (col, row)]
            if not isinstance(cell, MergedCell):
                slots.append(cell)
    return slots


# The MIN cell that belongs to each group's nominal. Used only to check the group's
# arithmetic is still live before trusting a nominal written into it.
_MIN_COL = {"L": "N", "S": "U", "Z": "AB", "AG": "AI", "AN": "AP"}


def _tolerance_is_live(ws, cell):
    """False when this group's MIN has been hand-typed over its array formula.

    Seen in the wild on V092 503838 '-244', where MIN/MAX were replaced with literal
    '40°'/'50°'. Writing a nominal into such a group leaves a number sitting next to
    somebody else's tolerance, which on a QA form is worse than not filling it - so
    the caller flags it instead of letting it pass silently.
    """
    from openpyxl.worksheet.formula import ArrayFormula

    col = _MIN_COL.get(cell.column_letter)
    if not col:
        return True
    value = ws["%s%d" % (col, cell.row)].value
    if isinstance(value, ArrayFormula):
        return True
    return isinstance(value, str) and value.startswith("=")


def write_nominals(ws, values):
    """Lay values across the form's five groups, wrapping to the next slot row."""
    slots = writable_slots(ws)
    written, stale = 0, []
    for cell, value in zip(slots, values):
        cell.value = value
        written += 1
        if not _tolerance_is_live(ws, cell):
            stale.append(cell.coordinate)
    return written, max(0, len(values) - len(slots)), stale


def _tab_part_number(ws):
    """The part a tab belongs to, from the part-number cell 911 Setup stamps."""
    raw = ws[_PART_CELL].value
    if not raw:
        return ""
    # the cell carries "<DYPN>\n<HULL>" (or the hull run on with spaces)
    return str(raw).replace("\r", "\n").split("\n")[0].strip().split()[0].upper()


def inspection_tabs(wb):
    """Per-part inspection tabs, as {sheet name: part number}."""
    tabs = {}
    for name in wb.sheetnames:
        if name.strip().upper() in _NON_PART_SHEETS:
            continue
        ws = wb[name]
        if ws.sheet_state != "visible":
            continue
        part = _tab_part_number(ws)
        if part:
            tabs[name] = part
    return tabs


def find_nest_workbook(folder):
    """The '911 BATCH <batch> <nest>.xlsx' in a nest folder, if there is one."""
    for pattern in _NEST_WORKBOOK_GLOBS:
        hits = [p for p in glob.glob(os.path.join(folder, pattern))
                if not os.path.basename(p).startswith("~$")]
        if hits:
            return sorted(hits)[0]
    return None


def fill_nest_workbook(workbook_path, parts, log, dry_run=False):
    """Write each part's nominals onto its inspection tab(s). Returns per-tab rows.

    A tab that already carries nominals is left exactly as it is - those are
    somebody's judgement calls and this reader has no business overwriting them.
    """
    results = []
    wb = sdk.load_workbook_resilient(workbook_path, log=log)
    try:
        tabs = inspection_tabs(wb)
        if not tabs:
            return [{"tab": "-", "part": "-", "status": "no inspection tabs in this workbook"}]

        by_part = {}
        for rec in parts:
            if rec.get("part") and not rec.get("problem"):
                by_part.setdefault(rec["part"].upper(), rec)

        touched = 0
        for name, part in tabs.items():
            ws = wb[name]
            rec = by_part.get(part)
            row = {"tab": name, "part": part, "written": 0, "status": ""}
            if rec is None:
                row["status"] = "no drawing read for this part"
                results.append(row)
                continue
            existing = sheet_nominals(ws)
            if existing:
                row["status"] = ("already filled in (%d value(s)) - left alone"
                                 % len(existing))
                results.append(row)
                continue
            values = nominals_for(rec["dims"], rec.get("welds"), log)
            if not values:
                row["status"] = "no dimensions found on the drawing"
                results.append(row)
                continue
            written, dropped, stale = write_nominals(ws, values)
            touched += 1
            row["written"] = written
            row["values"] = values[:MAX_NOMINALS]
            # echo it back onto the part record so the text report can show it
            if not rec.get("filled"):
                rec["filled"] = {"tab": name, "values": row["values"]}
            row["stale"] = stale
            row["status"] = "filled %d nominal(s)%s%s" % (
                written,
                " - %d did NOT fit on the form, write them in by hand" % dropped if dropped else "",
                " - CHECK %s: min/max there was typed in by hand, so it will NOT follow the"
                " new nominal" % ", ".join(stale) if stale else "")
            results.append(row)

        if touched and not dry_run:
            _save_workbook(wb, workbook_path, log)
        elif touched:
            log("   [dry run] %d tab(s) would be filled - nothing saved" % touched)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return results


def _save_workbook(wb, dest, log):
    """Save via a temp file then replace, so a failed write can't shred the original."""
    import tempfile

    handle, tmp = tempfile.mkstemp(suffix=".xlsx", prefix="techdeck_insp_",
                                   dir=sdk.long_path(os.path.dirname(dest)))
    os.close(handle)
    try:
        wb.save(sdk.long_path(tmp))
        wb.close()
        os.replace(sdk.long_path(tmp), sdk.long_path(dest))
        log("   saved %s" % os.path.basename(dest))
    except PermissionError as exc:
        raise sdk.locked_file_error(dest, exc)
    finally:
        if os.path.exists(sdk.long_path(tmp)):
            try:
                os.remove(sdk.long_path(tmp))
            except Exception:
                pass


def _packet_pdf(folder):
    """The nest package PDF in a folder - the MOVE TICKET OMIT copy if there is one."""
    pdfs = [p for p in sorted(glob.glob(os.path.join(folder, "*.pdf")))
            if not os.path.basename(p).startswith("~$")]
    if not pdfs:
        return None
    omit = [p for p in pdfs if "MOVE TICKET OMIT" in os.path.basename(p).upper()]
    return omit[0] if omit else None


def is_nest_folder(folder):
    """A nest folder we can actually read: one holding a stamped packet PDF."""
    return _packet_pdf(folder) is not None


def _subfolders(folder):
    try:
        entries = sorted(os.listdir(folder))
    except OSError:
        return []
    return [os.path.join(folder, e) for e in entries
            if sdk.is_dir(os.path.join(folder, e))]


# The tree is  911 QTDR \ <order> \ <nest> \ <nest> MOVE TICKET OMIT.pdf , so the
# roots below sit TWO levels above the work. Named only to make the refusal say why.
_PROGRAM_ROOT_NAMES = {"911 qtdr", "922 qtdr production packages",
                       "902 qtdr production packages", "pilot program"}


def classify_pick(folder):
    """What did the user actually pick? "nest", "order", "loose" or "too_high".

    Discovery used to scan one level down and, failing that, sweep up every loose
    PDF in the picked folder. Point that at the 911 QTDR ROOT and the last branch
    fired: the order folders one level down hold no packet of their own, so the
    ten stray QC PDFs sitting beside them became ten jobs and the run started
    across the whole program (reported 2026-09-01). The shape is now decided
    STRUCTURALLY, and a folder that is neither a nest nor a holder of nests is
    refused outright instead of guessed at.
    """
    if is_nest_folder(folder):
        return "nest"
    subs = _subfolders(folder)
    if any(is_nest_folder(sub) for sub in subs):
        return "order"
    if os.path.basename(os.path.normpath(folder)).strip().lower() in _PROGRAM_ROOT_NAMES:
        return "too_high"
    # A plain folder of packet PDFs is only ever a LEAF. If a subfolder holds PDFs
    # of its own we are standing above the work, not in it - so refuse.
    loose = [p for p in glob.glob(os.path.join(folder, "*.pdf"))
             if not os.path.basename(p).startswith("~$")]
    if loose and not any(glob.glob(os.path.join(sub, "*.pdf")) for sub in subs):
        return "loose"
    return "too_high"


def discover_jobs(folder, shape=None):
    """Pair each packet PDF with its workbook, for the shape the user picked.

    Three shapes are valid and are told apart by what is on disk, so the user
    never has to say which one they meant:
      * an ORDER folder - nest folders one level down, each with a workbook,
      * a single NEST folder,
      * a plain folder of packet PDFs with no workbooks (read-only report).
    """
    jobs = []
    seen = set()

    def add(pdf, workbook, label):
        key = os.path.normcase(os.path.abspath(pdf))
        if key not in seen:
            seen.add(key)
            jobs.append({"pdf": pdf, "workbook": workbook, "label": label,
                         "nest": os.path.basename(os.path.dirname(pdf))})

    if shape is None:
        shape = classify_pick(folder)

    if shape == "nest":
        own = _packet_pdf(folder)
        if own:
            add(own, find_nest_workbook(folder),
                os.path.basename(os.path.normpath(folder)))

    elif shape == "order":
        for sub in _subfolders(folder):
            pdf = _packet_pdf(sub)
            if pdf:
                add(pdf, find_nest_workbook(sub),
                    os.path.basename(os.path.normpath(sub)))

    elif shape == "loose":
        for pdf in sorted(glob.glob(os.path.join(folder, "*.pdf"))):
            if not os.path.basename(pdf).startswith("~$"):
                add(pdf, None, os.path.splitext(os.path.basename(pdf))[0])
    return jobs


# ------------------------------------------------------------------------- report
def _fmt_dim(d):
    label = {"linear": "", "radius": "R ", "diameter": "DIA ",
             "angle": "", "chamfer": ""}.get(d["kind"], "")
    suffix = " deg" if d["kind"] == "angle" else ""
    mods = (" " + " ".join(dict.fromkeys(d["mods"]))) if d["mods"] else ""
    flag = "   <-- CHECK (low confidence)" if d["score"] < 0.55 else ""
    return "%s%s%s%s%s" % (label, d["value"], suffix, mods, flag)


def process_pdf(pdf_path, log, cancel_event, progress=None):
    """Read one nest package. Returns a list of part records."""
    import fitz

    sdk.ensure_local(pdf_path, log=log)
    doc = fitz.open(sdk.long_path(pdf_path))
    parts = []
    try:
        for index, page in enumerate(doc):
            sdk.raise_if_cancelled(cancel_event)
            text = page.get_text()
            clip = drawing_clip(page)
            is_sketch = "PART SKETCH" in text

            if is_sketch:
                info = title_block(page)
                record = dict(info)
                record.update({"page": index + 1, "dims": [], "notes": [],
                               "welds": [], "views": 0, "problem": ""})
                parts.append(record)
                if clip is None:
                    record["problem"] = (
                        "no drawing on the page - the PDF says the drawing image could "
                        "not be loaded when the packet was printed"
                        if "not reachable" in text else "no drawing image on the page"
                    )
                    log("   page %d  %s - %s" % (index + 1, info.get("part", "?"), record["problem"]))
                    continue
            elif clip is not None and parts:
                record = parts[-1]           # a continuation view of the part above
                record["views"] += 1
            else:
                continue

            dims, notes, welds = read_drawing(page, clip, log)
            record["dims"].extend(dims)
            for note in notes:
                if note not in record["notes"]:
                    record["notes"].append(note)
            for weld in welds:
                if not any(w["code"] == weld["code"] and w["side"] == weld["side"]
                           for w in record["welds"]):
                    record["welds"].append(weld)
            if progress:
                progress()
            keep = len([d for d in dims if not d["ref"]])
            log("   page %d  %-18s %d dimension(s)%s%s"
                % (index + 1, record.get("part", "?"), keep,
                   ", %d REF" % (len(dims) - keep) if len(dims) - keep else "",
                   ", weld prep %s" % "/".join(w["code"] for w in welds) if welds else ""))
    finally:
        doc.close()
    return parts


def build_report(folder, results, fills, elapsed, failures=None):
    """Render the whole run as the text file the user gets."""
    now = datetime.datetime.now()
    out = []
    add = out.append
    add("911 INSPECTION DIMENSIONS")
    add("=" * 78)
    add("Folder : %s" % folder)
    add("Run    : %s" % now.strftime("%Y-%m-%d %H:%M"))
    add("Reader : TechDeck 911 Inspection Dimensions v%s" % VERSION)
    add("")
    add("Dimensions and weld preps are read off the drawing picture, so CHECK them")
    add("against the paperwork before they go on an inspection sheet. Anything the")
    add("reader was unsure about is marked CHECK. Dimensions marked REF on the")
    add("drawing are listed under 'Reference only' and are NOT inspection")
    add("dimensions. Weld preps are the KB codes called out on the drawing, with")
    add("the side each one applies to (NS = near side, FS = far side).")
    add("")

    total_parts = total_dims = total_ref = total_welds = 0
    problems = []
    low_conf = []
    bevel_problems = []

    for pdf_name, parts in results:
        add("=" * 78)
        add("PDF: %s" % pdf_name)
        add("=" * 78)
        if not parts:
            # An unreadable PDF is NOT an empty one - saying "no PART SKETCH pages"
            # for a packet that actually blew up sent the reader hunting the packet
            # instead of the error (FTOURIGNY-LT, 2026-09-01).
            reason = (failures or {}).get(pdf_name)
            if reason:
                add("  ** this PDF could not be read: %s" % reason)
            else:
                add("  (no PART SKETCH pages found)")
            add("")
            continue
        for rec in parts:
            total_parts += 1
            add("")
            add("  PART %s" % (rec.get("part") or "(part number not readable)"))
            meta = []
            if rec.get("work_order"):
                meta.append("Work order %s" % rec["work_order"])
            if rec.get("rev"):
                meta.append("Rev/Seq %s" % rec["rev"])
            if rec.get("qty"):
                meta.append("Qty %s" % rec["qty"])
            if rec.get("noun"):
                meta.append(rec["noun"].title())
            add("    %s" % ("  |  ".join(meta) if meta else "-"))
            if rec.get("size"):
                add("    Size     : %s" % rec["size"])
            if rec.get("fab_dim"):
                add("    FAB DIM  : %s" % rec["fab_dim"])
            add("    Sketch page %d%s"
                % (rec["page"], " (+%d extra view page(s))" % rec["views"] if rec["views"] else ""))

            if rec["problem"]:
                add("    ** %s" % rec["problem"])
                problems.append("%s - %s: %s" % (pdf_name, rec.get("part", "?"), rec["problem"]))
                continue

            keep = [d for d in rec["dims"] if not d["ref"]]
            refs = [d for d in rec["dims"] if d["ref"]]
            total_dims += len(keep)
            total_ref += len(refs)
            add("")
            if keep:
                add("    DIMENSIONS (%d)" % len(keep))
                for d in keep:
                    add("      - %s" % _fmt_dim(d))
                    if d["score"] < 0.55:
                        low_conf.append("%s - %s: %s" % (pdf_name, rec.get("part", "?"), d["value"]))
            else:
                add("    DIMENSIONS (0)   ** nothing readable - check this drawing by hand **")
                problems.append("%s - %s: no dimensions read" % (pdf_name, rec.get("part", "?")))
            if rec["welds"]:
                total_welds += len(rec["welds"])
                add("    WELD PREPS (%d)" % len(rec["welds"]))
                for weld in rec["welds"]:
                    side = ("  %s" % weld["side"]) if weld["side"] else "  (no side marked)"
                    flag = "   <-- CHECK (low confidence)" if weld["score"] < 0.55 else ""
                    add("      - %s%s%s" % (weld["code"], side, flag))
                    if weld["score"] < 0.55:
                        low_conf.append("%s - %s: weld prep %s"
                                        % (pdf_name, rec.get("part", "?"), weld["code"]))
                    # what the bevel book says this code is, and what went on the sheet
                    row = bevel_lookup(weld["code"])
                    if row is None:
                        near = nearest_bevel_code(weld["code"])
                        hint = (" - closest code in the book is %s, CHECK the drawing"
                                % near) if near else ""
                        add("          NOT IN THE BEVEL BOOK%s - nothing written, "
                            "look it up by hand" % hint)
                        bevel_problems.append(
                            "%s - %s: %s is not in the bevel book%s"
                            % (pdf_name, rec.get("part", "?"), weld["code"], hint))
                        continue
                    bits = []
                    if row["ns_angle"]:
                        bits.append("NS %s deg" % row["ns_angle"])
                    if row["fs_angle"]:
                        bits.append("FS %s deg" % row["fs_angle"])
                    if row["max_land"]:
                        bits.append("land %s" % row["max_land"])
                    add("          bevel book: %s%s"
                        % (row["type"] or "?", ("  -  " + ", ".join(bits)) if bits else ""))
                    if row["status"] == "VOID":
                        add("          *** %s IS VOID ***%s - nothing written"
                            % (weld["code"],
                               " use %s instead" % row["replacement"] if row["replacement"] else ""))
                        bevel_problems.append(
                            "%s - %s: %s is VOID%s"
                            % (pdf_name, rec.get("part", "?"), weld["code"],
                               " - use %s" % row["replacement"] if row["replacement"] else ""))
                    elif row["status"] == "NO SKETCH":
                        add("          *** %s has no sketch in the book *** - nothing written"
                            % weld["code"])
                        bevel_problems.append("%s - %s: %s has no sketch"
                                              % (pdf_name, rec.get("part", "?"), weld["code"]))
                    else:
                        angles = weld_prep_angles(weld)
                        if angles:
                            add("          -> wrote %s onto the inspection sheet"
                                % ", ".join(angles))
                        else:
                            # A taper-only sheet (KB702-729, KB738) or one cut square
                            # (SB610) genuinely prints no bevel angle. Nothing to
                            # write, but the reviewer has to be told - a prep that
                            # quietly contributes nothing looks the same on the form
                            # as one that was never on the drawing.
                            add("          -> nothing written: this sheet prints no "
                                "bevel angle (taper only / cut square)")
                            bevel_problems.append(
                                "%s - %s: %s prints no bevel angle (%s) - nothing written"
                                % (pdf_name, rec.get("part", "?"), weld["code"],
                                   row["type"] or "no type"))
            if rec.get("filled"):
                add("    -> written to inspection tab %s: %s"
                    % (rec["filled"]["tab"], ", ".join(str(v) for v in rec["filled"]["values"])))
            if refs:
                add("    Reference only (not inspected): %s"
                    % ", ".join(_fmt_dim(d).replace("   <-- CHECK (low confidence)", "") for d in refs))
            if rec["notes"]:
                add("    Drawing notes:")
                for note in rec["notes"]:
                    add("      %s" % note)
        add("")

    if fills:
        add("=" * 78)
        add("INSPECTION SHEETS FILLED IN")
        add("=" * 78)
        add("Nominals are written into the colour-filled TARGET cells only. Excel works")
        add("out MIN and MAX from each one by itself, so those are left alone.")
        add("A tab that already had numbers typed into it was NOT touched.")
        add("")
        add("TYP dimensions are written LAST in each part's run. The drawing prints them")
        add("once but the feature repeats, and it never says how many times - so copy the")
        add("last entries across as many times as the part actually needs.")
        add("")
        for book, path, rows in fills:
            add("  %s" % book)
            for row in rows:
                add("    %-16s %-18s %s" % (row["tab"], row["part"], row["status"]))
            add("")

    add("=" * 78)
    add("SUMMARY")
    add("=" * 78)
    add("PDFs read           : %d" % len(results))
    add("Parts found         : %d" % total_parts)
    add("Dimensions logged   : %d" % total_dims)
    add("Reference (skipped) : %d" % total_ref)
    add("Weld preps logged   : %d" % total_welds)
    add("Sheet tabs filled   : %d" % sum(1 for _, _, rows in fills for r in rows if r.get("written")))
    add("Time                : %.0f seconds" % elapsed)
    if problems:
        add("")
        add("NEEDS A HUMAN LOOK (%d)" % len(problems))
        for p in problems:
            add("  - %s" % p)
    if bevel_problems:
        add("")
        add("-" * 78)
        add("WELD PREPS THAT NEED A DECISION (%d) - nothing was written for these"
            % len(bevel_problems))
        add("-" * 78)
        for p in bevel_problems:
            add("  - %s" % p)

    if low_conf:
        add("")
        add("LOW CONFIDENCE READINGS (%d) - verify these against the drawing" % len(low_conf))
        for p in low_conf:
            add("  - %s" % p)
    return "\n".join(out) + "\n"


_PICK_TOO_HIGH = (
    "There are no readable nests in that folder, and nothing one level down "
    "either - so it is above the work, or its nests have no MOVE TICKET OMIT "
    "PDF yet.",
    "Pick ONE of:\n"
    "  * an order folder (the one holding the nest folders) - you then tick "
    "which nests to read, or\n"
    "  * a single nest folder (the one with the MOVE TICKET OMIT PDF in it).\n\n"
    "The 911 QTDR root is never accepted: every order under it would be read "
    "in one go.",
)


def choose_nests(params, folder, jobs, log):
    """Tick which nests to read. Returns the kept jobs, or None if cancelled.

    Shown for every multi-nest pick. Picking a single NEST folder is the one
    case that skips it - there is nothing to choose between.
    """
    labels, seen = [], {}
    for job in jobs:
        label = job["label"]
        if not job["workbook"]:
            label += "   (no inspection workbook - report only)"
        # SelectionDialog hands back the chosen STRINGS, so duplicates would be
        # indistinguishable coming back. Number them instead (as 911 Teams Cards does).
        if label in seen:
            seen[label] += 1
            label = "%s  (#%d)" % (label, seen[label])
        else:
            seen[label] = 1
        job["_pick"] = label
        labels.append(label)

    picked = sdk.request_selection(
        params, labels, None,
        window_title="911 Inspection Dimensions",
        header="Select Nests to Read",
        root_label="All nests in %s" % (os.path.basename(os.path.normpath(folder)) or folder),
        noun="nest",
        prompt_note=("Every nest found is ticked. Untick any you do not want read "
                     "yet - nothing in them is opened or changed."),
        run_button_text="Read Selected",
    )
    if picked is None:
        return None            # sdk.request_selection already flagged the cancel

    keep = set(picked)
    chosen = [j for j in jobs if j["_pick"] in keep]
    dropped = len(jobs) - len(chosen)
    if dropped:
        log("Skipping %d nest(s) you unticked - nothing in them was touched." % dropped)
    return chosen


# ---------------------------------------------------------------------------- run
def run(params, progress_callback, cancel_event):
    import time

    log = params.get("log", print)
    settings = params.get("settings", {})
    console = params.get("console")

    folder = sdk.request_directory(
        params,
        "Select the folder of nest PDFs to read",
        start_dir=settings.get("base_path", "") or "",
    )
    if not folder or cancel_event.is_set():
        cancel_event.set()
        return

    dry_run = bool(settings.get("dry_run"))
    if dry_run:
        log("DRY RUN - the report is written but no workbook is changed.")

    shape = classify_pick(folder)
    if shape == "too_high":
        raise sdk.UserFacingError(*_PICK_TOO_HIGH)

    jobs = discover_jobs(folder, shape)
    if not jobs:
        raise sdk.UserFacingError(
            "There are no nest package PDFs in that folder.",
            "Pick an order folder (the one holding the nest folders), a single nest "
            "folder, or a folder of nest package PDFs, then run it again.",
        )

    # A single nest folder is unambiguous; anything wider gets the tick-list first.
    if shape != "nest":
        jobs = choose_nests(params, folder, jobs, log)
        if jobs is None:
            cancel_event.set()
            return
        if not jobs:
            log("No nests were ticked - nothing to read.")
            cancel_event.set()
            return

    pdfs = [j["pdf"] for j in jobs]
    fillable = [j for j in jobs if j["workbook"]]

    log("Reading %d nest package(s) from %s" % (len(jobs), folder))
    if fillable:
        log("%d of them have a nest workbook - their inspection sheets will be filled in."
            % len(fillable))
    else:
        log("No nest workbooks alongside them, so this run is a read-only report.")
    log("")

    # rough page budget so the bar moves sensibly across the whole run
    import fitz

    total_pages = 0
    for path in pdfs:
        sdk.raise_if_cancelled(cancel_event)
        try:
            sdk.ensure_local(path, log=log)
            with fitz.open(sdk.long_path(path)) as doc:
                total_pages += sum(1 for pg in doc if drawing_clip(pg) is not None)
        except Exception:
            total_pages += 1
    done = [0]

    def tick():
        done[0] += 1
        progress_callback(min(97, int(5 + 92.0 * done[0] / max(1, total_pages))))

    started = time.time()
    results = []
    failed = []
    failures = {}
    fills = []
    for job in jobs:
        sdk.raise_if_cancelled(cancel_event)
        path = job["pdf"]
        name = os.path.basename(path)
        log("%s" % name)
        try:
            parts = process_pdf(path, log, cancel_event, tick)
        except sdk.PluginCancelled:
            raise
        except sdk.UserFacingError:
            raise
        except Exception as exc:
            log("   ! could not read this PDF: %s" % exc)
            failed.append("%s (%s)" % (name, exc))
            failures[name] = str(exc)
            results.append((name, []))
            continue
        results.append((name, parts))

        if not job["workbook"]:
            continue
        sdk.raise_if_cancelled(cancel_event)
        book = os.path.basename(job["workbook"])
        try:
            rows = fill_nest_workbook(job["workbook"], parts, log, dry_run=dry_run)
        except sdk.PluginCancelled:
            raise
        except sdk.UserFacingError as exc:
            log("   ! %s" % exc)
            failed.append("%s (%s)" % (book, exc))
            continue
        except Exception as exc:
            log("   ! could not fill %s: %s" % (book, exc))
            failed.append("%s (%s)" % (book, exc))
            continue
        fills.append((book, job["workbook"], rows))
        for row in rows:
            log("   %-14s %-18s %s" % (row["tab"], row["part"], row["status"]))
    elapsed = time.time() - started

    stamp = datetime.datetime.now().strftime("%Y-%m-%d")
    out_name = "911 Inspection Dimensions - %s - %s.txt" % (os.path.basename(folder.rstrip("\\/")), stamp)
    out_path = os.path.join(folder, out_name)
    with open(sdk.long_path(out_path), "w", encoding="utf-8") as handle:
        handle.write(build_report(folder, results, fills, elapsed, failures))

    parts = sum(len(p) for _, p in results)
    dims = sum(len([d for d in rec["dims"] if not d["ref"]]) for _, p in results for rec in p)
    filled_tabs = sum(1 for _, _, rows in fills for r in rows if r.get("written"))
    log("")
    log("Read %d part(s), logged %d dimension(s) in %.0f seconds." % (parts, dims, elapsed))
    if fills:
        log("Filled %d inspection sheet tab(s) across %d workbook(s)." % (filled_tabs, len(fills)))
    log("Saved: %s" % out_path)

    if console is not None and hasattr(console, "append_link"):
        try:
            console.append_link(out_name, out_path, prefix="REPORT", at_run_end=True)
        except TypeError:
            console.append_link(out_name, out_path)

    if failed and hasattr(sdk, "set_run_outcome"):
        sdk.set_run_outcome(
            params, sdk.RUN_OUTCOME_WARNING,
            "%d PDF(s) could not be read: %s" % (len(failed), "; ".join(failed)),
        )
    progress_callback(100)


if __name__ == "__main__":  # headless harness
    import sys
    import threading

    target = sys.argv[1] if len(sys.argv) > 1 else "."
    sdk.request_directory = lambda *a, **k: target
    run({"log": print, "settings": {}, "console": None}, lambda p: None, threading.Event())
