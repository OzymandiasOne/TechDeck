"""
911 Baked Beans Wild Ride Plugin for TechDeck
=============================================

Consolidates the per-part "NC style baked beans" calc workbooks (the
DSTV/linear-inch pricing sheets Antonio & Scott J's template fills out from
Vacam NC files) into ONE review list, so nobody has to open 60+ workbooks
one at a time. Named for Mr. Bones' Wild Ride (the RollerCoaster Tycoon
greentext saga): quoted rider thoughts mutter through the console during the run,
and the finish line — 'Mr Beans says: "The ride never ends!"' — is CLICKABLE
(console.append_link) and opens the bundled mr_beans_wild_ride.jpg, both
halves of the original saga merged into one long image.

  0. Opens a process picker (same SelectionDialog window as 911 Setup /
     902 DXF Prep) with two stops on the ride: "Run NC Style Baked Beans
     spreadsheet" — GREYED OUT, a teaser for when the manual-review workflow
     is ironed out enough to automate — and the consolidation, checked.
  1. Prompts for the folder holding the filled calc sheets via the native
     folder dialog — the same folder that holds the NC files; the output is
     saved right back into it (per the requester: "output to the same folder
     path that has all the nc files in it").
  2. For every .xlsm/.xlsx in the folder, reads off the summary (REF) sheet:
       - the DYPN — the "Part #"-labeled value (C6 today; shape-scan then
         file-name fallback, cross-checked against the file name with a
         mismatch flagged in the Notes column),
       - the BATCH + NEST — the "Batch/Nest"-labeled value ("  V085 P07819"
         at C3 today: batch token then nest token; two-token shape-scan
         fallback; missing gets flagged, not fatal), and
       - the three YELLOW totals: TOTAL Bevels / TOTAL Complex Bevels /
         TOTAL Cut Lin. Labels are FOUND BY TEXT anywhere on the sheet, never
         by fixed cell address (Hard Rules 1-2) — the template's label wording
         already drifted between versions, so each value is the first numeric
         cell to the right of its matched label.
  3. Writes "{batch} {nest} Consolidated NC Calcs.xlsx" into that folder —
     batch/nest come FROM the sheets (single pair → its name; several →
     "MULTI BATCH" / "{n} NESTS"; none anywhere → prompts as fallback). One
     row per part: DYPN | Batch | Nest | Total Bevels | Total Complex Bevels
     | Total Cut Lin | Notes, sorted batch → nest → DYPN (all natural order)
     so each nest's parts sit together, with a live =SUM() totals row.

Values come from openpyxl's CACHED formula results (data_only=True). A sheet
that was never recalculated/saved in Excel has no cache — those rows are
flagged instead of silently written as zero.

Future ask (not built yet): also RUN the template against raw NC files and
save the per-part workbooks. Blocked on the template's manual-review step
(unprogrammed bevels/features are still keyed in by hand).
"""

from __future__ import annotations

import datetime
import random
import re
from pathlib import Path

# SDK bootstrap (works in-process and for standalone CLI testing).
try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VERSION = "1.2.3"

# --- The ride ---------------------------------------------------------------
STEP_RUN_SHEET = "Run NC Style Baked Beans spreadsheet"
STEP_CONSOLIDATE = "Consolidate output sheets into one review list"

_THOUGHT_MAIN = '"I want off BAKED BEANS WILD RIDE"'
_THOUGHT_RARE = '"I have the strangest feeling someone is watching me"'
_FINALE_PREFIX = "Mr Beans says:"   # bold channel-tag style (egg purple)
_FINALE = '"The ride never ends!"'  # regular body text, but clickable
_RIDE_IMAGE = "mr_beans_wild_ride.jpg"   # bundled next to run.py
_THOUGHT_EVERY = 3                        # emit a rider thought every N files
_RARE_CHANCE = 0.125                      # odds a thought is the watching one

# The three yellow totals, keyed in the requester's column order. Anchored so
# "TOTAL Bevels" can never match "TOTAL Complex Bevels".
_TOTAL_PATTERNS = [
    ("Total Bevels (in)", re.compile(r"^TOTAL\s+BEVELS?\b", re.IGNORECASE)),
    ("Total Complex Bevels (in)", re.compile(r"^TOTAL\s+COMPLEX\s+BEVELS?\b", re.IGNORECASE)),
    ("Total Cut Lin (in)", re.compile(r"^TOTAL\s+CUT\s+LIN\b", re.IGNORECASE)),
]

# DYPN-shaped cell, e.g. "R4533287-109" (leading/trailing spaces tolerated).
_DYPN_RE = re.compile(r"^[A-Z]{0,3}\d{4,}-\d{1,4}$", re.IGNORECASE)

# The REF sheet's labeled fields: "Batch/Nest" -> "  V085 P07819" (batch then
# nest) and "Part #" -> the DYPN. Located by LABEL text, never a fixed cell;
# shape-scans below are the fallback for template versions without the labels.
_BATCHNEST_LABEL_RE = re.compile(r"^BATCH\s*/?\s*NEST$", re.IGNORECASE)
_PART_LABEL_RE = re.compile(r"^PART\s*#?$", re.IGNORECASE)
# Nest shape = the Hard-Rule-3 nest regex (numeric or alphanumeric-with-digit).
_NEST_RE = re.compile(r"^(?:[PS]?\d{3,}|(?=[A-Z0-9]*\d)[A-Z0-9]{4,8})$", re.IGNORECASE)

# How far into a sheet we look for the summary labels / the DYPN. The summary
# block sits in the top rows; generous bounds keep a moved block findable
# without scanning 300+ contour rows on every sheet.
_SCAN_ROWS = 40
_SCAN_COLS = 30
_DYPN_SCAN_ROWS = 15

_SKIP_NAME_HINT = "consolidated nc calcs"  # never re-ingest our own output


def _find_summary_sheet(wb):
    """The sheet holding the totals block — 'REF' in the current template, but
    located by content (the TOTAL Cut Lin label) so a rename survives."""
    ordered = sorted(wb.worksheets, key=lambda ws: ws.title.upper() != "REF")
    cut_lin_re = _TOTAL_PATTERNS[2][1]
    for ws in ordered:
        for row in ws.iter_rows(min_row=1, max_row=_SCAN_ROWS, max_col=_SCAN_COLS):
            for cell in row:
                if isinstance(cell.value, str) and cut_lin_re.match(cell.value.strip()):
                    return ws
    return None


def _value_right_of(ws, row_idx: int, col_idx: int):
    """First numeric cell to the right of a label cell (labels and their
    values are separated by merge padding that varies between versions)."""
    for cells in ws.iter_rows(min_row=row_idx, max_row=row_idx,
                              min_col=col_idx + 1, max_col=col_idx + 8):
        for cell in cells:
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return float(v)
    return None


def _string_right_of(ws, row_idx: int, col_idx: int):
    """First non-empty string cell to the right of a label cell."""
    for cells in ws.iter_rows(min_row=row_idx, max_row=row_idx,
                              min_col=col_idx + 1, max_col=col_idx + 8):
        for cell in cells:
            v = cell.value
            if isinstance(v, str) and v.strip():
                return v.strip()
    return None


def _extract_part(ws, file_stem: str):
    """Pull (dypn, batch, nest, {column: value}, notes) off a summary sheet."""
    notes = []
    totals = {name: None for name, _ in _TOTAL_PATTERNS}
    dypn = batch = nest = None
    dypn_shape = bn_shape = None

    for row in ws.iter_rows(min_row=1, max_row=_SCAN_ROWS, max_col=_SCAN_COLS):
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if not s:
                continue
            # Labeled fields (preferred)
            if dypn is None and _PART_LABEL_RE.match(s):
                right = _string_right_of(ws, cell.row, cell.column)
                if right and _DYPN_RE.match(right):
                    dypn = right.upper()
            if batch is None and _BATCHNEST_LABEL_RE.match(s):
                right = _string_right_of(ws, cell.row, cell.column)
                if right:
                    toks = right.upper().split()
                    batch = toks[0] if toks else None
                    nest = toks[1] if len(toks) > 1 else None
            # Shape fallbacks (older templates without the labels)
            if cell.row <= _DYPN_SCAN_ROWS:
                if dypn_shape is None and _DYPN_RE.match(s):
                    dypn_shape = s.upper()
                if bn_shape is None:
                    toks = s.upper().split()
                    if (len(toks) == 2 and toks[0].isalnum()
                            and _NEST_RE.match(toks[1])):
                        bn_shape = (toks[0], toks[1])
            # The three yellow totals
            for name, pat in _TOTAL_PATTERNS:
                if totals[name] is None and pat.match(s):
                    totals[name] = _value_right_of(ws, cell.row, cell.column)

    if dypn is None:
        dypn = dypn_shape
    if batch is None and bn_shape:
        batch, nest = bn_shape

    if dypn is None:
        dypn = file_stem.upper()
        notes.append("DYPN cell not found - used file name")
    elif dypn != file_stem.upper().strip():
        notes.append(f"DYPN differs from file name ({file_stem})")
    if batch is None:
        notes.append("Batch/Nest cell not found")
    elif nest is None:
        notes.append("nest missing from the Batch/Nest cell")

    missing = [name for name, v in totals.items() if v is None]
    if len(missing) == len(totals):
        notes.append("no cached totals - open the file in Excel, let it "
                     "calculate, and save")
    elif missing:
        notes.append("missing: " + ", ".join(m.replace(" (in)", "") for m in missing))

    return dypn, batch or "", nest or "", totals, notes


def _dypn_sort_key(dypn: str):
    """Natural sort: R4533287-110 after R4533287-109, not lexicographically."""
    m = re.match(r"^(.*?)(\d+)$", dypn)
    if m:
        return (m.group(1), int(m.group(2)))
    return (dypn, -1)


def _write_output(rows: list, out_path: Path, log) -> Path:
    """rows = [(dypn, batch, nest, {column: value}, notes)] already sorted
    (batch -> nest -> DYPN, so same-nest parts sit together)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    yellow = PatternFill("solid", start_color="FFFF00")
    bold = Font(bold=True)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = (["DYPN", "Batch", "Nest"]
               + [name for name, _ in _TOTAL_PATTERNS] + ["Notes"])
    val_first, val_last = 4, 3 + len(_TOTAL_PATTERNS)   # D..F
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=text)
        c.font = bold
        c.border = box
        c.alignment = Alignment(horizontal="center")
        if val_first <= col <= val_last:   # the value columns echo the source yellow
            c.fill = yellow

    for r, (dypn, batch, nest, totals, notes) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=dypn).border = box
        ws.cell(row=r, column=2, value=batch or None).border = box
        ws.cell(row=r, column=3, value=nest or None).border = box
        for col, (name, _) in enumerate(_TOTAL_PATTERNS, start=val_first):
            c = ws.cell(row=r, column=col, value=totals[name])
            c.border = box
            c.number_format = "0.000"
        ws.cell(row=r, column=val_last + 1,
                value="; ".join(notes) if notes else None).border = box

    # Live grand-total row so the sums stay honest if a value is corrected.
    total_row = len(rows) + 2
    label = ws.cell(row=total_row, column=1, value=f"TOTALS ({len(rows)} parts)")
    label.font = bold
    label.border = box
    for col in (2, 3, val_last + 1):
        ws.cell(row=total_row, column=col).border = box
    for col in range(val_first, val_last + 1):
        letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUM({letter}2:{letter}{total_row - 1})")
        c.font = bold
        c.border = box
        c.number_format = "0.000"
        c.fill = yellow

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{total_row - 1}"
    widths = {"A": 18, "B": 10, "C": 12, "D": 17, "E": 24, "F": 17, "G": 46}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        # Most likely the previous list is open in Excel — save alongside it.
        stamp = datetime.datetime.now().strftime("%H%M%S")
        alt = out_path.with_name(f"{out_path.stem} ({stamp}){out_path.suffix}")
        log(f"[WARN] {out_path.name} is locked (open in Excel?) - saving as "
            f"{alt.name} instead.")
        wb.save(alt)
        return alt


def _choose_process(params, log) -> bool:
    """The boarding gate: a picker matching 911 Setup's window. Only the
    consolidation is selectable — running the spreadsheet itself is shown
    greyed out so users can see where the track is headed. Returns True to
    proceed, False on cancel. Headless (no console) boards straight away."""
    console = params.get("console")
    if console is None or not hasattr(console, "request_selection"):
        return True
    picks = console.request_selection(
        [STEP_RUN_SHEET, STEP_CONSOLIDATE],
        None,
        window_title="911 Baked Beans Wild Ride",
        header="Select Processes to Run",
        root_label="All processes",
        noun="process",
        prompt_note=("Running the spreadsheet itself is a future stop — "
                     "coming once the manual review steps are ironed out."),
        run_button_text="Board the Ride",
        disabled_items={STEP_RUN_SHEET},
        disabled_label="coming soon",
    )
    if picks is None or STEP_CONSOLIDATE not in picks:
        log("Nobody boarded the ride.")
        return False
    return True


def run(params: dict, progress_callback, cancel_event):
    log = params.get("log", print)
    log(f"911 Baked Beans Wild Ride v{VERSION}")

    if not _choose_process(params, log):
        return

    folder_str = sdk.request_directory(
        params, "Select the folder holding the filled NC calc spreadsheets")
    if not folder_str:
        log("No folder selected - nothing to do.")
        return
    folder = Path(folder_str)
    progress_callback(5)

    candidates = sorted(
        p for pattern in ("*.xlsm", "*.xlsx") for p in folder.glob(pattern)
        if not p.name.startswith("~$")
        and _SKIP_NAME_HINT not in p.name.lower()
    )
    if not candidates:
        log(f"No .xlsm/.xlsx calc sheets found in {folder}")
        return
    log(f"Found {len(candidates)} workbook(s) in {folder.name} - "
        f"{len(candidates)} riders boarding...")

    rows, skipped = [], []
    for i, path in enumerate(candidates):
        if cancel_event is not None and cancel_event.is_set():
            log("Cancelled. (Some of them even want to get off.)")
            return
        if (i + 1) % _THOUGHT_EVERY == 0:
            log(_THOUGHT_RARE if random.random() < _RARE_CHANCE else _THOUGHT_MAIN)
        wb = None
        try:
            wb = sdk.load_workbook_resilient(
                path, log=log, data_only=True, read_only=True)
            ws = _find_summary_sheet(wb)
            if ws is None:
                skipped.append(f"{path.name}: no totals block found (not a calc sheet?)")
                continue
            dypn, batch, nest, totals, notes = _extract_part(ws, path.stem)
            rows.append((dypn, batch, nest, totals, notes))
            if notes:
                log(f"  [FLAG] {path.name}: {'; '.join(notes)}")
        except Exception as exc:
            skipped.append(f"{path.name}: {exc}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
        progress_callback(5 + int(85 * (i + 1) / len(candidates)))

    if not rows:
        log("No calc sheets could be read - nothing to consolidate.")
        for s in skipped:
            log(f"  [SKIPPED] {s}")
        return

    # Grouped sort: batch -> nest -> DYPN (all natural), rows missing a
    # batch/nest sink to the bottom of their group.
    rows.sort(key=lambda r: (_dypn_sort_key(r[1] or "~"),
                             _dypn_sort_key(r[2] or "~"),
                             _dypn_sort_key(r[0])))

    # Output name comes from the workbooks' own Batch/Nest cells; prompts are
    # only the fallback when no file declared one.
    batches = sorted({r[1] for r in rows if r[1]}, key=_dypn_sort_key)
    nests = sorted({r[2] for r in rows if r[2]}, key=_dypn_sort_key)
    if batches:
        bpart = batches[0] if len(batches) == 1 else "MULTI BATCH"
        npart = ((nests[0] if len(nests) == 1 else f"{len(nests)} NESTS")
                 if nests else "NO NEST")
        log(f"Batch/nest from the sheets: {', '.join(batches)} / "
            f"{', '.join(nests) if nests else '(none)'}")
    else:
        log("No sheet declared a Batch/Nest - asking for the output name.")
        bpart = (sdk.request_batch_number(params, "Enter batch number:") or "").strip()
        npart = (sdk.request_text(params, "Enter nest number:") or "").strip()
        if not bpart or not npart:
            log("Batch and nest numbers are required (they name the output file).")
            return
    safe = re.compile(r'[<>:"/\\|?*]')
    out_path = folder / (f"{safe.sub('_', bpart)} {safe.sub('_', npart)} "
                         "Consolidated NC Calcs.xlsx")

    saved = _write_output(rows, out_path, log)
    progress_callback(100)

    flagged = sum(1 for row in rows if row[4])
    log("")
    log(f"Consolidated {len(rows)} part(s) -> {saved}")
    if flagged:
        log(f"[REVIEW] {flagged} row(s) flagged - see the Notes column.")
    for s in skipped:
        log(f"  [SKIPPED] {s}")

    # The finish line. Clicking it opens the original Mr Bones' Wild Ride
    # saga (bundled, both halves merged into one long jpg).
    console = params.get("console")
    ride_img = Path(__file__).resolve().parent / _RIDE_IMAGE
    if (console is not None and hasattr(console, "append_link")
            and ride_img.exists()):
        try:
            console.append_link(_FINALE, str(ride_img), prefix=_FINALE_PREFIX)
        except TypeError:
            # Console predates the prefix argument (an older/still-running
            # TechDeck) - the finale must never fail the run.
            console.append_link(f"{_FINALE_PREFIX} {_FINALE}", str(ride_img))
    else:
        log(f"{_FINALE_PREFIX} {_FINALE}")
