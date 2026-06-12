"""
911 PO PDF Extractor Plugin for TechDeck v2.3.0
Extracts PO (Purchase Order) information from PDF packets into Excel.

NEW in v2.3.0:
- Updated column structure: 16 columns (A-P) with proper ordering
- Added blank placeholder columns: NEST (J), SOURCE (K), CUSTOMER (N), INTERCOMPANY (O), SUPPLIER (P)
- Renamed columns to match new headers (PO NO, LINE, PART REV, ORDER, BATCH, etc.)
- Added validation: missing items spreadsheet (.xlsx)
- Added validation: summary analysis text file (.txt)
- All outputs saved in dedicated folder named after Excel file

Previous updates:
- Added "Nest" column (extracted from Rev field - code after rev number like WJ182)
- Rev field now only contains part type and rev number (e.g., "OFLD-PLTCUT - Rev 108")
- Consolidated clause numbers into comma-separated string (one row per line)
- Auto-appends .xlsx extension if not provided

Prompts via console for:
- Folder path containing PO packet PDFs
- Output Excel filename (auto-adds .xlsx extension)
"""

import re
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
import threading

# Third-party imports
import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk


# ===== CONSTANTS =====
VERSION = "2.3.0"

# Month name to number mapping
MONTH_MAP = {
    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
}

# Column mapping: Excel column letter -> Header name
COLUMN_MAPPING = {
    'A': 'PO NO',
    'B': 'LINE',
    'C': 'PROMISE DATE',
    'D': 'DYPN',
    'E': 'Qty',
    'F': 'UNIT PRICE',
    'G': 'PART REV',
    'H': 'ORDER',
    'I': 'BATCH',
    'J': 'NEST',           # Blank placeholder
    'K': 'SOURCE',         # Blank placeholder
    'L': 'STANDARD CLAUSES',
    'M': 'SHIP',
    'N': 'CUSTOMER',       # Blank placeholder
    'O': 'INTERCOMPANY',   # Blank placeholder
    'P': 'SUPPLIER'        # Blank placeholder
}

# Reverse mapping for validation
HEADER_TO_COLUMN = {v: k for k, v in COLUMN_MAPPING.items()}

# Required columns (excluding blank placeholders)
REQUIRED_COLUMNS = [
    'PO NO', 'LINE', 'PROMISE DATE', 'DYPN', 'Qty', 'UNIT PRICE',
    'PART REV', 'ORDER', 'BATCH', 'STANDARD CLAUSES', 'SHIP'
]

# Compiled regex patterns
ORDER_PATTERN = re.compile(r'ORDER:\s*(\d+),\s*\d+')
ORDER_ALT_PATTERN = re.compile(r'ORDER\s+(\d+),\s+LINE NUMBER')
LINE_NO_PATTERN = re.compile(r'LINE\s+NO\s+PART NO / REVISION', re.DOTALL)
DATE_PATTERN = re.compile(r'(\d+)-([A-Z]+)-(\d{4})')


# ===== HELPER FUNCTIONS =====
def get_console_input(params: Dict[str, Any], prompt: str) -> str:
    """Get input from user via TechDeck console."""
    console = params.get('console')
    if console and hasattr(console, 'request_input'):
        return console.request_input(prompt)
    else:
        return input(f"{prompt}: ")


def format_promise_date(date_str: str) -> str:
    """Convert date from '21-OCT-2025' to 'MM/DD/YYYY'."""
    if not date_str or len(date_str) < 11:
        return ''
    
    match = DATE_PATTERN.match(date_str.strip())
    if not match:
        return date_str
    
    day = int(match.group(1))
    month_str = match.group(2)
    year = match.group(3)
    
    month = MONTH_MAP.get(month_str, 0)
    if month == 0:
        return date_str
    
    return f"{month:02d}/{day:02d}/{year}"


def extract_order_number(text: str) -> Optional[str]:
    """Extract order number from page text."""
    match = ORDER_PATTERN.search(text)
    if match:
        return match.group(1)
    return None


def build_line_index(doc: fitz.Document, log) -> Tuple[str, Dict[int, List[int]]]:
    """Build index mapping line numbers to pages."""
    line_index = {}
    order_no = None
    
    for page_idx, page in enumerate(doc):
        text = page.get_text("text")
        
        if not order_no:
            order_no = extract_order_number(text)
        
        line_pattern = r'LINE\s+NO\s+PART NO / REVISION.*?\n(\d{1,2})\s*\n\s*([\w-]+)\s*-\s*Rev'
        line_matches = re.finditer(line_pattern, text, re.DOTALL)
        
        for match in line_matches:
            line_no = int(match.group(1))
            if line_no >= 1:
                if line_no not in line_index:
                    line_index[line_no] = []
                if page_idx not in line_index[line_no]:
                    line_index[line_no].append(page_idx)
    
    return order_no, line_index


def extract_ship_to_address_once(all_page_texts: List[str]) -> str:
    """Extract Ship To address from first occurrence."""
    for page_text in all_page_texts:
        ship_pattern = r'SHIP TO LOCATION\s*\n(.*?)(?:\n\s*PAY ITEM|$)'
        match = re.search(ship_pattern, page_text, re.DOTALL)
        
        if not match:
            continue
        
        address_text = match.group(1).strip()
        address_lines = []
        
        for line in address_text.split('\n'):
            line_clean = line.strip()
            
            if not line_clean:
                continue
            
            skip_patterns = [
                'PROMISE', 'DATE', 'CONTRACT', 'DELIVERY', 'NEED BY',
                'DESCRIPTION', 'QUANTITY', 'PRICE', 'NOTE TO SELLER',
                'SOURCE DOCUMENT', 'ORDER ', 'ATTACHMENTS', 'REMARKS'
            ]
            if any(line_clean.upper().startswith(pattern) for pattern in skip_patterns):
                break
            
            if DATE_PATTERN.match(line_clean):
                continue
            
            if re.match(r'^\d+$', line_clean):
                continue
            
            if re.match(r'^\d+-\d+$', line_clean):
                continue
            
            if line_clean == '0':
                continue
            
            address_lines.append(line_clean)
        
        address = ' '.join(address_lines)
        
        if address and len(address) > 10:
            return address
    
    return ''


def extract_line_data(doc: fitz.Document, line_no: int, line_pages: List[int],
                     order_no: str, ship_to: str, all_page_texts: List[str],
                     log) -> Optional[Dict]:
    """Extract complete data for a line number."""
    pages_to_search = set(line_pages)
    for page_idx in line_pages:
        if page_idx + 1 < len(all_page_texts):
            pages_to_search.add(page_idx + 1)
        if page_idx > 0:
            pages_to_search.add(page_idx - 1)
    
    # Use join instead of += for better performance
    combined_text = "\n\n".join(all_page_texts[page_idx] for page_idx in sorted(pages_to_search))
    
    # Initialize record with new column structure
    record = {
        'PO NO': order_no,
        'LINE': str(line_no),
        'PROMISE DATE': '',
        'DYPN': '',
        'Qty': '',
        'UNIT PRICE': '',
        'PART REV': '',
        'ORDER': '',
        'BATCH': '',
        'NEST': '',           # Blank placeholder
        'SOURCE': '',         # Blank placeholder
        'STANDARD CLAUSES': '',
        'SHIP': ship_to,
        'CUSTOMER': '',       # Blank placeholder
        'INTERCOMPANY': '',   # Blank placeholder
        'SUPPLIER': ''        # Blank placeholder
    }
    
    # Extract Rev, WJ code (Batch), ORDER (PPN), DYPN
    rev_pattern = rf'{line_no}\s*\n\s*([\w-]+)\s*-\s*Rev\s*\n\s*(\d+)\s*\n\s*(\S+)/(\S+?)/([\w\d\s\n-]+?)/(?:CUT|PLTCUT)'
    rev_match = re.search(rev_pattern, combined_text, re.DOTALL)
    
    if rev_match:
        part_type = rev_match.group(1)
        rev_num = rev_match.group(2)
        batch_code = rev_match.group(3)  # WJ182, etc. (was "Nest", now "BATCH")
        order_code = rev_match.group(4)  # Was "PPN", now "ORDER"
        dypn_raw = rev_match.group(5)
        
        record['PART REV'] = f"{part_type} - Rev {rev_num}"
        record['BATCH'] = batch_code
        record['ORDER'] = order_code
        dypn_cleaned = ''.join(dypn_raw.split())
        record['DYPN'] = dypn_cleaned
    
    # Extract Quantity and Unit Price - line-specific
    line_marker_pattern = rf'{line_no}\s*\n\s*OFLD'
    line_marker_match = re.search(line_marker_pattern, combined_text)
    
    if line_marker_match:
        text_after_line = combined_text[line_marker_match.start():]
        
        qty_pattern = r'QUANTITY\s+UOM.*?\n(\d+)\s*\n'
        qty_match = re.search(qty_pattern, text_after_line, re.DOTALL)
        if qty_match:
            record['Qty'] = qty_match.group(1)
        
        price_pattern = r'UNIT PRICE.*?\n\$([0-9,.]+)'
        price_match = re.search(price_pattern, text_after_line, re.DOTALL)
        if price_match:
            record['UNIT PRICE'] = price_match.group(1)
    else:
        qty_pattern = r'QUANTITY\s+UOM.*?\n(\d+)\s*\n'
        qty_match = re.search(qty_pattern, combined_text, re.DOTALL)
        if qty_match:
            record['Qty'] = qty_match.group(1)
        
        price_pattern = r'UNIT PRICE.*?\n\$([0-9,.]+)'
        price_match = re.search(price_pattern, combined_text, re.DOTALL)
        if price_match:
            record['UNIT PRICE'] = price_match.group(1)
    
    # Extract Promise Date
    promise_pattern = r'PROMISE\s*\n?DATE.*?\n(\d+-[A-Z]+-\d{4})'
    promise_match = re.search(promise_pattern, combined_text, re.DOTALL)
    if promise_match:
        record['PROMISE DATE'] = format_promise_date(promise_match.group(1))
    
    # Extract Clause Numbers
    clause_pattern = rf'{line_no}\s*\n\s*OFLD-\s*\n\s*PLTCUT\s*\n\s*SC-([\d-]+)'
    clause_matches = re.findall(clause_pattern, combined_text, re.MULTILINE)
    
    if clause_matches:
        seen = set()
        clauses = []
        for clause in clause_matches:
            if clause not in seen:
                clauses.append(clause)
                seen.add(clause)
        record['STANDARD CLAUSES'] = ', '.join(clauses)
    
    return record


def extract_from_pdf(pdf_path: Path, log) -> List[Dict]:
    """Extract all PO data from a PDF."""
    doc = None
    try:
        sdk.ensure_local(pdf_path, log)  # OneDrive placeholder -> download first (Hard Rule 13)
        doc = fitz.open(str(pdf_path))
    except FileNotFoundError:
        raise ValueError(f"PDF file not found: {pdf_path}")
    except Exception as e:
        raise RuntimeError(f"Could not open PDF: {e}")
    
    try:
        # Pre-extract all page texts
        all_page_texts = [page.get_text("text") for page in doc]
        
        # Build line index
        order_no, line_index = build_line_index(doc, log)
        
        if not order_no:
            log(f"   WARNING: No order number found in {pdf_path.name}")
            return []
        
        # Extract Ship To address once
        ship_to = extract_ship_to_address_once(all_page_texts)
        
        # Extract data for each line
        records = []
        for line_no in sorted(line_index.keys()):
            line_pages = line_index[line_no]
            record = extract_line_data(doc, line_no, line_pages, order_no, ship_to, all_page_texts, log)
            
            if record:
                records.append(record)
        
        return records
    
    finally:
        if doc:
            doc.close()


def write_to_excel(output_path: Path, records: List[Dict], log):
    """Write or append records to Excel with new column structure."""
    # Define headers in correct order (A through P)
    headers = [
        'PO NO', 'LINE', 'PROMISE DATE', 'DYPN', 'Qty', 'UNIT PRICE',
        'PART REV', 'ORDER', 'BATCH', 'NEST', 'SOURCE', 'STANDARD CLAUSES',
        'SHIP', 'CUSTOMER', 'INTERCOMPANY', 'SUPPLIER'
    ]
    
    if output_path.exists():
        wb = load_workbook(str(output_path))
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "PO Data"
        ws.append(headers)
    
    for record in records:
        row = [record.get(h, '') for h in headers]
        ws.append(row)
    
    wb.save(str(output_path))


def check_sequential_lines(records: List[Dict]) -> Dict[str, List[int]]:
    """
    Check for missing sequential LINE numbers per PO.
    
    Returns:
        Dict mapping PO numbers to lists of missing LINE numbers
    """
    from collections import defaultdict
    
    # Group LINE numbers by PO
    po_lines = defaultdict(list)
    
    for record in records:
        po_no = record.get('PO NO', '').strip()
        line_no = record.get('LINE', '').strip()
        
        if po_no and line_no:
            try:
                po_lines[po_no].append(int(line_no))
            except (ValueError, TypeError):
                pass
    
    # Find missing sequential numbers for each PO
    missing_per_po = {}
    
    for po_no, lines in po_lines.items():
        if not lines:
            continue
        
        lines_sorted = sorted(lines)
        min_line = min(lines_sorted)
        max_line = max(lines_sorted)
        
        # Expected sequence
        expected_lines = set(range(min_line, max_line + 1))
        actual_lines = set(lines_sorted)
        
        # Find gaps
        missing_lines = sorted(expected_lines - actual_lines)
        
        if missing_lines:
            missing_per_po[po_no] = missing_lines
    
    return missing_per_po


def validate_extracted_data(records: List[Dict]) -> Tuple[List[Dict], Dict[str, int]]:
    """
    Validate extracted records for missing data.
    
    Returns:
        - List of dicts for missing items Excel (one per PO with issues)
        - Dict of summary counts by column header
    """
    missing_items = []
    summary_counts = {col: 0 for col in REQUIRED_COLUMNS}
    
    for row_idx, record in enumerate(records, start=2):  # Start at row 2 (after header)
        missing_in_row = []
        
        for col_name in REQUIRED_COLUMNS:
            value = record.get(col_name, '').strip()
            if not value:
                col_letter = HEADER_TO_COLUMN[col_name]
                cell_ref = f"{col_letter}{row_idx}"
                missing_in_row.append(f"{col_name} ({cell_ref})")
                summary_counts[col_name] += 1
        
        if missing_in_row:
            # Build PO identifier
            po_value = record.get('PO NO', '').strip()
            if po_value:
                po_identifier = f"PO {po_value} (A{row_idx})"
            else:
                po_identifier = f"PO NO (A{row_idx})"
            
            missing_items.append({
                'po_identifier': po_identifier,
                'missing_fields': missing_in_row
            })
    
    return missing_items, summary_counts


def write_missing_items_excel(output_path: Path, missing_items: List[Dict], 
                              missing_lines_per_po: Dict[str, List[int]], log):
    """Write missing items validation to Excel, including missing LINE numbers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Items"
    
    # Header for missing cell data
    ws['A1'] = "PO'S MISSING ITEMS"
    
    current_row = 2
    
    # Section 1: Missing cell data
    for item in missing_items:
        # Column A: PO identifier
        ws.cell(row=current_row, column=1, value=item['po_identifier'])
        
        # Columns B onwards: Missing fields
        for col_idx, missing_field in enumerate(item['missing_fields'], start=2):
            ws.cell(row=current_row, column=col_idx, value=missing_field)
        
        current_row += 1
    
    # Section 2: Missing LINE numbers (if any)
    if missing_lines_per_po:
        # Add blank row separator
        current_row += 1
        
        # Header for missing LINE numbers section
        ws.cell(row=current_row, column=1, value="MISSING LINE NUMBERS")
        current_row += 1
        
        # List each missing LINE number
        for po_no in sorted(missing_lines_per_po.keys()):
            missing_lines = sorted(missing_lines_per_po[po_no])
            
            for line_no in missing_lines:
                ws.cell(row=current_row, column=1, value=f"PO {po_no}")
                ws.cell(row=current_row, column=2, value=f"LINE {line_no}")
                current_row += 1
    
    wb.save(str(output_path))
    log(f"   Missing items report: {output_path.name}")


def write_analysis_txt(output_path: Path, summary_counts: Dict[str, int], 
                       missing_lines_per_po: Dict[str, List[int]], log):
    """Write summary analysis to text file including missing LINE numbers."""
    with open(output_path, 'w', encoding='utf-8') as f:
        # Section 1: Missing cell data counts
        f.write("=" * 60 + "\n")
        f.write("MISSING CELL DATA\n")
        f.write("=" * 60 + "\n")
        for col_name in REQUIRED_COLUMNS:
            count = summary_counts.get(col_name, 0)
            f.write(f"Missing {col_name}: {count}\n")
        
        # Section 2: Missing sequential LINE numbers
        f.write("\n" + "=" * 60 + "\n")
        f.write("MISSING SEQUENTIAL LINE NUMBERS\n")
        f.write("(Entire rows not extracted from PDFs)\n")
        f.write("=" * 60 + "\n")
        
        if missing_lines_per_po:
            for po_no in sorted(missing_lines_per_po.keys()):
                missing_lines = missing_lines_per_po[po_no]
                f.write(f"\nPO {po_no}:\n")
                f.write(f"  Missing {len(missing_lines)} line(s): {', '.join(map(str, missing_lines))}\n")
        else:
            f.write("\nNo missing sequential LINE numbers - all LINE sequences are complete.\n")
    
    log(f"   Analysis report: {output_path.name}")


# ===== MAIN PLUGIN FUNCTION =====
def run(params: Dict[str, Any], progress_callback, cancel_event: threading.Event) -> None:
    """Main plugin execution."""
    log = params.get('log', print)
    
    log("Starting 911 PO PDF Extractor v2.3.0...")
    progress_callback(0)
    
    # === PROMPT 1: FOLDER PATH ===
    log("Input required from user...")
    folder_input = get_console_input(params, "Enter folder path containing PO packet PDFs")
    
    folder = Path(folder_input.strip()).expanduser().resolve()
    if not folder.exists():
        log(f"ERROR: Folder does not exist: {folder}")
        raise ValueError(f"Folder does not exist: {folder}")
    if not folder.is_dir():
        log(f"ERROR: Path is not a directory: {folder}")
        raise ValueError(f"Path is not a directory: {folder}")
    
    log(f"Folder: {folder}")
    progress_callback(5)
    
    # === PROMPT 2: OUTPUT FILENAME ===
    output_input = get_console_input(params, "Enter output Excel filename (e.g., 'po_data')")
    
    if not output_input.strip():
        log("ERROR: Output filename cannot be empty!")
        raise ValueError("Output filename is required")
    
    # Auto-append .xlsx if not present
    output_filename = output_input.strip()
    if '/' in output_filename or '\\' in output_filename:
        log("ERROR: Filename cannot contain path separators!")
        raise ValueError("Filename cannot contain path separators")
    
    if not output_filename.lower().endswith('.xlsx'):
        output_filename += '.xlsx'
    
    # Extract base name for folder (without .xlsx extension)
    output_base_name = output_filename[:-5] if output_filename.lower().endswith('.xlsx') else output_filename
    
    # Create output folder
    output_folder = folder / output_base_name
    output_folder.mkdir(exist_ok=True)
    
    # Define output paths
    xlsx_path = output_folder / output_filename
    missing_items_path = output_folder / f"{output_base_name}_missing_items.xlsx"
    analysis_path = output_folder / f"{output_base_name}_analysis.txt"
    
    log(f"Output folder: {output_folder.name}/")
    log(f"Main output: {output_filename}")
    log("")
    progress_callback(10)
    
    if cancel_event.is_set():
        log("Operation cancelled")
        return
    
    # === FIND PDFs ===
    log("Finding PDFs...")
    pdfs = sorted([p for p in folder.iterdir() if p.suffix.lower() == ".pdf"])
    
    if not pdfs:
        log("ERROR: No PDFs found in folder")
        raise ValueError("No PDFs found")
    
    log(f"Found {len(pdfs)} PDF(s)")
    progress_callback(15)
    
    # === EXTRACT DATA ===
    log("Extracting PO packet data...")
    log("   This may take a few minutes...")
    log("")
    
    all_records = []
    total = len(pdfs)
    
    for idx, pdf in enumerate(pdfs):
        if cancel_event.is_set():
            log("WARNING: Cancelled by user")
            return
        
        try:
            log(f"   Processing: {pdf.name}")
            records = extract_from_pdf(pdf, log)
            all_records.extend(records)
            log(f"   Extracted {len(records)} record(s)")
        except Exception as e:
            log(f"   WARNING: Error: {e}")
        
        progress = 15 + int((idx + 1) / total * 60)
        progress_callback(progress)
    
    log("")
    log(f"Total records extracted: {len(all_records)}")
    progress_callback(75)
    
    if not all_records:
        log("WARNING: No records extracted")
        progress_callback(100)
        return
    
    # === WRITE TO EXCEL ===
    log("Writing to Excel...")
    try:
        write_to_excel(xlsx_path, all_records, log)
        log(f"Wrote {len(all_records)} records to {xlsx_path.name}")
    except Exception as e:
        log(f"ERROR: Failed to write Excel: {e}")
        raise
    
    progress_callback(85)
    
    # === VALIDATE DATA ===
    log("")
    log("Validating extracted data...")
    
    # Check for missing cell data
    missing_items, summary_counts = validate_extracted_data(all_records)
    
    # Check for missing sequential LINE numbers
    missing_lines_per_po = check_sequential_lines(all_records)
    
    if missing_items or missing_lines_per_po:
        if missing_items:
            log(f"WARNING: Found {len(missing_items)} PO(s) with missing cell data")
        if missing_lines_per_po:
            total_missing_lines = sum(len(lines) for lines in missing_lines_per_po.values())
            log(f"WARNING: Found {total_missing_lines} missing LINE number(s) across {len(missing_lines_per_po)} PO(s)")
        
        log("")
        log("Generating validation reports...")
        
        # Write missing items Excel (includes both missing cells and missing LINE numbers)
        if missing_items or missing_lines_per_po:
            write_missing_items_excel(missing_items_path, missing_items, missing_lines_per_po, log)
        
        # Write analysis text (includes both cell data and line sequences)
        write_analysis_txt(analysis_path, summary_counts, missing_lines_per_po, log)
    else:
        log("All records complete - no missing data or LINE gaps")
    
    progress_callback(95)
    
    # === SUMMARY ===
    log("")
    log("=" * 50)
    log("EXTRACTION SUMMARY")
    log("=" * 50)
    log(f"Processed: {len(pdfs)} PDFs")
    log(f"Extracted: {len(all_records)} records")
    log(f"Output folder: {output_folder.name}/")
    log(f"   • Main data: {output_filename}")
    
    # List validation files that were created
    validation_files_created = []
    if missing_items or missing_lines_per_po:
        validation_files_created.append(f"Missing items: {missing_items_path.name}")
    if missing_items or missing_lines_per_po:
        validation_files_created.append(f"Analysis: {analysis_path.name}")
    
    if validation_files_created:
        for file_desc in validation_files_created:
            log(f"   • {file_desc}")
    
    log("=" * 50)
    log("911 PO PDF Extractor completed successfully!")
    
    progress_callback(100)


if __name__ == "__main__":
    import threading
    cancel_event = threading.Event()
    
    def progress(p):
        print(f"Progress: {p}%")
    
    try:
        run(params={'console': None}, progress_callback=progress, cancel_event=cancel_event)
        print("\nDone!")
    except Exception as e:
        print(f"\nERROR: Failed: {e}")
