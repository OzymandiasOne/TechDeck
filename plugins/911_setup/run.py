"""
911 Setup Plugin
================
Automates the full 911 QTDR batch setup workflow:

  1. Prompt for batch number -> locate batch folder
  2. Read BATCH LIST -> extract unique nest numbers from "Nest Pkg Nbr" header column
  2.5 Show the nest-selection dialog -> user picks which nests to run (nests
     with an existing folder are flagged "already set up"). Nothing below runs
     until the user submits.
  3. Create a subfolder per selected nest inside the batch folder
  4. Copy "911 BATCH _.xlsx" template into each nest folder, rename it, and
     drop the scribe-verification doc (QF-QU-15 ... SHAPES.docx) into a
     PRODUCTION PAPERWORK subfolder of each
  5. Copy Working Forecast List -> extract rows for each nest -> paste into
     NEST sheet cols A-C starting row 4
  6. Parse the nest's own NEST PACKAGES packet PDF -> read the labeled
     MIL SPEC field (-> D4) and MATERIAL field (-> E4) off the MOVE TICKET page
  7. Read BATCH LIST -> filter rows by nest number -> paste into NEST cols F-K
     starting row 4
  8. Read part rows (Work Order / DYPN / DYPN QTY) from NEST -> copy
     INSPECTION SHEET tab for each row; fill one Part Number slot per piece
     (qty 3 -> slots A16/A18/A20) with the hull code (= work order's first
     2 chars) under the part number in every slot; slots 2+ mirror slot 1's
     description; template tab hidden afterwards (v1.6.0). Sheet name =
     suffix (e.g. "-80"); on collisions, both colliding sheets are renamed
     using the last 2 chars of the preceding segment (e.g. H4533321-80 ->
     "21-80", H4533322-80 -> "22-80").
  9. Save every nest excel, repeat for all nests in the batch
 10. Build MOVE TICKET OMIT PDF for each nest (removes MOVE TICKET pages, keeps
     MIL-SPEC/HULL); first page stamped BATCH/NEST + Material Type filled
     (v1.5.0, via the sibling 911_remove_ticket helpers)

v1.2.0 changes
  - Filesystem roots are configurable via plugin settings:
      * qtdr_base_path     -> 911 QTDR root
      * forecast_dir       -> Forecast and Inventory Reports root
      * template_subdir    -> subpath under QTDR for the 911 BATCH template
      * forecast_filename  -> name of the working forecast workbook
    Blank/missing values fall back to the original Path.home() defaults,
    so existing installs keep working without any setup.

  - DYPN QTY verification (v1.4.0): some upstream BATCH LISTs arrive with
    the 'DYPN QTY' and 'Material Amount (Total)' headers swapped. Before
    pasting batch rows, both columns are scored against the per-part
    quantities in each nest packet PDF's SUMMARY OF NEST page (keyed by
    DYPN + Work Order); the column the packets agree with feeds the nest
    workbooks regardless of its label, the swapped headers are repaired in
    the BATCH LIST file (skipped with a warning if it's open in Excel),
    and any row whose qty still disagrees with the packet keeps the BATCH
    LIST value but is flagged three ways: its DYPN QTY cell (NEST col I)
    is filled yellow in the workbook, a QTY MISMATCH console warning is
    logged, and an end-of-run QTY VERIFICATION summary repeats every
    mismatch. Parts missing from the packet summary are reported as
    UNVERIFIED in the same summary.

  - Nest number regex accepts legacy numeric and alphanumeric formats
    (case-insensitive):
      * P07866, S013      (P/S-prefixed format)
      * 503682            (digits-only format)
      * 5CDAVW, 9FANDR    (alphanumeric IDs, 4-8 chars with at least one
                           digit -- first seen in batch GX030)
    Rejects stray small numbers (< 3 digits) that could appear in totals
    or footer rows.

  - QTDR root existence is validated before trying to find the batch
    folder, with an error message that points the user at the relevant
    setting.

  - FIXED: forecast row matching used a raw `==` between the batch-list
    nest (str) and the forecast cell value. Digits-only nests like
    "503627" failed silently because the forecast stores them as int.
    Match now normalizes both sides to stripped, upper-cased strings.

v1.3.0 changes
  - Nest selection: after the batch number is entered, a dialog lists every
    nest in the batch with checkboxes (batch root toggles all; existing nest
    folders are flagged "already set up"). The user picks which nests to run
    and submits; the setup work runs only on the chosen nests. Cancelling the
    dialog runs nothing. With no console (CLI/test) the optional 'nests'
    param selects a subset, otherwise all nests run.

  - FIXED: MIL spec and MATERIAL (cols D4/E4) are now read from the nest's
    OWN packet PDF (the NEST PACKAGES file whose name contains the nest
    number) instead of scanning every PDF in NEST PACKAGES and taking the
    first match -- which let one nest inherit another nest's spec.

  - FIXED: both values are read from the labeled MOVE TICKET fields
    "MIL SPEC:" and "MATERIAL:". The old code read MIL via a bare
    `MIL-S-\\d+` token (missing QQ-/ASTM-/AISI- specs entirely) and read
    material from the PART SKETCH "MATL:" field, which is blank on many
    parts -- so the old `MATL:\\s*(\\S+)` regex skipped the blank and
    captured the next field's label (e.g. "LVL:"). A blank MATERIAL now
    stays blank.

  - Scribe-verification doc: "QF-QU-15 REV B - SCRIBE VERIFICATION -
    SHAPES.docx" is copied from the SACO template dir into a
    "PRODUCTION PAPERWORK" subfolder of every generated nest folder (one copy
    per nest -- skipped if already present; a loose copy in the nest root
    from an earlier version is moved into the subfolder). Missing source is
    logged as a warning and does not abort the run.

v1.3.1 changes
  - Nests are sorted in Windows Explorer ascending order (StrCmpLogicalW
    natural sort) so the selection dialog and processing match the file
    system's ordering instead of BATCH LIST order.

  - NEST cols A-E are filled down for every part row, not just row 4. The
    template ships D/E as formulas keyed off column C (the forecast BATCH),
    which is usually blank, so MIL spec / material never appeared below
    row 4. Now the row-4 values are written down every part row (formula
    cells in D/E overwritten with the literal value; blank A-C cells filled,
    real forecast data preserved). The SCRIBE VERIFICATION sheet mirrors
    NEST via formulas, so it fills automatically once Excel recalculates.

v1.8.0 changes
  - New FIRST stage, "Generate Teams Cards": one card per nest the EB 922
    Schedule marks NEED TEAMS/SETUP, posted to the MODELING bucket of the
    SOPO D911 PIPELINE plan (D922 channel) via a Power Automate webhook --
    the same indirection 922 Setup uses, so TechDeck never touches Planner.

    The stage is BATCH-INDEPENDENT: its work list is the schedule's
    CURRENT PIPELINE sheet (DEPT. 911 + STATUS NEED TEAMS/SETUP), which
    spans whatever batches are queued. It therefore runs BEFORE the batch
    prompt, and checking it alone is a complete run that never asks for a
    batch number.

    Card = "BATCH: {batch} - NEST: {nest} ({source material})", due on the
    schedule's DATE, labelled with its difficulty (SIMPLE / MEDIUM /
    DIFFICULT, read from the RATING cell's COLOUR by the same sibling
    911 Remove Ticket helpers that stamp the packet) and its machine
    (SAW CUT / TUBE LASER, decided from the NOTES text). The "(...)" is
    the nest's EB source-material stock code, which lives in the batch's
    own BATCH LIST 'Material' column -- not on the schedule -- so each
    referenced batch's BATCH LIST is read once and cached.

    The "Program" checklist item is dropped on non-tube stock. Card layout,
    checklist, label slots and the machine rules all live in the sibling
    card_template.json; the flow recipe is docs/TEAMS_CARDS.md (flow #3).
"""

import ctypes
import datetime as _dt
import json
import re
import shutil
import threading
from functools import cmp_to_key
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

import importlib.util


# ---------------------------------------------------------------------------
# Nest number pattern
#
# Two real-world formats:
#   - legacy numeric: optional P or S prefix, then 3+ digits (P0123, 503627)
#   - alphanumeric IDs: 4-8 letters/digits with at least one digit
#     (5CDAVW, 9FANDR -- first seen in batch GX030, May 2026)
# The at-least-one-digit requirement keeps footer/total text like "TOTALS"
# out, which is the whole reason this filter exists.
# ---------------------------------------------------------------------------
_NEST_RE = re.compile(
    r'^(?:[PS]?\d{3,}|(?=[A-Z0-9]*\d)[A-Z0-9]{4,8})$',
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Windows-style natural sort
#
# Nests are displayed (and processed) in the same ascending order Windows
# Explorer shows files: digit runs compared numerically, case-insensitive.
# StrCmpLogicalW is the exact shell API Explorer uses; off Windows we fall
# back to a manual natural-sort key.
# ---------------------------------------------------------------------------
try:
    _StrCmpLogicalW = ctypes.windll.shlwapi.StrCmpLogicalW
    _StrCmpLogicalW.argtypes = (ctypes.c_wchar_p, ctypes.c_wchar_p)
    _StrCmpLogicalW.restype = ctypes.c_int
except (AttributeError, OSError):
    _StrCmpLogicalW = None


def _windows_natural_sorted(values: list) -> list:
    """Sort like Windows Explorer (logical/natural ascending order)."""
    if _StrCmpLogicalW is not None:
        return sorted(values, key=cmp_to_key(
            lambda a, b: _StrCmpLogicalW(str(a), str(b))))
    return sorted(
        values,
        key=lambda s: [int(t) if t.isdigit() else t.lower()
                       for t in re.split(r'(\d+)', str(s))],
    )


# ---------------------------------------------------------------------------
# Path helpers
#
# Each helper takes an optional override string. If the override is non-empty
# it wins; otherwise the helper falls back to the default Path.home() path
# that the plugin has always used. This means the new settings are purely
# additive - blank settings preserve the previous behavior exactly.
# ---------------------------------------------------------------------------

_DEFAULT_TEMPLATE_SUBDIR = "03 - Processing Forms & Templates\\00 - SACO"
_DEFAULT_FORECAST_FILENAME = "Working Forecast List.xlsx"

# Scribe-verification form dropped into a PRODUCTION PAPERWORK subfolder of
# every generated nest folder (subfolder per user feedback 2026-07-21). Lives
# in the same SACO template directory as the 911 BATCH template (template_dir).
_SCRIBE_DOC_FILENAME = "QF-QU-15 REV B - SCRIBE VERIFICATION - SHAPES.docx"
_SCRIBE_SUBFOLDER = "PRODUCTION PAPERWORK"


def _base_qtdr(override: str = "") -> Path:
    """911 QTDR root. Override wins; otherwise auto-discover across every
    OneDrive path variant. Falls back to the canonical default path (which
    may not exist) so the caller's existence check reports something useful."""
    root = sdk.resolve_911_qtdr_root(override)
    if root is not None:
        return root
    return sdk.pilot_program_roots()[0] / "911 QTDR"


def _forecast_dir(override: str = "") -> Path:
    """Forecast and Inventory Reports root. Override wins; otherwise auto-discover."""
    root = sdk.resolve_forecast_dir(override)
    if root is not None:
        return root
    return sdk.pilot_program_roots()[0] / "Forecast and Inventory Reports"


def _template_dir(qtdr_override: str = "",
                  template_subdir: str = _DEFAULT_TEMPLATE_SUBDIR) -> Path:
    """Template directory, relative to the 911 QTDR root."""
    return _base_qtdr(qtdr_override) / template_subdir


# ---------------------------------------------------------------------------
# BATCH LIST helpers
# ---------------------------------------------------------------------------

def _find_batch_list(batch_folder: Path, batch_number: str) -> Path:
    """
    Locate the BATCH LIST excel inside the batch folder.
    Expected name: "<batch_number> BATCH LIST.xlsx" (case-insensitive).
    Falls back to any .xlsx containing 'BATCH LIST' if exact name not found.
    """
    exact = batch_folder / f"{batch_number} BATCH LIST.xlsx"
    if exact.exists():
        return exact

    for f in batch_folder.iterdir():
        if (f.is_file()
                and f.suffix.lower() == ".xlsx"
                and "BATCH LIST" in f.name.upper()
                and not f.name.startswith("~")):
            return f

    raise FileNotFoundError(
        f"No BATCH LIST file found in {batch_folder}. "
        f"Expected '{batch_number} BATCH LIST.xlsx'."
    )


def _find_header_col(ws, header_name: str, header_row: int):
    """
    Scan a specific row for a column whose value matches header_name
    (case-insensitive, stripped). Returns the 1-based column index or None.
    """
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is not None and str(val).strip().upper() == header_name.upper():
            return col
    return None


def _get_unique_nests_from_batch_list(batch_list_path: Path) -> list:
    """
    Read the BATCH LIST 'BATCH' sheet.
    Headers in row 3, data from row 4.
    Locates 'Nest Pkg Nbr' column by header name (case-insensitive).
    Returns the unique nest numbers matching _NEST_RE, sorted in
    Windows Explorer ascending order so the selection dialog and processing
    follow the same ordering the operator sees in the file system.
    """
    wb = sdk.load_workbook_resilient(batch_list_path, data_only=True)
    ws = wb["BATCH"] if "BATCH" in wb.sheetnames else wb.active

    nest_col = _find_header_col(ws, "Nest Pkg Nbr", header_row=3)
    if nest_col is None:
        wb.close()
        raise ValueError(
            f"Could not find 'Nest Pkg Nbr' column header in row 3 of {batch_list_path.name}."
        )

    seen = []
    seen_set = set()
    for row in range(4, ws.max_row + 1):
        val = ws.cell(row, nest_col).value
        if val is not None:
            val = str(val).strip()
            if val and _NEST_RE.match(val) and val not in seen_set:
                seen_set.add(val)
                seen.append(val)

    wb.close()
    return _windows_natural_sorted(seen)


# ---------------------------------------------------------------------------
# DYPN QTY verification against the nest packet PDFs
#
# Upstream BATCH LISTs sometimes ship with the 'DYPN QTY' and
# 'Material Amount (Total)' headers swapped, so the column labeled
# 'DYPN QTY' carries material amounts and vice versa (first confirmed in
# GX030, Jun 2026). The packet PDF's SUMMARY OF NEST page is the ground
# truth for per-part quantities, so before pasting batch rows we check
# both candidate columns against the packets, use whichever column the
# packets agree with, and repair the swapped headers in the BATCH LIST.
# ---------------------------------------------------------------------------

_SUMMARY_REF_RE = re.compile(r'^\d{1,4}$')
_SUMMARY_WO_RE = re.compile(r'^[A-Z]{1,2}\d{5,8}$', re.IGNORECASE)
_SUMMARY_PART_RE = re.compile(r'^[A-Z0-9][A-Z0-9.\-]{3,}$', re.IGNORECASE)


def _as_int(value):
    """Coerce a cell/PDF value to int, or None if it isn't numeric."""
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _parse_packet_summary_qtys(pdf_path: Path) -> dict:
    """
    Parse the SUMMARY OF NEST table(s) in a nest packet PDF.

    Returns {(PART NUMBER, WORK ORDER): qty} (keys uppercased, qty summed
    across duplicate rows). The table renders as a flat line sequence --
    REF / PART NUMBER / QTY / WORK ORDER / SK headers, then per part:
      <ref int> <part number> <qty int> <work order> [sk]
    -- so rows are recovered by sliding a 4-line window over the page text
    and keeping windows that match that shape.
    """
    qtys = {}
    if not PYMUPDF_AVAILABLE:
        return qtys
    sdk.ensure_local(pdf_path)  # OneDrive placeholder -> download first (Hard Rule 13)
    doc = fitz.open(pdf_path)
    try:
        for page in doc:
            text = page.get_text()
            if "SUMMARY OF NEST" not in text.upper():
                continue
            lines = [ln.strip() for ln in text.splitlines()]
            i = 0
            while i + 3 < len(lines):
                ref, part, qty, wo = lines[i:i + 4]
                if (_SUMMARY_REF_RE.match(ref)
                        and _SUMMARY_PART_RE.match(part)
                        and not _SUMMARY_REF_RE.match(part)
                        and _SUMMARY_REF_RE.match(qty)
                        and _SUMMARY_WO_RE.match(wo)):
                    key = (part.upper(), wo.upper())
                    qtys[key] = qtys.get(key, 0) + int(qty)
                    i += 4
                else:
                    i += 1
    finally:
        doc.close()
    return qtys


def _find_header_col_prefix(ws, prefix: str, header_row: int):
    """Like _find_header_col but matches on a case-insensitive prefix, so
    'Material Amount (Total)' and any future suffix variants both hit."""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is not None and str(val).strip().upper().startswith(prefix.upper()):
            return col
    return None


def _resolve_dypn_qty_col(batch_list_path: Path, nest_packages_folder: Path,
                          nests: list, log, cancel_event=None) -> tuple:
    """
    Decide which BATCH LIST column truly holds the DYPN quantities by
    scoring the column labeled 'DYPN QTY' and the one labeled
    'Material Amount (Total)' against every available packet PDF's
    SUMMARY OF NEST quantities (keyed by DYPN + Work Order).

    Returns (qty_col, swap_with, packet_qtys):
      qty_col     -- 1-based column to read DYPN QTY from
      swap_with   -- column whose header must be swapped with qty_col in
                     the BATCH LIST file (None when the label is correct)
      packet_qtys -- {nest: {(part, wo): qty}} for later per-row warnings
    Falls back to the labeled column (with a warning) when there is
    nothing to verify against.
    """
    HEADER_ROW = 3
    wb = sdk.load_workbook_resilient(batch_list_path, data_only=True)
    ws = wb["BATCH"] if "BATCH" in wb.sheetnames else wb.active

    col_labeled = _find_header_col(ws, "DYPN QTY", HEADER_ROW)
    col_mat = _find_header_col_prefix(ws, "Material Amount", HEADER_ROW)
    col_dypn = _find_header_col(ws, "DYPN", HEADER_ROW)
    col_wo = _find_header_col(ws, "Work Order", HEADER_ROW)
    col_nest = _find_header_col(ws, "Nest Pkg Nbr", HEADER_ROW)

    packet_qtys = {}
    for nest in nests:
        sdk.raise_if_cancelled(cancel_event)
        pdf = _find_nest_pdf(nest_packages_folder, nest)
        if pdf is None:
            log(f"  WARNING: No packet PDF for nest {nest} -- skipping it in QTY verification.")
            continue
        try:
            qmap = _parse_packet_summary_qtys(pdf)
        except Exception as e:
            log(f"  WARNING: Could not parse {pdf.name} for quantities: {e}")
            continue
        if qmap:
            packet_qtys[nest] = qmap
        else:
            log(f"  WARNING: No SUMMARY OF NEST quantities found in {pdf.name}.")

    if col_labeled is None or None in (col_dypn, col_wo, col_nest) or not packet_qtys:
        wb.close()
        log("  WARNING: Could not verify the DYPN QTY column against the nest "
            "packets -- using the column labeled 'DYPN QTY' as-is.")
        return col_labeled, None, packet_qtys

    match_labeled = match_mat = compared = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        nest_val = ws.cell(row, col_nest).value
        qmap = packet_qtys.get(str(nest_val).strip()) if nest_val is not None else None
        if not qmap:
            continue
        key = (str(ws.cell(row, col_dypn).value).strip().upper(),
               str(ws.cell(row, col_wo).value).strip().upper())
        pdf_qty = qmap.get(key)
        if pdf_qty is None:
            continue
        compared += 1
        if _as_int(ws.cell(row, col_labeled).value) == pdf_qty:
            match_labeled += 1
        if col_mat is not None and _as_int(ws.cell(row, col_mat).value) == pdf_qty:
            match_mat += 1

    mat_header = str(ws.cell(HEADER_ROW, col_mat).value).strip() if col_mat else ""
    wb.close()

    if compared == 0:
        log("  WARNING: No BATCH LIST rows could be matched to the nest packets "
            "(DYPN/Work Order mismatch?) -- using the column labeled 'DYPN QTY' as-is.")
        return col_labeled, None, packet_qtys

    log(f"  'DYPN QTY' column matches the packets on {match_labeled}/{compared} rows; "
        f"'{mat_header or 'Material Amount'}' on {match_mat}/{compared}.")

    if match_mat > match_labeled:
        log(f"  Headers are SWAPPED: true DYPN QTY is column "
            f"{get_column_letter(col_mat)} ('{mat_header}').")
        return col_mat, col_labeled, packet_qtys

    if match_labeled == 0:
        log("  WARNING: NEITHER column matches the packet quantities -- "
            "using the column labeled 'DYPN QTY'; verify the paperwork by hand.")
    else:
        log("  Verified: the column labeled 'DYPN QTY' holds the true quantities.")
    return col_labeled, None, packet_qtys


def _swap_batch_list_headers(batch_list_path: Path, col_a: int, col_b: int, log):
    """
    Swap the two row-3 header cells in the BATCH LIST file so 'DYPN QTY'
    sits over the column that actually matches the nest packets. A locked
    file (open in Excel) is non-fatal: setup still reads the verified
    column either way, only the on-disk label stays wrong.
    """
    try:
        wb = sdk.load_workbook_resilient(batch_list_path)
        ws = wb["BATCH"] if "BATCH" in wb.sheetnames else wb.active
        a, b = ws.cell(3, col_a).value, ws.cell(3, col_b).value
        ws.cell(3, col_a).value = b
        ws.cell(3, col_b).value = a
        wb.save(batch_list_path)
        wb.close()
        log(f"  Swapped BATCH LIST headers: '{b}' <-> '{a}' "
            f"(cols {get_column_letter(col_a)}/{get_column_letter(col_b)}).")
        return True
    except PermissionError:
        log("  WARNING: BATCH LIST is open or locked -- headers NOT swapped. "
            "Close it and re-run, or swap the two headers by hand.")
        return False
    except Exception as e:
        log(f"  WARNING: Could not swap BATCH LIST headers: {e}")
        return False


def _get_batch_rows_for_nest(batch_list_path: Path, nest_number: str,
                             qty_col_override: int = None) -> list:
    """
    Read the BATCH LIST 'BATCH' sheet and return rows matching nest_number
    in the 'Nest Pkg Nbr' column.

    Headers in row 3, data from row 4.
    Pulls these columns by header name (case-insensitive):
      Work Order, DYPN, Material, DYPN QTY, Nest Pkg Nbr, SCOPE OF WORK

    qty_col_override, when given, is the packet-verified DYPN QTY column
    from _resolve_dypn_qty_col and is used instead of the 'DYPN QTY'
    header lookup (the header can sit on the wrong column upstream).

    Returns list of 6-tuples in that order.
    """
    wb = sdk.load_workbook_resilient(batch_list_path, data_only=True)
    ws = wb["BATCH"] if "BATCH" in wb.sheetnames else wb.active

    HEADER_ROW = 3

    col_work_order = _find_header_col(ws, "Work Order",    HEADER_ROW)
    col_dypn       = _find_header_col(ws, "DYPN",          HEADER_ROW)
    col_material   = _find_header_col(ws, "Material",      HEADER_ROW)
    col_dypn_qty   = qty_col_override if qty_col_override is not None \
        else _find_header_col(ws, "DYPN QTY", HEADER_ROW)
    col_nest       = _find_header_col(ws, "Nest Pkg Nbr",  HEADER_ROW)
    col_scope      = _find_header_col(ws, "SCOPE OF WORK", HEADER_ROW)

    missing = [name for name, col in [
        ("Work Order",    col_work_order),
        ("DYPN",          col_dypn),
        ("Material",      col_material),
        ("DYPN QTY",      col_dypn_qty),
        ("Nest Pkg Nbr",  col_nest),
        ("SCOPE OF WORK", col_scope),
    ] if col is None]

    if missing:
        wb.close()
        raise ValueError(
            f"Could not find these column headers in row 3 of {batch_list_path.name}: "
            + ", ".join(missing)
        )

    rows = []
    for row in range(4, ws.max_row + 1):
        nest_val = ws.cell(row, col_nest).value
        if nest_val is not None and str(nest_val).strip().upper() == nest_number.upper():
            rows.append((
                ws.cell(row, col_work_order).value,
                ws.cell(row, col_dypn).value,
                ws.cell(row, col_material).value,
                ws.cell(row, col_dypn_qty).value,
                ws.cell(row, col_nest).value,
                ws.cell(row, col_scope).value,
            ))

    wb.close()
    return rows


def _paste_batch_rows_into_nest(nest_ws, batch_rows: list):
    """
    Paste batch_rows into NEST sheet cols F-K (6-11) starting at row 4.
    Each item is a 6-tuple:
      (Work Order, DYPN, Material, DYPN QTY, Nest Pkg Nbr, SCOPE OF WORK)
    """
    for i, row_data in enumerate(batch_rows):
        dest_row = 4 + i
        for j, val in enumerate(row_data):
            nest_ws.cell(dest_row, 6 + j).value = val  # F=6 through K=11


# Yellow fill marking a DYPN QTY cell that disagrees with the nest packet.
_QTY_MISMATCH_FILL = PatternFill("solid", fgColor="FFFF00")


def _flag_qty_mismatches(nest_ws, batch_rows: list, pmap: dict, nest: str,
                         workbook_name: str, log, mismatches: list,
                         unverified: list):
    """
    Compare each pasted batch row's qty against the nest packet's
    SUMMARY OF NEST quantities (pmap, keyed (DYPN, WORK ORDER) upper).

    A row whose qty disagrees with the packet keeps the BATCH LIST value
    but gets its DYPN QTY cell (NEST col I) filled yellow plus a console
    warning. A row absent from a non-empty packet summary is reported as
    unverifiable. Both are appended to the run-level mismatches /
    unverified lists that feed the end-of-run summary block.
    """
    for i, (wo, dypn, _mat, qty, _nestval, _scope) in enumerate(batch_rows):
        pdf_qty = pmap.get((str(dypn).strip().upper(),
                            str(wo).strip().upper()))
        if pdf_qty is None:
            if pmap:
                unverified.append((nest, dypn, wo, qty))
                log(f"  WARNING: {dypn} / {wo} is not in the nest packet "
                    f"summary -- qty {qty} could NOT be verified.")
        elif _as_int(qty) != pdf_qty:
            mismatches.append((nest, dypn, wo, qty, pdf_qty))
            nest_ws.cell(4 + i, 9).fill = _QTY_MISMATCH_FILL  # col I = DYPN QTY
            log(f"  WARNING: QTY MISMATCH for {dypn} / {wo}: BATCH LIST says "
                f"{qty}, nest packet says {pdf_qty} -- qty cell highlighted "
                f"yellow in {workbook_name}.")


def _fill_nest_part_rows(nest_ws, num_parts: int):
    """
    Replicate the row-4 header values (cols A-E: forecast PO/line/batch, MIL
    spec, material type) down to every part row so each part's row is fully
    populated.

    The template ships D5:E.. as formulas keyed off column C (=IF(C5<>"",
    $D$4,"")). Column C is the forecast BATCH, which is blank whenever the
    forecast has no matching row, so those formulas resolve to "" and the
    MIL spec / material never appear below row 4. We overwrite the formula
    cells (and fill any blank A-C cell) with the literal row-4 value.

    A-C cells that already hold real forecast data are left untouched (only
    blanks are filled); D-E formula cells are always replaced with the literal
    nest-level value, matching the template's own intent that MIL spec and
    material repeat per row.

    The SCRIBE VERIFICATION sheet is entirely formula-driven off NEST
    (D3=IF(ISBLANK(NEST!D5),"",NEST!D5), etc.), so filling NEST fills SCRIBE
    too once Excel recalculates (the inspection-sheet COM save does this).
    """
    if num_parts <= 1:
        return
    header = [nest_ws.cell(4, c).value for c in range(1, 6)]  # A4..E4
    for i in range(1, num_parts):
        row = 4 + i
        for c in range(1, 6):
            cur = nest_ws.cell(row, c).value
            is_blank = (
                cur is None
                or (isinstance(cur, str) and (cur.strip() == "" or cur.startswith("=")))
            )
            if is_blank:
                nest_ws.cell(row, c).value = header[c - 1]


# ---------------------------------------------------------------------------
# Step 5: Working Forecast List -> NEST cols A-C
# ---------------------------------------------------------------------------

def _nest_matches(cell_val, target_upper: str) -> bool:
    """True if a forecast 'Nest' cell equals the batch-list nest. Compares as
    upper/stripped strings AND as integers, so a nest stored as a number
    (503712, or even 503712.0 on some machines) matches the string '503712'."""
    if cell_val is None:
        return False
    if str(cell_val).strip().upper() == target_upper:
        return True
    a, b = _as_int(cell_val), _as_int(target_upper)
    return a is not None and a == b


def _locate_forecast(forecast_wb, log):
    """Find the 911 forecast sheet, its header row, the 'Nest' match column, and
    the PO/Line/Batch output columns -- all by HEADER NAME (Hard Rules 1 & 2),
    never a fixed index, so an inserted/moved column can't silently break the
    lookup. Returns (ws, header_row, nest_col, (po, line, batch)). Raises a clear
    error if the sheet or the 'Nest' column is absent -- failing loudly beats
    silently leaving every nest's columns A-C blank."""
    ws = None
    for name in forecast_wb.sheetnames:
        if name.strip().lower() == "911 forecast":
            ws = forecast_wb[name]
            break
    if ws is None:                       # tolerate a renamed-but-recognizable sheet
        # Fuzzy fallback scans VISIBLE sheets only - a hidden old copy
        # ("Copy of 911 Forecast") must never shadow the live sheet.
        for name in sdk.visible_sheetnames(forecast_wb):
            low = name.lower()
            if "911" in low and "forecast" in low:
                ws = forecast_wb[name]
                break
    if ws is None:
        raise RuntimeError(
            "Could not find a '911 Forecast' sheet in the Working Forecast List. "
            f"Sheets present: {forecast_wb.sheetnames}")

    header_row = nest_col = None
    for r in range(1, min(ws.max_row, 8) + 1):
        col = _find_header_col(ws, "Nest", r)
        if col is not None:
            header_row, nest_col = r, col
            break
    if nest_col is None:
        raise RuntimeError(
            f"Could not find a 'Nest' column header in the '{ws.title}' sheet of "
            "the Working Forecast List (looked in the first 8 rows).")

    po = _find_header_col(ws, "PO", header_row) or 1
    line = _find_header_col(ws, "Line", header_row) or 2
    batch = _find_header_col_prefix(ws, "Batch", header_row) or 3
    log(f"  Forecast lookup: sheet '{ws.title}', nest column "
        f"{get_column_letter(nest_col)} (headers in row {header_row}); A/B/C <- "
        f"{get_column_letter(po)}/{get_column_letter(line)}/{get_column_letter(batch)}.")
    return ws, header_row, nest_col, (po, line, batch)


def _copy_forecast_rows(ws, header_row: int, nest_col: int,
                        out_cols: tuple, nest_number: str) -> list:
    """All (PO, Line, Batch) tuples whose 'Nest' cell matches nest_number, read
    from the header-located columns (see _locate_forecast). Tolerant of nests
    stored as int/str/float. These land in NEST cols A/B/C."""
    target = str(nest_number).strip().upper()
    po, line, batch = out_cols
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        if _nest_matches(ws.cell(row, nest_col).value, target):
            rows.append((
                ws.cell(row, po).value,
                ws.cell(row, line).value,
                ws.cell(row, batch).value,
            ))
    return rows


def _paste_forecast_into_nest(nest_ws, forecast_rows: list):
    """
    Paste forecast_rows into NEST sheet cols A, B, C starting at row 4.
    """
    for i, (a, b, c) in enumerate(forecast_rows):
        dest_row = 4 + i
        nest_ws.cell(dest_row, 1).value = a
        nest_ws.cell(dest_row, 2).value = b
        nest_ws.cell(dest_row, 3).value = c


# ---------------------------------------------------------------------------
# Step 6: PDF -> MIL spec + MATERIAL
#
# The authoritative source for both values is the MOVE TICKET page in the
# nest's own work packet, which carries explicit labeled fields:
#       MIL SPEC: MIL-S-22698
#       MATERIAL:
#       HSS
# (the value sits on the same line for MIL SPEC, the next line for MATERIAL).
#
# The previous version read MIL-S via a bare `MIL-S-\d+` scan across EVERY PDF
# in NEST PACKAGES (so a nest could inherit another nest's spec) and read the
# material from the PART SKETCH `MATL:` field -- which is blank on many parts,
# letting the old `MATL:\s*(\S+)` regex skip the blank and grab the next field
# label (e.g. "LVL:"). Both are fixed by scoping to the nest's own packet and
# reading the labeled MOVE TICKET fields.
# ---------------------------------------------------------------------------

_LABEL_RE = re.compile(r'^[A-Z][A-Z0-9 ./#-]*:')


def _looks_like_label(s: str) -> bool:
    """True if s reads as another field label (e.g. 'DRA REV:', 'LVL:')."""
    return bool(_LABEL_RE.match(s.strip()))


def _labeled_value(text: str, label: str):
    """
    Return the value following 'label:' in PDF text. Handles both
    'LABEL: value' (same line) and 'LABEL:\\nvalue' (value on the next line).
    Returns None when the field is blank or the following token is itself
    another field label.
    """
    pat = re.compile(
        re.escape(label) + r'\s*:[ \t]*(?P<same>[^\n]*)(?:\n[ \t]*(?P<next>[^\n]*))?',
        re.IGNORECASE,
    )
    m = pat.search(text)
    if not m:
        return None
    same = (m.group("same") or "").strip()
    if same and not _looks_like_label(same):
        return same
    nxt = (m.group("next") or "").strip()
    if nxt and not _looks_like_label(nxt):
        return nxt
    return None


def _find_nest_pdf(nest_packages_folder: Path, nest_number: str):
    """Return the PDF in NEST PACKAGES whose stem contains the nest number,
    or None. Shared by the MIL/MATERIAL read (Step 6) and the drawings
    extraction (Step 9) so both look at the same nest-scoped packet."""
    if not nest_packages_folder.exists():
        return None
    nest_upper = nest_number.upper()
    for f in nest_packages_folder.iterdir():
        if (f.is_file() and f.suffix.lower() == ".pdf"
                and nest_upper in f.stem.upper()):
            return f
    return None


def _extract_pdf_data(pdf_path: Path) -> tuple:
    """
    Parse a single nest packet PDF and return (mil_spec, matl_type).

    MIL spec is read from the labeled 'MIL SPEC:' field, falling back to a
    bare 'MIL-S-...' token if the label is absent. Material is read from the
    labeled 'MATERIAL:' field only (no fallback -- a blank field stays blank).
    Returns (None, None) if neither is found.
    """
    if not PYMUPDF_AVAILABLE:
        raise ImportError(
            "PyMuPDF is required for PDF parsing. "
            "Install it with: pip install pymupdf"
        )

    sdk.ensure_local(pdf_path)  # OneDrive placeholder -> download first (Hard Rule 13)
    doc = fitz.open(str(pdf_path))
    full_text = "".join(page.get_text() for page in doc)
    doc.close()

    mil_spec = _labeled_value(full_text, "MIL SPEC")
    if not mil_spec:
        m = re.search(r'MIL-S-\S+', full_text, re.IGNORECASE)
        mil_spec = m.group(0) if m else None

    matl_type = _labeled_value(full_text, "MATERIAL")

    return mil_spec, matl_type


def _get_pdf_data_for_nest(nest_packages_folder: Path, nest_number: str, log) -> tuple:
    """
    Read (mil_spec, matl_type) from THIS nest's own packet PDF in NEST PACKAGES
    (the one whose filename contains the nest number). Scoping to the nest's
    own packet prevents one nest from inheriting another nest's MIL spec.
    """
    if not nest_packages_folder.exists():
        log(f"  WARNING: NEST PACKAGES folder not found: {nest_packages_folder}")
        return None, None

    nest_pdf = _find_nest_pdf(nest_packages_folder, nest_number)
    if nest_pdf is None:
        log(f"  WARNING: No packet PDF containing '{nest_number}' found in NEST PACKAGES.")
        return None, None

    log(f"  Parsing packet PDF: {nest_pdf.name}")
    try:
        return _extract_pdf_data(nest_pdf)
    except Exception as e:
        log(f"  WARNING: Could not parse {nest_pdf.name}: {e}")
        return None, None


# ---------------------------------------------------------------------------
# Step 7b: Part Sketch extraction -> MOVE TICKET OMIT PDF
# ---------------------------------------------------------------------------

_omit_stamps = None


def _load_omit_stamp_helpers(log):
    """
    Import the sibling 911_remove_ticket plugin, the single home of the
    cover-stamp helpers (batch/nest text box + Material Type fill). Both
    plugins ship in the same plugins dir so they update together. Returns
    the module, or None (with one logged warning) if unavailable.
    """
    global _omit_stamps
    if _omit_stamps is not None:
        return _omit_stamps or None
    try:
        path = Path(__file__).resolve().parents[1] / "911_remove_ticket" / "run.py"
        spec = importlib.util.spec_from_file_location(
            "techdeck_911_remove_ticket_for_setup", str(path))
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _omit_stamps = mod
    except Exception as e:
        log(f"  WARNING: 911 Remove Ticket helpers unavailable ({e}) - omit PDFs will not be stamped.")
        _omit_stamps = False
    return _omit_stamps or None


def _extract_nest_drawings(nest_packages_folder: Path, nest_number: str,
                            dest_path: Path, log, batch: str = "",
                            material_hint: str = "", difficulty=None) -> bool:
    """
    Build the MOVE TICKET OMIT PDF for a nest.

    Starts with every page in the nest's source PDF and removes pages that
    contain "MOVE TICKET" text.  Pages that contain "MIL-SPEC" or "HULL"
    are always kept, even if they also contain "MOVE TICKET".

    The first page is then stamped (via the sibling 911_remove_ticket
    helpers): "BATCH {batch} - NEST {nest}" under the Quality Requirements
    grid, and the blank Material Type cell filled with the MATERIAL read
    off the removed move tickets (falling back to material_hint, the Step-6
    value).

    Returns True if the output PDF was written successfully.
    """
    if not nest_packages_folder.exists():
        log("  WARNING: NEST PACKAGES folder not found -- skipping drawings.")
        return False

    matching_pdf = _find_nest_pdf(nest_packages_folder, nest_number)

    if matching_pdf is None:
        log(f"  WARNING: No PDF containing '{nest_number}' found in NEST PACKAGES.")
        return False

    log(f"  Found nest PDF: {matching_pdf.name}")

    doc = None
    try:
        sdk.ensure_local(matching_pdf)  # OneDrive placeholder -> download first (Hard Rule 13)
        doc = fitz.open(str(matching_pdf))
        total_pages = len(doc)

        stamps = _load_omit_stamp_helpers(log)
        if stamps:
            remove_pages, materials = stamps._scan_document(doc)
        else:
            materials = set()
            remove_pages = set()
            for i in range(total_pages):
                text = (doc[i].get_text("text") or "").upper()
                if "MOVE TICKET" in text and "MIL-SPEC" not in text and "HULL" not in text:
                    remove_pages.add(i)

        if len(remove_pages) >= total_pages:
            log(f"  WARNING: All {total_pages} pages are MOVE TICKET pages - nothing to write for {nest_number}.")
            return False

        if remove_pages:
            doc.delete_pages(sorted(remove_pages))

        if stamps:
            material = " / ".join(sorted(materials)) or (material_hint or "")
            for w in stamps._stamp_first_page(doc[0], batch, nest_number, material,
                                              log, difficulty):
                log(f"  WARNING: {w}")

        doc.save(str(dest_path), garbage=3, deflate=True)

        kept = total_pages - len(remove_pages)
        log(f"  Drawings PDF: {dest_path.name} ({kept} page(s), removed {len(remove_pages)} MOVE TICKET page(s))")
        return True

    except Exception as e:
        log(f"  WARNING: Could not write drawings PDF: {e}")
        return False
    finally:
        if doc is not None:
            doc.close()


# ---------------------------------------------------------------------------
# Step 8: Inspection sheet naming helpers
# ---------------------------------------------------------------------------

def _basic_suffix(dypn: str) -> str:
    """Sheet name = part after the last '-' (e.g. 'H4533321-80' -> '-80')."""
    parts = dypn.rsplit('-', 1)
    return f"-{parts[-1]}" if len(parts) == 2 else dypn


def _disambiguated_name(dypn: str) -> str:
    """
    Used when two DYPNs share a suffix.
    H4533321-80 -> '21-80', H4533322-80 -> '22-80'
    (last 2 chars of the head + '-' + suffix tail)
    """
    segments = dypn.rsplit('-', 1)
    if len(segments) == 2:
        head, tail = segments
        return f"{head[-2:]}-{tail}"
    return dypn


def _plan_sheet_names(dypns: list, existing_sheet_names: set = None) -> list:
    """
    Build the list of (full_dypn, sheet_name) pairs for the inspection
    sheet copy step. If two DYPNs share a suffix, *both* get the
    disambiguated name. Falls back to '<name> (n)' for any remaining
    collision (e.g. identical DYPNs).
    """
    if existing_sheet_names is None:
        existing_sheet_names = set()

    suffix_counts = {}
    for dypn in dypns:
        s = _basic_suffix(dypn)
        suffix_counts[s] = suffix_counts.get(s, 0) + 1

    used = set()
    plan = []
    for full_dypn in dypns:
        suffix = _basic_suffix(full_dypn)
        name = _disambiguated_name(full_dypn) if suffix_counts.get(suffix, 0) > 1 else suffix
        name = name[:31]

        base = name
        counter = 2
        while name in used or name in existing_sheet_names:
            name = f"{base} ({counter})"[:31]
            counter += 1
        used.add(name)
        plan.append((full_dypn, name))
    return plan


def _get_dypn_rows(nest_ws) -> list:
    """
    Read NEST cols F/G/I (Work Order / DYPN / DYPN QTY) from row 4 downward.
    Return a list of (dypn, work_order, qty) triples for every row with a
    DYPN. qty falls back to 1 on blank/junk (the sheet still gets one slot).
    """
    result = []
    for row in range(4, nest_ws.max_row + 1):
        val = nest_ws.cell(row, 7).value  # Column G = DYPN
        if not (val and str(val).strip()):
            continue
        wo = str(nest_ws.cell(row, 6).value or "").strip()   # Column F
        try:
            qty = int(float(str(nest_ws.cell(row, 9).value).strip()))  # Column I
        except (TypeError, ValueError):
            qty = 1
        result.append((str(val).strip(), wo, max(qty, 1)))
    return result


# ---------------------------------------------------------------------------
# Inspection sheet copy step uses Excel COM, not openpyxl.
#
# openpyxl's copy_worksheet does not propagate conditional formatting rules
# to the new sheet (verified empirically: 350 CF rules on the template ->
# 0 CF rules on every copy). The template's grey->white field-fill logic
# is driven entirely by those CF rules, so the copies look "completely
# white". Excel's native Sheet.Copy preserves the CF perfectly.
#
# openpyxl is still used for the NEST data writes -- those don't go through
# copy_worksheet, and the CF on the source INSPECTION SHEET tab survives
# the openpyxl load/save round-trip intact.
# ---------------------------------------------------------------------------

def _build_inspection_sheets_via_excel(workbook_path: Path, parts: list, log):
    """
    Copy the INSPECTION SHEET tab once per NEST part row. parts is the
    list of (full_dypn, work_order, qty) triples from _get_dypn_rows.

    v1.6.0 (QA feedback, C.D. 2026-07): each copy now fills one
    Part Number slot PER PIECE — qty 3 writes the part into the first 3
    of the form's 10 two-row slots (A16, A18, ... A34) so the inspector
    records each piece; qty > 10 fills all 10 with a warning. Every slot
    shows the HULL CODE under the part number (= the work order's first
    two characters — verified identical to the part sketches' HULL field
    across V094). Each duplicate slot is un-greyed (its conditional-format
    helper cell BB{r} is stamped, since the form only whitens slot 1 off
    A16) and rebuilt to slot 1's TWO-row Part Description layout — source
    material CODE (=D$16) on top, DESCRIPTION (=D$17) below — replacing the
    template's broken single-cell #REF! stub so every piece shows the same
    code + description as the first. After the copies, the INSPECTION SHEET
    template tab is hidden (her April ask) so the workbook opens straight
    onto real sheets.

    Excel COM is used so conditional formatting on the copies is
    preserved (openpyxl's copy_worksheet drops it).

    Implementation note: Excel SaveAs writing back to a OneDrive-synced
    path fails with "Cannot access" because OneDrive is tracking the
    file. So we do all COM work on a local temp copy outside OneDrive,
    then atomically replace the original via os.replace.

    Sheet name = suffix after the last '-' (e.g. '-80'). When two DYPNs
    share a suffix, both names are disambiguated (e.g. H4533321-80 ->
    '21-80', H4533322-80 -> '22-80').
    """
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        raise RuntimeError(
            "Excel COM bindings (pywin32) are required by 911 Setup."
        )

    import os
    import tempfile

    plan = _plan_sheet_names([p[0] for p in parts])

    xlCalculationManual = -4135
    xlCalculationAutomatic = -4105
    msoAutomationSecurityForceDisable = 3

    def _silence(app):
        """Disable every dialog category Excel can throw at us."""
        app.Visible = False
        app.DisplayAlerts = False
        app.ScreenUpdating = False
        app.EnableEvents = False
        app.AskToUpdateLinks = False
        for prop, val in (
            ("AlertBeforeOverwriting", False),
            ("FeatureInstall", 0),
            ("AutomationSecurity", msoAutomationSecurityForceDisable),
            ("Calculation", xlCalculationManual),
        ):
            try:
                setattr(app, prop, val)
            except Exception:
                pass

    pythoncom.CoInitialize()
    excel = None
    wb = None
    tmp_dir = None
    try:
        # Stage the file outside OneDrive so Excel never sees a synced path
        tmp_dir = tempfile.mkdtemp(prefix="techdeck_911_")
        local_copy = Path(tmp_dir) / workbook_path.name
        sdk.copy_resilient(workbook_path, local_copy, log)  # hydrate + locked-open message
        log(f"  Staging via local temp: {local_copy}")

        excel = win32com.client.DispatchEx("Excel.Application")
        _silence(excel)

        wb = excel.Workbooks.Open(
            Filename=str(local_copy),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            Notify=False,
            AddToMru=False,
        )
        try:
            wb.CheckCompatibility = False
        except Exception:
            pass

        template = wb.Sheets("INSPECTION SHEET")

        # The QF-QU-09 form has 10 two-row Part Number slots: A16:C17,
        # A18:C19, ... A34:C35.
        SLOT_ROWS = list(range(16, 36, 2))

        for (full_dypn, sheet_name), (_, work_order, qty) in zip(plan, parts):
            count_before = wb.Sheets.Count
            last_sheet = wb.Sheets(count_before)

            # Excel COM signature is Worksheet.Copy(Before, After).
            # Passing After= as a kwarg via pywin32 late binding has been
            # observed to silently fall through to Copy() with no args
            # (which spawns a new orphan workbook). Use positional args.
            template.Copy(None, last_sheet)

            try:
                excel.CutCopyMode = False
            except Exception:
                pass

            count_after = wb.Sheets.Count
            if count_after != count_before + 1:
                # Fallback: Copy()-with-no-args creates a new workbook;
                # move that sheet back into our workbook.
                log(f"  WARNING: in-place Copy did not grow sheet count "
                    f"({count_before} -> {count_after}); trying Copy()+Move fallback")
                try:
                    template.Copy()
                    new_wb = excel.ActiveWorkbook
                    if new_wb is not None and new_wb.Name != wb.Name:
                        new_wb.Sheets(1).Move(None, wb.Sheets(wb.Sheets.Count))
                        try:
                            new_wb.Close(SaveChanges=False)
                        except Exception:
                            pass
                except Exception as e:
                    raise RuntimeError(
                        f"Copy+Move fallback failed for {full_dypn}: {e}"
                    )
                if wb.Sheets.Count != count_before + 1:
                    raise RuntimeError(
                        f"Sheet.Copy did not create a new sheet for {full_dypn}. "
                        f"Workbook still has {wb.Sheets.Count} sheets."
                    )

            # After Copy/Move the new sheet is the active sheet. Don't
            # index by position -- the new copy isn't necessarily at the
            # end of the tab order, so wb.Sheets(wb.Sheets.Count) can
            # return the wrong sheet and we'd rename SOURCE MATERIAL INFO
            # 22 times instead of the actual copies.
            new_sheet = excel.ActiveSheet
            new_sheet.Name = sheet_name

            # Hull code shows on its own line under the part number in
            # every filled slot (merged 2-row cell, so wrap displays it).
            hull = work_order[:2].upper() if work_order else ""
            slot_text = f"{full_dypn}\n{hull}" if hull else full_dypn

            n_slots = min(qty, len(SLOT_ROWS))
            if qty > len(SLOT_ROWS):
                log(f"  WARNING: {full_dypn} qty {qty} exceeds the form's "
                    f"{len(SLOT_ROWS)} sample slots - filling {len(SLOT_ROWS)}.")

            for slot in range(n_slots):
                r = SLOT_ROWS[slot]
                cell = new_sheet.Range(f"A{r}")
                cell.Value = slot_text
                cell.WrapText = True
                if slot > 0:
                    # --- Un-grey the duplicate slot (QA feedback 2026-07) --
                    # The form's conditional formatting greys the whole
                    # A{r}:I{r+1} block while helper cell BB{r} is blank
                    # (slot 1 keys off A16 instead, so filling A16 already
                    # whitens it). The template leaves BB{r} empty for slots
                    # 2+, so every duplicate piece stayed grey. Stamp BB{r}
                    # with slot 1's part reference so the rule turns the row
                    # white, matching slot 1.
                    new_sheet.Range(f"BB{r}").Formula = "=BB$14"

                    # --- Two-row description like slot 1 ------------------
                    # Slot 1 shows the source-material CODE (D16) on the top
                    # row and its DESCRIPTION (D17) on the row below, as two
                    # separate merged cells. The template pre-merges slots 2+
                    # into ONE cell (D{r}:G{r+1}) carrying a broken #REF!
                    # formula, so duplicates showed only the code on one line.
                    # Rebuild slot 1's two-row layout and mirror it (every
                    # slot on the sheet is the same part).
                    xlPasteFormats = -4122
                    try:
                        new_sheet.Range(f"D{r}:G{r+1}").UnMerge()
                    except Exception:
                        pass
                    # Copy slot 1's formatting (borders + the two 1-row merges)
                    new_sheet.Range("D16:G17").Copy()
                    new_sheet.Range(f"D{r}").PasteSpecial(Paste=xlPasteFormats)
                    excel.CutCopyMode = False
                    # Guarantee the two 1-row merges regardless of paste behaviour
                    for rr in (r, r + 1):
                        try:
                            new_sheet.Range(f"D{rr}:G{rr}").Merge()
                        except Exception:
                            pass
                    new_sheet.Range(f"D{r}").Formula = "=D$16"     # code
                    new_sheet.Range(f"D{r + 1}").Formula = "=D$17"  # description

            log(f"  Creating inspection sheet '{sheet_name}' for {full_dypn} "
                f"(qty {qty}, hull {hull or '?'})")

        # Hide the template tab so the workbook opens onto the real
        # inspection sheets (April QA ask). Copies were already made,
        # and at least one other sheet is always visible.
        if plan:
            try:
                template.Visible = 0  # xlSheetHidden
            except Exception:
                pass

        try:
            excel.Calculation = xlCalculationAutomatic
        except Exception:
            pass

        wb.Save()
        wb.Close(SaveChanges=False)
        wb = None

        # Tear Excel down before moving the file back -- ensures all
        # handles are released so os.replace can overwrite cleanly.
        try:
            excel.Quit()
        except Exception:
            pass
        excel = None

        # Atomic replace back into OneDrive. os.replace is atomic on
        # Windows when source and target are on the same volume.
        os.replace(str(local_copy), str(workbook_path))
        log(f"  Wrote final workbook -> {workbook_path}")
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.EnableEvents = True
                excel.DisplayAlerts = True
                excel.ScreenUpdating = True
            except Exception:
                pass
            try:
                excel.Quit()
            except Exception:
                pass
        excel = None
        if tmp_dir is not None:
            try:
                shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception:
                pass
        pythoncom.CoUninitialize()


# ---------------------------------------------------------------------------
# Template finder
# ---------------------------------------------------------------------------

def _find_template_911(template_dir: Path) -> Path:
    """
    Find '911 BATCH _.xlsx' in the template directory.
    Matches any .xlsx whose name starts with '911 BATCH' (case-insensitive).
    """
    for f in template_dir.iterdir():
        if (f.is_file()
                and f.suffix.lower() == ".xlsx"
                and f.stem.upper().startswith("911 BATCH")
                and not f.name.startswith("~")):
            return f

    raise FileNotFoundError(
        f"Could not find a '911 BATCH _.xlsx' template in {template_dir}"
    )


# ===========================================================================
# Generate Teams Cards (v1.8.0)
# ===========================================================================
# One card per nest that is WAITING to be set up, posted to the MODELING
# bucket of the SOPO D911 PIPELINE plan (D922 channel) through a Power
# Automate webhook -- the same "TechDeck never touches Planner directly"
# pattern 922 Setup uses (flow recipe: docs/TEAMS_CARDS.md, flow #3).
#
# The work list is NOT the batch folder: it is the EB 922 Schedule's
# "CURRENT PIPELINE" sheet, every row where DEPT. is 911 and STATUS is
# "NEED TEAMS/SETUP". That makes the stage batch-independent -- it cards the
# whole 911 queue, whichever batch each nest belongs to -- so it runs before
# (and without) the batch prompt.
#
# Per row:
#   col B "BATCH / NEST"  "V092 503836" -> batch V092, nest 503836
#   col C "DATE"          -> the card's due date
#   col D "NOTES"         -> source-material text -> SAW CUT / TUBE LASER
#                            label, and whether the "Program" checklist item
#                            applies (tube stock only)
#   col E "RATING"        -> SIMPLE / MEDIUM / DIFFICULT, read from the CELL
#                            COLOUR (the cell holds no text) via the sibling
#                            911 Remove Ticket helpers -- the single home of
#                            that logic, shared with the packet stamp
#
# The "(#)" in the card title is the nest's EB source-material stock code
# (e.g. 211076345), which lives in the batch's own BATCH LIST 'Material'
# column -- not on the schedule. Each referenced batch's BATCH LIST is read
# once and cached.
# ---------------------------------------------------------------------------

# The 'TechDeck 911 Setup - Create Modeling Cards' Power Automate flow.
# EMPTY until that flow is built (docs/TEAMS_CARDS.md, flow #3): with no URL
# the stage forces a dry run, logs every card it WOULD create and writes the
# payload preview, so nothing posts by accident. Bake the real URL in here
# once the flow exists (same pattern as 922 Setup's DEFAULT_WEBHOOK_URL).
DEFAULT_CARD_WEBHOOK_URL = ""

_CARD_PREVIEW_FILENAME = "last_911_setup_payload.json"

# Schedule sheet + the headers we read, all looked up BY NAME (Hard Rules 1-2)
# -- the live sheet ships them with trailing spaces ('DEPT. ', 'STATUS ').
_SCHED_SHEET = "CURRENT PIPELINE"
_SCHED_DEPT = "DEPT."
_SCHED_KEY = "BATCH / NEST"
_SCHED_DATE = "DATE"
_SCHED_NOTES = "NOTES"
_SCHED_RATING = "RATING"
_SCHED_STATUS = "STATUS"


def _load_card_template() -> dict:
    """Load card_template.json sitting next to this file."""
    with open(Path(__file__).with_name("card_template.json"),
              "r", encoding="utf-8") as fh:
        return json.load(fh)


def _norm_text(value) -> str:
    """Uppercase, whitespace-collapsed form used for every schedule match."""
    return re.sub(r"\s+", " ", str(value if value is not None else "")).strip().upper()


def _norm_status(value) -> str:
    """_norm_text plus tightened slashes, so a hand-typed 'NEED TEAMS / SETUP'
    still matches the canonical 'NEED TEAMS/SETUP'."""
    return re.sub(r"\s*/\s*", "/", _norm_text(value))


def _split_batch_nest(value):
    """'V092 503836' -> ('V092', '503836'); 'V085 S20085' -> ('V085','S20085').

    Batch first, nest LAST -- never a digits-only match, because nests are not
    always numeric (Hard Rule 3's alphanumeric-nest class). Returns
    (None, None) when the cell does not carry both halves.
    """
    tokens = _norm_text(value).split()
    if len(tokens) < 2:
        return None, None
    return tokens[0], tokens[-1]


def _due_date_iso(value) -> str:
    """Schedule DATE cell -> an ISO-8601 instant Planner accepts, or "".

    Non-dates are expected and fine: the column also carries 'HOLD' and 'N/A'
    for nests with no scheduled date, which simply means no due date.
    """
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%dT00:00:00Z")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%dT00:00:00Z")
    return ""


def _resolve_machine(material_text: str, template: dict):
    """'SAW CUT' / 'TUBE LASER' / None for a source-material description.

    The rule, in order (C.D. 2026-08-03):
      1. the material says TUBE -> it IS a tube -> TUBE LASER
      2. otherwise a '(TL)' note means it is going on the tube laser anyway
      3. everything else -> SAW CUT

    Shape does NOT imply the tube laser -- an angle is an angle, and the
    '(TL)' marker is exactly how the schedule flags the exceptions. Only a
    completely EMPTY material text returns None, so that card is created
    unlabelled and reported rather than guessed at.
    """
    text = _norm_text(material_text)
    if not text:
        return None
    if _is_tube(text, template):
        return "TUBE LASER"
    for marker in (template.get("machine_tube_markers") or ["TL"]):
        if re.search(r"\(\s*" + re.escape(_norm_text(marker)) + r"\s*\)", text):
            return "TUBE LASER"
    return template.get("machine_default", "SAW CUT")


def _is_tube(material_text: str, template: dict) -> bool:
    """True when the source material is tube stock (drives the 'Program'
    checklist item -- a tube-laser program is meaningless on saw stock)."""
    text = _norm_text(material_text)
    return any(_norm_text(k) in text
               for k in (template.get("tube_keywords") or ["TUBE"]))


def _read_schedule_rows(params, template, log, cancel_event):
    """Every CURRENT PIPELINE row that needs a card -> (rows, problem_text).

    ``problem_text`` is non-empty when the whole lookup was unavailable (the
    schedule is missing, open in Excel, or has no such sheet); the caller
    reports it and skips the stage rather than posting a half-built payload.

    Each row dict: batch, nest, date, notes, difficulty, excel_row.
    """
    stamps = _load_omit_stamp_helpers(log)
    if stamps is None:
        return [], ("The 911 Remove Ticket helpers could not be loaded, so the "
                    "EB 922 Schedule's difficulty colours cannot be read and "
                    "no Teams cards were created.")

    path = stamps._schedule_path(params)
    if path is None:
        return [], ("The EB 922 Schedule workbook could not be found, so no "
                    "Teams cards were created. Check that the '922 QTDR "
                    "Production Packages' folder is synced, or set the "
                    "'EB 922 Schedule' path in this plugin's Settings.")
    log(f"Schedule      : {path}")
    try:
        wb = sdk.load_workbook_resilient(path, log=log, data_only=True)
    except Exception as exc:
        return [], (f"The Teams-card work list could not be read from "
                    f"{path.name}:\n\n{exc}\n\nNo Teams cards were created.")

    try:
        if _SCHED_SHEET not in wb.sheetnames:
            return [], (f"{path.name} has no '{_SCHED_SHEET}' sheet, so no "
                        f"Teams cards were created.")
        ws = wb[_SCHED_SHEET]
        hdr_row, hdr = sdk.find_header_row(ws, [_SCHED_KEY, _SCHED_STATUS])
        if not hdr_row:
            return [], (f"{path.name} has no row containing both "
                        f"'{_SCHED_KEY}' and '{_SCHED_STATUS}' on the "
                        f"'{_SCHED_SHEET}' sheet, so no Teams cards were "
                        f"created.")
        c_key = hdr.get(_SCHED_KEY)
        c_status = hdr.get(_SCHED_STATUS)
        c_dept = hdr.get(_SCHED_DEPT)
        c_date = hdr.get(_SCHED_DATE)
        c_notes = hdr.get(_SCHED_NOTES)
        c_rating = hdr.get(_SCHED_RATING)
        theme_rgbs = stamps._theme_rgbs(wb)

        want_dept = _norm_text(template.get("schedule_dept", "911"))
        want_status = _norm_status(template.get("schedule_status",
                                                "NEED TEAMS/SETUP"))

        rows = []
        for r in range(hdr_row + 1, ws.max_row + 1):
            if r % 64 == 0:
                sdk.raise_if_cancelled(cancel_event)
            if c_dept and _norm_text(ws.cell(r, c_dept).value) != want_dept:
                continue
            if _norm_status(ws.cell(r, c_status).value) != want_status:
                continue
            batch, nest = _split_batch_nest(ws.cell(r, c_key).value)
            difficulty = None
            if c_rating:
                difficulty = stamps._match_fill(
                    stamps._fill_rgb(ws.cell(r, c_rating), theme_rgbs))
            rows.append({
                "excel_row": r,
                "raw_key": str(ws.cell(r, c_key).value or "").strip(),
                "batch": batch,
                "nest": nest,
                "date": ws.cell(r, c_date).value if c_date else None,
                "notes": ws.cell(r, c_notes).value if c_notes else None,
                "difficulty": difficulty,
            })
        return rows, ""
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_batch_list_materials(qtdr_root: Path, batch: str, log):
    """{NEST: {'code': EB stock code, 'desc': description}} for one batch.

    The card title's "(#)" is this stock code -- the schedule does not carry
    it, the batch's own BATCH LIST does ('Material' column, keyed by
    'Nest Pkg Nbr'). The description doubles as the source-material fallback
    for schedule rows whose NOTES cell is blank.

    Returns (mapping, warning_text); an unreadable/absent BATCH LIST yields
    ({}, why) and the batch's cards are simply created without a code.
    """
    batch_folder = qtdr_root / batch
    if not batch_folder.is_dir():
        return {}, (f"No batch folder '{batch}' under {qtdr_root} - its cards "
                    f"have no source-material code.")
    try:
        path = _find_batch_list(batch_folder, batch)
    except FileNotFoundError as exc:
        return {}, f"{exc} Cards for batch {batch} have no source-material code."

    try:
        wb = sdk.load_workbook_resilient(path, log=log, data_only=True,
                                         read_only=True)
    except Exception as exc:
        return {}, (f"Could not read {path.name}: {exc} - cards for batch "
                    f"{batch} have no source-material code.")
    try:
        ws = wb["BATCH"] if "BATCH" in wb.sheetnames else wb[wb.sheetnames[0]]
        hdr_row, hdr = sdk.find_header_row(ws, ["NEST PKG NBR", "MATERIAL"])
        if not hdr_row:
            return {}, (f"{path.name} has no 'Nest Pkg Nbr' + 'Material' header "
                        f"row - cards for batch {batch} have no source-material "
                        f"code.")
        c_nest = hdr["NEST PKG NBR"]
        c_mat = hdr["MATERIAL"]
        c_desc = hdr.get("DESCRIPTION")
        out: dict = {}
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            nest = row[c_nest - 1] if c_nest - 1 < len(row) else None
            key = _norm_text(nest)
            if not key or key in out:
                continue
            code = row[c_mat - 1] if c_mat - 1 < len(row) else None
            desc = row[c_desc - 1] if c_desc and c_desc - 1 < len(row) else None
            if code is None and desc is None:
                continue
            out[key] = {"code": str(code).strip() if code is not None else "",
                        "desc": str(desc).strip() if desc is not None else ""}
        return out, ""
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _build_card(row: dict, material: dict, template: dict, label_map: dict,
                warnings: list, unlabelled: list):
    """One schedule row -> one Teams card dict (plus its label NAMES for the log)."""
    batch, nest = row["batch"], row["nest"]
    code = (material or {}).get("code", "")
    # NOTES is the authority (it carries the (TL)/(SAW) overrides); the BATCH
    # LIST description is the fallback for the rows planning has not annotated
    # yet, so a blank NOTES cell still resolves a machine instead of nothing.
    material_text = str(row.get("notes") or "").strip() or (material or {}).get("desc", "")

    if code:
        title = template.get("title_format",
                             "BATCH: {batch} - NEST: {nest} ({material})").format(
            batch=batch, nest=nest, material=code)
    else:
        title = template.get("title_format_no_material",
                             "BATCH: {batch} - NEST: {nest}").format(
            batch=batch, nest=nest)

    # Checklist: the tube-only items (Program) drop out on saw stock.
    checklist = list(template.get("checklist", []))
    if not _is_tube(material_text, template):
        tube_only = {_norm_text(t) for t in (template.get("tube_only_checklist") or [])}
        checklist = [t for t in checklist if _norm_text(t) not in tube_only]

    names = []
    if row.get("difficulty"):
        names.append(row["difficulty"])
    machine = _resolve_machine(material_text, template)
    if machine:
        names.append(machine)
    else:
        unlabelled.append(f"{batch} {nest}"
                          + (f" ({material_text})" if material_text else " (no material listed)"))

    slots = []
    for name in names:
        slot = label_map.get(_norm_text(name))
        if slot:
            slots.append(slot)
        else:
            warnings.append(f"No Teams label mapped for '{name}' - skipped on "
                            f"{batch} {nest} (add it to card_template.json's "
                            f"label_map AND to the plan's labels).")

    card = {
        "title": title,
        "bucket": template.get("bucket", "MODELING"),
        "priority": template.get("priority", "Medium"),
        "status": template.get("status", "Not started"),
        "dueDate": _due_date_iso(row.get("date")),
        "checklist": checklist,
        "labels": slots,
    }
    return card, names


def _run_teams_cards(params: dict, progress_callback, cancel_event,
                     qtdr_override: str, lo: int = 0, hi: int = 100) -> bool:
    """The Generate Teams Cards stage. Returns True when the payload posted
    (or dry-ran) cleanly, False when the stage could not run."""
    log = params.get("log", print)
    settings = params.get("settings", {}) or {}

    log(f"\n{'='*50}")
    log("Generate Teams Cards")
    log(f"{'='*50}")

    try:
        template = _load_card_template()
    except (OSError, json.JSONDecodeError) as exc:
        log(f"ERROR: could not read card_template.json: {exc}")
        return False

    def _pct(frac):
        progress_callback(lo + int((hi - lo) * frac))

    rows, problem = _read_schedule_rows(params, template, log, cancel_event)
    if problem:
        log(f"ERROR: {problem}")
        sdk.show_warning(params, "911 Setup - Teams cards not created", problem)
        return False
    if not rows:
        log(f"No rows on '{_SCHED_SHEET}' are DEPT. "
            f"{template.get('schedule_dept', '911')} + STATUS "
            f"'{template.get('schedule_status', 'NEED TEAMS/SETUP')}' - "
            f"nothing to card.")
        _pct(1.0)
        return True
    log(f"Work list     : {len(rows)} nest(s) marked "
        f"'{template.get('schedule_status', 'NEED TEAMS/SETUP')}'.")
    _pct(0.25)

    warnings: list = []
    unlabelled: list = []
    skipped = [r for r in rows if not (r["batch"] and r["nest"])]
    for r in skipped:
        warnings.append(f"Schedule row {r['excel_row']} ('{r['raw_key']}') does "
                        f"not read as 'BATCH NEST' - no card created.")
    rows = [r for r in rows if r["batch"] and r["nest"]]

    # --- Source-material stock codes: one BATCH LIST read per batch ---------
    qtdr_root = _base_qtdr(qtdr_override)
    materials: dict = {}
    for batch in sorted({r["batch"] for r in rows}):
        sdk.raise_if_cancelled(cancel_event)
        mapping, warn = _read_batch_list_materials(qtdr_root, batch, log)
        materials[batch] = mapping
        if warn:
            warnings.append(warn)
    _pct(0.6)

    # --- Build the cards ----------------------------------------------------
    label_map = {_norm_text(name): slot
                 for name, slot in (template.get("label_map") or {}).items()}
    cards, card_labels, no_code = [], [], []
    for r in rows:
        sdk.raise_if_cancelled(cancel_event)
        material = (materials.get(r["batch"]) or {}).get(r["nest"])
        if not material or not material.get("code"):
            no_code.append(f"{r['batch']} {r['nest']}")
        card, names = _build_card(r, material, template, label_map,
                                  warnings, unlabelled)
        cards.append(card)
        card_labels.append(names)

    # A card with no difficulty label looks identical to a rated one at a
    # glance, so an uncoloured RATING cell is surfaced the same way the packet
    # stamp surfaces it -- loudly -- rather than passing as a clean run.
    unrated = [f"{r['batch']} {r['nest']}" for r in rows if not r.get("difficulty")]
    if unrated:
        warnings.append(f"No difficulty rating colour on the EB 922 Schedule "
                        f"(CURRENT PIPELINE, column E) for {len(unrated)} "
                        f"nest(s), so their cards carry NO difficulty label: "
                        f"{', '.join(unrated[:15])}"
                        + (" ..." if len(unrated) > 15 else ""))
    if no_code:
        warnings.append(f"No BATCH LIST 'Material' code found for "
                        f"{len(no_code)} nest(s), so their card titles carry no "
                        f"'(code)': {', '.join(no_code[:15])}"
                        + (" ..." if len(no_code) > 15 else ""))
    if unlabelled:
        warnings.append(f"No SAW CUT / TUBE LASER label could be decided for "
                        f"{len(unlabelled)} nest(s) - those cards were created "
                        f"without a machine label: {'; '.join(unlabelled[:15])}"
                        + (" ..." if len(unlabelled) > 15 else ""))

    payload = {
        "plan": template.get("plan", "SOPO D911 PIPELINE"),
        "bucket": template.get("bucket", "MODELING"),
        # Posted in REVERSE: Planner's "Create a task" top-inserts each new
        # card, so posting the schedule order straight through makes the
        # bucket read bottom-up (same fix as 922 Setup's _order_for_planner).
        "tasks": list(reversed(cards)),
    }

    log(f"\nWill create {len(cards)} card(s) in plan '{payload['plan']}', "
        f"bucket '{payload['bucket']}':")
    for card, names in zip(cards, card_labels):
        bits = ", ".join(names) if names else "no labels"
        due = card["dueDate"][:10] if card["dueDate"] else "no due date"
        prog = "" if "Program" in card["checklist"] else "  (no Program)"
        log(f"  - {card['title']}   [{bits}]  due {due}{prog}")

    if warnings:
        log("\nWarnings:")
        for w in warnings:
            log(f"  ! {w}")
        sdk.show_warning(params, "911 Setup - Teams card warnings",
                         "\n\n".join(warnings))
    _pct(0.8)
    if cancel_event.is_set():
        return False

    # --- Post (or dry-run) --------------------------------------------------
    url = (settings.get("card_webhook_url", "") or "").strip() or DEFAULT_CARD_WEBHOOK_URL
    dry_run = bool(settings.get("card_dry_run", False))
    if not url:
        log("\nNo Teams card webhook URL is configured - running as a DRY RUN.")
        log("Build the flow (docs/TEAMS_CARDS.md, flow #3) and paste its URL "
            "into Settings -> '911 Setup' -> 'Teams Webhook URL'.")
        dry_run = True
    elif dry_run:
        log("\nDry run enabled in Settings -> not posting.")

    if dry_run:
        sdk.write_payload_preview(payload, _CARD_PREVIEW_FILENAME, log)
        _pct(1.0)
        log("\nTeams cards: DONE (dry run - nothing was posted).")
        return True

    log("\nPosting cards to the webhook...")
    ok = sdk.post_webhook(url, payload, log)
    _pct(1.0)
    if ok:
        log(f"\nTeams cards: DONE. Requested {len(cards)} card(s) in "
            f"'{payload['bucket']}'.")
        log(f"Check the {payload['plan']} tab in the D922 channel to confirm.")
        return True
    log("\nTeams cards: FAILED - see the errors above. No cards were created.")
    return False


# ---------------------------------------------------------------------------
# Main run() function -- TechDeck plugin interface
# ---------------------------------------------------------------------------

def _dialog_groups() -> list:
    """The master window's plain-data spec (sdk.request_grouped_toggles).

    Grouped by ARTEFACT rather than by the numbered internal steps, because
    steps 5-7 all write into the same NEST sheet and are meaningless apart.
    "Difficulty label" is a child of PDF Stamping and defaults ON (C.D.,
    2026-07-31).

    "Generate Teams Cards" leads, and is the one stage that is NOT about the
    batch you are setting up: it cards every nest the EB 922 Schedule marks
    NEED TEAMS/SETUP, so it runs before the batch prompt (v1.8.0). Check it
    alone and the plugin never asks for a batch number.
    """
    return [
        {"key": "teams_cards",
         "label": "Generate Teams Cards (from the EB 922 Schedule)",
         "checked": True,
         "children": []},
        {"key": "folder_setup",
         "label": "Nest Folder Setup",
         "checked": True,
         "children": []},
        {"key": "nest_data",
         "label": "Nest Workbook Data (forecast, mil spec, batch list)",
         "checked": True,
         "children": []},
        {"key": "inspection_sheets",
         "label": "Inspection Sheets",
         "checked": True,
         "children": []},
        {"key": "pdf_stamping",
         "label": "PDF Stamping (Move Ticket Omit)",
         "checked": True,
         "children": [
             {"key": "difficulty", "label": "Difficulty label", "checked": True},
         ]},
    ]


def run(params: dict, progress_callback, cancel_event: threading.Event):
    """
    TechDeck plugin entry point.

    params keys:
      - 'settings': dict of plugin settings from plugin.json (optional overrides)
      - 'console':  TechDeck console (used for batch number prompt)
      - 'log':      callable injected by executor
      - 'batch_number': str (used when running without a console, e.g. CLI test)
    """
    log = params.get("log", print)
    settings = params.get("settings", {}) or {}

    # ------------------------------------------------------------------ #
    # Master toggle window (v1.7.0) -- same GroupedToggleDialog pattern as
    # 922 Setup, so both setups are driven by a checklist of what to run.
    #
    # The steps are far more interdependent than 922's stages: nothing below
    # "Nest Folder Setup" can run on a batch that has never been set up. The
    # stages are here so a RE-RUN can redo one part (most often just the PDF
    # stamping) without repeating the whole batch, and each stage warns rather
    # than crashing when its inputs are missing.
    # ------------------------------------------------------------------ #
    stages = sdk.request_grouped_toggles(
        params, _dialog_groups(),
        window_title="911 Setup",
        header="911 Setup - Select Actions",
        subtext=("Uncheck anything you do not want to run. Re-running a batch "
                 "with only PDF Stamping checked re-stamps the packets without "
                 "touching the nest workbooks."),
        run_button_text="Run 911 Setup")
    if stages is None:
        log("Cancelled - nothing was run.")
        return

    def _on(key):
        return bool((stages.get(key) or {}).get("enabled"))

    def _opt(key, child, default=True):
        g = stages.get(key) or {}
        return bool((g.get("options") or {}).get(child, default))

    do_cards = _on("teams_cards")
    do_folders = _on("folder_setup")
    do_nest_data = _on("nest_data")
    do_inspection = _on("inspection_sheets")
    do_stamping = _on("pdf_stamping")
    stamp_difficulty = do_stamping and _opt("pdf_stamping", "difficulty", True)

    enabled = [n for n, f in (("Generate Teams Cards", do_cards),
                              ("Nest Folder Setup", do_folders),
                              ("Nest Workbook Data", do_nest_data),
                              ("Inspection Sheets", do_inspection),
                              ("PDF Stamping", do_stamping)) if f]
    if not enabled:
        log("No actions selected - nothing to do.")
        return
    log("Actions: " + " -> ".join(enabled))
    if do_stamping:
        log(f"  Difficulty label: {'ON' if stamp_difficulty else 'OFF'}")

    # ------------------------------------------------------------------ #
    # Read configurable paths from settings.
    # Blank/missing values fall back to the original Path.home() defaults.
    # ------------------------------------------------------------------ #
    qtdr_override     = (settings.get("qtdr_base_path") or "").strip()
    forecast_override = (settings.get("forecast_dir") or "").strip()
    template_subdir   = (settings.get("template_subdir") or "").strip() or _DEFAULT_TEMPLATE_SUBDIR
    forecast_filename = (settings.get("forecast_filename") or "").strip() or _DEFAULT_FORECAST_FILENAME

    # ------------------------------------------------------------------ #
    # Stage 0 -- Generate Teams Cards (v1.8.0)
    #
    # Runs FIRST and independently of the batch: its work list is the EB 922
    # Schedule's NEED TEAMS/SETUP rows, which span whatever batches happen to
    # be queued. Checking only this stage is a complete run -- no batch
    # number is asked for, nothing on disk is touched.
    # ------------------------------------------------------------------ #
    batch_stages = do_folders or do_nest_data or do_inspection or do_stamping
    if do_cards:
        _run_teams_cards(params, progress_callback, cancel_event,
                         qtdr_override, lo=0, hi=8 if batch_stages else 100)
        if cancel_event.is_set():
            return
    if not batch_stages:
        log(f"\n{'='*50}")
        log("911 Setup complete (Teams cards only).")
        log(f"{'='*50}")
        return

    # ------------------------------------------------------------------ #
    # Step 1 -- Prompt user for batch number, then resolve batch folder
    # ------------------------------------------------------------------ #
    console = params.get("console")
    if console and hasattr(console, "request_input"):
        raw = sdk.request_batch_number(params, "Enter Batch Number (e.g. V060, V086):")
    else:
        raw = params.get("batch_number", "")

    batch_number = raw.strip().upper()
    if not batch_number:
        raise ValueError("No batch number provided. Please enter a batch number.")

    qtdr_root = _base_qtdr(qtdr_override)
    batch_folder = qtdr_root / batch_number

    log(f"Batch number : {batch_number}")
    log(f"QTDR root    : {qtdr_root}")
    log(f"Batch folder : {batch_folder}")

    if not qtdr_root.exists():
        raise FileNotFoundError(
            f"911 QTDR root not found: {qtdr_root}\n"
            f"Check the '911 QTDR Base Directory' setting for this plugin, or "
            f"verify your OneDrive sync."
        )

    if not batch_folder.exists():
        raise FileNotFoundError(
            f"Batch folder not found: {batch_folder}\n"
            f"Verify the batch number '{batch_number}' is correct and the folder exists."
        )

    progress_callback(5)
    if cancel_event.is_set():
        return

    # ------------------------------------------------------------------ #
    # Step 2 -- Find BATCH LIST and extract unique nest numbers
    # ------------------------------------------------------------------ #
    log("Locating BATCH LIST...")
    try:
        batch_list_path = _find_batch_list(batch_folder, batch_number)
    except FileNotFoundError as e:
        raise FileNotFoundError(str(e))

    log(f"BATCH LIST    : {batch_list_path.name}")

    try:
        all_nests = _get_unique_nests_from_batch_list(batch_list_path)
    except ValueError as e:
        raise ValueError(str(e))

    if not all_nests:
        raise ValueError(
            f"No nest numbers found in the 'Nest Pkg Nbr' column of {batch_list_path.name}. "
            "Check that the BATCH LIST has data starting from row 4."
        )

    log(f"Nests found   : {', '.join(all_nests)}")

    # ------------------------------------------------------------------ #
    # Step 2.5 -- Let the user choose which nests to run.
    #
    # Nests whose folder already exists in the batch are flagged as
    # "already set up" so the operator can tell a fresh run from a re-run.
    # Re-running a nest overwrites it. Selection happens up front: nothing
    # below this point runs until the user submits the dialog. Order is
    # preserved by filtering against all_nests (a set membership test).
    # ------------------------------------------------------------------ #
    existing_nests = {n for n in all_nests if (batch_folder / n).is_dir()}
    if existing_nests:
        log(f"Already set up: {', '.join(n for n in all_nests if n in existing_nests)}")

    if console is not None and hasattr(console, "request_nest_selection"):
        selection = console.request_nest_selection(batch_number, all_nests, existing_nests)
        if selection is None:
            log("Nest selection cancelled -- nothing was run.")
            cancel_event.set()  # user cancel: don't count as a successful (ticket-earning) run
            return
        chosen = set(selection)
        nest_numbers = [n for n in all_nests if n in chosen]
    else:
        # No console (CLI/test): honor an explicit 'nests' override, else all.
        override = params.get("nests")
        if override:
            chosen = {str(n).strip().upper() for n in override}
            nest_numbers = [n for n in all_nests if n.upper() in chosen]
        else:
            nest_numbers = list(all_nests)

    if not nest_numbers:
        log("No nests selected -- nothing to do.")
        cancel_event.set()  # nothing ran: don't count as a successful (ticket-earning) run
        return

    log(f"Running nests : {', '.join(nest_numbers)}")
    progress_callback(10)

    # ------------------------------------------------------------------ #
    # Step 2.7 -- Verify which BATCH LIST column truly holds DYPN QTY.
    #
    # The 'DYPN QTY' / 'Material Amount (Total)' headers arrive swapped on
    # some batches; the packet PDFs' SUMMARY OF NEST quantities are ground
    # truth. Score both columns against every available packet, use the
    # winner for the nest workbooks, and repair the headers in the file.
    # ------------------------------------------------------------------ #
    nest_packages_folder = batch_folder / "NEST PACKAGES"

    log("Verifying DYPN QTY column against the nest packet PDFs...")
    qty_col, swap_with, packet_qtys = _resolve_dypn_qty_col(
        batch_list_path, nest_packages_folder, all_nests, log, cancel_event)
    if swap_with is not None:
        _swap_batch_list_headers(batch_list_path, qty_col, swap_with, log)

    # ------------------------------------------------------------------ #
    # Step 3 -- Create nest subfolders
    # ------------------------------------------------------------------ #
    if do_folders:
        log("Creating nest folders...")
        for nest in nest_numbers:
            nest_dir = batch_folder / nest
            nest_dir.mkdir(exist_ok=True)
            log(f"  Folder: {nest_dir.name}")
    else:
        log("[skipped] Nest Folder Setup unchecked.")
        missing = [n for n in nest_numbers if not (batch_folder / n).is_dir()]
        if missing:
            log(f"  WARNING: {len(missing)} nest folder(s) do not exist and were "
                f"NOT created: {', '.join(missing[:10])}"
                + (" ..." if len(missing) > 10 else ""))

    progress_callback(15)
    if cancel_event.is_set():
        return

    # ------------------------------------------------------------------ #
    # Step 4 -- Copy & rename 911 BATCH template into each nest folder
    # ------------------------------------------------------------------ #
    # NOTE: copying the template OVERWRITES an existing nest workbook. That is
    # the intended behaviour for a first run, but it is why this is gated on
    # "Nest Folder Setup" — unchecking it is what makes a re-run (e.g. to
    # re-stamp the packets) safe, because the filled-in workbooks survive.
    nest_excel_paths = {}
    scribe_available = False
    if not do_folders:
        log("[skipped] 911 BATCH template copy (nest workbooks left untouched).")
        for nest in nest_numbers:
            dest_path = batch_folder / nest / f"911 BATCH {batch_number} {nest}.xlsx"
            nest_excel_paths[nest] = dest_path
        absent = [n for n, p in nest_excel_paths.items() if not p.exists()]
        if absent:
            log(f"  WARNING: no existing nest workbook for {len(absent)} nest(s): "
                f"{', '.join(absent[:10])}" + (" ..." if len(absent) > 10 else ""))
    else:
        log("Copying 911 BATCH template...")
        template_dir = _template_dir(qtdr_override, template_subdir)

        if not template_dir.exists():
            raise FileNotFoundError(
                f"Template directory not found: {template_dir}\n"
                f"Check the 'Template Subfolder' setting for this plugin."
            )

        template_path = _find_template_911(template_dir)
        log(f"Template      : {template_path.name}")

        # Scribe-verification doc copied into each nest folder (one copy per
        # nest). Same SACO template dir; a missing source is non-fatal.
        scribe_src = template_dir / _SCRIBE_DOC_FILENAME
        scribe_available = scribe_src.exists()
        if not scribe_available:
            log(f"  WARNING: Scribe doc not found ({scribe_src.name}) -- skipping for all nests.")

        for nest in nest_numbers:
            dest_name = f"911 BATCH {batch_number} {nest}.xlsx"
            dest_path = batch_folder / nest / dest_name
            shutil.copy2(template_path, dest_path)
            nest_excel_paths[nest] = dest_path
            log(f"  Copied -> {dest_name}")

            if scribe_available:
                scribe_dir = batch_folder / nest / _SCRIBE_SUBFOLDER
                scribe_dest = scribe_dir / _SCRIBE_DOC_FILENAME
                # Earlier versions dropped the doc loose in the nest root;
                # relocate such a copy instead of duplicating it.
                legacy_dest = batch_folder / nest / _SCRIBE_DOC_FILENAME
                if scribe_dest.exists():
                    log(f"  Scribe doc already in {nest} -- skipped")
                else:
                    try:
                        scribe_dir.mkdir(exist_ok=True)
                        if legacy_dest.exists():
                            legacy_dest.replace(scribe_dest)
                            log(f"  Scribe doc moved -> {nest}\\{_SCRIBE_SUBFOLDER}")
                        else:
                            shutil.copy2(scribe_src, scribe_dest)
                            log(f"  Scribe doc -> {nest}\\{_SCRIBE_SUBFOLDER}")
                    except Exception as e:
                        log(f"  WARNING: Could not copy scribe doc into {nest}: {e}")

    progress_callback(20)
    if cancel_event.is_set():
        return

    # ------------------------------------------------------------------ #
    # Step 5 -- Copy Working Forecast List into batch folder, load it
    # ------------------------------------------------------------------ #
    log("Copying Working Forecast List...")
    forecast_src = _forecast_dir(forecast_override) / forecast_filename
    if not forecast_src.exists():
        raise FileNotFoundError(
            f"Working Forecast List not found at:\n{forecast_src}\n"
            "Check the 'Forecast and Inventory Reports Directory' and "
            "'Working Forecast Filename' settings, and that the folder is synced."
        )

    forecast_copy = batch_folder / forecast_filename
    sdk.copy_resilient(forecast_src, forecast_copy, log)  # hydrate + locked-open message (Hard Rule 13)
    log(f"Forecast copy : {forecast_copy}")

    log("Loading 911 Forecast sheet...")
    forecast_wb = load_workbook(forecast_copy, data_only=True)
    # Locate the forecast columns once, by header name (Hard Rules 1 & 2).
    forecast_ws, forecast_hdr, forecast_nest_col, forecast_out_cols = \
        _locate_forecast(forecast_wb, log)

    progress_callback(25)
    if cancel_event.is_set():
        forecast_wb.close()
        return

    progress_callback(30)

    # ------------------------------------------------------------------ #
    # Per-nest processing loop
    # ------------------------------------------------------------------ #
    qty_mismatches = []   # (nest, dypn, wo, batch qty, packet qty)
    qty_unverified = []   # (nest, dypn, wo, batch qty) -- not in packet summary
    missing_forecast = []  # nests with no forecast row (cols A-C left blank)
    total_nests = len(nest_numbers)

    # -- Difficulty ratings: read the schedule ONCE for the whole batch -----
    # Owned by the sibling 911_remove_ticket (the single home of the stamp
    # helpers), so the lookup and the colour legend live in one place.
    diff_map, diff_problem, unrated_nests = {}, "", []
    _stamps = None
    if do_stamping and stamp_difficulty:
        _stamps = _load_omit_stamp_helpers(log)
        if _stamps is not None and hasattr(_stamps, "_load_difficulty_map"):
            diff_map, diff_problem = _stamps._load_difficulty_map(params, log)
        else:
            diff_problem = ("The 911 Remove Ticket helpers could not be loaded, "
                            "so no difficulty labels were stamped.")
        if diff_problem:
            log(f"  WARNING: {diff_problem.splitlines()[0]}")
    for nest_idx, nest in enumerate(nest_numbers):
        if cancel_event.is_set():
            break

        log(f"\n{'='*50}")
        log(f"Processing nest {nest_idx+1}/{total_nests}: {nest}")
        log(f"{'='*50}")

        nest_excel = nest_excel_paths[nest]
        wb = sdk.load_workbook_resilient(nest_excel, log=log)
        nest_ws = wb["NEST"]

        # Steps 5-7 all write into the same NEST sheet, so they are one
        # toggle. matl_type still feeds Step 9's material fill, so it is read
        # from the packet even when the workbook write is skipped.
        mil_spec, matl_type, batch_rows = None, None, []

        if not do_nest_data:
            log("  [skipped] Nest Workbook Data unchecked.")
            if do_stamping:
                _, matl_type = _get_pdf_data_for_nest(nest_packages_folder, nest, log)
        else:
            # -- Step 5: Forecast data -> NEST cols A-C, starting row 4 ------
            log(f"  [Step 5] Extracting forecast rows for {nest}...")
            forecast_rows = _copy_forecast_rows(
                forecast_ws, forecast_hdr, forecast_nest_col, forecast_out_cols, nest)
            if not forecast_rows:
                missing_forecast.append(nest)
                log(f"  WARNING: No forecast rows found for nest {nest} in the "
                    f"'{forecast_ws.title}' sheet (column "
                    f"{get_column_letter(forecast_nest_col)}).")
            else:
                log(f"  Found {len(forecast_rows)} forecast rows.")
                _paste_forecast_into_nest(nest_ws, forecast_rows)

            # -- Step 6: PDF -> MIL-S spec (D4) + MATL (E4) ------------------
            log(f"  [Step 6] Reading MIL SPEC / MATERIAL from nest packet...")
            mil_spec, matl_type = _get_pdf_data_for_nest(nest_packages_folder, nest, log)

            if mil_spec:
                nest_ws.cell(4, 4).value = mil_spec   # D4
                log(f"  MIL Spec -> D4: {mil_spec}")
            else:
                log(f"  WARNING: MIL spec not found in nest packet.")

            if matl_type:
                nest_ws.cell(4, 5).value = matl_type  # E4
                log(f"  Material -> E4: {matl_type}")
            else:
                log(f"  WARNING: MATERIAL not found in nest packet (left blank).")

            # -- Step 7: BATCH LIST -> NEST cols F-K, starting row 4 ---------
            log(f"  [Step 7] Extracting batch rows for {nest}...")
            try:
                batch_rows = _get_batch_rows_for_nest(batch_list_path, nest,
                                                      qty_col_override=qty_col)
            except ValueError as e:
                log(f"  WARNING: {e} -- skipping batch data for {nest}")
                batch_rows = []

            if not batch_rows:
                log(f"  WARNING: No batch rows found for nest {nest}.")
            else:
                log(f"  Found {len(batch_rows)} batch rows.")
                _paste_batch_rows_into_nest(nest_ws, batch_rows)

                # Residual QTY check: even the verified column can disagree
                # with the packet on individual rows -- highlight those cells
                # and collect them for the end-of-run summary.
                _flag_qty_mismatches(nest_ws, batch_rows,
                                     packet_qtys.get(nest) or {}, nest,
                                     nest_excel.name, log,
                                     qty_mismatches, qty_unverified)

            # -- Fill A-E down for every part row (SCRIBE mirrors NEST) -------
            num_parts = len(batch_rows)
            if num_parts > 1:
                _fill_nest_part_rows(nest_ws, num_parts)
                log(f"  Filled MIL spec / material / forecast down {num_parts} part rows.")

        # -- Step 8a: Collect part rows (WO / DYPN / qty) from NEST ------
        log(f"  [Step 8] Reading DYPN values from NEST col G...")
        part_rows = _get_dypn_rows(nest_ws)
        if not part_rows:
            log(f"  WARNING: No DYPN values found in NEST col G.")
        else:
            log(f"  Found {len(part_rows)} DYPN rows.")

        # -- Save and close openpyxl workbook before Excel COM opens it --
        log(f"  Saving {nest_excel.name}...")
        wb.save(nest_excel)
        wb.close()

        # -- Step 8b: Copy inspection sheets via Excel COM ---------------
        # openpyxl's copy_worksheet drops conditional formatting rules
        # on copy (verified: 350 CF rules on template -> 0 on copy).
        # Excel's native Sheet.Copy preserves them.
        if part_rows and do_inspection:
            log(f"  Building {len(part_rows)} inspection sheet(s) via Excel...")
            try:
                _build_inspection_sheets_via_excel(nest_excel, part_rows, log)
            except Exception as e:
                log(f"  ERROR: Excel COM inspection sheet build failed: {e}")
                raise
        elif part_rows:
            log("  [skipped] Inspection Sheets unchecked.")

        log(f"  Done: {nest_excel.name}")

        # -- Step 9: Build MOVE TICKET OMIT PDF (remove MOVE TICKET pages, keep MIL-SPEC/HULL) --
        if do_stamping:
            log(f"  [Step 9] Extracting drawings for {nest}...")
            drawings_dest = batch_folder / nest / f"{nest} MOVE TICKET OMIT.pdf"
            difficulty = None
            if stamp_difficulty and not diff_problem and _stamps is not None:
                difficulty = _stamps._lookup_difficulty(diff_map, nest)
                if difficulty:
                    log(f"  Difficulty: {difficulty}")
                else:
                    unrated_nests.append(str(nest))
            _extract_nest_drawings(nest_packages_folder, nest, drawings_dest, log,
                                   batch=batch_number, material_hint=matl_type or "",
                                   difficulty=difficulty)
        else:
            log("  [skipped] PDF Stamping unchecked.")

        pct = 30 + int(65 * (nest_idx + 1) / total_nests)
        progress_callback(pct)

    # ------------------------------------------------------------------ #
    # Cleanup
    # ------------------------------------------------------------------ #
    forecast_wb.close()

    # ------------------------------------------------------------------ #
    # Difficulty disclaimer -- nothing is printed on a packet whose nest has
    # no rating, so the omission is surfaced in a popup the user must dismiss
    # rather than a log line that scrolls past (C.D. 2026-07-31).
    # ------------------------------------------------------------------ #
    if do_stamping and stamp_difficulty and (diff_problem or unrated_nests):
        if diff_problem:
            body = diff_problem
        else:
            shown = "\n".join(f"  - {n}" for n in unrated_nests[:25])
            body = (f"{len(unrated_nests)} of {total_nests} nest(s) were stamped "
                    f"WITHOUT a difficulty label because the nest has no rating "
                    f"colour on the EB 922 Schedule (CURRENT PIPELINE, column E), "
                    f"is marked N/A, or is not listed:\n\n{shown}"
                    + ("\n  ..." if len(unrated_nests) > 25 else "")
                    + "\n\nEverything else on those packets stamped normally.")
        sdk.show_warning(params, "Difficulty label - not stamped", body)

    # ------------------------------------------------------------------ #
    # Forecast-coverage summary -- a missing forecast row leaves NEST cols
    # A-C blank, which previously looked like a clean run. Surface it loudly
    # and actionably so it can't be mistaken for success.
    # ------------------------------------------------------------------ #
    if missing_forecast:
        log(f"\n{'!'*50}")
        log(f"FORECAST DATA MISSING -- NEST columns A-C left BLANK for "
            f"{len(missing_forecast)} of {total_nests} nest(s):")
        log(f"  {', '.join(missing_forecast)}")
        log(f"These nests were not found in the Working Forecast List's "
            f"'{forecast_ws.title}' sheet. Most often the forecast is out of "
            "date or not fully synced on this machine.")
        log("FIX: in OneDrive, right-click the 'Forecast and Inventory Reports' "
            "folder -> 'Always keep on this device', let it finish syncing, then "
            "re-run 911 Setup. (If these nests have not been added to the "
            "forecast yet, add them there first.)")
        log(f"{'!'*50}")

    # ------------------------------------------------------------------ #
    # QTY verification summary -- repeated at the end so warnings that
    # scrolled by during the run cannot be missed.
    # ------------------------------------------------------------------ #
    if qty_mismatches or qty_unverified:
        log(f"\n{'!'*50}")
        log("QTY VERIFICATION -- MANUAL REVIEW NEEDED:")
        for nest, dypn, wo, qty, pdf_qty in qty_mismatches:
            log(f"  MISMATCH    {nest}  {dypn} / {wo}: workbook has {qty}, "
                f"nest packet says {pdf_qty} (cell highlighted yellow)")
        for nest, dypn, wo, qty in qty_unverified:
            log(f"  UNVERIFIED  {nest}  {dypn} / {wo}: qty {qty} not found "
                f"in the packet summary")
        log("Cross-check these against page 2 of the nest packet PDF "
            "before releasing paperwork.")
        log(f"{'!'*50}")
    elif packet_qtys:
        log("\nAll part quantities verified against the nest packet PDFs.")

    progress_callback(100)
    log(f"\n{'='*50}")
    log(f"911 Setup complete for batch {batch_number}.")
    log(f"Processed {total_nests} nest(s): {', '.join(nest_numbers)}")
    log(f"{'='*50}")


# ---------------------------------------------------------------------------
# CLI entry point for local testing
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    batch = input("Enter batch number: ").strip()
    if not batch:
        print("No batch number entered.")
        sys.exit(1)

    ev = threading.Event()

    def _prog(v):
        print(f"  [Progress] {v}%")

    run({"batch_number": batch, "log": print, "settings": {}}, _prog, ev)
