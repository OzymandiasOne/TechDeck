"""
911 Bean Counter Plugin for TechDeck
====================================

Consolidates the per-part "NC style baked beans" calc workbooks (the
DSTV/linear-inch pricing sheets Antonio & Scott J's template fills out from
Vacam NC files) into ONE review list, so nobody has to open 60+ workbooks
one at a time.

  1. Prompts for the folder holding the filled calc sheets via the native
     folder dialog — the same folder that holds the NC files; the output is
     saved right back into it (per the requester: "output to the same folder
     path that has all the nc files in it").
  2. Prompts for the batch number (family-shared) + nest number — used only
     to name the output file.
  3. For every .xlsm/.xlsx in the folder, reads off the summary (REF) sheet:
       - the DYPN (a part-number-shaped cell near the top-left, e.g.
         "  R4533287-109" at C6; falls back to the file name and the two are
         cross-checked — a mismatch is flagged in the Notes column), and
       - the three YELLOW totals: TOTAL Bevels / TOTAL Complex Bevels /
         TOTAL Cut Lin. Labels are FOUND BY TEXT anywhere on the sheet, never
         by fixed cell address (Hard Rules 1-2) — the template's label wording
         already drifted between versions, so each value is the first numeric
         cell to the right of its matched label.
  4. Writes "{batch} {nest} Consolidated NC Calcs.xlsx" into that folder:
     one row per part (DYPN | Total Bevels | Total Complex Bevels |
     Total Cut Lin | Notes, column order per the requester), sorted naturally
     by DYPN, with a live =SUM() totals row at the bottom.

Values come from openpyxl's CACHED formula results (data_only=True). A sheet
that was never recalculated/saved in Excel has no cache — those rows are
flagged instead of silently written as zero.

Future ask (not built yet): also RUN the template against raw NC files and
save the per-part workbooks. Blocked on the template's manual-review step
(unprogrammed bevels/features are still keyed in by hand).
"""

from __future__ import annotations

import datetime
import re
from pathlib import Path

# SDK bootstrap (works in-process and for standalone CLI testing).
try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VERSION = "1.0.0"

# The three yellow totals, keyed in the requester's column order. Anchored so
# "TOTAL Bevels" can never match "TOTAL Complex Bevels".
_TOTAL_PATTERNS = [
    ("Total Bevels (in)", re.compile(r"^TOTAL\s+BEVELS?\b", re.IGNORECASE)),
    ("Total Complex Bevels (in)", re.compile(r"^TOTAL\s+COMPLEX\s+BEVELS?\b", re.IGNORECASE)),
    ("Total Cut Lin (in)", re.compile(r"^TOTAL\s+CUT\s+LIN\b", re.IGNORECASE)),
]

# DYPN-shaped cell, e.g. "R4533287-109" (leading/trailing spaces tolerated).
_DYPN_RE = re.compile(r"^[A-Z]{0,3}\d{4,}-\d{1,4}$", re.IGNORECASE)

# How far into a sheet we look for the summary labels / the DYPN. The summary
# block sits in the top rows; generous bounds keep a moved block findable
# without scanning 300+ contour rows on every sheet.
_SCAN_ROWS = 40
_SCAN_COLS = 30
_DYPN_SCAN_ROWS = 15

_SKIP_NAME_HINT = "consolidated nc calcs"  # never re-ingest our own output


def _find_summary_sheet(wb):
    """The sheet holding the totals block — 'REF' in the current template, but
    located by content (the TOTAL Cut Lin label) so a rename survives."""
    ordered = sorted(wb.worksheets, key=lambda ws: ws.title.upper() != "REF")
    cut_lin_re = _TOTAL_PATTERNS[2][1]
    for ws in ordered:
        for row in ws.iter_rows(min_row=1, max_row=_SCAN_ROWS, max_col=_SCAN_COLS):
            for cell in row:
                if isinstance(cell.value, str) and cut_lin_re.match(cell.value.strip()):
                    return ws
    return None


def _value_right_of(ws, row_idx: int, col_idx: int):
    """First numeric cell to the right of a label cell (labels and their
    values are separated by merge padding that varies between versions)."""
    for cells in ws.iter_rows(min_row=row_idx, max_row=row_idx,
                              min_col=col_idx + 1, max_col=col_idx + 8):
        for cell in cells:
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return float(v)
    return None


def _extract_part(ws, file_stem: str):
    """Pull (dypn, {column: value}, notes) off a summary sheet."""
    notes = []
    totals = {name: None for name, _ in _TOTAL_PATTERNS}

    dypn = None
    for row in ws.iter_rows(min_row=1, max_row=_DYPN_SCAN_ROWS, max_col=_SCAN_COLS):
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if dypn is None and _DYPN_RE.match(s):
                dypn = s.upper()
            for name, pat in _TOTAL_PATTERNS:
                if totals[name] is None and pat.match(s):
                    totals[name] = _value_right_of(ws, cell.row, cell.column)

    # Labels can sit below the DYPN scan window; finish the label scan deeper.
    if any(v is None for v in totals.values()):
        for row in ws.iter_rows(min_row=_DYPN_SCAN_ROWS + 1, max_row=_SCAN_ROWS,
                                max_col=_SCAN_COLS):
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                s = v.strip()
                for name, pat in _TOTAL_PATTERNS:
                    if totals[name] is None and pat.match(s):
                        totals[name] = _value_right_of(ws, cell.row, cell.column)

    if dypn is None:
        dypn = file_stem.upper()
        notes.append("DYPN cell not found - used file name")
    elif dypn != file_stem.upper().strip():
        notes.append(f"DYPN differs from file name ({file_stem})")

    missing = [name for name, v in totals.items() if v is None]
    if len(missing) == len(totals):
        notes.append("no cached totals - open the file in Excel, let it "
                     "calculate, and save")
    elif missing:
        notes.append("missing: " + ", ".join(m.replace(" (in)", "") for m in missing))

    return dypn, totals, notes


def _dypn_sort_key(dypn: str):
    """Natural sort: R4533287-110 after R4533287-109, not lexicographically."""
    m = re.match(r"^(.*?)(\d+)$", dypn)
    if m:
        return (m.group(1), int(m.group(2)))
    return (dypn, -1)


def _write_output(rows: list, out_path: Path, log) -> Path:
    """rows = [(dypn, {column: value}, notes)] already sorted."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    yellow = PatternFill("solid", start_color="FFFF00")
    bold = Font(bold=True)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["DYPN"] + [name for name, _ in _TOTAL_PATTERNS] + ["Notes"]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=text)
        c.font = bold
        c.border = box
        c.alignment = Alignment(horizontal="center")
        if 2 <= col <= 4:            # the three value columns echo the source yellow
            c.fill = yellow

    for r, (dypn, totals, notes) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=dypn).border = box
        for col, (name, _) in enumerate(_TOTAL_PATTERNS, start=2):
            c = ws.cell(row=r, column=col, value=totals[name])
            c.border = box
            c.number_format = "0.000"
        ws.cell(row=r, column=5, value="; ".join(notes) if notes else None).border = box

    # Live grand-total row so the sums stay honest if a value is corrected.
    total_row = len(rows) + 2
    label = ws.cell(row=total_row, column=1, value=f"TOTALS ({len(rows)} parts)")
    label.font = bold
    label.border = box
    for col in range(2, 5):
        letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUM({letter}2:{letter}{total_row - 1})")
        c.font = bold
        c.border = box
        c.number_format = "0.000"
        c.fill = yellow
    ws.cell(row=total_row, column=5).border = box

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{total_row - 1}"
    widths = {"A": 18, "B": 17, "C": 24, "D": 17, "E": 46}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        # Most likely the previous list is open in Excel — save alongside it.
        stamp = datetime.datetime.now().strftime("%H%M%S")
        alt = out_path.with_name(f"{out_path.stem} ({stamp}){out_path.suffix}")
        log(f"[WARN] {out_path.name} is locked (open in Excel?) - saving as "
            f"{alt.name} instead.")
        wb.save(alt)
        return alt


def run(params: dict, progress_callback, cancel_event):
    log = params.get("log", print)
    log(f"911 Bean Counter v{VERSION}")

    folder_str = sdk.request_directory(
        params, "Select the folder holding the filled NC calc spreadsheets")
    if not folder_str:
        log("No folder selected - nothing to do.")
        return
    folder = Path(folder_str)

    batch = (sdk.request_batch_number(params, "Enter batch number:") or "").strip()
    nest = (sdk.request_text(params, "Enter nest number:") or "").strip()
    if not batch or not nest:
        log("Batch and nest numbers are required (they name the output file).")
        return
    safe = re.compile(r'[<>:"/\\|?*]')
    out_name = f"{safe.sub('_', batch)} {safe.sub('_', nest)} Consolidated NC Calcs.xlsx"
    out_path = folder / out_name
    progress_callback(5)

    candidates = sorted(
        p for pattern in ("*.xlsm", "*.xlsx") for p in folder.glob(pattern)
        if not p.name.startswith("~$")
        and _SKIP_NAME_HINT not in p.name.lower()
    )
    if not candidates:
        log(f"No .xlsm/.xlsx calc sheets found in {folder}")
        return
    log(f"Found {len(candidates)} workbook(s) in {folder.name} - consolidating...")

    rows, skipped = [], []
    for i, path in enumerate(candidates):
        if cancel_event is not None and cancel_event.is_set():
            log("Cancelled.")
            return
        wb = None
        try:
            wb = sdk.load_workbook_resilient(
                path, log=log, data_only=True, read_only=True)
            ws = _find_summary_sheet(wb)
            if ws is None:
                skipped.append(f"{path.name}: no totals block found (not a calc sheet?)")
                continue
            dypn, totals, notes = _extract_part(ws, path.stem)
            rows.append((dypn, totals, notes))
            if notes:
                log(f"  [FLAG] {path.name}: {'; '.join(notes)}")
        except Exception as exc:
            skipped.append(f"{path.name}: {exc}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
        progress_callback(5 + int(85 * (i + 1) / len(candidates)))

    if not rows:
        log("No calc sheets could be read - nothing to consolidate.")
        for s in skipped:
            log(f"  [SKIPPED] {s}")
        return

    rows.sort(key=lambda r: _dypn_sort_key(r[0]))
    saved = _write_output(rows, out_path, log)
    progress_callback(100)

    flagged = sum(1 for _, _, notes in rows if notes)
    log("")
    log(f"Consolidated {len(rows)} part(s) -> {saved}")
    if flagged:
        log(f"[REVIEW] {flagged} row(s) flagged - see the Notes column.")
    for s in skipped:
        log(f"  [SKIPPED] {s}")
