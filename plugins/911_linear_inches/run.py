"""
911 Linear Inches plugin for TechDeck.

Closed-loop cut-length report for a 911 QTDR batch, plus an approximate nest-layout
viewer.

The user points the plugin at a batch's ``IGES FILES`` folder (or the batch folder). Each
IGES subfolder is named by the nest package number (``N<nest>S`` -> nest ``<nest>``) and
holds that nest's part DXFs. Cut length per part is computed EXACTLY from the DXF vector
geometry; the per-part quantity is read from the nest package PDF's "SUMMARY OF NEST"
table (WORK ORDER -> qty), matched to the DXF by the work order embedded in its filename.

Part 1 - Excel report: per-part rows, per-nest subtotals, a batch total.
Part 2 - Layout viewer: a window that simulates each nest (bottom-left-fill rectangle
         packing using DXF bounding boxes) and draws it to scale with measurements, next
         to a per-nest measurements table. This is an APPROXIMATION of the SigmaNest
         layout for deriving data (utilization, fit), not an exact reproduction.

This is a GUI plugin (``requires_main_thread: true``): run() executes on the Qt main
thread, so it uses QFileDialog/QMessageBox directly (not the worker-thread SDK console
helpers).

The DXF parser is vendored from the ``customer_dxf_quoting`` plugin (stdlib only, ASCII
DXF, LINE/ARC/CIRCLE/LWPOLYLINE/POLYLINE + bulge). It should later be centralized into
``plugin_sdk`` so both plugins share one implementation.
"""

import os
import re
import glob
import math
from pathlib import Path

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QComboBox,
    QGraphicsView, QGraphicsScene, QGraphicsRectItem, QGraphicsSimpleTextItem,
    QGraphicsItem, QSplitter, QFileDialog, QMessageBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QGroupBox,
)
from PySide6.QtCore import Qt, QRectF, QRect, QTimer
from PySide6.QtGui import QPen, QBrush, QColor, QPainter

from techdeck.core.plugin_window import PluginWindow

_window = None  # module-level - prevents the window from being garbage collected

# Work order token embedded in DXF filenames + the summary table, e.g. "3X24-398".
WO_RE = re.compile(r"\d[A-Z]\d+-\d+")
# Layers that are reference-only and never count as cut (matches customer_dxf_quoting).
EXCLUDE_LAYERS = {"BOUNDING_BOX", "IGNORE"}
# Entity types we do NOT measure; their presence means the length may be under-counted.
UNSUPPORTED_FLAGGED = {"SPLINE", "ELLIPSE"}
PART_GAP_IN = 0.5  # spacing left between packed parts (approx kerf/handling gap)


# ---------------------------------------------------------------------------
# DXF parsing (vendored from customer_dxf_quoting; TODO: centralize in plugin_sdk)
# ---------------------------------------------------------------------------

def _decode_dxf(path):
    data = Path(path).read_bytes()
    if data.startswith(b"AutoCAD Binary DXF"):
        raise ValueError("Binary DXF is not supported - re-export as ASCII DXF.")
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not decode file - is this a DXF?")


def _read_pairs(path):
    lines = _decode_dxf(path).splitlines()
    pairs = []
    it = iter(lines)
    for code_line in it:
        try:
            value = next(it)
        except StopIteration:
            break
        try:
            code = int(code_line.strip())
        except ValueError:
            raise ValueError(f"Malformed DXF (expected numeric group code, got {code_line.strip()!r}).")
        pairs.append((code, value.strip()))
    if not pairs:
        raise ValueError("File is empty.")
    return pairs


def _bulge_length(p1, p2, bulge):
    chord = math.hypot(p2[0] - p1[0], p2[1] - p1[1])
    if abs(bulge) < 1e-12 or chord < 1e-12:
        return chord
    theta = 4.0 * math.atan(bulge)
    radius = chord * (1.0 + bulge * bulge) / (4.0 * abs(bulge))
    return radius * abs(theta)


def _collect_until_next_entity(pairs, j):
    groups = []
    n = len(pairs)
    while j < n and pairs[j][0] != 0:
        groups.append(pairs[j])
        j += 1
    return groups, j


def _build_line(groups):
    g = {}
    for code, val in groups:
        if code in (8, 10, 20, 11, 21):
            g[code] = val
    try:
        p1 = (float(g[10]), float(g[20]))
        p2 = (float(g[11]), float(g[21]))
    except KeyError:
        return None
    return {"layer": g.get(8, "0"), "points": [p1, p2], "length": math.dist(p1, p2)}


def _build_arc(groups, full_circle=False):
    g = {}
    for code, val in groups:
        if code in (8, 10, 20, 40, 50, 51):
            g[code] = val
    try:
        cx, cy, r = float(g[10]), float(g[20]), float(g[40])
    except KeyError:
        return None
    if full_circle:
        a0, sweep = 0.0, math.tau
    else:
        a0 = math.radians(float(g.get(50, 0.0)))
        a1 = math.radians(float(g.get(51, 360.0)))
        sweep = (a1 - a0) % math.tau or math.tau
    n = max(8, int(abs(sweep) / math.tau * 64) + 1)
    pts = [(cx + r * math.cos(a0 + sweep * t / n), cy + r * math.sin(a0 + sweep * t / n))
           for t in range(n + 1)]
    return {"layer": g.get(8, "0"), "points": pts, "length": r * sweep}


def _polyline_entity(layer, verts, closed):
    verts = [v for v in verts if v[1] is not None]
    if len(verts) < 2:
        return None
    total = 0.0
    seg_pairs = list(zip(verts, verts[1:]))
    if closed:
        seg_pairs.append((verts[-1], verts[0]))
    for v1, v2 in seg_pairs:
        total += _bulge_length((v1[0], v1[1]), (v2[0], v2[1]), v1[2])
    points = [(v[0], v[1]) for v in verts]
    return {"layer": layer, "points": points, "length": total}


def _build_lwpolyline(groups):
    layer, flags = "0", 0
    verts = []  # [x, y, bulge]
    for code, val in groups:
        if code == 8:
            layer = val
        elif code == 70:
            flags = int(val)
        elif code == 10:
            verts.append([float(val), None, 0.0])
        elif code == 20 and verts:
            verts[-1][1] = float(val)
        elif code == 42 and verts:
            verts[-1][2] = float(val)
    return _polyline_entity(layer, verts, closed=bool(flags & 1))


def _build_polyline(pairs, j):
    groups, j = _collect_until_next_entity(pairs, j)
    layer, flags = "0", 0
    for code, val in groups:
        if code == 8:
            layer = val
        elif code == 70:
            flags = int(val)
    verts = []
    n = len(pairs)
    while j < n:
        code, val = pairs[j]
        if code == 0 and val == "VERTEX":
            vgroups, j = _collect_until_next_entity(pairs, j + 1)
            v = [None, None, 0.0]
            vflags = 0
            for c, vv in vgroups:
                if c == 10:
                    v[0] = float(vv)
                elif c == 20:
                    v[1] = float(vv)
                elif c == 42:
                    v[2] = float(vv)
                elif c == 70:
                    vflags = int(vv)
            if v[0] is not None and v[1] is not None and not (vflags & 16):
                verts.append(v)
        elif code == 0 and val == "SEQEND":
            _, j = _collect_until_next_entity(pairs, j + 1)
            break
        else:
            j += 1
    return _polyline_entity(layer, verts, closed=bool(flags & 1)), j


def parse_dxf(path):
    """Return (entities, skipped, insunits). entity = {layer, points, length}."""
    pairs = _read_pairs(path)
    n = len(pairs)
    entities, skipped = [], {}
    insunits = None
    section = None
    i = 0
    while i < n:
        code, val = pairs[i]
        if code == 0 and val == "SECTION" and i + 1 < n and pairs[i + 1][0] == 2:
            section = pairs[i + 1][1]
            i += 2
            continue
        if code == 0 and val == "ENDSEC":
            section = None
            i += 1
            continue
        if section == "HEADER" and code == 9 and val == "$INSUNITS":
            if i + 1 < n and pairs[i + 1][0] == 70:
                insunits = int(pairs[i + 1][1])
            i += 1
            continue
        if section == "ENTITIES" and code == 0:
            if val == "LINE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_line(groups)
            elif val == "ARC":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_arc(groups)
            elif val == "CIRCLE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_arc(groups, full_circle=True)
            elif val == "LWPOLYLINE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_lwpolyline(groups)
            elif val == "POLYLINE":
                ent, i = _build_polyline(pairs, i + 1)
            else:
                skipped[val] = skipped.get(val, 0) + 1
                i += 1
                continue
            if ent is not None:
                entities.append(ent)
            continue
        i += 1
    return entities, skipped, insunits


def dxf_metrics(path):
    """Cut length (in), bounding box (in), and anomaly info for one part DXF."""
    entities, skipped, insunits = parse_dxf(path)
    scale, unit_note = 1.0, None
    if insunits == 4:            # millimetres
        scale, unit_note = 1.0 / 25.4, "units=mm (converted to in)"
    elif insunits not in (1, 0, None):
        unit_note = f"unexpected $INSUNITS={insunits}"
    total = 0.0
    xs, ys, layers = [], [], set()
    for e in entities:
        layers.add(e["layer"])
        for (x, y) in e["points"]:
            xs.append(x)
            ys.append(y)
        if e["layer"].strip().upper() in EXCLUDE_LAYERS:
            continue
        total += e["length"]
    total *= scale
    bbox = ((max(xs) - min(xs)) * scale, (max(ys) - min(ys)) * scale) if xs else (0.0, 0.0)
    return {
        "linear_inches": total,
        "bbox": bbox,
        "skipped": skipped,
        "layers": layers,
        "unit_note": unit_note,
        "n_entities": len(entities),
    }


# ---------------------------------------------------------------------------
# Folder resolution + nest package lookup
# ---------------------------------------------------------------------------

def _find_child_dir(parent, name):
    try:
        for d in os.listdir(parent):
            if d.strip().upper() == name.upper() and os.path.isdir(os.path.join(parent, d)):
                return os.path.join(parent, d)
    except OSError:
        pass
    return None


def resolve_roots(selected):
    """Return (iges_root, batch_dir) from whatever the user picked."""
    selected = os.path.abspath(selected)
    if os.path.basename(selected).strip().upper() == "IGES FILES":
        return selected, os.path.dirname(selected)
    child = _find_child_dir(selected, "IGES FILES")
    if child:
        return child, selected
    return selected, os.path.dirname(selected)


def derive_nest(subfolder_name):
    """N5CDARQS -> 5CDARQ; anything else returned unchanged."""
    s = subfolder_name
    if len(s) > 2 and s[0].upper() == "N" and s[-1].upper() == "S":
        return s[1:-1]
    return s


def _plate_from_text(text):
    """Pull (PLATE LENGTH, PLATE WIDTH) in inches from a nest-drawing page's text."""
    toks = [t.strip() for t in text.splitlines() if t.strip()]
    length = width = None

    def _num_after(label):
        for i, t in enumerate(toks):
            if t.upper() == label:
                for u in toks[i + 1:i + 3]:
                    try:
                        return float(u)
                    except ValueError:
                        continue
        return None
    length = _num_after("PLATE LENGTH")
    width = _num_after("PLATE WIDTH")
    if length and width:
        return (length, width)
    return None


def read_nest_package(nest, nest_pkg_dir, batch_dir, log):
    """Return (summary {wo: (part, qty)}, plate (L, W) or None)."""
    pdf = None
    if nest_pkg_dir:
        cands = glob.glob(os.path.join(nest_pkg_dir, f"{nest}.pdf")) + \
            glob.glob(os.path.join(nest_pkg_dir, f"{nest} *.pdf"))
        if cands:
            pdf = cands[0]
    if pdf is None:  # fall back to the nest folder's MOVE TICKET OMIT pdf
        nest_folder = _find_child_dir(batch_dir, nest)
        if nest_folder:
            cands = glob.glob(os.path.join(nest_folder, "*MOVE TICKET OMIT*.pdf"))
            if cands:
                pdf = cands[0]
    if pdf is None:
        return {}, None
    try:
        sdk.ensure_local(pdf, log=log)
        doc = fitz.open(pdf)
    except Exception as exc:
        log(f"  ! could not open package PDF for {nest}: {exc}")
        return {}, None
    summary, plate = {}, None
    try:
        for page in doc:
            text = page.get_text()
            if not summary and "SUMMARY OF NEST" in text:
                toks = [t.strip() for t in text.splitlines() if t.strip()]
                for i, tk in enumerate(toks):
                    if WO_RE.fullmatch(tk) and i >= 2:
                        try:
                            qty = int(toks[i - 1])
                        except ValueError:
                            continue
                        summary[tk] = (toks[i - 2], qty)
            if plate is None and "PLATE LENGTH" in text:
                plate = _plate_from_text(text)
    finally:
        doc.close()
    return summary, plate


# ---------------------------------------------------------------------------
# Part 2 - approximate nest layout (bottom-left-fill rectangle packing)
# ---------------------------------------------------------------------------

def pack_bottom_left(rects, plate_w, plate_h, gap=PART_GAP_IN):
    """rects: list of (w, h, label). Returns (placements, overflow_labels).

    placements: (x, y, w, h, label) with origin at bottom-left of the plate.
    Clusters toward the origin; parts that don't fit are returned as overflow.
    """
    placed, overflow = [], []
    candidates = [(0.0, 0.0)]

    def collides(x, y, w, h):
        for (px, py, pw, ph, _) in placed:
            if x < px + pw + gap and x + w + gap > px and \
               y < py + ph + gap and y + h + gap > py:
                return True
        return False

    for (w, h, label) in rects:
        spot = None
        for (cx, cy) in sorted(set(candidates), key=lambda c: (c[1], c[0])):
            if cx + w <= plate_w + 1e-6 and cy + h <= plate_h + 1e-6 \
                    and not collides(cx, cy, w, h):
                spot = (cx, cy)
                break
        if spot is None:
            overflow.append(label)
            continue
        x, y = spot
        placed.append((x, y, w, h, label))
        candidates.append((x + w + gap, y))
        candidates.append((x, y + h + gap))
    return placed, overflow


def build_nest_layout(nest, plate, parts):
    """Compute the simulated placement + derived metrics for one nest."""
    rects = []
    for p in parts:
        w, h = p["bbox"]
        for _ in range(max(p["qty"], 1)):
            rects.append((w, h, p["wo"]))
    rects.sort(key=lambda r: -(r[0] * r[1]))
    placements, overflow = ([], [])
    util = None
    if plate:
        placements, overflow = pack_bottom_left(rects, plate[0], plate[1])
        part_area = sum(w * h for (_, _, w, h, _) in placements)
        plate_area = plate[0] * plate[1]
        util = (part_area / plate_area * 100.0) if plate_area else None
    return {
        "nest": nest,
        "plate": plate,
        "parts": parts,
        "placements": placements,
        "overflow": overflow,
        "total_li": sum(p["total"] for p in parts),
        "util": util,
    }


# ---------------------------------------------------------------------------
# Excel report (Part 1)
# ---------------------------------------------------------------------------

def write_report(out_path, batch, nests, log):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Linear Inches"
    header = ["Batch", "Nest", "Part Number", "Work Order", "Qty",
              "Individual Linear Inches", "Total Linear Inches (Part x Qty)"]
    hf, hb = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="305496")
    nf, nb = Font(bold=True), PatternFill("solid", fgColor="D9E1F2")
    bf, bb = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F3864")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(thin, thin, thin, thin)

    ws.append(header)
    for c in ws[1]:
        c.font, c.fill = hf, hb
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border
    r = 2
    batch_total = 0.0
    for nest, parts in nests:
        nest_total = sum(p["total"] for p in parts)
        for p in parts:
            ws.append([batch, nest, p["part"], p["wo"], p["qty"],
                       round(p["indiv"], 2), round(p["total"], 2)])
            for c in ws[r]:
                c.border = border
            r += 1
        ws.append(["", "", "", "", f"NEST {nest} TOTAL", "", round(nest_total, 2)])
        for c in ws[r]:
            c.font, c.fill, c.border = nf, nb, border
        r += 1
        batch_total += nest_total
    ws.append(["", "", "", "", f"BATCH {batch} TOTAL", "", round(batch_total, 2)])
    for c in ws[r]:
        c.font, c.fill, c.border = bf, bb, border

    for i, w in enumerate([10, 12, 22, 13, 8, 16, 20], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    try:
        wb.save(out_path)
    except PermissionError:
        alt = out_path[:-5] + " (new).xlsx"
        wb.save(alt)
        log(f"  ! '{os.path.basename(out_path)}' was open/locked - saved as '{os.path.basename(alt)}' instead.")
        return alt, batch_total
    return out_path, batch_total


# ---------------------------------------------------------------------------
# Part 2 - layout viewer window
# ---------------------------------------------------------------------------

class NestView(QGraphicsView):
    """Graphics view with wheel zoom and middle/right-drag pan."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._pan_origin = None
        self.setRenderHint(QPainter.Antialiasing)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setBackgroundBrush(QColor("#14141c"))
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

    def wheelEvent(self, event):
        factor = 1.25 if event.angleDelta().y() > 0 else 0.8
        self.scale(factor, factor)

    def mousePressEvent(self, event):
        if event.button() in (Qt.MiddleButton, Qt.RightButton):
            self._pan_origin = event.position().toPoint()
            self.setCursor(Qt.ClosedHandCursor)
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._pan_origin is not None:
            delta = event.position().toPoint() - self._pan_origin
            self._pan_origin = event.position().toPoint()
            self.horizontalScrollBar().setValue(self.horizontalScrollBar().value() - delta.x())
            self.verticalScrollBar().setValue(self.verticalScrollBar().value() - delta.y())
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self._pan_origin is not None and event.button() in (Qt.MiddleButton, Qt.RightButton):
            self._pan_origin = None
            self.setCursor(Qt.ArrowCursor)
            return
        super().mouseReleaseEvent(event)


class NestLayoutWindow(PluginWindow):

    def __init__(self, batch, layouts, report_path, log=print, on_success=None):
        super().__init__("911_linear_inches", f"911 Linear Inches - {batch}")
        self.resize(1280, 820)
        self._batch = batch
        self._layouts = layouts            # list of build_nest_layout dicts
        self._report_path = report_path
        self._log = log
        self._on_success = on_success
        self._plate_rect = QRectF()
        self._build_ui()
        if layouts:
            self.cmb_nest.setCurrentIndex(0)
            self._show_nest(0)

    def _build_ui(self):
        container = QWidget()
        root = QVBoxLayout(container)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        toolbar = QHBoxLayout()
        toolbar.addWidget(QLabel("Nest:"))
        self.cmb_nest = QComboBox()
        for lay in self._layouts:
            self.cmb_nest.addItem(lay["nest"])
        self.cmb_nest.currentIndexChanged.connect(self._show_nest)
        toolbar.addWidget(self.cmb_nest)
        self.btn_fit = QPushButton("Fit View")
        self.btn_fit.clicked.connect(self.fit_view)
        toolbar.addWidget(self.btn_fit)
        self.btn_report = QPushButton("Open Excel Report")
        self.btn_report.clicked.connect(self._open_report)
        toolbar.addWidget(self.btn_report)
        toolbar.addSpacing(12)
        self.lbl_summary = QLabel("")
        self.lbl_summary.setStyleSheet("font-weight: bold;")
        toolbar.addWidget(self.lbl_summary)
        toolbar.addStretch(1)
        root.addLayout(toolbar)

        splitter = QSplitter(Qt.Horizontal)
        self.scene = QGraphicsScene()
        self.view = NestView()
        self.view.setScene(self.scene)
        splitter.addWidget(self.view)

        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(6)
        grp = QGroupBox("Parts in this nest")
        gl = QVBoxLayout(grp)
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(
            ["Part Number", "Work Order", "Qty", "Indiv in", "Total in"])
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        gl.addWidget(self.table)
        self.lbl_totals = QLabel("")
        self.lbl_totals.setStyleSheet("font-size: 14px; font-weight: bold;")
        gl.addWidget(self.lbl_totals)
        self.lbl_note = QLabel(
            "Simulated bottom-left-fill layout - approximate configuration, "
            "not the exact SigmaNest placement.")
        self.lbl_note.setWordWrap(True)
        self.lbl_note.setStyleSheet("color: #888;")
        gl.addWidget(self.lbl_note)
        rl.addWidget(grp, 1)
        splitter.addWidget(right)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 2)
        splitter.setSizes([760, 460])
        root.addWidget(splitter, 1)

        self.set_content(container)

    # ----- rendering --------------------------------------------------------

    def _add_dim_text(self, text, x, y, color="#cfd6ff"):
        item = QGraphicsSimpleTextItem(text)
        item.setBrush(QBrush(QColor(color)))
        item.setFlag(QGraphicsItem.ItemIgnoresTransformations)
        item.setPos(x, y)
        item.setZValue(20)
        self.scene.addItem(item)
        return item

    def _show_nest(self, index):
        if index < 0 or index >= len(self._layouts):
            return
        lay = self._layouts[index]
        self.scene.clear()
        plate = lay["plate"]

        if plate:
            pw, ph = plate
            # Plate outline (y-up -> negate y for Qt's y-down scene)
            plate_item = QGraphicsRectItem(0, -ph, pw, ph)
            pen = QPen(QColor("#e8e8e8"))
            pen.setCosmetic(True)
            pen.setWidthF(2.0)
            plate_item.setPen(pen)
            self.scene.addItem(plate_item)
            self._plate_rect = QRectF(0, -ph, pw, ph)
            # Plate dimension labels
            self._add_dim_text(f'{pw:.2f}"', pw / 2.0, 6, "#ffd479")
            self._add_dim_text(f'{ph:.2f}"', -2, -ph / 2.0, "#ffd479")
        else:
            self._plate_rect = QRectF()

        part_pen = QPen(QColor("#5aa0ff"))
        part_pen.setCosmetic(True)
        part_pen.setWidthF(1.6)
        part_brush = QBrush(QColor(90, 160, 255, 60))
        for (x, y, w, h, label) in lay["placements"]:
            rect = QGraphicsRectItem(x, -(y + h), w, h)
            rect.setPen(part_pen)
            rect.setBrush(part_brush)
            self.scene.addItem(rect)
            t = self._add_dim_text(f'{w:.1f}x{h:.1f}', x + w / 2.0, -(y + h / 2.0), "#dfe8ff")
            # rough-center the fixed-size label on the part
            br = t.boundingRect()
            t.setPos(x + w / 2.0 - br.width() * 0.0, -(y + h / 2.0))

        if not self._plate_rect.isNull():
            pad_x = max(self._plate_rect.width() * 0.15, 2.0)
            pad_y = max(self._plate_rect.height() * 0.15, 2.0)
            self.scene.setSceneRect(self._plate_rect.adjusted(-pad_x, -pad_y, pad_x, pad_y))
        QTimer.singleShot(0, self.fit_view)

        # summary + table
        n_parts = len(lay["placements"])
        util = lay["util"]
        util_s = f" | util {util:.1f}%" if util is not None else ""
        over = f" | OVERFLOW {len(lay['overflow'])}" if lay["overflow"] else ""
        plate_s = f'{plate[0]:.2f} x {plate[1]:.2f} in' if plate else "plate size unknown"
        self.lbl_summary.setText(
            f"{lay['nest']}: plate {plate_s} | {n_parts} parts | "
            f"cut {lay['total_li']:.1f} in{util_s}{over}")

        self.table.setRowCount(0)
        for p in lay["parts"]:
            row = self.table.rowCount()
            self.table.insertRow(row)
            vals = [p["part"], p["wo"], str(p["qty"]),
                    f'{p["indiv"]:.2f}', f'{p["total"]:.2f}']
            for col, v in enumerate(vals):
                item = QTableWidgetItem(v)
                if col >= 2:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(row, col, item)
        self.lbl_totals.setText(f"Nest total: {lay['total_li']:.2f} in "
                                f"({lay['total_li'] / 12.0:.2f} ft)")

    def fit_view(self):
        if not self._plate_rect.isNull():
            mx = self._plate_rect.width() * 0.12
            my = self._plate_rect.height() * 0.12
            self.view.fitInView(self._plate_rect.adjusted(-mx, -my, mx, my),
                                Qt.KeepAspectRatio)

    def _open_report(self):
        if self._report_path and os.path.exists(self._report_path):
            os.startfile(self._report_path)  # noqa: S606 (Windows-only app)

    def fire_success(self):
        if callable(self._on_success):
            self._on_success()


# ---------------------------------------------------------------------------
# Entry point (GUI plugin - runs on the Qt main thread)
# ---------------------------------------------------------------------------

def run(params, progress_callback, cancel_event):
    global _window
    log = params.get("log", print)
    settings = params.get("settings", {}) or {}

    start = (settings.get("default_root") or "").strip()
    selected = QFileDialog.getExistingDirectory(
        None, "Select the batch's 'IGES FILES' folder (or the batch folder)", start)
    if not selected:
        log("Cancelled - no folder selected.")
        return

    iges_root, batch_dir = resolve_roots(selected)
    batch = os.path.basename(batch_dir.rstrip("\\/")) or batch_dir
    nest_pkg_dir = _find_child_dir(batch_dir, "NEST PACKAGES")
    log(f"Batch: {batch}")
    log(f"IGES root: {iges_root}")
    log(f"Nest packages: {nest_pkg_dir or '(not found - will try MOVE TICKET OMIT pdfs)'}")

    try:
        subs = sorted(d for d in os.listdir(iges_root)
                      if os.path.isdir(os.path.join(iges_root, d)))
    except OSError as exc:
        QMessageBox.critical(None, "911 Linear Inches", f"Could not read:\n{iges_root}\n\n{exc}")
        return
    if not subs:
        QMessageBox.warning(None, "911 Linear Inches", f"No nest subfolders found in:\n{iges_root}")
        return

    nests = []       # (nest, [part dicts]) for the report
    layouts = []     # build_nest_layout dicts for the viewer
    flags = []
    for idx, sub in enumerate(subs):
        progress_callback(int(5 + 80 * idx / len(subs)))
        nest = derive_nest(sub)
        if nest == sub:
            flags.append(f"{sub}: folder name is not the expected N<nest>S format")
        summary, plate = read_nest_package(nest, nest_pkg_dir, batch_dir, log)
        if not summary:
            flags.append(f"{nest}: no SUMMARY OF NEST found - quantities unknown")
        if plate is None:
            flags.append(f"{nest}: plate size not found - layout skipped")

        dxfs = sorted(glob.glob(os.path.join(iges_root, sub, "*.dxf")))
        if not dxfs:
            flags.append(f"{nest}: subfolder has no DXF files")
        parts = []
        for dxf in dxfs:
            fn = os.path.basename(dxf)
            try:
                sdk.ensure_local(dxf, log=log)
                m = dxf_metrics(dxf)
            except Exception as exc:
                flags.append(f"{nest}/{fn}: could not read DXF ({exc})")
                continue
            wo_match = WO_RE.search(fn)
            wo = wo_match.group(0) if wo_match else "?"
            part, qty = summary.get(wo, ("(unmatched)", 0))
            if qty == 0:
                flags.append(f"{nest}/{fn}: no quantity match for work order {wo}")
            if m["skipped"]:
                bad = ", ".join(f"{k}x{v}" for k, v in m["skipped"].items()
                                if k in UNSUPPORTED_FLAGGED)
                if bad:
                    flags.append(f"{nest}/{fn}: unmeasured geometry ({bad}) - length may be low")
            if len(m["layers"]) > 1:
                flags.append(f"{nest}/{fn}: {len(m['layers'])} layers - verify all count as cut")
            if m["unit_note"]:
                flags.append(f"{nest}/{fn}: {m['unit_note']}")
            indiv = m["linear_inches"]
            parts.append({"part": part, "wo": wo, "qty": qty, "bbox": m["bbox"],
                          "indiv": indiv, "total": indiv * qty})
            log(f"  {nest}  {part}  WO {wo}  {indiv:.2f} in x {qty} = {indiv * qty:.2f} in")
        if parts:
            nests.append((nest, parts))
            layout = build_nest_layout(nest, plate, parts)
            if layout["overflow"]:
                flags.append(f"{nest}: {len(layout['overflow'])} part(s) did not fit the "
                             f"plate in the simulated layout")
            layouts.append(layout)

    if not nests:
        QMessageBox.warning(None, "911 Linear Inches",
                            "No part DXFs could be measured. Check the folder selection.")
        return

    progress_callback(92)
    out_path = os.path.join(batch_dir, f"{batch} Linear Inches.xlsx")
    saved_path, batch_total = write_report(out_path, batch, nests, log)
    log("")
    log(f"BATCH {batch} TOTAL = {batch_total:.2f} in = {batch_total / 12:.2f} ft")
    log(f"Report: {saved_path}")

    console = params.get("console")
    if console is not None and hasattr(console, "append_link"):
        try:
            console.append_link(os.path.basename(saved_path), saved_path,
                                prefix="REPORT", at_run_end=True)
        except TypeError:
            console.append_link(os.path.basename(saved_path), saved_path)

    progress_callback(97)
    _window = NestLayoutWindow(batch, layouts, saved_path, log=log,
                               on_success=params.get("on_success"))
    _window.show()
    _window.fire_success()

    if flags:
        log("")
        log(f"{len(flags)} item(s) flagged for review:")
        for f in flags:
            log(f"  ! {f}")
        QMessageBox.warning(_window, "911 Linear Inches - review needed",
                            "Report saved, but some items need a look:\n\n" +
                            "\n".join(f"- {f}" for f in flags[:20]) +
                            ("\n..." if len(flags) > 20 else ""))
    progress_callback(100)
    log("Layout viewer opened.")
