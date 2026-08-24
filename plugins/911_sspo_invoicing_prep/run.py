"""
911 SSPO Invoicing Prep  (plugin id: 911_sspo_invoicing_prep, family: 911)
==========================================================================
Takes a big SSPO pricing sheet (like "SSPO Pricing Back Up") and splits it into one
workbook per Batch + Nest, named "{BATCH} {NEST} Pricing Back Up.xlsx", each saved in
its own "{BATCH} {NEST} Invoicing Docs" subfolder of a new folder next to the file
the user picks.

The top level of that output folder also gets a "D911 Workorder Close Outs
{m-d-yyyy}.xlsx" reconstruction of the sheet previously ripped by hand from the
pricing master: the source's columns from A through "Machine", values AND cell
styles copied verbatim, with every row's Scheduling Group set to "Closed".

Each output workbook is built ENTIRELY from scratch (no template file) and gets:

  Tab 1 "Pricing Back Up"    - the split rows for that Batch+Nest (value snapshot,
                               number formats preserved) + a SUM total under the
                               "Total price per WO" column.
  Tab 2 "Invoice Supplement" - the ASA invoice sheet (logo from the bundled
                               asa_logo.png, blue header band, accounting formats),
                               one line per work order:
                                 A PO / B PO Line  <- looked up from the Working
                                   Forecast List ('911 Forecast' sheet, falling back
                                   to 'Complete 911 QTDR'), matched on Batch + Nest,
                                   copied verbatim (Line may be text like "SSPO")
                                 C-H               <- Batch, Work Order, DYPN,
                                   Material, DYPN QTY, Nest Pkg Nbr from tab 1
                                 I Price/Ea        <- formula =J{r}/G{r}
                                 J Ext. Price      <- live formula into tab 1's
                                   "Total price per WO" cell
                               plus a Grand Total =SUM over column J. Invoice # and
                               Invoice Date are left blank for manual fill.

The live Working Forecast List is NEVER opened directly: it's copied into the
output folder first, the copy is read, and the copy is deleted once every workbook
has been written (even on error/cancel).

GUI plugin: pops a file picker on the main Qt thread (requires_main_thread), splits,
then opens the results folder. A "back up" = a value snapshot (cached numbers, not live
formulas) with each cell's number format preserved so $ and dates still display right.
"""

import os
import re
import shutil
from copy import copy
from datetime import date
from pathlib import Path

# --- SDK bootstrap (works under the app and for headless CLI testing) -------------
try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys
    import pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from PySide6.QtWidgets import QMessageBox

# Hard Rule 3 nest-number regex: accept legacy numeric IDs (503633, P08229) AND
# alphanumeric IDs that contain a digit (5CDAWK); reject footer/total/junk text.
NEST_RE = sdk.NEST_ID_RE  # single home in the SDK — never re-type the pattern

REQUIRED_HEADERS = ["Batch", "Nest Pkg Nbr"]

PBU_SHEET = "Pricing Back Up"
INV_SHEET = "Invoice Supplement"
TOTAL_HEADER = "TOTAL PRICE PER WO"
LOGO_NAME = "asa_logo.png"

# Workorder Close Outs sheet: the source's columns A..Machine, verbatim, with
# Scheduling Group forced to "Closed" (reconstructs the sheet previously hand-ripped
# from the pricing master).
CLOSEOUT_LAST_HEADER = "MACHINE"
CLOSEOUT_STATUS_HEADER = "SCHEDULING GROUP"
CLOSEOUT_STATUS_VALUE = "Closed"

# Forecast sheets searched for the PO / PO Line, in priority order (active
# forecast first; finished batches roll off to the Complete sheet).
FORECAST_SHEETS = ["911 Forecast", "Complete 911 QTDR"]
DEFAULT_FORECAST_FILENAME = "Working Forecast List.xlsx"
FORECAST_COPY_NAME = "~ Working Forecast List (temp copy).xlsx"

# ---- Invoice Supplement look (reproduced from the hand-made ASA sheet) ------------
INV_HEADERS = ["PO ", "PO Line", "Batch", "Workorder", "DYPN",
               "Source Material", "Qty", "Nest ", "Price/Ea", "Ext. Price"]
INV_COL_WIDTHS = {"A": 11.0, "B": 9.71, "C": 13.86, "D": 13.0, "E": 14.43,
                  "F": 13.86, "H": 14.43, "I": 14.29, "J": 11.57}
INV_HEADER_ROW = 7
INV_DATA_START = 8
# px, anchor A1. Was (553, 202) = the logo's natural size, which sprawled over
# the invoice fields (F2/F3) and the header band; resized 2026-07-13 to match
# invoicing's corrected sheet (434x121 = 4133850x1152525 EMU) so it stays
# inside its A1:D5 home.
LOGO_SIZE = (434, 121)

ACCT_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
_THIN = Side(style="thin")
_MEDIUM = Side(style="medium")
HDR_FONT = Font(name="Tahoma", size=8, bold=True, color="FFFFFFFF")
HDR_FILL = PatternFill("solid", fgColor="FF273991")
HDR_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
LABEL_FONT = Font(name="Tahoma", size=10, bold=True)
DATA_FONT = Font(name="Tahoma", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")
GT_LABEL_BORDER = Border(left=_THIN, right=_MEDIUM, top=_THIN, bottom=_THIN)
GT_VALUE_BORDER = Border(right=_THIN, top=_THIN, bottom=_THIN)


def _as_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _safe_filename(name):
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()


def _pick_sheet(wb, log):
    """Find the sheet + header row that has both Batch and Nest Pkg Nbr (Hard Rules
    1-2: locate by name, scan for the header row). Returns (ws, header_row, header_map)."""
    for ws in wb.worksheets:
        hdr_row, hmap = sdk.find_header_row(ws, REQUIRED_HEADERS, max_scan=12)
        if hdr_row:
            log(f"Using sheet '{ws.title}' (header on row {hdr_row}).")
            return ws, hdr_row, hmap
    raise ValueError(
        "Could not find a sheet with 'Batch' and 'Nest Pkg Nbr' columns.")


# ---------------------------------------------------------------------------------
# PO / PO Line lookup from the Working Forecast List
# ---------------------------------------------------------------------------------
def _copy_forecast(settings, out_dir, log):
    """Copy the Working Forecast List into out_dir so the live file is never read
    directly. Returns the copy's Path, or None (with a logged warning) if the
    forecast can't be located — the split still runs, PO/Line just stay blank."""
    override = str(settings.get("forecast_dir", "") or "").strip()
    fdir = sdk.resolve_forecast_dir(override)
    if fdir is None:
        log("WARNING: could not locate the 'Forecast and Inventory Reports' folder - "
            "PO / PO Line will be left blank. Set it in this app's Settings.")
        return None
    fname = str(settings.get("forecast_filename", "") or "").strip() or DEFAULT_FORECAST_FILENAME
    src = Path(fdir) / fname
    if not sdk.exists(src):
        log(f"WARNING: forecast workbook not found at {src} - "
            "PO / PO Line will be left blank.")
        return None
    dest = out_dir / FORECAST_COPY_NAME
    log(f"Copying {src.name} into the output folder (the live file is never opened)...")
    sdk.copy_resilient(src, dest, log)  # hydrate + locked-open message (Hard Rule 13)
    return dest


def _read_po_map(copy_path, log, cancel_event):
    """{(BATCH, NEST): (po, po_line)} from the local forecast copy.

    Reads the '911 Forecast' sheet first, then 'Complete 911 QTDR' (older batches
    roll off to it) — first sheet wins on duplicate keys. PO and Line are copied
    verbatim (Line is sometimes the text 'SSPO', not a number).
    """
    # Resilient even on the local copy: it can be open in Excel, and the
    # resilient loader is a cheap no-op on a healthy local file (Hard Rule 13).
    wb = sdk.load_workbook_resilient(copy_path, log=log, data_only=True,
                                     read_only=True)
    po_map = {}
    try:
        for sheet_name in FORECAST_SHEETS:
            if sheet_name not in wb.sheetnames:
                log(f"WARNING: sheet '{sheet_name}' not found in the forecast.")
                continue
            ws = wb[sheet_name]
            cols = None  # (po, line, batch, nest) 0-based indices, set at header row
            for i, row in enumerate(ws.iter_rows(max_col=60)):
                if cancel_event is not None and i % 256 == 0 and cancel_event.is_set():
                    return {}
                vals = [c.value for c in row]
                if cols is None:
                    # Scan for the header row (Hard Rule 2). Batch header varies
                    # ('Batch /DR' vs 'Batch/Order'), so match any BATCH* header.
                    hmap = {str(v).strip().upper(): j
                            for j, v in enumerate(vals) if isinstance(v, str) and v.strip()}
                    batch_col = next((j for h, j in hmap.items()
                                      if h.startswith("BATCH")), None)
                    if "PO" in hmap and "LINE" in hmap and "NEST" in hmap \
                            and batch_col is not None:
                        cols = (hmap["PO"], hmap["LINE"], batch_col, hmap["NEST"])
                    if i >= 8 and cols is None:
                        log(f"WARNING: no PO/Line/Batch/Nest header row found in "
                            f"'{sheet_name}' (looked in the first 8 rows).")
                        break
                    continue
                i_po, i_line, i_batch, i_nest = cols
                batch = _as_str(vals[i_batch] if i_batch < len(vals) else None).upper()
                nest = _as_str(vals[i_nest] if i_nest < len(vals) else None).upper()
                po = vals[i_po] if i_po < len(vals) else None
                if not batch or not NEST_RE.match(nest) or po in (None, ""):
                    continue
                line = vals[i_line] if i_line < len(vals) else None
                po_map.setdefault((batch, nest), (po, line))
    finally:
        wb.close()
    log(f"  found PO numbers for {len(po_map)} batch+nest combinations.")
    return po_map


# ---------------------------------------------------------------------------------
# Output workbook (built from scratch: split data on tab 1, invoice supplement on 2)
# ---------------------------------------------------------------------------------
def _build_supplement(wb, hmap, rows, po_info, logo_path):
    """Create the Invoice Supplement sheet from scratch in `wb` (whose first sheet
    is the filled Pricing Back Up tab). `hmap` = {UPPER HEADER: 1-based col} of the
    source slice; `po_info` is (po, line) or None."""
    ws = wb.create_sheet(INV_SHEET)
    for letter, width in INV_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    # Title block: ASA logo over merged A1:D5, manual-fill invoice fields at F2/F3.
    ws.merge_cells("A1:D5")
    if sdk.exists(logo_path):
        img = XLImage(str(logo_path))
        img.width, img.height = LOGO_SIZE
        ws.add_image(img, "A1")
    ws["F2"] = "Invoice #"
    ws["F2"].font = LABEL_FONT
    ws["F3"] = "Invoice Date:"
    ws["F3"].font = LABEL_FONT
    ws["G3"].font = DATA_FONT
    ws["G3"].number_format = "mm-dd-yy"

    # Header band.
    for j, title in enumerate(INV_HEADERS, start=1):
        c = ws.cell(row=INV_HEADER_ROW, column=j, value=title)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = HDR_BORDER
        c.alignment = HDR_ALIGN

    # Data rows. Column C-H sources on tab 1, looked up by header NAME (Hard Rule 1).
    src_cols = [hmap.get(h) for h in
                ("BATCH", "WORK ORDER", "DYPN", "MATERIAL", "DYPN QTY", "NEST PKG NBR")]
    total_col = hmap.get(TOTAL_HEADER)
    total_letter = get_column_letter(total_col) if total_col else None
    po, po_line = po_info if po_info else (None, None)

    for idx, src_row in enumerate(rows):
        r = INV_DATA_START + idx
        src_r = 2 + idx                       # matching data row on tab 1
        for j in range(1, 11):
            c = ws.cell(row=r, column=j)
            c.font = DATA_FONT
            c.alignment = DATA_ALIGN
        ws.cell(row=r, column=1, value=po)
        ws.cell(row=r, column=2, value=po_line)
        for k, sc in enumerate(src_cols):     # C..H
            if sc:
                ws.cell(row=r, column=3 + k, value=src_row[sc - 1].value)
        pea = ws.cell(row=r, column=9, value=f"=J{r}/G{r}")      # I: Price/Ea
        pea.number_format = ACCT_FMT
        ext = ws.cell(row=r, column=10)                          # J: Ext. Price
        ext.number_format = ACCT_FMT
        if total_letter:
            ext.value = f"='{PBU_SHEET}'!{total_letter}{src_r}"

    # Blank spacer row, then the boxed Grand Total.
    gt_row = INV_DATA_START + len(rows) + 1
    lbl = ws.cell(row=gt_row, column=9, value="Grand Total")
    lbl.font = LABEL_FONT
    lbl.border = GT_LABEL_BORDER
    lbl.alignment = Alignment(horizontal="center", vertical="top")
    val = ws.cell(row=gt_row, column=10,
                  value=f"=SUM(J{INV_DATA_START}:J{gt_row - 1})")
    val.font = Font(bold=True)
    val.border = GT_VALUE_BORDER
    val.number_format = ACCT_FMT


def _copy_cell(src_c, dst_c, value=None):
    """Copy a source cell's value + full style (font/fill/border/alignment/number
    format) onto dst_c; `value` overrides the copied value."""
    dst_c.value = src_c.value if value is None else value
    dst_c.font = copy(src_c.font)
    dst_c.fill = copy(src_c.fill)
    dst_c.border = copy(src_c.border)
    dst_c.alignment = copy(src_c.alignment)
    dst_c.number_format = src_c.number_format


def _write_closeouts(src_ws, hdr_row, hmap, valid_rows, out_dir, log):
    """Write the 'D911 Workorder Close Outs {m-d-yyyy}.xlsx' workbook at the top of
    out_dir: the source sheet's columns A through "Machine" (values and cell styles
    verbatim, so the hand-ripped original is reproduced exactly) with every row's
    Scheduling Group set to "Closed". Returns the filename written."""
    last_col = hmap.get(CLOSEOUT_LAST_HEADER)
    if not last_col:
        last_col = max(hmap.values())
        log(f"  WARNING: no '{CLOSEOUT_LAST_HEADER}' column - the Close Outs sheet "
            "will include every column instead.")
    status_col = hmap.get(CLOSEOUT_STATUS_HEADER)
    if not status_col or status_col > last_col:
        status_col = None
        log(f"  WARNING: no '{CLOSEOUT_STATUS_HEADER}' column - no rows marked "
            f"'{CLOSEOUT_STATUS_VALUE}'.")

    wb = openpyxl.Workbook()
    ws = wb.active                      # stays "Sheet1", like the hand-made original
    for j in range(1, last_col + 1):
        letter = get_column_letter(j)
        dim = src_ws.column_dimensions.get(letter)
        if dim is not None and dim.width:
            ws.column_dimensions[letter].width = dim.width
        _copy_cell(src_ws.cell(row=hdr_row, column=j), ws.cell(row=1, column=j))
    for r_i, src_row in enumerate(valid_rows, start=2):
        for j in range(1, last_col + 1):
            _copy_cell(src_row[j - 1], ws.cell(row=r_i, column=j),
                       value=CLOSEOUT_STATUS_VALUE if j == status_col else None)

    today = date.today()
    fname = _safe_filename(
        f"D911 Workorder Close Outs {today.month}-{today.day}-{today.year}.xlsx")
    sdk.save_workbook(wb, out_dir / fname)
    return fname


def _write_output(headers, hmap, rows, po_info, logo_path, out_path, log):
    """Build one output workbook from scratch for a single Batch+Nest group."""
    wb = openpyxl.Workbook()

    # ---- Tab 1: the split pricing rows (value snapshot) -------------------------
    ws1 = wb.active
    ws1.title = PBU_SHEET
    last_col = len(headers)
    for j, h in enumerate(headers, start=1):
        c = ws1.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
    for r_i, src_row in enumerate(rows, start=2):
        for j, src_c in enumerate(src_row, start=1):
            dst = ws1.cell(row=r_i, column=j, value=src_c.value)
            if src_c.number_format and src_c.number_format != "General":
                dst.number_format = src_c.number_format
    for j in range(1, last_col + 1):
        ws1.column_dimensions[get_column_letter(j)].width = 16

    last_data = 1 + len(rows)
    total_col = hmap.get(TOTAL_HEADER)
    if total_col:
        letter = get_column_letter(total_col)
        tot = ws1.cell(row=last_data + 1, column=total_col,
                       value=f"=SUM({letter}2:{letter}{last_data})")
        tot.number_format = '"$"#,##0.00'
        tot.font = Font(bold=True)
    else:
        log(f"  WARNING: no '{TOTAL_HEADER}' column - skipped the tab-1 total.")

    # ---- Tab 2: Invoice Supplement, built from scratch ---------------------------
    _build_supplement(wb, hmap, rows, po_info, logo_path)
    sdk.save_workbook(wb, out_path)


def split_workbook(src_path, out_dir, settings, log,
                   progress_callback=None, cancel_event=None):
    """Split src_path into one from-scratch workbook per (Batch, Nest), each in its
    own "{BATCH} {NEST} Invoicing Docs" subfolder, plus the Workorder Close Outs
    workbook at the top of out_dir. Returns (written, missing_po, closeout_name)
    where written = [(relative path, row_count)], missing_po = [display names of
    groups whose PO wasn't in the forecast], and closeout_name is the Close Outs
    filename (None if cancelled before it was written)."""
    src_path = Path(src_path)
    out_dir = Path(out_dir)
    logo_path = Path(__file__).resolve().parent / LOGO_NAME

    sdk.ensure_dir(out_dir)
    forecast_copy = None
    try:
        forecast_copy = _copy_forecast(settings, out_dir, log)
        if progress_callback:
            progress_callback(8)
        po_map = _read_po_map(forecast_copy, log, cancel_event) if forecast_copy else {}
        if cancel_event is not None and cancel_event.is_set():
            log("Cancelled.")
            return [], [], None
        if progress_callback:
            progress_callback(12)

        sdk.ensure_local(src_path, log=log)      # hydrate OneDrive placeholders (Rule 13)
        log(f"Opening {src_path.name} ...")
        wb = sdk.load_workbook_resilient(src_path, log=log, data_only=True)
        ws, hdr_row, hmap = _pick_sheet(wb, log)

        last_col = max(hmap.values())             # 1-based last real header column
        headers = [ws.cell(row=hdr_row, column=j).value for j in range(1, last_col + 1)]
        i_batch = hmap["BATCH"]                    # 1-based
        i_nest = hmap["NEST PKG NBR"]

        # Group data rows by (batch, nest), preserving first-seen order; keep the
        # flat source-order row list too for the Close Outs sheet.
        groups, order, valid_rows, skipped = {}, [], [], 0
        for i, row in enumerate(ws.iter_rows(min_row=hdr_row + 1, max_col=last_col)):
            if cancel_event is not None and i % 256 == 0 and cancel_event.is_set():
                log("Cancelled.")
                return [], [], None
            if all(c.value in (None, "") for c in row):
                continue
            nest = _as_str(row[i_nest - 1].value)
            batch = _as_str(row[i_batch - 1].value)
            if not NEST_RE.match(nest):
                skipped += 1
                continue
            key = (batch, nest)
            if key not in groups:
                groups[key] = []
                order.append(key)
            groups[key].append(row)
            valid_rows.append(row)

        if not order:
            raise ValueError("No rows with a valid nest number were found.")

        closeout_name = _write_closeouts(ws, hdr_row, hmap, valid_rows, out_dir, log)
        log(f"  wrote {closeout_name}  ({len(valid_rows)} rows, "
            f"Scheduling Group -> '{CLOSEOUT_STATUS_VALUE}')")

        written, missing_po = [], []
        for gi, (batch, nest) in enumerate(order):
            if cancel_event is not None and cancel_event.is_set():
                log("Cancelled.")
                break
            rows = groups[(batch, nest)]
            po_info = po_map.get((batch.upper(), nest.upper()))
            if po_info is None:
                missing_po.append(f"{batch} {nest}")
                log(f"  WARNING: {batch} {nest} not found in the forecast - "
                    "PO / PO Line left blank.")
            sub_dir = out_dir / _safe_filename(f"{batch} {nest} Invoicing Docs")
            sdk.ensure_dir(sub_dir)
            fname = _safe_filename(f"{batch} {nest} Pricing Back Up.xlsx")
            _write_output(headers, hmap, rows, po_info, logo_path,
                          sub_dir / fname, log)
            rel = f"{sub_dir.name}\\{fname}"
            written.append((rel, len(rows)))
            log(f"  wrote {rel}  ({len(rows)} row{'s' if len(rows) != 1 else ''})")
            if progress_callback:
                progress_callback(15 + int(80 * (gi + 1) / len(order)))

        if skipped:
            log(f"Skipped {skipped} row(s) without a valid nest number (footers/blanks).")
        return written, missing_po, closeout_name
    finally:
        # The forecast copy is working scratch only - always remove it, even on
        # error/cancel, so it never lingers next to the real output files.
        if forecast_copy is not None and sdk.exists(forecast_copy):
            try:
                forecast_copy.unlink()
                log("Removed the temporary forecast copy.")
            except OSError as e:
                log(f"WARNING: could not delete the temporary forecast copy "
                    f"({forecast_copy.name}): {e} - delete it by hand.")


# ---------------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------------
def run(params, progress_callback, cancel_event):
    log = params.get("log", print)
    settings = params.get("settings", {}) or {}

    start_dir = str(Path.home() / "Downloads")
    # GUI plugin (main thread), so the console's blocking picker is off-limits:
    # sdk.pick_file_gui gives the same Sentry Drone contract (kill-cam when the
    # drone is owned + enabled for this app, plain dialog otherwise).
    src = sdk.pick_file_gui(
        params, "Choose the pricing workbook to split", start_dir,
        "Excel files (*.xlsx *.xlsm);;All files (*.*)")
    if not src:
        log("No file chosen - nothing to do.")
        return

    src = Path(src)
    out_dir = src.parent / f"{src.stem} - Back Ups"
    progress_callback(5)

    try:
        written, missing_po, closeout_name = split_workbook(
            src, out_dir, settings, log, progress_callback, cancel_event)
    except Exception as e:
        log(f"ERROR: {e}")
        QMessageBox.critical(None, "911 SSPO Invoicing Prep",
                             f"Could not split the file:\n\n{e}")
        return

    if not written:                                # cancelled
        return

    total = sum(n for _, n in written)
    progress_callback(100)
    log(f"Done. Wrote {len(written)} file(s), {total} data row(s) total.")
    log(f"Folder: {out_dir}")
    # GUI plugins suppress the shell's auto success chime; fire it ourselves now
    # that the files are actually written (the meaningful moment).
    on_success = params.get("on_success")
    if callable(on_success):
        on_success()
    msg = (f"Done! Wrote {len(written)} back-up file(s) "
           f"({total} rows total), each in its own Invoicing Docs folder, to:"
           f"\n\n{out_dir}")
    if closeout_name:
        msg += f"\n\nWorkorder Close Outs sheet (top level): {closeout_name}"
    if missing_po:
        msg += ("\n\nNo PO found in the Working Forecast List for:\n  "
                + "\n  ".join(missing_po)
                + "\n\nTheir PO / PO Line columns were left blank - fill them in "
                  "by hand or fix the forecast and re-run.")
    QMessageBox.information(None, "911 SSPO Invoicing Prep", msg)
    try:
        os.startfile(str(out_dir))                 # open the results folder in Explorer
    except Exception:
        pass
