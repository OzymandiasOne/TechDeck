"""
911 Repeater Plugin
===================
Finds repeat parts for a 911 QTDR batch by cross-referencing DYPN numbers
from each nest's 911 Batch Excel against the Master Parts and Inspection Libraries.

Output structure per nest:
  <batch>/<nest>/NC Repeats/          <- .NC files
  <batch>/<nest>/Inspection Repeats/  <- PDFs with batch/nest text updated

Workflow:
  1. Prompt for batch number (e.g. V071)
  2. Locate batch folder under 911 QTDR root
  3. For each nest subfolder, read DYPNs from NEST sheet col K
  4. Match each DYPN against Master Parts Library (.NC) and Inspection Library (.pdf)
  5. Copy matches into per-nest output folders under REPEATER/<batch>/<nest>/
  6. In each copied PDF, find the old "batch nest" pair and replace with new values
  7. Report summary
"""

import re
import shutil
from pathlib import Path

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

def _repeater_root(qtdr_override: str = "") -> Path | None:
    """REPEATER folder under the 911 QTDR root. A configured override wins;
    otherwise auto-discovers the QTDR root across all OneDrive path variants;
    returns None if it can't be found."""
    qtdr = sdk.resolve_911_qtdr_root(qtdr_override)
    if qtdr is None:
        return None
    return (
        qtdr
        / "04 - Notes - Protocols - Tutorials"
        / "TECH SERVICES"
        / "REPEATER"
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Matches e.g. "S025 P07750", "V086 P07685", "V071 P08001"
_BATCH_NEST_RE = re.compile(r'[VS]\d{2,4}\s+P0\d+', re.IGNORECASE)


def _find_nest_excel(nest_folder: Path, batch_number: str, nest_number: str) -> Path | None:
    expected = nest_folder / f"911 Batch {batch_number} {nest_number}.xlsx"
    if expected.exists():
        return expected
    candidates = [
        f for f in nest_folder.glob("*.xlsx")
        if "batch" in f.name.lower() and not f.name.startswith("~")
    ]
    return candidates[0] if candidates else None


def _read_dypns(excel_path: Path) -> list[str]:
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    try:
        if "NEST" not in wb.sheetnames:
            return []
        ws = wb["NEST"]
        dypns = []
        for row in range(3, ws.max_row + 1):
            val = ws.cell(row, 11).value  # column K
            if val is not None:
                stripped = str(val).strip()
                if stripped:
                    dypns.append(stripped)
        return dypns
    finally:
        wb.close()


def _build_library_index(library_dir: Path, extension: str) -> dict[str, Path]:
    index: dict[str, Path] = {}
    if not library_dir.exists():
        return index
    ext_lower = extension.lower()
    for f in library_dir.iterdir():
        if f.is_file() and f.suffix.lower() == ext_lower:
            index[f.stem.upper()] = f
    return index


def _find_match_rect(line: dict, fitz) -> object | None:
    """
    Given a text line dict from get_text('dict'), find the bounding rect
    covering all spans that form the regex match.
    """
    spans = line.get("spans", [])
    if not spans:
        return None

    offsets = []
    pos = 0
    for span in spans:
        offsets.append(pos)
        pos += len(span["text"])

    full_text = "".join(s["text"] for s in spans)
    m = _BATCH_NEST_RE.search(full_text)
    if not m:
        return None

    match_start, match_end = m.start(), m.end()

    involved = []
    for i, span in enumerate(spans):
        s_start = offsets[i]
        s_end = s_start + len(span["text"])
        if s_start < match_end and s_end > match_start:
            involved.append(span)

    if not involved:
        return None

    x0 = min(s["bbox"][0] for s in involved)
    y0 = min(s["bbox"][1] for s in involved)
    x1 = max(s["bbox"][2] for s in involved)
    y1 = max(s["bbox"][3] for s in involved)
    return fitz.Rect(x0, y0, x1, y1)


def _replace_batch_nest_in_pdf(pdf_path: Path, new_batch: str, new_nest: str, log) -> bool:
    """
    Find the single old "batch nest" pair in the PDF and replace with
    "new_batch new_nest". Overwrites the file in place.
    Returns True if a replacement was made.
    """
    import fitz  # PyMuPDF

    replacement = f"{new_batch} {new_nest}"
    doc = fitz.open(str(pdf_path))
    replaced = False

    try:
        for page in doc:
            blocks = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]
            for block in blocks:
                if block.get("type") != 0:
                    continue
                for line in block.get("lines", []):
                    line_text = "".join(s["text"] for s in line.get("spans", []))
                    if not _BATCH_NEST_RE.search(line_text):
                        continue

                    old_text = _BATCH_NEST_RE.search(line_text).group(0)
                    log(f"    Found '{old_text}' -> '{replacement}'")

                    match_rect = _find_match_rect(line, fitz)
                    if match_rect is None:
                        log(f"    WARNING: Could not locate rect for '{old_text}' - skipping")
                        continue

                    font_size = line["spans"][0]["size"] if line["spans"] else 10

                    page.add_redact_annot(match_rect, fill=(1, 1, 1))
                    page.apply_redactions()
                    page.insert_text(
                        (match_rect.x0, match_rect.y1 - 1),
                        replacement,
                        fontsize=font_size,
                        color=(0, 0, 0),
                    )
                    replaced = True
                    break
                if replaced:
                    break

        if replaced:
            # Save to a temp file then replace - required when using redactions
            tmp = pdf_path.with_suffix(".tmp.pdf")
            doc.save(str(tmp), incremental=False, encryption=fitz.PDF_ENCRYPT_NONE)
            doc.close()
            tmp.replace(pdf_path)
            return True
        else:
            log(f"    WARNING: No batch/nest pattern found in PDF - file unchanged")
    finally:
        if not doc.is_closed:
            doc.close()

    return replaced


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run(params, progress_callback, cancel_event):
    log = params.get("log", print)
    console = params.get("console", None)
    settings = params.get("settings", {}) or {}

    log("911 Repeater starting...")
    progress_callback(0)

    # ------------------------------------------------------------------ #
    # Step 1 - Prompt for batch number
    # ------------------------------------------------------------------ #
    if console is not None:
        batch_number = console.request_input(
            "Enter Batch Number (e.g. V072, S045):"
        ).strip().upper()
    else:
        batch_number = input("Enter Batch Number (e.g. V072, S045): ").strip().upper()

    if not batch_number:
        log("No batch number entered - aborting.")
        return

    log(f"Batch         : {batch_number}")

    # ------------------------------------------------------------------ #
    # Step 2 - Resolve and validate paths
    # ------------------------------------------------------------------ #
    repeater_root = _repeater_root(settings.get("qtdr_base_path", ""))
    if repeater_root is None or not repeater_root.exists():
        log("ERROR: Could not locate the 911 QTDR REPEATER folder. "
            "Verify OneDrive is synced, or set the 911 QTDR root in "
            "Settings > Apps > 911 Batch Repeater.")
        return
    log(f"Repeater root : {repeater_root}")

    parts_lib = repeater_root / "MASTER PARTS LIBRARY"
    insp_lib = repeater_root / "MASTER INSPECTION LIBRARY"

    if not parts_lib.exists():
        log(f"ERROR: Master Parts Library not found:\n  {parts_lib}")
        return

    if not insp_lib.exists():
        log(f"ERROR: Master Inspection Library not found:\n  {insp_lib}")
        return

    # ------------------------------------------------------------------ #
    # Step 3 - Locate batch folder in QTDR
    # ------------------------------------------------------------------ #
    batch_folder = sdk.find_911_batch_folder(repeater_root, batch_number)
    if batch_folder is None:
        log(f"ERROR: Batch folder '{batch_number}' not found under:\n"
            f"  {repeater_root}\n"
            "Check that the batch folder exists inside the REPEATER directory.")
        return

    log(f"Batch folder  : {batch_folder}")
    progress_callback(5)

    if cancel_event.is_set():
        return

    # ------------------------------------------------------------------ #
    # Step 4 - Build library indexes once
    # ------------------------------------------------------------------ #
    log("Indexing Master Parts Library...")
    nc_index = _build_library_index(parts_lib, ".NC")
    log(f"  .NC files found  : {len(nc_index)}")

    log("Indexing Master Inspection Library...")
    pdf_index = _build_library_index(insp_lib, ".pdf")
    log(f"  .pdf files found : {len(pdf_index)}")

    progress_callback(10)

    if cancel_event.is_set():
        return

    # ------------------------------------------------------------------ #
    # Step 5 - Collect DYPNs per nest
    # ------------------------------------------------------------------ #
    nest_folders = sorted([d for d in batch_folder.iterdir() if d.is_dir()])

    if not nest_folders:
        log(f"No nest subfolders found in {batch_folder.name}. Has 911 Setup been run?")
        return

    log(f"Nest folders  : {len(nest_folders)} found")

    nest_dypns: dict[str, list[str]] = {}

    for nest_dir in nest_folders:
        nest_name = nest_dir.name
        excel_path = _find_nest_excel(nest_dir, batch_number, nest_name)

        if excel_path is None:
            log(f"  [{nest_name}] WARNING: No 911 Batch Excel found - skipping")
            continue

        try:
            dypns = _read_dypns(excel_path)
        except Exception as exc:
            log(f"  [{nest_name}] WARNING: Could not read Excel ({exc}) - skipping")
            continue

        if not dypns:
            log(f"  [{nest_name}] No DYPNs in NEST col K")
            continue

        log(f"  [{nest_name}] DYPNs: {', '.join(dypns)}")
        nest_dypns[nest_name] = dypns

    progress_callback(30)

    if not nest_dypns:
        log("No DYPNs found across any nest - nothing to do.")
        return

    if cancel_event.is_set():
        return

    # ------------------------------------------------------------------ #
    # Step 6 - Per nest: match, copy, edit PDFs
    # ------------------------------------------------------------------ #
    total_nc = 0
    total_pdf = 0
    total_edited = 0
    total_new = 0
    total_errors = 0

    nest_list = list(nest_dypns.items())
    for nest_idx, (nest_name, dypns) in enumerate(nest_list):
        if cancel_event.is_set():
            log("Cancelled by user.")
            return

        log(f"\n[{nest_name}] Processing {len(dypns)} DYPNs...")

        nc_out = repeater_root / batch_number / nest_name / "NC Repeats"
        insp_out = repeater_root / batch_number / nest_name / "Inspection Repeats"
        nc_out.mkdir(parents=True, exist_ok=True)
        insp_out.mkdir(parents=True, exist_ok=True)

        seen: set[str] = set()
        nest_nc = nest_pdf = nest_edited = nest_errors = 0

        for dypn in dypns:
            dypn_key = dypn.upper()
            if dypn_key in seen:
                continue
            seen.add(dypn_key)

            nc_src = nc_index.get(dypn_key)
            if nc_src is None:
                log(f"  {dypn} - not in parts library (new part, skipped)")
                total_new += 1
                continue

            # Copy NC
            nc_dest = nc_out / nc_src.name
            try:
                if nc_dest.exists():
                    log(f"  {dypn} - NC already exists, skipping")
                else:
                    shutil.copy2(nc_src, nc_dest)
                    nest_nc += 1
                    log(f"  {dypn} - NC copied")
            except Exception as exc:
                log(f"  {dypn} - ERROR copying NC: {exc}")
                nest_errors += 1

            # Copy and edit PDF
            pdf_src = pdf_index.get(dypn_key)
            if pdf_src is None:
                log(f"  {dypn} - no PDF in Inspection Library (NC only)")
                continue

            pdf_dest = insp_out / pdf_src.name
            try:
                if pdf_dest.exists():
                    log(f"  {dypn} - PDF already exists, skipping")
                else:
                    shutil.copy2(pdf_src, pdf_dest)
                    nest_pdf += 1
                    log(f"  {dypn} - PDF copied, updating batch/nest text...")
                    if _replace_batch_nest_in_pdf(pdf_dest, batch_number, nest_name, log):
                        nest_edited += 1
            except Exception as exc:
                log(f"  {dypn} - ERROR with PDF: {exc}")
                nest_errors += 1

        log(f"  [{nest_name}] NC: {nest_nc} copied | PDF: {nest_pdf} copied, {nest_edited} edited")
        total_nc += nest_nc
        total_pdf += nest_pdf
        total_edited += nest_edited
        total_errors += nest_errors

        progress_callback(30 + int((nest_idx + 1) / len(nest_list) * 65))

    # ------------------------------------------------------------------ #
    # Step 7 - Summary
    # ------------------------------------------------------------------ #
    progress_callback(100)
    log("\n" + "=" * 50)
    log("REPEATER SUMMARY")
    log("=" * 50)
    log(f"Batch          : {batch_number}")
    log(f"Nests processed: {len(nest_dypns)}")
    log(f"NC files copied: {total_nc}")
    log(f"PDFs copied    : {total_pdf}")
    log(f"PDFs edited    : {total_edited}")
    if total_new:
        log(f"New parts (no NC): {total_new}")
    if total_errors:
        log(f"Errors         : {total_errors}")
    log(f"\nOutput root    : {repeater_root / batch_number}")
    log("=" * 50)
    log("Done." if total_errors == 0 else f"Done with {total_errors} error(s).")
