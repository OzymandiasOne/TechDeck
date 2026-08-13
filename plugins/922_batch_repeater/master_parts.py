"""Master Parts List engine, shared by the 922 Batch Repeater plugin and the
one-off build scripts in tools/ (mpl_extract_batches.py / mpl_build_master.py).

The MASTER PARTS sheet in 922 MPL.xlsx is a true per-piece catalog: one row per
(PPN, DYPN) with repeat counts, batch history, alternate suffixes (the same
physical piece renamed between batches, e.g. -4 -> -4A), and a part-type
classification driven by the EB 922 H# Quote MATERIAL PRICING sheet.

Lives inside the plugin folder (NOT techdeck.core) so a plugin-folder copy to
installed machines ships it without a TechDeck release.
"""

import re
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk


MASTER_SHEET_NAME = "MASTER PARTS"

# Column order of the MASTER PARTS sheet. DYPN is the first-seen (canonical)
# spelling; LAST DYPN is the spelling used in the most recent batch - the
# merge engine needs it to align renamed pieces against the *previous*
# occurrence, and humans need it to know what the part is called today.
MASTER_HEADERS = [
    "PPN", "DYPN", "SUFFIX", "ALT SUFFIXES", "LAST DYPN", "SOURCE MATERIAL",
    "PART TYPE", "PART TYPE SOURCE", "MATERIAL TYPE (RAW)", "PART DESCRIPTION",
    "QTY (LATEST)", "TIMES MADE", "TOTAL PIECES", "FIRST BATCH", "LAST BATCH",
    "BATCHES", "LAST WO", "HOURS (LATEST)", "NOTES", "LAST UPDATED",
    # helper for the ANALYSIS sheet: 1 on the first row of each PPN, 0 after -
    # lets plain COUNTIF/SUMPRODUCT formulas count DISTINCT PPNs (TIMES MADE
    # repeats on every row of a PPN, so summing it raw would overcount).
    # Written as static values by write_master_sheet, which rewrites the whole
    # sheet on every update, so it can never go stale.
    "PPN FIRST",
]

ANALYSIS_SHEET_NAME = "ANALYSIS"

PO_SHEET_REQUIRED = ("ORDER", "PPN", "DYPN", "SOURCE MATERIAL")


# ─────────────────────────────────────────────────────────────────────────────
# PO workbook discovery (the batch's OWN "PO {n} QF-QU-09 REV C" workbook)
# ─────────────────────────────────────────────────────────────────────────────

def _own_po_regexes(batch_num: int):
    """(rev_c_re, any_rev_re) for the batch's own PO workbook filename.
    Accepts 'PO 334' and 'PO H334' styles plus copy suffixes like ' (2)'."""
    base = rf"^PO\s+H?0*{batch_num}\s+QF[\s-]?QU[\s-]?09"
    rev_c = re.compile(base + r".*REV\s*C.*\.xlsx$", re.IGNORECASE)
    any_rev = re.compile(base + r".*\.xlsx$", re.IGNORECASE)
    return rev_c, any_rev


def find_batch_po_candidates(batch_dir: Path,
                             batch_num: int) -> Tuple[list, bool]:
    """Every copy of the batch's own REV C PO workbook, in preference order:
    'Batch N - Documentation' folder copies first, then the direct files of
    each ORDER folder. Copies loose in the batch root or anywhere under
    REPEAT BATCHES are never used (they are other batches' POs riding along
    with repeat orders).

    Returns ([(path, location), ...], old_rev_seen). Some Documentation
    copies are blank 'PASTE HERE' templates (bit batches 319/367), so callers
    should try candidates until one parses to rows.
    """
    batch_dir = Path(batch_dir)
    rev_c_re, any_rev_re = _own_po_regexes(batch_num)
    old_rev_seen = False
    candidates: list = []

    def scan(folder: Path, location: str):
        nonlocal old_rev_seen
        try:
            files = [p for p in folder.iterdir() if p.is_file()]
        except OSError:
            return
        for f in files:
            if f.name.startswith("~$"):
                continue
            if rev_c_re.match(f.name):
                candidates.append((f, location))
            elif any_rev_re.match(f.name):
                old_rev_seen = True

    try:
        children = sorted(p for p in batch_dir.iterdir() if p.is_dir())
    except OSError:
        children = []

    doc_dirs = [d for d in children if "documentation" in d.name.lower()]
    for d in doc_dirs:
        scan(d, "documentation")
    for d in children:
        if d in doc_dirs or "repeat" in d.name.lower():
            continue
        scan(d, "order_folder")
    return candidates, old_rev_seen


def find_batch_po_workbook(batch_dir: Path, batch_num: int) -> Dict[str, Any]:
    """First candidate from find_batch_po_candidates as a status dict:
    {"status": "ok"|"old_rev"|"no_po_workbook", "path", "location"}."""
    candidates, old_rev_seen = find_batch_po_candidates(batch_dir, batch_num)
    if candidates:
        path, location = candidates[0]
        return {"status": "ok", "path": path, "location": location}
    return {"status": "old_rev" if old_rev_seen else "no_po_workbook",
            "path": None, "location": None}


def extract_batch(batch_dir: Path, batch_num: int, log=None,
                  max_candidates: int = 6) -> Dict[str, Any]:
    """Locate AND parse the batch's PO data, trying candidate workbook copies
    until one yields rows (a blank Documentation template falls through to the
    order-folder copies). The one entry point both the Phase-1 extraction and
    the repeater's Phase-2 MPL update use."""
    candidates, old_rev_seen = find_batch_po_candidates(batch_dir, batch_num)
    if not candidates:
        return {"status": "old_rev" if old_rev_seen else "no_po_workbook",
                "path": None, "location": None, "rows": [], "warnings": []}
    failures: List[str] = []
    for path, location in candidates[:max_candidates]:
        try:
            rows, warnings = parse_po_sheet(path, log=log)
        except Exception as exc:
            failures.append(f"{path.name} [{location}]: "
                            f"{type(exc).__name__}: {exc}")
            continue
        if rows:
            if failures:
                warnings = warnings + [f"earlier copy unusable - {f}"
                                       for f in failures]
            return {"status": "ok", "path": path, "location": location,
                    "rows": rows, "warnings": warnings}
        failures.append(f"{path.name} [{location}]: parsed to zero rows")
    return {"status": "error", "path": None, "location": None, "rows": [],
            "warnings": [], "error": "; ".join(failures)}


# ─────────────────────────────────────────────────────────────────────────────
# PO sheet parsing
# ─────────────────────────────────────────────────────────────────────────────

def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _norm_hdr(v) -> str:
    """Header text with ALL whitespace removed, uppercased - some REV C copies
    lost the spaces in their headers ('SOURCEMATERIAL', 'P.OITEM'; bit batch
    407), so name matching must be whitespace-insensitive."""
    return re.sub(r"\s+", "", str(v)).upper()


# normalized header -> canonical key ('ORDER#' is the BOM-sheet spelling)
_HDR_KEYS = {"ORDER": "ORDER", "ORDER#": "ORDER", "PPN": "PPN", "DYPN": "DYPN",
             "SOURCEMATERIAL": "SOURCE MATERIAL", "QTY": "QTY",
             "NOTES": "NOTES", "HOURS": "HOURS"}


def _scan_po_headers(ws, max_scan: int = 15):
    """(header_row, {canonical key: 1-based col}) for the first row holding all
    of PO_SHEET_REQUIRED (whitespace-insensitive), or (None, {})."""
    limit = min(ws.max_row or 0, max_scan)
    for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=limit,
                                           values_only=True), 1):
        found: Dict[str, int] = {}
        for c_i, v in enumerate(row, 1):
            if v is None:
                continue
            key = _HDR_KEYS.get(_norm_hdr(v))
            if key and key not in found:
                found[key] = c_i
        if all(k in found for k in PO_SHEET_REQUIRED):
            return r_i, found
    return None, {}


def parse_po_sheet(path: Path, log=None) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Read the 'PO' sheet of a QF-QU-09 REV C workbook into row dicts:
    {order, ppn, dypn, qty, source_material, notes, hours}.

    Header row/columns are scanned by NAME (Hard Rules 1+2), tolerating the
    space-less header variants. The PO-named sheet is tried first, then any
    sheet whose headers match. ORDER is forward-filled (it repeats on every
    row of an order group, but be defensive). Rows end after 5 consecutive
    rows with no DYPN/PPN/ORDER.
    """
    warnings: List[str] = []
    wb = sdk.load_workbook_resilient(path, log=log, read_only=True, data_only=True)
    try:
        # Visible sheets only: REV C copies can carry hidden staging/"PASTE
        # HERE" sheets whose headers also match - the scan must not read one.
        names = sorted(sdk.visible_sheetnames(wb),
                       key=lambda s: 0 if s.strip().casefold() == "po" else 1)
        ws, hdr, hmap = None, None, {}
        for s in names:
            r, m = _scan_po_headers(wb[s])
            if r is not None:
                ws, hdr, hmap = wb[s], r, m
                break
        if ws is None or hdr is None:
            raise ValueError(
                f"no sheet with headers {PO_SHEET_REQUIRED} in '{Path(path).name}'")

        c_order, c_ppn, c_dypn = hmap["ORDER"], hmap["PPN"], hmap["DYPN"]
        c_sm = hmap["SOURCE MATERIAL"]
        c_qty = hmap.get("QTY")
        c_notes = hmap.get("NOTES")
        c_hours = hmap.get("HOURS")

        rows: List[Dict[str, Any]] = []
        blanks = 0
        last_order = None
        for r in ws.iter_rows(min_row=hdr + 1, values_only=False):
            def cell(col):
                return r[col - 1].value if col and col <= len(r) else None

            order = _clean(cell(c_order))
            ppn = _clean(cell(c_ppn))
            dypn = _clean(cell(c_dypn))
            if not (order or ppn or dypn):
                blanks += 1
                if blanks >= 5:
                    break
                continue
            blanks = 0
            if not dypn:
                continue  # spacer / annotation row
            if order:
                last_order = order
            elif last_order:
                order = last_order
            if not order or not ppn:
                warnings.append(f"row with DYPN {dypn} missing "
                                f"{'ORDER' if not order else 'PPN'} - skipped")
                continue

            qty_raw = cell(c_qty)
            try:
                qty = int(float(qty_raw)) if qty_raw is not None else 1
            except (TypeError, ValueError):
                warnings.append(f"non-numeric QTY {qty_raw!r} on {dypn} - using 1")
                qty = 1
            hours_raw = cell(c_hours)
            try:
                hours = float(hours_raw) if hours_raw is not None else None
            except (TypeError, ValueError):
                hours = None

            rows.append({
                "order": order, "ppn": ppn, "dypn": dypn, "qty": max(qty, 1),
                "source_material": _clean(cell(c_sm)),
                "notes": _clean(cell(c_notes)), "hours": hours,
            })
        return rows, warnings
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# Suffix helpers
# ─────────────────────────────────────────────────────────────────────────────

def suffix_of(dypn: str, ppn: str) -> str:
    """The DYPN's piece suffix = what follows 'PPN-' (e.g. '4A'). Falls back to
    the text after the last dash if the DYPN doesn't start with the PPN."""
    d, p = dypn.strip(), ppn.strip()
    if d.lower().startswith(p.lower() + "-"):
        return d[len(p) + 1:]
    return d.rsplit("-", 1)[-1] if "-" in d else d


_NAT_RE = re.compile(r"(\d+)")


def natural_suffix_key(suffix: str):
    """Natural sort key: '4A' -> ('', 4, 'A'), '10' before '9A' sorts 9A first.
    Used to pair renamed pieces within a source-material group in order."""
    parts = _NAT_RE.split(suffix.strip().upper())
    return tuple(int(p) if p.isdigit() else p for p in parts)


# ─────────────────────────────────────────────────────────────────────────────
# Part typing (EB 922 H# Quote.xlsx -> MATERIAL PRICING)
# ─────────────────────────────────────────────────────────────────────────────

# Display order for composite types: a multi-material piece's families join in
# THIS order (so plate+rod is always 'Mount Plate / Rod', never the reverse).
PART_TYPE_FAMILIES = (
    "Assembly", "Clevis", "Mount Plate", "Lug", "Flat Bar", "Tube",
    "Rod", "Angle Iron", "Square Bar", "Round Bar",
)

# Ordered: first hit wins. 'Flat Bar Assembly' must land on Assembly, and
# 'Bar Round' on Round Bar before the bare 'rod'/'bar' words can misfire.
_TYPE_RULES = [
    (re.compile(r"assy|assembly", re.I), "Assembly"),
    (re.compile(r"clevis", re.I), "Clevis"),
    (re.compile(r"tube", re.I), "Tube"),
    (re.compile(r"flat\s*bar", re.I), "Flat Bar"),
    (re.compile(r"angle", re.I), "Angle Iron"),
    (re.compile(r"square\s*bar", re.I), "Square Bar"),
    (re.compile(r"bar\s*round|round\s*bar|round", re.I), "Round Bar"),
    (re.compile(r"rod", re.I), "Rod"),
    (re.compile(r"lug", re.I), "Lug"),
    (re.compile(r"plate", re.I), "Mount Plate"),
]


def normalize_material_type(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    for rx, family in _TYPE_RULES:
        if rx.search(raw):
            return family
    return None


def _is_a_suffix(source_material: Optional[str]) -> bool:
    """True when the source material's suffix starts with 'A'
    (e.g. 262028653-A51) - user rule: that's a clevis assembly."""
    if not source_material or "-" not in source_material:
        return False
    return source_material.split("-", 1)[1].strip().upper().startswith("A")


def load_pricing_map(quote_path: Path, log=None) -> Dict[str, Dict[str, Any]]:
    """{SOURCE MATERIAL: {raw, description, family, source}} from the quote
    workbook's MATERIAL PRICING sheet.

    Rows with a blank Material Type are resolved in place: A-suffix source
    materials become Clevis (user rule), everything else inherits the nearest
    typed neighbor above/below in the sheet (user rule: 'infer what it is from
    the other parts around it'), tagged source='inferred'.

    The quote workbook is a live document someone often has open in Excel, so
    it's read via a TEMP COPY (a locked file blocks a direct open but not a
    copy) - the copy-to-temp pattern, same as 911 SSPO's forecast read.
    """
    import os
    import tempfile
    tmp = Path(tempfile.gettempdir()) / f"techdeck_quote_{os.getpid()}.xlsx"
    sdk.copy_resilient(quote_path, tmp, log=log)
    wb = sdk.load_workbook_resilient(tmp, log=log, read_only=True,
                                     data_only=True)
    try:
        ws = None
        for s in wb.sheetnames:
            if s.strip().casefold() == "material pricing":
                ws = wb[s]
                break
        if ws is None:
            raise ValueError("MATERIAL PRICING sheet not found in quote workbook")
        hdr, hmap = sdk.find_header_row(ws, ("SOURCE MATERIAL", "MATERIAL TYPE"))
        if hdr is None:
            raise ValueError("MATERIAL PRICING header row not found")
        c_sm = hmap["SOURCE MATERIAL"]
        c_type = hmap["MATERIAL TYPE"]
        c_desc = hmap.get("PART DESCRIPTION")

        ordered: List[Tuple[str, Optional[str], Optional[str]]] = []
        for r in ws.iter_rows(min_row=hdr + 1, values_only=True):
            def val(col):
                return r[col - 1] if col and col <= len(r) else None
            sm = _clean(val(c_sm))
            if not sm:
                continue
            ordered.append((sm, _clean(val(c_type)),
                            _clean(val(c_desc)) if c_desc else None))
    finally:
        wb.close()
        try:
            tmp.unlink()
        except OSError:
            pass

    families = [normalize_material_type(raw) for _, raw, _ in ordered]
    out: Dict[str, Dict[str, Any]] = {}
    for i, (sm, raw, desc) in enumerate(ordered):
        family, source = families[i], "explicit"
        if family is None:
            if _is_a_suffix(sm):
                family, source = "Clevis", "A-suffix"
            else:
                above = next((families[j] for j in range(i - 1, -1, -1)
                              if families[j]), None)
                below = next((families[j] for j in range(i + 1, len(families))
                              if families[j]), None)
                family = above or below
                source = "inferred" if family else "unknown"
        out.setdefault(sm, {"raw": raw, "description": desc,
                            "family": family or "Unknown", "source": source})
    return out


def _classify_single(source_material: Optional[str],
                     pricing: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """Part-type answer for ONE source-material serial.

    Precedence (user-confirmed): the pricing sheet's explicit/neighbor-resolved
    entry wins; then the SDK tube-serial sets (bare serials like 218019941
    aren't in the pricing sheet); then the A-suffix->Clevis rule; else Unknown.
    """
    if source_material:
        hit = pricing.get(source_material)
        if hit:
            return {"family": hit["family"], "source": hit["source"],
                    "raw": hit["raw"], "description": hit["description"]}
        base = source_material.split("-", 1)[0].strip()
        if source_material in sdk.ALL_TUBE_MATERIALS or base in sdk.ALL_TUBE_MATERIALS:
            return {"family": "Tube", "source": "tube-serial",
                    "raw": None, "description": None}
        if _is_a_suffix(source_material):
            return {"family": "Clevis", "source": "A-suffix",
                    "raw": None, "description": None}
    return {"family": "Unknown", "source": "unknown", "raw": None,
            "description": None}


def classify_part(source_material: Optional[str],
                  pricing: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """Part-type answer for a piece's source material, which may be a
    COMPOSITE '; '-joined signature (one PO row per constituent material,
    e.g. a plate part cut from 262028653-11 AND -12). Families are combined:
    all-same -> that family, mixed -> 'A / B'."""
    sms = [s.strip() for s in (source_material or "").split(";") if s.strip()]
    if len(sms) <= 1:
        return _classify_single(sms[0] if sms else None, pricing)
    infos = [_classify_single(s, pricing) for s in sms]
    families = []
    for i in infos:
        if i["family"] not in families:
            families.append(i["family"])
    known = sorted(
        (f for f in families if f != "Unknown"),
        key=lambda f: (PART_TYPE_FAMILIES.index(f)
                       if f in PART_TYPE_FAMILIES else len(PART_TYPE_FAMILIES), f))
    best = next((i for i in infos if i["family"] != "Unknown"), infos[0])
    return {
        "family": " / ".join(known) if known else "Unknown",
        "source": best["source"],
        "raw": "; ".join(dict.fromkeys(i["raw"] for i in infos if i["raw"]))
               or None,
        "description": "; ".join(
            dict.fromkeys(i["description"] for i in infos if i["description"]))
               or None,
    }


def group_pieces(rows: List[Dict[str, Any]],
                 warnings: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    """One piece per (DYPN, source material) PO row. The same DYPN on two rows
    with different materials is TWO SEPARATE PARTS that happen to share a name
    (plate + rod pairs are common; user-confirmed 2026-07-20) - each gets its
    own catalog row. A repeated (DYPN, material) pair is a true duplicate and
    sums qty with a warning."""
    pieces: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for r in rows:
        sm = (r.get("source_material") or "").strip() or None
        k = (r["dypn"].upper(), (sm or "").upper())
        p = pieces.get(k)
        if p is None:
            pieces[k] = {"dypn": r["dypn"], "source_material": sm,
                         "qty": r["qty"], "notes": r.get("notes"),
                         "hours": r.get("hours")}
        else:
            p["qty"] += r["qty"]
            if warnings is not None:
                warnings.append(f"duplicate row {r['dypn']} / "
                                f"{sm or '(no SM)'} (qty summed)")
            if r.get("notes") and not p.get("notes"):
                p["notes"] = r["notes"]
            if r.get("hours") is not None and p.get("hours") is None:
                p["hours"] = r["hours"]
    return list(pieces.values())


# ─────────────────────────────────────────────────────────────────────────────
# The merge engine (repeat counting + alternate-suffix alignment)
# ─────────────────────────────────────────────────────────────────────────────

def _rk(ppn: str, dypn: str, sm: Optional[str]) -> Tuple[str, str, str]:
    """Catalog row key. A row = one (PPN, DYPN, source material) - the same
    DYPN with two materials is two separate parts sharing a name."""
    return (ppn.upper(), dypn.upper(), (sm or "").upper())


class Catalog:
    """In-memory MASTER PARTS catalog.

    rows: {(ppn_u, canonical_dypn_u, sm_u): row dict} - one row per part,
    where a part is a (DYPN, source material) line (same-name parts with
    different materials each get their own row; user-confirmed 2026-07-20).
    _ppn: {ppn_upper: {"batches": set[int], "prev": [(dypn, sm, rowkey)],
                       "dypn_map": {(dypn_upper, sm_upper): rowkey}}}

    Feed occurrences (one WO's pieces for one PPN) in batch order via
    merge_occurrence(); the same code path serves the Phase 1 bulk build and
    the Phase 2 per-run repeater update.
    """

    def __init__(self):
        self.rows: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
        self._ppn: Dict[str, Dict[str, Any]] = {}
        self.warnings: List[str] = []

    # -- row plumbing --------------------------------------------------------

    def _new_row(self, ppn: str, piece: Dict[str, Any], batch: int, wo: str,
                 today: str) -> Tuple[str, str, str]:
        key = _rk(ppn, piece["dypn"], piece.get("source_material"))
        if key in self.rows:  # same part reappearing as a "new" row - just update
            self._absorb(key, piece, batch, wo, today)
            return key
        self.rows[key] = {
            "ppn": ppn, "dypn": piece["dypn"],
            "suffix": suffix_of(piece["dypn"], ppn),
            "alt_suffixes": [], "last_dypn": piece["dypn"],
            "source_material": piece.get("source_material"),
            "qty_latest": piece["qty"], "times_made": 1,
            "total_pieces": piece["qty"], "first_batch": batch,
            "last_batch": batch, "batches": [batch], "last_wo": wo,
            "hours_latest": piece.get("hours"), "notes": piece.get("notes"),
            "last_updated": today,
        }
        return key

    def _absorb(self, rowkey: Tuple[str, str, str], piece: Dict[str, Any],
                batch: int, wo: str, today: str) -> None:
        row = self.rows[rowkey]
        row["qty_latest"] = piece["qty"]
        row["total_pieces"] += piece["qty"]
        row["last_batch"] = batch
        if not row["batches"] or row["batches"][-1] != batch:
            row["batches"].append(batch)
        row["last_wo"] = wo
        row["last_dypn"] = piece["dypn"]
        if piece.get("hours") is not None:
            row["hours_latest"] = piece["hours"]
        if piece.get("notes"):
            row["notes"] = piece["notes"]
        if piece.get("source_material") and not row.get("source_material"):
            row["source_material"] = piece["source_material"]
        row["last_updated"] = today
        suf = suffix_of(piece["dypn"], row["ppn"])
        if (suf.upper() != row["suffix"].upper()
                and suf.upper() not in (a.upper() for a in row["alt_suffixes"])):
            row["alt_suffixes"].append(suf)

    # -- the occurrence merge ------------------------------------------------

    def merge_occurrence(self, ppn: str, batch: int, wo: str,
                         rows: List[Dict[str, Any]], today: str) -> None:
        """Fold one occurrence (one WO's PO rows for one PPN in one batch) in.

        Rows are first grouped into pieces - one per (DYPN, source material)
        line; the same DYPN with two materials is two separate parts.

        Alignment rule (user-specified): a piece renamed between batches can
        only be re-identified through its SOURCE MATERIAL group - and only when
        the group has the SAME piece count in the previous occurrence and this
        one (an added piece of the same material makes the mapping ambiguous,
        so those become new rows instead). Within an aligned group, pieces pair
        in natural suffix order (5A,5B vs 5,6 -> 5A~5, 5B~6).
        """
        pu = ppn.upper()
        state = self._ppn.setdefault(
            pu, {"batches": set(), "prev": [], "dypn_map": {}})

        warn: List[str] = []
        new_pieces = group_pieces(rows, warn)
        self.warnings.extend(f"batch {batch} {wo}: {w}" for w in warn)

        exact, remainder = [], []
        for p in new_pieces:
            smu = (p.get("source_material") or "").upper()
            rk = state["dypn_map"].get((p["dypn"].upper(), smu))
            if rk is None:
                direct = _rk(ppn, p["dypn"], p.get("source_material"))
                if direct in self.rows:
                    rk = direct
            (exact if rk else remainder).append((p, rk))

        matched_rowkeys = set()
        this_occ: List[Tuple[str, Optional[str], Tuple[str, str, str]]] = []
        for p, rk in exact:
            self._absorb(rk, p, batch, wo, today)
            state["dypn_map"][(p["dypn"].upper(),
                              (p.get("source_material") or "").upper())] = rk
            matched_rowkeys.add(rk)
            this_occ.append((p["dypn"], p.get("source_material"), rk))

        if remainder:
            prev_rem = [t for t in state["prev"] if t[2] not in matched_rowkeys]
            prev_by_sm: Dict[Optional[str], list] = {}
            for t in prev_rem:
                prev_by_sm.setdefault(t[1], []).append(t)
            new_by_sm: Dict[Optional[str], list] = {}
            for p, _ in remainder:
                new_by_sm.setdefault(p.get("source_material"), []).append(p)

            for sm, group in new_by_sm.items():
                prev_group = prev_by_sm.get(sm, [])
                if sm is not None and prev_group and len(prev_group) == len(group):
                    # unambiguous rename group: pair in natural suffix order
                    prev_sorted = sorted(
                        prev_group,
                        key=lambda t: natural_suffix_key(suffix_of(t[0], ppn)))
                    new_sorted = sorted(
                        group,
                        key=lambda p: natural_suffix_key(
                            suffix_of(p["dypn"], ppn)))
                    for (pd, _psm, prk), np in zip(prev_sorted, new_sorted):
                        self._absorb(prk, np, batch, wo, today)
                        state["dypn_map"][(np["dypn"].upper(),
                                           (np.get("source_material") or "")
                                           .upper())] = prk
                        this_occ.append((np["dypn"],
                                         np.get("source_material"), prk))
                else:
                    for np in group:
                        rk = self._new_row(ppn, np, batch, wo, today)
                        state["dypn_map"][(np["dypn"].upper(),
                                           (np.get("source_material") or "")
                                           .upper())] = rk
                        this_occ.append((np["dypn"],
                                         np.get("source_material"), rk))

        state["prev"] = this_occ
        state["batches"].add(batch)

    def has_batch(self, ppn: str, batch: int) -> bool:
        """True when this batch is already recorded for the PPN - the
        repeater's idempotency guard against double-counting a re-run."""
        return batch in self._ppn.get(ppn.upper(), {}).get("batches", set())

    def finalize_times_made(self) -> None:
        """TIMES MADE = distinct batches the PPN ran in (PPN-level, same value
        on every row of the PPN - the user's definition of a repeat)."""
        for key, row in self.rows.items():
            row["times_made"] = len(
                self._ppn.get(key[0], {}).get("batches", ())) or 1

    def apply_part_types(self, pricing: Dict[str, Dict[str, Any]],
                         only_untyped: bool = False) -> None:
        """only_untyped=True leaves rows that already carry a real type alone -
        used when the pricing sheet couldn't be read, so existing types are
        never clobbered to Unknown."""
        for row in self.rows.values():
            if only_untyped and row.get("part_type") not in (None, "", "Unknown"):
                continue
            info = classify_part(row.get("source_material"), pricing)
            row["part_type"] = info["family"]
            row["part_type_source"] = info["source"]
            row["material_type_raw"] = info["raw"]
            row["part_description"] = info["description"]

    def sorted_rows(self) -> List[Dict[str, Any]]:
        # PPN, then suffix, then material - same-name parts sit adjacent
        return sorted(
            self.rows.values(),
            key=lambda r: (r["ppn"].upper(), natural_suffix_key(r["suffix"]),
                           r.get("source_material") or ""))


# ─────────────────────────────────────────────────────────────────────────────
# MASTER PARTS sheet I/O (openpyxl, normal mode)
# ─────────────────────────────────────────────────────────────────────────────

def _row_to_cells(row: Dict[str, Any]) -> list:
    return [
        row["ppn"], row["dypn"], row["suffix"],
        "; ".join(row["alt_suffixes"]) or None, row["last_dypn"],
        row.get("source_material"), row.get("part_type"),
        row.get("part_type_source"), row.get("material_type_raw"),
        row.get("part_description"), row["qty_latest"], row["times_made"],
        row["total_pieces"], row["first_batch"], row["last_batch"],
        ", ".join(str(b) for b in row["batches"]), row["last_wo"],
        row.get("hours_latest"), row.get("notes"), row["last_updated"],
    ]


def write_master_sheet(wb, rows: List[Dict[str, Any]],
                       after_sheet: str = "PO 321+"):
    """(Re)write the MASTER PARTS sheet in an open workbook. An existing sheet
    is cleared and refilled IN PLACE - never deleted - because the ANALYSIS
    sheet's formulas reference 'MASTER PARTS'!... and a delete/recreate would
    turn every one of them into #REF!. Every other sheet is untouched. Caller
    saves."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if MASTER_SHEET_NAME in wb.sheetnames:
        ws = wb[MASTER_SHEET_NAME]
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
    else:
        try:
            idx = wb.sheetnames.index(after_sheet) + 1
        except ValueError:
            idx = len(wb.sheetnames)
        ws = wb.create_sheet(MASTER_SHEET_NAME, idx)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E5F")
    for c, name in enumerate(MASTER_HEADERS, 1):
        cell = ws.cell(1, c, name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    seen_ppns: set = set()
    for r, row in enumerate(rows, 2):
        cells = _row_to_cells(row)
        pu = row["ppn"].upper()
        cells.append(0 if pu in seen_ppns else 1)  # PPN FIRST helper
        seen_ppns.add(pu)
        for c, v in enumerate(cells, 1):
            ws.cell(r, c, v)

    widths = {"A": 18, "B": 22, "C": 8, "D": 14, "E": 22, "F": 16, "G": 12,
              "H": 12, "I": 20, "J": 26, "K": 8, "L": 8, "M": 8, "N": 8,
              "O": 8, "P": 24, "Q": 12, "R": 9, "S": 32, "T": 12, "U": 6}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(MASTER_HEADERS))}{max(len(rows) + 1, 2)}")
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# ANALYSIS sheet (formulas + native charts; auto-recalcs on every MPL update)
# ─────────────────────────────────────────────────────────────────────────────

# Palette (light-surface validated). Navy matches the MASTER PARTS header so
# the workbook reads as one designed document; the 12 slice hues are the
# fixed-order categorical palette + 3 blue-ramp extras + gray for 'Other'.
_VIZ_BLUE = "2A78D6"
_VIZ_BLUE_LIGHT = "B7D3F6"
_VIZ_NAVY = "1F4E5F"
_TILE_FILL = "EEF4FB"
_TILE_EDGE = "C9D9EE"
_ROW_BAND = "F5F7FA"
_INK_SECONDARY = "52514E"
_VIZ_SLICES = ["2A78D6", "1BAF7A", "EDA100", "008300", "4A3AA7", "E34948",
               "E87BA4", "EB6834", "86B6EF", "1C5CAB", "104281", "898781"]

# Fixed part-type row order for the type table (roughly by expected volume;
# 'Other' catches any family not listed so the total always reconciles).
_ANALYSIS_TYPES = ["Rod", "Mount Plate", "Tube", "Flat Bar", "Lug", "Unknown",
                   "Clevis", "Assembly", "Angle Iron", "Square Bar",
                   "Round Bar"]


def write_analysis_sheet(wb, after_sheet: str = MASTER_SHEET_NAME):
    """(Re)create the ANALYSIS sheet as a dashboard: a title band, KPI stat
    tiles, two pies, and styled breakdown tables - all plain Excel FORMULAS
    over 'MASTER PARTS' (via its PPN FIRST helper column) and 'PO 321+', so
    everything recalcs the moment those sheets change; no macros, no manual
    refresh. Native charts reference the formula cells and follow along.
    Safe to delete and rebuild (nothing references ANALYSIS)."""
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if ANALYSIS_SHEET_NAME in wb.sheetnames:
        del wb[ANALYSIS_SHEET_NAME]
    try:
        idx = wb.sheetnames.index(after_sheet) + 1
    except ValueError:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet(ANALYSIS_SHEET_NAME, idx)
    ws.sheet_view.showGridLines = False

    M = f"'{MASTER_SHEET_NAME}'"
    N = 60000  # formula row cap (~3x headroom over today's row count)
    # shared sub-expressions (kept as strings so every tile is self-contained)
    DISTINCT = f"SUM({M}!$U$2:$U${N})"
    INSTANCES = f"SUMPRODUCT(({M}!$U$2:$U${N})*({M}!$L$2:$L${N}))"
    REPEAT_PPNS = f"SUMPRODUCT(({M}!$U$2:$U${N})*({M}!$L$2:$L${N}>1))"
    UNIQUE = f"COUNTA({M}!$A$2:$A${N})"

    navy = _VIZ_NAVY
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tile_fill = PatternFill("solid", fgColor=_TILE_FILL)
    band_fill = PatternFill("solid", fgColor=_ROW_BAND)
    hdr_fill = PatternFill("solid", fgColor=navy)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    small_hdr = Font(bold=True, size=9, color=_INK_SECONDARY)
    edge = Side(style="thin", color=_TILE_EDGE)

    # --- column grid: five 3-wide tile bays with slim spacers --------------
    tile_starts = [1, 5, 9, 13, 17]  # A, E, I, M, Q
    for c0 in tile_starts:
        for c in range(c0, c0 + 3):
            ws.column_dimensions[get_column_letter(c)].width = 10.2
    for c in (4, 8, 12, 16, 20):
        ws.column_dimensions[get_column_letter(c)].width = 1.6

    # --- title band --------------------------------------------------------
    for r in (1, 2):
        for c in range(1, 20):
            ws.cell(r, c).fill = hdr_fill
    ws.merge_cells("A1:S2")
    t = ws["A1"]
    t.value = "922 MASTER PARTS  —  ANALYSIS"
    t.font = Font(bold=True, size=20, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A3:S3")
    sub = ws["A3"]
    sub.value = ("Live view of the MASTER PARTS catalog and PO 321+ matrix "
                 "— every figure recalculates automatically when they "
                 "change.")
    sub.font = Font(italic=True, size=9, color=_INK_SECONDARY)
    sub.alignment = Alignment(horizontal="left", indent=1)

    # --- KPI tiles ---------------------------------------------------------
    def tile(row0, col0, caption, formula, fmt, accent=navy):
        for r in range(row0, row0 + 3):
            for c in range(col0, col0 + 3):
                cell = ws.cell(r, c)
                cell.fill = tile_fill
                b = {}
                if r == row0:
                    b["top"] = edge
                if r == row0 + 2:
                    b["bottom"] = edge
                if c == col0:
                    b["left"] = edge
                if c == col0 + 2:
                    b["right"] = edge
                if b:
                    cell.border = Border(**b)
        ws.merge_cells(start_row=row0, start_column=col0,
                       end_row=row0 + 1, end_column=col0 + 2)
        ws.merge_cells(start_row=row0 + 2, start_column=col0,
                       end_row=row0 + 2, end_column=col0 + 2)
        v = ws.cell(row0, col0, formula)
        v.font = Font(bold=True, size=19, color=accent)
        v.alignment = center
        v.number_format = fmt
        cap = ws.cell(row0 + 2, col0, caption)
        cap.font = Font(bold=True, size=8, color=_INK_SECONDARY)
        cap.alignment = center

    for r, h in ((5, 20), (6, 20), (7, 24), (9, 20), (10, 20), (11, 24)):
        ws.row_dimensions[r].height = h

    tiles_top = [
        ("UNIQUE PARTS CATALOGUED", f"={UNIQUE}", "#,##0", navy),
        ("DISTINCT PART NUMBERS (PPNs)", f"={DISTINCT}", "#,##0", navy),
        ("REPEAT PPNs (MADE IN 2+ BATCHES)", f"={REPEAT_PPNS}", "#,##0",
         _VIZ_BLUE),
        ("REPEAT SHARE OF ALL ORDERS",
         f"=IFERROR(({INSTANCES}-{DISTINCT})/{INSTANCES},0)", "0.0%",
         _VIZ_BLUE),
        ("REPEAT ORDERS PULLED (RE-MAKES)", f"={INSTANCES}-{DISTINCT}",
         "#,##0", _VIZ_BLUE),
    ]
    tiles_bottom = [
        ("AVG TIMES EACH PPN IS MADE",
         f"=IFERROR({INSTANCES}/{DISTINCT},0)", "0.00", navy),
        ("PARTS WITH ALTERNATE NAMES",
         '=COUNTIF(' + M + '!$D$2:$D$' + str(N) + ',"?*")', "#,##0", navy),
        ("TOTAL PIECES MADE", f"=SUM({M}!$M$2:$M${N})", "#,##0", navy),
        ("BATCHES IN THE PO 321+ MATRIX",
         '=COUNTIF(\'PO 321+\'!$1:$5,"PO *")', "#,##0", navy),
        ("ORDERS LOGGED IN THE MATRIX",
         '=SUMPRODUCT(--(\'PO 321+\'!$A$4:$ZZ$600<>""))', "#,##0", navy),
    ]
    for (cap, formula, fmt, accent), c0 in zip(tiles_top, tile_starts):
        tile(5, c0, cap, formula, fmt, accent)
    for (cap, formula, fmt, accent), c0 in zip(tiles_bottom, tile_starts):
        tile(9, c0, cap, formula, fmt, accent)

    # --- breakdown tables (also feed the pies) -----------------------------
    def section(row, col0, col1, label):
        ws.merge_cells(start_row=row, start_column=col0,
                       end_row=row, end_column=col1)
        for c in range(col0, col1 + 1):
            ws.cell(row, c).fill = hdr_fill
        s = ws.cell(row, col0, label)
        s.font = hdr_font
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    T = 40   # tables start row (the pies float over rows 13-38)
    section(T, 1, 3, "PART TYPE BREAKDOWN")
    for c, h in ((1, "PART TYPE"), (2, "PARTS"), (3, "% OF PARTS")):
        ws.cell(T + 1, c, h).font = small_hdr
    t0 = T + 2                          # first data row (42)
    for i, fam in enumerate(_ANALYSIS_TYPES):
        r = t0 + i
        ws.cell(r, 1, fam)
        ws.cell(r, 2, f"=COUNTIF({M}!$G$2:$G${N},A{r})").number_format = "#,##0"
        ws.cell(r, 3, f"=IFERROR(B{r}/{UNIQUE},0)").number_format = "0.0%"
    r_other = t0 + len(_ANALYSIS_TYPES)
    ws.cell(r_other, 1, "Other")
    ws.cell(r_other, 2,
            f"=MAX(0,{UNIQUE}-SUM(B{t0}:B{r_other - 1}))").number_format = "#,##0"
    ws.cell(r_other, 3,
            f"=IFERROR(B{r_other}/{UNIQUE},0)").number_format = "0.0%"
    for r in range(t0, r_other + 1):
        if (r - t0) % 2 == 1:
            for c in (1, 2, 3):
                ws.cell(r, c).fill = band_fill

    section(T, 5, 7, "TIMES MADE (DISTINCT PPNs)")
    ws.cell(T + 1, 5, "TIMES MADE").font = small_hdr
    ws.cell(T + 1, 6, "PPNS").font = small_hdr
    dist = [("1x", "=1"), ("2x", "=2"), ("3x", "=3"), ("4x", "=4"),
            ("5x +", ">=5")]
    for i, (label, cond) in enumerate(dist):
        r = t0 + i
        ws.cell(r, 5, label)
        expr = (f"({M}!$L$2:$L${N}>=5)" if cond == ">=5"
                else f"({M}!$L$2:$L${N}={cond[1:]})")
        ws.cell(r, 6,
                f"=SUMPRODUCT(({M}!$U$2:$U${N})*{expr})").number_format = "#,##0"
        if i % 2 == 1:
            for c in (5, 6, 7):
                ws.cell(r, c).fill = band_fill

    section(T, 9, 11, "PPN MIX")
    ws.cell(T + 1, 9, "GROUP").font = small_hdr
    ws.cell(T + 1, 10, "PPNS").font = small_hdr
    ws.cell(t0, 9, "Repeat")
    ws.cell(t0, 10, f"={REPEAT_PPNS}").number_format = "#,##0"
    ws.cell(t0 + 1, 9, "First-time")
    ws.cell(t0 + 1, 10, f"={DISTINCT}-J{t0}").number_format = "#,##0"

    # --- charts ------------------------------------------------------------
    # Every series gets an explicit name - an unnamed series surfaces as
    # Excel's meaningless "Series 1" default in tooltips/labels.
    type_pie = PieChart()
    type_pie.title = "Parts by type"
    type_pie.height, type_pie.width = 12, 17
    data = Reference(ws, min_col=2, min_row=t0, max_row=r_other)
    cats = Reference(ws, min_col=1, min_row=t0, max_row=r_other)
    type_pie.add_data(data, titles_from_data=False)
    type_pie.set_categories(cats)
    ser = type_pie.series[0]
    ser.tx = SeriesLabel(v="Parts")
    ser.data_points = [
        DataPoint(idx=i, spPr=GraphicalProperties(solidFill=hx))
        for i, hx in enumerate(_VIZ_SLICES)]
    # 12 slices, several near zero: identity via the legend (labels on tiny
    # slices collide); exact counts live in the table below it.
    if type_pie.legend is not None:
        type_pie.legend.position = "r"
    ws.add_chart(type_pie, "A13")

    mix_pie = PieChart()
    mix_pie.title = "Repeat vs first-time PPNs"
    mix_pie.legend = None
    mix_pie.height, mix_pie.width = 12, 13
    data = Reference(ws, min_col=10, min_row=t0, max_row=t0 + 1)
    cats = Reference(ws, min_col=9, min_row=t0, max_row=t0 + 1)
    mix_pie.add_data(data, titles_from_data=False)
    mix_pie.set_categories(cats)
    ser = mix_pie.series[0]
    ser.tx = SeriesLabel(v="PPNs")
    ser.data_points = [
        DataPoint(idx=0, spPr=GraphicalProperties(solidFill=_VIZ_BLUE)),
        DataPoint(idx=1, spPr=GraphicalProperties(solidFill=_VIZ_BLUE_LIGHT)),
    ]
    mix_pie.dataLabels = DataLabelList(showCatName=True, showPercent=True,
                                       showVal=False)
    ws.add_chart(mix_pie, "M13")
    return ws


def read_master_sheet(ws) -> Catalog:
    """Rebuild a Catalog from an existing MASTER PARTS sheet so the repeater
    can merge a new batch into it (Phase 2).

    prev-occurrence pieces are reconstructed from the rows whose LAST BATCH is
    the PPN's newest batch, using LAST DYPN as the spelling seen there - which
    is exactly what the alignment step compares against.
    """
    hdr, hmap = sdk.find_header_row(ws, ("PPN", "DYPN", "SUFFIX", "BATCHES"))
    if hdr is None:
        raise ValueError(f"'{MASTER_SHEET_NAME}' sheet has no header row")

    def col(name):
        c = hmap.get(name)
        if c is None:
            raise ValueError(f"'{MASTER_SHEET_NAME}' is missing column '{name}'")
        return c

    cat = Catalog()
    for r in ws.iter_rows(min_row=hdr + 1, values_only=False):
        def val(name, required=False):
            c = hmap.get(name)
            v = r[c - 1].value if c and c <= len(r) else None
            return _clean(v) if isinstance(v, str) else v
        ppn, dypn = val("PPN"), val("DYPN")
        if not ppn or not dypn:
            continue
        alt_raw = val("ALT SUFFIXES")
        alts = [a.strip() for a in str(alt_raw).split(";") if a.strip()] \
            if alt_raw else []
        batches_raw = val("BATCHES")
        batches = [int(b) for b in re.findall(r"\d+", str(batches_raw or ""))]
        row = {
            "ppn": str(ppn), "dypn": str(dypn),
            "suffix": str(val("SUFFIX") or suffix_of(str(dypn), str(ppn))),
            "alt_suffixes": alts,
            "last_dypn": str(val("LAST DYPN") or dypn),
            "source_material": val("SOURCE MATERIAL"),
            "part_type": val("PART TYPE"),
            "part_type_source": val("PART TYPE SOURCE"),
            "material_type_raw": val("MATERIAL TYPE (RAW)"),
            "part_description": val("PART DESCRIPTION"),
            "qty_latest": int(val("QTY (LATEST)") or 1),
            "times_made": int(val("TIMES MADE") or 1),
            "total_pieces": int(val("TOTAL PIECES") or 1),
            "first_batch": int(val("FIRST BATCH") or 0),
            "last_batch": int(val("LAST BATCH") or 0),
            "batches": batches, "last_wo": val("LAST WO"),
            "hours_latest": val("HOURS (LATEST)"),
            "notes": val("NOTES"),
            "last_updated": val("LAST UPDATED"),
        }
        key = _rk(row["ppn"], row["dypn"], row["source_material"])
        cat.rows[key] = row
        state = cat._ppn.setdefault(
            row["ppn"].upper(), {"batches": set(), "prev": [], "dypn_map": {}})
        state["batches"].update(batches)
        smu = (row["source_material"] or "").upper() \
            if isinstance(row["source_material"], str) else ""
        state["dypn_map"][(row["dypn"].upper(), smu)] = key
        state["dypn_map"][(row["last_dypn"].upper(), smu)] = key
        for a in alts:
            state["dypn_map"][(f"{row['ppn']}-{a}".upper(), smu)] = key

    # prev-occurrence reconstruction: the PPN's newest batch's pieces
    for pu, state in cat._ppn.items():
        if not state["batches"]:
            continue
        newest = max(state["batches"])
        state["prev"] = [
            (row["last_dypn"], row.get("source_material"), key)
            for key, row in cat.rows.items()
            if key[0] == pu and row["last_batch"] == newest]
    return cat
