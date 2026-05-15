"""
LST Organizer Plugin — v15
Single-file TechDeck plugin. All logic runs in-process (no subprocess or external script).

Changes from v14:
- Merged into one file; no LSTOrganizer_full.py dependency
- Post-gather case-insensitive duplicate removal
- organize_by_material setting wired up
- Auto-discover base path when setting is empty
- Granular progress reporting (10/20/30/40/60/75/90/100)
- Cancellation checkpoints throughout
- All output routed through log()
- Fixed output filenames to include actual batch number
"""
from __future__ import annotations

import glob
import json
import os
import re
import shutil
import time
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

# ── Material sets ──────────────────────────────────────────────────────────────

STANDARD_TUBE_MATERIALS = {
    '218002867', '218003095', '218004492', '218012302', '218019939',
    '218019941', '218021954', '218026206', '218026875', '218026962',
    '218033112', '218136140', '40-00-2020', '40-07-1003',
}

OVERSIZED_TUBE_MATERIALS = {
    '181361440', '218000414', '218001209', '218001975', '218002101',
    '218002191', '218002642', '218002778', '218004273', '218012555',
    '218017859', '218018574', '218019943', '218020020', '218020119',
    '218026338', '218026500', '218032982',
}

ALL_TUBE_MATERIALS = STANDARD_TUBE_MATERIALS | OVERSIZED_TUBE_MATERIALS

# ── Path helpers ───────────────────────────────────────────────────────────────

def _candidate_roots() -> List[Path]:
    home = Path.home()
    rel = "Communication site - Electric Boat ASA Docs/Pilot Program/922 QTDR Production Packages"
    cands = [
        home / "American Steel & Alum" / rel,
        home / "OneDrive - American Steel & Alum" / rel,
    ]
    od = os.environ.get("ONEDRIVE") or os.environ.get("OneDrive")
    if od:
        cands.append(Path(od) / rel)
    seen: set = set(); out = []
    for c in cands:
        s = str(c)
        if s not in seen:
            out.append(c); seen.add(s)
    return out


def _locate_root() -> Optional[Path]:
    for r in _candidate_roots():
        if r.exists():
            return r
    return None


def _find_batch_path(root: Path, batch: str) -> Optional[Path]:
    live = root / f"Batch {batch}"
    if live.exists():
        return live
    archived = root / "1 - Completed" / f"Batch {batch}"
    if archived.exists():
        return archived
    return None


def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


ORDER_PREFIX = re.compile(r"^(X|BJ|BK|XY)[0-9]+$", re.IGNORECASE)
_BATCH_INPUT_RE = re.compile(r'^(?:batch|po)\s+([0-9]+)$', re.IGNORECASE)


def _parse_batch_input(raw: str) -> Optional[str]:
    """
    Accept 'Batch 403', 'PO 403', or bare '403' (all case-insensitive).
    Returns the numeric batch string, or None if unrecognisable.
    """
    raw = raw.strip()
    m = _BATCH_INPUT_RE.match(raw)
    if m:
        return m.group(1)
    if re.match(r'^[0-9]+$', raw):
        return raw
    return None

# ── LST discovery ──────────────────────────────────────────────────────────────

def _find_lsts_for_order(order_dir: Path) -> Tuple[List[Path], str]:
    """
    Find .lst files inside any CAD-AND-SHOP-PRINTS/*/7000/ subtree.
    Prefers non-repeat paths; falls back to repeat paths if nothing else found.
    """
    primary: List[Path] = []
    secondary: List[Path] = []
    for p in order_dir.rglob("*"):
        if not (p.is_dir() and p.name.lower() == "7000"):
            continue
        parts_lower = [seg.lower() for seg in p.parts]
        if "cad-and-shop-prints" not in parts_lower:
            continue
        in_repeat = any("repeat" in seg for seg in parts_lower)
        lsts = [f for f in p.glob("*.lst") if f.is_file()]
        if not lsts:
            continue
        (secondary if in_repeat else primary).extend(lsts)
    if primary:
        return primary, "CAD-AND-SHOP-PRINTS"
    if secondary:
        return secondary, "REPEATS*"
    return [], "NONE"

# ── Excel helpers ──────────────────────────────────────────────────────────────

def _strip_step(dypn: str) -> str:
    return dypn[:-5] if dypn.upper().endswith("-STEP") else dypn


def _strip_step_suffix(tokens: List[str]) -> List[str]:
    return tokens[:-1] if tokens and tokens[-1].upper() == "STEP" else tokens


def _sheetmap(wb) -> dict:
    return {s.strip().casefold(): s for s in wb.sheetnames}


def _find_excel_header_row(
    sheet, required_headers: List[str], max_search_rows: int = 20
) -> Tuple[int, Dict[str, int]]:
    required_lower = {h.lower().strip() for h in required_headers}
    for row_num in range(1, max_search_rows + 1):
        row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        row_headers: Dict[str, int] = {}
        for col_idx, cell_value in enumerate(row):
            if cell_value:
                row_headers[str(cell_value).lower().strip()] = col_idx
        if required_lower.issubset(row_headers.keys()):
            return row_num, {h: row_headers[h.lower().strip()] for h in required_headers}
    raise ValueError(
        f"Could not find row with headers {required_headers} in first {max_search_rows} rows"
    )


def _read_po_maps(xlsx_path: Path, debug_fp) -> Tuple[Dict, Dict]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    smap = _sheetmap(wb)
    if not {"po", "source material"}.issubset(smap.keys()):
        raise ValueError("Workbook missing required sheets: PO / SOURCE MATERIAL")
    sh_po = wb[smap["po"]]
    sh_src = wb[smap["source material"]]

    try:
        po_header_row, po_cols = _find_excel_header_row(
            sh_po, ["ORDER", "DYPN", "SOURCE MATERIAL"]
        )
        debug_fp.write(json.dumps({"event": "found_po_headers", "row": po_header_row, "cols": po_cols}) + "\n")
    except ValueError as e:
        raise ValueError(f"PO sheet header error: {e}")

    dypn_to_order_serial: Dict = {}
    for row in sh_po.iter_rows(min_row=po_header_row + 1, values_only=True):
        def _cell(idx: int) -> Optional[str]:
            return str(row[idx]).strip() if len(row) > idx and row[idx] not in (None, "") else None
        order = _cell(po_cols["ORDER"])
        part = _cell(po_cols["DYPN"])
        serial = _cell(po_cols["SOURCE MATERIAL"])
        if order and part:
            dypn_to_order_serial[_strip_step(part)] = (order, serial)

    try:
        src_header_row, src_cols = _find_excel_header_row(sh_src, ["Serial", "Description"])
        debug_fp.write(json.dumps({"event": "found_src_headers", "row": src_header_row, "cols": src_cols}) + "\n")
    except ValueError as e:
        debug_fp.write(json.dumps({"event": "src_header_fallback", "error": str(e)}) + "\n")
        src_header_row = 1
        src_cols = {"Serial": 0, "Description": 5}

    serial_to_desc: Dict = {}
    for row in sh_src.iter_rows(min_row=src_header_row + 1, values_only=True):
        def _cell(idx: int) -> Optional[str]:
            return str(row[idx]).strip() if len(row) > idx and row[idx] not in (None, "") else None
        serial = _cell(src_cols["Serial"])
        desc = _cell(src_cols["Description"])
        if serial and desc:
            serial_to_desc[serial] = desc

    wb.close()
    debug_fp.write(
        json.dumps({"event": "po_maps_ok", "dypn": len(dypn_to_order_serial), "serial": len(serial_to_desc)}) + "\n"
    )
    return dypn_to_order_serial, serial_to_desc


def _autodiscover_master_po(batch_path: Path, debug_fp) -> Optional[Path]:
    doc_folder = batch_path / f"{batch_path.name} - Documentation"
    patterns = []
    if doc_folder.exists():
        patterns.append(str(doc_folder / "**" / "*.xlsx"))
    patterns.append(str(batch_path / "**" / "*.xlsx"))
    seen: set = set(); candidates: List[Path] = []
    for pat in patterns:
        for x in glob.glob(pat, recursive=True):
            if x not in seen:
                candidates.append(Path(x)); seen.add(x)

    def _score(p: Path) -> int:
        s = 0; n = p.name.lower()
        if "qf" in n or "qu" in n: s += 5
        if "pallet" in n or "organizer" in n: s -= 10
        if "po " in n: s += 2
        return s

    ranked = sorted(candidates, key=_score, reverse=True)
    debug_fp.write(json.dumps({"event": "autodiscover_scan", "candidates": len(ranked)}) + "\n")
    return ranked[0] if ranked else None

# ── Row building ───────────────────────────────────────────────────────────────

def _parse_name_parts(stem: str) -> Tuple[Optional[str], str]:
    tokens = _strip_step_suffix(stem.split("-"))
    if not tokens:
        return None, stem
    if ORDER_PREFIX.match(tokens[0]):
        if len(tokens) >= 4:
            return "-".join(tokens[:3]), "-".join(tokens[3:])
        return None, "-".join(tokens)
    return None, "-".join(tokens)


def _build_rows(
    copied_files: List[Path], dypn_map: Dict, serial_map: Dict, debug_fp
) -> Tuple[list, List[str]]:
    rows = []; problems = []
    for p in copied_files:
        order, part = _parse_name_parts(p.stem)
        po_hit = dypn_map.get(part)
        serial = None
        if po_hit:
            order_from_po, serial = po_hit
            if order is None:
                order = order_from_po
        else:
            problems.append(f"Missing PO mapping for DYPN '{part}'")
        desc = serial_map.get(serial) if serial else None
        if serial and not desc:
            problems.append(f"Missing description for serial '{serial}'")
        newname = f"{desc}_{order}-{part}.lst" if (desc and order and part) else None
        rows.append((p.name, order, part, serial, desc, newname))
        debug_fp.write(
            json.dumps({"event": "row", "orig": p.name, "order": order, "part": part,
                        "serial": serial, "desc": desc, "new": newname}) + "\n"
        )
    return rows, problems

# ── Counting ───────────────────────────────────────────────────────────────────

def _compute_counts(
    rows: list, dypn_map: Dict, batch_orders: Set[str], serial_map: Dict
) -> Tuple:
    moved_files = len(rows)
    moved_unique = len({r[2] for r in rows})
    order_nums = {name.split("-")[0] for name in batch_orders} if batch_orders else set()

    standard_tubes: Dict = {}; oversized_tubes: Dict = {}

    def _categorize(part: str, ordnum: str, serial: Optional[str]) -> None:
        s = str(serial).strip() if serial else None
        desc = serial_map.get(s, "UNKNOWN") if s else "UNKNOWN"
        if s in STANDARD_TUBE_MATERIALS:
            standard_tubes[part] = (ordnum, serial, desc)
        elif s in OVERSIZED_TUBE_MATERIALS:
            oversized_tubes[part] = (ordnum, serial, desc)

    target_map = {
        part: (ordnum, serial)
        for part, (ordnum, serial) in dypn_map.items()
        if not order_nums or ordnum in order_nums
    }
    for part, (ordnum, serial) in target_map.items():
        _categorize(part, ordnum, serial)

    moved_tube_parts = {r[2] for r in rows if r[3] and str(r[3]).strip() in ALL_TUBE_MATERIALS}
    missing_std = [
        (p, o, s, d) for p, (o, s, d) in standard_tubes.items() if p not in moved_tube_parts
    ]
    oversized_list = [(p, o, s, d) for p, (o, s, d) in oversized_tubes.items()]

    return (
        moved_files, moved_unique,
        len(standard_tubes), len(oversized_tubes),
        len(standard_tubes) + len(oversized_tubes),
        missing_std, oversized_list,
    )

# ── Output writers ─────────────────────────────────────────────────────────────

def _write_overview_txt(
    txt_path: Path, rows: list, issues: List[str], info: dict,
    moved_files: int, moved_unique: int,
    std_count: int, over_count: int, total_count: int,
    missing_std: list, oversized: list, batch_orders: Set[str],
) -> None:
    with open(txt_path, "w", encoding="utf-8", newline="") as f:
        f.write("# LST TXT work file (generated)\n")
        f.write("# " + "=" * 76 + "\n")
        for k, v in info.items():
            f.write(f"# {k}: {v}\n")
        f.write(f"# moved_files: {moved_files}\n")
        f.write(f"# moved_unique_parts: {moved_unique}\n")
        f.write(f"# standard_tubes_in_po: {std_count}\n")
        f.write(f"# oversized_tubes_in_po: {over_count} (>0.375 NOM - NOT gathered)\n")
        f.write(f"# total_tubes_in_po: {total_count}\n")

        if oversized:
            f.write("#\n")
            f.write(f"# OVERSIZED TUBES (>0.375\" NOM) - {len(oversized)} parts\n")
            f.write("# " + "=" * 76 + "\n")
            f.write("# Tracked in PO but NOT gathered (require special handling).\n#\n")
            by_order: Dict = defaultdict(list)
            for part, order, serial, desc in oversized:
                by_order[order].append((part, serial, desc))
            for order in sorted(by_order):
                f.write(f"#\n# Order: {order}\n")
                for part, serial, desc in by_order[order]:
                    f.write(f"#   - {part} (Serial: {serial}, Desc: {desc})\n")
            f.write("# " + "=" * 76 + "\n")

        if missing_std:
            f.write("#\n")
            f.write(f"# MISSING STANDARD TUBE FILES - {len(missing_std)} not found\n")
            f.write("# " + "=" * 76 + "\n")
            f.write("# These parts are in the PO but no .lst was found.\n#\n")
            by_order = defaultdict(list)
            for part, order, serial, desc in missing_std:
                by_order[order].append((part, serial, desc))
            for order in sorted(by_order):
                f.write(f"#\n# Order: {order}\n")
                for part, serial, desc in by_order[order]:
                    f.write(f"#   - {part} (Serial: {serial}, Desc: {desc})\n")
                    matching = [fo for fo in batch_orders if fo.startswith(order)]
                    if matching:
                        f.write(f"#     Look in: {matching[0]}/*/CAD-AND-SHOP-PRINTS/*/7000/{part}.lst\n")
            f.write("# " + "=" * 76 + "\n")

        f.write("\n")
        f.write("original\torder\tpart\tserial\tdescription\tnew_name\tstatus\n")
        for (orig, order, part, serial, desc, newname) in rows:
            status = "OK" if newname else "SKIPPED"
            f.write(
                f"{orig}\t{order or ''}\t{part or ''}\t{serial or ''}"
                f"\t{desc or ''}\t{newname or ''}\t{status}\n"
            )

        if issues:
            f.write("\n# Issues\n")
            for it in issues:
                f.write(f"# {it}\n")


def _write_grouped_txt(out_path: Path, batch_num: str, rows: list) -> int:
    items = []
    for (orig, order, part, serial, desc, _) in rows:
        if not serial or not desc:
            continue
        if str(serial).strip() not in STANDARD_TUBE_MATERIALS:
            continue
        items.append((desc, str(serial).strip(), order or "", part or ""))

    items.sort(key=lambda t: (t[0].lower(), t[1], t[2], t[3]))
    groups: Dict = defaultdict(list)
    for desc, s, order, part in items:
        groups[(desc, s)].append((order, part))

    with open(out_path, "w", encoding="utf-8", newline="") as f:
        f.write(f"Batch {batch_num}\n")
        f.write(f"Total standard tube parts: {len(items)}\n\n")
        for (desc, s) in sorted(groups, key=lambda k: k[0].lower()):
            f.write(f"{desc} ({s}):\n")
            for order, part in groups[(desc, s)]:
                f.write(f"{order}\t{part}\n")
            f.write("\n")
    return len(items)

# ── File operations ────────────────────────────────────────────────────────────

def _retry_fileop(fn, *args, **kwargs):
    import errno
    tries = kwargs.pop("tries", 5); delay = kwargs.pop("delay", 0.2)
    for i in range(tries):
        try:
            return fn(*args, **kwargs)
        except OSError as e:
            if e.errno in (errno.EACCES, errno.EPERM, errno.EBUSY):
                time.sleep(delay * (i + 1)); continue
            raise


def _gather_and_copy(
    batch_path: Path, batch_num: str, dry_run: bool,
    debug_fp, log, cancel_event,
) -> Tuple[List[Path], List[str], Path, Set[str]]:
    issues: List[str] = []; copied: List[Path] = []; orders: Set[str] = set()
    dest = batch_path / f"Batch {batch_num} - Documentation" / "LST"
    _ensure_dir(dest)
    seen_lower: Set[str] = set()  # case-insensitive guard during gather

    for child in sorted(d for d in batch_path.iterdir() if d.is_dir()):
        if cancel_event.is_set():
            break
        name = child.name
        if name == f"Batch {batch_num} - Documentation":
            continue
        if name.strip().lower() == "repeat batches":
            debug_fp.write(json.dumps({"event": "skip_repeat_batches", "dir": str(child)}) + "\n")
            continue
        if len(name.split("-")) >= 3:
            orders.add(name)

        lsts, src = _find_lsts_for_order(child)
        debug_fp.write(
            json.dumps({"event": "order_probe", "dir": str(child), "source": src, "count": len(lsts)}) + "\n"
        )
        if not lsts:
            continue

        for f in lsts:
            if cancel_event.is_set():
                break
            fname_lower = f.name.lower()
            target = dest / f.name
            try:
                if fname_lower in seen_lower:
                    debug_fp.write(json.dumps({"event": "skip_duplicate_gather", "src": str(f)}) + "\n")
                    continue
                if dry_run:
                    copied.append(target)
                    seen_lower.add(fname_lower)
                    debug_fp.write(json.dumps({"event": "dry_run", "src": str(f), "dst": str(target)}) + "\n")
                else:
                    if target.exists():
                        debug_fp.write(json.dumps({"event": "skip_exists", "dst": str(target)}) + "\n")
                        continue
                    _retry_fileop(shutil.copy2, f, target)
                    copied.append(target)
                    seen_lower.add(fname_lower)
                    debug_fp.write(json.dumps({"event": "copied", "src": str(f), "dst": str(target)}) + "\n")
            except Exception as e:
                issues.append(f"Failed to copy {f.name}: {e}")
                debug_fp.write(json.dumps({"event": "copy_error", "src": str(f), "error": str(e)}) + "\n")

    return copied, issues, dest, orders


def _dedup_destination(dest_dir: Path, log, debug_fp) -> int:
    """
    Case-insensitive duplicate removal on .lst files already in dest_dir.
    For each collision group, keeps the first name alphabetically; deletes the rest.
    Returns the count of files removed.
    """
    by_lower: Dict[str, List[Path]] = defaultdict(list)
    for p in dest_dir.glob("*.lst"):
        by_lower[p.name.lower()].append(p)

    removed = 0
    for group in by_lower.values():
        if len(group) <= 1:
            continue
        group.sort(key=lambda p: p.name)
        keep = group[0]
        for dupe in group[1:]:
            try:
                dupe.unlink()
                log(f"  Removed duplicate: {dupe.name} (kept {keep.name})")
                debug_fp.write(json.dumps({"event": "dedup_removed", "file": dupe.name, "kept": keep.name}) + "\n")
                removed += 1
            except Exception as e:
                log(f"  Warning: could not remove {dupe.name}: {e}")
                debug_fp.write(json.dumps({"event": "dedup_error", "file": dupe.name, "error": str(e)}) + "\n")
    return removed


def _sanitize_folder_name(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "-", name).strip()


def _organize_by_material(dest_dir: Path, rows: list, debug_fp, cancel_event) -> None:
    by_fname: Dict = {}; by_stem: Dict = {}
    for (orig, _, _, serial, desc, _) in rows:
        if not (orig and serial and desc):
            continue
        lname = orig.lower()
        stem = orig[:-4].lower() if lname.endswith(".lst") else lname
        by_fname[lname] = (desc, str(serial).strip())
        by_stem[stem] = (desc, str(serial).strip())

    moved = 0
    for p in sorted(dest_dir.glob("*.lst")):
        if cancel_event.is_set():
            break
        lname = p.name.lower(); stem = p.stem
        clean_stem = stem
        if "(" in stem and stem.endswith(")"):
            try:
                clean_stem = stem[: stem.rindex("(")]
            except ValueError:
                pass

        pair = by_fname.get(lname) or by_stem.get(stem.lower()) or by_stem.get(clean_stem.lower())
        folder = _sanitize_folder_name(f"{pair[0]} ({pair[1]})") if pair else "Uncategorized"
        target_dir = dest_dir / folder
        _ensure_dir(target_dir)

        try:
            final = target_dir / p.name
            k = 2
            while final.exists():
                final = target_dir / f"{p.stem}({k}){p.suffix}"; k += 1
            shutil.move(str(p), str(final))
            moved += 1
            debug_fp.write(json.dumps({"event": "organized", "file": p.name, "folder": folder}) + "\n")
        except Exception as e:
            debug_fp.write(json.dumps({"event": "organize_error", "file": p.name, "error": str(e)}) + "\n")

    debug_fp.write(json.dumps({"event": "organize_done", "moved": moved}) + "\n")

# ── TechDeck plugin entry point ────────────────────────────────────────────────

def run(params: dict, progress_callback, cancel_event) -> None:
    settings = params.get('settings', {})
    log = params.get('log', print)

    log("Starting LST Organizer...")
    progress_callback(0)

    base_path_str = settings.get('base_path', '').strip()
    dry_run = settings.get('dry_run', False)
    do_organize = settings.get('organize_by_material', True)

    # Always prompt for batch number at run time
    console = params.get('console')
    if console and hasattr(console, 'request_input'):
        raw = console.request_input('Enter batch number (e.g. "403", "Batch 403", or "PO 403"):')
    else:
        raw = input('Enter batch number: ')
    batch_no = _parse_batch_input(raw or '')
    if not batch_no:
        raise ValueError(f"Unrecognised batch number input: {raw!r}")

    # Resolve base path — auto-discover if not set
    if base_path_str:
        root = Path(base_path_str)
        if not root.is_dir():
            raise ValueError(f"Base directory not found: {base_path_str}")
        log(f"Base directory: {root}")
    else:
        log("Base path not set — searching OneDrive paths...")
        root = _locate_root()
        if not root:
            raise RuntimeError(
                "Could not auto-locate '922 QTDR Production Packages'. "
                "Set Base Directory in plugin settings."
            )
        log(f"Auto-discovered: {root}")

    batch_path = _find_batch_path(root, batch_no)
    if not batch_path:
        raise RuntimeError(
            f"Batch {batch_no} not found under {root} (also checked '1 - Completed'). "
            "Verify the batch exists and OneDrive is synced."
        )

    log(f"Batch {batch_no}: {batch_path}")
    if dry_run:
        log("DRY RUN MODE — no files will be copied or moved.")

    # Output file paths
    ts = time.strftime("%Y%m%d_%H%M%S")
    lst_dir = batch_path / f"Batch {batch_no} - Documentation" / "LST"
    _ensure_dir(lst_dir)
    debug_path = lst_dir / f"debug_gather_lsts_{ts}.jsonl"
    txt_path = lst_dir / f"LST_Overview_Batch_{batch_no}.txt"
    group_path = lst_dir / f"Tube_Parts_Batch_{batch_no}.txt"

    with open(debug_path, "a", encoding="utf-8") as debug_fp:
        debug_fp.write(
            json.dumps({"event": "start", "ts": ts, "batch": batch_no,
                        "root": str(root), "batch_path": str(batch_path)}) + "\n"
        )

        if cancel_event.is_set():
            return

        # ── Phase 1: Gather LST files ──────────────────────────────────────────
        log("Scanning order folders for .lst files...")
        copied, copy_issues, dest_dir, batch_orders = _gather_and_copy(
            batch_path, batch_no, dry_run, debug_fp, log, cancel_event
        )
        if cancel_event.is_set():
            log("Cancelled."); return
        log(f"Gathered {len(copied)} file(s) across {len(batch_orders)} order folder(s).")
        progress_callback(10)

        # ── Phase 2: Deduplicate destination ───────────────────────────────────
        if not dry_run:
            log("Checking for case-insensitive duplicate filenames...")
            removed = _dedup_destination(dest_dir, log, debug_fp)
            if removed:
                log(f"Removed {removed} duplicate file(s).")
                copied = [p for p in copied if p.exists()]
            else:
                log("No duplicates found.")
        progress_callback(20)

        if cancel_event.is_set():
            log("Cancelled."); return

        # ── Phase 3: Locate Master PO ──────────────────────────────────────────
        log("Auto-discovering Master PO workbook...")
        master_po = _autodiscover_master_po(batch_path, debug_fp)

        if not master_po or not master_po.exists():
            log("WARNING: Master PO not found — writing seed file (filenames only).")
            seed_rows = [(p.name, None, _strip_step(p.stem), None, None, None) for p in copied]
            _write_overview_txt(
                txt_path, seed_rows, ["Master PO workbook NOT FOUND"],
                {"batch": batch_no, "note": "seed only"},
                len(seed_rows), len({r[2] for r in seed_rows}),
                0, 0, 0, [], [], batch_orders,
            )
            raise RuntimeError(
                f"Master PO not found. Seed report written to:\n  {txt_path}\n"
                "Ensure the Master PO workbook is present in the batch folder and re-run."
            )

        debug_fp.write(json.dumps({"event": "master_po", "path": str(master_po)}) + "\n")
        progress_callback(30)

        # ── Phase 4: Read PO data ──────────────────────────────────────────────
        log(f"Reading Master PO: {master_po.name}...")
        try:
            dypn_map, serial_map = _read_po_maps(master_po, debug_fp)
        except Exception as e:
            debug_fp.write(json.dumps({"event": "po_read_fail", "error": str(e)}) + "\n")
            raise RuntimeError(f"Could not read Master PO: {e}")

        log(f"Loaded {len(dypn_map)} DYPN entries and {len(serial_map)} material descriptions.")
        progress_callback(40)

        if cancel_event.is_set():
            log("Cancelled."); return

        # ── Phase 5: Build rows ────────────────────────────────────────────────
        log("Mapping files to PO data...")
        rows, problems = _build_rows(copied, dypn_map, serial_map, debug_fp)
        progress_callback(60)

        # ── Phase 6: Compute tube counts ───────────────────────────────────────
        (moved_files, moved_unique, std_count, over_count, total_count,
         missing_std, oversized_list) = _compute_counts(rows, dypn_map, batch_orders, serial_map)

        # ── Phase 7: Write output files ────────────────────────────────────────
        log("Writing output files...")
        issues = copy_issues + problems
        _write_overview_txt(
            txt_path, rows, issues,
            {"batch": batch_no, "root": str(root), "master_po": str(master_po),
             "dest_dir": str(dest_dir), "orders": ",".join(sorted(batch_orders))},
            moved_files, moved_unique, std_count, over_count, total_count,
            missing_std, oversized_list, batch_orders,
        )
        _write_grouped_txt(group_path, batch_no, rows)
        progress_callback(75)

        if cancel_event.is_set():
            log("Cancelled."); return

        # ── Phase 8: Organize by material ──────────────────────────────────────
        if do_organize and not dry_run:
            log("Organizing files by material type...")
            _organize_by_material(dest_dir, rows, debug_fp, cancel_event)
        elif dry_run:
            log("Dry run: skipping file organization.")
        progress_callback(90)

        debug_fp.write(json.dumps({"event": "done"}) + "\n")

    # ── Console summary ────────────────────────────────────────────────────────
    log("=" * 60)
    log(f"LST Organizer — Batch {batch_no} Complete")
    log("=" * 60)
    log(f"Files gathered:  {moved_files}")
    log(f"Unique parts:    {moved_unique}")
    log("")
    log("Tube counts:")
    log(f"  Standard (gathered):      {std_count}")
    log(f"  Oversized (>0.375\" NOM):  {over_count}")
    log(f"  Total in PO:              {total_count}")

    if oversized_list:
        log(f"\n{len(oversized_list)} oversized tube part(s) tracked but NOT gathered.")

    if missing_std:
        log(f"\nWARNING: {len(missing_std)} standard tube file(s) not found!")
        for part, order, *_ in sorted(missing_std)[:5]:
            log(f"  - {part} (Order: {order})")
        if len(missing_std) > 5:
            log(f"  ... and {len(missing_std) - 5} more. See report for full list.")
    else:
        log("\nAll expected standard tube files found.")

    if issues:
        log(f"\n{len(issues)} issue(s) logged — see report for details.")

    log("")
    log(f"Main report:   {txt_path}")
    log(f"Grouped tubes: {group_path}")
    log(f"Debug log:     {debug_path}")
    log("=" * 60)
    progress_callback(100)


# ── Standalone test harness ────────────────────────────────────────────────────
if __name__ == "__main__":
    import threading
    cancel = threading.Event()
    run(
        params={
            'log': print,
            'settings': {
                'base_path': '',
                'batch_number': input("Batch number: ").strip(),
                'master_po_path': '',
                'dry_run': False,
                'organize_by_material': True,
            },
        },
        progress_callback=lambda p: print(f"[{p}%]"),
        cancel_event=cancel,
    )
