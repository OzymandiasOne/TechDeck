"""
911 SSPO Award Review Plugin for TechDeck
=========================================

(Renamed from '911 Runtime Estimator' 2026-07-13 at planning's request -- the
tool now covers the whole award review, not just plate time estimating.)

Given a ROOT directory containing one subfolder per order (e.g. F124, P003,
S029, V092 ...), this plugin:

  1. Walks each order folder, locating the nest work-package PDFs and the
     order's BATCH LIST workbook (folder names are resolved DEFENSIVELY -- the
     PDFs may live in 'NEST PACKAGES' or 'WPDD SKETCHES' under a
     'CUI- TECH DATA READ ME' folder, and layouts vary between orders).
  2. Parses each nest packet PDF (PyMuPDF) for thickness, pieces, material,
     stock size, mil spec, and plate weight -- cross-checking fields across the
     page-1 data row, the SUMMARY page, the part sketch, and the MOVE TICKET,
     exactly like the 911 Setup plugin. Shape (1D) nest orders additionally
     yield the 'Summary of Batches' lengths -> 'Shape Ft Req' (each length /
     12, rounded UP to the whole foot, summed per nest) plus the stock
     Material Size / Type, feeding receiving's material cheat sheet: the
     per-nest value repeats on the nest's rows, and a 'Shape Ft Req' sheet
     rolls up one line per shape nest with a grand total.
  3. Computes a per-row plate cutting-time estimate from ACTUAL THROUGHPUT:
     a thickness-band 'pieces per hour' table derived from 18.5 months of
     closed D911 actuals (planning's table, 2026-07) -- see CALCULATION below.
     Structural-shape rows keep the legacy thickness-band estimate. The old
     linear-inch times are still computed for every plate row that resolves a
     DXF (Work Order matched in the order's 'IGES FILES' folder, cut length
     measured from the DXF vector geometry / a thickness-driven feed rate) and
     ride along as REFERENCE columns at the far right of the table.
  4. Writes ONE consolidated workbook with two sheets (Plates / Non-Plates).
     Each is a flat data table that reproduces EVERY batch-list column in its
     native order, then our generated columns. 'Material' on the table is the
     MOVE TICKET designation; the batch list's own 'Material' (an EB stock code)
     is shown as 'Source Material'.
  5. Adds a 'Working Forecast Input' sheet -- one row per nest (plates AND
     shapes) in the Working Forecast List's input-column layout (Order,
     Source Material, Pieces, Nest Pkg Nbr, Orders, REM Used, REM Made,
     Total Ft Req, Operations) so a reviewed award copy-pastes straight into
     the forecast. Source Material is the packet's page-1 'Source Code' (EB
     stock identifier), Pieces/Orders are the page-1 data-row counts, Total
     Ft Req repeats the shape nest feet; the REM pair + Operations stay blank
     for manual input.
  6. Drops a REAL, refreshable Excel PivotTable below the data on each sheet --
     ONE line per nest: Parts (true piece count for the whole nest) + the same
     Cut Time total in both hours and minutes, grand total at the bottom -- by
     driving Excel via COM (pywin32). A plain-English caveat stating the
     throughput basis (includes setup/handling -- it's derived from closed
     actuals) sits above it. If Excel/COM is unavailable or errors, it falls back
     to a static per-nest summary of the same shape so a run never hard-crashes.

Runs off an AWARD PACKAGE (the SPARS-received folder of order folders, before
batching) so work-scope review / division + machine assignment can happen up
front; the plate cut times are the actuals-derived throughput estimates (with
the exact linear-inch times as far-right reference columns). Replaced the
standalone 911_linear_inch_cuttime tool (removed) whose engine is vendored here.

CALCULATION (per plate batch-list row; v2.7.0 -- actuals-derived throughput):
    pcs/hr       = lookup_pieces_per_hour(thickness)  [PIECES_PER_HOUR bands]
    Est Min/PC   = 60 / pcs/hr
    Est Min/WO   = Est Min/PC * (row PPN Quantity)
    Est Cut Hours = Est Min/WO / 60
An 'Equation' column (right before Est Cut Hours) prints this chain with the
row's own numbers plugged in so the arithmetic is auditable cell-by-cell.
The prior primary estimate -- EXACT DXF linear inches / thickness feed rate --
is kept per planning's request as reference columns at the FAR RIGHT of the
table (Feed Rate (IPM), LI Est Min/PC, LI Est Min/WO, LI Est Cut Hours):
    LI Est Min/PC = linear_in_pc / lookup_feed(thickness)  [rounds thk UP]
Bevel adds NO time (the legacy +180 min/pc surcharge was removed per planning
direction); bevel scope is still detected and flagged for review. Structural
shape rows use the legacy band for their primary estimate:
    band(t): t<0.5->6 ; <1.0->9 ; <2.0->12 ; >=2.0->18 (min/pc); est = t*band*qty.

Prompts for the ROOT via the native folder dialog only (no pre-configured settings
required). Optional 'output_dir' setting picks where the workbook is saved;
blank saves it inside ROOT.
"""

from __future__ import annotations

import os
import re
import glob
import math
import statistics
import datetime
from pathlib import Path
from typing import Optional

# SDK bootstrap (works in-process and for standalone CLI testing).
try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList


# ══════════════════════════════════════════════════════════════════════════
# Linear-inch cut-time engine (vendored from the removed 911_linear_inch_cuttime)
# ──────────────────────────────────────────────────────────────────────────
# Plate cut time = EXACT DXF cut length / feed rate (IPM). Cut length per part
# comes from the order's IGES DXF vector geometry; the feed rate is looked up by
# thickness in one combined chart, rounding UP to the next listed thickness so
# oddball thicknesses stay covered. TODO: centralize this parser into plugin_sdk
# (customer_dxf_analysis + this module carry copies).
# ══════════════════════════════════════════════════════════════════════════

# Work order token in DXF filenames (PART.REV.WORKORDER) + the batch list. Two
# SigmaNest formats: current "X4549775"/"BK514922" (1-2 letters + 5-8 digits) and
# legacy "3X24-398"; same form 911_setup's summary parser uses.
WO_RE = re.compile(r"(?:[A-Z]{1,2}\d{5,8}|\d[A-Z]\d+-\d+)", re.I)
# Layers that are reference-only and never count as cut (matches customer_dxf_analysis).
# Only these NAMED reference layers are dropped; ALL numeric geometry layers count.
# Confirmed by shop validation in SigmaNest (5 award parts, 2026-07-09): the cut is
# spread across every numeric layer (1/2 carry perimeter+holes inconsistently, 4/5
# carry other cut features) - so cut length = the sum of ALL non-reference layers.
# An earlier "layer 1 only" rule was WRONG (it rested on one Kinetic profile-only
# sheet that happened to list just a part's outer perimeter).
EXCLUDE_LAYERS = {"BOUNDING_BOX", "IGNORE"}
# Entity types we do NOT measure; their presence means the length may be under-counted.
UNSUPPORTED_FLAGGED = {"SPLINE", "ELLIPSE"}

FEED_RATES_SOURCE = "Plasma & Oxy Feed Rates (single combined chart, rev 2026-07-09)"
FEED_RATES = [  # (thickness in, speed IPM) - ascending
    (0.125, 180), (0.188, 140), (0.25, 100), (0.313, 80), (0.375, 100),
    (0.438, 90), (0.5, 75), (0.625, 55), (0.75, 80), (0.875, 70),
    (1.0, 50), (1.125, 45), (1.25, 40), (1.375, 35), (1.5, 30),
    (1.625, 45), (1.75, 40), (2.0, 30), (2.25, 25), (2.5, 20),
    (2.75, 9.25), (3.0, 9.25), (3.25, 9.0), (3.5, 8.5), (3.75, 8.5),
    (4.0, 8.5), (4.25, 8.0), (4.5, 8.0), (5.0, 7.25), (5.25, 6.5),
    (5.5, 6.0), (5.75, 6.0), (6.0, 6.0), (6.25, 6.0), (6.5, 6.0),
    (6.75, 6.0), (7.0, 6.0),
]


def lookup_feed(thickness):
    """(ipm, note) for a thickness, rounding UP to the next listed thickness (thicker
    = slower = more time = conservative). Clamps below/above the chart (flagged via
    note). None only when thickness is None."""
    if thickness is None:
        return None
    for t, ipm in FEED_RATES:
        if thickness <= t + 1e-9:
            note = None if abs(t - thickness) < 1e-6 else f"rounded up to {t}\" row"
            return ipm, note
    t, ipm = FEED_RATES[-1]
    return ipm, f"above chart max {t}\" - used {t}\" row"


# PRIMARY estimate driver since v2.7.0 (planning's table, 2026-07-20): pieces
# per hour by thickness band, derived from 18.5 months of closed D911 actuals
# ('D911 Closed as of 7-14-2026' -> 'Pivot Summ_Piv': pcs/hr = $500/hr shop
# rate / avg billed value per piece in the band, Jan 2025 - Jul 2026). Being
# billing-derived it bakes in setup/handling/downtime, which is why it replaced
# the geometry-only linear-inch time as the primary number (LI times remain as
# reference columns). Bands are contiguous (source sheet's gaps/typos cleaned
# per planning: <=0.313, >0.313-0.50, >0.50-1, >1-2, >2). Scott's call
# (2026-07-21, via planning): 5/16" is the group-1/group-2 break -- 0.313 is
# the TOP of group 1, so 5/16 (0.3125) plate lands in group 1.
PIECES_PER_HOUR_SOURCE = ("D911 closed actuals Jan 2025 - Jul 2026 "
                          "('D911 Closed as of 7-14-2026', Pivot Summ_Piv)")
PIECES_PER_HOUR = [  # (band max thickness in, pieces per hour) - ascending
    (0.313, 5.73),
    (0.50, 6.00),
    (1.0, 5.33),
    (2.0, 2.85),
    (6.5, 1.05),
]


def lookup_pieces_per_hour(thickness):
    """(pcs_hr, note) for a thickness; band = first row whose max covers it.
    Above the table max, clamps to the last band (flagged via note). None only
    when thickness is None/non-positive."""
    if thickness is None or thickness <= 0:
        return None
    for tmax, pph in PIECES_PER_HOUR:
        if thickness <= tmax + 1e-9:
            return pph, None
    tmax, pph = PIECES_PER_HOUR[-1]
    return pph, f"above throughput table max {tmax}\" - used {tmax}\" band"


def _decode_dxf(path):
    data = Path(path).read_bytes()
    if data.startswith(b"AutoCAD Binary DXF"):
        raise ValueError("Binary DXF is not supported - re-export as ASCII DXF.")
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not decode file - is this a DXF?")


def _read_pairs(path):
    lines = _decode_dxf(path).splitlines()
    pairs = []
    it = iter(lines)
    for code_line in it:
        try:
            value = next(it)
        except StopIteration:
            break
        try:
            code = int(code_line.strip())
        except ValueError:
            raise ValueError(f"Malformed DXF (expected numeric group code, got {code_line.strip()!r}).")
        pairs.append((code, value.strip()))
    if not pairs:
        raise ValueError("File is empty.")
    return pairs


def _bulge_length(p1, p2, bulge):
    chord = math.hypot(p2[0] - p1[0], p2[1] - p1[1])
    if abs(bulge) < 1e-12 or chord < 1e-12:
        return chord
    theta = 4.0 * math.atan(bulge)
    radius = chord * (1.0 + bulge * bulge) / (4.0 * abs(bulge))
    return radius * abs(theta)


def _collect_until_next_entity(pairs, j):
    groups = []
    n = len(pairs)
    while j < n and pairs[j][0] != 0:
        groups.append(pairs[j])
        j += 1
    return groups, j


def _build_line(groups):
    g = {}
    for code, val in groups:
        if code in (8, 10, 20, 11, 21):
            g[code] = val
    try:
        p1 = (float(g[10]), float(g[20]))
        p2 = (float(g[11]), float(g[21]))
    except KeyError:
        return None
    return {"layer": g.get(8, "0"), "points": [p1, p2], "length": math.dist(p1, p2)}


def _build_arc(groups, full_circle=False):
    g = {}
    for code, val in groups:
        if code in (8, 10, 20, 40, 50, 51):
            g[code] = val
    try:
        cx, cy, r = float(g[10]), float(g[20]), float(g[40])
    except KeyError:
        return None
    if full_circle:
        a0, sweep = 0.0, math.tau
    else:
        a0 = math.radians(float(g.get(50, 0.0)))
        a1 = math.radians(float(g.get(51, 360.0)))
        sweep = (a1 - a0) % math.tau or math.tau
    n = max(8, int(abs(sweep) / math.tau * 64) + 1)
    pts = [(cx + r * math.cos(a0 + sweep * t / n), cy + r * math.sin(a0 + sweep * t / n))
           for t in range(n + 1)]
    return {"layer": g.get(8, "0"), "points": pts, "length": r * sweep}


def _polyline_entity(layer, verts, closed):
    verts = [v for v in verts if v[1] is not None]
    if len(verts) < 2:
        return None
    total = 0.0
    seg_pairs = list(zip(verts, verts[1:]))
    if closed:
        seg_pairs.append((verts[-1], verts[0]))
    for v1, v2 in seg_pairs:
        total += _bulge_length((v1[0], v1[1]), (v2[0], v2[1]), v1[2])
    points = [(v[0], v[1]) for v in verts]
    return {"layer": layer, "points": points, "length": total}


def _build_lwpolyline(groups):
    layer, flags = "0", 0
    verts = []  # [x, y, bulge]
    for code, val in groups:
        if code == 8:
            layer = val
        elif code == 70:
            flags = int(val)
        elif code == 10:
            verts.append([float(val), None, 0.0])
        elif code == 20 and verts:
            verts[-1][1] = float(val)
        elif code == 42 and verts:
            verts[-1][2] = float(val)
    return _polyline_entity(layer, verts, closed=bool(flags & 1))


def _build_polyline(pairs, j):
    groups, j = _collect_until_next_entity(pairs, j)
    layer, flags = "0", 0
    for code, val in groups:
        if code == 8:
            layer = val
        elif code == 70:
            flags = int(val)
    verts = []
    n = len(pairs)
    while j < n:
        code, val = pairs[j]
        if code == 0 and val == "VERTEX":
            vgroups, j = _collect_until_next_entity(pairs, j + 1)
            v = [None, None, 0.0]
            vflags = 0
            for c, vv in vgroups:
                if c == 10:
                    v[0] = float(vv)
                elif c == 20:
                    v[1] = float(vv)
                elif c == 42:
                    v[2] = float(vv)
                elif c == 70:
                    vflags = int(vv)
            if v[0] is not None and v[1] is not None and not (vflags & 16):
                verts.append(v)
        elif code == 0 and val == "SEQEND":
            _, j = _collect_until_next_entity(pairs, j + 1)
            break
        else:
            j += 1
    return _polyline_entity(layer, verts, closed=bool(flags & 1)), j


def parse_dxf(path):
    """Return (entities, skipped, insunits). entity = {layer, points, length}."""
    pairs = _read_pairs(path)
    n = len(pairs)
    entities, skipped = [], {}
    insunits = None
    section = None
    i = 0
    while i < n:
        code, val = pairs[i]
        if code == 0 and val == "SECTION" and i + 1 < n and pairs[i + 1][0] == 2:
            section = pairs[i + 1][1]
            i += 2
            continue
        if code == 0 and val == "ENDSEC":
            section = None
            i += 1
            continue
        if section == "HEADER" and code == 9 and val == "$INSUNITS":
            if i + 1 < n and pairs[i + 1][0] == 70:
                insunits = int(pairs[i + 1][1])
            i += 1
            continue
        if section == "ENTITIES" and code == 0:
            if val == "LINE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_line(groups)
            elif val == "ARC":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_arc(groups)
            elif val == "CIRCLE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_arc(groups, full_circle=True)
            elif val == "LWPOLYLINE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_lwpolyline(groups)
            elif val == "POLYLINE":
                ent, i = _build_polyline(pairs, i + 1)
            else:
                skipped[val] = skipped.get(val, 0) + 1
                i += 1
                continue
            if ent is not None:
                entities.append(ent)
            continue
        i += 1
    return entities, skipped, insunits


def dxf_metrics(path):
    """Cut length (in) + anomaly info for one part DXF. Sums the geometry on ALL
    non-reference layers (the cut is spread across every numeric layer). Layers
    that hold no geometry (0 in) are disregarded - they don't count toward the
    layer total nor toward the multi-layer determination (`n_layers`)."""
    entities, skipped, insunits = parse_dxf(path)
    scale, unit_note = 1.0, None
    if insunits == 4:            # millimetres
        scale, unit_note = 1.0 / 25.4, "units=mm (converted to in)"
    elif insunits not in (1, 0, None):
        unit_note = f"unexpected $INSUNITS={insunits}"
    per_layer = {}
    for e in entities:
        ly = e["layer"].strip()
        if ly.upper() in EXCLUDE_LAYERS:
            continue
        per_layer[ly] = per_layer.get(ly, 0.0) + e["length"]
    # Disregard empty (0 in) layers so they can't be counted or skew anything,
    # but count them (a part carrying a present-but-zero-length layer is a
    # data-quality signal we surface on the Analysis sheet).
    cut_layers = {ly: ln for ly, ln in per_layer.items() if ln > 1e-9}
    n_empty_layers = len(per_layer) - len(cut_layers)
    total = sum(cut_layers.values())
    return {
        "linear_inches": total * scale,
        "skipped": skipped,
        "layers": set(cut_layers),
        "n_layers": len(cut_layers),
        "n_empty_layers": n_empty_layers,
        "unit_note": unit_note,
        "n_entities": len(entities),
    }


def _dxf_work_order(filename):
    """The work order token embedded in a DXF filename (PART.REV.WORKORDER): the
    last WO-shaped '.'-separated component, so the part number's leading digits
    aren't mistaken for it. None if none present."""
    stem = os.path.splitext(os.path.basename(filename))[0].upper()
    for comp in reversed(re.split(r"[.\s_]+", stem)):
        if WO_RE.fullmatch(comp):
            return comp
    return None


def build_dxf_index(iges_folder, log, cancel_event=None):
    """Map WORK ORDER (upper) -> linear inches per piece from an order's IGES DXFs.
    Also returns `suspect` (work orders whose geometry may be under-counted:
    unmeasured SPLINE/ELLIPSE or non-inch units), `multilayer` (work order ->
    True when the part's cut spans more than one non-empty layer), and `empty`
    (work order -> True when the part carries a present-but-zero-length layer)."""
    index, suspect, multilayer, empty = {}, {}, {}, {}
    if not iges_folder:
        return index, suspect, multilayer, empty
    for i, dxf in enumerate(glob.glob(os.path.join(iges_folder, "**", "*.dxf"), recursive=True)):
        if i % 64 == 0:
            sdk.raise_if_cancelled(cancel_event)
        wo = _dxf_work_order(dxf)
        if not wo:
            continue
        try:
            sdk.ensure_local(dxf, log=log)
            m = dxf_metrics(dxf)
        except Exception as exc:
            log(f"  FLAG: could not read DXF {os.path.basename(dxf)}: {exc}")
            continue
        index[wo] = m["linear_inches"]
        multilayer[wo] = m["n_layers"] > 1
        empty[wo] = m["n_empty_layers"] > 0
        notes = []
        bad = [k for k in m["skipped"] if k in UNSUPPORTED_FLAGGED]
        if bad:
            notes.append("unmeasured " + "/".join(sorted(bad)))
        if m["unit_note"] and "converted" not in m["unit_note"]:
            notes.append(m["unit_note"])
        if notes:
            suspect[wo] = "; ".join(notes)
    return index, suspect, multilayer, empty


VERSION = "2.8.0"

# We reproduce EVERY batch-list column verbatim, in its native order, then
# append our own generated columns after them. The batch list's own 'Material'
# column is an EB stock code; we surface it as 'Source Material' to keep it
# distinct from the part-material designation we pull off the MOVE TICKET pages.
BATCH_HEADER_RENAME = {"Material": "Source Material"}
# 'Order' (which order folder a row came from) leads the table; our other
# generated columns follow the batch-list block (in this order). 'Material' is
# the MOVE TICKET designation (sits right after MIL SPEC on the packet).
GENERATED_COLS = ["Division", "Process", "Thickness (in)",
                  "Remnant Length", "Remnant Width", "Stock L", "Stock W",
                  "Shape Ft Req",
                  "Mil Spec", "Material", "Linear In/PC", "Total Linear In",
                  "Multi-Layered", "LI Dev (SD)",
                  "Pcs/Hr", "Est Min/PC", "Est Min/WO", "Multiplier",
                  "Equation", "Est Cut Hours", "Flags",
                  "",   # deliberate BLANK spacer column: clean visual division
                        # between the live estimate and the old LI reference
                  "Feed Rate (IPM)", "LI Est Min/PC", "LI Est Min/WO",
                  "LI Est Cut Hours",
                  "EB Machine", "EB Fuel"]
# 'Remnant Length'/'Remnant Width' are the nest sheet's page-1 Length/Width --
# the remnant's dimensions -- filled ONLY on rows whose batch-list 'Rem Used'
# (AA) is populated (a remnant was used); blank otherwise.
# 'Shape Ft Req' (v2.5.0, receiving's ask) = the SHAPE nest's total stock feet:
# each 'Summary of Batches' table row's Length (inches) / 12, rounded UP to the
# whole foot, summed across the table (e.g. 143 + 20 -> 12 ft + 2 ft = 14 ft).
# It's a per-NEST value repeated on every row of the nest (like Thickness), so
# don't SUM the column -- the per-nest/total rollup lives on the 'Shape Ft Req'
# sheet. Blank on plate rows (plate packets carry no Summary of Batches table).
# 'Multiplier' is an intentionally blank placeholder the planner fills in by hand
# (a future classification factor, e.g. base=1 / std bevel=2), like 'Division'.
# 'Pcs/Hr' (v2.7.0) is the row's PIECES_PER_HOUR band rate -- the primary
# estimate driver: Est Min/PC = 60 / Pcs/Hr for every plate row with a known
# thickness (no DXF needed). Structural shapes keep the legacy band estimate.
# 'Equation' spells out the FULL calculation for the row with this row's own
# numbers plugged in (Est Min/PC -> Est Min/WO -> Est Cut Hours), so the math is
# auditable cell-by-cell -- it sits right before 'Est Cut Hours' (the result).
# To make that audit hold, the Est columns are rounded PROGRESSIVELY (each derived
# from the previous rounded cell) so e.g. Est Min/PC x Qty reproduces Est Min/WO
# exactly rather than diverging by an independent-rounding cent.
# 'Division' is an intentionally blank placeholder the planner fills in by hand.
# The FAR-RIGHT block after 'Flags' -- Feed Rate (IPM) / LI Est Min/PC /
# LI Est Min/WO / LI Est Cut Hours -- is the RETIRED primary estimate (exact DXF
# linear inches / feed rate), kept as a reference comparison per planning's
# request (v2.7.0: "keep the original runtimes, moved to a separate location on
# the right"). LI Est Min/PC = Linear In/PC / Feed Rate; the chain matches the
# old Equation column. Populated only for plate rows that resolve a DXF.
# 'EB Machine' / 'EB Fuel' (v2.8.0) close the table: the packet's page-1
# header-table Machine (labelled 'Gantry' on some packets) and Fuel cells --
# per-NEST values repeated on the nest's rows, read positionally by
# _page1_value_below (the flat text interleaves that table's cells).
# The linear-inch measurement columns (Linear In/PC, Total Linear In,
# Multi-Layered, LI Dev (SD)) stay in place -- they're geometry facts, not
# runtimes. 'Multi-Layered' = Y/N (does
# the part's cut span >1 non-empty layer; multi-layer rows are highlighted).
# 'LI Dev (SD)' = how many standard deviations the part's Linear In/PC sits from
# the mean of all plate parts in the run; parts past LI_OUTLIER_SD are flagged +
# highlighted so an inflated/odd part is easy to catch.
LI_OUTLIER_SD = 3.0
# For each OUTPUT header, the UPPERCASE row key its value is read from. Most map
# to their own upper-cased name; these two are indirections (the display name
# differs from the key the value is stored under).
HEADER_SOURCE_KEY = {
    "Source Material": "MATERIAL",      # batch list 'Material' (EB stock code)
    "Material": "MT MATERIAL",          # MOVE TICKET material designation (new)
}

# The column the pivot sums (our formula-driven estimate, in hours).
EST_COL = "Est Cut Hours"


# ──────────────────────────────────────────────────────────────────────────
# Calculation
# ──────────────────────────────────────────────────────────────────────────

def _eqnum(x):
    """Format a number for the human-readable Equation cell: no trailing zeros,
    integers shown without a decimal point ('4.0' -> '4', '0.750' -> '0.75') so
    the printed arithmetic matches what Excel shows in the neighbouring cells."""
    if x is None:
        return "?"
    try:
        xf = float(x)
    except (TypeError, ValueError):
        return str(x)
    if xf == int(xf):
        return str(int(xf))
    return f"{xf:g}"


def band_minutes(thickness: float) -> int:
    """Minutes-per-piece band for a plate thickness (inches).

    Lower-inclusive / upper-exclusive; the >= band wins on a boundary.
    """
    if thickness < 0.5:
        return 6
    if thickness < 1.0:
        return 9
    if thickness < 2.0:
        return 12
    return 18


def row_estimate_hours(thickness: Optional[float], ppn_qty: float, scope: str):
    """Legacy thickness-band estimate, used only as a fallback for rows with no
    DXF geometry (structural shapes / a missing file). Returns
    (est_hours, factor_min_per_pc, 0.0), or (None, None, None) when thickness is
    unknown. Bevel adds no time (removed per planning direction); ``scope`` is
    kept for signature stability.
    """
    if thickness is None or thickness <= 0:
        return None, None, None
    qty = ppn_qty if (ppn_qty and ppn_qty > 0) else 0
    factor = thickness * band_minutes(thickness)          # min/pc
    est_min = factor * qty
    return est_min / 60.0, factor, 0.0


def throughput_estimate(thickness, ppn_qty):
    """PRIMARY plate estimate since v2.7.0: actuals-derived pieces-per-hour.

    Returns a dict with the band rate, the two Est Min breakouts, and Est Cut
    Hours -- or None when thickness is unknown. Needs NO DXF (the band covers
    setup/handling because it's billing-derived), so every plate row with a
    thickness gets a primary estimate:
        Est Min/PC   = 60 / pcs_hr
        Est Min/WO   = Est Min/PC * qty
        Est Cut Hours = Est Min/WO / 60
    """
    pph = lookup_pieces_per_hour(thickness)
    if pph is None:
        return None
    pcs_hr, note = pph
    qty = ppn_qty if (ppn_qty and ppn_qty > 0) else 0
    min_pc = 60.0 / pcs_hr
    min_wo = min_pc * qty
    return {
        "pcs_hr": pcs_hr,
        "pph_note": note,
        "min_pc": min_pc,
        "min_wo": min_wo,
        "hours": min_wo / 60.0,
    }


def plate_cut_estimate(linear_in_pc, thickness, ppn_qty):
    """Linear-inch cut-time estimate for a plate row that resolved a DXF.

    Returns a dict with the per-piece cut length, feed rate, the two Est Min
    breakouts, and Est Cut Hours -- or None when it can't be computed (no DXF
    match or no thickness). Cut time = cut length / feed rate (IPM); the feed
    rate rounds thickness UP to the next chart row. Bevel adds NO time (the old
    +180 min/pc surcharge was removed per planning direction); bevel scope is
    still flagged for review.
        Est Min/PC   = linear_in_pc / feed_ipm
        Est Min/WO   = Est Min/PC * qty
        Est Cut Hours = Est Min/WO / 60
    """
    if linear_in_pc is None or thickness is None or thickness <= 0:
        return None
    feed = lookup_feed(thickness)
    if feed is None or not feed[0]:
        return None
    ipm, note = feed
    qty = ppn_qty if (ppn_qty and ppn_qty > 0) else 0
    min_pc = linear_in_pc / ipm
    min_wo = min_pc * qty
    return {
        "linear_in_pc": linear_in_pc,
        "total_linear_in": linear_in_pc * qty,
        "feed_ipm": ipm,
        "feed_note": note,
        "min_pc": min_pc,
        "min_wo": min_wo,
        "hours": min_wo / 60.0,
    }


def _is_bevel(scope) -> bool:
    """True if the scope-of-work text contains 'BEVEL' (case-insensitive).

    Real batch lists use compound scopes like 'CUT/BEVEL', so this is a
    CONTAINS test, not exact-equals.
    """
    return scope is not None and "BEVEL" in str(scope).strip().upper()


# Material-form tokens that mark a row as a structural shape (NOT a plate),
# matched against the Description's LEADING form word. Real batch-list
# descriptions lead with the stock form, e.g. 'ANGLE, HIGH STR, , STL, ...',
# 'TUBE, SQUARE, ...', 'T-SECTION, ...', 'BAR, FLAT, ...', 'ROD, ROUND, ...'.
# Shapes routinely carry a thickness token ('2.500 X 2.500 X 0.188 THK',
# '0.500THK3.000W') in their description, so they MUST be classified by FORM,
# not by whether a thickness/estimate was found.
_NONPLATE_FORMS = ("ANGLE", "TUBE", "PIPE", "BAR", "ROD", "CHANNEL", "BEAM",
                   "T-SECTION", "T SECTION", "I-SECTION", "I SECTION",
                   "W-SECTION", "W SECTION", "SECTION")


def _desc_is_shape_form(description) -> bool:
    """True if a Description's LEADING form token is a structural shape.

    Shared by _is_plate_row (sheet split) and the build loop (which estimate
    drives the row: shapes keep the legacy band, everything else gets the
    throughput table) so the two can never disagree.
    """
    desc = str(description or "").upper()
    lead = re.split(r"[;,]", desc, maxsplit=1)[0].strip()
    return any(lead.startswith(f) for f in _NONPLATE_FORMS)


def _is_plate_row(row: dict) -> bool:
    """True if a data row is a plate (vs a structural shape/bar/tube).

    Classification is DESCRIPTION-driven, keyed on the leading material-form
    token (the text before the first ',' or ';'):
      * leading form is a known shape (ANGLE/TUBE/T-SECTION/BAR/ROD/...) -> NOT a
        plate. This is the fix: shapes carry a thickness token in their
        description (e.g. 'TUBE, SQUARE, ..., 0.188 THK') so the old
        'estimate-was-computed' rule misfiled every dimensioned shape onto the
        Plates sheet;
      * description says PLATE -> plate;
      * otherwise (blank/unknown description) fall back to 'an estimate was
        computed' -- a nest-packet plate thickness implies a plate.
    """
    desc = str(row.get("Description") or "").upper()
    if _desc_is_shape_form(desc):
        return False
    if "PLATE" in desc:
        return True
    return row.get(EST_COL) is not None


# ──────────────────────────────────────────────────────────────────────────
# PDF field parsing (mirrors 911 Setup's labeled-value approach)
# ──────────────────────────────────────────────────────────────────────────

_LABEL_RE = re.compile(r'^[A-Z][A-Z0-9 ./#-]*:')
_NUM_RE = re.compile(r'^-?\d*\.?\d+$')
_CUT_METHOD_RE = re.compile(r'\b(PLASMA|LASER|WATER\s*JET|WATERJET|WJET|TELESIS|OXY\w*)\b',
                            re.IGNORECASE)
# "1.000 THK #109X360"  /  "0.500X90X360"  /  "0.375 THICK"
_SIZE_LW_HASH_RE = re.compile(r'#\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)')
_SIZE_LW_PLAIN_RE = re.compile(
    r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)')
_THK_FROM_DESC_RE = re.compile(r'(\d*\.?\d+)\s*TH(?:K|ICK)', re.IGNORECASE)


def _looks_like_label(s: str) -> bool:
    return bool(_LABEL_RE.match(s.strip()))


def _labeled_value(text: str, label: str):
    """Value following 'label:' -- same line or the next line. Returns None
    when blank or when the following token is itself another field label."""
    pat = re.compile(
        re.escape(label) + r'\s*:[ \t]*(?P<same>[^\n]*)(?:\n[ \t]*(?P<next>[^\n]*))?',
        re.IGNORECASE,
    )
    m = pat.search(text)
    if not m:
        return None
    same = (m.group("same") or "").strip()
    if same and not _looks_like_label(same):
        return same
    nxt = (m.group("next") or "").strip()
    if nxt and not _looks_like_label(nxt):
        return nxt
    return None


def _find_nest_pdf(folder: Path, nest_number: str):
    """The PDF in `folder` whose stem contains the nest number, or None."""
    if not folder.exists():
        return None
    nest_upper = nest_number.upper()
    for f in folder.iterdir():
        if (f.is_file() and f.suffix.lower() == ".pdf"
                and nest_upper in f.stem.upper()):
            return f
    return None


def _page1_pieces_thickness(page1_text: str):
    """Pull (pieces, thickness, length, width) from the page-1 'Orders Pieces
    Thickness Length Width' data row. Labels appear as a block, then the values
    as a block, so we find the label block and read the value sequence after it.
    Length/Width are the nest sheet dimensions (the remnant's size when the nest
    is cut from a remnant).

    Returns (pieces|None, thickness|None, length|None, width|None).
    """
    lines = [ln.strip() for ln in page1_text.splitlines()]
    labels = ["ORDERS", "PIECES", "THICKNESS", "LENGTH", "WIDTH"]
    start = None
    for i in range(len(lines) - len(labels) + 1):
        if [lines[i + k].upper() for k in range(len(labels))] == labels:
            start = i + len(labels)
            break
    if start is None:
        return None, None, None, None
    # Collect the numeric value tokens that follow (skip any trailing label like
    # 'Remnant Used', whose own value comes AFTER Width). Order is Orders, Pieces,
    # Thickness, Length, Width -- so the first five numbers are what we need.
    nums = []
    for ln in lines[start:]:
        if _NUM_RE.match(ln):
            nums.append(ln)
        if len(nums) >= 5:
            break
    pieces = _to_float(nums[1]) if len(nums) >= 2 else None
    thickness = _to_float(nums[2]) if len(nums) >= 3 else None
    length = _to_float(nums[3]) if len(nums) >= 4 else None
    width = _to_float(nums[4]) if len(nums) >= 5 else None
    # Pieces is a count -> int-ish
    if pieces is not None:
        pieces = int(round(pieces))
    return pieces, thickness, length, width


def _page1_orders_pieces(page1_text: str):
    """(orders, pieces) from the page-1 data row -- works for BOTH packet
    layouts. Plate nest sheets label the row 'Orders Pieces Thickness Length
    Width'; shape (1D) nest sheets label it 'Orders Pieces Material Size
    Material Type Material Style'. Either way the labels come as a block with
    ORDERS then PIECES adjacent and the values follow it, so the first two
    numeric lines after the label block are the Orders and Pieces counts. The
    scan window is bounded so numbers further down the page (e.g. a shape
    sheet's Summary of Batches) can't be misread as the data row."""
    lines = [ln.strip() for ln in page1_text.splitlines()]
    for i in range(len(lines) - 1):
        if lines[i].upper() == "ORDERS" and lines[i + 1].upper() == "PIECES":
            nums = []
            for ln in lines[i + 2:i + 12]:
                if _NUM_RE.match(ln):
                    nums.append(_to_float(ln))
                    if len(nums) == 2:
                        return (int(round(nums[0])) if nums[0] is not None else None,
                                int(round(nums[1])) if nums[1] is not None else None)
            break
    return None, None


def _page1_source_code(page1_text: str):
    """The page-1 'Source Code' value -- the stock identifier the Working
    Forecast's Source Mat'l. column carries. Same forms as the batch list's
    Material column: 'EB218014388A', bare '218019613', '42-07-3105'. The
    packet prints it in a stacked label layout (no colon): the 'Source Code' /
    'Alternate Source Code' label lines, then the value(s). Accept the first
    following token that contains a digit (every stock code does; field labels
    don't) and stop at the next field block ('Orders')."""
    lines = [ln.strip() for ln in page1_text.splitlines()]
    for i, ln in enumerate(lines):
        if ln.upper() == "SOURCE CODE":
            for cand in lines[i + 1:i + 6]:
                cu = cand.upper()
                if cu in ("", "ALTERNATE SOURCE CODE"):
                    continue
                if cu == "ORDERS":  # ran into the next field block: no value
                    break
                if any(ch.isdigit() for ch in cand):
                    return cand
            break
    return None


# Date-shaped words ('May' / '23,' / '2026' / '1:39' / 'PM') -- the 'Date
# Package Printed' cell is the Fuel cell's right-hand neighbour, so these are
# the likeliest junk to drift into a value column window; dismiss them.
_DATE_WORD_RE = re.compile(
    r'^(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|June?|July?|'
    r'Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?|'
    r'\d{1,2},?|\d{4}|\d{1,2}:\d{2}|[AP]M)$', re.IGNORECASE)


def _page1_value_below(words, label_texts):
    """The value cell directly BELOW a page-1 header-table label ('Machine' /
    'Gantry' / 'Fuel'), read positionally from the page's word boxes
    (fitz get_text('words') tuples: x0, y0, x1, y1, text, ...). The flat text
    extraction interleaves this table's cells out of label order, so line
    scanning can't pair them. Takes the topmost word matching a label, then
    the first text line that starts below it within one row height and
    horizontally centered on the label's column; the row-height cap keeps a
    blank cell from reading the next row's labels ('Source Code'), and
    date-shaped words are dismissed so the neighbouring 'Date Package
    Printed' value can never bleed in. Verified on all 142 Award 8/9/10
    packets (both the 'Gantry' and 'Machine' label variants). Returns None
    when the label or its value is missing."""
    cands = [w for w in words if w[4].strip().lower() in label_texts]
    if not cands:
        return None
    lab = min(cands, key=lambda w: w[1])
    lab_cx = (lab[0] + lab[2]) / 2
    below = [w for w in words if w[1] > lab[3] - 1 and w[1] < lab[3] + 15
             and abs((w[0] + w[2]) / 2 - lab_cx) < 70
             and not _DATE_WORD_RE.match(w[4].strip())]
    if not below:
        return None
    top = min(w[1] for w in below)
    line = sorted((w for w in below if w[1] - top < 3), key=lambda w: w[0])
    return " ".join(w[4] for w in line).strip() or None


def _parse_size(full_text: str):
    """Return (stock_L, stock_W) from a part-sketch 'SIZE:' field, normalized
    L=max, W=min. Thickness-only sizes (e.g. '0.375 THICK') -> (None, None).

    A packet can carry more than one 'SIZE:' field (e.g. the nest sheet's
    usable area '85.06 X 53.74' AND the part-sketch stock size
    '1.000 THK #109X360'). We scan every SIZE value and prefer a stock-format
    match: the '#L X W' hash form first, then the 'thk X L X W' three-number
    form.
    """
    values = []
    for m in re.finditer(r'SIZE\s*:[ \t]*([^\n]*)(?:\n[ \t]*([^\n]*))?', full_text,
                         re.IGNORECASE):
        same = (m.group(1) or "").strip()
        cand = same if (same and not _looks_like_label(same)) else (m.group(2) or "").strip()
        if cand and not _looks_like_label(cand):
            values.append(cand)
    # Prefer the '#L X W' hash form.
    for v in values:
        mm = _SIZE_LW_HASH_RE.search(v)
        if mm:
            a, b = float(mm.group(1)), float(mm.group(2))
            return max(a, b), min(a, b)
    # Then the 'thk X L X W' three-number form (ignore the leading thickness).
    for v in values:
        mm = _SIZE_LW_PLAIN_RE.search(v)
        if mm:
            dims = sorted((float(mm.group(2)), float(mm.group(3))))
            return dims[1], dims[0]
    return None, None


def _summary_batch_lengths(pages):
    """Lengths (inches) from the 'Summary of Batches' table on a SHAPE nest
    order page (1D nests; plate packets don't carry the table).

    Extracted text layout: the labels 'Batch' / '# of' / 'Parts' / 'Length'
    each on their own line, then the table rows as a flat run of numbers in
    triples (batch, # of parts, length), terminated by the next label
    ('Laser NS'). Returns (lengths, malformed): lengths = one value per table
    row, or None when no packet page has the table; malformed = True when the
    table was found but its value run wasn't clean triples.
    """
    for txt in pages:
        if "SUMMARY OF BATCHES" not in txt.upper():
            continue
        lines = [ln.strip() for ln in txt.splitlines()]
        up = [ln.upper() for ln in lines]
        start = None
        for i, u in enumerate(up):
            if u != "BATCH":
                continue
            j = i + 1
            # '# of' + 'Parts' wrap onto one or two lines depending on export.
            if j < len(up) and up[j].startswith("# OF"):
                j += 1
                if j < len(up) and up[j] == "PARTS":
                    j += 1
            if j < len(up) and up[j] == "LENGTH":
                start = j + 1
                break
        if start is None:
            continue
        nums = []
        for ln in lines[start:]:
            if not _NUM_RE.match(ln):
                break                     # first non-numeric line ends the table
            nums.append(float(ln.replace(",", "")))
        if not nums or len(nums) % 3:
            return None, True             # found the table but couldn't read it
        return [nums[k + 2] for k in range(0, len(nums), 3)], False
    return None, False


def shape_feet_required(lengths):
    """Receiving's spec: each Summary-of-Batches row's Length (inches) / 12,
    rounded UP to the whole foot, then summed (143 + 20 -> 12 + 2 = 14 ft)."""
    return sum(math.ceil(l / 12.0) for l in lengths)


def _material_size_type(pages):
    """Best-effort (size, type) from the 'Material Size / Material Type /
    Material Style' block on a shape nest order page. Size formats vary --
    '2.500 X 2.500 X 0.375 THK' (angle/tube) but also '0.500THK2.000W' (bar,
    no X separators) -- so the size is the first value line mixing digits and
    letters; the type ('ANGLE'/'BAR'/...) is the letters-only line after it.
    Either may be None."""
    for txt in pages:
        lines = [ln.strip() for ln in txt.splitlines()]
        up = [ln.upper() for ln in lines]
        for i, u in enumerate(up):
            if (u == "MATERIAL SIZE" and i + 2 < len(up)
                    and up[i + 1] == "MATERIAL TYPE"
                    and up[i + 2] == "MATERIAL STYLE"):
                size = mtype = None
                for ln in lines[i + 3:i + 13]:
                    if size is None:
                        if (any(ch.isdigit() for ch in ln)
                                and any(ch.isalpha() for ch in ln)):
                            size = ln
                        continue
                    # First letters-only line after the size = the type. Real
                    # types are ALL CAPS ('ANGLE', 'I SECTION'); a mixed-case
                    # hit is the next form label ('Mark Operator') leaking in
                    # because the type cell was blank -- treat as no type.
                    if ln and not any(ch.isdigit() for ch in ln):
                        mtype = ln if ln == ln.upper() else None
                        break
                return size, mtype
    return None, None


def parse_nest_pdf(pdf_path: Path) -> dict:
    """Parse a nest packet PDF. Returns a dict of extracted fields (any may be
    None). Thickness is the only field the calculation needs."""
    out = {"thickness": None, "pieces": None, "material": None,
           "stock_l": None, "stock_w": None, "mil_spec": None,
           "plate_weight": None, "process": None, "mt_material": None,
           "sheet_length": None, "sheet_width": None,
           "batch_lengths": None, "summary_malformed": False,
           "mat_size": None, "mat_type": None,
           "orders": None, "source_code": None,
           "machine": None, "fuel": None}
    sdk.ensure_local(pdf_path)  # OneDrive placeholder -> download first (Hard Rule 13)
    doc = fitz.open(str(pdf_path))
    try:
        pages = [p.get_text() for p in doc]
        # Word boxes for page 1 only: the Machine/Fuel header table needs
        # position-based reads (the flat text interleaves its cells).
        words1 = doc[0].get_text("words") if doc.page_count else []
    finally:
        doc.close()
    full = "\n".join(pages)
    page1 = pages[0] if pages else ""

    (out["pieces"], out["thickness"],
     out["sheet_length"], out["sheet_width"]) = _page1_pieces_thickness(page1)

    # Working Forecast Input fields: the page-1 Orders count + the Source Code
    # (EB stock identifier). Both packet layouts (plate + shape) carry them;
    # shape sheets lack the Thickness/Length/Width labels, so their Pieces
    # only resolves here.
    out["orders"], _wf_pieces = _page1_orders_pieces(page1)
    if out["pieces"] is None:
        out["pieces"] = _wf_pieces
    out["source_code"] = _page1_source_code(page1)

    # Cut method / process from page 1 (PLASMA / LASER / ...).
    m = _CUT_METHOD_RE.search(page1)
    if m:
        out["process"] = m.group(1).upper().replace("  ", " ")

    # EB Machine / EB Fuel: the page-1 header table's Machine (labelled
    # 'Gantry' on some packets) and Fuel cells, read positionally.
    out["machine"] = _page1_value_below(words1, ("machine", "gantry"))
    out["fuel"] = _page1_value_below(words1, ("fuel",))

    # Material: REMNANT 'TYPE:' OR part-sketch 'MATL:' OR move-ticket 'MATERIAL:'.
    out["material"] = (_labeled_value(full, "TYPE")
                       or _labeled_value(full, "MATL")
                       or _labeled_value(full, "MATERIAL"))
    # MOVE TICKET material designation (e.g. 'HSS'), distinct from the batch
    # list's stock 'Material' code -- pulled only from MOVE TICKET pages.
    out["mt_material"] = _move_ticket_material(pages)
    # Mil spec.
    out["mil_spec"] = _labeled_value(full, "MIL SPEC")
    if not out["mil_spec"]:
        mm = re.search(r'MIL-S-\S+', full, re.IGNORECASE)
        out["mil_spec"] = mm.group(0) if mm else None
    # Stock L x W from a part sketch.
    out["stock_l"], out["stock_w"] = _parse_size(full)
    # Plate weight: REMNANT 'WEIGHT: n LBS' else page-1 'Plate/Max Part Weight'.
    out["plate_weight"] = _parse_plate_weight(full, page1)

    # Thickness fallback from REMNANT 'THICK:' if page-1 row missed it.
    if out["thickness"] is None:
        tv = _labeled_value(full, "THICK")
        if tv:
            out["thickness"] = _to_float(tv)

    # Shape (1D) nest order data: the Summary of Batches lengths that drive
    # 'Shape Ft Req', plus the stock Material Size / Type for the Ft Req sheet.
    out["batch_lengths"], out["summary_malformed"] = _summary_batch_lengths(pages)
    out["mat_size"], out["mat_type"] = _material_size_type(pages)
    return out


def _move_ticket_material(pages) -> Optional[str]:
    """The 'MATERIAL:' value from a MOVE TICKET page (the part-material
    designation, e.g. 'HSS') -- the value that sits right after 'MIL SPEC:' on
    the move ticket. Returns the first non-empty one found, scanning only pages
    that mention MOVE TICKET so a stray 'MATERIAL' elsewhere can't win."""
    for txt in pages:
        if "MOVE TICKET" in txt.upper():
            val = _labeled_value(txt, "MATERIAL")
            if val:
                return val
    return None


def _parse_plate_weight(full: str, page1: str):
    w = _labeled_value(full, "WEIGHT")
    if w:
        mm = re.search(r'(\d+(?:\.\d+)?)', w)
        if mm:
            return _to_float(mm.group(1))
    # page-1 'Plate/Max Part Weight  1181 / 3' -> first number before slash.
    mm = re.search(r'Plate/Max Part Weight\s*[^0-9]*(\d+(?:\.\d+)?)', page1,
                   re.IGNORECASE)
    if mm:
        return _to_float(mm.group(1))
    return None


def _to_float(v):
    if v is None:
        return None
    try:
        return float(str(v).strip().replace(",", ""))
    except (ValueError, TypeError):
        return None


def thickness_from_description(desc) -> Optional[float]:
    """Fallback thickness parse from a batch-list Description like
    'PLATE ; STL ; 1.125 THK #90 X 360 ; SF'."""
    if not desc:
        return None
    m = _THK_FROM_DESC_RE.search(str(desc))
    return _to_float(m.group(1)) if m else None


# ──────────────────────────────────────────────────────────────────────────
# Folder + batch-list resolution
# ──────────────────────────────────────────────────────────────────────────

def _looks_like_order(folder: Path) -> bool:
    """True if `folder` is a single order/batch folder — it directly holds a
    'CUI- TECH DATA READ ME' folder, a *BATCH*LIST*.xlsx, or the packet
    folders themselves."""
    try:
        for sub in folder.iterdir():
            name_up = sub.name.upper()
            if sub.is_dir():
                if "CUI" in name_up and "TECH" in name_up:
                    return True
                if name_up in ("NEST PACKAGES", "WPDD SKETCHES"):
                    return True
            elif (sub.suffix.lower() == ".xlsx" and "BATCH" in name_up
                    and "LIST" in name_up and not sub.name.startswith("~$")):
                return True
    except OSError:
        return False
    return False


def discover_order_folders(root: Path, log) -> list:
    """Collect the order folders under ROOT, tolerating a ROOT picked a level
    (or several) too high.

    Award folders and zip extractions add wrapper levels (e.g.
    'Award 10\\1000129724 SSPO Award 10\\1000129724 SSPO Award 10\\{L028,...}').
    The old ROOT.iterdir()-only discovery treated a wrapper as ONE order, then
    the CUI/batch-list fallbacks silently processed a single arbitrary batch
    list from somewhere inside it (bit A.T.'s Award 10 runs: 1 of 4 batches,
    no packet PDFs, everything mis-filed). Now: a ROOT subfolder that doesn't
    itself look like an order is descended into (depth-capped) until real
    order folders are found; picking an order folder itself as ROOT works too.
    """
    if _looks_like_order(root):
        log("ROOT is itself a single order folder.")
        return [root]

    found = []

    def walk(folder: Path, depth: int) -> None:
        try:
            subs = sorted((d for d in folder.iterdir()
                           if d.is_dir() and not d.name.startswith(".")),
                          key=lambda p: p.name.upper())
        except OSError:
            return
        for d in subs:
            if _looks_like_order(d):
                found.append(d)
            elif depth < 4:
                before = len(found)
                walk(d, depth + 1)
                if len(found) > before:
                    log(f"  NOTE: '{d.name}' is a wrapper folder - using the "
                        f"order folders inside it.")

    walk(root, 1)
    return found


def find_cui_folder(order_folder: Path) -> Optional[Path]:
    """Locate the 'CUI- TECH DATA READ ME' folder (name varies in spacing/dash).
    Falls back to any subfolder that itself contains a NEST PACKAGES / WPDD
    SKETCHES folder or a *BATCH*LIST*.xlsx. Returns the order folder itself if
    the packets live directly under it."""
    for sub in order_folder.iterdir():
        if sub.is_dir() and "CUI" in sub.name.upper() and "TECH" in sub.name.upper():
            return sub
    # Fallback: a subfolder holding the packet folders or a batch list.
    for sub in order_folder.iterdir():
        if not sub.is_dir():
            continue
        names = {c.name.upper() for c in sub.iterdir() if c.is_dir()} if sub.exists() else set()
        if "NEST PACKAGES" in names or "WPDD SKETCHES" in names:
            return sub
        if any("BATCH" in f.name.upper() and "LIST" in f.name.upper()
               for f in sub.glob("*.xlsx")):
            return sub
    # Maybe packets are directly under the order folder.
    return order_folder


def find_pdf_folder(cui_folder: Path) -> Optional[Path]:
    """Prefer 'NEST PACKAGES' (nest-named packet PDFs); fall back to
    'WPDD SKETCHES', then the CUI folder itself."""
    for name in ("NEST PACKAGES", "WPDD SKETCHES"):
        cand = cui_folder / name
        if cand.exists() and any(cand.glob("*.pdf")):
            return cand
    # Case-insensitive scan.
    for sub in cui_folder.iterdir():
        if sub.is_dir() and sub.name.upper() in ("NEST PACKAGES", "WPDD SKETCHES"):
            if any(sub.glob("*.pdf")):
                return sub
    if any(cui_folder.glob("*.pdf")):
        return cui_folder
    return None


def find_iges_folder(cui_folder: Path) -> Optional[Path]:
    """Locate the 'IGES FILES' folder (holds the nest DXFs) under the CUI folder,
    case-insensitively. None if absent (e.g. a structural-only order)."""
    if cui_folder is None:
        return None
    for sub in cui_folder.iterdir():
        if sub.is_dir() and sub.name.strip().upper() == "IGES FILES":
            return sub
    return None


def find_batch_list(cui_folder: Path, log=None) -> Optional[Path]:
    """Glob for the '*BATCH*LIST*.xlsx' workbook, skipping Excel lock files."""
    candidates = [f for f in cui_folder.glob("*.xlsx")
                  if "BATCH" in f.name.upper() and "LIST" in f.name.upper()
                  and not f.name.startswith("~$")]
    if candidates:
        return candidates[0]
    # Deep fallback. Multiple hits at different depths is the wrong-ROOT
    # signature (a wrapper folder holding several batches) — never silently
    # pick one; take the shallowest and say so.
    deep = [f for f in cui_folder.rglob("*.xlsx")
            if "BATCH" in f.name.upper() and "LIST" in f.name.upper()
            and not f.name.startswith("~$")]
    if not deep:
        return None
    deep.sort(key=lambda f: (len(f.parts), str(f).upper()))
    if len(deep) > 1 and log:
        log(f"  WARNING: {len(deep)} BATCH LIST workbooks found under "
            f"'{cui_folder.name}' - it looks like a wrapper of several "
            f"batches, not one order. Using {deep[0].name}; check the ROOT "
            f"you selected.")
    return deep[0]


def load_batch_rows(batch_list_path: Path):
    """Return (headers_in_order, rows) from the batch-list sheet.

    headers_in_order : list[str]  -- the header names in column order
    rows             : list[dict] -- {HEADER_UPPER: value} per data row

    Locates the sheet + header row by scanning for 'Nest Pkg Nbr'/'PPN Quantity'
    so it survives column/row shifts and extra sheets (SCRAP, Sheet3 ...).
    """
    wb = sdk.load_workbook_resilient(batch_list_path, data_only=True)
    try:
        target_ws = None
        header_row = None
        cols = {}
        for ws in wb.worksheets:
            hr, cmap = sdk.find_header_row(ws, ["Nest Pkg Nbr", "PPN Quantity"])
            if hr is not None:
                target_ws, header_row, cols = ws, hr, cmap
                break
        if target_ws is None:
            return [], []

        # Header names in column order (preserve original casing for output).
        ordered = []
        for c in range(1, target_ws.max_column + 1):
            v = target_ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip():
                ordered.append((c, v.strip()))
        headers_in_order = [name for _, name in ordered]

        rows = []
        nest_col = cols.get("NEST PKG NBR")
        for r in range(header_row + 1, target_ws.max_row + 1):
            # Skip fully blank rows and rows without a nest.
            nest_val = target_ws.cell(r, nest_col).value if nest_col else None
            if nest_val is None or str(nest_val).strip() == "":
                continue
            row = {}
            for c, name in ordered:
                row[name.upper()] = target_ws.cell(r, c).value
            rows.append(row)
        return headers_in_order, rows
    finally:
        wb.close()


# ──────────────────────────────────────────────────────────────────────────
# Workbook output
# ──────────────────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="305496")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_PIVOT_FILL = PatternFill("solid", fgColor="1F4E78")
_TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
_MISSING_FILL = PatternFill("solid", fgColor="FFFF00")   # no estimate (yellow)
_OUTLIER_FILL = PatternFill("solid", fgColor="F4B084")   # LI outlier (orange)
_MULTI_FILL = PatternFill("solid", fgColor="DDEBF7")     # multi-layered (light blue)
_NODXF_FILL = PatternFill("solid", fgColor="FF6666")     # no DXF match / band estimate (red)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _is_date(v):
    return isinstance(v, (datetime.datetime, datetime.date))


def write_analysis_sheet(wb, plate_rows, log):
    """Add an 'Analysis' sheet of summary charts driven by the plate rows.

    The little aggregate data tables the charts read from are parked in far-right
    columns (T+) out of the way; the charts float over the left grid. Charts are
    skipped cleanly when their category has no data (e.g. no DXF-resolved rows).
    """
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    title = ws.cell(1, 1, "Batch Analysis")
    title.font = Font(bold=True, size=14)
    ws.cell(2, 1, "Counts are per part (one batch-list row); charts read the "
                  "tables parked to the right (columns T+).").font = Font(italic=True,
                                                                          color="808080")

    # ── aggregations ──────────────────────────────────────────────────────
    thick = {}
    for r in plate_rows:
        t = r.get("Thickness (in)")
        if isinstance(t, (int, float)):
            k = round(float(t), 3)
            thick[k] = thick.get(k, 0) + 1
    thick_items = [(f'{k:g}"', v) for k, v in sorted(thick.items())]

    single = sum(1 for r in plate_rows if r.get("Multi-Layered") == "N")
    multi = sum(1 for r in plate_rows if r.get("Multi-Layered") == "Y")
    layer_items = [("Single-Layer", single), ("Multi-Layered", multi)]

    has_empty = sum(1 for r in plate_rows if r.get("_empty_layers") is True)
    no_empty = sum(1 for r in plate_rows if r.get("_empty_layers") is False)
    empty_items = [("All layers cut", no_empty), ('Has 0" layer', has_empty)]

    # Est Cut Hours grouped by plate thickness (the "time per material" view --
    # material here means thickness). Sits beside the Plate Thickness pie so the
    # part-count split and the cut-time split read against the same buckets.
    thick_hours = {}
    for r in plate_rows:
        t = r.get("Thickness (in)")
        if isinstance(t, (int, float)):
            k = round(float(t), 3)
            thick_hours[k] = thick_hours.get(k, 0.0) + (r.get("Est Cut Hours") or 0.0)
    thick_hours_items = [(f'{k:g}"', round(h, 3)) for k, h in sorted(thick_hours.items())]

    mat = {}
    for r in plate_rows:
        name = (str(r.get("Material") or r.get("Source Material") or "").strip()
                or "(blank)")
        mat[name] = mat.get(name, 0) + 1
    mat_items = sorted(mat.items(), key=lambda kv: -kv[1])

    # Data-quality / flag breakdown: how many parts carry each attention flag
    # (a part can carry more than one, so bars needn't sum to the part count).
    # Benign feed-rounding notes ('rounded up to ..'/'above chart max') are NOT
    # counted, so a row carrying only those still reads as 'Clean'.
    QUALITY = [
        ("no thickness", "No thickness"),
        ("thk from desc", "Thickness from desc"),
        ("no dxf match", "No DXF match"),
        ("unmeasured", "Unmeasured geometry"),
        ("outlier", "LI outlier"),
        ("insunits", "Unexpected units"),
    ]
    flag_counts, clean = {}, 0
    for r in plate_rows:
        fl = str(r.get("Flags") or "").lower()
        hits = {label for needle, label in QUALITY if needle in fl}
        if hits:
            for label in hits:
                flag_counts[label] = flag_counts.get(label, 0) + 1
        else:
            clean += 1
    flag_items = [("Clean", clean)] + sorted(flag_counts.items(), key=lambda kv: -kv[1])

    # Bevel scope is review-worthy but not a data defect, so it gets its own
    # chart rather than swamping the Data Quality bars.
    bevel = sum(1 for r in plate_rows if "bevel" in str(r.get("Flags") or "").lower())
    bevel_items = [("Straight cut", len(plate_rows) - bevel), ("Bevel scope", bevel)]

    # ── data tables (columns T/U, stacked) ────────────────────────────────
    data_col = 20  # column T
    state = {"row": 1}

    def put_table(heading, items, value_title="Parts", fmt=None):
        hr = state["row"]
        ws.cell(hr, data_col, heading).font = Font(bold=True)
        ws.cell(hr, data_col + 1, value_title).font = Font(bold=True)
        first = hr + 1
        for i, (k, v) in enumerate(items):
            ws.cell(first + i, data_col, str(k))
            cell = ws.cell(first + i, data_col + 1, v)
            if fmt:
                cell.number_format = fmt
        last = first + len(items) - 1
        state["row"] = (last if items else hr) + 3
        if not items or sum(float(v) for _, v in items) <= 0:
            return None, None
        cats = Reference(ws, min_col=data_col, min_row=first, max_row=last)
        data = Reference(ws, min_col=data_col + 1, min_row=hr, max_row=last)
        return cats, data

    def pie(chart_title, cats, data, anchor):
        if cats is None or data is None:
            return
        ch = PieChart()
        ch.title = chart_title
        ch.height, ch.width = 8, 12
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        # The category names live in the legend (right); the on-slice labels
        # show ONLY the percentage so many-slice pies (thickness/material) don't
        # pile category + series + value text on top of each other. "bestFit"
        # lets Excel scatter the labels to reduce overlap.
        dl = DataLabelList()
        dl.showPercent = True
        dl.showCatName = False
        dl.showSerName = False
        dl.showVal = False
        dl.showLegendKey = False
        dl.numFmt = "0%"
        dl.dLblPos = "bestFit"
        ch.dataLabels = dl
        ch.legend.position = "r"
        ch.legend.overlay = False
        ws.add_chart(ch, anchor)

    def bar(chart_title, cats, data, anchor, value_labels=True):
        if cats is None or data is None:
            return
        ch = BarChart()
        ch.type = "col"
        ch.title = chart_title
        ch.height, ch.width = 8, 14
        ch.legend = None
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        # openpyxl leaves both axes flagged deleted by default, so Excel hides
        # the category (x) tick labels -- that's why the bars looked "keyless".
        # Forcing delete=False restores the labels under each bar.
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        if value_labels:
            dl = DataLabelList()
            dl.showVal = True
            dl.showCatName = False
            dl.showSerName = False
            dl.showLegendKey = False
            ch.dataLabels = dl
        ws.add_chart(ch, anchor)

    c_thick = put_table("Plate Thickness", thick_items)
    c_layer = put_table("Layering", layer_items)
    c_empty = put_table("Zero-inch Layers", empty_items)
    c_thick_hours = put_table("Est Cut Hours by Thickness", thick_hours_items,
                              "Est Cut Hours", "0.000")
    c_mat = put_table("Material", mat_items)
    c_flag = put_table("Data Quality", flag_items)
    c_bevel = put_table("Bevel Scope", bevel_items)

    # ── charts (2-wide grid over the left of the sheet) ────────────────────
    pie("Plate Thickness (in)", *c_thick, anchor="A4")
    bar("Est Cut Hours by Thickness", *c_thick_hours, anchor="J4", value_labels=False)
    pie("Single vs Multi-Layered", *c_layer, anchor="A22")
    pie("Material Distribution", *c_mat, anchor="J22")
    pie('Parts with a 0" Layer', *c_empty, anchor="A40")
    bar("Data Quality (parts per flag)", *c_flag, anchor="J40")
    pie("Bevel vs Straight Cut", *c_bevel, anchor="A58")

    # ── DXF coverage callout (bottom of sheet) ────────────────────────────
    # Plate parts with NO DXF match still get the throughput estimate (v2.7.0),
    # but their far-right LI reference columns are empty; group them by
    # order+nest so the missing IGES/DXF files can be chased. These same rows
    # are filled red on the Plates sheet.
    missing_dxf = {}
    for r in plate_rows:
        if "no dxf match" not in str(r.get("Flags") or "").lower():
            continue
        key = (str(r.get("Order") or "").strip(), str(r.get("Nest Pkg Nbr") or "").strip())
        d = missing_dxf.setdefault(key, {"count": 0, "thk": set()})
        d["count"] += 1
        t = r.get("Thickness (in)")
        if isinstance(t, (int, float)):
            d["thk"].add(round(float(t), 3))

    mr = 78
    ws.cell(mr, 1, "DXF Coverage").font = Font(bold=True, size=12)
    if not missing_dxf:
        ws.cell(mr + 1, 1, "All plate parts matched a DXF -- every row has the "
                           "linear-inch reference columns alongside its "
                           "throughput estimate.").font = Font(italic=True,
                                                               color="548235")
        next_row = mr + 3
    else:
        total_missing = sum(d["count"] for d in missing_dxf.values())
        ws.cell(mr + 1, 1,
                f"{total_missing} plate part(s) across {len(missing_dxf)} nest(s) "
                f"had NO DXF match -- their cut time still estimates from the "
                f"pieces/hour table, but the linear-inch reference columns are "
                f"empty (rows highlighted red on the Plates sheet). Chase down "
                f"these IGES/DXF files:").font = Font(bold=True, color="C00000")
        for c, name in enumerate(("Order", "Nest", "Parts w/o DXF", "Thickness(es)"),
                                 start=1):
            cell = ws.cell(mr + 3, c, name)
            cell.font = _HDR_FONT
            cell.fill = _HDR_FILL
            cell.border = _BORDER
        rr = mr + 4
        for (order, nest), d in sorted(missing_dxf.items()):
            thks = ", ".join(f'{t:g}"' for t in sorted(d["thk"])) or "-"
            for c, v in enumerate((order or "-", nest or "-", d["count"], thks),
                                  start=1):
                cell = ws.cell(rr, c, v)
                cell.border = _BORDER
                cell.fill = _NODXF_FILL
            rr += 1
        next_row = rr + 2

    # ── flagged deviations (LI outliers) ──────────────────────────────────
    # Parts whose measured Linear In/PC sits past LI_OUTLIER_SD from the mean --
    # the orange rows on the Plates sheet. Listed most-extreme first so an
    # inflated/odd geometry is easy to eyeball against its neighbours.
    outliers = [r for r in plate_rows
                if "outlier" in str(r.get("Flags") or "").lower()]

    def _abs_sd(r):
        sd = r.get("LI Dev (SD)")
        return abs(sd) if isinstance(sd, (int, float)) else 0.0
    outliers.sort(key=_abs_sd, reverse=True)

    dr = next_row
    ws.cell(dr, 1, "Flagged Deviations (LI outliers)").font = Font(bold=True, size=12)
    if not outliers:
        ws.cell(dr + 1, 1, "No linear-inch outliers -- every measured part sits "
                           "within tolerance of the mean.").font = Font(italic=True,
                                                                        color="548235")
    else:
        ws.cell(dr + 1, 1,
                f"{len(outliers)} measured plate part(s) sit past the outlier "
                f"threshold (rows highlighted orange on the Plates sheet). Review "
                f"the geometry on these:").font = Font(bold=True, color="C55A11")
        cols = ("Order", "Nest", "Work Order", "Thickness", "Linear In/PC",
                "LI Dev (SD)")
        for c, name in enumerate(cols, start=1):
            cell = ws.cell(dr + 3, c, name)
            cell.font = _HDR_FONT
            cell.fill = _HDR_FILL
            cell.border = _BORDER
        er = dr + 4
        for r in outliers:
            t = r.get("Thickness (in)")
            thk = f'{float(t):g}"' if isinstance(t, (int, float)) else "-"
            li = r.get("Linear In/PC")
            sd = r.get("LI Dev (SD)")
            vals = (str(r.get("Order") or "-").strip(),
                    str(r.get("Nest Pkg Nbr") or "-").strip(),
                    str(r.get("Work Order") or "-").strip(), thk,
                    round(li, 1) if isinstance(li, (int, float)) else "-",
                    round(sd, 2) if isinstance(sd, (int, float)) else "-")
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(er, c, v)
                cell.border = _BORDER
                cell.fill = _OUTLIER_FILL
            er += 1

    log("  Added Analysis sheet (thickness, cut hours by thickness, layering, "
        "material, 0-inch layers, data-quality flags, bevel scope, DXF coverage, "
        "flagged deviations).")


def write_workbook(out_path: Path, data_headers, plate_rows, nonplate_rows,
                   shape_summary, forecast_summary, log):
    """Write the workbook with two identically-structured sheets: 'Plates' and
    'Non-Plates'. Each holds ONLY the flat data table (every batch-list column
    plus our generated columns) with the yellow missing-data highlighting; the
    real PivotTables are added afterwards via Excel COM (see add_real_pivots).
    A 'Shape Ft Req' rollup sheet (one line per shape nest) follows when any
    shape nest carried a Summary of Batches table, then the 'Working Forecast
    Input' sheet (one line per nest, forecast-input column layout).

    Returns sheets_meta: [(sheet_name, data_row_count), ...] for the pivot pass.
    """
    wb = Workbook()
    ws_plate = wb.active
    ws_plate.title = "Plates"
    write_data_table(ws_plate, data_headers, plate_rows)

    ws_other = wb.create_sheet("Non-Plates")
    write_data_table(ws_other, data_headers, nonplate_rows)

    if shape_summary:
        write_shape_ft_sheet(wb, shape_summary, log)

    if forecast_summary:
        write_working_forecast_sheet(wb, forecast_summary, log)

    write_analysis_sheet(wb, plate_rows, log)

    wb.save(str(out_path))
    return [("Plates", len(plate_rows)), ("Non-Plates", len(nonplate_rows))]


SHAPE_FT_HEADERS = ["Order", "Nest Pkg Nbr", "Material Type", "Material Size",
                    "Description", "Lengths (in)", "Ft (rounded up)",
                    "Total Ft Req"]


def write_shape_ft_sheet(wb, shape_summary, log):
    """The receiving cheat-sheet feed: one row per shape nest with its stock
    material and the Summary-of-Batches feet math spelled out ('143, 20' ->
    '12 + 2' -> 14), plus a grand-total row. Feet round UP per table row
    BEFORE summing (receiving's spec)."""
    ws = wb.create_sheet("Shape Ft Req")
    for c, name in enumerate(SHAPE_FT_HEADERS, start=1):
        cell = ws.cell(1, c, name)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER
    ri = 1
    for entry in shape_summary:
        ri += 1
        feet = [math.ceil(l / 12.0) for l in entry["lengths"]]
        vals = [entry["order"], entry["nest"], entry["mat_type"],
                entry["mat_size"], entry["desc"],
                ", ".join(_eqnum(l) for l in entry["lengths"]),
                " + ".join(str(f) for f in feet),
                entry["ft"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(ri, c, v)
            cell.border = _BORDER
    total_cell = ws.cell(ri + 1, len(SHAPE_FT_HEADERS) - 1, "TOTAL")
    total_cell.font = Font(bold=True)
    total_cell.border = _BORDER
    val_cell = ws.cell(ri + 1, len(SHAPE_FT_HEADERS),
                       sum(e["ft"] for e in shape_summary))
    val_cell.font = Font(bold=True)
    val_cell.border = _BORDER
    ws.freeze_panes = "A2"
    _autosize(ws, SHAPE_FT_HEADERS)
    log(f"  Added Shape Ft Req sheet ({len(shape_summary)} shape nest(s), "
        f"{sum(e['ft'] for e in shape_summary)} ft total).")


# The Working Forecast List's input-column layout ('911 Forecast' tab), so a
# reviewed award copy-pastes straight into the forecast. 'REM Used'/'REM Made'
# and 'Operations' are intentionally blank manual-input columns (the REM pair
# is the plate-work remnant info the planner re-enters per nest).
WFI_HEADERS = ["Order", "Source Material", "Pieces", "Nest Pkg Nbr", "Orders",
               "REM Used", "REM Made", "Total Ft Req", "Operations"]


def write_working_forecast_sheet(wb, forecast_summary, log):
    """One row per nest -- plates AND shapes -- in the Working Forecast's
    input-column layout (planning's ask, 2026-07-13). Source Material = the
    packet's page-1 'Source Code' (EB stock identifier; falls back to the
    batch list's Material codes); Pieces / Orders = the packet's page-1
    data-row counts (fall back to the nest's batch-list rows); Total Ft Req =
    the shape nest's Summary-of-Batches feet (blank on plate nests, with a
    grand total at the bottom). REM Used / REM Made / Operations stay blank
    for manual input before the paste."""
    ws = wb.create_sheet("Working Forecast Input")
    for c, name in enumerate(WFI_HEADERS, start=1):
        cell = ws.cell(1, c, name)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER
    ri = 1
    for entry in forecast_summary:
        ri += 1
        vals = [entry["order"], entry["source"], entry["pieces"],
                entry["nest"], entry["orders"], None, None,
                entry["ft"], None]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(ri, c, v)
            cell.border = _BORDER
    ft_col = WFI_HEADERS.index("Total Ft Req") + 1
    total_cell = ws.cell(ri + 1, ft_col - 1, "TOTAL")
    total_cell.font = Font(bold=True)
    total_cell.border = _BORDER
    val_cell = ws.cell(ri + 1, ft_col,
                       sum(e["ft"] for e in forecast_summary if e["ft"]))
    val_cell.font = Font(bold=True)
    val_cell.border = _BORDER
    ws.freeze_panes = "A2"
    _autosize(ws, WFI_HEADERS)
    log(f"  Added Working Forecast Input sheet ({len(forecast_summary)} "
        f"nest(s)).")


def write_data_table(ws, data_headers, data_rows):
    """Write one sheet's flat data table (header row 1, then data). Row highlights:
    red = plate with no DXF match (geometry missing -- the LI reference block is
    empty; the pcs/hr primary estimate still computes), orange = LI outlier
    (past LI_OUTLIER_SD), yellow = no estimate, light blue = multi-layered. No
    summary block -- pivots are layered on later."""
    for c, name in enumerate(data_headers, start=1):
        if name == "":
            continue   # spacer column: no header text or styling
        cell = ws.cell(1, c, name)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER

    # Hours columns get 3-decimal formatting: the primary Est Cut Hours + the
    # far-right LI reference copy.
    hour_idxs = {data_headers.index(c) + 1
                 for c in (EST_COL, "LI Est Cut Hours") if c in data_headers}
    for ri, row in enumerate(data_rows, start=2):
        # Pick one row highlight (most-important wins): no-DXF-match (red) >
        # LI outlier > no-estimate > multi-layered. Red leads because a plate
        # with no geometry (empty LI reference) is the one to chase.
        is_nodxf = "no dxf match" in str(row.get("Flags") or "").lower()
        is_outlier = "outlier" in str(row.get("Flags") or "").lower()
        missing = row.get(EST_COL) is None
        is_multi = row.get("Multi-Layered") == "Y"
        row_fill = (_NODXF_FILL if is_nodxf else _OUTLIER_FILL if is_outlier
                    else _MISSING_FILL if missing
                    else _MULTI_FILL if is_multi else None)
        for c, name in enumerate(data_headers, start=1):
            if name == "":
                continue   # spacer stays truly empty: no border, no highlight
            val = row.get(name)
            cell = ws.cell(ri, c, val)
            cell.border = _BORDER
            if row_fill is not None:
                cell.fill = row_fill
            if _is_date(val):
                cell.number_format = "MM/DD/YYYY"
            if c in hour_idxs and isinstance(val, (int, float)):
                cell.number_format = "0.000"
    ws.freeze_panes = "A2"
    _autosize(ws, data_headers)


def _nest_sums(data_rows):
    """Aggregate rows to nest -> summed Est Cut Hours. Returns
    (sums: dict[nest -> hours], grand_est, grand_qty)."""
    sums = {}
    grand_est = 0.0
    grand_qty = 0.0
    for row in data_rows:
        est = row.get(EST_COL) or 0.0
        qty = _to_float(row.get("PPN Quantity")) or 0.0
        nest = str(row.get("Nest Pkg Nbr") or "").strip()
        sums[nest] = sums.get(nest, 0.0) + est
        grand_est += est
        grand_qty += qty
    return sums, grand_est, grand_qty


# Excel COM enum constants we need under late binding (no makepy types).
_XL_DATABASE = 1        # PivotCache SourceType (xlDatabase)
_XL_ROW_FIELD = 1       # PivotField.Orientation (xlRowField)
_XL_SUM = -4157         # xlConsolidationFunction (xlSum)
_XL_CALC_MANUAL = -4135 # xlCalculationManual


def add_real_pivots(out_path, data_headers, sheets_meta, log):
    """Open the saved workbook in Excel via COM and drop a REAL PivotTable below
    the data table on each sheet: row field = Nest Pkg Nbr, data field =
    Sum of Est Cut Hours, with the grand-total row turned on. Returns True on
    success. Requires Excel + pywin32; on any failure returns False so the
    caller can fall back to a static summary.

    Runs on the plugin's worker thread, so COM MUST be initialised on this
    thread (CoInitialize) before any Dispatch call.
    """
    try:
        import pythoncom
        import win32com.client as win32
    except ImportError:
        log("  Excel COM (pywin32) unavailable; writing static summary instead.")
        return False

    if EST_COL not in data_headers:
        return False
    nest_header = next((h for h in data_headers if h.upper() == "NEST PKG NBR"),
                       "Nest Pkg Nbr")
    qty_header = next((h for h in data_headers if h.upper() == "PPN QUANTITY"), None)
    # Pivot source range stops at the blank spacer column: Excel refuses a
    # PivotCache whose source has a blank header, and every pivoted field
    # (Nest Pkg Nbr, PPN Quantity, Est Cut Hours) sits left of it anyway.
    n_src_cols = data_headers.index("") if "" in data_headers else len(data_headers)
    last_col = get_column_letter(n_src_cols)

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        # Dedicated, isolated instance (DispatchEx) -- kill the churn that makes
        # COM pivot work slow: no screen repaints, no events, no auto-recalc
        # after every edit. We Quit the instance afterwards, so no need to
        # restore these.
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        wb = excel.Workbooks.Open(str(out_path))
        # Application.Calculation can only be set once a workbook is open.
        # Setting it on a fresh, workbook-less instance raises "Unable to set
        # the Calculation property of the Application class" -- which aborted
        # the ENTIRE pivot pass, so every run silently fell back to the static
        # summary and no real PivotTable was ever created. Set it AFTER Open.
        excel.Calculation = _XL_CALC_MANUAL
        for sheet_name, n_rows in sheets_meta:
            if n_rows < 1:
                continue
            ws = wb.Worksheets(sheet_name)
            last_row = n_rows + 1  # +1 for the header row
            src = ws.Range(f"A1:{last_col}{last_row}")
            # Plain-English caveat right above the pivot stating what the
            # number is: throughput-derived, so it INCLUDES handling (v2.7.0
            # flipped this from the old torch-on-only linear-inch basis).
            ws.Cells(last_row + 2, 1).Value = (
                "Estimated from actual D911 throughput (pieces/hour by "
                "thickness, closed actuals Jan 2025 - Jul 2026); includes "
                "typical setup, positioning, and handling time.")
            dest = ws.Cells(last_row + 4, 1)  # blank + caveat rows under the data
            cache = wb.PivotCaches().Create(SourceType=_XL_DATABASE, SourceData=src)
            table_name = sheet_name.replace("-", "") + "Pivot"
            pt = cache.CreatePivotTable(TableDestination=dest, TableName=table_name)
            # Defer recalculation until every field is configured. With the
            # default ManualUpdate=False, Excel rebuilds the WHOLE pivot after
            # each field change. Batching them into one recalc is the fix.
            pt.ManualUpdate = True
            # ONE row per nest. Under it, three data columns so the number is
            # self-explanatory: Parts = the true piece count for the WHOLE nest
            # (all quantities summed), then the same cut-time total in BOTH
            # hours and minutes. (PPN Quantity is no longer a row field -- that
            # split each nest into confusing per-quantity slivers, e.g.
            # "5CDAZP / 8 / 0.013", which read as "8 parts = 0.013 hr".)
            pt.PivotFields(nest_header).Orientation = _XL_ROW_FIELD
            if qty_header:
                dparts = pt.AddDataField(pt.PivotFields(qty_header),
                                         "Parts", _XL_SUM)
                dparts.NumberFormat = "0"
            dhr = pt.AddDataField(pt.PivotFields(EST_COL),
                                  "Cut Time (hr)", _XL_SUM)
            dhr.NumberFormat = "0.000"
            # Minutes = hours x 60, as a calculated field so the minutes column
            # is always an EXACT conversion of the hours column (never drifts).
            try:
                pt.CalculatedFields().Add("CutMinutes", f"='{EST_COL}' *60")
                dmin = pt.AddDataField(pt.PivotFields("CutMinutes"),
                                       "Cut Time (min)", _XL_SUM)
                dmin.NumberFormat = "0.0"
            except Exception:
                pass  # minutes is a convenience column; hours is the source of truth
            # ColumnGrand = the bottom total row (total of all nests) -- what we
            # want. RowGrand would add a redundant rightmost total column.
            pt.ColumnGrand = True
            pt.RowGrand = False
            pt.ManualUpdate = False   # one recalc/re-layout, now that we're done
        wb.Save()
        wb.Close(SaveChanges=True)
        log("  Added real Excel PivotTables (per nest -> Parts + Cut Time "
            "in hr and min).")
        return True
    except Exception as e:
        log(f"  Excel pivot creation failed ({e}); writing static summary instead.")
        return False
    finally:
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def add_static_pivots(out_path, plate_rows, nonplate_rows, log):
    """Fallback when Excel COM is unavailable: reopen the workbook with openpyxl
    and append a static nest -> Sum of Est Cut Hours summary (+ grand total) to
    each sheet, mirroring what the real PivotTable would have shown."""
    # Resilient: the just-written workbook may already be open in Excel or
    # syncing as cloud-only (Hard Rule 13).
    wb = sdk.load_workbook_resilient(out_path, log=log)
    for sheet_name, rows in (("Plates", plate_rows), ("Non-Plates", nonplate_rows)):
        if sheet_name in wb.sheetnames:
            _write_static_summary(wb[sheet_name], rows)
    wb.save(out_path)
    log("  Wrote static nest summary (Excel PivotTable unavailable).")


def _write_static_summary(ws, data_rows):
    """Append a per-nest summary (+ Grand Total) below the data table, mirroring
    the real pivot: ONE line per nest with Parts (the true piece count for the
    whole nest), then the same cut-time total in BOTH hours and minutes."""
    agg, order = {}, []           # nest -> [parts, hours]
    for row in data_rows:
        est = row.get(EST_COL) or 0.0
        qty = _to_float(row.get("PPN Quantity")) or 0.0
        nest = str(row.get("Nest Pkg Nbr") or "").strip()
        if nest not in agg:
            agg[nest] = [0.0, 0.0]
            order.append(nest)
        agg[nest][0] += qty
        agg[nest][1] += est
    start = ws.max_row + 3
    # Plain-English caveat above the summary (matches the COM pivot's).
    ws.cell(start, 1, "Estimated from actual D911 throughput (pieces/hour by "
                      "thickness, closed actuals Jan 2025 - Jul 2026); includes "
                      "typical setup, positioning, and handling time.")
    start += 1
    for c, label in ((1, "Nest Pkg Nbr"), (2, "Parts"),
                     (3, "Cut Time (hr)"), (4, "Cut Time (min)")):
        cell = ws.cell(start, c, label)
        cell.fill = _PIVOT_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = _BORDER
    r = start + 1
    tot_parts = tot_hr = 0.0
    for nest in sorted(order):
        parts, hours = agg[nest]
        tot_parts += parts
        tot_hr += hours
        ws.cell(r, 1, nest)
        p = ws.cell(r, 2, round(parts)); p.number_format = "0"
        h = ws.cell(r, 3, round(hours, 3)); h.number_format = "0.000"
        m = ws.cell(r, 4, round(hours * 60, 1)); m.number_format = "0.0"
        r += 1
    ws.cell(r, 1, "Grand Total")
    gp = ws.cell(r, 2, round(tot_parts)); gp.number_format = "0"
    gh = ws.cell(r, 3, round(tot_hr, 3)); gh.number_format = "0.000"
    gm = ws.cell(r, 4, round(tot_hr * 60, 1)); gm.number_format = "0.0"
    for c in (1, 2, 3, 4):
        ws.cell(r, c).fill = _TOTAL_FILL
        ws.cell(r, c).font = Font(bold=True)
        ws.cell(r, c).border = _BORDER


def _autosize(ws, data_headers):
    widths = {}
    for c, name in enumerate(data_headers, start=1):
        widths[c] = 4 if name == "" else max(len(str(name)), 10)
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 400)):
        for cell in row:
            if cell.value is not None and cell.column <= len(data_headers):
                widths[cell.column] = max(widths.get(cell.column, 10),
                                          min(len(str(cell.value)) + 2, 45))
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ──────────────────────────────────────────────────────────────────────────
# Main entry point
# ──────────────────────────────────────────────────────────────────────────

def run(params: dict, progress_callback, cancel_event):
    log = params.get("log", print)
    settings = params.get("settings", {}) or {}

    if not PYMUPDF_AVAILABLE:
        # An install/build problem, not something the user can fix -> generic
        # error path (try again, then send a debug report).
        raise RuntimeError("PyMuPDF (fitz) is not available; cannot parse PDFs.")

    log("=" * 60)
    log(f"911 SSPO Award Review v{VERSION}")
    log(f"Cut-time basis: pieces/hour by thickness ({PIECES_PER_HOUR_SOURCE});")
    log("linear-inch times kept as far-right reference columns.")
    log("=" * 60)

    start_dir = ""
    try:
        roots = sdk.pilot_program_roots()
        if roots:
            start_dir = str(roots[0])
    except Exception:
        pass
    raw = sdk.request_directory(params, "Select the ROOT folder of order folders",
                                start_dir)
    if cancel_event.is_set():
        return
    if not raw:
        log("Folder selection cancelled - nothing was run.")
        cancel_event.set()  # user cancel: not a successful (ticket-earning) run
        return
    root = Path(raw.strip().strip('"'))
    if not root.exists() or not root.is_dir():
        raise sdk.UserFacingError(
            f"That folder doesn't exist: {root}",
            "Check the path and pick the ROOT folder of order folders, then run again.")
    log(f"ROOT: {root}")

    order_folders = discover_order_folders(root, log)
    if not order_folders:
        raise sdk.UserFacingError(
            "No order folders were found under that folder.",
            "Pick the folder whose subfolders are the orders (each holds a "
            "'CUI- TECH DATA READ ME' folder, a BATCH LIST workbook, or NEST "
            "PACKAGES), then run again.")
    log(f"Found {len(order_folders)} order folder(s): "
        f"{', '.join(o.name for o in order_folders)}")
    progress_callback(5)

    # Union of all batch-list headers (in first-seen column order) -> data cols.
    header_order = []           # batch-list headers, ordered
    header_seen = set()
    data_rows = []
    flags_summary = []
    shape_summary = []          # one entry per shape nest, for the Ft Req sheet
    forecast_summary = []       # one entry per nest, for Working Forecast Input

    for oi, order_folder in enumerate(order_folders):
        if cancel_event.is_set():
            log("Cancelled.")
            return
        order = order_folder.name
        log(f"\n--- Order {oi + 1}/{len(order_folders)}: {order} ---")

        cui = find_cui_folder(order_folder)
        pdf_folder = find_pdf_folder(cui) if cui else None
        batch_list = find_batch_list(cui, log=log) if cui else None

        if batch_list is None:
            log(f"  WARNING: No BATCH LIST workbook found for {order} -- skipping.")
            flags_summary.append(f"{order}: no batch list")
            continue
        log(f"  Batch list: {batch_list.name}")
        if pdf_folder:
            log(f"  PDF folder: {pdf_folder.name}")
        else:
            log(f"  WARNING: No PDF folder with packets found for {order}.")

        # Exact per-part cut length from the order's IGES DXFs, keyed by work
        # order (matches the batch list's 'Work Order' column). Absent for
        # structural-only orders -> those rows fall back to the band estimate.
        iges_folder = find_iges_folder(cui) if cui else None
        dxf_index, dxf_suspect, dxf_multi, dxf_empty = build_dxf_index(
            str(iges_folder) if iges_folder else None, log, cancel_event)
        if iges_folder:
            log(f"  IGES DXFs: {len(dxf_index)} work order(s) measured.")

        headers, rows = load_batch_rows(batch_list)
        if not rows:
            log(f"  WARNING: No data rows in {batch_list.name}.")
            continue
        for h in headers:
            if h.upper() not in header_seen:
                header_seen.add(h.upper())
                header_order.append(h)

        # Parse each unique nest's PDF once.
        nests_in_list = []
        seen_nests = []
        for row in rows:
            n = row.get("NEST PKG NBR")
            if n is not None and str(n).strip():
                ns = str(n).strip()
                if ns not in seen_nests:
                    seen_nests.append(ns)
        nest_pdf_data = {}
        for ns in seen_nests:
            sdk.raise_if_cancelled(cancel_event)
            pdf = _find_nest_pdf(pdf_folder, ns) if pdf_folder else None
            if pdf is None:
                nest_pdf_data[ns] = None
                log(f"  FLAG: nest {ns} in batch list has no packet PDF "
                    f"(thickness from Description).")
                flags_summary.append(f"{order}/{ns}: no PDF")
                continue
            try:
                nest_pdf_data[ns] = parse_nest_pdf(pdf)
            except Exception as e:
                nest_pdf_data[ns] = None
                log(f"  WARNING: could not parse {pdf.name}: {e}")

        # Flag PDFs present but not referenced by the batch list.
        if pdf_folder:
            for pdf in pdf_folder.glob("*.pdf"):
                if not any(ns.upper() in pdf.stem.upper() for ns in seen_nests):
                    log(f"  FLAG: packet {pdf.name} not referenced in batch list.")
                    flags_summary.append(f"{order}: orphan PDF {pdf.name}")

        # Shape (1D) nests: one Ft Req rollup entry per nest whose packet
        # carries a Summary of Batches table (receiving's cheat-sheet feed).
        for ns in seen_nests:
            pdfd = nest_pdf_data.get(ns)
            if not pdfd:
                continue
            if pdfd.get("summary_malformed"):
                log(f"  FLAG: nest {ns}: Summary of Batches table unreadable - "
                    "Shape Ft Req left blank.")
                flags_summary.append(f"{order}/{ns}: Summary of Batches unreadable")
                continue
            lengths = pdfd.get("batch_lengths")
            if not lengths:
                continue
            desc = next((row.get("DESCRIPTION") for row in rows
                         if str(row.get("NEST PKG NBR") or "").strip() == ns
                         and row.get("DESCRIPTION")), None)
            shape_summary.append({
                "order": order, "nest": ns,
                "mat_type": pdfd.get("mat_type"),
                "mat_size": pdfd.get("mat_size"),
                "desc": desc, "lengths": lengths,
                "ft": shape_feet_required(lengths),
            })

        # Working Forecast Input rollup: one entry per nest (plates AND
        # shapes) mirroring the forecast's input columns so a reviewed award
        # copy-pastes straight into the Working Forecast List. Packet page-1
        # values lead; the nest's batch-list rows are the fallback when the
        # packet is missing/unparsed.
        for ns in seen_nests:
            pdfd = nest_pdf_data.get(ns)
            nest_rows = [row for row in rows
                         if str(row.get("NEST PKG NBR") or "").strip() == ns]
            src = (pdfd or {}).get("source_code")
            if not src:
                # Batch-list 'Material' carries the same EB stock code.
                mats = []
                for row in nest_rows:
                    mv = str(row.get("MATERIAL") or "").strip()
                    if mv and mv not in mats:
                        mats.append(mv)
                src = ", ".join(mats) or None
            pieces = (pdfd or {}).get("pieces")
            if pieces is None:
                ppn_sum = sum(_to_float(row.get("PPN QUANTITY")) or 0.0
                              for row in nest_rows)
                pieces = int(round(ppn_sum)) if ppn_sum else None
            n_orders = (pdfd or {}).get("orders")
            if n_orders is None:
                n_orders = len(nest_rows) or None
            if isinstance(src, str) and src.isdigit():
                src = int(src)  # bare numeric codes are numbers in the forecast
            _wf_lengths = (pdfd or {}).get("batch_lengths")
            ft = (shape_feet_required(_wf_lengths)
                  if _wf_lengths and not (pdfd or {}).get("summary_malformed")
                  else None)
            forecast_summary.append({
                "order": order, "nest": ns, "source": src,
                "pieces": pieces, "orders": n_orders, "ft": ft,
            })

        # Build a data row per batch-list row.
        for row in rows:
            nest = str(row.get("NEST PKG NBR") or "").strip()
            pdfd = nest_pdf_data.get(nest)
            ppn = _to_float(row.get("PPN QUANTITY")) or 0.0
            scope = row.get("SCOPE OF WORK") or row.get("SCOPE OF WORK ")
            thickness = (pdfd or {}).get("thickness") if pdfd else None
            thk_src = "PDF"
            if thickness is None:
                thickness = thickness_from_description(row.get("DESCRIPTION"))
                thk_src = "Description" if thickness is not None else "MISSING"

            # Primary cut time (v2.7.0): actuals-derived pieces/hour by
            # thickness band for every non-shape row; structural shapes keep
            # the legacy band. The DXF linear-inch time is still computed as
            # the far-right reference block when the work order resolves a DXF.
            wo = str(row.get("WORK ORDER") or "").strip().upper()
            li_pc = dxf_index.get(wo)
            bevel = _is_bevel(scope)
            is_shape = _desc_is_shape_form(row.get("DESCRIPTION"))
            est = None if is_shape else throughput_estimate(thickness, ppn)
            est_li = plate_cut_estimate(li_pc, thickness, ppn)

            out_row = {}
            # Copy every batch-list field by header.
            for h in headers:
                out_row[h] = row.get(h.upper())
            out_row["Order"] = order
            out_row["Division"] = ""      # blank placeholder (planner fills in)
            # Location / Process: from batch list if present, else PDF-derived.
            out_row["Location"] = row.get("LOCATION")
            proc = row.get("PROCESS")
            if not proc and pdfd:
                proc = pdfd.get("process")
            out_row["Process"] = proc
            out_row["Thickness (in)"] = thickness
            # Remnant Length/Width: the nest sheet's page-1 dimensions, filled
            # ONLY when this row used a remnant (batch-list 'Rem Used' populated).
            rem_used = row.get("REM USED")
            has_rem = rem_used is not None and str(rem_used).strip() != ""
            out_row["Remnant Length"] = ((pdfd or {}).get("sheet_length")
                                         if (has_rem and pdfd) else None)
            out_row["Remnant Width"] = ((pdfd or {}).get("sheet_width")
                                        if (has_rem and pdfd) else None)
            out_row["Multiplier"] = ""    # blank placeholder (planner fills in)
            # Shape Ft Req: per-NEST stock feet from the packet's Summary of
            # Batches table (shape/1D nests only) -- repeated on the nest's rows.
            _sb_lengths = (pdfd or {}).get("batch_lengths")
            out_row["Shape Ft Req"] = (shape_feet_required(_sb_lengths)
                                       if _sb_lengths else None)
            # Preserved (non-column) field for the Analysis sheet: did this part's
            # DXF carry a present-but-zero-length layer? True/False for DXF-resolved
            # plate rows, None when there's no DXF match.
            out_row["_empty_layers"] = dxf_empty.get(wo) if wo in dxf_index else None
            out_row["Stock L"] = (pdfd or {}).get("stock_l") if pdfd else None
            out_row["Stock W"] = (pdfd or {}).get("stock_w") if pdfd else None
            out_row["Plate Weight"] = (pdfd or {}).get("plate_weight") if pdfd else None
            out_row["Mil Spec"] = (pdfd or {}).get("mil_spec") if pdfd else None
            out_row["MT Material"] = (pdfd or {}).get("mt_material") if pdfd else None
            # Per-NEST machine/fuel from the packet's page-1 header table,
            # repeated on the nest's rows (like Thickness).
            out_row["EB Machine"] = pdfd.get("machine") if pdfd else None
            out_row["EB Fuel"] = pdfd.get("fuel") if pdfd else None

            flags = []
            # LI Dev (SD) is filled in after all rows are built (needs the mean/std
            # across the whole run); leave a placeholder here.
            out_row["LI Dev (SD)"] = None
            # Far-right LI reference block (the retired primary estimate) --
            # filled whenever the row's work order resolved a DXF. Progressive
            # rounding as below so LI Est Min/PC x Qty reproduces LI Est Min/WO.
            if est_li is not None:
                li_disp = round(est_li["linear_in_pc"], 2)
                li_mpc_disp = round(est_li["min_pc"], 2)
                li_mwo_disp = round(li_mpc_disp * ppn, 2)
                out_row["Linear In/PC"] = li_disp
                out_row["Total Linear In"] = round(li_disp * ppn, 2)
                out_row["Multi-Layered"] = "Y" if dxf_multi.get(wo) else "N"
                out_row["Feed Rate (IPM)"] = est_li["feed_ipm"]
                out_row["LI Est Min/PC"] = li_mpc_disp
                out_row["LI Est Min/WO"] = li_mwo_disp
                out_row["LI Est Cut Hours"] = round(li_mwo_disp / 60.0, 4)
                if est_li["feed_note"]:
                    flags.append(est_li["feed_note"])
                if wo in dxf_suspect:
                    flags.append(dxf_suspect[wo])
            else:
                out_row["Linear In/PC"] = None
                out_row["Total Linear In"] = None
                out_row["Multi-Layered"] = ""     # N/A without a DXF
                out_row["Feed Rate (IPM)"] = None
                out_row["LI Est Min/PC"] = None
                out_row["LI Est Min/WO"] = None
                out_row["LI Est Cut Hours"] = None
                # A plate SHOULD have an IGES DXF; flag the miss (the primary
                # estimate no longer needs it, but the LI reference is lost).
                if li_pc is None and "PLATE" in str(row.get("DESCRIPTION") or "").upper():
                    flags.append("no DXF match - no LI reference")

            if est is not None:
                # Primary estimate: pieces/hour band. Progressive rounding:
                # each cell is derived from the previous ROUNDED cell so the
                # printed Equation reproduces every column exactly
                # (Est Min/PC x Qty == Est Min/WO, etc.).
                mpc_disp = round(est["min_pc"], 2)
                mwo_disp = round(mpc_disp * ppn, 2)
                hours_disp = round(mwo_disp / 60.0, 4)
                out_row["Pcs/Hr"] = est["pcs_hr"]
                out_row["Est Min/PC"] = mpc_disp
                out_row["Est Min/WO"] = mwo_disp
                out_row["Est Cut Hours"] = hours_disp
                out_row["Equation"] = (
                    f"Est Min/PC = 60 / Pcs-Hr = "
                    f"60 / {_eqnum(est['pcs_hr'])} = {_eqnum(mpc_disp)} min   |   "
                    f"Est Min/WO = Est Min/PC x Qty = "
                    f"{_eqnum(mpc_disp)} x {_eqnum(ppn)} = {_eqnum(mwo_disp)} min   |   "
                    f"Est Cut Hours = Est Min/WO / 60 = "
                    f"{_eqnum(mwo_disp)} / 60 = {hours_disp:.3f} hr"
                )
                if est["pph_note"]:
                    flags.append(est["pph_note"])
            else:
                # Structural shape (legacy band estimate), or thickness unknown.
                est_hr, factor, _bevel = row_estimate_hours(thickness, ppn, scope)
                out_row["Pcs/Hr"] = None
                if est_hr is not None and factor is not None:
                    band = band_minutes(thickness)
                    factor_disp = round(factor, 2)
                    mwo_disp = round(factor_disp * ppn, 2)
                    hours_disp = round(mwo_disp / 60.0, 4)
                    out_row["Est Min/PC"] = factor_disp
                    out_row["Est Min/WO"] = mwo_disp
                    out_row["Est Cut Hours"] = hours_disp
                    out_row["Equation"] = (
                        f"Band fallback (structural shape). "
                        f"Est Min/PC = Thickness x band-factor = "
                        f"{_eqnum(thickness)} x {band} = {_eqnum(factor_disp)} min   |   "
                        f"Est Min/WO = Est Min/PC x Qty = "
                        f"{_eqnum(factor_disp)} x {_eqnum(ppn)} = {_eqnum(mwo_disp)} min   |   "
                        f"Est Cut Hours = Est Min/WO / 60 = "
                        f"{_eqnum(mwo_disp)} / 60 = {hours_disp:.3f} hr"
                    )
                else:
                    out_row["Est Min/PC"] = None
                    out_row["Est Min/WO"] = None
                    out_row["Est Cut Hours"] = None
                    out_row["Equation"] = "No estimate (thickness unknown)."

            if thk_src == "MISSING":
                flags.append("no thickness")
            elif thk_src == "Description":
                flags.append("thk from desc")
            if bevel:
                flags.append("bevel scope (no time added)")
            if pdfd and pdfd.get("summary_malformed"):
                flags.append("Summary of Batches unreadable - no Shape Ft Req")
            out_row["Flags"] = "; ".join(flags)
            # Make uppercase-keyed lookups work for pivot/derived fields.
            out_row["Nest Pkg Nbr"] = nest
            out_row["PPN Quantity"] = row.get("PPN QUANTITY")
            out_row["Description"] = row.get("DESCRIPTION")
            data_rows.append(out_row)

        progress_callback(5 + int(85 * (oi + 1) / len(order_folders)))

    if not data_rows:
        log("\nERROR: No rows produced; nothing to write.")
        return

    # Column order: 'Order' first, then every batch-list column verbatim in
    # native order (with the batch list's 'Material' shown as 'Source
    # Material'), then our generated columns. Nothing from the batch list is
    # dropped.
    batch_headers = [BATCH_HEADER_RENAME.get(h, h) for h in header_order]
    seen = {h.upper() for h in batch_headers} | {"ORDER"}
    data_headers = (["Order"] + batch_headers
                    + [c for c in GENERATED_COLS if c.upper() not in seen])

    # Normalize each row's keys to the exact header strings used in output.
    norm_rows = []
    for row in data_rows:
        nr = {}
        low = {k.upper(): v for k, v in row.items()}
        for h in data_headers:
            nr[h] = low.get(HEADER_SOURCE_KEY.get(h, h.upper()))
        # Preserve the canonical pivot keys.
        nr["Nest Pkg Nbr"] = row.get("Nest Pkg Nbr") or low.get("NEST PKG NBR")
        nr["PPN Quantity"] = row.get("PPN Quantity") if row.get("PPN Quantity") is not None else low.get("PPN QUANTITY")
        nr["Description"] = row.get("Description") if row.get("Description") is not None else low.get("DESCRIPTION")
        nr["Est Cut Hours"] = low.get("EST CUT HOURS")
        nr["_empty_layers"] = row.get("_empty_layers")   # Analysis sheet only
        norm_rows.append(nr)

    # Split into plate vs non-plate (shapes/bars/tubes) — each gets its own
    # identically-structured sheet. A row is a plate if its Description says
    # PLATE or if we managed to compute an estimate (thickness was found).
    plate_rows = [r for r in norm_rows if _is_plate_row(r)]
    nonplate_rows = [r for r in norm_rows if not _is_plate_row(r)]
    log(f"\nBuilding workbook: {len(plate_rows)} plate rows, "
        f"{len(nonplate_rows)} non-plate rows.")

    # Linear-inch deviation: how many standard deviations each plate part's
    # Linear In/PC sits from the mean of all measured plate parts. Parts beyond
    # LI_OUTLIER_SD are flagged + highlighted so an inflated/odd part stands out.
    li_vals = [r.get("Linear In/PC") for r in plate_rows
               if isinstance(r.get("Linear In/PC"), (int, float))]
    if len(li_vals) >= 2:
        li_mean = statistics.mean(li_vals)
        li_sd = statistics.pstdev(li_vals)
        n_outliers = 0
        for r in plate_rows:
            li = r.get("Linear In/PC")
            if isinstance(li, (int, float)) and li_sd > 0:
                z = (li - li_mean) / li_sd
                r["LI Dev (SD)"] = round(z, 2)
                if abs(z) >= LI_OUTLIER_SD:
                    n_outliers += 1
                    note = f"LI outlier {z:+.1f}SD"
                    r["Flags"] = f"{r['Flags']}; {note}" if r.get("Flags") else note
        log(f"  Linear-inch mean {li_mean:.0f} in, SD {li_sd:.0f} in; "
            f"{n_outliers} part(s) beyond {LI_OUTLIER_SD:g} SD flagged.")

    # Output path.
    out_dir = settings.get("output_dir", "").strip()
    out_base = Path(out_dir) if out_dir else root
    stamp = datetime.datetime.now().strftime("%Y-%m-%d")
    out_path = out_base / f"911 SSPO AWARD REVIEW - {root.name} - {stamp}.xlsx"
    try:
        sheets_meta = write_workbook(out_path, data_headers, plate_rows,
                                     nonplate_rows, shape_summary,
                                     forecast_summary, log)
    except PermissionError:
        log(f"  Output file is open/locked: {out_path.name}. Close it and re-run.")
        return

    # Layer on the real Excel PivotTables; fall back to a static summary if
    # Excel COM isn't available or errors.
    if not add_real_pivots(out_path, data_headers, sheets_meta, log):
        try:
            add_static_pivots(out_path, plate_rows, nonplate_rows, log)
        except PermissionError:
            log(f"  Output file is open/locked: {out_path.name}. Close it and re-run.")
            return

    _, grand_est, grand_qty = _nest_sums(plate_rows)
    progress_callback(100)
    log("\n" + "=" * 60)
    log(f"DONE. Wrote: {out_path}")
    log(f"  Plate rows: {len(plate_rows)}  |  Non-plate rows: {len(nonplate_rows)}")
    log(f"  Plate cut time (throughput basis, incl. handling): {grand_est:.2f} hr "
        f"({grand_est * 60:.0f} min) across {int(grand_qty)} pieces")
    if shape_summary:
        log(f"  Shape stock required: {sum(e['ft'] for e in shape_summary)} ft "
            f"across {len(shape_summary)} shape nest(s) (Shape Ft Req sheet)")
    if forecast_summary:
        log(f"  Working Forecast Input: {len(forecast_summary)} nest(s) ready "
            f"to paste into the forecast")
    if flags_summary:
        log(f"  Flags ({len(flags_summary)}):")
        for f in flags_summary[:25]:
            log(f"    - {f}")
        if len(flags_summary) > 25:
            log(f"    ... and {len(flags_summary) - 25} more")
    log("=" * 60)


if __name__ == "__main__":
    import threading
    run({"log": print}, lambda p: print(f"[{p}%]"), threading.Event())
