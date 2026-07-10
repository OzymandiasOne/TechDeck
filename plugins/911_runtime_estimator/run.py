"""
911 Runtime Estimator Plugin for TechDeck
=========================================

Given a ROOT directory containing one subfolder per order (e.g. F124, P003,
S029, V092 ...), this plugin:

  1. Walks each order folder, locating the nest work-package PDFs and the
     order's BATCH LIST workbook (folder names are resolved DEFENSIVELY -- the
     PDFs may live in 'NEST PACKAGES' or 'WPDD SKETCHES' under a
     'CUI- TECH DATA READ ME' folder, and layouts vary between orders).
  2. Parses each nest packet PDF (PyMuPDF) for thickness, pieces, material,
     stock size, mil spec, and plate weight -- cross-checking fields across the
     page-1 data row, the SUMMARY page, the part sketch, and the MOVE TICKET,
     exactly like the 911 Setup plugin.
  3. Computes a per-row plate cutting-time estimate from EXACT linear inches of
     cut. Each batch-list row's 'Work Order' is matched to its DXF in the order's
     'IGES FILES' folder; the cut length is measured from the DXF vector
     geometry and divided by a thickness-driven feed rate (IPM) -- see
     CALCULATION below. Rows with no DXF (structural shapes, or a missing file)
     fall back to the legacy thickness-band estimate so hours aren't lost.
  4. Writes ONE consolidated workbook with two sheets (Plates / Non-Plates).
     Each is a flat data table that reproduces EVERY batch-list column in its
     native order, then our generated columns. 'Material' on the table is the
     MOVE TICKET designation; the batch list's own 'Material' (an EB stock code)
     is shown as 'Source Material'.
  5. Drops a REAL, refreshable Excel PivotTable below the data on each sheet --
     ONE line per nest: Parts (true piece count for the whole nest) + the same
     Cut Time total in both hours and minutes, grand total at the bottom -- by
     driving Excel via COM (pywin32). A plain-English "machine cut time only"
     caveat sits above it. If Excel/COM is unavailable or errors, it falls back
     to a static per-nest summary of the same shape so a run never hard-crashes.

Runs off an AWARD PACKAGE (the SPARS-received folder of order folders, before
batching) so work-scope review / division + machine assignment can happen up
front; the plate cut times are the exact linear-inch estimates. Replaced the
standalone 911_linear_inch_cuttime tool (removed) whose engine is vendored here.

CALCULATION (per plate batch-list row; linear inches drive the estimate):
    linear_in_pc = exact DXF cut length for the row's Work Order (inches)
    feed (IPM)   = lookup_feed(thickness), rounding thickness UP to the next chart row
    Est Min/PC   = linear_in_pc / feed
    Est Min/WO   = Est Min/PC * (row PPN Quantity)
    Est Cut Hours = Est Min/WO / 60
An 'Equation' column (right before Est Cut Hours) prints this chain with the
row's own numbers plugged in so the arithmetic is auditable cell-by-cell.
Bevel adds NO time (the legacy +180 min/pc surcharge was removed per planning
direction); bevel scope is still detected and flagged for review. Non-plate /
DXF-less rows use the legacy band as a fallback:
    band(t): t<0.5->6 ; <1.0->9 ; <2.0->12 ; >=2.0->18 (min/pc); est = t*band*qty.

Prompts for the ROOT via the native folder dialog only (no pre-configured settings
required). Optional 'output_dir' setting picks where the workbook is saved;
blank saves it inside ROOT.
"""

from __future__ import annotations

import os
import re
import glob
import math
import statistics
import datetime
from pathlib import Path
from typing import Optional

# SDK bootstrap (works in-process and for standalone CLI testing).
try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════
# Linear-inch cut-time engine (vendored from the removed 911_linear_inch_cuttime)
# ──────────────────────────────────────────────────────────────────────────
# Plate cut time = EXACT DXF cut length / feed rate (IPM). Cut length per part
# comes from the order's IGES DXF vector geometry; the feed rate is looked up by
# thickness in one combined chart, rounding UP to the next listed thickness so
# oddball thicknesses stay covered. TODO: centralize this parser into plugin_sdk
# (customer_dxf_quoting + this module carry copies).
# ══════════════════════════════════════════════════════════════════════════

# Work order token in DXF filenames (PART.REV.WORKORDER) + the batch list. Two
# SigmaNest formats: current "X4549775"/"BK514922" (1-2 letters + 5-8 digits) and
# legacy "3X24-398"; same form 911_setup's summary parser uses.
WO_RE = re.compile(r"(?:[A-Z]{1,2}\d{5,8}|\d[A-Z]\d+-\d+)", re.I)
# Layers that are reference-only and never count as cut (matches customer_dxf_quoting).
# Only these NAMED reference layers are dropped; ALL numeric geometry layers count.
# Confirmed by shop validation in SigmaNest (5 award parts, 2026-07-09): the cut is
# spread across every numeric layer (1/2 carry perimeter+holes inconsistently, 4/5
# carry other cut features) - so cut length = the sum of ALL non-reference layers.
# An earlier "layer 1 only" rule was WRONG (it rested on one Kinetic profile-only
# sheet that happened to list just a part's outer perimeter).
EXCLUDE_LAYERS = {"BOUNDING_BOX", "IGNORE"}
# Entity types we do NOT measure; their presence means the length may be under-counted.
UNSUPPORTED_FLAGGED = {"SPLINE", "ELLIPSE"}

FEED_RATES_SOURCE = "Plasma & Oxy Feed Rates (single combined chart, rev 2026-07-09)"
FEED_RATES = [  # (thickness in, speed IPM) - ascending
    (0.125, 180), (0.188, 140), (0.25, 100), (0.313, 80), (0.375, 100),
    (0.438, 90), (0.5, 75), (0.625, 55), (0.75, 80), (0.875, 70),
    (1.0, 50), (1.125, 45), (1.25, 40), (1.375, 35), (1.5, 30),
    (1.625, 45), (1.75, 40), (2.0, 30), (2.25, 25), (2.5, 20),
    (2.75, 9.25), (3.0, 9.25), (3.25, 9.0), (3.5, 8.5), (3.75, 8.5),
    (4.0, 8.5), (4.25, 8.0), (4.5, 8.0), (5.0, 7.25), (5.25, 6.5),
    (5.5, 6.0), (5.75, 6.0), (6.0, 6.0), (6.25, 6.0), (6.5, 6.0),
    (6.75, 6.0), (7.0, 6.0),
]


def lookup_feed(thickness):
    """(ipm, note) for a thickness, rounding UP to the next listed thickness (thicker
    = slower = more time = conservative). Clamps below/above the chart (flagged via
    note). None only when thickness is None."""
    if thickness is None:
        return None
    for t, ipm in FEED_RATES:
        if thickness <= t + 1e-9:
            note = None if abs(t - thickness) < 1e-6 else f"rounded up to {t}\" row"
            return ipm, note
    t, ipm = FEED_RATES[-1]
    return ipm, f"above chart max {t}\" - used {t}\" row"


def _decode_dxf(path):
    data = Path(path).read_bytes()
    if data.startswith(b"AutoCAD Binary DXF"):
        raise ValueError("Binary DXF is not supported - re-export as ASCII DXF.")
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not decode file - is this a DXF?")


def _read_pairs(path):
    lines = _decode_dxf(path).splitlines()
    pairs = []
    it = iter(lines)
    for code_line in it:
        try:
            value = next(it)
        except StopIteration:
            break
        try:
            code = int(code_line.strip())
        except ValueError:
            raise ValueError(f"Malformed DXF (expected numeric group code, got {code_line.strip()!r}).")
        pairs.append((code, value.strip()))
    if not pairs:
        raise ValueError("File is empty.")
    return pairs


def _bulge_length(p1, p2, bulge):
    chord = math.hypot(p2[0] - p1[0], p2[1] - p1[1])
    if abs(bulge) < 1e-12 or chord < 1e-12:
        return chord
    theta = 4.0 * math.atan(bulge)
    radius = chord * (1.0 + bulge * bulge) / (4.0 * abs(bulge))
    return radius * abs(theta)


def _collect_until_next_entity(pairs, j):
    groups = []
    n = len(pairs)
    while j < n and pairs[j][0] != 0:
        groups.append(pairs[j])
        j += 1
    return groups, j


def _build_line(groups):
    g = {}
    for code, val in groups:
        if code in (8, 10, 20, 11, 21):
            g[code] = val
    try:
        p1 = (float(g[10]), float(g[20]))
        p2 = (float(g[11]), float(g[21]))
    except KeyError:
        return None
    return {"layer": g.get(8, "0"), "points": [p1, p2], "length": math.dist(p1, p2)}


def _build_arc(groups, full_circle=False):
    g = {}
    for code, val in groups:
        if code in (8, 10, 20, 40, 50, 51):
            g[code] = val
    try:
        cx, cy, r = float(g[10]), float(g[20]), float(g[40])
    except KeyError:
        return None
    if full_circle:
        a0, sweep = 0.0, math.tau
    else:
        a0 = math.radians(float(g.get(50, 0.0)))
        a1 = math.radians(float(g.get(51, 360.0)))
        sweep = (a1 - a0) % math.tau or math.tau
    n = max(8, int(abs(sweep) / math.tau * 64) + 1)
    pts = [(cx + r * math.cos(a0 + sweep * t / n), cy + r * math.sin(a0 + sweep * t / n))
           for t in range(n + 1)]
    return {"layer": g.get(8, "0"), "points": pts, "length": r * sweep}


def _polyline_entity(layer, verts, closed):
    verts = [v for v in verts if v[1] is not None]
    if len(verts) < 2:
        return None
    total = 0.0
    seg_pairs = list(zip(verts, verts[1:]))
    if closed:
        seg_pairs.append((verts[-1], verts[0]))
    for v1, v2 in seg_pairs:
        total += _bulge_length((v1[0], v1[1]), (v2[0], v2[1]), v1[2])
    points = [(v[0], v[1]) for v in verts]
    return {"layer": layer, "points": points, "length": total}


def _build_lwpolyline(groups):
    layer, flags = "0", 0
    verts = []  # [x, y, bulge]
    for code, val in groups:
        if code == 8:
            layer = val
        elif code == 70:
            flags = int(val)
        elif code == 10:
            verts.append([float(val), None, 0.0])
        elif code == 20 and verts:
            verts[-1][1] = float(val)
        elif code == 42 and verts:
            verts[-1][2] = float(val)
    return _polyline_entity(layer, verts, closed=bool(flags & 1))


def _build_polyline(pairs, j):
    groups, j = _collect_until_next_entity(pairs, j)
    layer, flags = "0", 0
    for code, val in groups:
        if code == 8:
            layer = val
        elif code == 70:
            flags = int(val)
    verts = []
    n = len(pairs)
    while j < n:
        code, val = pairs[j]
        if code == 0 and val == "VERTEX":
            vgroups, j = _collect_until_next_entity(pairs, j + 1)
            v = [None, None, 0.0]
            vflags = 0
            for c, vv in vgroups:
                if c == 10:
                    v[0] = float(vv)
                elif c == 20:
                    v[1] = float(vv)
                elif c == 42:
                    v[2] = float(vv)
                elif c == 70:
                    vflags = int(vv)
            if v[0] is not None and v[1] is not None and not (vflags & 16):
                verts.append(v)
        elif code == 0 and val == "SEQEND":
            _, j = _collect_until_next_entity(pairs, j + 1)
            break
        else:
            j += 1
    return _polyline_entity(layer, verts, closed=bool(flags & 1)), j


def parse_dxf(path):
    """Return (entities, skipped, insunits). entity = {layer, points, length}."""
    pairs = _read_pairs(path)
    n = len(pairs)
    entities, skipped = [], {}
    insunits = None
    section = None
    i = 0
    while i < n:
        code, val = pairs[i]
        if code == 0 and val == "SECTION" and i + 1 < n and pairs[i + 1][0] == 2:
            section = pairs[i + 1][1]
            i += 2
            continue
        if code == 0 and val == "ENDSEC":
            section = None
            i += 1
            continue
        if section == "HEADER" and code == 9 and val == "$INSUNITS":
            if i + 1 < n and pairs[i + 1][0] == 70:
                insunits = int(pairs[i + 1][1])
            i += 1
            continue
        if section == "ENTITIES" and code == 0:
            if val == "LINE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_line(groups)
            elif val == "ARC":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_arc(groups)
            elif val == "CIRCLE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_arc(groups, full_circle=True)
            elif val == "LWPOLYLINE":
                groups, i = _collect_until_next_entity(pairs, i + 1)
                ent = _build_lwpolyline(groups)
            elif val == "POLYLINE":
                ent, i = _build_polyline(pairs, i + 1)
            else:
                skipped[val] = skipped.get(val, 0) + 1
                i += 1
                continue
            if ent is not None:
                entities.append(ent)
            continue
        i += 1
    return entities, skipped, insunits


def dxf_metrics(path):
    """Cut length (in) + anomaly info for one part DXF. Sums the geometry on ALL
    non-reference layers (the cut is spread across every numeric layer). Layers
    that hold no geometry (0 in) are disregarded - they don't count toward the
    layer total nor toward the multi-layer determination (`n_layers`)."""
    entities, skipped, insunits = parse_dxf(path)
    scale, unit_note = 1.0, None
    if insunits == 4:            # millimetres
        scale, unit_note = 1.0 / 25.4, "units=mm (converted to in)"
    elif insunits not in (1, 0, None):
        unit_note = f"unexpected $INSUNITS={insunits}"
    per_layer = {}
    for e in entities:
        ly = e["layer"].strip()
        if ly.upper() in EXCLUDE_LAYERS:
            continue
        per_layer[ly] = per_layer.get(ly, 0.0) + e["length"]
    # Disregard empty (0 in) layers so they can't be counted or skew anything.
    cut_layers = {ly: ln for ly, ln in per_layer.items() if ln > 1e-9}
    total = sum(cut_layers.values())
    return {
        "linear_inches": total * scale,
        "skipped": skipped,
        "layers": set(cut_layers),
        "n_layers": len(cut_layers),
        "unit_note": unit_note,
        "n_entities": len(entities),
    }


def _dxf_work_order(filename):
    """The work order token embedded in a DXF filename (PART.REV.WORKORDER): the
    last WO-shaped '.'-separated component, so the part number's leading digits
    aren't mistaken for it. None if none present."""
    stem = os.path.splitext(os.path.basename(filename))[0].upper()
    for comp in reversed(re.split(r"[.\s_]+", stem)):
        if WO_RE.fullmatch(comp):
            return comp
    return None


def build_dxf_index(iges_folder, log):
    """Map WORK ORDER (upper) -> linear inches per piece from an order's IGES DXFs.
    Also returns `suspect` (work orders whose geometry may be under-counted:
    unmeasured SPLINE/ELLIPSE or non-inch units) and `multilayer` (work order ->
    True when the part's cut spans more than one non-empty layer)."""
    index, suspect, multilayer = {}, {}, {}
    if not iges_folder:
        return index, suspect, multilayer
    for dxf in glob.glob(os.path.join(iges_folder, "**", "*.dxf"), recursive=True):
        wo = _dxf_work_order(dxf)
        if not wo:
            continue
        try:
            sdk.ensure_local(dxf, log=log)
            m = dxf_metrics(dxf)
        except Exception as exc:
            log(f"  FLAG: could not read DXF {os.path.basename(dxf)}: {exc}")
            continue
        index[wo] = m["linear_inches"]
        multilayer[wo] = m["n_layers"] > 1
        notes = []
        bad = [k for k in m["skipped"] if k in UNSUPPORTED_FLAGGED]
        if bad:
            notes.append("unmeasured " + "/".join(sorted(bad)))
        if m["unit_note"] and "converted" not in m["unit_note"]:
            notes.append(m["unit_note"])
        if notes:
            suspect[wo] = "; ".join(notes)
    return index, suspect, multilayer


VERSION = "2.2.0"

# We reproduce EVERY batch-list column verbatim, in its native order, then
# append our own generated columns after them. The batch list's own 'Material'
# column is an EB stock code; we surface it as 'Source Material' to keep it
# distinct from the part-material designation we pull off the MOVE TICKET pages.
BATCH_HEADER_RENAME = {"Material": "Source Material"}
# 'Order' (which order folder a row came from) leads the table; our other
# generated columns follow the batch-list block (in this order). 'Material' is
# the MOVE TICKET designation (sits right after MIL SPEC on the packet).
GENERATED_COLS = ["Division", "Process", "Thickness (in)", "Stock L", "Stock W",
                  "Mil Spec", "Material", "Linear In/PC", "Total Linear In",
                  "Multi-Layered", "LI Dev (SD)",
                  "Feed Rate (IPM)", "Est Min/PC", "Est Min/WO", "Equation",
                  "Est Cut Hours", "Flags"]
# 'Equation' spells out the FULL calculation for the row with this row's own
# numbers plugged in (Est Min/PC -> Est Min/WO -> Est Cut Hours), so the math is
# auditable cell-by-cell -- it sits right before 'Est Cut Hours' (the result).
# To make that audit hold, the Est columns are rounded PROGRESSIVELY (each derived
# from the previous rounded cell) so e.g. Est Min/PC x Qty reproduces Est Min/WO
# exactly rather than diverging by an independent-rounding cent.
# 'Division' is an intentionally blank placeholder the planner fills in by hand.
# The linear-inch columns + the two Est Min breakouts are populated for plate rows
# that resolve a DXF; non-plate rows leave them blank. 'Multi-Layered' = Y/N (does
# the part's cut span >1 non-empty layer; multi-layer rows are highlighted).
# 'LI Dev (SD)' = how many standard deviations the part's Linear In/PC sits from
# the mean of all plate parts in the run; parts past LI_OUTLIER_SD are flagged +
# highlighted so an inflated/odd part is easy to catch.
LI_OUTLIER_SD = 3.0
# For each OUTPUT header, the UPPERCASE row key its value is read from. Most map
# to their own upper-cased name; these two are indirections (the display name
# differs from the key the value is stored under).
HEADER_SOURCE_KEY = {
    "Source Material": "MATERIAL",      # batch list 'Material' (EB stock code)
    "Material": "MT MATERIAL",          # MOVE TICKET material designation (new)
}

# The column the pivot sums (our formula-driven estimate, in hours).
EST_COL = "Est Cut Hours"


# ──────────────────────────────────────────────────────────────────────────
# Calculation
# ──────────────────────────────────────────────────────────────────────────

def _eqnum(x):
    """Format a number for the human-readable Equation cell: no trailing zeros,
    integers shown without a decimal point ('4.0' -> '4', '0.750' -> '0.75') so
    the printed arithmetic matches what Excel shows in the neighbouring cells."""
    if x is None:
        return "?"
    try:
        xf = float(x)
    except (TypeError, ValueError):
        return str(x)
    if xf == int(xf):
        return str(int(xf))
    return f"{xf:g}"


def band_minutes(thickness: float) -> int:
    """Minutes-per-piece band for a plate thickness (inches).

    Lower-inclusive / upper-exclusive; the >= band wins on a boundary.
    """
    if thickness < 0.5:
        return 6
    if thickness < 1.0:
        return 9
    if thickness < 2.0:
        return 12
    return 18


def row_estimate_hours(thickness: Optional[float], ppn_qty: float, scope: str):
    """Legacy thickness-band estimate, used only as a fallback for rows with no
    DXF geometry (structural shapes / a missing file). Returns
    (est_hours, factor_min_per_pc, 0.0), or (None, None, None) when thickness is
    unknown. Bevel adds no time (removed per planning direction); ``scope`` is
    kept for signature stability.
    """
    if thickness is None or thickness <= 0:
        return None, None, None
    qty = ppn_qty if (ppn_qty and ppn_qty > 0) else 0
    factor = thickness * band_minutes(thickness)          # min/pc
    est_min = factor * qty
    return est_min / 60.0, factor, 0.0


def plate_cut_estimate(linear_in_pc, thickness, ppn_qty):
    """Linear-inch cut-time estimate for a plate row that resolved a DXF.

    Returns a dict with the per-piece cut length, feed rate, the two Est Min
    breakouts, and Est Cut Hours -- or None when it can't be computed (no DXF
    match or no thickness). Cut time = cut length / feed rate (IPM); the feed
    rate rounds thickness UP to the next chart row. Bevel adds NO time (the old
    +180 min/pc surcharge was removed per planning direction); bevel scope is
    still flagged for review.
        Est Min/PC   = linear_in_pc / feed_ipm
        Est Min/WO   = Est Min/PC * qty
        Est Cut Hours = Est Min/WO / 60
    """
    if linear_in_pc is None or thickness is None or thickness <= 0:
        return None
    feed = lookup_feed(thickness)
    if feed is None or not feed[0]:
        return None
    ipm, note = feed
    qty = ppn_qty if (ppn_qty and ppn_qty > 0) else 0
    min_pc = linear_in_pc / ipm
    min_wo = min_pc * qty
    return {
        "linear_in_pc": linear_in_pc,
        "total_linear_in": linear_in_pc * qty,
        "feed_ipm": ipm,
        "feed_note": note,
        "min_pc": min_pc,
        "min_wo": min_wo,
        "hours": min_wo / 60.0,
    }


def _is_bevel(scope) -> bool:
    """True if the scope-of-work text contains 'BEVEL' (case-insensitive).

    Real batch lists use compound scopes like 'CUT/BEVEL', so this is a
    CONTAINS test, not exact-equals.
    """
    return scope is not None and "BEVEL" in str(scope).strip().upper()


# Material-form tokens that mark a row as a structural shape (NOT a plate),
# matched against the Description's LEADING form word. Real batch-list
# descriptions lead with the stock form, e.g. 'ANGLE, HIGH STR, , STL, ...',
# 'TUBE, SQUARE, ...', 'T-SECTION, ...', 'BAR, FLAT, ...', 'ROD, ROUND, ...'.
# Shapes routinely carry a thickness token ('2.500 X 2.500 X 0.188 THK',
# '0.500THK3.000W') in their description, so they MUST be classified by FORM,
# not by whether a thickness/estimate was found.
_NONPLATE_FORMS = ("ANGLE", "TUBE", "PIPE", "BAR", "ROD", "CHANNEL", "BEAM",
                   "T-SECTION", "T SECTION", "I-SECTION", "I SECTION",
                   "W-SECTION", "W SECTION", "SECTION")


def _is_plate_row(row: dict) -> bool:
    """True if a data row is a plate (vs a structural shape/bar/tube).

    Classification is DESCRIPTION-driven, keyed on the leading material-form
    token (the text before the first ',' or ';'):
      * leading form is a known shape (ANGLE/TUBE/T-SECTION/BAR/ROD/...) -> NOT a
        plate. This is the fix: shapes carry a thickness token in their
        description (e.g. 'TUBE, SQUARE, ..., 0.188 THK') so the old
        'estimate-was-computed' rule misfiled every dimensioned shape onto the
        Plates sheet;
      * description says PLATE -> plate;
      * otherwise (blank/unknown description) fall back to 'an estimate was
        computed' -- a nest-packet plate thickness implies a plate.
    """
    desc = str(row.get("Description") or "").upper()
    lead = re.split(r"[;,]", desc, maxsplit=1)[0].strip()
    if any(lead.startswith(f) for f in _NONPLATE_FORMS):
        return False
    if "PLATE" in desc:
        return True
    return row.get(EST_COL) is not None


# ──────────────────────────────────────────────────────────────────────────
# PDF field parsing (mirrors 911 Setup's labeled-value approach)
# ──────────────────────────────────────────────────────────────────────────

_LABEL_RE = re.compile(r'^[A-Z][A-Z0-9 ./#-]*:')
_NUM_RE = re.compile(r'^-?\d*\.?\d+$')
_CUT_METHOD_RE = re.compile(r'\b(PLASMA|LASER|WATER\s*JET|WATERJET|WJET|TELESIS|OXY\w*)\b',
                            re.IGNORECASE)
# "1.000 THK #109X360"  /  "0.500X90X360"  /  "0.375 THICK"
_SIZE_LW_HASH_RE = re.compile(r'#\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)')
_SIZE_LW_PLAIN_RE = re.compile(
    r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)')
_THK_FROM_DESC_RE = re.compile(r'(\d*\.?\d+)\s*TH(?:K|ICK)', re.IGNORECASE)


def _looks_like_label(s: str) -> bool:
    return bool(_LABEL_RE.match(s.strip()))


def _labeled_value(text: str, label: str):
    """Value following 'label:' -- same line or the next line. Returns None
    when blank or when the following token is itself another field label."""
    pat = re.compile(
        re.escape(label) + r'\s*:[ \t]*(?P<same>[^\n]*)(?:\n[ \t]*(?P<next>[^\n]*))?',
        re.IGNORECASE,
    )
    m = pat.search(text)
    if not m:
        return None
    same = (m.group("same") or "").strip()
    if same and not _looks_like_label(same):
        return same
    nxt = (m.group("next") or "").strip()
    if nxt and not _looks_like_label(nxt):
        return nxt
    return None


def _find_nest_pdf(folder: Path, nest_number: str):
    """The PDF in `folder` whose stem contains the nest number, or None."""
    if not folder.exists():
        return None
    nest_upper = nest_number.upper()
    for f in folder.iterdir():
        if (f.is_file() and f.suffix.lower() == ".pdf"
                and nest_upper in f.stem.upper()):
            return f
    return None


def _page1_pieces_thickness(page1_text: str):
    """Pull (pieces, thickness) from the page-1 'Orders Pieces Thickness Length
    Width' data row. Labels appear as a block, then the values as a block, so
    we find the label block and read the value sequence after it.

    Returns (pieces|None, thickness|None).
    """
    lines = [ln.strip() for ln in page1_text.splitlines()]
    labels = ["ORDERS", "PIECES", "THICKNESS", "LENGTH", "WIDTH"]
    start = None
    for i in range(len(lines) - len(labels) + 1):
        if [lines[i + k].upper() for k in range(len(labels))] == labels:
            start = i + len(labels)
            break
    if start is None:
        return None, None
    # Collect the numeric value tokens that follow (skip any trailing label
    # like 'Remnant Used'). Order is Orders, Pieces, Thickness, Length, Width.
    nums = []
    for ln in lines[start:]:
        if _NUM_RE.match(ln):
            nums.append(ln)
        if len(nums) >= 3:
            break
    pieces = _to_float(nums[1]) if len(nums) >= 2 else None
    thickness = _to_float(nums[2]) if len(nums) >= 3 else None
    # Pieces is a count -> int-ish
    if pieces is not None:
        pieces = int(round(pieces))
    return pieces, thickness


def _parse_size(full_text: str):
    """Return (stock_L, stock_W) from a part-sketch 'SIZE:' field, normalized
    L=max, W=min. Thickness-only sizes (e.g. '0.375 THICK') -> (None, None).

    A packet can carry more than one 'SIZE:' field (e.g. the nest sheet's
    usable area '85.06 X 53.74' AND the part-sketch stock size
    '1.000 THK #109X360'). We scan every SIZE value and prefer a stock-format
    match: the '#L X W' hash form first, then the 'thk X L X W' three-number
    form.
    """
    values = []
    for m in re.finditer(r'SIZE\s*:[ \t]*([^\n]*)(?:\n[ \t]*([^\n]*))?', full_text,
                         re.IGNORECASE):
        same = (m.group(1) or "").strip()
        cand = same if (same and not _looks_like_label(same)) else (m.group(2) or "").strip()
        if cand and not _looks_like_label(cand):
            values.append(cand)
    # Prefer the '#L X W' hash form.
    for v in values:
        mm = _SIZE_LW_HASH_RE.search(v)
        if mm:
            a, b = float(mm.group(1)), float(mm.group(2))
            return max(a, b), min(a, b)
    # Then the 'thk X L X W' three-number form (ignore the leading thickness).
    for v in values:
        mm = _SIZE_LW_PLAIN_RE.search(v)
        if mm:
            dims = sorted((float(mm.group(2)), float(mm.group(3))))
            return dims[1], dims[0]
    return None, None


def parse_nest_pdf(pdf_path: Path) -> dict:
    """Parse a nest packet PDF. Returns a dict of extracted fields (any may be
    None). Thickness is the only field the calculation needs."""
    out = {"thickness": None, "pieces": None, "material": None,
           "stock_l": None, "stock_w": None, "mil_spec": None,
           "plate_weight": None, "process": None, "mt_material": None}
    sdk.ensure_local(pdf_path)  # OneDrive placeholder -> download first (Hard Rule 13)
    doc = fitz.open(str(pdf_path))
    try:
        pages = [p.get_text() for p in doc]
    finally:
        doc.close()
    full = "\n".join(pages)
    page1 = pages[0] if pages else ""

    out["pieces"], out["thickness"] = _page1_pieces_thickness(page1)

    # Cut method / process from page 1 (PLASMA / LASER / ...).
    m = _CUT_METHOD_RE.search(page1)
    if m:
        out["process"] = m.group(1).upper().replace("  ", " ")

    # Material: REMNANT 'TYPE:' OR part-sketch 'MATL:' OR move-ticket 'MATERIAL:'.
    out["material"] = (_labeled_value(full, "TYPE")
                       or _labeled_value(full, "MATL")
                       or _labeled_value(full, "MATERIAL"))
    # MOVE TICKET material designation (e.g. 'HSS'), distinct from the batch
    # list's stock 'Material' code -- pulled only from MOVE TICKET pages.
    out["mt_material"] = _move_ticket_material(pages)
    # Mil spec.
    out["mil_spec"] = _labeled_value(full, "MIL SPEC")
    if not out["mil_spec"]:
        mm = re.search(r'MIL-S-\S+', full, re.IGNORECASE)
        out["mil_spec"] = mm.group(0) if mm else None
    # Stock L x W from a part sketch.
    out["stock_l"], out["stock_w"] = _parse_size(full)
    # Plate weight: REMNANT 'WEIGHT: n LBS' else page-1 'Plate/Max Part Weight'.
    out["plate_weight"] = _parse_plate_weight(full, page1)

    # Thickness fallback from REMNANT 'THICK:' if page-1 row missed it.
    if out["thickness"] is None:
        tv = _labeled_value(full, "THICK")
        if tv:
            out["thickness"] = _to_float(tv)
    return out


def _move_ticket_material(pages) -> Optional[str]:
    """The 'MATERIAL:' value from a MOVE TICKET page (the part-material
    designation, e.g. 'HSS') -- the value that sits right after 'MIL SPEC:' on
    the move ticket. Returns the first non-empty one found, scanning only pages
    that mention MOVE TICKET so a stray 'MATERIAL' elsewhere can't win."""
    for txt in pages:
        if "MOVE TICKET" in txt.upper():
            val = _labeled_value(txt, "MATERIAL")
            if val:
                return val
    return None


def _parse_plate_weight(full: str, page1: str):
    w = _labeled_value(full, "WEIGHT")
    if w:
        mm = re.search(r'(\d+(?:\.\d+)?)', w)
        if mm:
            return _to_float(mm.group(1))
    # page-1 'Plate/Max Part Weight  1181 / 3' -> first number before slash.
    mm = re.search(r'Plate/Max Part Weight\s*[^0-9]*(\d+(?:\.\d+)?)', page1,
                   re.IGNORECASE)
    if mm:
        return _to_float(mm.group(1))
    return None


def _to_float(v):
    if v is None:
        return None
    try:
        return float(str(v).strip().replace(",", ""))
    except (ValueError, TypeError):
        return None


def thickness_from_description(desc) -> Optional[float]:
    """Fallback thickness parse from a batch-list Description like
    'PLATE ; STL ; 1.125 THK #90 X 360 ; SF'."""
    if not desc:
        return None
    m = _THK_FROM_DESC_RE.search(str(desc))
    return _to_float(m.group(1)) if m else None


# ──────────────────────────────────────────────────────────────────────────
# Folder + batch-list resolution
# ──────────────────────────────────────────────────────────────────────────

def _looks_like_order(folder: Path) -> bool:
    """True if `folder` is a single order/batch folder — it directly holds a
    'CUI- TECH DATA READ ME' folder, a *BATCH*LIST*.xlsx, or the packet
    folders themselves."""
    try:
        for sub in folder.iterdir():
            name_up = sub.name.upper()
            if sub.is_dir():
                if "CUI" in name_up and "TECH" in name_up:
                    return True
                if name_up in ("NEST PACKAGES", "WPDD SKETCHES"):
                    return True
            elif (sub.suffix.lower() == ".xlsx" and "BATCH" in name_up
                    and "LIST" in name_up and not sub.name.startswith("~$")):
                return True
    except OSError:
        return False
    return False


def discover_order_folders(root: Path, log) -> list:
    """Collect the order folders under ROOT, tolerating a ROOT picked a level
    (or several) too high.

    Award folders and zip extractions add wrapper levels (e.g.
    'Award 10\\1000129724 SSPO Award 10\\1000129724 SSPO Award 10\\{L028,...}').
    The old ROOT.iterdir()-only discovery treated a wrapper as ONE order, then
    the CUI/batch-list fallbacks silently processed a single arbitrary batch
    list from somewhere inside it (bit A.T.'s Award 10 runs: 1 of 4 batches,
    no packet PDFs, everything mis-filed). Now: a ROOT subfolder that doesn't
    itself look like an order is descended into (depth-capped) until real
    order folders are found; picking an order folder itself as ROOT works too.
    """
    if _looks_like_order(root):
        log("ROOT is itself a single order folder.")
        return [root]

    found = []

    def walk(folder: Path, depth: int) -> None:
        try:
            subs = sorted((d for d in folder.iterdir()
                           if d.is_dir() and not d.name.startswith(".")),
                          key=lambda p: p.name.upper())
        except OSError:
            return
        for d in subs:
            if _looks_like_order(d):
                found.append(d)
            elif depth < 4:
                before = len(found)
                walk(d, depth + 1)
                if len(found) > before:
                    log(f"  NOTE: '{d.name}' is a wrapper folder - using the "
                        f"order folders inside it.")

    walk(root, 1)
    return found


def find_cui_folder(order_folder: Path) -> Optional[Path]:
    """Locate the 'CUI- TECH DATA READ ME' folder (name varies in spacing/dash).
    Falls back to any subfolder that itself contains a NEST PACKAGES / WPDD
    SKETCHES folder or a *BATCH*LIST*.xlsx. Returns the order folder itself if
    the packets live directly under it."""
    for sub in order_folder.iterdir():
        if sub.is_dir() and "CUI" in sub.name.upper() and "TECH" in sub.name.upper():
            return sub
    # Fallback: a subfolder holding the packet folders or a batch list.
    for sub in order_folder.iterdir():
        if not sub.is_dir():
            continue
        names = {c.name.upper() for c in sub.iterdir() if c.is_dir()} if sub.exists() else set()
        if "NEST PACKAGES" in names or "WPDD SKETCHES" in names:
            return sub
        if any("BATCH" in f.name.upper() and "LIST" in f.name.upper()
               for f in sub.glob("*.xlsx")):
            return sub
    # Maybe packets are directly under the order folder.
    return order_folder


def find_pdf_folder(cui_folder: Path) -> Optional[Path]:
    """Prefer 'NEST PACKAGES' (nest-named packet PDFs); fall back to
    'WPDD SKETCHES', then the CUI folder itself."""
    for name in ("NEST PACKAGES", "WPDD SKETCHES"):
        cand = cui_folder / name
        if cand.exists() and any(cand.glob("*.pdf")):
            return cand
    # Case-insensitive scan.
    for sub in cui_folder.iterdir():
        if sub.is_dir() and sub.name.upper() in ("NEST PACKAGES", "WPDD SKETCHES"):
            if any(sub.glob("*.pdf")):
                return sub
    if any(cui_folder.glob("*.pdf")):
        return cui_folder
    return None


def find_iges_folder(cui_folder: Path) -> Optional[Path]:
    """Locate the 'IGES FILES' folder (holds the nest DXFs) under the CUI folder,
    case-insensitively. None if absent (e.g. a structural-only order)."""
    if cui_folder is None:
        return None
    for sub in cui_folder.iterdir():
        if sub.is_dir() and sub.name.strip().upper() == "IGES FILES":
            return sub
    return None


def find_batch_list(cui_folder: Path, log=None) -> Optional[Path]:
    """Glob for the '*BATCH*LIST*.xlsx' workbook, skipping Excel lock files."""
    candidates = [f for f in cui_folder.glob("*.xlsx")
                  if "BATCH" in f.name.upper() and "LIST" in f.name.upper()
                  and not f.name.startswith("~$")]
    if candidates:
        return candidates[0]
    # Deep fallback. Multiple hits at different depths is the wrong-ROOT
    # signature (a wrapper folder holding several batches) — never silently
    # pick one; take the shallowest and say so.
    deep = [f for f in cui_folder.rglob("*.xlsx")
            if "BATCH" in f.name.upper() and "LIST" in f.name.upper()
            and not f.name.startswith("~$")]
    if not deep:
        return None
    deep.sort(key=lambda f: (len(f.parts), str(f).upper()))
    if len(deep) > 1 and log:
        log(f"  WARNING: {len(deep)} BATCH LIST workbooks found under "
            f"'{cui_folder.name}' - it looks like a wrapper of several "
            f"batches, not one order. Using {deep[0].name}; check the ROOT "
            f"you selected.")
    return deep[0]


def load_batch_rows(batch_list_path: Path):
    """Return (headers_in_order, rows) from the batch-list sheet.

    headers_in_order : list[str]  -- the header names in column order
    rows             : list[dict] -- {HEADER_UPPER: value} per data row

    Locates the sheet + header row by scanning for 'Nest Pkg Nbr'/'PPN Quantity'
    so it survives column/row shifts and extra sheets (SCRAP, Sheet3 ...).
    """
    wb = sdk.load_workbook_resilient(batch_list_path, data_only=True)
    try:
        target_ws = None
        header_row = None
        cols = {}
        for ws in wb.worksheets:
            hr, cmap = sdk.find_header_row(ws, ["Nest Pkg Nbr", "PPN Quantity"])
            if hr is not None:
                target_ws, header_row, cols = ws, hr, cmap
                break
        if target_ws is None:
            return [], []

        # Header names in column order (preserve original casing for output).
        ordered = []
        for c in range(1, target_ws.max_column + 1):
            v = target_ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip():
                ordered.append((c, v.strip()))
        headers_in_order = [name for _, name in ordered]

        rows = []
        nest_col = cols.get("NEST PKG NBR")
        for r in range(header_row + 1, target_ws.max_row + 1):
            # Skip fully blank rows and rows without a nest.
            nest_val = target_ws.cell(r, nest_col).value if nest_col else None
            if nest_val is None or str(nest_val).strip() == "":
                continue
            row = {}
            for c, name in ordered:
                row[name.upper()] = target_ws.cell(r, c).value
            rows.append(row)
        return headers_in_order, rows
    finally:
        wb.close()


# ──────────────────────────────────────────────────────────────────────────
# Workbook output
# ──────────────────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="305496")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_PIVOT_FILL = PatternFill("solid", fgColor="1F4E78")
_TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
_MISSING_FILL = PatternFill("solid", fgColor="FFFF00")   # no estimate (yellow)
_OUTLIER_FILL = PatternFill("solid", fgColor="F4B084")   # LI outlier (orange)
_MULTI_FILL = PatternFill("solid", fgColor="DDEBF7")     # multi-layered (light blue)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _is_date(v):
    return isinstance(v, (datetime.datetime, datetime.date))


def write_workbook(out_path: Path, data_headers, plate_rows, nonplate_rows, log):
    """Write the workbook with two identically-structured sheets: 'Plates' and
    'Non-Plates'. Each holds ONLY the flat data table (every batch-list column
    plus our generated columns) with the yellow missing-data highlighting; the
    real PivotTables are added afterwards via Excel COM (see add_real_pivots).

    Returns sheets_meta: [(sheet_name, data_row_count), ...] for the pivot pass.
    """
    wb = Workbook()
    ws_plate = wb.active
    ws_plate.title = "Plates"
    write_data_table(ws_plate, data_headers, plate_rows)

    ws_other = wb.create_sheet("Non-Plates")
    write_data_table(ws_other, data_headers, nonplate_rows)

    wb.save(str(out_path))
    return [("Plates", len(plate_rows)), ("Non-Plates", len(nonplate_rows))]


def write_data_table(ws, data_headers, data_rows):
    """Write one sheet's flat data table (header row 1, then data). Row highlights:
    orange = LI outlier (past LI_OUTLIER_SD), yellow = no estimate, light blue =
    multi-layered. No summary block -- pivots are layered on later."""
    for c, name in enumerate(data_headers, start=1):
        cell = ws.cell(1, c, name)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER

    est_idx = data_headers.index(EST_COL) + 1
    for ri, row in enumerate(data_rows, start=2):
        # Pick one row highlight (most-important wins): LI outlier > no-estimate
        # > multi-layered.
        is_outlier = "outlier" in str(row.get("Flags") or "").lower()
        missing = row.get(EST_COL) is None
        is_multi = row.get("Multi-Layered") == "Y"
        row_fill = (_OUTLIER_FILL if is_outlier else _MISSING_FILL if missing
                    else _MULTI_FILL if is_multi else None)
        for c, name in enumerate(data_headers, start=1):
            val = row.get(name)
            cell = ws.cell(ri, c, val)
            cell.border = _BORDER
            if row_fill is not None:
                cell.fill = row_fill
            if _is_date(val):
                cell.number_format = "MM/DD/YYYY"
            if c == est_idx and isinstance(val, (int, float)):
                cell.number_format = "0.000"
    ws.freeze_panes = "A2"
    _autosize(ws, data_headers)


def _nest_sums(data_rows):
    """Aggregate rows to nest -> summed Est Cut Hours. Returns
    (sums: dict[nest -> hours], grand_est, grand_qty)."""
    sums = {}
    grand_est = 0.0
    grand_qty = 0.0
    for row in data_rows:
        est = row.get(EST_COL) or 0.0
        qty = _to_float(row.get("PPN Quantity")) or 0.0
        nest = str(row.get("Nest Pkg Nbr") or "").strip()
        sums[nest] = sums.get(nest, 0.0) + est
        grand_est += est
        grand_qty += qty
    return sums, grand_est, grand_qty


# Excel COM enum constants we need under late binding (no makepy types).
_XL_DATABASE = 1        # PivotCache SourceType (xlDatabase)
_XL_ROW_FIELD = 1       # PivotField.Orientation (xlRowField)
_XL_SUM = -4157         # xlConsolidationFunction (xlSum)
_XL_CALC_MANUAL = -4135 # xlCalculationManual


def add_real_pivots(out_path, data_headers, sheets_meta, log):
    """Open the saved workbook in Excel via COM and drop a REAL PivotTable below
    the data table on each sheet: row field = Nest Pkg Nbr, data field =
    Sum of Est Cut Hours, with the grand-total row turned on. Returns True on
    success. Requires Excel + pywin32; on any failure returns False so the
    caller can fall back to a static summary.

    Runs on the plugin's worker thread, so COM MUST be initialised on this
    thread (CoInitialize) before any Dispatch call.
    """
    try:
        import pythoncom
        import win32com.client as win32
    except ImportError:
        log("  Excel COM (pywin32) unavailable; writing static summary instead.")
        return False

    if EST_COL not in data_headers:
        return False
    nest_header = next((h for h in data_headers if h.upper() == "NEST PKG NBR"),
                       "Nest Pkg Nbr")
    qty_header = next((h for h in data_headers if h.upper() == "PPN QUANTITY"), None)
    last_col = get_column_letter(len(data_headers))

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        # Dedicated, isolated instance (DispatchEx) -- kill the churn that makes
        # COM pivot work slow: no screen repaints, no events, no auto-recalc
        # after every edit. We Quit the instance afterwards, so no need to
        # restore these.
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        wb = excel.Workbooks.Open(str(out_path))
        # Application.Calculation can only be set once a workbook is open.
        # Setting it on a fresh, workbook-less instance raises "Unable to set
        # the Calculation property of the Application class" -- which aborted
        # the ENTIRE pivot pass, so every run silently fell back to the static
        # summary and no real PivotTable was ever created. Set it AFTER Open.
        excel.Calculation = _XL_CALC_MANUAL
        for sheet_name, n_rows in sheets_meta:
            if n_rows < 1:
                continue
            ws = wb.Worksheets(sheet_name)
            last_row = n_rows + 1  # +1 for the header row
            src = ws.Range(f"A1:{last_col}{last_row}")
            # Plain-English caveat right above the pivot so a small per-nest
            # number is never misread as total labor time.
            ws.Cells(last_row + 2, 1).Value = (
                "Machine cut time only (torch-on). Excludes pierce, "
                "positioning, and handling time.")
            dest = ws.Cells(last_row + 4, 1)  # blank + caveat rows under the data
            cache = wb.PivotCaches().Create(SourceType=_XL_DATABASE, SourceData=src)
            table_name = sheet_name.replace("-", "") + "Pivot"
            pt = cache.CreatePivotTable(TableDestination=dest, TableName=table_name)
            # Defer recalculation until every field is configured. With the
            # default ManualUpdate=False, Excel rebuilds the WHOLE pivot after
            # each field change. Batching them into one recalc is the fix.
            pt.ManualUpdate = True
            # ONE row per nest. Under it, three data columns so the number is
            # self-explanatory: Parts = the true piece count for the WHOLE nest
            # (all quantities summed), then the same cut-time total in BOTH
            # hours and minutes. (PPN Quantity is no longer a row field -- that
            # split each nest into confusing per-quantity slivers, e.g.
            # "5CDAZP / 8 / 0.013", which read as "8 parts = 0.013 hr".)
            pt.PivotFields(nest_header).Orientation = _XL_ROW_FIELD
            if qty_header:
                dparts = pt.AddDataField(pt.PivotFields(qty_header),
                                         "Parts", _XL_SUM)
                dparts.NumberFormat = "0"
            dhr = pt.AddDataField(pt.PivotFields(EST_COL),
                                  "Cut Time (hr)", _XL_SUM)
            dhr.NumberFormat = "0.000"
            # Minutes = hours x 60, as a calculated field so the minutes column
            # is always an EXACT conversion of the hours column (never drifts).
            try:
                pt.CalculatedFields().Add("CutMinutes", f"='{EST_COL}' *60")
                dmin = pt.AddDataField(pt.PivotFields("CutMinutes"),
                                       "Cut Time (min)", _XL_SUM)
                dmin.NumberFormat = "0.0"
            except Exception:
                pass  # minutes is a convenience column; hours is the source of truth
            # ColumnGrand = the bottom total row (total of all nests) -- what we
            # want. RowGrand would add a redundant rightmost total column.
            pt.ColumnGrand = True
            pt.RowGrand = False
            pt.ManualUpdate = False   # one recalc/re-layout, now that we're done
        wb.Save()
        wb.Close(SaveChanges=True)
        log("  Added real Excel PivotTables (per nest -> Parts + Cut Time "
            "in hr and min).")
        return True
    except Exception as e:
        log(f"  Excel pivot creation failed ({e}); writing static summary instead.")
        return False
    finally:
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def add_static_pivots(out_path, plate_rows, nonplate_rows, log):
    """Fallback when Excel COM is unavailable: reopen the workbook with openpyxl
    and append a static nest -> Sum of Est Cut Hours summary (+ grand total) to
    each sheet, mirroring what the real PivotTable would have shown."""
    wb = openpyxl.load_workbook(out_path)
    for sheet_name, rows in (("Plates", plate_rows), ("Non-Plates", nonplate_rows)):
        if sheet_name in wb.sheetnames:
            _write_static_summary(wb[sheet_name], rows)
    wb.save(out_path)
    log("  Wrote static nest summary (Excel PivotTable unavailable).")


def _write_static_summary(ws, data_rows):
    """Append a per-nest summary (+ Grand Total) below the data table, mirroring
    the real pivot: ONE line per nest with Parts (the true piece count for the
    whole nest), then the same cut-time total in BOTH hours and minutes."""
    agg, order = {}, []           # nest -> [parts, hours]
    for row in data_rows:
        est = row.get(EST_COL) or 0.0
        qty = _to_float(row.get("PPN Quantity")) or 0.0
        nest = str(row.get("Nest Pkg Nbr") or "").strip()
        if nest not in agg:
            agg[nest] = [0.0, 0.0]
            order.append(nest)
        agg[nest][0] += qty
        agg[nest][1] += est
    start = ws.max_row + 3
    # Plain-English caveat above the summary.
    ws.cell(start, 1, "Machine cut time only (torch-on). Excludes pierce, "
                      "positioning, and handling time.")
    start += 1
    for c, label in ((1, "Nest Pkg Nbr"), (2, "Parts"),
                     (3, "Cut Time (hr)"), (4, "Cut Time (min)")):
        cell = ws.cell(start, c, label)
        cell.fill = _PIVOT_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = _BORDER
    r = start + 1
    tot_parts = tot_hr = 0.0
    for nest in sorted(order):
        parts, hours = agg[nest]
        tot_parts += parts
        tot_hr += hours
        ws.cell(r, 1, nest)
        p = ws.cell(r, 2, round(parts)); p.number_format = "0"
        h = ws.cell(r, 3, round(hours, 3)); h.number_format = "0.000"
        m = ws.cell(r, 4, round(hours * 60, 1)); m.number_format = "0.0"
        r += 1
    ws.cell(r, 1, "Grand Total")
    gp = ws.cell(r, 2, round(tot_parts)); gp.number_format = "0"
    gh = ws.cell(r, 3, round(tot_hr, 3)); gh.number_format = "0.000"
    gm = ws.cell(r, 4, round(tot_hr * 60, 1)); gm.number_format = "0.0"
    for c in (1, 2, 3, 4):
        ws.cell(r, c).fill = _TOTAL_FILL
        ws.cell(r, c).font = Font(bold=True)
        ws.cell(r, c).border = _BORDER


def _autosize(ws, data_headers):
    widths = {}
    for c, name in enumerate(data_headers, start=1):
        widths[c] = max(len(str(name)), 10)
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 400)):
        for cell in row:
            if cell.value is not None and cell.column <= len(data_headers):
                widths[cell.column] = max(widths.get(cell.column, 10),
                                          min(len(str(cell.value)) + 2, 45))
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ──────────────────────────────────────────────────────────────────────────
# Main entry point
# ──────────────────────────────────────────────────────────────────────────

def run(params: dict, progress_callback, cancel_event):
    log = params.get("log", print)
    settings = params.get("settings", {}) or {}

    if not PYMUPDF_AVAILABLE:
        log("ERROR: PyMuPDF (fitz) is not available; cannot parse PDFs.")
        return

    log("=" * 60)
    log(f"911 Runtime Estimator v{VERSION}")
    log("=" * 60)

    start_dir = ""
    try:
        roots = sdk.pilot_program_roots()
        if roots:
            start_dir = str(roots[0])
    except Exception:
        pass
    raw = sdk.request_directory(params, "Select the ROOT folder of order folders",
                                start_dir)
    if cancel_event.is_set():
        return
    if not raw:
        log("Folder selection cancelled - nothing was run.")
        cancel_event.set()  # user cancel: not a successful (ticket-earning) run
        return
    root = Path(raw.strip().strip('"'))
    if not root.exists() or not root.is_dir():
        log(f"ERROR: ROOT directory not found: {root}")
        return
    log(f"ROOT: {root}")

    order_folders = discover_order_folders(root, log)
    if not order_folders:
        log("ERROR: No order folders found under ROOT (no folder holding a "
            "CUI- TECH DATA READ ME folder, a BATCH LIST workbook, or NEST "
            "PACKAGES). Pick the folder whose subfolders are the orders.")
        return
    log(f"Found {len(order_folders)} order folder(s): "
        f"{', '.join(o.name for o in order_folders)}")
    progress_callback(5)

    # Union of all batch-list headers (in first-seen column order) -> data cols.
    header_order = []           # batch-list headers, ordered
    header_seen = set()
    data_rows = []
    flags_summary = []

    for oi, order_folder in enumerate(order_folders):
        if cancel_event.is_set():
            log("Cancelled.")
            return
        order = order_folder.name
        log(f"\n--- Order {oi + 1}/{len(order_folders)}: {order} ---")

        cui = find_cui_folder(order_folder)
        pdf_folder = find_pdf_folder(cui) if cui else None
        batch_list = find_batch_list(cui, log=log) if cui else None

        if batch_list is None:
            log(f"  WARNING: No BATCH LIST workbook found for {order} -- skipping.")
            flags_summary.append(f"{order}: no batch list")
            continue
        log(f"  Batch list: {batch_list.name}")
        if pdf_folder:
            log(f"  PDF folder: {pdf_folder.name}")
        else:
            log(f"  WARNING: No PDF folder with packets found for {order}.")

        # Exact per-part cut length from the order's IGES DXFs, keyed by work
        # order (matches the batch list's 'Work Order' column). Absent for
        # structural-only orders -> those rows fall back to the band estimate.
        iges_folder = find_iges_folder(cui) if cui else None
        dxf_index, dxf_suspect, dxf_multi = build_dxf_index(
            str(iges_folder) if iges_folder else None, log)
        if iges_folder:
            log(f"  IGES DXFs: {len(dxf_index)} work order(s) measured.")

        headers, rows = load_batch_rows(batch_list)
        if not rows:
            log(f"  WARNING: No data rows in {batch_list.name}.")
            continue
        for h in headers:
            if h.upper() not in header_seen:
                header_seen.add(h.upper())
                header_order.append(h)

        # Parse each unique nest's PDF once.
        nests_in_list = []
        seen_nests = []
        for row in rows:
            n = row.get("NEST PKG NBR")
            if n is not None and str(n).strip():
                ns = str(n).strip()
                if ns not in seen_nests:
                    seen_nests.append(ns)
        nest_pdf_data = {}
        for ns in seen_nests:
            pdf = _find_nest_pdf(pdf_folder, ns) if pdf_folder else None
            if pdf is None:
                nest_pdf_data[ns] = None
                log(f"  FLAG: nest {ns} in batch list has no packet PDF "
                    f"(thickness from Description).")
                flags_summary.append(f"{order}/{ns}: no PDF")
                continue
            try:
                nest_pdf_data[ns] = parse_nest_pdf(pdf)
            except Exception as e:
                nest_pdf_data[ns] = None
                log(f"  WARNING: could not parse {pdf.name}: {e}")

        # Flag PDFs present but not referenced by the batch list.
        if pdf_folder:
            for pdf in pdf_folder.glob("*.pdf"):
                if not any(ns.upper() in pdf.stem.upper() for ns in seen_nests):
                    log(f"  FLAG: packet {pdf.name} not referenced in batch list.")
                    flags_summary.append(f"{order}: orphan PDF {pdf.name}")

        # Build a data row per batch-list row.
        for row in rows:
            nest = str(row.get("NEST PKG NBR") or "").strip()
            pdfd = nest_pdf_data.get(nest)
            ppn = _to_float(row.get("PPN QUANTITY")) or 0.0
            scope = row.get("SCOPE OF WORK") or row.get("SCOPE OF WORK ")
            thickness = (pdfd or {}).get("thickness") if pdfd else None
            thk_src = "PDF"
            if thickness is None:
                thickness = thickness_from_description(row.get("DESCRIPTION"))
                thk_src = "Description" if thickness is not None else "MISSING"

            # Cut time from EXACT DXF linear inches / feed rate when the row's
            # work order resolves a DXF; else fall back to the thickness band.
            wo = str(row.get("WORK ORDER") or "").strip().upper()
            li_pc = dxf_index.get(wo)
            bevel = _is_bevel(scope)
            est = plate_cut_estimate(li_pc, thickness, ppn)

            out_row = {}
            # Copy every batch-list field by header.
            for h in headers:
                out_row[h] = row.get(h.upper())
            out_row["Order"] = order
            out_row["Division"] = ""      # blank placeholder (planner fills in)
            # Location / Process: from batch list if present, else PDF-derived.
            out_row["Location"] = row.get("LOCATION")
            proc = row.get("PROCESS")
            if not proc and pdfd:
                proc = pdfd.get("process")
            out_row["Process"] = proc
            out_row["Thickness (in)"] = thickness
            out_row["Stock L"] = (pdfd or {}).get("stock_l") if pdfd else None
            out_row["Stock W"] = (pdfd or {}).get("stock_w") if pdfd else None
            out_row["Plate Weight"] = (pdfd or {}).get("plate_weight") if pdfd else None
            out_row["Mil Spec"] = (pdfd or {}).get("mil_spec") if pdfd else None
            out_row["MT Material"] = (pdfd or {}).get("mt_material") if pdfd else None

            flags = []
            # LI Dev (SD) is filled in after all rows are built (needs the mean/std
            # across the whole run); leave a placeholder here.
            out_row["LI Dev (SD)"] = None
            if est is not None:
                # Progressive rounding: each cell is derived from the previous
                # ROUNDED cell so the printed Equation reproduces every column
                # exactly (Est Min/PC x Qty == Est Min/WO, etc.).
                li_disp = round(est["linear_in_pc"], 2)
                ipm = est["feed_ipm"]
                mpc_disp = round(est["min_pc"], 2)
                tot_li_disp = round(li_disp * ppn, 2)
                mwo_disp = round(mpc_disp * ppn, 2)
                hours_disp = round(mwo_disp / 60.0, 4)
                out_row["Linear In/PC"] = li_disp
                out_row["Total Linear In"] = tot_li_disp
                out_row["Multi-Layered"] = "Y" if dxf_multi.get(wo) else "N"
                out_row["Feed Rate (IPM)"] = ipm
                out_row["Est Min/PC"] = mpc_disp
                out_row["Est Min/WO"] = mwo_disp
                out_row["Est Cut Hours"] = hours_disp
                out_row["Equation"] = (
                    f"Est Min/PC = Linear In/PC / Feed Rate = "
                    f"{_eqnum(li_disp)} in / {_eqnum(ipm)} IPM = {_eqnum(mpc_disp)} min   |   "
                    f"Est Min/WO = Est Min/PC x Qty = "
                    f"{_eqnum(mpc_disp)} x {_eqnum(ppn)} = {_eqnum(mwo_disp)} min   |   "
                    f"Est Cut Hours = Est Min/WO / 60 = "
                    f"{_eqnum(mwo_disp)} / 60 = {hours_disp:.3f} hr"
                )
                if est["feed_note"]:
                    flags.append(est["feed_note"])
                if wo in dxf_suspect:
                    flags.append(dxf_suspect[wo])
            else:
                # No DXF geometry: structural row, or a plate whose DXF is
                # missing. Keep the thickness-band estimate so hours aren't lost.
                est_hr, factor, _bevel = row_estimate_hours(thickness, ppn, scope)
                out_row["Linear In/PC"] = None
                out_row["Total Linear In"] = None
                out_row["Multi-Layered"] = ""     # N/A without a DXF
                out_row["Feed Rate (IPM)"] = None
                out_row["Est Min/PC"] = None
                out_row["Est Min/WO"] = None
                if est_hr is not None and factor is not None:
                    band = band_minutes(thickness)
                    factor_disp = round(factor, 2)
                    mwo_disp = round(factor_disp * ppn, 2)
                    hours_disp = round(mwo_disp / 60.0, 4)
                    out_row["Est Cut Hours"] = hours_disp
                    out_row["Equation"] = (
                        f"Band fallback (no DXF geometry). "
                        f"Est Min/PC = Thickness x band-factor = "
                        f"{_eqnum(thickness)} x {band} = {_eqnum(factor_disp)} min   |   "
                        f"Est Min/WO = Est Min/PC x Qty = "
                        f"{_eqnum(factor_disp)} x {_eqnum(ppn)} = {_eqnum(mwo_disp)} min   |   "
                        f"Est Cut Hours = Est Min/WO / 60 = "
                        f"{_eqnum(mwo_disp)} / 60 = {hours_disp:.3f} hr"
                    )
                else:
                    out_row["Est Cut Hours"] = None
                    out_row["Equation"] = "No estimate (thickness unknown)."
                if li_pc is None and "PLATE" in str(row.get("DESCRIPTION") or "").upper():
                    flags.append("no DXF match - band estimate")

            if thk_src == "MISSING":
                flags.append("no thickness")
            elif thk_src == "Description":
                flags.append("thk from desc")
            if bevel:
                flags.append("bevel scope (no time added)")
            out_row["Flags"] = "; ".join(flags)
            # Make uppercase-keyed lookups work for pivot/derived fields.
            out_row["Nest Pkg Nbr"] = nest
            out_row["PPN Quantity"] = row.get("PPN QUANTITY")
            out_row["Description"] = row.get("DESCRIPTION")
            data_rows.append(out_row)

        progress_callback(5 + int(85 * (oi + 1) / len(order_folders)))

    if not data_rows:
        log("\nERROR: No rows produced; nothing to write.")
        return

    # Column order: 'Order' first, then every batch-list column verbatim in
    # native order (with the batch list's 'Material' shown as 'Source
    # Material'), then our generated columns. Nothing from the batch list is
    # dropped.
    batch_headers = [BATCH_HEADER_RENAME.get(h, h) for h in header_order]
    seen = {h.upper() for h in batch_headers} | {"ORDER"}
    data_headers = (["Order"] + batch_headers
                    + [c for c in GENERATED_COLS if c.upper() not in seen])

    # Normalize each row's keys to the exact header strings used in output.
    norm_rows = []
    for row in data_rows:
        nr = {}
        low = {k.upper(): v for k, v in row.items()}
        for h in data_headers:
            nr[h] = low.get(HEADER_SOURCE_KEY.get(h, h.upper()))
        # Preserve the canonical pivot keys.
        nr["Nest Pkg Nbr"] = row.get("Nest Pkg Nbr") or low.get("NEST PKG NBR")
        nr["PPN Quantity"] = row.get("PPN Quantity") if row.get("PPN Quantity") is not None else low.get("PPN QUANTITY")
        nr["Description"] = row.get("Description") if row.get("Description") is not None else low.get("DESCRIPTION")
        nr["Est Cut Hours"] = low.get("EST CUT HOURS")
        norm_rows.append(nr)

    # Split into plate vs non-plate (shapes/bars/tubes) — each gets its own
    # identically-structured sheet. A row is a plate if its Description says
    # PLATE or if we managed to compute an estimate (thickness was found).
    plate_rows = [r for r in norm_rows if _is_plate_row(r)]
    nonplate_rows = [r for r in norm_rows if not _is_plate_row(r)]
    log(f"\nBuilding workbook: {len(plate_rows)} plate rows, "
        f"{len(nonplate_rows)} non-plate rows.")

    # Linear-inch deviation: how many standard deviations each plate part's
    # Linear In/PC sits from the mean of all measured plate parts. Parts beyond
    # LI_OUTLIER_SD are flagged + highlighted so an inflated/odd part stands out.
    li_vals = [r.get("Linear In/PC") for r in plate_rows
               if isinstance(r.get("Linear In/PC"), (int, float))]
    if len(li_vals) >= 2:
        li_mean = statistics.mean(li_vals)
        li_sd = statistics.pstdev(li_vals)
        n_outliers = 0
        for r in plate_rows:
            li = r.get("Linear In/PC")
            if isinstance(li, (int, float)) and li_sd > 0:
                z = (li - li_mean) / li_sd
                r["LI Dev (SD)"] = round(z, 2)
                if abs(z) >= LI_OUTLIER_SD:
                    n_outliers += 1
                    note = f"LI outlier {z:+.1f}SD"
                    r["Flags"] = f"{r['Flags']}; {note}" if r.get("Flags") else note
        log(f"  Linear-inch mean {li_mean:.0f} in, SD {li_sd:.0f} in; "
            f"{n_outliers} part(s) beyond {LI_OUTLIER_SD:g} SD flagged.")

    # Output path.
    out_dir = settings.get("output_dir", "").strip()
    out_base = Path(out_dir) if out_dir else root
    stamp = datetime.datetime.now().strftime("%Y-%m-%d")
    out_path = out_base / f"911 RUNTIME ESTIMATOR - {root.name} - {stamp}.xlsx"
    try:
        sheets_meta = write_workbook(out_path, data_headers, plate_rows,
                                     nonplate_rows, log)
    except PermissionError:
        log(f"  Output file is open/locked: {out_path.name}. Close it and re-run.")
        return

    # Layer on the real Excel PivotTables; fall back to a static summary if
    # Excel COM isn't available or errors.
    if not add_real_pivots(out_path, data_headers, sheets_meta, log):
        try:
            add_static_pivots(out_path, plate_rows, nonplate_rows, log)
        except PermissionError:
            log(f"  Output file is open/locked: {out_path.name}. Close it and re-run.")
            return

    _, grand_est, grand_qty = _nest_sums(plate_rows)
    progress_callback(100)
    log("\n" + "=" * 60)
    log(f"DONE. Wrote: {out_path}")
    log(f"  Plate rows: {len(plate_rows)}  |  Non-plate rows: {len(nonplate_rows)}")
    log(f"  Plate cut time (machine only): {grand_est:.2f} hr "
        f"({grand_est * 60:.0f} min) across {int(grand_qty)} pieces")
    if flags_summary:
        log(f"  Flags ({len(flags_summary)}):")
        for f in flags_summary[:25]:
            log(f"    - {f}")
        if len(flags_summary) > 25:
            log(f"    ... and {len(flags_summary) - 25} more")
    log("=" * 60)


if __name__ == "__main__":
    import threading
    run({"log": print}, lambda p: print(f"[{p}%]"), threading.Event())
