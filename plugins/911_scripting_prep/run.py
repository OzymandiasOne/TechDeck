"""
911 Scripting Prep Plugin for TechDeck v1.0.0
Builds the SSPO ERP scripting workbook (PO Data + Part Data) from an award
package's SSPO AWARD REVIEW output and the Working Forecast List.

A SEPARATE app from the 911 PO PDF Extractor, not a replacement for it. That
extractor still reads an INCOMING PO packet from EB and is unchanged; this one
builds an OUTGOING scripting sheet from data a reviewer has already checked.
Different direction, different inputs, different output -- so they ship side by
side. (This briefly shipped as a rewrite of that plugin, 2026-09-02; reverted
the same day.)

What it does:
- Award data comes from the 911 SSPO Award Review workbook's 'Working Forecast
  Input' sheet -- already deduplicated to one line per nest, and already
  reviewed, so nothing here re-derives from the packet PDFs.
- ONE workbook, TWO sheets, matching the hand-built scripting format:
  'PO Data' (17 cols A-Q) and 'Part Data' (18 cols A-R, laid out like the 911
  Sketch Extractor's sheet but keyed per NEST rather than per part).
- A third 'Unresolved' sheet appears only when something needs hand entry --
  same workbook, so the exception list can't get separated from its data, and
  the user deletes the sheet once the values are in.
- MATL / DESC resolve out of the Working Forecast List's 'Inventory Listing'
  sheet, keyed on Source Material -- read from a staged COPY of the live
  workbook, so a run can never touch the shared file (as 911 Setup does).
- Headers on row 1, data from row 2. App-filled columns are written in red so
  the sheet shows at a glance which values it produced. (The hand-built
  reference had a source-tag row above the headers; those were the author's
  notes and are not written -- the tags still drive the code.)

Prompts at runtime for:
- The 911 SSPO AWARD REVIEW workbook (the award package's review output)
"""

import re
import threading
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Dict, List, Tuple

# Third-party imports
from openpyxl import Workbook
from openpyxl.styles import Font

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk


# ===== CONSTANTS =====
VERSION = "1.0.0"

_DEFAULT_FORECAST_FILENAME = "Working Forecast List.xlsx"

# Sheet the award review writes its per-nest, forecast-shaped rows onto. This
# is the ONE sheet we read: it is already deduplicated to one line per nest,
# and it is what the reviewer signed off on.
AWARD_SHEET = "Working Forecast Input"
AWARD_REQUIRED_HEADERS = ("Order", "Source Material", "Nest Pkg Nbr")

# Sheet inside the Working Forecast List that resolves a Source Material code
# to its material designation + size.
INVENTORY_SHEET = "Inventory Listing"
INV_KEY_HEADER = "Source Mat'l"
INV_MATL_HEADER = "Mat'l Des"
# The size string ('6 X 4 X 0.500', 'L3X3X0.500') lives under the column
# HEADED 'Thickness'. That header is a misnomer the forecast has always
# carried -- for structural shapes the cell holds the full profile, and the
# hand-built scripting sheet's DESC column is an XLOOKUP straight at it.
# Looked up by that header NAME anyway (Hard Rule 1), never by index.
INV_DESC_HEADER = "Thickness"

# ---------------------------------------------------------------------------
# Output layout: row 1 = header, row 2+ = data.
#
# The second element of each pair is the column's SOURCE, which decides what
# the app writes there:
#
#   AWARD    - computed from the award review workbook
#   FORECAST - looked up in the Working Forecast List
#   FIXED    - the same value on every award, written from the tables below
#   MANUAL   - typed by the user afterwards; header only, cells left blank
#
# These tags are internal. The hand-built reference printed them on a row above
# the headers, but those were the author's working notes, so the app does not
# write them (user, 2026-09-02). AWARD/FORECAST columns ARE written in red,
# matching the reference's colour coding.
# ---------------------------------------------------------------------------
RED = Font(color="FFFF0000")

PO_COLUMNS: List[Tuple[str, str]] = [
    ("PO NO",            "MANUAL"),
    ("LINE",             "MANUAL"),
    ("PROMISE DATE",     "MANUAL"),
    ("DYPN",             "AWARD"),
    ("QTY",              "FIXED"),
    ("UNIT PRICE",       "FIXED"),
    ("PART REV",         "MANUAL"),
    ("BATCH",            "AWARD"),
    ("NEST",             "FIXED"),
    ("ORDER",            "FIXED"),
    ("SOURCE",           "AWARD"),
    ("STANDARD CLAUSES", "MANUAL"),
    ("SHIP",             "MANUAL"),
    ("CUSTOMER",         "FIXED"),
    ("INTERCOMPANY",     "MANUAL"),
    ("SUPPLIER",         "MANUAL"),
    ("SINGLE",           "MANUAL"),
]

PART_COLUMNS: List[Tuple[str, str]] = [
    ("ORDER",             "MANUAL"),
    ("DYPN",              "AWARD"),
    ("SOURCE",            "AWARD"),
    ("MATL",              "FORECAST"),
    ("DESC",              "FORECAST"),
    ("SIZE",              "MANUAL"),
    ("WIDTH",             "FIXED"),
    ("LENGTH",            "FIXED"),
    ("WEIGHT",            "FIXED"),
    ("MIL SPEC",          "MANUAL"),
    ("NEST ",             "MANUAL"),
    ("PageNumber",        "MANUAL"),
    ("CATALOG",           "FIXED"),
    ("CATEGORY",          "FIXED"),
    ("CUSTOMER/SUPPLIER", "FIXED"),
    ("DIVISION 1",        "FIXED"),
    ("DIVISION 2",        "AWARD"),
    ("DIVISION 3 ",       "MANUAL"),
]

# Values for every FIXED column. Identical on every row of every award, so
# writing them saves 30-odd rows of fill-down; anything award-SPECIFIC stays
# MANUAL above.
#
# QTY = 1 is the PO LINE's quantity, NOT a piece count -- the reference sheet
# carries 1 on every row while PCS ranges 1-26 and ORDERS 1-20, and those two
# already have their own columns (BATCH, SOURCE). Nothing in the award review
# or the forecast is all-1s, so there is no column to source it from; it is a
# constant of the line format. Exposed as a setting anyway (`po_qty`) so an
# award that needs a different value is a settings change, not a code edit.
PO_FIXED = {
    "QTY": 1,
    "UNIT PRICE": 0,
    "NEST": ";",
    "ORDER": "SSPO ",
    "CUSTOMER": "ELECTRIC BOAT",
}
PART_FIXED = {
    "WIDTH": 0,
    "LENGTH": 0,
    "WEIGHT": 0,
    "CATALOG": "ELECTRIC BOAT",
    "CATEGORY": "911 PRODUCTION",
    "CUSTOMER/SUPPLIER": "ELECTRIC BOAT",
    "DIVISION 1": "AUBURN",
}

# Optional settings that FILL an otherwise-MANUAL column when set. Blank by
# default, because each cites something award-specific (a part revision, the
# award's VIR numbers, the ship-to for this line of work).
_SETTING_FILLS = [
    ("part_rev",         "PO", "PART REV"),
    ("standard_clauses", "PO", "STANDARD CLAUSES"),
    ("ship_to",          "PO", "SHIP"),
]

# '911 SSPO AWARD REVIEW - 1000129724 SSPO Award 13 - 2026-08-28' -> the middle
_AWARD_NAME_RE = re.compile(
    r"^\s*911\s+SSPO\s+AWARD\s+REVIEW\s*-\s*(.+?)\s*-\s*\d{4}-\d{2}-\d{2}\s*$",
    re.IGNORECASE)


# ===== HELPER FUNCTIONS =====
def _forecast_dir(override: str = "") -> Path:
    """Forecast and Inventory Reports root. Override wins; otherwise
    auto-discover across every OneDrive path variant."""
    root = sdk.resolve_forecast_dir(override)
    if root is not None:
        return root
    return sdk.pilot_program_roots()[0] / "Forecast and Inventory Reports"


def _clean(value: Any) -> str:
    """Cell -> trimmed string. Numeric source codes arrive as ints/floats from
    openpyxl; a trailing '.0' would break the inventory key match."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def award_identity(award_path: Path) -> str:
    """The award identity out of the review workbook's filename, so the output
    sits beside it under a matching name. Falls back to the whole stem."""
    match = _AWARD_NAME_RE.match(award_path.stem)
    return match.group(1).strip() if match else award_path.stem


def read_award_rows(award_path: Path, cancel_event, log) -> List[Dict[str, str]]:
    """One dict per nest from the award review's 'Working Forecast Input'."""
    wb = sdk.load_workbook_resilient(award_path, log=log, data_only=True)
    try:
        if AWARD_SHEET not in wb.sheetnames:
            raise sdk.UserFacingError(
                "That workbook has no '%s' sheet." % AWARD_SHEET,
                "Pick the '911 SSPO AWARD REVIEW ...' workbook that the 911 SSPO "
                "Award Review app produced for this award package, then run again.")
        ws = wb[AWARD_SHEET]

        header_row, hdr = sdk.find_header_row(ws, AWARD_REQUIRED_HEADERS)
        if header_row is None:
            raise sdk.UserFacingError(
                "Could not find the header row on '%s'." % AWARD_SHEET,
                "It needs Order, Source Material and Nest Pkg Nbr columns. Re-run "
                "the 911 SSPO Award Review app to regenerate the workbook.")

        col = {name: sdk.header_col(hdr, name) for name in
               ("Order", "Source Material", "PCS", "Nest Pkg Nbr", "ORDERS",
                "Division")}

        def cell(row_idx: int, name: str) -> str:
            idx = col[name]
            return _clean(ws.cell(row_idx, idx).value) if idx else ""

        rows: List[Dict[str, str]] = []
        for r in range(header_row + 1, ws.max_row + 1):
            sdk.raise_if_cancelled(cancel_event)

            order = cell(r, "Order")
            nest = cell(r, "Nest Pkg Nbr")
            # Hard Rule 3: the nest regex drops footer/total/junk rows.
            if not order or not sdk.is_nest_id(nest):
                continue

            rows.append({
                "order":    order,
                "nest":     nest,
                "source":   cell(r, "Source Material"),
                "pcs":      cell(r, "PCS"),
                "orders":   cell(r, "ORDERS"),
                "division": cell(r, "Division"),
            })
        return rows
    finally:
        wb.close()


def read_inventory(forecast_copy: Path, cancel_event, log) -> Dict[str, Tuple[str, str]]:
    """{source material code: (MATL, DESC)} from the forecast's Inventory
    Listing. First entry wins -- the listing carries one row per code."""
    wb = sdk.load_workbook_resilient(forecast_copy, log=log, data_only=True)
    try:
        if INVENTORY_SHEET not in wb.sheetnames:
            raise sdk.UserFacingError(
                "The Working Forecast List has no '%s' sheet." % INVENTORY_SHEET,
                "Check the 'Forecast and Inventory Reports Directory' and "
                "'Working Forecast Filename' settings for this app.")
        ws = wb[INVENTORY_SHEET]

        header_row, hdr = sdk.find_header_row(
            ws, (INV_KEY_HEADER, INV_MATL_HEADER, INV_DESC_HEADER), prefix_ok=True)
        if header_row is None:
            raise sdk.UserFacingError(
                "Could not find the header row on '%s'." % INVENTORY_SHEET,
                "It needs %s, %s and %s columns."
                % (INV_KEY_HEADER, INV_MATL_HEADER, INV_DESC_HEADER))

        wanted = (INV_KEY_HEADER, INV_MATL_HEADER, INV_DESC_HEADER)
        found = {name: sdk.header_col(hdr, name) for name in wanted}
        missing = [name for name in wanted if found[name] is None]
        if missing:
            raise sdk.UserFacingError(
                "The Inventory Listing is missing the %s column(s)."
                % ", ".join("'%s'" % name for name in missing),
                "Those columns are where the material designation and size come "
                "from. Check the Working Forecast List, then run again.")
        key_col: int = found[INV_KEY_HEADER]      # type: ignore[assignment]
        matl_col: int = found[INV_MATL_HEADER]    # type: ignore[assignment]
        desc_col: int = found[INV_DESC_HEADER]    # type: ignore[assignment]

        lookup: Dict[str, Tuple[str, str]] = {}
        for r in range(header_row + 1, ws.max_row + 1):
            if r % 64 == 0:
                sdk.raise_if_cancelled(cancel_event)
            key = _clean(ws.cell(r, key_col).value)
            if not key or key in lookup:
                continue
            lookup[key] = (_clean(ws.cell(r, matl_col).value),
                           _clean(ws.cell(r, desc_col).value))
        return lookup
    finally:
        wb.close()


def _number(raw, default):
    """A numeric setting, falling back to `default` on blank/garbage. Written
    as an int when it is one, so QTY lands as 1 rather than 1.0."""
    text = str(raw if raw is not None else "").strip()
    if not text:
        return default
    try:
        value = float(text)
    except ValueError:
        return default
    return int(value) if value.is_integer() else value


def build_records(award_rows: List[Dict[str, str]],
                  inventory: Dict[str, Tuple[str, str]],
                  fills: Dict[str, str],
                  qty=PO_FIXED["QTY"],
                  unit_price=PO_FIXED["UNIT PRICE"]) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """(PO Data rows, Part Data rows, unresolved-source rows).

    A source code the inventory doesn't carry leaves MATL/DESC blank and lands
    in the validation report -- never guessed at.
    """
    po_rows: List[Dict] = []
    part_rows: List[Dict] = []
    unresolved: List[Dict] = []

    for row in award_rows:
        dypn = ("%s %s" % (row["order"], row["nest"])).strip()
        matl, desc = inventory.get(row["source"], ("", ""))

        if row["source"] and not matl and not desc:
            unresolved.append({"dypn": dypn, "source": row["source"],
                               "reason": "Source Material not in Inventory Listing"})
        elif not desc:
            unresolved.append({"dypn": dypn, "source": row["source"],
                               "reason": "No %s value on the Inventory Listing row"
                                         % INV_DESC_HEADER})

        po = dict(PO_FIXED)
        po.update({
            "QTY":        qty,
            "UNIT PRICE": unit_price,
            "DYPN":   dypn,
            "BATCH":  "%s PCS" % row["pcs"] if row["pcs"] else "",
            "SOURCE": "%s ORDERS" % row["orders"] if row["orders"] else "",
        })
        po_rows.append(po)

        part = dict(PART_FIXED)
        part.update({
            "DYPN":       dypn,
            "SOURCE":     row["source"],
            "MATL":       matl,
            "DESC":       desc,
            "DIVISION 2": row["division"],
        })
        part_rows.append(part)

    for key, sheet, header in _SETTING_FILLS:
        value = fills.get(key, "")
        if not value:
            continue
        for record in (po_rows if sheet == "PO" else part_rows):
            record[header] = value

    return po_rows, part_rows, unresolved


def _write_sheet(ws, columns: List[Tuple[str, str]], records: List[Dict]) -> None:
    """Row 1 headers, row 2+ data.

    The hand-built reference carried a source-tag row above the headers
    (AWARD / FORECAST / FIXED); those were the author's working notes, not part
    of the format, so they are not written (user, 2026-09-02). The tags still
    drive the code -- which columns get filled and which are left for a human --
    they just don't ship in the sheet. AWARD/FORECAST cells stay red so a
    reviewer can still see at a glance which values the app produced.
    """
    for idx, (header, tag) in enumerate(columns, start=1):
        head_cell = ws.cell(row=1, column=idx, value=header)
        if tag in ("AWARD", "FORECAST"):
            head_cell.font = RED

    for offset, record in enumerate(records):
        for idx, (header, tag) in enumerate(columns, start=1):
            if header not in record:
                continue
            cell = ws.cell(row=2 + offset, column=idx, value=record[header])
            if tag in ("AWARD", "FORECAST"):
                cell.font = RED

    for idx, (header, _tag) in enumerate(columns, start=1):
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = max(12, min(len(header) + 4, 40))


def _write_unresolved_sheet(wb, unresolved: List[Dict], total: int) -> None:
    """Third sheet: one row per nest whose Source Material didn't resolve.

    It rides in the SAME workbook rather than a sibling file so the exception
    list can't get separated from the data it belongs to -- and so it can be
    deleted once the values are filled in, which a second file can't be.
    Only created when there is something to report.
    """
    ws = wb.create_sheet("Unresolved")
    ws["A1"] = "SOURCE MATERIALS NOT RESOLVED IN THE WORKING FORECAST LIST"
    ws["A1"].font = Font(bold=True)
    ws.append([])
    ws.append(["DYPN", "Source Material", "Reason"])
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for item in unresolved:
        ws.append([item["dypn"], item["source"], item["reason"]])
    ws.append([])
    ws.append(["%d of %d nests need MATL / DESC filled in by hand."
               % (len(unresolved), total)])
    ws.append(["Delete this sheet once they are filled in."])
    for letter, width in (("A", 18), ("B", 20), ("C", 52)):
        ws.column_dimensions[letter].width = width


def write_output(output_path: Path, po_rows: List[Dict], part_rows: List[Dict],
                 unresolved: List[Dict], total: int, log) -> None:
    """The scripting workbook: PO Data, Part Data, and an Unresolved sheet
    when anything needs hand entry."""
    wb = Workbook()
    ws_po = wb.active or wb.create_sheet()
    ws_po.title = "PO Data"
    _write_sheet(ws_po, PO_COLUMNS, po_rows)

    ws_part = wb.create_sheet("Part Data")
    _write_sheet(ws_part, PART_COLUMNS, part_rows)

    if unresolved:
        _write_unresolved_sheet(wb, unresolved, total)

    sdk.save_workbook(wb, output_path, log)


# ===== MAIN PLUGIN FUNCTION =====
def run(params: Dict[str, Any], progress_callback, cancel_event: threading.Event) -> None:
    """Main plugin execution."""
    log = params.get('log', print)
    settings = params.get('settings', {}) or {}

    log("Starting 911 Scripting Prep v%s..." % VERSION)
    progress_callback(0)

    # === PROMPT: THE AWARD REVIEW WORKBOOK ===
    log("Input required from user...")
    picked = sdk.request_file(
        params,
        "Select the 911 SSPO AWARD REVIEW workbook",
        name_filter="Excel workbooks (*.xlsx)")
    if not picked or cancel_event.is_set():
        log("Cancelled - no award review workbook picked.")
        return

    award_path = Path(picked).expanduser()
    if not sdk.exists(award_path):
        raise sdk.UserFacingError(
            "That file doesn't exist: %s" % award_path,
            "Pick the award review workbook again, then run again.")

    log("Award review : %s" % award_path.name)
    progress_callback(10)

    # === READ THE AWARD REVIEW ===
    log("Reading the award review's Working Forecast Input...")
    award_rows = read_award_rows(award_path, cancel_event, log)
    if not award_rows:
        raise sdk.UserFacingError(
            "No nest rows were found on that award review's '%s' sheet."
            % AWARD_SHEET,
            "Re-run the 911 SSPO Award Review app on the award package, then "
            "run this app on the workbook it writes.")
    log("   %d nest(s)" % len(award_rows))
    progress_callback(30)

    # === STAGE A COPY OF THE LIVE FORECAST, THEN READ IT ===
    # The Working Forecast List is a shared, live workbook. Copy it to a temp
    # folder and read the COPY so a run can never write to, lock, or otherwise
    # disturb the file everyone else is in (the 911 Setup pattern).
    log("Copying the Working Forecast List...")
    forecast_filename = (settings.get("forecast_filename")
                         or _DEFAULT_FORECAST_FILENAME)
    forecast_src = _forecast_dir(settings.get("forecast_dir", "")) / forecast_filename
    if not sdk.exists(forecast_src):
        raise sdk.UserFacingError(
            "Working Forecast List not found at:\n%s" % forecast_src,
            "Check this app's 'Forecast and Inventory Reports Directory' and "
            "'Working Forecast Filename' settings, and that the folder is synced.")

    with TemporaryDirectory(prefix="sspo_scripting_") as tmp_dir:
        forecast_copy = Path(tmp_dir) / forecast_filename
        sdk.copy_resilient(forecast_src, forecast_copy, log)  # hydrate + locked msg
        log("   Read-only copy staged (the live file is never opened)")
        progress_callback(45)

        log("Reading the Inventory Listing...")
        inventory = read_inventory(forecast_copy, cancel_event, log)
        log("   %d source material code(s)" % len(inventory))
    progress_callback(60)

    # === BUILD THE ROWS ===
    log("Building the scripting rows...")
    fills = {key: str(settings.get(key, "") or "").strip()
             for key, _sheet, _header in _SETTING_FILLS}
    po_rows, part_rows, unresolved = build_records(
        award_rows, inventory, fills,
        qty=_number(settings.get("po_qty"), PO_FIXED["QTY"]),
        unit_price=_number(settings.get("po_unit_price"), PO_FIXED["UNIT PRICE"]))
    progress_callback(75)

    # === WRITE ===
    identity = award_identity(award_path)
    stamp = datetime.now().strftime("%Y-%m-%d")
    output_name = "SSPO SCRIPTING - %s - %s.xlsx" % (identity, stamp)

    output_dir = str(settings.get("output_dir", "") or "").strip()
    output_folder = Path(output_dir) if output_dir else award_path.parent
    sdk.ensure_dir(output_folder)
    output_path = output_folder / output_name

    log("Writing the scripting workbook...")
    write_output(output_path, po_rows, part_rows, unresolved, len(award_rows), log)
    log("   %s" % output_path.name)
    progress_callback(90)

    if unresolved:
        log("")
        log("WARNING: %d nest(s) had no MATL / DESC in the forecast"
            % len(unresolved))
        log("   Listed on the 'Unresolved' sheet - delete it once they're filled in.")

    console = params.get("console")
    if console is not None and hasattr(console, "append_link"):
        try:
            console.append_link(output_path.name, str(output_path),
                                prefix="OUTPUT", at_run_end=True)
        except TypeError:
            # Older console signature without prefix/at_run_end.
            console.append_link(output_path.name, str(output_path))

    # === SUMMARY ===
    resolved = len(award_rows) - len(unresolved)
    log("")
    log("=" * 50)
    log("SCRIPTING PREP SUMMARY")
    log("=" * 50)
    log("Nests           : %d" % len(award_rows))
    log("MATL/DESC found : %d" % resolved)
    log("Needs hand entry: %d" % len(unresolved))
    log("Output folder   : %s" % output_folder)
    log("   - %s" % output_path.name)
    log("Sheets          : PO Data, Part Data%s"
        % (", Unresolved" if unresolved else ""))
    log("")
    log("Blank by design (manual entry): PO NO, LINE, PROMISE DATE, PART REV,")
    log("STANDARD CLAUSES, SHIP, and the other columns shown in black.")
    log("=" * 50)
    log("911 Scripting Prep completed successfully!")

    progress_callback(100)


if __name__ == "__main__":
    cancel = threading.Event()

    def progress(pct):
        print("Progress: %d%%" % pct)

    try:
        run(params={'console': None}, progress_callback=progress, cancel_event=cancel)
        print("\nDone!")
    except Exception as exc:
        print("\nERROR: Failed: %s" % exc)
