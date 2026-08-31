"""
922 FormingFinder
=================
Discovers formed plate AND formed flat bar PDFs in a 922 QTDR batch using
three methods:
  1. Filename scan for "PLT F" / "BAR F" (DriveWorks appends " F" to a formed
     plate's filename, and - since flat bars are formed in-house - to a flat
     bar's when its Material State is "Formed")
  2. PO workbook NOTES column containing "bend"
  3. PDF content / spatial analysis for unlabeled formed parts

Copies all discovered PDFs into Batch {n} - Documentation/Forming {n}/, merges
them into Forming {n}.pdf, and writes one row per part to the Bent Plates
sheet of the Pallet & Rod Organizer workbook (sorted by SOURCE MATERIAL).

Everything here is keyed by (ORDER, DYPN), never DYPN alone: two orders in one
batch routinely carry the SAME formed part, and each order needs its own binder
copy and its own Bent Plates row (Batch 490: BL427447 and BM352088 both build
R7211262-H2-3; DYPN-only keying dropped the second one, and only 922 Kitting -
which walks per order - noticed).
"""
from __future__ import annotations

import os
import re
import shutil
import threading
from pathlib import Path
from typing import Optional

import fitz
import openpyxl
from openpyxl.styles import Font

try:
    from techdeck.core import plugin_sdk as sdk
except ModuleNotFoundError:
    import sys, pathlib
    sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
    from techdeck.core import plugin_sdk as sdk


# Regex patterns ---------------------------------------------------------------

_UP_DEG_R_RE = re.compile(r'UP\s*\d+(?:\.\d+)?\s*°\s*R', re.IGNORECASE)
_DEG_WORD_RE = re.compile(r'\d+\s*°')
_NUMERIC_WORD_RE = re.compile(r'^\d+\.?\d*$')

PURPLE_HEX = "FF7030A0"


# DYPN helpers -----------------------------------------------------------------

def _dypn_from_filename(filename: str) -> str:
    """Extract DYPN from filename: the portion before ' PLT' or ' BAR'."""
    for token in (" PLT", " BAR"):
        idx = filename.find(token)
        if idx != -1:
            return filename[:idx].strip()
    return Path(filename).stem


def _has_formed_suffix(name: str) -> bool:
    """DriveWorks marks a formed part's filename with ' F' after the type
    token: '{DYPN} PLT F' (formed mounting plate project) and '{DYPN} BAR F'
    (FLAT BAR project with Material State = Formed). Case-sensitive on
    purpose, same as the original PLT-only scan."""
    return "PLT F" in name or "BAR F" in name


def _is_part_pdf(name: str) -> bool:
    """A plate or flat-bar part PDF. ' BAR' keeps its leading space so a DYPN
    that merely contains the letters can't misfire ('PLT' predates this and
    stays as-is for back-compat)."""
    return "PLT" in name or " BAR" in name


def _bent_suffix(dypn: str) -> str:
    """E.g. 'E6492162-H22-3' -> 'H22-3 BEND' (suffix is everything from first -H)."""
    idx = dypn.find("-H")
    if idx == -1:
        return f"{dypn} BEND"
    return f"{dypn[idx + 1:]} BEND"


def _is_skipped(pdf: Path, batch_path: Path, batch_no: str) -> bool:
    try:
        rel = pdf.relative_to(batch_path)
    except ValueError:
        return False
    parts_lower = [seg.casefold() for seg in rel.parts]
    if "repeat batches" in parts_lower:
        return True
    doc_name = f"batch {batch_no} - documentation".casefold()
    if doc_name in parts_lower:
        return True
    return False


# ORDER helpers ----------------------------------------------------------------

def _order_from_folder_name(folder: str) -> str:
    """Order number from a batch order folder name - the DYPN always follows
    the first hyphen ('BL427447-R7211262-H2' -> 'BL427447')."""
    return folder.split("-", 1)[0].strip()


def _order_for_pdf(pdf: Path, batch_path: Path, known_orders: set[str]) -> str:
    """The ORDER a discovered PDF belongs to = its top-level batch folder.
    Prefer a name the PO actually knows (whole folder, then the leading token);
    fall back to the leading token so a missing PO still separates orders.
    '' when the PDF isn't under an order folder at all."""
    try:
        rel = pdf.relative_to(batch_path)
    except ValueError:
        return ""
    if len(rel.parts) < 2:
        return ""
    folder = rel.parts[0]
    guess = _order_from_folder_name(folder)
    if known_orders:
        if folder.casefold() in known_orders:
            return folder
        if guess.casefold() in known_orders:
            return guess
    return guess


def _unique_dest(forming_dir: Path, name: str, used: set[str]) -> Path:
    """Destination for a binder copy, de-duplicated by name. Two orders can
    carry the same formed part, so their drawings share a filename - both
    belong in the binder, so the second becomes '<name> (2).pdf' (the Explorer
    convention) and both get merged into Forming {n}.pdf."""
    stem, suffix = Path(name).stem, Path(name).suffix
    candidate, i = name, 1
    while candidate.casefold() in used:
        i += 1
        candidate = f"{stem} ({i}){suffix}"
    used.add(candidate.casefold())
    return forming_dir / candidate


# PO workbook ------------------------------------------------------------------

def _find_po_workbook(doc_folder: Path, batch_no: str) -> Optional[Path]:
    matches = [
        p for p in doc_folder.glob(f"PO H{batch_no} QF-QU-09 REV C*.xlsx")
        if not p.name.startswith("~")
    ]
    return sorted(matches)[0] if matches else None


def _find_organizer_workbook(doc_folder: Path, batch_no: str) -> Optional[Path]:
    matches = [
        p for p in doc_folder.glob(f"PO H{batch_no} Pallet & Rod Organizer*.xlsx")
        if not p.name.startswith("~")
    ]
    return sorted(matches)[0] if matches else None


# Source-material classification off the PO's SOURCE MATERIAL sheet
# descriptions. Real Batch 485 text: plates read 'OSS 0.625 THK', rods read
# 'OSS 0.625 OD ASTM A36'. \bOD\b (word-bounded) so 'ROD'/'BODY' can't misfire.
# Flat bars (formable in-house since 2026-08) are matched on 'FLAT BAR' only -
# a bare \bBAR\b would also hit 'BAR STOCK'-style rod text.
_THK_RE = re.compile(r"\bTHK\b|\bPLATE\b", re.IGNORECASE)
_OD_RE = re.compile(r"\bOD\b|\bROD\b", re.IGNORECASE)
_BAR_RE = re.compile(r"\bFLAT\s*BAR\b", re.IGNORECASE)

# The kinds that can be a FORMED part (a rod never is).
_FORMABLE_KINDS = ("plate", "bar")


def _material_kind(desc: str) -> str:
    """'plate' | 'bar' | 'rod' | 'unknown' from a source-material description.
    Conflicting formable-vs-rod evidence (one 485 row says THK in Scribe and
    OD in Part Description) is 'unknown' — never guess a formed part onto a
    rod. THK + FLAT BAR together is fine (both say formable) and reads as
    'plate'."""
    plate = bool(_THK_RE.search(desc))
    bar = bool(_BAR_RE.search(desc))
    rod = bool(_OD_RE.search(desc))
    if (plate or bar) and not rod:
        return "plate" if plate else "bar"
    if rod and not (plate or bar):
        return "rod"
    return "unknown"


def _load_material_descs(wb, log) -> dict[str, str]:
    """{serial_casefold: 'scribe desc / part desc'} from the SOURCE MATERIAL
    sheet. Empty (with a log line) when the sheet or headers are missing."""
    smap = {s.strip().casefold(): s for s in wb.sheetnames}
    name = smap.get('source material')
    if not name:
        log("PO workbook: no SOURCE MATERIAL sheet - source-material "
            "descriptions unavailable for plate/rod disambiguation.")
        return {}
    ws = wb[name]
    hdr, cols = None, {}
    for r in range(1, min(ws.max_row, 10) + 1):
        row_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                row_map[v.strip().upper()] = c
        if 'SOURCE MATERIAL' in row_map:
            hdr, cols = r, row_map
            break
    if hdr is None:
        log("PO workbook: SOURCE MATERIAL sheet has no header row.")
        return {}
    ser_col = cols['SOURCE MATERIAL']
    desc_cols = [cols[k] for k in ('SCRIBE DESCRIPTION', 'PART DESCRIPTION')
                 if k in cols]
    out: dict[str, str] = {}
    for r in range(hdr + 1, ws.max_row + 1):
        ser = ws.cell(row=r, column=ser_col).value
        if ser in (None, ''):
            continue
        text = " / ".join(str(ws.cell(row=r, column=c).value).strip()
                          for c in desc_cols
                          if ws.cell(row=r, column=c).value not in (None, ''))
        out.setdefault(str(ser).strip().casefold(), text)
    return out


def _load_po_lookup(po_path: Path, log) -> dict:
    """Return {(order_casefold, dypn_casefold): row metadata} for Bent Plates.

    Keyed by ORDER *and* DYPN because one batch can run the same part for two
    orders (Batch 490: BL427447 + BM352088 both build R7211262-H2-3) and each
    needs its own row. Within one order:

    A DYPN can appear on SEVERAL PO rows with DIFFERENT source materials — the
    dual-material parts carry a formed PLATE and a ROD under one part number
    (Batch 485: H7441366-H417-3 = 262028653-50 plate + -48 rod). Row choice,
    in order:
      1. a row whose NOTES say 'bend' (the PO told us outright);
      2. the row whose source-material description reads as FORMABLE (a plate
         via THK, or a flat bar via FLAT BAR) and not a ROD (OD) - a formed
         part is never a rod, so when the PO is silent the material type
         decides. First-row-wins here is what put FORMED on the rods of
         Kitting 485 pages 30/35 (2026-08-05);
      3. anything not rod-like; else the first row, each with a warning.
    """
    # Resilient load: on a fresh Files-On-Demand sync the PO workbook can be a
    # cloud-only placeholder — a plain load_workbook dies in zipfile with
    # OSError [Errno 22] even though the file exists (bit NRAPINI-LT, Batch 481).
    wb = sdk.load_workbook_resilient(po_path, log=log, data_only=True)
    if 'PO' not in wb.sheetnames:
        log("PO workbook: 'PO' sheet not found.")
        wb.close()
        return {}
    ws = wb['PO']

    # Locate header row by scanning for cells named ORDER + DYPN
    header_row = None
    headers: dict[str, int] = {}
    for r in range(1, min(ws.max_row, 30) + 1):
        row_map: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                row_map[v.strip().upper()] = c
        if 'DYPN' in row_map and 'ORDER' in row_map:
            header_row = r
            headers = row_map
            break

    if header_row is None:
        log("PO workbook: could not locate header row.")
        wb.close()
        return {}

    required = ['ORDER', 'PPN', 'DYPN', 'NOTES', 'SOURCE MATERIAL']
    missing = [h for h in required if h not in headers]
    if missing:
        log(f"PO workbook: missing header columns: {missing}")

    def cell_val(r: int, name: str):
        col = headers.get(name)
        if col is None:
            return None
        return ws.cell(row=r, column=col).value

    descs = _load_material_descs(wb, log)

    # Collect EVERY row per (ORDER, DYPN) - first-wins loses the plate/rod
    # distinction, and dropping ORDER loses the second order entirely.
    rows_by_key: dict[tuple[str, str], list[dict]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        dypn = cell_val(r, 'DYPN')
        if not dypn:
            continue
        order = cell_val(r, 'ORDER')
        key = (str(order).strip().casefold() if order not in (None, '') else '',
               str(dypn).strip().casefold())
        rows_by_key.setdefault(key, []).append({
            'ORDER': order,
            'PPN': cell_val(r, 'PPN'),
            'DYPN': str(dypn).strip(),
            'NOTES': cell_val(r, 'NOTES'),
            'SOURCE MATERIAL': cell_val(r, 'SOURCE MATERIAL'),
        })
    wb.close()

    def _kind(row: dict) -> str:
        sm = row.get('SOURCE MATERIAL')
        if sm in (None, ''):
            return "unknown"
        return _material_kind(descs.get(str(sm).strip().casefold(), ""))

    lookup: dict[tuple[str, str], dict] = {}
    for key, rows in rows_by_key.items():
        bend = [r for r in rows if r.get('NOTES')
                and 'bend' in str(r['NOTES']).lower()]
        if bend:
            lookup[key] = bend[0]
            continue
        if len(rows) == 1:
            lookup[key] = rows[0]
            continue
        # Silent PO, several materials: the formed part is the formable one
        # (plate or flat bar), never the rod. The choice note is STASHED on
        # the row, not logged here - the lookup covers every PO DYPN, but only
        # the few actually discovered as formed matter, and the caller logs
        # the note when it uses the row.
        formable = [r for r in rows if _kind(r) in _FORMABLE_KINDS]
        if len(formable) == 1:
            chosen = formable[0]
            word = "plate" if _kind(chosen) == "plate" else "flat bar"
            others = ", ".join(str(r['SOURCE MATERIAL']) for r in rows
                               if r is not chosen)
            desc = descs.get(str(chosen['SOURCE MATERIAL']).strip().casefold(), '?')
            chosen['_PICK_NOTE'] = (
                f"{chosen['DYPN']}: {len(rows)} source materials on the PO; "
                f"using {word} {chosen['SOURCE MATERIAL']} ({desc}) over {others}")
        else:
            non_rod = [r for r in rows if _kind(r) != "rod"]
            chosen = (formable or non_rod or rows)[0]
            chosen['_PICK_WARN'] = (
                f"WARNING: {chosen['DYPN']} has {len(rows)} source materials "
                f"on the PO and none is unambiguously the plate or flat bar - "
                f"using {chosen['SOURCE MATERIAL']}; verify the FORMED tag on "
                f"this part's kit page.")
        lookup[key] = chosen

    return lookup


# Method 1 ---------------------------------------------------------------------

def _method1(batch_path: Path, batch_no: str, known_orders: set[str],
             log=lambda *_: None, cancel_event=None) -> dict[tuple[str, str], dict]:
    """Recursively scan for PDFs with 'PLT F' or 'BAR F' in the filename
    (case-sensitive). Keyed by (ORDER, DYPN) so the same part in two orders
    yields two entries."""
    found: dict[tuple[str, str], dict] = {}
    # Poll cancel_event during the rglob walk — over a OneDrive batch tree this
    # single call can run for a long time, and without the check a Cancel click
    # can't interrupt it (the worker thread is stuck in this loop). Emit a
    # heartbeat every ~500 files so the walk doesn't look frozen (a silent
    # full-tree rglob is what made LST Organizer look hung).
    for i, pdf in enumerate(batch_path.rglob("*.pdf")):
        if cancel_event is not None and i % 64 == 0 and cancel_event.is_set():
            break
        if i and i % 500 == 0:
            log(f"  Method 1: scanned {i} PDFs so far...")
        if not _has_formed_suffix(pdf.name):
            continue
        if _is_skipped(pdf, batch_path, batch_no):
            continue
        dypn = _dypn_from_filename(pdf.name)
        if not dypn:
            continue
        order = _order_for_pdf(pdf, batch_path, known_orders)
        key = (order.casefold(), dypn.casefold())
        if key not in found:
            found[key] = {'order': order, 'dypn': dypn,
                          'pdf_path': pdf, 'methods': {1}}
    return found


# Method 2 ---------------------------------------------------------------------

def _resolve_method2_pdf(batch_path: Path, order: str, dypn: str, cancel_event=None) -> Optional[Path]:
    """Find a PDF for DYPN under {batch_root}/{order folder}/CAD-AND-SHOP-PRINTS/.
    Plates and flat bars both qualify; prefer the ' F'-suffixed (formed) file
    over the flat one if both exist."""
    order_folder = None
    for child in batch_path.iterdir():
        if not sdk.is_dir(child):
            continue
        if child.name == order or child.name.startswith(f"{order}-"):
            order_folder = child
            break
    if not order_folder:
        return None

    # CAD-AND-SHOP-PRINTS case-insensitive
    cad_root = None
    for child in order_folder.iterdir():
        if sdk.is_dir(child) and child.name.lower() == "cad-and-shop-prints":
            cad_root = child
            break
    if not cad_root:
        return None

    formed_match: Optional[Path] = None
    flat_match: Optional[Path] = None
    needle = f"{dypn} "
    for i, pdf in enumerate(cad_root.rglob("*.pdf")):
        if cancel_event is not None and i % 64 == 0 and cancel_event.is_set():
            break
        name = pdf.name
        if needle not in name or not _is_part_pdf(name):
            continue
        if _has_formed_suffix(name):
            if formed_match is None:
                formed_match = pdf
        else:
            if flat_match is None:
                flat_match = pdf
    return formed_match or flat_match


def _method2(batch_path: Path, po_lookup: dict, log, cancel_event=None) -> dict[tuple[str, str], dict]:
    """Find DYPNs where PO NOTES contain 'bend'; resolve each to a PDF on disk.
    po_lookup is already keyed by (ORDER, DYPN), so a part shared by two orders
    resolves once per order."""
    found: dict[tuple[str, str], dict] = {}
    for key, meta in po_lookup.items():
        if cancel_event is not None and cancel_event.is_set():
            break
        notes = meta.get('NOTES')
        if not notes or 'bend' not in str(notes).lower():
            continue
        order = meta.get('ORDER')
        dypn = meta['DYPN']
        if not order:
            log(f"  Method 2: row for {dypn} has no ORDER value; cannot resolve PDF")
            continue
        pdf = _resolve_method2_pdf(batch_path, str(order).strip(), dypn, cancel_event)
        if not pdf:
            log(f"  Method 2: PDF not found on disk for {dypn} (order {order})")
            continue
        found[key] = {'order': str(order).strip(), 'dypn': dypn,
                      'pdf_path': pdf, 'methods': {2}}
    return found


# Method 3 ---------------------------------------------------------------------

def _is_formed_pdf(pdf: Path) -> bool:
    """Open PDF; return True if Check A (UP X deg R) or Check B (spatial) passes."""
    try:
        sdk.ensure_local(pdf)  # OneDrive placeholder -> download first (Hard Rule 13)
        doc = fitz.open(sdk.long_path(pdf))
    except Exception:
        return False
    try:
        page = doc[0]
        text = page.get_text("text") or ""

        if _UP_DEG_R_RE.search(text):
            return True

        words = page.get_text("words") or []
        tb_top: Optional[float] = None
        for w in words:
            if 'WELDMENT' in w[4].upper() and 'FAB' in w[4].upper():
                tb_top = w[1]
                break
        if tb_top is None:
            for w in words:
                if 'WELDMENT' in w[4].upper():
                    tb_top = w[1]
                    break
        if tb_top is None:
            return False

        x_min = page.rect.width / 2.0
        x_max = page.rect.width
        y_min = tb_top - (tb_top * 0.45)
        y_max = tb_top
        x_span = x_max - x_min
        bevel_cutoff = x_min + x_span * 0.15
        gap_cutoff = page.rect.width * 0.15

        def _in_region(w):
            cx = (w[0] + w[2]) / 2.0
            cy = (w[1] + w[3]) / 2.0
            return x_min <= cx <= x_max and y_min <= cy <= y_max

        sv_words = [w for w in words if _in_region(w)]
        deg_words = [w for w in sv_words if _DEG_WORD_RE.search(w[4])]

        for dw in deg_words:
            if dw[0] < bevel_cutoff:
                continue
            dw_yc = (dw[1] + dw[3]) / 2.0
            for nw in sv_words:
                if not _NUMERIC_WORD_RE.fullmatch(nw[4]):
                    continue
                if nw[2] >= dw[0]:
                    continue
                if dw[0] - nw[2] >= gap_cutoff:
                    continue
                nw_yc = (nw[1] + nw[3]) / 2.0
                if abs(dw_yc - nw_yc) > 8:
                    continue
                return True
        return False
    finally:
        doc.close()


def _method3(
    batch_path: Path,
    batch_no: str,
    known_orders: set[str],
    already_found: set[tuple[str, str]],
    log,
    progress_callback,
    cancel_event: threading.Event,
    progress_start: int,
    progress_end: int,
) -> dict[tuple[str, str], dict]:
    found: dict[tuple[str, str], dict] = {}
    candidates: list[Path] = []
    for i, pdf in enumerate(batch_path.rglob("*.pdf")):
        if i % 64 == 0 and cancel_event.is_set():
            return found
        if not _is_part_pdf(pdf.name):
            continue
        if _has_formed_suffix(pdf.name):
            continue
        if _is_skipped(pdf, batch_path, batch_no):
            continue
        candidates.append(pdf)

    log(f"  Method 3 candidate PDFs: {len(candidates)}")
    total = max(1, len(candidates))
    for i, pdf in enumerate(candidates):
        if cancel_event.is_set():
            return found
        progress_callback(progress_start + int((i / total) * (progress_end - progress_start)))
        dypn = _dypn_from_filename(pdf.name)
        if not dypn:
            continue
        order = _order_for_pdf(pdf, batch_path, known_orders)
        key = (order.casefold(), dypn.casefold())
        if key in already_found:
            continue
        if _is_formed_pdf(pdf):
            found[key] = {'order': order, 'dypn': dypn,
                          'pdf_path': pdf, 'methods': {3}}
            log(f"  Method 3: formed part detected -- {dypn} (order {order or '?'})")
    return found


# Phase 2 ----------------------------------------------------------------------

def _update_bent_plates(
    organizer_path: Path,
    batch_no: str,
    rows_data: list[dict],
    log,
) -> None:
    wb = sdk.load_workbook_resilient(organizer_path, log=log)
    if 'Bent Plates' not in wb.sheetnames:
        wb.close()
        raise RuntimeError("'Bent Plates' sheet not found in organizer workbook")
    ws = wb['Bent Plates']

    header_row = 2
    headers: dict[str, int] = {}
    scan_cols = max(ws.max_column, 8)
    for c in range(1, scan_cols + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str):
            headers[v.strip().upper()] = c

    required = ['ORDER', 'PPN', 'DYPN', 'SOURCE MATERIAL', 'NOTES']
    missing = [h for h in required if h not in headers]
    if missing:
        wb.close()
        raise RuntimeError(f"Bent Plates header row missing columns: {missing}")

    # Clear existing data rows from row 3 down across all known header columns
    last_col = max(headers.values())
    clear_to = max(ws.max_row, 3)
    for r in range(3, clear_to + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.value = None

    purple_font = Font(color=PURPLE_HEX)
    for i, row in enumerate(rows_data):
        r = 3 + i
        ws.cell(row=r, column=headers['ORDER'], value=row.get('ORDER'))
        ws.cell(row=r, column=headers['PPN'], value=row.get('PPN'))
        ws.cell(row=r, column=headers['DYPN'], value=row.get('DYPN'))
        ws.cell(row=r, column=headers['SOURCE MATERIAL'], value=row.get('SOURCE MATERIAL'))
        notes_cell = ws.cell(row=r, column=headers['NOTES'], value=row.get('NOTES'))
        notes_cell.font = purple_font

    # Row 1 title (merged cell; write to top-left)
    ws.cell(row=1, column=1, value=f"BATCH {batch_no}: FORMED PARTS")

    sdk.save_workbook(wb, organizer_path)
    wb.close()


# Plugin entry -----------------------------------------------------------------

def run(params: dict, progress_callback, cancel_event: threading.Event) -> None:
    log = params.get('log', print)
    console = params.get('console')
    settings = params.get('settings', {}) or {}

    log("Starting 922 FormingFinder...")
    progress_callback(0)

    # Batch input = pick the 'Batch NNN' folder (Sentry Drone capable,
    # family-cache aware — a queued 922 run prompts at most once).
    picked = sdk.request_922_batch_folder(params, settings.get('base_path', ''))
    if picked is None or cancel_event.is_set():
        return  # user cancelled — the helper already flagged the run
    batch_no, batch_path = picked
    doc_folder = batch_path / f"Batch {batch_no} - Documentation"
    progress_callback(5)

    # The moment the batch is known, start hydrating everything this run will
    # read (the PO/organizer workbooks, then every PDF in the batch tree) in
    # the background — the downloads overlap the overwrite prompt below and
    # Methods 1-3's own serial walk, so Method 3's per-PDF content sniff hits
    # local files instead of stalling on OneDrive one file at a time.
    def _read_set():
        yield from doc_folder.glob("*.xlsx")
        yield from batch_path.rglob("*.pdf")
    prefetch = sdk.prefetch_paths(_read_set(), cancel_event=cancel_event)

    # Existing-forming check: prompt before doing any work.
    forming_dir = doc_folder / f"Forming {batch_no}"
    if sdk.exists(forming_dir) and any(forming_dir.iterdir()):
        prompt = "Looks like this batch already has forming. Overwrite existing? Y/N"
        if console and hasattr(console, 'request_input'):
            ans = console.request_input(prompt)
        else:
            ans = input(prompt + ": ")
        if (ans or "").strip().lower() not in {"y", "yes"}:
            prefetch.stop()
            log("Aborted -- existing forming preserved.")
            return
        log("Overwriting existing forming.")

    # PO workbook lookup (drives Method 2 and Bent Plates metadata)
    po_path = _find_po_workbook(doc_folder, batch_no)
    if po_path:
        log(f"PO workbook: {po_path.name}")
        po_lookup = _load_po_lookup(po_path, log)
        log(f"  PO (order, DYPN) pairs indexed: {len(po_lookup)}")
    else:
        log("WARNING: PO workbook (QF-QU-09) not found. Skipping Method 2 and PO metadata lookup.")
        po_lookup = {}

    # Orders the PO knows about - used to read the ORDER off a discovered PDF's
    # folder, and as a DYPN-only metadata fallback when a PDF sits outside any
    # order folder.
    known_orders = {k[0] for k in po_lookup if k[0]}
    po_by_dypn: dict[str, dict] = {}
    for (_order, dypn_key), meta in po_lookup.items():
        po_by_dypn.setdefault(dypn_key, meta)
    progress_callback(10)
    if cancel_event.is_set():
        log("Cancelled.")
        return

    # Method 1
    log("Method 1: scanning filenames for 'PLT F' / 'BAR F'...")
    m1 = _method1(batch_path, batch_no, known_orders, log, cancel_event)
    log(f"  Method 1 found: {len(m1)} DYPN(s)")
    progress_callback(25)
    if cancel_event.is_set():
        log("Cancelled.")
        return

    # Method 2
    if po_lookup:
        log("Method 2: scanning PO NOTES for 'bend'...")
        m2 = _method2(batch_path, po_lookup, log, cancel_event)
        log(f"  Method 2 found: {len(m2)} DYPN(s)")
    else:
        m2 = {}
    progress_callback(40)
    if cancel_event.is_set():
        log("Cancelled.")
        return

    # Combine Methods 1 and 2 results before running Method 3 (which skips duplicates)
    combined: dict[tuple[str, str], dict] = {}
    for src in (m1, m2):
        for key, entry in src.items():
            if key in combined:
                combined[key]['methods'].update(entry['methods'])
            else:
                combined[key] = {
                    'order': entry.get('order', ''),
                    'dypn': entry['dypn'],
                    'pdf_path': entry['pdf_path'],
                    'methods': set(entry['methods']),
                }

    # Method 3
    log("Method 3: scanning PDFs for formed-part visual signature...")
    m3 = _method3(
        batch_path, batch_no, known_orders, set(combined.keys()),
        log, progress_callback, cancel_event,
        progress_start=40, progress_end=70,
    )
    log(f"  Method 3 found: {len(m3)} additional DYPN(s)")
    for key, entry in m3.items():
        combined[key] = entry
    progress_callback(70)
    if cancel_event.is_set():
        log("Cancelled.")
        return

    if not combined:
        log("No formed parts discovered by any method. Nothing to write.")
        progress_callback(100)
        return

    dupes = len(combined) - len({k[1] for k in combined})
    log(f"Total formed parts (order + DYPN): {len(combined)}")
    if dupes:
        log(f"  ({dupes} of them are the same part run for more than one order "
            f"- each gets its own binder copy and Bent Plates row)")

    # Phase 2: copy PDFs into Forming subfolder
    sdk.ensure_dir(forming_dir)
    log(f"Output folder: {forming_dir}")

    # Sort by DYPN then ORDER so the ' (2)' suffix lands on the same order
    # every run (the lower order number keeps the plain filename).
    sorted_for_copy = sorted(
        combined.values(),
        key=lambda e: (e['dypn'].casefold(), (e.get('order') or '').casefold()))
    copied_paths: list[Path] = []
    used_names: set[str] = set()
    for f in sorted_for_copy:
        if cancel_event.is_set():
            log("Cancelled.")
            return
        src = f['pdf_path']
        dest = _unique_dest(forming_dir, src.name, used_names)
        shutil.copy2(sdk.long_path(src), sdk.long_path(dest))
        methods_str = ",".join(str(m) for m in sorted(f['methods']))
        log(f"  copied {dest.name}  (order {f.get('order') or '?'}, "
            f"methods: {methods_str})")
        copied_paths.append(dest)
    progress_callback(85)

    # Merge
    merged_path = forming_dir / f"Forming {batch_no}.pdf"
    log(f"Merging {len(copied_paths)} PDF(s) into {merged_path.name}")
    sdk.merge_pdfs(copied_paths, merged_path)
    progress_callback(90)

    # Build Bent Plates row data (sorted by SOURCE MATERIAL)
    rows_data: list[dict] = []
    for f in sorted_for_copy:
        order = f.get('order') or ''
        meta = po_lookup.get((order.casefold(), f['dypn'].casefold()))
        if meta is None:
            # PDF outside any known order folder - fall back to DYPN-only
            # metadata so the row still carries PPN / SOURCE MATERIAL.
            meta = po_by_dypn.get(f['dypn'].casefold(), {})
        # Surface the plate-vs-rod choice only for parts actually discovered
        # as formed - it decides which kit row 922 Kitting stamps FORMED.
        if meta.get('_PICK_NOTE'):
            log(f"  {meta['_PICK_NOTE']}")
        if meta.get('_PICK_WARN'):
            log(f"  {meta['_PICK_WARN']}")
        rows_data.append({
            'ORDER': meta.get('ORDER') or (order or None),
            'PPN': meta.get('PPN'),
            'DYPN': f['dypn'],
            'SOURCE MATERIAL': meta.get('SOURCE MATERIAL'),
            'NOTES': _bent_suffix(f['dypn']),
        })

    def _sort_key(row):
        sm = row.get('SOURCE MATERIAL')
        dypn = str(row.get('DYPN') or '')
        order = str(row.get('ORDER') or '')
        if sm is None or sm == '':
            return (1, '', dypn, order)
        return (0, str(sm), dypn, order)
    rows_data.sort(key=_sort_key)

    organizer_path = _find_organizer_workbook(doc_folder, batch_no)
    if not organizer_path:
        raise RuntimeError(
            f"Pallet & Rod Organizer workbook not found in {doc_folder}. "
            "Halted Phase 2 write."
        )
    log(f"Updating Bent Plates in: {organizer_path.name}")
    _update_bent_plates(organizer_path, batch_no, rows_data, log)
    progress_callback(100)

    log("=" * 50)
    log(f"922 FormingFinder -- Batch {batch_no}")
    log(f"  Total formed parts  : {len(rows_data)}")
    log("=" * 50)


# Standalone test harness ------------------------------------------------------
if __name__ == "__main__":
    cancel = threading.Event()
    run(
        params={'log': print, 'settings': {}, 'console': None},
        progress_callback=lambda p: print(f"[{p}%]"),
        cancel_event=cancel,
    )
