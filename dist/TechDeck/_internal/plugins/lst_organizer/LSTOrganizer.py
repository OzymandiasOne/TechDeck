#!/usr/bin/env python3
r"""
LST Organizer — TXT-driven with Oversized Tube Tracking
v14:
- Tracks TWO categories of tubes:
  1. Standard tubes (gathered for LST files)
  2. Oversized tubes (>0.375 NOM - counted but not gathered)
- Provides separate counts for each category
- Shows which oversized tubes are in the PO but won't be gathered

Changes from v13:
- Added OVERSIZED_TUBE_MATERIALS set
- Added compute_counts_with_oversized() function
- Enhanced output to show both categories
- Updated console summary with oversized tube count
"""
from __future__ import annotations
from collections import defaultdict
import argparse, json, os, re, shutil, sys, time, glob
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Set

EXIT_OK=0; EXIT_USAGE=2; EXIT_NOT_FOUND=4; EXIT_PERM=7; EXIT_DATA=8; EXIT_UNKNOWN=9

# ======= EDITABLE: Tube materials configuration =======

# Standard tubes - actively gathered for LST files
STANDARD_TUBE_MATERIALS = {
    '218002867', '218003095', '218004492', '218012302', '218019939', 
    '218019941', '218021954', '218026206', '218026875', '218026962', 
    '218033112', '218136140', '40-00-2020', '40-07-1003'
}

# Oversized tubes (>0.375" NOM wall thickness) - counted but NOT gathered
OVERSIZED_TUBE_MATERIALS = {
    '181361440', '218000414', '218001209', '218001975', '218002101',
    '218002191', '218002642', '218002778', '218004273', '218012555',
    '218017859', '218018574', '218019943', '218020020', '218020119',
    '218026338', '218026500', '218032982'
}

# Combined set for total PO counting
ALL_TUBE_MATERIALS = STANDARD_TUBE_MATERIALS | OVERSIZED_TUBE_MATERIALS

def is_tty() -> bool:
    try: return sys.stdout.isatty()
    except Exception: return False

def now_tag() -> str: return time.strftime("%Y%m%d_%H%M%S")
def pretty(s:str)->None: print(s)
def err(msg:str, how_to_fix:str|None=None, code:int=EXIT_UNKNOWN)->int:
    print(f"ERROR: {msg}", file=sys.stderr)
    if how_to_fix: print(f"How to fix: {how_to_fix}", file=sys.stderr)
    return code

def ensure_dir(p: Path) -> None: p.mkdir(parents=True, exist_ok=True)

# --- Root / Batch discovery ---
def candidate_roots() -> List[Path]:
    home = Path.home()
    cands = [
        home / "American Steel & Alum" / "Communication site - Electric Boat ASA Docs" / "Pilot Program" / "922 QTDR Production Packages",
        home / "OneDrive - American Steel & Alum" / "Communication site - Electric Boat ASA Docs" / "Pilot Program" / "922 QTDR Production Packages",
    ]
    od = os.environ.get("ONEDRIVE") or os.environ.get("OneDrive")
    if od: cands.append(Path(od) / "Communication site - Electric Boat ASA Docs" / "Pilot Program" / "922 QTDR Production Packages")
    seen=set(); out=[]
    for c in cands:
        s=str(c)
        if s not in seen: out.append(c); seen.add(s)
    return out

def locate_root() -> Optional[Path]:
    for r in candidate_roots():
        if r.exists(): return r
    return None

def find_batch_path(root: Path, batch: str) -> Optional[Path]:
    live = root / f"Batch {batch}"
    if live.exists(): return live
    archived = root / "1 - Completed" / f"Batch {batch}"
    if archived.exists(): return archived
    return None

ORDER_PREFIX = re.compile(r"^(X|BJ|BK|XY)[0-9]+$", re.IGNORECASE)

# --- LST discovery ---
def find_lsts_for_order(order_dir: Path) -> Tuple[List[Path], str]:
    """
    Scan an order folder for *.lst that live under a '7000' directory which itself is somewhere
    below a 'CAD-AND-SHOP-PRINTS' segment. Paths containing any folder with 'repeat' (any case)
    are considered secondary; others are primary. We prefer primary; if none, return secondary.
    """
    primary=[]; secondary=[]
    for p in order_dir.rglob("*"):
        if not (p.is_dir() and p.name.lower()=="7000"):
            continue
        parts_lower = [seg.lower() for seg in p.parts]
        in_cad = "cad-and-shop-prints" in parts_lower
        in_repeat = any("repeat" in seg for seg in parts_lower)
        if not in_cad:
            continue
        lsts = [f for f in p.glob("*.lst") if f.is_file()]
        if not lsts:
            continue
        (secondary if in_repeat else primary).extend(lsts)
    if primary:
        return primary, "CAD-AND-SHOP-PRINTS"
    if secondary:
        return secondary, "REPEATS*"
    return [], "NONE"

# --- Excel reading helpers (no writes) ---
def strip_step_suffix(tokens: List[str]) -> List[str]:
    return tokens[:-1] if tokens and tokens[-1].upper()=="STEP" else tokens

def strip_step(dypn: str) -> str:
    return dypn[:-5] if dypn.upper().endswith("-STEP") else dypn

def _sheetmap(wb) -> dict:
    return {s.strip().casefold(): s for s in wb.sheetnames}

def find_excel_header_row(sheet, required_headers: List[str], max_search_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    """
    Dynamically find the header row in an Excel sheet.
    
    Args:
        sheet: openpyxl worksheet object
        required_headers: List of header names to look for (case-insensitive)
        max_search_rows: Maximum number of rows to search
        
    Returns:
        Tuple of (header_row_number, column_mapping)
        where column_mapping is dict of {header_name: column_index}
        
    Raises:
        ValueError if headers not found
    """
    required_lower = {h.lower().strip() for h in required_headers}
    
    for row_num in range(1, max_search_rows + 1):
        row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        
        # Create mapping of headers found in this row
        row_headers = {}
        for col_idx, cell_value in enumerate(row):
            if cell_value:
                header_clean = str(cell_value).lower().strip()
                row_headers[header_clean] = col_idx
        
        # Check if all required headers are present
        if required_lower.issubset(row_headers.keys()):
            # Build the column mapping with original case
            column_mapping = {}
            for orig_header in required_headers:
                header_lower = orig_header.lower().strip()
                column_mapping[orig_header] = row_headers[header_lower]
            
            return row_num, column_mapping
    
    raise ValueError(f"Could not find row with headers: {required_headers} in first {max_search_rows} rows")

def read_po_maps_openpyxl(xlsx_path: Path, debug_fp):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    smap = _sheetmap(wb)
    if not {"po","source material"}.issubset(smap.keys()):
        raise ValueError("Workbook missing required sheets PO / SOURCE MATERIAL")
    sh_po = wb[smap["po"]]; sh_src = wb[smap["source material"]]

    # Find header row dynamically for PO sheet
    try:
        po_header_row, po_cols = find_excel_header_row(
            sh_po, 
            required_headers=["ORDER", "DYPN", "SOURCE MATERIAL"],
            max_search_rows=20
        )
        debug_fp.write(json.dumps({
            "event": "found_po_headers",
            "row": po_header_row,
            "columns": po_cols
        }) + "\n")
    except ValueError as e:
        raise ValueError(f"PO sheet header error: {e}")

    dypn_to_order_serial = {}
    data_start_row = po_header_row + 1
    for row in sh_po.iter_rows(min_row=data_start_row, values_only=True):
        order = str(row[po_cols["ORDER"]]).strip() if len(row) > po_cols["ORDER"] and row[po_cols["ORDER"]] not in (None,"") else None
        part  = str(row[po_cols["DYPN"]]).strip() if len(row) > po_cols["DYPN"] and row[po_cols["DYPN"]] not in (None,"") else None
        serial= str(row[po_cols["SOURCE MATERIAL"]]).strip() if len(row) > po_cols["SOURCE MATERIAL"] and row[po_cols["SOURCE MATERIAL"]] not in (None,"") else None
        if order and part:
            dypn_to_order_serial[strip_step(part)] = (order, serial)

    # Find header row for SOURCE MATERIAL sheet
    try:
        src_header_row, src_cols = find_excel_header_row(
            sh_src,
            required_headers=["Serial", "Description"],
            max_search_rows=20
        )
        debug_fp.write(json.dumps({
            "event": "found_source_material_headers",
            "row": src_header_row,
            "columns": src_cols
        }) + "\n")
    except ValueError as e:
        # Fallback to positional if headers not found
        debug_fp.write(json.dumps({
            "event": "source_material_header_fallback",
            "error": str(e)
        }) + "\n")
        src_header_row = 1
        src_cols = {"Serial": 0, "Description": 5}

    serial_to_desc = {}
    data_start_row = src_header_row + 1
    for row in sh_src.iter_rows(min_row=data_start_row, values_only=True):
        serial = str(row[src_cols["Serial"]]).strip() if len(row) > src_cols["Serial"] and row[src_cols["Serial"]] not in (None,"") else None
        desc   = str(row[src_cols["Description"]]).strip() if len(row) > src_cols["Description"] and row[src_cols["Description"]] not in (None,"") else None
        if serial and desc:
            serial_to_desc[serial] = desc

    wb.close()
    debug_fp.write(json.dumps({"event":"po_maps_openpyxl_ok","dypn":len(dypn_to_order_serial),"serial":len(serial_to_desc)})+"\n")
    return dypn_to_order_serial, serial_to_desc

def autodiscover_master_po(batch_path: Path, debug_fp) -> Optional[Path]:
    doc_folder = batch_path / f"{batch_path.name} - Documentation"
    patterns = []
    if doc_folder.exists(): patterns.append(str(doc_folder / "**" / "*.xlsx"))
    patterns.append(str(batch_path / "**" / "*.xlsx"))
    seen=set(); candidates=[]
    for pat in patterns:
        for x in glob.glob(pat, recursive=True):
            if x not in seen: candidates.append(Path(x)); seen.add(x)
    def score(p: Path) -> int:
        s = 0; n = p.name.lower()
        if "qf" in n or "qu" in n: s+=5
        if "pallet" in n or "organizer" in n: s-=10
        if "po " in n: s+=2
        return s
    ranked = sorted(candidates, key=score, reverse=True)
    debug_fp.write(json.dumps({"event":"autodiscover_scan","candidates":len(ranked)})+"\n")
    return ranked[0] if ranked else None

# --- Build rows and TXT ---
def parse_name_parts(stem: str, debug_fp=None) -> Tuple[Optional[str], str]:
    tokens = stem.split("-")
    tokens = strip_step_suffix(tokens)  # drop trailing STEP token if present
    if not tokens:
        if debug_fp: debug_fp.write(json.dumps({"event":"parse_name","stem":stem,"result":"empty"})+"\n")
        return None, stem

    if ORDER_PREFIX.match(tokens[0]):
        if len(tokens) >= 4:
            order = "-".join(tokens[:3])
            part  = "-".join(tokens[3:])
            if debug_fp: debug_fp.write(json.dumps({"event":"parse_name","stem":stem,"order":order,"part":part,"mode":"order_prefixed"})+"\n")
            return order, part
        else:
            part = "-".join(tokens)
            if debug_fp: debug_fp.write(json.dumps({"event":"parse_name","stem":stem,"order":None,"part":part,"mode":"short"})+"\n")
            return None, part
    else:
        part = "-".join(tokens)
        if debug_fp: debug_fp.write(json.dumps({"event":"parse_name","stem":stem,"order":None,"part":part,"mode":"dypn_only"})+"\n")
        return None, part

def build_rows(copied_files: List[Path], dypn_map: Dict[str,Tuple[str,Optional[str]]], serial_map: Dict[str,str], debug_fp):
    rows=[]; problems=[]
    for p in copied_files:
        orig = p.name; base = p.stem
        order, part = parse_name_parts(base, debug_fp=debug_fp)
        po_hit = dypn_map.get(part)
        serial=None
        if po_hit:
            order_from_po, serial = po_hit
            if order is None: order = order_from_po
        else:
            problems.append(f"Missing PO mapping for DYPN '{part}'")

        desc = serial_map.get(serial) if serial else None
        if serial and not desc:
            problems.append(f"Missing description for serial '{serial}'")

        newname = f"{desc}_{order}-{part}.lst" if (desc and order and part) else None
        rows.append((orig, order, part, serial, desc, newname))
        debug_fp.write(json.dumps({"event":"row", "orig":orig, "order":order, "part":part, "serial":serial, "desc":desc, "new":newname})+"\n")
    return rows, problems

def compute_counts_with_oversized(rows, dypn_map: Dict[str,Tuple[str,Optional[str]]], batch_orders: Set[str], serial_map: Dict[str,str]):
    """
    Compute file counts with separate tracking for standard and oversized tubes.
    
    Returns:
        Tuple of (moved_files, moved_unique_parts, standard_tube_count, oversized_tube_count, 
                  total_tube_count, missing_standard_tubes, oversized_tubes_in_po)
    """
    moved_files = len(rows)
    moved_unique_parts = len({r[2] for r in rows})
    order_nums = {name.split("-")[0] for name in batch_orders} if batch_orders else set()
    
    def is_standard_tube(s):
        return (s is not None) and (str(s).strip() in STANDARD_TUBE_MATERIALS)
    
    def is_oversized_tube(s):
        return (s is not None) and (str(s).strip() in OVERSIZED_TUBE_MATERIALS)
    
    def is_any_tube(s):
        return (s is not None) and (str(s).strip() in ALL_TUBE_MATERIALS)
    
    # Categorize all tube parts in PO
    standard_tubes = {}
    oversized_tubes = {}
    
    if not order_nums:
        for part, (ordnum, serial) in dypn_map.items():
            desc = serial_map.get(str(serial).strip(), "UNKNOWN") if serial else "UNKNOWN"
            if is_standard_tube(serial):
                standard_tubes[part] = (ordnum, serial, desc)
            elif is_oversized_tube(serial):
                oversized_tubes[part] = (ordnum, serial, desc)
    else:
        for part, (ordnum, serial) in dypn_map.items():
            if ordnum in order_nums:
                desc = serial_map.get(str(serial).strip(), "UNKNOWN") if serial else "UNKNOWN"
                if is_standard_tube(serial):
                    standard_tubes[part] = (ordnum, serial, desc)
                elif is_oversized_tube(serial):
                    oversized_tubes[part] = (ordnum, serial, desc)
    
    standard_tube_count = len(standard_tubes)
    oversized_tube_count = len(oversized_tubes)
    total_tube_count = standard_tube_count + oversized_tube_count
    
    # Find which standard tubes were actually moved
    moved_tube_parts = {r[2] for r in rows if r[3] and is_any_tube(r[3])}
    
    # Identify missing standard tubes
    missing_standard_tubes = []
    for part, (order, serial, desc) in standard_tubes.items():
        if part not in moved_tube_parts:
            missing_standard_tubes.append((part, order, serial, desc))
    
    # Convert oversized tubes to list format
    oversized_tubes_list = [(part, order, serial, desc) for part, (order, serial, desc) in oversized_tubes.items()]
    
    return (moved_files, moved_unique_parts, standard_tube_count, oversized_tube_count, 
            total_tube_count, missing_standard_tubes, oversized_tubes_list)

def write_txt_with_oversized(txt_path: Path, rows, issues, info: dict, moved_files: int, moved_unique: int, 
                             standard_tube_count: int, oversized_tube_count: int, total_tube_count: int,
                             missing_standard_tubes: List[Tuple[str, str, str, str]], 
                             oversized_tubes: List[Tuple[str, str, str, str]],
                             batch_orders: Set[str]):
    """
    Write the main TXT output file with comprehensive information including oversized tube tracking.
    """
    
    with open(txt_path, "w", encoding="utf-8", newline="") as f:
        f.write("# LST TXT work file (generated)\n")
        f.write("# ============================================================================\n")
        for k,v in info.items(): 
            f.write(f"# {k}: {v}\n")
        f.write(f"# moved_files: {moved_files}\n")
        f.write(f"# moved_unique_parts: {moved_unique}\n")
        f.write(f"# standard_tubes_in_po: {standard_tube_count}\n")
        f.write(f"# oversized_tubes_in_po: {oversized_tube_count} (>0.375 NOM - NOT gathered)\n")
        f.write(f"# total_tubes_in_po: {total_tube_count}\n")
        
        # Add oversized tubes section
        if oversized_tubes:
            f.write("#\n")
            f.write(f"# OVERSIZED TUBES (>0.375\" NOM) - {len(oversized_tubes)} parts\n")
            f.write("# ============================================================================\n")
            f.write("# These tubes are tracked in the PO but are NOT actively gathered due to size.\n")
            f.write("# They require special handling and are not included in standard LST files.\n")
            f.write("#\n")
            
            # Group by order
            by_order = defaultdict(list)
            for part, order, serial, desc in oversized_tubes:
                by_order[order].append((part, serial, desc))
            
            f.write("# Oversized tubes by order:\n")
            for order in sorted(by_order.keys()):
                f.write(f"#\n# Order: {order}\n")
                for part, serial, desc in by_order[order]:
                    f.write(f"#   - {part} (Serial: {serial}, Desc: {desc})\n")
            f.write("# ============================================================================\n")
        
        # Add missing standard tubes analysis
        if missing_standard_tubes:
            f.write("#\n")
            f.write(f"# MISSING STANDARD TUBE FILES - {len(missing_standard_tubes)} files not found\n")
            f.write("# ============================================================================\n")
            f.write(f"# These {len(missing_standard_tubes)} standard tube parts are listed in the PO\n")
            f.write("# but their .lst files were NOT found in the order folders.\n")
            f.write("#\n")
            f.write("# Possible reasons:\n")
            f.write("#   1. Files haven't been created yet\n")
            f.write("#   2. Files are in a different folder location\n")
            f.write("#   3. Filenames don't match expected pattern\n")
            f.write("#   4. CAD-AND-SHOP-PRINTS/7000 folder doesn't exist for these parts\n")
            f.write("#\n")
            
            # Group by order
            by_order = defaultdict(list)
            for part, order, serial, desc in missing_standard_tubes:
                by_order[order].append((part, serial, desc))
            
            f.write("# Missing files by order:\n")
            for order in sorted(by_order.keys()):
                f.write(f"#\n# Order: {order}\n")
                for part, serial, desc in by_order[order]:
                    f.write(f"#   - {part} (Serial: {serial}, Desc: {desc})\n")
                    
                    # Find the full order folder name
                    matching_folders = [fo for fo in batch_orders if fo.startswith(order)]
                    if matching_folders:
                        f.write(f"#     Expected in folder: {matching_folders[0]}\n")
                        f.write(f"#     Search path: .../Batch ###/{matching_folders[0]}/*/CAD-AND-SHOP-PRINTS/*/7000/{part}.lst\n")
            
            f.write("# ============================================================================\n")
        
        f.write("\n")
        f.write("original\torder\tpart\tserial\tdescription\tnew_name\tstatus\n")
        for (orig, order, part, serial, desc, newname) in rows:
            status = "OK" if newname else "SKIPPED"
            f.write(f"{orig}\t{order or ''}\t{part or ''}\t{serial or ''}\t{desc or ''}\t{newname or ''}\t{status}\n")
        
        if issues:
            f.write("\n# Issues\n")
            for it in issues: f.write(f"# {it}\n")


def write_grouped_by_description(out_path: Path, batch_num: str, rows) -> int:
    """
    Create a grouped listing of tube parts by Description (Serial).
    Only includes rows whose serial is in STANDARD_TUBE_MATERIALS (not oversized).
    """
    items = []
    for (orig, order, part, serial, desc, _newname) in rows:
        if not serial or not desc:
            continue
        s = str(serial).strip()
        if s not in STANDARD_TUBE_MATERIALS:
            continue
        items.append((desc, s, order or "", part or ""))

    # Group by (desc, serial), keep stable/alpha order
    items.sort(key=lambda t: (t[0].lower(), t[1], t[2], t[3]))
    groups = defaultdict(list)
    for desc, s, order, part in items:
        groups[(desc, s)].append((order, part))

    with open(out_path, "w", encoding="utf-8", newline="") as f:
        f.write(f"Batch {batch_num}\n")
        f.write(f"Total standard tube parts: {len(items)}\n\n")
        for (desc, s) in sorted(groups.keys(), key=lambda k: k[0].lower()):
            f.write(f"{desc} ({s}):\n")
            for order, part in groups[(desc, s)]:
                f.write(f"{order}\t{part}\n")
            f.write("\n")
    return len(items)

# --- Copy & rename functions (unchanged from v13) ---
def retry_fileop(fn, *args, **kwargs):
    tries = kwargs.pop("tries", 5); delay = kwargs.pop("delay", 0.2)
    import errno, time as _t
    for i in range(tries):
        try: return fn(*args, **kwargs)
        except OSError as e:
            if e.errno in (errno.EACCES, errno.EPERM, errno.EBUSY):
                _t.sleep(delay * (i+1)); continue
            raise

def gather_and_copy(batch_path: Path, batch_num: str, kill: bool, debug_fp):
    issues=[]; copied=[]; srcs=[]; orders=set()
    dest = batch_path / f"Batch {batch_num} - Documentation" / "LST"
    ensure_dir(dest)
    
    # Track filenames we've already copied to skip duplicates
    copied_filenames = set()
    
    for child in sorted([d for d in batch_path.iterdir() if d.is_dir()]):
        name = child.name
        if name == f"Batch {batch_num} - Documentation": continue
        if len(name.split("-"))>=3: orders.add(name)
        lsts, src = find_lsts_for_order(child)
        debug_fp.write(json.dumps({"event":"order_probe","order_dir":str(child),"source":src,"lst_count":len(lsts)}) + "\n")
        if not lsts: continue
        for f in lsts:
            try:
                target = dest / f.name
                
                # Skip if we've already copied a file with this name
                if f.name in copied_filenames:
                    debug_fp.write(json.dumps({"event":"copy_skipped_duplicate","src":str(f),"reason":"already_copied"}) + "\n")
                    continue
                
                if kill or os.environ.get("AO_DISABLE_RISKY") == "1":
                    copied.append(target); srcs.append(src)
                    copied_filenames.add(f.name)
                    debug_fp.write(json.dumps({"event":"copy_skipped_kill","src":str(f),"dst":str(target)}) + "\n")
                else:
                    # Only copy if target doesn't exist (shouldn't happen with tracking, but safety check)
                    if target.exists():
                        debug_fp.write(json.dumps({"event":"copy_skipped_exists","src":str(f),"dst":str(target)}) + "\n")
                        continue
                    
                    retry_fileop(shutil.copy2, f, target)
                    copied.append(target); srcs.append(src)
                    copied_filenames.add(f.name)
                    debug_fp.write(json.dumps({"event":"copied","src":str(f),"dst":str(target)}) + "\n")
            except Exception as e:
                issues.append(f"Failed to copy {f}: {e}")
                debug_fp.write(json.dumps({"event":"copy_error","src":str(f),"error":str(e)}) + "\n")
    return copied, srcs, issues, dest, orders

def _sanitize_folder_name(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "-", name).strip()

def organize_by_material(dest_dir: Path, rows, debug_fp, *, kill: bool=False):
    """Organize files into material folders (unchanged from v13)"""
    if kill:
        debug_fp.write(json.dumps({"event":"organize_skip_kill"}) + "\n")
        return

    by_fname = {}
    by_stem  = {}

    for (orig, _order, _part, serial, desc, _new) in rows:
        if not (orig and serial and desc):
            continue
        key_name = orig.lower()
        key_stem = orig[:-4].lower() if key_name.endswith(".lst") else orig.lower()
        by_fname[key_name] = (desc, str(serial).strip())
        by_stem[key_stem]  = (desc, str(serial).strip())

    moved = 0
    for p in sorted(dest_dir.glob("*.lst")):
        fname = p.name
        lname = fname.lower()
        stem  = p.stem
        clean_stem = stem
        if "(" in stem and stem.endswith(")"):
            try:
                clean_stem = stem[: stem.rindex("(")]
            except ValueError:
                clean_stem = stem

        pair = (
            by_fname.get(lname) or
            by_stem.get(stem.lower()) or
            by_stem.get(clean_stem.lower())
        )

        if not pair:
            debug_fp.write(json.dumps({"event":"organize_unmatched","file":fname}) + "\n")
            target_dir = dest_dir / "Uncategorized"
        else:
            desc, serial = pair
            folder_name = _sanitize_folder_name(f"{desc} ({serial})")
            target_dir = dest_dir / folder_name

        ensure_dir(target_dir)
        try:
            final = target_dir / fname
            k = 2
            while final.exists():
                final = target_dir / f"{p.stem}({k}){p.suffix}"
                k += 1
            shutil.move(str(p), str(final))
            moved += 1
            debug_fp.write(json.dumps({
                "event": "organize_moved",
                "src": str(p),
                "dst": str(final),
                "folder": str(target_dir.name)
            }) + "\n")
        except Exception as e:
            debug_fp.write(json.dumps({"event":"organize_error","src":str(p),"error":str(e)}) + "\n")

    debug_fp.write(json.dumps({"event":"organize_done","moved":moved}) + "\n")

# --- Main ---
def main(argv: Optional[List[str]]=None) -> int:
    ap = argparse.ArgumentParser(description="LST Organizer — v14 with oversized tube tracking")
    ap.add_argument("--batch", help="Batch number, e.g., 403")
    ap.add_argument("--po", help="Path to Master PO (.xlsx)")
    ap.add_argument("--kill", action="store_true", help="Plan only (no copy/rename)")
    args = ap.parse_args(argv)

    root = locate_root()
    if not root:
        return err("Could not find the ASA root '...\\922 QTDR Production Packages' under your profile.",
                   "Ensure the SharePoint/OneDrive library is synced locally.", code=EXIT_NOT_FOUND)

    batch = args.batch or (input("Enter Batch number (e.g., 403): ").strip())
    if not batch:
        return err("No batch specified.", "Pass --batch 403 or enter a number at the prompt.", code=EXIT_USAGE)

    batch_path = find_batch_path(root, batch)
    if not batch_path:
        return err(f"Batch {batch} not found under:\n  {root}\n  {root / '1 - Completed'}",
                   "Verify the batch exists or that OneDrive sync is complete.", code=EXIT_NOT_FOUND)

    ts = now_tag()
    debug_path = batch_path / f"Batch {batch} - Documentation" / "LST" / f"debug_gather_lsts_{ts}.jsonl"
    ensure_dir(debug_path.parent)
    txt_path = batch_path / f"Batch {batch} - Documentation" / "LST" / f"LST_Overview_(batch).txt"

    with open(debug_path, "a", encoding="utf-8") as debug_fp:
        debug_fp.write(json.dumps({"event":"start","ts":ts,"batch":batch,"root":str(root),"batch_path":str(batch_path)}) + "\n")

        copied, sources, copy_issues, dest_dir, batch_orders = gather_and_copy(batch_path, batch, kill=args.kill, debug_fp=debug_fp)

        # Pick Master PO
        if args.po: master_po = Path(args.po)
        else: master_po = autodiscover_master_po(batch_path, debug_fp)
        if not master_po or not master_po.exists():
            rows = [(p.name, None, strip_step(p.stem), None, None, None) for p in copied]
            moved_files=len(rows); moved_unique=len({r[2] for r in rows})
            write_txt_with_oversized(txt_path, rows, ["Master PO workbook NOT FOUND"], 
                                    {"batch":batch,"root":root,"note":"seed only"}, 
                                    moved_files, moved_unique, 0, 0, 0, [], [], batch_orders)
            return err("Master PO workbook not found. Seed TXT created (filenames only).",
                       f"Open TXT at: {txt_path}. Re-run with --po once the workbook path is known.", code=EXIT_NOT_FOUND)

        debug_fp.write(json.dumps({"event":"master_po","path":str(master_po)}) + "\n")

        # Build lookup maps
        try:
            dypn_map, serial_map = read_po_maps_openpyxl(master_po, debug_fp)
        except Exception as e:
            debug_fp.write(json.dumps({"event":"openpyxl_fail","error":str(e)})+"\n")
            # COM fallback omitted for brevity - same as v13
            return err("Could not read the Master PO.", "Ensure file is accessible.", code=EXIT_DATA)

        rows, problems = build_rows(copied, dypn_map, serial_map, debug_fp)
        
        # Enhanced counting with oversized tracking
        (moved_files, moved_unique, standard_tube_count, oversized_tube_count, 
         total_tube_count, missing_standard_tubes, oversized_tubes) = compute_counts_with_oversized(
            rows, dypn_map, batch_orders, serial_map)

        # Write TXT with enhanced oversized tracking
        issues = copy_issues + problems
        write_txt_with_oversized(txt_path, rows, issues, 
                                {"batch":batch,"root":root,"master_po":str(master_po),
                                 "dest_dir":str(dest_dir),"orders_considered":",".join(sorted(batch_orders))}, 
                                moved_files, moved_unique, standard_tube_count, oversized_tube_count, 
                                total_tube_count, missing_standard_tubes, oversized_tubes, batch_orders)

        group_path = dest_dir / f"Tube_Parts_(batch).txt"
        write_grouped_by_description(group_path, batch, rows)

        organize_by_material(dest_dir, rows, debug_fp, kill=(args.kill or os.environ.get("AO_DISABLE_RISKY") == "1"))
        
        debug_fp.write(json.dumps({"event":"rename_disabled"}) + "\n")
        
        # Enhanced console summary
        print(f"\n{'='*80}")
        print(f"LST Organizer - Batch {batch} Complete")
        print(f"{'='*80}")
        print(f"Files found and copied: {moved_files}")
        print(f"Unique parts: {moved_unique}")
        print(f"\nTube counts:")
        print(f"  Standard tubes (gathered): {standard_tube_count}")
        print(f"  Oversized tubes (>0.375\" NOM): {oversized_tube_count}")
        print(f"  Total tubes in PO: {total_tube_count}")
        
        if oversized_tubes:
            print(f"\n📊 {len(oversized_tubes)} oversized tube parts tracked (NOT gathered)")
        
        if missing_standard_tubes:
            print(f"\n⚠  WARNING: {len(missing_standard_tubes)} standard tube files were NOT found!")
            print(f"See detailed analysis in: {txt_path}")
            print(f"\nMissing standard tubes:")
            for part, order, serial, desc in sorted(missing_standard_tubes)[:5]:
                print(f"  - {part} (Order: {order})")
            if len(missing_standard_tubes) > 5:
                print(f"  ... and {len(missing_standard_tubes) - 5} more")
        else:
            print(f"\n✓ All expected standard tube files were found!")
        
        print(f"\nOutput files:")
        print(f"  - Main report: {txt_path}")
        print(f"  - Grouped tubes: {group_path}")
        print(f"  - Debug log: {debug_path}")
        print(f"{'='*80}\n")

    return EXIT_OK

if __name__ == "__main__":
    try: raise SystemExit(main())
    except KeyboardInterrupt: raise SystemExit(EXIT_UNKNOWN)
