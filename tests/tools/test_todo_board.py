"""Tests for the DevKit Dev Board (source-only dev tool).

Data-layer tests use a temp state file; widget tests use the offscreen `qapp`
and stub `load_feedback` so they never read the maintainer's real telemetry
workbook.
"""

from tools.devkit.todo_board.model import (
    BoardStore, FeedbackEntry, FeedbackLoad,
)


def _entry(idx: int, which_feature: str, source: str) -> FeedbackEntry:
    return FeedbackEntry(
        key=f"{source}{idx}", title=f"suggestion {idx}", user="tester",
        date_iso=f"2026-07-2{idx}T09:00:00", which_feature=which_feature,
        version="0.8.6.9", orig_status="Needs Review", source=source)


def _load(feedback=(), archive=()) -> FeedbackLoad:
    return FeedbackLoad(feedback=list(feedback), archive=list(archive),
                        path=None, ok=True, message="stub")


# ---- local-copy freshness (wedged-download detector) -----------------------

def test_load_is_stale_only_past_threshold():
    from datetime import datetime, timedelta
    from tools.devkit.todo_board.model import STALE_AFTER_HOURS
    fresh = FeedbackLoad(ok=True, mtime_iso=(
        datetime.now() - timedelta(hours=1)).isoformat())
    assert fresh.is_stale is False
    assert 0.5 < fresh.age_hours < 1.5

    wedged = FeedbackLoad(ok=True, mtime_iso=(
        datetime.now() - timedelta(hours=STALE_AFTER_HOURS + 1)).isoformat())
    assert wedged.is_stale is True


def test_load_without_mtime_is_never_stale():
    """Stubbed/unknown mtime must not raise a false alarm."""
    assert FeedbackLoad(ok=True).age_hours is None
    assert FeedbackLoad(ok=True).is_stale is False
    assert FeedbackLoad(ok=True, mtime_iso="not-a-date").is_stale is False


def test_freshness_prefers_data_over_file_mtime():
    """mtime lies on a synced file: OneDrive stamped a freshly downloaded copy
    2026-08-06 15:55 while its rows ran through 2026-08-07 10:14. Judging by
    mtime called it a day stale seconds after it arrived."""
    from datetime import datetime, timedelta
    now = datetime.now()
    load = FeedbackLoad(
        ok=True,
        mtime_iso=(now - timedelta(hours=20)).isoformat(),      # misleading
        data_iso=(now - timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S"))
    assert load.freshness_iso == load.data_iso
    assert 1.5 < load.age_hours < 2.5
    assert load.is_stale is False            # the data is 2h old, not 20h

    # With no data timestamps at all, mtime is still the fallback.
    only_mtime = FeedbackLoad(ok=True, mtime_iso=(now - timedelta(hours=20)).isoformat())
    assert only_mtime.freshness_iso == only_mtime.mtime_iso
    assert only_mtime.is_stale is True


def test_newest_data_timestamp_scans_usage_first(tmp_path):
    """The Usage sheet gains a row on every run by every user, so it is the
    truest freshness signal — and it must be found by header name."""
    import openpyxl
    from tools.devkit.todo_board.model import _newest_data_timestamp
    xlsx = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook()
    us = wb.active
    us.title = "Usage"
    us.append(["Timestamp", "User", "Machine"])
    us.append(["2026-08-07 10:14:33", "amy", "M1"])
    us.append(["2026-08-05 08:00:00", "bob", "M2"])
    fb = wb.create_sheet("Feedback")
    fb.append(["Timestamp", "User", "Machine", "Suggestion"])
    fb.append(["2026-08-06 09:15:40", "cd", "M3", "idea"])
    wb.save(xlsx)
    wb.close()
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        assert _newest_data_timestamp(wb) == "2026-08-07 10:14:33"
    finally:
        wb.close()


def test_load_feedback_reports_both_timestamps(tmp_path, monkeypatch):
    """Freshness follows the DATA, not the file. This workbook is written right
    now but holds only a July row, so it is correctly reported stale — the
    inverse of the bug this replaced, where a file downloaded seconds ago was
    called stale because OneDrive backdated its mtime."""
    from tools.devkit.todo_board.model import load_feedback
    xlsx = tmp_path / "TechDeck Telemetry.xlsx"
    _make_telemetry_workbook(xlsx, [
        ["2026-07-01 10:00:00", "amy", "M1", "add dark mode",
         "Shell", "0.8.6", "Needs Review"]])
    monkeypatch.setenv("TECHDECK_TELEMETRY_XLSX", str(xlsx))
    load = load_feedback()
    assert load.ok and load.mtime_iso                    # file mtime captured
    assert load.data_iso.startswith("2026-07-01")        # newest row in the data
    assert load.freshness_iso == load.data_iso           # data wins
    assert load.is_stale is True                         # July data is stale


def _stale_board(qapp, tmp_path, monkeypatch, days=6):
    from datetime import datetime, timedelta
    from tools.devkit.todo_board import model
    from tools.devkit.todo_board import board as board_mod
    monkeypatch.setattr(model, "_state_path", lambda: tmp_path / "todo_board.json")
    monkeypatch.setattr(board_mod, "load_feedback", lambda: FeedbackLoad(
        feedback=[], archive=[], path=None, ok=True, message="stub",
        mtime_iso=(datetime.now() - timedelta(days=days)).isoformat()))
    board = board_mod.TodoBoard()
    qapp.processEvents()
    return board


def test_board_status_line_warns_on_stale_local_copy(qapp, tmp_path, monkeypatch):
    """A wedged download must SAY so — an empty board silently reading a
    week-old file is the bug this guards against. The toolbar chip stays SHORT
    (prose there widens the window); the explanation lives behind the details
    button."""
    board = _stale_board(qapp, tmp_path, monkeypatch, days=6)
    assert "⚠ data 6 days old" in board._status_full
    assert "stopped syncing" not in board._status_full     # short chip only
    assert "stopped syncing" in board._status_detail       # the why, on demand
    assert board._status_ok is False              # rendered in the warning color


def test_stale_chip_is_short_and_never_widens_the_toolbar(qapp, tmp_path, monkeypatch):
    """The rendered label must stay inside the pixel budget and carry the full
    text only as a tooltip — this is what keeps the DevKit page from growing."""
    from PySide6.QtGui import QFontMetrics
    from tools.devkit.todo_board.board import STATUS_MAX_PX
    board = _stale_board(qapp, tmp_path, monkeypatch, days=6)
    shown = board._status.text()
    cap = board._status.maximumWidth()
    assert QFontMetrics(board._status.font()).horizontalAdvance(shown) <= cap
    # The cap stretches only as far as the warning needs, never to the full line.
    # (Exact pixels vary with the font, so the meaningful guard is the second
    # assertion: the cap must stay well under the untruncated status width.)
    full_px = QFontMetrics(board._status.font()).horizontalAdvance(board._status_full)
    assert STATUS_MAX_PX <= cap <= STATUS_MAX_PX + 200
    assert cap < full_px
    assert board._status.toolTip() == board._status_full


def test_stale_warning_survives_elision(qapp, tmp_path, monkeypatch):
    """Elision cuts the TAIL, so the warning must lead — otherwise the one
    message explaining an empty board is the first thing truncated away (it was,
    on the first cut of this fix: the chip rendered '0 open · 0 archi…')."""
    from datetime import datetime, timedelta
    from tools.devkit.todo_board import model
    from tools.devkit.todo_board import board as board_mod
    monkeypatch.setattr(model, "_state_path", lambda: tmp_path / "todo_board.json")
    # A deliberately long status: many counts plus a write-back note.
    monkeypatch.setattr(board_mod, "load_feedback", lambda: FeedbackLoad(
        feedback=[_entry(i, "911 Setup", "feedback") for i in range(1, 6)],
        archive=[_entry(7, "922 Kitting", "archive")], path=None, ok=True,
        message="stub",
        mtime_iso=(datetime.now() - timedelta(days=6)).isoformat()))
    board = board_mod.TodoBoard()
    qapp.processEvents()
    assert board._status_full.startswith("⚠ data 6 days old")
    assert "data 6 days old" in board._status.text()   # survived the cut


def test_age_phrase_singular_and_plural():
    from tools.devkit.todo_board.board import TodoBoard
    assert TodoBoard._age_phrase(25) == "1 day old"
    assert TodoBoard._age_phrase(6 * 24) == "6 days old"
    assert TodoBoard._age_phrase(13) == "13 hours old"
    assert TodoBoard._age_phrase(1) == "1 hour old"


def test_status_details_dialog_opens_with_full_text(qapp, tmp_path, monkeypatch):
    from PySide6.QtWidgets import QPlainTextEdit
    board = _stale_board(qapp, tmp_path, monkeypatch, days=2)
    board._show_status_details()
    qapp.processEvents()
    dlg = board._details_dialog
    assert dlg is not None
    body = dlg.findChild(QPlainTextEdit).toPlainText()
    assert "data 2 days old" in body and "stopped syncing" in body
    dlg.close()


def test_board_status_line_shows_as_of_when_fresh(qapp, tmp_path, monkeypatch):
    from datetime import datetime, timedelta
    from tools.devkit.todo_board import model
    from tools.devkit.todo_board import board as board_mod
    monkeypatch.setattr(model, "_state_path", lambda: tmp_path / "todo_board.json")
    monkeypatch.setattr(board_mod, "load_feedback", lambda: FeedbackLoad(
        feedback=[], archive=[], path=None, ok=True, message="stub",
        mtime_iso=(datetime.now() - timedelta(minutes=20)).isoformat()))
    board = board_mod.TodoBoard()
    qapp.processEvents()
    assert "as of " in board._status_full
    assert "⚠" not in board._status_full
    assert board._status_detail == ""             # nothing to explain when fresh
    assert board._status_ok is True


# ---- registry --------------------------------------------------------------

def test_todo_board_is_default_devkit_page():
    from tools.devkit.registry import DEV_TOOLS
    assert DEV_TOOLS[0].key == "todo_board"        # mounts first => default page
    assert DEV_TOOLS[0].label == "Dev Board"


# ---- data layer ------------------------------------------------------------

def test_sync_routes_feedback_to_todo_and_archive_to_done(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    added = store.sync(_load(
        feedback=[_entry(1, "911 Setup", "feedback"), _entry(2, "Other", "feedback")],
        archive=[_entry(3, "922 Kitting", "archive")]))
    assert added == 3
    assert len(store.bucket("todo")["cards"]) == 2
    assert len(store.bucket("done")["cards"]) == 1
    # feedback label seeded from Which Feature
    todo_key = store.bucket("todo")["cards"][0]
    assert "911 Setup" in store.cards[todo_key]["labels"] or \
           "Other" in store.cards[todo_key]["labels"]


def test_sync_is_idempotent(tmp_path):
    load = _load(feedback=[_entry(1, "911 Setup", "feedback")])
    store = BoardStore(path=tmp_path / "b.json")
    assert store.sync(load) == 1
    assert BoardStore(path=tmp_path / "b.json").sync(load) == 0   # reloaded, no dupes


def test_newest_feedback_lands_on_top(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "A", "feedback"), _entry(5, "B", "feedback")]))
    top = store.bucket("todo")["cards"][0]
    assert store.cards[top]["date"] == "2026-07-25T09:00:00"   # idx 5 is newest


def test_move_card_reorder_contract(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.add_manual_card("todo", "C")
    kb = store.add_manual_card("todo", "B")
    ka = store.add_manual_card("todo", "A")   # add_manual inserts at 0 => [A, B, C]
    titles = lambda: [store.cards[k]["title"] for k in store.bucket("todo")["cards"]]
    assert titles() == ["A", "B", "C"]
    store.move_card(ka, "todo", 2)            # A to end among [B, C]
    assert titles() == ["B", "C", "A"]
    store.move_card(kb, "done", 0)            # cross-bucket
    assert store.bucket_of(kb)["id"] == "done"
    assert titles() == ["C", "A"]


def test_checklist_mutations(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    k = store.add_manual_card("todo", "task")
    store.add_checklist_item(k, "one")
    store.add_checklist_item(k, "two")
    store.toggle_checklist_item(k, 0)
    items = store.cards[k]["checklist"]
    assert [i["text"] for i in items] == ["one", "two"]
    assert items[0]["done"] is True
    store.remove_checklist_item(k, 0)
    assert [i["text"] for i in store.cards[k]["checklist"]] == ["two"]


def test_label_add_remove(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    k = store.add_manual_card("todo", "task")
    store.add_label(k, "bug")
    store.add_label(k, "bug")            # dedup
    assert store.cards[k]["labels"] == ["bug"]
    store.remove_label(k, "bug")
    assert store.cards[k]["labels"] == []


def test_delete_bucket_reassigns_cards(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    k = store.add_manual_card("done", "orphan-me")
    store.remove_bucket("done")
    assert store.bucket("done") is None
    assert store.bucket_of(k) is store.buckets[0]    # fell back to first bucket


def test_reconcile_files_orphan_cards(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.cards["ghost"] = {"title": "x", "checklist": [], "labels": []}
    store._reconcile()
    assert store.bucket_of("ghost") is not None      # placed, not lost


# ---- external status sync (mark-complete-off-board -> Done) ----------------

def _same_key(key: str, status: str, source: str = "feedback") -> FeedbackEntry:
    """A feedback/archive row for an EXISTING card key, with a chosen Status."""
    return FeedbackEntry(
        key=key, title="suggestion 1", user="tester",
        date_iso="2026-07-21T09:00:00", which_feature="911 Setup",
        version="0.8.6.9", orig_status=status, source=source)


def test_external_complete_moves_todo_card_to_done(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    moved = store.apply_external_status(_load(feedback=[_same_key(key, "Complete")]))
    assert moved == 1
    assert key in store.bucket("done")["cards"]
    assert key not in store.bucket("todo")["cards"]
    # Recorded as in-sync so the write-back doesn't redundantly rewrite it.
    assert store.cards[key]["written_status"] == "Complete"
    assert (key, "Complete") not in store.pending_writebacks()


def test_external_archive_moves_existing_card_to_done(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    # Row was moved to the Archive sheet off-board.
    moved = store.apply_external_status(
        _load(archive=[_same_key(key, "Complete", source="archive")]))
    assert moved == 1
    assert key in store.bucket("done")["cards"]


def test_external_wont_do_moves_to_wont_do(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    moved = store.apply_external_status(_load(feedback=[_same_key(key, "Won't Do")]))
    assert moved == 1
    assert key in store.bucket("wont_do")["cards"]


def test_external_status_respects_terminal_placement(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    store.move_card(key, "wont_do", 0)               # user parked it terminally
    moved = store.apply_external_status(_load(feedback=[_same_key(key, "Complete")]))
    assert moved == 0                                 # not yanked out of Won't Do
    assert key in store.bucket("wont_do")["cards"]


def test_external_status_leaves_working_buckets_when_open(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    store.move_card(key, "in_progress", 0)
    moved = store.apply_external_status(_load(feedback=[_same_key(key, "Needs Review")]))
    assert moved == 0                                 # open status -> no move
    assert key in store.bucket("in_progress")["cards"]


def test_external_status_ignores_manual_cards(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    mkey = store.add_manual_card("todo", "hand-written")
    moved = store.apply_external_status(_load(feedback=[_same_key(mkey, "Complete")]))
    assert moved == 0                                 # manual cards never auto-move
    assert mkey in store.bucket("todo")["cards"]


# ---- clear completed (Done pile) -------------------------------------------

def test_clear_bucket_empties_and_remembers(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(archive=[_entry(1, "911 Setup", "archive"),
                              _entry(2, "922 Kitting", "archive")]))
    assert len(store.bucket("done")["cards"]) == 2
    removed = store.clear_bucket("done")
    assert removed == 2
    assert store.bucket("done")["cards"] == []
    assert store.cleared == {"archive1", "archive2"}


def test_cleared_archived_rows_do_not_resync(tmp_path):
    load = _load(archive=[_entry(1, "911 Setup", "archive")])
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(load)
    store.clear_bucket("done")
    assert store.sync(load) == 0                       # stays cleared, not re-added
    assert store.bucket("done")["cards"] == []


def test_cleared_state_persists_across_reload(tmp_path):
    load = _load(archive=[_entry(1, "911 Setup", "archive")])
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(load)
    store.clear_bucket("done")
    reloaded = BoardStore(path=tmp_path / "b.json")
    assert "archive1" in reloaded.cleared
    assert reloaded.sync(load) == 0                    # still suppressed after reload


def test_cleared_item_resurrects_when_reopened(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(archive=[_entry(1, "911 Setup", "archive")]))
    store.clear_bucket("done")
    # The same item comes back as an OPEN feedback row (reopened).
    reopened = _same_key("archive1", "Needs Review", source="feedback")
    added = store.sync(_load(feedback=[reopened]))
    assert added == 1
    assert "archive1" in store.bucket("todo")["cards"]
    assert "archive1" not in store.cleared


def test_clear_bucket_no_op_when_empty(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    assert store.clear_bucket("done") == 0


# ---- widgets ---------------------------------------------------------------

def test_label_color_mapping():
    from tools.devkit.todo_board.widgets import label_color
    assert label_color("911 Setup") == "#3B82F6"     # family color
    assert label_color("922 Kitting") == "#F59E0B"
    assert label_color("Other") == "#6B7280"         # neutral
    assert label_color("Custom Tag") == label_color("Custom Tag")   # stable


def test_menu_stylesheet_is_opaque_and_themed():
    """The card / bucket menus set their own QMenu style so Win11 doesn't render
    them with a translucent popup (the app-wide rule doesn't reach a menu whose
    parent carries its own stylesheet)."""
    from types import SimpleNamespace
    from tools.devkit.todo_board.widgets import menu_stylesheet
    pal = SimpleNamespace(surface="#111111", text="#EEEEEE", border_strong="#444444",
                          surface_hover="#222222", accent="#00AAFF",
                          text_secondary="#888888", border="#333333")
    qss = menu_stylesheet(pal)
    assert "QMenu {" in qss and "background-color: #111111" in qss   # opaque bg
    assert "QMenu::item:selected" in qss and "#00AAFF" in qss
    assert "QMenu::separator" in qss


def test_task_card_renders_title_labels_checklist(qapp, tmp_path):
    from techdeck.ui.theme_manager import get_theme_manager
    from tools.devkit.todo_board.widgets import TaskCard, LabelChip
    store = BoardStore(path=tmp_path / "b.json")
    k = store.add_manual_card("todo", "Fix the thing")
    store.add_label(k, "911 Setup")
    store.add_checklist_item(k, "step")
    card = TaskCard(store, k, get_theme_manager().get_current_palette())
    card.show()
    qapp.processEvents()
    assert card._title.text() == "Fix the thing"
    assert card.findChildren(LabelChip)


def test_board_builds_with_stubbed_feedback(qapp, tmp_path, monkeypatch):
    from tools.devkit.todo_board import model
    from tools.devkit.todo_board import board as board_mod
    monkeypatch.setattr(model, "_state_path", lambda: tmp_path / "todo_board.json")
    monkeypatch.setattr(board_mod, "load_feedback", lambda: _load(
        feedback=[_entry(1, "911 Setup", "feedback"), _entry(2, "Other", "feedback")],
        archive=[_entry(3, "922 Kitting", "archive")]))
    board = board_mod.TodoBoard()
    board.show()
    qapp.processEvents()
    assert list(board._columns) == ["todo", "in_progress", "done", "wont_do"]
    assert len(board.store.bucket("todo")["cards"]) == 2
    assert len(board.store.bucket("done")["cards"]) == 1

    # a drop moves the card and the board rebuilds cleanly
    key = board.store.bucket("todo")["cards"][0]
    board._on_card_dropped(key, "in_progress", 0)
    board._do_rebuild()
    qapp.processEvents()
    assert board.store.bucket_of(key)["id"] == "in_progress"

    # theme rebuild must not raise
    board.apply_theme()
    qapp.processEvents()


# ---- write-back ------------------------------------------------------------

def _make_telemetry_workbook(path, rows):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback"
    ws.append(["Timestamp", "User", "Machine", "Suggestion",
               "Which Feature", "TechDeck Version", "Status"])
    for row in rows:
        ws.append(row)
    arch = wb.create_sheet("Archive")
    arch.append(["Timestamp", "User", "Machine", "Suggestion",
                 "Which Feature", "TechDeck Version", "Status"])
    wb.save(path)


def _status_cell(path, row=2):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    try:
        return wb["Feedback"].cell(row=row, column=7).value
    finally:
        wb.close()


def test_writeback_done_then_restore(tmp_path, monkeypatch):
    from tools.devkit.todo_board.model import flush_writebacks, load_feedback
    xlsx = tmp_path / "TechDeck Telemetry.xlsx"
    _make_telemetry_workbook(xlsx, [
        ["2026-07-01 10:00:00", "amy", "M1", "add dark mode",
         "Shell", "0.8.6", "Needs Review"]])
    monkeypatch.setenv("TECHDECK_TELEMETRY_XLSX", str(xlsx))
    store = BoardStore(path=tmp_path / "b.json")
    load = load_feedback()
    assert load.ok
    store.sync(load)
    key = store.bucket("todo")["cards"][0]

    store.move_card(key, "done", 0)
    assert store.pending_writebacks() == [(key, "Complete")]
    res = flush_writebacks(store)
    assert res.ok and res.written == 1
    assert _status_cell(xlsx) == "Complete"
    assert store.cards[key]["written_status"] == "Complete"
    assert store.pending_writebacks() == []           # idempotent — no re-write

    store.move_card(key, "todo", 0)                   # drag back out
    res = flush_writebacks(store)
    assert res.ok and res.written == 1
    assert _status_cell(xlsx) == "Needs Review"
    assert store.cards[key]["written_status"] == ""


def test_writeback_manual_edit_wins_on_restore(tmp_path, monkeypatch):
    import openpyxl
    from tools.devkit.todo_board.model import flush_writebacks, load_feedback
    xlsx = tmp_path / "TechDeck Telemetry.xlsx"
    _make_telemetry_workbook(xlsx, [
        ["2026-07-01 10:00:00", "amy", "M1", "add dark mode",
         "Shell", "0.8.6", "Needs Review"]])
    monkeypatch.setenv("TECHDECK_TELEMETRY_XLSX", str(xlsx))
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(load_feedback())
    key = store.bucket("todo")["cards"][0]

    store.move_card(key, "wont_do", 0)
    assert flush_writebacks(store).written == 1
    assert _status_cell(xlsx) == "Won't Do"

    # Maintainer overrides the cell in Excel...
    wb = openpyxl.load_workbook(xlsx)
    wb["Feedback"].cell(row=2, column=7).value = "Deferred"
    wb.save(xlsx)
    wb.close()

    # ...so dragging back out must NOT stomp it.
    store.move_card(key, "todo", 0)
    res = flush_writebacks(store)
    assert res.ok and res.written == 0 and res.adopted == 1
    assert _status_cell(xlsx) == "Deferred"
    assert store.cards[key]["written_status"] == ""   # cleared; not retried
    assert store.pending_writebacks() == []


def test_writeback_adopts_row_missing_from_sheet(tmp_path, monkeypatch):
    from tools.devkit.todo_board.model import flush_writebacks
    xlsx = tmp_path / "TechDeck Telemetry.xlsx"
    _make_telemetry_workbook(xlsx, [])                # header only — row archived
    monkeypatch.setenv("TECHDECK_TELEMETRY_XLSX", str(xlsx))
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    store.move_card(key, "done", 0)
    res = flush_writebacks(store)
    assert res.ok and res.missing == 1 and res.written == 0
    assert store.cards[key]["written_status"] == "Complete"   # adopted
    assert store.pending_writebacks() == []           # never stuck pending


def test_writeback_skips_manual_and_archive_cards(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    store.add_manual_card("done", "hand-written task")
    store.sync(_load(archive=[_entry(3, "922 Kitting", "archive")]))  # lands in done
    assert store.pending_writebacks() == []


# ---- webhook triage mode ----------------------------------------------------

def test_webhook_triage_default_off_and_persists(tmp_path):
    store = BoardStore(path=tmp_path / "b.json")
    assert store.webhook_triage is False               # off until the flow branch exists
    store.set_webhook_triage(True)
    assert BoardStore(path=tmp_path / "b.json").webhook_triage is True


def test_webhook_triage_flush_sends_events_and_records(tmp_path, monkeypatch):
    from techdeck.core import usage_tracker
    from tools.devkit.todo_board.model import flush_writebacks
    sent = []
    monkeypatch.setattr(usage_tracker, "post_triage_events",
                        lambda events: sent.append(events) or True)
    store = BoardStore(path=tmp_path / "b.json")
    store.set_webhook_triage(True)
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    store.move_card(key, "done", 0)

    res = flush_writebacks(store)
    assert res.ok and res.via_webhook and res.written == 1
    ev = sent[0][0]
    # row_key = "{timestamp} {user}" joined with | — the flow's Key column
    # composes the same string from its Timestamp + User cells.
    assert ev["row_key"] == "2026-07-21 09:00:00|tester"
    assert ev["new_status"] == "Complete"
    assert store.cards[key]["written_status"] == "Complete"
    assert store.pending_writebacks() == []            # idempotent — no re-send

    store.move_card(key, "todo", 0)                    # drag back out
    res = flush_writebacks(store)
    assert res.ok and res.written == 1
    assert sent[1][0]["new_status"] == "Needs Review"
    assert store.cards[key]["written_status"] == ""


def test_webhook_triage_failure_stays_pending(tmp_path, monkeypatch):
    from techdeck.core import usage_tracker
    from tools.devkit.todo_board.model import flush_writebacks
    monkeypatch.setattr(usage_tracker, "post_triage_events", lambda events: False)
    store = BoardStore(path=tmp_path / "b.json")
    store.set_webhook_triage(True)
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    store.move_card(key, "done", 0)
    res = flush_writebacks(store)
    assert not res.ok and res.written == 0
    assert store.cards[key]["written_status"] == ""    # nothing recorded
    assert store.pending_writebacks() == [(key, "Complete")]   # retried next flush


def test_requeue_stale_writeback_when_row_still_needs_review(tmp_path):
    """A 202'd triage event whose flow run failed (or a file-mode write the
    workbook reverted) leaves the row at Needs Review — readback re-queues."""
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    store.move_card(key, "done", 0)
    store.cards[key]["written_status"] = "Complete"    # delivery looked done
    assert store.pending_writebacks() == []
    requeued = store.requeue_stale_writebacks(
        _load(feedback=[_same_key(key, "Needs Review")]))
    assert requeued == 1
    assert store.pending_writebacks() == [(key, "Complete")]


def test_requeue_leaves_manual_edits_alone(tmp_path):
    """Deferred (or any non-Needs-Review value) is a manual edit and wins —
    only the exact 'Needs Review' re-queues."""
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    store.move_card(key, "done", 0)
    store.cards[key]["written_status"] = "Complete"
    assert store.requeue_stale_writebacks(
        _load(feedback=[_same_key(key, "Deferred")])) == 0
    assert store.requeue_stale_writebacks(
        _load(feedback=[_same_key(key, "Complete")])) == 0
    assert store.pending_writebacks() == []


def test_external_status_skips_card_with_restore_in_flight(tmp_path):
    """Webhook lag: a card dragged OUT of Done still has written_status set
    until the restore lands; its row's stale 'Complete' must not yank the
    card straight back into Done."""
    store = BoardStore(path=tmp_path / "b.json")
    store.sync(_load(feedback=[_entry(1, "911 Setup", "feedback")]))
    key = store.bucket("todo")["cards"][0]
    store.move_card(key, "done", 0)
    store.cards[key]["written_status"] = "Complete"    # terminal write delivered
    store.move_card(key, "todo", 0)                    # dragged back out; restore pending
    moved = store.apply_external_status(_load(feedback=[_same_key(key, "Complete")]))
    assert moved == 0
    assert key in store.bucket("todo")["cards"]

