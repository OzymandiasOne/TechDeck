"""Tests for the pure plugin_sdk helpers (Hard Rules 1-3 header lookup, batch
parsing, tube classification, the batch cache, run-outcome params, and the
user-facing error mapping)."""

import builtins
import threading

import pytest

from techdeck.core import plugin_sdk
from techdeck.core.plugin_executor import PluginExecutor


# ── tiny openpyxl-like worksheet stub ────────────────────────────────────────
class _Cell:
    def __init__(self, value):
        self.value = value


class _WS:
    """rows is a 1-based-by-index list of lists (row-major)."""
    def __init__(self, rows):
        self._rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=0)

    def cell(self, row=None, column=None):
        try:
            v = self._rows[row - 1][column - 1]
        except IndexError:
            v = None
        return _Cell(v)


def _sheet():
    return _WS([
        [None, None, None],
        ["ORDER", "DYPN", "Source Material"],   # header row 2, mixed case
        ["A1", "D1", "218002867"],
    ])


# ── header helpers ───────────────────────────────────────────────────────────
def test_find_header_col_case_insensitive():
    ws = _sheet()
    assert plugin_sdk.find_header_col(ws, "order", 2) == 1
    assert plugin_sdk.find_header_col(ws, "  source material ", 2) == 3
    assert plugin_sdk.find_header_col(ws, "missing", 2) is None


def test_header_map():
    assert plugin_sdk.header_map(_sheet(), 2) == {
        "ORDER": 1, "DYPN": 2, "SOURCE MATERIAL": 3}


def test_find_header_row_scans_for_required():
    ws = _sheet()
    row, mp = plugin_sdk.find_header_row(ws, ["order", "dypn"])
    assert row == 2 and mp["DYPN"] == 2
    assert plugin_sdk.find_header_row(ws, ["nonexistent"]) == (None, {})


# ── batch parsing ────────────────────────────────────────────────────────────
def test_parse_922_batch():
    f = plugin_sdk.parse_922_batch
    assert f("473") == "473"
    assert f("Batch 473") == "473"
    assert f("PO #473") == "473"
    assert f("#473") == "473"
    assert f("  batch  473 ") == "473"
    assert f("nope") is None
    assert f(None) is None


def test_normalize_911_batch():
    assert plugin_sdk.normalize_911_batch("  v060 ") == "V060"
    assert plugin_sdk.normalize_911_batch(None) == ""


# ── tube classification ──────────────────────────────────────────────────────
def test_tube_class():
    std = next(iter(plugin_sdk.STANDARD_TUBE_MATERIALS))
    ovr = next(iter(plugin_sdk.OVERSIZED_TUBE_MATERIALS))
    assert plugin_sdk.tube_class(std) == "standard"
    assert plugin_sdk.tube_class(f"  {ovr}  ") == "oversized"    # stripped
    assert plugin_sdk.tube_class("not-a-serial") is None
    assert plugin_sdk.tube_class(None) is None


# ── batch-number family cache ────────────────────────────────────────────────
def test_request_batch_number_caches_within_family(monkeypatch):
    calls = []
    monkeypatch.setattr(builtins, "input",
                        lambda prompt="": (calls.append(prompt) or "473"))
    shared = {"922": {}}
    params = {"shared_state": shared, "plugin_family": "922"}

    assert plugin_sdk.request_batch_number(params, "Batch?") == "473"
    assert shared["922"]["batch_number"] == "473"
    # Second same-family plugin reuses the cache — no second prompt.
    assert plugin_sdk.request_batch_number(params, "Batch?") == "473"
    assert len(calls) == 1


# ── run-outcome params round-trip with the executor ──────────────────────────
def test_ticket_units_and_run_outcome_roundtrip():
    p = {}
    plugin_sdk.set_ticket_units(p, 3)
    assert PluginExecutor._ticket_units_from(p) == 3
    plugin_sdk.set_ticket_units(p, 0)                 # clamps to 1
    assert PluginExecutor._ticket_units_from(p) == 1

    p2 = {}
    plugin_sdk.set_run_outcome(p2, plugin_sdk.RUN_OUTCOME_WARNING, "heads up")
    assert PluginExecutor._run_outcome_from(p2) == ("warning", "heads up")


# ── user-facing error mapping ────────────────────────────────────────────────
def test_user_facing_error_normalizes_whitespace():
    e = plugin_sdk.UserFacingError("problem   text", "the   fix")
    assert e.problem == "problem text"
    assert e.fix == "the fix"


def test_as_user_facing_recognizes_and_rejects():
    e = plugin_sdk.UserFacingError("problem", "fix")
    assert plugin_sdk.as_user_facing(e) is e

    perm = PermissionError(13, "denied")
    perm.filename = "C:/x/pricing.xlsm"
    mapped = plugin_sdk.as_user_facing(perm)
    assert mapped is not None
    assert "open in another program" in mapped.problem.lower()

    assert plugin_sdk.as_user_facing(ValueError("plain code bug")) is None


# ── cooperative cancellation ─────────────────────────────────────────────────
def test_cancellation_helpers():
    ev = threading.Event()
    # None + unset -> not cancelled; helpers are no-ops.
    assert plugin_sdk.cancelled(None) is False
    assert plugin_sdk.cancelled(ev) is False
    plugin_sdk.raise_if_cancelled(ev)
    plugin_sdk.raise_if_cancelled(None)
    # Set -> cancelled; raise_if_cancelled raises PluginCancelled (an Exception,
    # so the executor's generic handler can't miss it, but it's ordered first).
    ev.set()
    assert plugin_sdk.cancelled(ev) is True
    # BaseException (not Exception) so a plugin's `except Exception` around
    # per-file work can't swallow a cancel; the executor still catches it.
    assert issubclass(plugin_sdk.PluginCancelled, BaseException)
    assert not issubclass(plugin_sdk.PluginCancelled, Exception)
    with pytest.raises(plugin_sdk.PluginCancelled):
        plugin_sdk.raise_if_cancelled(ev)


def test_as_user_facing_walks_cause_chain():
    inner = plugin_sdk.UserFacingError("inner problem", "fix")
    try:
        try:
            raise inner
        except plugin_sdk.UserFacingError as e:
            raise RuntimeError("wrapped") from e
    except RuntimeError as outer:
        assert plugin_sdk.as_user_facing(outer) is inner


# ── visible_sheetnames (hidden staging sheets must not win sheet scans) ─────
def test_visible_sheetnames_filters_hidden():
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "FORMING"
    wb.active.sheet_state = "hidden"
    wb.create_sheet("PO- 4130")
    hidden2 = wb.create_sheet("PRICING")
    hidden2.sheet_state = "veryHidden"
    wb.create_sheet("Notes")
    assert plugin_sdk.visible_sheetnames(wb) == ["PO- 4130", "Notes"]


def test_nest_id_re_is_the_hard_rule_3_shape():
    """The canonical nest-id pattern (single home since 2026-08-17; six
    plugins alias it as their local _NEST_RE). Positive: legacy numeric,
    P/S-prefixed, alphanumeric-with-digit. Negative: the junk that used to
    leak in as nests."""
    from techdeck.core.plugin_sdk import NEST_ID_RE, is_nest_id
    for good in ("503633", "P08229", "S045123", "5CDAVW", "9FANDR", "GX030"):
        assert NEST_ID_RE.match(good), good
        assert is_nest_id(f"  {good}  "), good
    for junk in ("TOTALS", "NEST PACKAGES", "", "12", "ABCDEFGH", "WPDD"):
        assert not is_nest_id(junk), junk


def test_plugin_nest_regexes_are_aliases_not_copies():
    """The pattern was revised once (the alphanumeric branch) and the fix
    had to land in six files. Guard the dedup: no plugin may re-compile a
    pattern containing the nest shape's tell ([PS]?\d{3,})."""
    from pathlib import Path
    plugins = Path(__file__).resolve().parents[2] / "plugins"
    offenders = []
    for py in plugins.rglob("*.py"):
        for i, line in enumerate(
                py.read_text(encoding="utf-8", errors="replace").splitlines(), 1):
            # Only actual pattern strings count - a comment DESCRIBING the
            # shape is fine (batch_auditor documents it in prose).
            if "[PS]?" in line and ("re.compile" in line
                                    or 'r"' in line or "r'" in line):
                offenders.append(f"{py.relative_to(plugins.parent)}:{i}")
    assert offenders == [], (
        "re-typed Hard Rule 3 nest regex (alias sdk.NEST_ID_RE instead): "
        + ", ".join(offenders))
