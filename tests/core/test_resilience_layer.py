"""First tests for the OneDrive resilience layer (Hard Rule 13).

This is the most-called code in the SDK (~57 plugin call sites), exists
BECAUSE it failed in the field on four documented machines (NRAPINI-LT,
LAPTOP-1AMLBK7B, CPENG_TOWERPC, + the v0.8.6.1 sweep), is now
build-MANDATED by gate E13 — and had zero tests until this file.

The cloud conditions are simulated: `is_cloud_placeholder` / hydration are
monkeypatched, locked files raise PermissionError the way Excel's
deny-share mode does. No OneDrive needed.
"""

import zipfile

import pytest

from techdeck.core import plugin_sdk as sdk


# ---------------------------------------------------------------- primitives

def test_is_cloud_placeholder_false_on_a_normal_file(tmp_path):
    f = tmp_path / "normal.txt"
    f.write_text("content", encoding="utf-8")
    assert sdk.is_cloud_placeholder(f) is False


def test_is_cloud_placeholder_false_on_a_missing_file(tmp_path):
    assert sdk.is_cloud_placeholder(tmp_path / "nope.txt") is False


def test_is_cloud_placeholder_reads_the_attribute_bits(monkeypatch, tmp_path):
    class _Stat:
        st_file_attributes = 0x00400000  # FILE_ATTRIBUTE_RECALL_ON_DATA_ACCESS

    monkeypatch.setattr(sdk.os, "stat", lambda p: _Stat())
    assert sdk.is_cloud_placeholder(tmp_path / "anything.xlsx") is True


def test_hydrate_reads_a_real_file_clean(tmp_path):
    f = tmp_path / "data.bin"
    f.write_bytes(b"x" * 4096)
    sdk.hydrate_cloud_file(f)  # must not raise


def test_hydrate_failure_is_user_facing_with_the_onedrive_fix(tmp_path):
    missing = tmp_path / "cloud_only.xlsx"
    with pytest.raises(sdk.UserFacingError) as exc_info:
        sdk.hydrate_cloud_file(missing, attempts=2, delay=0)
    assert "cloud" in exc_info.value.problem
    assert "Always keep on this device" in exc_info.value.fix


def test_ensure_local_is_a_noop_on_a_normal_file(monkeypatch, tmp_path):
    f = tmp_path / "local.xlsx"
    f.write_text("x", encoding="utf-8")
    hydrated = []
    monkeypatch.setattr(sdk, "hydrate_cloud_file",
                        lambda *a, **k: hydrated.append(True))
    sdk.ensure_local(f)
    assert hydrated == []


def test_ensure_local_hydrates_a_placeholder_and_logs(monkeypatch, tmp_path):
    f = tmp_path / "cloudy.xlsx"
    hydrated, lines = [], []
    monkeypatch.setattr(sdk, "is_cloud_placeholder", lambda p: True)
    monkeypatch.setattr(sdk, "hydrate_cloud_file",
                        lambda *a, **k: hydrated.append(True))
    sdk.ensure_local(f, log=lines.append)
    assert hydrated == [True]
    assert any("cloud-only" in line for line in lines)


def test_is_share_lock_signatures():
    assert sdk._is_share_lock(PermissionError(13, "denied")) is True
    err22 = OSError(22, "Invalid argument")
    assert sdk._is_share_lock(err22) is False       # that's the placeholder
    err13 = OSError()
    err13.errno = 13
    assert sdk._is_share_lock(err13) is True
    assert sdk._is_share_lock(ValueError("x")) is False


def test_locked_file_error_names_the_file_and_the_lock_file(tmp_path):
    err = sdk.locked_file_error(tmp_path / "PO H473 Pricing.xlsm")
    assert "PO H473 Pricing.xlsm" in err.problem
    assert "open in another program" in err.problem
    assert "~$PO H473 Pricing" in err.fix


# ---------------------------------------------------------------- the core

def _fail_then_succeed(exc, result="loaded"):
    """An op that raises `exc` on the first call, returns `result` after."""
    calls = []

    def op():
        calls.append(True)
        if len(calls) == 1:
            raise exc
        return result

    return op, calls


def test_core_happy_path_no_retry(tmp_path):
    f = tmp_path / "fine.xlsx"
    f.write_text("x", encoding="utf-8")
    calls = []
    result = sdk._resilient_read(lambda: calls.append(True) or "ok", f)
    assert result == "ok"
    assert len(calls) == 1


def test_core_locked_file_becomes_user_facing(tmp_path):
    f = tmp_path / "open_in_excel.xlsx"
    f.write_text("x", encoding="utf-8")

    def op():
        raise PermissionError(13, "The process cannot access the file")

    with pytest.raises(sdk.UserFacingError) as exc_info:
        sdk._resilient_read(op, f)
    assert "open in another program" in exc_info.value.problem


@pytest.mark.parametrize("disguise", [
    OSError(22, "Invalid argument"),
    zipfile.BadZipFile("File is not a zip file"),
], ids=["errno22", "badzipfile"])
def test_core_placeholder_disguises_hydrate_and_retry(monkeypatch, tmp_path,
                                                      disguise):
    """The two field-observed shapes of the same failure: raw Errno 22, and
    zipfile eating that OSError and re-raising it as BadZipFile."""
    f = tmp_path / "cloudy.xlsx"
    f.write_text("x", encoding="utf-8")
    monkeypatch.setattr(sdk, "is_cloud_placeholder", lambda p: True)
    monkeypatch.setattr(sdk, "hydrate_cloud_file", lambda *a, **k: None)

    op, calls = _fail_then_succeed(disguise)
    lines = []
    # ensure_local also sees the placeholder and "hydrates" (patched no-op);
    # the retry after the failing eager op is what this asserts.
    assert sdk._resilient_read(op, f, log=lines.append) == "loaded"
    assert len(calls) == 2
    assert any("cloud-only" in line for line in lines)


def test_core_non_placeholder_oserror_reraises(monkeypatch, tmp_path):
    """An OSError on a file that is NOT a placeholder is a real error —
    hydrating can't fix it and must not mask it."""
    f = tmp_path / "healthy_but_broken.xlsx"
    f.write_text("x", encoding="utf-8")
    monkeypatch.setattr(sdk, "is_cloud_placeholder", lambda p: False)

    def op():
        raise OSError(22, "Invalid argument")

    with pytest.raises(OSError):
        sdk._resilient_read(op, f)


def test_core_unrelated_exceptions_pass_straight_through(tmp_path):
    f = tmp_path / "fine.xlsx"
    f.write_text("x", encoding="utf-8")

    def op():
        raise ValueError("a genuine parsing bug")

    with pytest.raises(ValueError):
        sdk._resilient_read(op, f)


# ---------------------------------------------------------------- public fns

def test_load_workbook_resilient_round_trip(tmp_path):
    import openpyxl
    src = openpyxl.Workbook()
    src.active.append(["NEST PKG NBR", "QTY"])
    xlsx = tmp_path / "real.xlsx"
    src.save(xlsx)

    wb = sdk.load_workbook_resilient(xlsx)
    assert wb.active.cell(row=1, column=1).value == "NEST PKG NBR"


def test_load_workbook_resilient_locked_is_user_facing(monkeypatch, tmp_path):
    import openpyxl
    f = tmp_path / "locked.xlsx"
    f.write_text("x", encoding="utf-8")

    def locked(*a, **k):
        raise PermissionError(13, "in use")
    monkeypatch.setattr(openpyxl, "load_workbook", locked)

    with pytest.raises(sdk.UserFacingError):
        sdk.load_workbook_resilient(f)


def test_read_excel_resilient_round_trip(tmp_path):
    import openpyxl
    src = openpyxl.Workbook()
    src.active.append(["DYPN"])
    src.active.append(["R123-H1-2"])
    xlsx = tmp_path / "real.xlsx"
    src.save(xlsx)

    df = sdk.read_excel_resilient(xlsx)
    assert list(df.columns) == ["DYPN"]


def test_open_excel_resilient_round_trip(tmp_path):
    import openpyxl
    src = openpyxl.Workbook()
    src.active.title = "PO"
    xlsx = tmp_path / "real.xlsx"
    src.save(xlsx)

    xls = sdk.open_excel_resilient(xlsx)
    assert "PO" in xls.sheet_names


# ---------------------------------------------------------------- copy

def test_copy_resilient_copies(tmp_path):
    src = tmp_path / "src.xlsx"
    src.write_bytes(b"workbook bytes")
    dest = tmp_path / "out"
    dest.mkdir()
    result = sdk.copy_resilient(src, dest / "copy.xlsx")
    assert result.read_bytes() == b"workbook bytes"


def test_copy_resilient_lock_names_the_destination(monkeypatch, tmp_path):
    """copy2 raising a share lock while the SOURCE still reads fine means the
    open handle is on the destination (the write-back case)."""
    src = tmp_path / "src.xlsx"
    src.write_bytes(b"data")
    dest = tmp_path / "dest.xlsx"

    def locked(*a, **k):
        raise PermissionError(13, "in use")
    monkeypatch.setattr(sdk.shutil, "copy2", locked)

    with pytest.raises(sdk.UserFacingError) as exc_info:
        sdk.copy_resilient(src, dest)
    assert "dest.xlsx" in exc_info.value.problem


def test_copy_resilient_non_lock_errors_reraise(tmp_path):
    with pytest.raises(FileNotFoundError):
        sdk.copy_resilient(tmp_path / "missing.xlsx", tmp_path / "out.xlsx")
