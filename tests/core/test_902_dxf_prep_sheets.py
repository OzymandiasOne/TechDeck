"""Regression tests for 902 DXF Prep's PO-sheet selection + part-number scheme.

Batch 4130 (2026-07-30): the pricing workbook's PO tab was named 'PO- 4130'
(hyphen AND space) and the old ^PO[-_ ]?\\d+ regex missed it, so the fallback
scan read a hidden 3-row FORMING staging sheet instead of the 231-row PO
sheet — three "successful" runs built a 3-part QTY workbook. The fix widens
the separator match to [-_ ]* and restricts every scan to visible sheets.

Same day, a PO carried part numbers the original 1-2-letters+digits DYPN
scheme rejected (263500220-5001, H5730805A230-1, H5731304C49-13, LTG-2213AE):
every one was flagged "unrecognized filename" and dumped in EXTRA\\ while the
PO-row filter skipped them too. _PART_RE now accepts any dash-separated
alphanumeric ID containing a digit.
"""

import importlib.util
from pathlib import Path

import openpyxl
import pytest

_RUN_PY = Path(__file__).resolve().parents[2] / "plugins" / "902_dxf_prep" / "run.py"


@pytest.fixture(scope="module")
def plugin():
    spec = importlib.util.spec_from_file_location("run902_test", _RUN_PY)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _po_headers(ws, rows):
    ws.append(["HULL", "ORDER#", "DYPN", "SOURCE MATERIAL", "QTY ORDERED"])
    for dypn, qty in rows:
        ws.append(["BJ", "BJA00000", dypn, "EB218001123", qty])


def _write(tmp_path, name, build):
    wb = openpyxl.Workbook()
    build(wb)
    path = tmp_path / name
    wb.save(path)
    return path


def test_po_tab_with_space_beats_hidden_forming_sheet(tmp_path, plugin):
    """The 4130 repro: hidden FORMING sheet first, real tab named 'PO- 4130'."""
    def build(wb):
        forming = wb.active
        forming.title = "FORMING"
        _po_headers(forming, [("H4141313-31", 5), ("H4144001-14", 40),
                              ("H5731615-142", 2)])
        forming.sheet_state = "hidden"
        po = wb.create_sheet("PO- 4130")
        _po_headers(po, [("R4111921-42", 4), ("R4138924-154", 4),
                         ("R4140922-1M", 1), ("R4140933-16M", 1)])

    path = _write(tmp_path, "4130 PRICING.xlsx", build)
    batch, rows = plugin._load_po_rows(path, log=lambda m: None)
    assert batch == "4130"
    assert [d for d, _ in rows] == ["R4111921-42", "R4138924-154",
                                    "R4140922-1M", "R4140933-16M"]


@pytest.mark.parametrize("tab", ["PO-4131", "PO 4131", "PO_4131",
                                 "PO4131", "PO -  4131", "po- 4131"])
def test_po_tab_separator_variants_all_match(tmp_path, plugin, tab):
    def build(wb):
        wb.active.title = "Notes"
        _po_headers(wb.create_sheet(tab), [("R4111921-42", 1)])

    path = _write(tmp_path, f"{tab.replace(' ', '_')}.xlsx", build)
    batch, rows = plugin._load_po_rows(path, log=lambda m: None)
    assert batch == "4131"
    assert len(rows) == 1


# ── Part-number scheme ──────────────────────────────────────────────────────

@pytest.mark.parametrize("pn", [
    # legacy 1-2 letters + digits + dash segments
    "H4130810-484", "R4533683-5-3", "K5731417-325", "R4140922-1M",
    # all-numeric head (first seen on the 2026-07-30 PO)
    "263500220-5001", "263500220-7001", "263511908-301",
    # letters embedded mid-number
    "H5730805A230-1", "H5731304C49-13", "H5731304C49-19",
    # letter-only head, digits in a later segment
    "LTG-2213AE", "LTG-2213N", "LTG-2213V",
])
def test_part_re_accepts_real_po_part_numbers(plugin, pn):
    assert plugin._PART_RE.match(pn)
    assert plugin._PART_RE.match(pn.lower())


@pytest.mark.parametrize("junk", [
    "NAME OF VENDOR REPRESENTATIVE",   # footer prose (spaces)
    "TOTAL",                           # no dash segment
    "LTG-ABC",                         # no digit anywhere
    "N/A", "",                         # slash / empty
    "SEE NOTE 1",
])
def test_part_re_still_rejects_junk_rows(plugin, junk):
    assert not plugin._PART_RE.match(junk)


@pytest.mark.parametrize("stem,expected", [
    # export rev + counter suffixes strip down to the PO part number
    ("H5731304C49-13_A_1", "H5731304C49-13"),
    ("LTG-2213AE_E_1", "LTG-2213AE"),
    ("LTG-2213N_FLAT-PATTERN#1", "LTG-2213N"),
    ("(4x) 263500220-5001", "263500220-5001"),
    # already-clean names come back unchanged
    ("263511908-301", "263511908-301"),
    ("H5730805A230-1", "H5730805A230-1"),
])
def test_clean_stem_resolves_new_pn_shapes(plugin, stem, expected):
    assert plugin.clean_stem(stem) == expected


def test_fallback_scan_skips_hidden_sheets_and_warns(tmp_path, plugin):
    """No PO-named tab at all: the fallback must pick the VISIBLE sheet with
    matching headers, never a hidden one earlier in the book, and must warn."""
    def build(wb):
        hidden = wb.active
        hidden.title = "FORMING"
        _po_headers(hidden, [("H9999999-1", 1)])
        hidden.sheet_state = "hidden"
        _po_headers(wb.create_sheet("order list"), [("R4111921-42", 2)])

    path = _write(tmp_path, "no_po_tab.xlsx", build)
    logs = []
    batch, rows = plugin._load_po_rows(path, log=logs.append)
    assert batch is None                       # no PO tab -> no batch from name
    assert [d for d, _ in rows] == ["R4111921-42"]
    assert any("WARNING" in m and "order list" in m for m in logs)
