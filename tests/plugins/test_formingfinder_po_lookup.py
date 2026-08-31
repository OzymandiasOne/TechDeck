"""FormingFinder's PO lookup — picking the PLATE on dual-material DYPNs.

The fixture replicates the real Batch 485 PO: a part number can appear on two
PO rows with different source materials — a formed PLATE ('OSS 0.625 THK') and
a ROD ('OSS 0.625 OD ASTM A36') sharing one DYPN. First-row-wins recorded the
rod's serial on the Bent Plates sheet, and 922 Kitting (whose raw-material
guard trusts that sheet) stamped FORMED on the rods of Kitting 485 pages 30/35
(2026-08-05). A BEND note still wins outright; with the PO silent, the
source-material DESCRIPTION decides; conflicting evidence warns instead of
guessing.

The lookup is keyed by (ORDER, DYPN): one batch can build the same part for two
orders, and each order needs its own binder copy and Bent Plates row (Batch 490:
BL427447 and BM352088 both build R7211262-H2-3, and DYPN-only keying silently
dropped the second - only 922 Kitting, which walks per order, caught it).
"""

import importlib.util
from pathlib import Path

import openpyxl
import pytest

PLUGINS = Path(__file__).resolve().parents[2] / "plugins"


@pytest.fixture(scope="module")
def ff():
    spec = importlib.util.spec_from_file_location(
        "ff_run", PLUGINS / "922_formingfinder" / "run.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _po_workbook(tmp_path, po_rows, materials):
    """A minimal QF-QU-09 lookalike: PO sheet (header on row 5, like the real
    485 workbook) + SOURCE MATERIAL sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO"
    ws.append([]); ws.append([]); ws.append([]); ws.append([])
    ws.append(["ORDER", "PPN", "DYPN", "SOURCE MATERIAL", "NOTES"])
    for row in po_rows:
        ws.append(row)
    sm = wb.create_sheet("SOURCE MATERIAL")
    sm.append(["Source Material", "", "Weight/FT", "Quanity",
               "Scribe Description", "Part Description"])
    for serial, scribe, part in materials:
        sm.append([serial, "", "1LB", 1, scribe, part])
    path = tmp_path / "PO H485 QF-QU-09 REV C.xlsx"
    wb.save(path)
    return path


# Real Batch 485 material descriptions.
_MATERIALS = [
    ("262028653-11", "OSS 0.50 THK ", "0.50 THK X 10.60 X 4.25 Length"),
    ("262028653-12", "OSS 0.375 OD ASTM A36", "0.375 OD X 18.00 Length"),
    ("262028653-48", "OSS 0.625 OD ASTM A36", "0.625 OD X 24.00 Length"),
    ("262028653-50", "OSS 0.625 THK", "0.625 THK X 9.60 X 6.60 Length"),
    ("262028653-18", "OSS 0.50 THK ", "0.375 OD X 2.00 Length"),  # conflict
]


def test_bend_note_still_wins(ff, tmp_path):
    # BK499434's H51: the rod row comes FIRST, but -11 carries the BEND note.
    po = _po_workbook(tmp_path, [
        ["BK499434", "R6455461-H51", "R6455461-H51-2", "262028653-12", ""],
        ["BK499434", "R6455461-H51", "R6455461-H51-2", "262028653-11", "BEND"],
    ], _MATERIALS)
    lookup = ff._load_po_lookup(po, lambda *a: None)
    assert lookup[("bk499434", "r6455461-h51-2")]["SOURCE MATERIAL"] == "262028653-11"


def test_silent_po_picks_the_plate_over_the_rod(ff, tmp_path):
    # X7371983's H417-3 (Kitting 485 page 30): rod -48 listed FIRST, no notes.
    po = _po_workbook(tmp_path, [
        ["X7371983", "H7441366-H417", "H7441366-H417-3", "262028653-48", ""],
        ["X7371983", "H7441366-H417", "H7441366-H417-3", "262028653-50", ""],
    ], _MATERIALS)
    lookup = ff._load_po_lookup(po, lambda *a: None)
    row = lookup[("x7371983", "h7441366-h417-3")]
    assert row["SOURCE MATERIAL"] == "262028653-50"
    # The choice note is stashed for the caller to log when the part is
    # actually discovered as formed (not spammed for every PO DYPN).
    assert "using plate 262028653-50" in row["_PICK_NOTE"]


def test_silent_po_picks_the_plate_in_either_row_order(ff, tmp_path):
    # BL392386's H10-2 (page 35): rod -12 first there too, but don't rely on it.
    po = _po_workbook(tmp_path, [
        ["BL392386", "R7653561-H10", "R7653561-H10-2", "262028653-11", ""],
        ["BL392386", "R7653561-H10", "R7653561-H10-2", "262028653-12", ""],
    ], _MATERIALS)
    lookup = ff._load_po_lookup(po, lambda *a: None)
    assert lookup[("bl392386", "r7653561-h10-2")]["SOURCE MATERIAL"] == "262028653-11"


def test_single_row_needs_no_description(ff, tmp_path):
    po = _po_workbook(tmp_path, [
        ["X8390862", "H7655461-H12", "H7655461-H12-4", 218019941, ""],
    ], _MATERIALS)
    lookup = ff._load_po_lookup(po, lambda *a: None)
    assert lookup[("x8390862", "h7655461-h12-4")]["SOURCE MATERIAL"] == 218019941


def test_conflicting_descriptions_warn_instead_of_guessing(ff, tmp_path):
    # -18's scribe says THK, its part description says OD (real 485 row) -> no
    # unambiguous plate; keep a deterministic pick but tell the user to verify.
    po = _po_workbook(tmp_path, [
        ["X1", "P-1", "P-1-2", "262028653-18", ""],
        ["X1", "P-1", "P-1-2", "262028653-12", ""],
    ], _MATERIALS)
    lookup = ff._load_po_lookup(po, lambda *a: None)
    row = lookup[("x1", "p-1-2")]
    assert row["SOURCE MATERIAL"] == "262028653-18"  # not the rod
    assert "verify the FORMED tag" in row["_PICK_WARN"]


def test_two_plates_warn_instead_of_guessing(ff, tmp_path):
    po = _po_workbook(tmp_path, [
        ["X1", "P-1", "P-1-2", "262028653-11", ""],
        ["X1", "P-1", "P-1-2", "262028653-50", ""],
    ], _MATERIALS)
    lookup = ff._load_po_lookup(po, lambda *a: None)
    row = lookup[("x1", "p-1-2")]
    assert row["SOURCE MATERIAL"] == "262028653-11"
    assert "_PICK_WARN" in row


def test_missing_source_material_sheet_degrades_to_first_row(ff, tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO"
    ws.append(["ORDER", "PPN", "DYPN", "SOURCE MATERIAL", "NOTES"])
    ws.append(["X1", "P-1", "P-1-2", "262028653-48", ""])
    ws.append(["X1", "P-1", "P-1-2", "262028653-50", ""])
    path = tmp_path / "po.xlsx"
    wb.save(path)
    lookup = ff._load_po_lookup(path, lambda *a: None)
    row = lookup[("x1", "p-1-2")]
    assert row["SOURCE MATERIAL"] == "262028653-48"
    assert "_PICK_WARN" in row


@pytest.mark.parametrize("desc,kind", [
    ("OSS 0.625 THK", "plate"),
    ("OSS 0.625 OD ASTM A36", "rod"),
    ("0.50 THK X 10.60 X 4.25 Length", "plate"),
    ("HSS 1.5 THK / 1.75 X 1.50 THK X 15.00 Length", "plate"),
    ("OSS 0.50 THK  / 0.375 OD X 2.00 Length", "unknown"),   # conflict
    ("", "unknown"),
])
def test_material_kind(ff, desc, kind):
    assert ff._material_kind(desc) == kind


# --- flat bars are formable too (formed in-house since 2026-08) ---------------

@pytest.mark.parametrize("desc,kind", [
    ("OSS 0.375 X 2.00 FLAT BAR ASTM A36", "bar"),
    ("0.375 X 2.00 FLATBAR", "bar"),
    ("OSS 0.375 OD FLAT BAR", "unknown"),   # bar + rod evidence = conflict
])
def test_material_kind_flat_bar(ff, desc, kind):
    assert ff._material_kind(desc) == kind


def test_silent_po_picks_the_flat_bar_over_the_rod(ff, tmp_path):
    materials = _MATERIALS + [
        ("262028652-77", "OSS 0.375 X 2.00 FLAT BAR",
         "0.375 X 2.00 FLAT BAR X 24.00 Length"),
    ]
    po = _po_workbook(tmp_path, [
        ["X1", "P-2", "P-2-3", "262028653-48", ""],   # rod
        ["X1", "P-2", "P-2-3", "262028652-77", ""],   # flat bar
    ], materials)
    lookup = ff._load_po_lookup(po, lambda *a: None)
    row = lookup[("x1", "p-2-3")]
    assert row["SOURCE MATERIAL"] == "262028652-77"
    assert "using flat bar 262028652-77" in row["_PICK_NOTE"]


@pytest.mark.parametrize("name,dypn", [
    ("E6492162-H22-3 PLT F.pdf", "E6492162-H22-3"),
    ("E6492162-H22-3 BAR F.pdf", "E6492162-H22-3"),
    ("E6492162-H22-3 BAR.pdf", "E6492162-H22-3"),
])
def test_dypn_from_filename(ff, name, dypn):
    assert ff._dypn_from_filename(name) == dypn


@pytest.mark.parametrize("name,formed", [
    ("E6492162-H22-3 PLT F.pdf", True),
    ("E6492162-H22-3 BAR F.pdf", True),
    ("E6492162-H22-3 PLT.pdf", False),
    ("E6492162-H22-3 BAR.pdf", False),
])
def test_formed_suffix(ff, name, formed):
    assert ff._has_formed_suffix(name) is formed


# --- the same part built for two orders --------------------------------------

def test_same_dypn_in_two_orders_keeps_both(ff, tmp_path):
    """Batch 490: BL427447 and BM352088 both build R7211262-H2-3, each with the
    same plate/rod pair. Keying by DYPN alone collapsed them to one row."""
    po = _po_workbook(tmp_path, [
        ["BL427447", "R7211262-H2", "R7211262-H2-3", "262028653-48", ""],
        ["BL427447", "R7211262-H2", "R7211262-H2-3", "262028653-50", ""],
        ["BM352088", "R7211262-H2", "R7211262-H2-3", "262028653-48", ""],
        ["BM352088", "R7211262-H2", "R7211262-H2-3", "262028653-50", ""],
    ], _MATERIALS)
    lookup = ff._load_po_lookup(po, lambda *a: None)
    assert ("bl427447", "r7211262-h2-3") in lookup
    assert ("bm352088", "r7211262-h2-3") in lookup
    # Each order still picks the PLATE over the rod, independently.
    for order in ("bl427447", "bm352088"):
        row = lookup[(order, "r7211262-h2-3")]
        assert row["SOURCE MATERIAL"] == "262028653-50"
        assert row["ORDER"].casefold() == order


def test_row_with_no_order_still_indexes(ff, tmp_path):
    po = _po_workbook(tmp_path, [
        [None, "P-9", "P-9-1", "262028653-50", ""],
    ], _MATERIALS)
    lookup = ff._load_po_lookup(po, lambda *a: None)
    assert lookup[("", "p-9-1")]["SOURCE MATERIAL"] == "262028653-50"


@pytest.mark.parametrize("folder,expected", [
    ("BL427447-R7211262-H2", "BL427447"),
    ("X6524642-H7918266-H27", "X6524642"),
    ("BM347140-R8653362-H15L", "BM347140"),
])
def test_order_from_folder_name(ff, folder, expected):
    assert ff._order_from_folder_name(folder) == expected


def test_order_for_pdf_uses_the_top_level_folder(ff, tmp_path):
    batch = tmp_path / "Batch 490"
    pdf = batch / "BL427447-R7211262-H2" / "CAD-AND-SHOP-PRINTS" / "x PLT F.pdf"
    assert ff._order_for_pdf(pdf, batch, {"bl427447"}) == "BL427447"
    # No PO to check against -> still splits on the first hyphen.
    assert ff._order_for_pdf(pdf, batch, set()) == "BL427447"
    # A PDF sitting loose in the batch root belongs to no order.
    assert ff._order_for_pdf(batch / "loose.pdf", batch, set()) == ""


def test_unique_dest_suffixes_repeat_filenames(ff, tmp_path):
    used: set[str] = set()
    names = [ff._unique_dest(tmp_path, "R7211262-H2-3 PLT F.pdf", used).name
             for _ in range(3)]
    assert names == ["R7211262-H2-3 PLT F.pdf",
                     "R7211262-H2-3 PLT F (2).pdf",
                     "R7211262-H2-3 PLT F (3).pdf"]
