"""911 Setup PLATE mode (v2.1.0, coworker feedback 2026-09-03).

The plugin was tailored exclusively to SHAPE batches; plate batches were set
up by hand. The three reported inaccuracies these tests pin down:

1. MIL SPEC must be the literal 'N/A' for FERROUS (carbon) plate — decided
   by the packet's own MOVE TICKET 'FERROUS:' flag ('F'), never a keyword
   guess. Non-ferrous plate ('N': CRES/IN625/MONEL/CUNI/BRASS, 'A': aluminum)
   keeps the real spec. SHAPE runs are byte-for-byte unaffected.
2. The SCRIBE sheet's UNIQUE - TRACE column is real data on plate (manually
   entered from the forecast today, hardcoded to "N/A" by the SHAPE
   template's formulas) — plate runs fill it from the forecast's TRACE/MIC
   column, and a blank forecast leaves it blank, never "N/A".
3. Plate runs must use the plate templates: '911 PLATE BATCH _.xlsx' and the
   PLATES scribe form, both of which already lived in the SACO dir but were
   unreachable — the finder's '911 BATCH' prefix can never match them.
"""

import importlib.util
from pathlib import Path

import openpyxl
import pytest

PLUGIN = (Path(__file__).resolve().parents[2]
          / "plugins" / "911_setup" / "run.py")


@pytest.fixture(scope="module")
def su():
    spec = importlib.util.spec_from_file_location("su_911_setup", PLUGIN)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ── template finder: the two prefixes must never cross-match ────────────────
@pytest.fixture
def saco(tmp_path):
    """A SACO template dir with both workbook templates, as on the real share."""
    (tmp_path / "911 BATCH _.xlsx").write_bytes(b"shape")
    (tmp_path / "911 PLATE BATCH _.xlsx").write_bytes(b"plate")
    return tmp_path


def test_shape_run_finds_the_shape_template(su, saco):
    assert su._find_template_911(saco).name == "911 BATCH _.xlsx"


def test_plate_run_finds_the_plate_template(su, saco):
    assert su._find_template_911(saco, plate=True).name == \
        "911 PLATE BATCH _.xlsx"


def test_shape_prefix_cannot_grab_the_plate_template(su, tmp_path):
    """Only the plate template present: a SHAPE run must fail loudly, not
    silently set a shape batch up on the plate workbook."""
    (tmp_path / "911 PLATE BATCH _.xlsx").write_bytes(b"plate")
    with pytest.raises(FileNotFoundError, match="911 BATCH"):
        su._find_template_911(tmp_path)


def test_missing_plate_template_names_the_plate_pattern(su, tmp_path):
    (tmp_path / "911 BATCH _.xlsx").write_bytes(b"shape")
    with pytest.raises(FileNotFoundError, match="911 PLATE BATCH"):
        su._find_template_911(tmp_path, plate=True)


# ── MIL SPEC: the packet's own FERROUS flag decides ─────────────────────────
# Real observed flag values: F (STL/HSS/OSS/HY-80/HY-100), N (CRES304/316L,
# CRES 2205, IN625, K-MONEL, CUNI, BRASS), A (AL).
@pytest.mark.parametrize("mil, ferrous, plate, expected", [
    ("MIL-S-22698", "F", True, "N/A"),          # carbon plate -> N/A
    ("MIL-S-22698", "f", True, "N/A"),          # case-tolerant
    ("ASTM-A240", "N", True, "ASTM-A240"),      # stainless plate -> real spec
    ("ASTM-B209", "A", True, "ASTM-B209"),      # aluminum plate -> real spec
    ("MIL-S-22698", None, True, "MIL-S-22698"), # no flag -> keep, human strikes
    ("MIL-S-22698", "F", False, "MIL-S-22698"), # SHAPE untouched, even ferrous
    (None, "F", True, "N/A"),                   # carbon plate, blank spec field
    (None, "N", True, None),                    # nothing to write
])
def test_effective_mil_spec(su, mil, ferrous, plate, expected):
    assert su._effective_mil_spec(mil, ferrous, plate) == expected


def test_ferrous_flag_reads_from_real_move_ticket_layout(su):
    """The flag sits mid-line on the MOVE TICKET ('FERROUS:' then the value
    on the next extracted line) — same labeled-field read as MIL SPEC."""
    text = "LEVEL: N\nSUBSAFE:\nN\nFERROUS:\nF\nPART WT:\n41\n"
    assert su._labeled_value(text, "FERROUS") == "F"


# ── scribe doc: plate form for plate runs ───────────────────────────────────
def test_scribe_doc_filename_per_mode(su):
    assert "SHAPES" in su._scribe_doc_filename(False)
    assert "PLATES" in su._scribe_doc_filename(True)
    assert su._scribe_doc_filename(True) == \
        "QF-QU-15 REV B - SCRIBE VERIFICATION - PLATES.docx"


# ── UNIQUE - TRACE: forecast TRACE/MIC -> plate SCRIBE sheet ────────────────
def _plate_workbook():
    """A workbook with the plate template's SCRIBE header row (PART ID shifts
    UNIQUE - TRACE to col G — the writer must find it by NAME, not position)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SCRIBE VERIFICATION"
    for col, hdr in enumerate(["PART ID", "QTY", "DYPN", "HULL CODE",
                               "MILL - SPEC", "MAT TYPE", "UNIQUE - TRACE"], 1):
        ws.cell(1, col).value = hdr
    return wb


def test_trace_fills_every_part_row_from_one_forecast_row(su):
    """The usual case: one forecast row per nest, several part rows — the
    trace replicates down exactly like _fill_nest_part_rows does for A-E."""
    wb = _plate_workbook()
    rows = [("PO1", 1, "V092", "DL34270")]
    su._fill_scribe_trace(wb, rows, num_parts=3, log=lambda *a: None)
    ws = wb["SCRIBE VERIFICATION"]
    assert [ws.cell(r, 7).value for r in (2, 3, 4)] == ["DL34270"] * 3


def test_trace_keeps_per_row_values_when_forecast_has_several(su):
    wb = _plate_workbook()
    rows = [("PO1", 1, "V092", "DL34270"), ("PO1", 2, "V092", "XL30183")]
    su._fill_scribe_trace(wb, rows, num_parts=2, log=lambda *a: None)
    ws = wb["SCRIBE VERIFICATION"]
    assert ws.cell(2, 7).value == "DL34270"
    assert ws.cell(3, 7).value == "XL30183"


def test_blank_forecast_trace_stays_blank_for_manual_entry(su):
    """The reported hazard is 'N/A' being overlooked — a blank forecast must
    leave the column EMPTY, never invent a value."""
    wb = _plate_workbook()
    rows = [("PO1", 1, "V092", None), ("PO1", 2, "V092", "")]
    su._fill_scribe_trace(wb, rows, num_parts=2, log=lambda *a: None)
    ws = wb["SCRIBE VERIFICATION"]
    assert ws.cell(2, 7).value is None
    assert ws.cell(3, 7).value is None


def test_trace_write_survives_a_workbook_without_the_column(su):
    """An older/edited template: warn and skip, never crash the batch run."""
    wb = openpyxl.Workbook()
    wb.active.title = "SCRIBE VERIFICATION"      # headerless sheet
    logs = []
    su._fill_scribe_trace(wb, [("a", "b", "c", "T1")], 1, logs.append)
    assert any("UNIQUE - TRACE" in m for m in logs)


# ── forecast reader carries the trace column through ────────────────────────
def test_copy_forecast_rows_reads_the_trace_column(su):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, hdr in enumerate(["PO", "Line", "Batch /DR", "NEST",
                               "TRACE/MIC"], 1):
        ws.cell(1, col).value = hdr
    ws.append(["PO9", 4, "V092", "503836", "DL28965"])
    rows = su._copy_forecast_rows(ws, 1, 4, (1, 2, 3, 5), "503836")
    assert rows == [("PO9", 4, "V092", "DL28965")]


def test_copy_forecast_rows_tolerates_a_forecast_without_trace(su):
    """trace_col None (no TRACE/MIC header found) must not break SHAPE runs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, hdr in enumerate(["PO", "Line", "Batch /DR", "NEST"], 1):
        ws.cell(1, col).value = hdr
    ws.append(["PO9", 4, "V092", "503836"])
    rows = su._copy_forecast_rows(ws, 1, 4, (1, 2, 3, None), "503836")
    assert rows == [("PO9", 4, "V092", None)]


# ── the toggle itself: defaults SHAPE, never sticky ─────────────────────────
def test_plate_toggle_declares_no_memory_and_defaults_off(su):
    g = next(x for x in su._dialog_groups() if x["key"] == "plate_batch")
    assert g["checked"] is False          # SHAPE is the default, every run
    assert g["remember"] is False         # and memory may never change that
