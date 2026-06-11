"""
TechDeck Feedback Writer
========================
Appends a feedback row to the shared TechDeck_Suggestions.xlsx workbook in
the OneDrive-synced SharePoint folder. Each Windows user gets the same target
path resolved via Path.home(), so the file lands inside their own OneDrive
cache and syncs up to SharePoint automatically.

Schema (matches the existing workbook):
    Sheet:   "Feedback"
    Cols:    A=Suggestion, B=Which Feature?, C=Date Logged, D=Action Taken?

Submissions write to the next empty row in the "Feedback" sheet. Column D
(Action Taken?) is left blank for the maintainer to fill in during triage.

Concurrent-write conflicts are tolerated: if two users save within the same
OneDrive sync window, OneDrive will create a conflict copy alongside the
original. The maintainer reviews and merges these manually.

If the file is locked (open in Excel), the write retries a few times then
surfaces a clear error.
"""

import logging
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------

# Workbook location relative to the Pilot Program root. The root itself is
# resolved via plugin_sdk.pilot_program_roots(), which tries every OneDrive
# folder-name variant ('American Steel & Alum', 'OneDrive - American Steel &
# Alum', and the ONEDRIVE env var). The old code hardcoded ONE variant under
# Path.home(), which broke on machines where OneDrive named the folder
# differently — the exact problem the SDK resolver exists to solve (and why
# the plugins never hit this).
_FEEDBACK_REL = ("922 QTDR Production Packages", "Notes")
_FEEDBACK_FILENAME = "TechDeck_Suggestions.xlsx"
_FEEDBACK_SHEET_NAME = "Feedback"


def feedback_file_path() -> Optional[Path]:
    """First EXISTING candidate for the shared feedback workbook across all
    OneDrive folder-name variants, or None if it isn't synced anywhere."""
    from techdeck.core import plugin_sdk as sdk
    return sdk.resolve_under_pilot_program(*_FEEDBACK_REL, _FEEDBACK_FILENAME)


def feedback_candidate_paths() -> list:
    """Every location we look for the workbook (for error messages)."""
    from techdeck.core import plugin_sdk as sdk
    return [root.joinpath(*_FEEDBACK_REL, _FEEDBACK_FILENAME)
            for root in sdk.pilot_program_roots()]


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _append_row_with_retry(path: Path, row: list, max_attempts: int = 4,
                           delay_seconds: float = 0.75) -> None:
    """
    Open the workbook, append to the Feedback sheet, save. Retries on
    PermissionError (file locked by another process - usually Excel itself
    or an OS sync). After max_attempts, re-raises the last error so the
    caller can show it to the user.
    """
    last_err: Optional[Exception] = None
    for attempt in range(1, max_attempts + 1):
        try:
            wb = load_workbook(path)
            try:
                if _FEEDBACK_SHEET_NAME not in wb.sheetnames:
                    raise ValueError(
                        f"Sheet '{_FEEDBACK_SHEET_NAME}' not found in "
                        f"{path.name}. Existing sheets: {wb.sheetnames}"
                    )
                ws = wb[_FEEDBACK_SHEET_NAME]
                ws.append(row)
                wb.save(path)
            finally:
                wb.close()
            return
        except PermissionError as e:
            last_err = e
            logger.warning(
                f"Feedback write attempt {attempt}/{max_attempts} failed "
                f"(file locked): {e}"
            )
            time.sleep(delay_seconds)
        except Exception:
            # Non-lock errors aren't worth retrying
            raise

    assert last_err is not None
    raise last_err


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def submit_feedback(suggestion: str, which_feature: str) -> Path:
    """
    Append a feedback row to the shared workbook and return the path that
    was written to.

    Args:
        suggestion: The user's suggestion or bug report text (col A).
        which_feature: Which feature/plugin it relates to (col B).

    The "Date Logged" column (C) is filled with today's date.
    The "Action Taken?" column (D) is left blank for triage.

    Raises:
        FileNotFoundError: if the workbook does not exist (OneDrive not
            synced, or file was renamed/moved).
        PermissionError: if the file is locked and all retries failed.
        ValueError: if the workbook is missing the expected Feedback sheet.
        OSError: for other I/O problems.
    """
    path = feedback_file_path()

    if path is None:
        tried = "\n".join(f"  {p}" for p in feedback_candidate_paths())
        raise FileNotFoundError(
            f"{_FEEDBACK_FILENAME} was not found in any OneDrive location. "
            f"Checked:\n{tried}\n\n"
            "Make sure the 'Communication site - Electric Boat ASA Docs' "
            "SharePoint library is synced to OneDrive on this machine, and "
            "that the workbook exists under "
            "922 QTDR Production Packages\\Notes."
        )

    # Match the existing schema: A=Suggestion, B=Which Feature, C=Date Logged,
    # D=Action Taken? (left blank for the maintainer).
    row = [
        suggestion,
        which_feature,
        datetime.now(),
        None,
    ]

    _append_row_with_retry(path, row)
    logger.info(f"Feedback row appended to {path}")
    return path
