"""
TechDeck Plugin SDK
===================
Shared helpers for plugins. Plugins run inside the TechDeck process (or the
frozen exe, which bundles the techdeck package), so they can import this
module directly:

    from techdeck.core import plugin_sdk as sdk

For standalone CLI testing (`python plugins/foo/run.py`), the repo root may
not be on sys.path. Use this bootstrap at the top of a plugin so both modes
work:

    try:
        from techdeck.core import plugin_sdk as sdk
    except ModuleNotFoundError:
        import sys, pathlib
        sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
        from techdeck.core import plugin_sdk as sdk

This module is the single source of truth for things every plugin used to
reimplement: locating the OneDrive "Pilot Program" roots, parsing a batch
number, finding a batch folder, Excel header lookups, and PDF merge/save.
Fix a bug here once instead of in eight copies.
"""
from __future__ import annotations

import os
import re
import shutil
import tempfile
import time
from pathlib import Path
from typing import Iterable, NamedTuple, Optional


# ─────────────────────────────────────────────────────────────────────────────
# User-facing errors
#
# A failure the USER can fix (a file left open in Excel, a cloud file that
# won't download, an empty/wrong folder) — NOT a code bug. Raising this instead
# of a bare Exception lets the console print a clean "here's the problem, here's
# the fix" in plain English across every plugin, instead of a raw traceback or a
# cryptic errno. Raise via a factory (locked_file_error, ...) or directly:
#
#     raise UserFacingError("The folder has no DXF files in it.",
#                           "Pick the folder that holds the part files.")
# ─────────────────────────────────────────────────────────────────────────────

class UserFacingError(RuntimeError):
    """An error with a plain-language `problem` + `fix` for the user."""

    def __init__(self, problem: str, fix: str = "", *, detail: str = ""):
        self.problem = " ".join(str(problem).split())
        self.fix = " ".join(str(fix).split())
        self.detail = str(detail).strip()  # e.g. the full file path (for logs)
        msg = self.problem + (f"  Fix: {self.fix}" if self.fix else "")
        super().__init__(msg)


def as_user_facing(exc: BaseException) -> "UserFacingError | None":
    """Map an exception to a UserFacingError when it's a KNOWN user-caused
    failure, else None. This is the executor's safety net so every app shows a
    clean message, not a traceback.

    Walks the exception chain (`__cause__`/`__context__`): a plugin that catches
    a UserFacingError and re-raises `RuntimeError(f"...: {e}")` still surfaces
    cleanly, and a raw share-lock (Errno 13) from an un-converted read site is
    recognized too (open()/copy set exc.filename)."""
    seen: set[int] = set()
    cur: BaseException | None = exc
    while cur is not None and id(cur) not in seen:
        seen.add(id(cur))
        if isinstance(cur, UserFacingError):
            return cur
        if _is_share_lock(cur):
            fn = getattr(cur, "filename", None)
            if fn:
                return locked_file_error(fn, cur)
            return UserFacingError(
                "A file TechDeck needs is open in another program (usually Excel).",
                "Close any open batch or pricing files, then run again.",
            )
        cur = cur.__cause__ or cur.__context__
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Cooperative cancellation
#
# The executor sets a threading.Event (run()'s 3rd arg) when the user hits
# Cancel; it CANNOT interrupt a running plugin, so plugins must poll it. Call
# raise_if_cancelled(cancel_event) inside any loop that can run for a while
# (over files, PDFs, DXFs, rows, nests, ...) and Cancel responds promptly
# instead of waiting for the whole batch to finish. The executor maps
# PluginCancelled to a clean CANCELLED status (not an error).
# ─────────────────────────────────────────────────────────────────────────────

class PluginCancelled(BaseException):
    """Raised by raise_if_cancelled() when the user cancels a run. Caught by the
    executor and reported as CANCELLED, never ERROR.

    Subclasses BaseException (like KeyboardInterrupt), NOT Exception, so a
    plugin's own `try/except Exception` around per-file work (used to skip a bad
    file and continue) can never accidentally swallow a cancel."""


def cancelled(cancel_event) -> bool:
    """True if the run has been cancelled. Tolerates cancel_event=None (a plugin
    run standalone / in tests), returning False."""
    return bool(cancel_event is not None and cancel_event.is_set())


def raise_if_cancelled(cancel_event) -> None:
    """Bail out of a long operation the instant the user cancels. Call once per
    iteration of any loop that can run a while (over files/PDFs/DXFs/rows/nests).
    Cheap — a plain flag read — so per-item calls are fine. Raises
    PluginCancelled, which the executor turns into a clean CANCELLED."""
    if cancelled(cancel_event):
        raise PluginCancelled()


# ─────────────────────────────────────────────────────────────────────────────
# OneDrive / Pilot Program root resolution
#
# Every machine syncs the same SharePoint library, but the local cache path
# differs by how OneDrive named the folder. We check every known base so a
# plugin never has to make the user type a directory.
# ─────────────────────────────────────────────────────────────────────────────

_PILOT_REL = (
    "Communication site - Electric Boat ASA Docs",
    "Pilot Program",
)


def pilot_program_roots() -> list[Path]:
    """Candidate 'Pilot Program' directories, in priority order, de-duplicated.

    The fixed, known layouts come first and are NOT existence-checked (callers
    like 911 Setup use [0] as the default for error messages). After those,
    DISCOVERED layouts are appended — globbed from disk, so they exist — to
    cover machines where OneDrive named things differently. Colleagues' fresh
    installs hit exactly this (v0.8.6.1): their sync base wasn't one of the
    two hardcoded names, so every batch app failed to locate the root until
    the directory was set manually per app. Discovery covers:

    * any tenant display-name variant: ``~\\<Tenant>\\<site>\\Pilot Program``
    * a personal-OneDrive folder with any tenant suffix: ``~\\OneDrive - *``
    * "Add shortcut to My files" layouts, where the site folder — or the
      'Pilot Program' folder itself — sits inside the personal OneDrive
    * a OneDrive relocated off the user profile (siblings of ``%OneDrive%``)
    * the parent of any manually-set per-app directory override saved in
      Settings > Apps (a user-confirmed location outranks globbed guesses)
    """
    home = Path.home()
    site, pilot = _PILOT_REL
    od = os.environ.get("ONEDRIVE") or os.environ.get("OneDrive")

    out: list[Path] = []
    seen: set[str] = set()

    def add(root: Path) -> None:
        key = str(root).casefold()
        if key not in seen:
            seen.add(key)
            out.append(root)

    # Known fixed layouts — keep these first so machines that work today keep
    # resolving to the exact same path.
    add(home / "American Steel & Alum" / site / pilot)
    add(home / "OneDrive - American Steel & Alum" / site / pilot)
    if od:
        add(Path(od) / site / pilot)

    # Roots derived from saved per-app directory overrides. On a machine
    # where auto-detect failed, the user manually set some app's 922/911
    # folder in Settings > Apps — that override is ground truth for where the
    # library lives, so every other consumer (all batch apps, the feedback
    # writer) should benefit from it too. Those overrides point at a package
    # root directly under Pilot Program, so the override's PARENT is the
    # Pilot Program root. Keys with other semantics (pdf_directory,
    # output_dir, ...) are deliberately not consulted.
    try:
        from techdeck.core.settings import SettingsManager
        _ROOT_OVERRIDE_KEYS = ("base_path", "base_directory", "qtdr_922_root",
                               "qtdr_911_root", "qtdr_base_path", "forecast_dir")
        per_plugin_settings = SettingsManager().data.get("plugin_settings", {})
        for plugin_conf in per_plugin_settings.values():
            if not isinstance(plugin_conf, dict):
                continue
            for key in _ROOT_OVERRIDE_KEYS:
                raw = str(plugin_conf.get(key) or "").strip()
                if not raw:
                    continue
                parent = Path(raw).expanduser().parent
                if parent.is_dir():
                    add(parent)
    except Exception:
        # Settings must never break root resolution (e.g. plugin_sdk used
        # standalone in CLI testing outside the app).
        pass

    # Discovered layouts. Globs are shallow with fixed tails (no ** walks —
    # cheap even on OneDrive). Discovery must never break resolution, so any
    # filesystem oddity just skips it.
    try:
        containers = [home]
        if od:
            od_parent = Path(od).parent
            if str(od_parent).casefold() != str(home).casefold():
                containers.append(od_parent)
        for container in containers:
            # Any tenant-named sync root (also catches the site folder synced
            # inside a personal 'OneDrive - <tenant>' folder).
            for hit in container.glob(f"*/{site}/{pilot}"):
                if hit.is_dir():
                    add(hit)
        # 'Pilot Program' shortcut sitting directly inside a personal OneDrive
        # ("Add shortcut to My files" on the Pilot Program folder itself).
        personals = list(home.glob("OneDrive - *"))
        if od:
            personals.append(Path(od))
        for base in personals:
            for hit in base.glob(pilot):
                if hit.is_dir():
                    add(hit)
    except OSError:
        pass

    return out


def resolve_under_pilot_program(*parts: str) -> Optional[Path]:
    """Return the first existing '<pilot program root>/<parts...>' across all
    candidate roots, or None if none exist."""
    for root in pilot_program_roots():
        candidate = root.joinpath(*parts)
        if candidate.exists():
            return candidate
    return None


def library_roots(library_tail: str) -> list[Path]:
    """All EXISTING on-disk roots of a SharePoint QTDR library, regardless of how
    OneDrive synced it, de-duplicated. ``library_tail`` is the canonical library
    folder name, e.g. ``"922 QTDR Production Packages"``.

    pilot_program_roots() only covers the WHOLE-SITE sync layout
    (``<site>/Pilot Program/<library>``). But a colleague can instead sync a
    single document library directly, which OneDrive names ``<site> - <library>``
    (e.g. ``Communication site - 922 QTDR Production Packages``) with NO
    'Pilot Program' parent — so the canonical-name lookup misses it, and any
    consumer without a manual override (notably the feedback writer) can't find
    files in it. This covers both layouts plus saved per-app overrides."""
    out: list[Path] = []
    seen: set[str] = set()
    tail_cf = library_tail.casefold()

    def add(p: Path) -> None:
        try:
            if not p.is_dir():
                return
        except OSError:
            return
        key = str(p).casefold()
        if key not in seen:
            seen.add(key)
            out.append(p)

    # 1) Saved per-app base_path overrides whose folder IS this library (the
    #    override value itself is the library root — covers the 'Communication
    #    site - <library>' direct-sync naming a user pointed an app at).
    try:
        from techdeck.core.settings import SettingsManager
        per_plugin = SettingsManager().data.get("plugin_settings", {})
        for conf in per_plugin.values():
            if not isinstance(conf, dict):
                continue
            for key in ("base_path", "base_directory"):
                raw = str(conf.get(key) or "").strip()
                if raw and Path(raw).name.casefold().endswith(tail_cf):
                    add(Path(raw).expanduser())
    except Exception:
        pass

    # 2) Whole-site sync: <pilot program root>/<library>.
    for root in pilot_program_roots():
        add(root / library_tail)

    # 3) Direct single-library sync: '<site> - <library>' under the home dir, a
    #    tenant sync folder, or a personal 'OneDrive - <tenant>'. Shallow globs
    #    with a fixed tail (no ** walks), so cheap even on OneDrive.
    home = Path.home()
    od = os.environ.get("ONEDRIVE") or os.environ.get("OneDrive")
    containers = [home, *home.glob("OneDrive - *")]
    if od:
        containers += [Path(od), Path(od).parent]
    try:
        for container in containers:
            for hit in container.glob(f"*{library_tail}"):      # <container>/<site> - <library>
                add(hit)
            for hit in container.glob(f"*/*{library_tail}"):    # <container>/<tenant>/<site> - <library>
                add(hit)
    except OSError:
        pass

    return out


def _resolve_root(override: str, *parts: str) -> Optional[Path]:
    """Shared override-or-autodiscover logic. A non-empty override always wins
    (returned even if it doesn't exist, so the caller can show a precise
    error). Otherwise auto-discover under the Pilot Program roots, and — for a
    single-segment library — fall back to library_roots() so a directly-synced
    'Communication site - <library>' is found even without a manual override."""
    if override and override.strip():
        return Path(override.strip()).expanduser()
    hit = resolve_under_pilot_program(*parts)
    if hit is not None:
        return hit
    if len(parts) == 1:                       # single library folder
        roots = library_roots(parts[0])
        if roots:
            return roots[0]
    return None


def resolve_922_root(override: str = "") -> Optional[Path]:
    """'922 QTDR Production Packages' root."""
    return _resolve_root(override, "922 QTDR Production Packages")


def resolve_911_qtdr_root(override: str = "") -> Optional[Path]:
    """'911 QTDR' root."""
    return _resolve_root(override, "911 QTDR")


def resolve_forecast_dir(override: str = "") -> Optional[Path]:
    """'Forecast and Inventory Reports' root."""
    return _resolve_root(override, "Forecast and Inventory Reports")


# ─────────────────────────────────────────────────────────────────────────────
# Batch number parsing / batch folder location
# ─────────────────────────────────────────────────────────────────────────────

# Accepts "473", "Batch 473", "PO 473", "PO #473", "#473" (case-insensitive).
_BATCH_INPUT_RE = re.compile(
    r'^(?:(?:batch|po)\s+#?\s*|#\s*)([0-9]+)$', re.IGNORECASE
)


def parse_922_batch(raw: str) -> Optional[str]:
    """Parse a 922 batch number (digits) from free-form input.

    Returns the digit string, or None if nothing usable was entered.
    """
    if raw is None:
        return None
    raw = raw.strip()
    m = _BATCH_INPUT_RE.match(raw)
    if m:
        return m.group(1)
    if re.match(r'^[0-9]+$', raw):
        return raw
    return None


def normalize_911_batch(raw: str) -> str:
    """911 batches are alphanumeric (e.g. 'V060', 'S045'). Just trim + upcase."""
    return (raw or "").strip().upper()


# Hard Rule 3: THE nest-id shape. Legacy numeric nests ([PS]?\d{3,}) plus
# alphanumeric IDs like 5CDAVW (4-8 chars, must contain a digit so footer
# text like "TOTALS" is rejected; first seen GX030, May 2026). This is the
# single home — plugins alias it (`_NEST_RE = sdk.NEST_ID_RE`), never
# re-type it: the pattern has been revised once already (the alphanumeric
# branch), and the next revision must not need six separate edits.
NEST_ID_RE = re.compile(
    r"^(?:[PS]?\d{3,}|(?=[A-Z0-9]*\d)[A-Z0-9]{4,8})$", re.IGNORECASE)


def is_nest_id(text) -> bool:
    """True when `text` is nest-id shaped per Hard Rule 3 (trimmed first)."""
    return bool(NEST_ID_RE.match(str(text or "").strip()))


def load_sibling(plugin_id: str, anchor=None):
    """Import a sibling plugin's run.py as a module and return it.

    Plugins compose — 911 Setup runs the 911 Teams Cards engine in-process,
    both it and Teams Cards import 911 Remove Ticket's stamp helpers, and
    922 Setup runs its stamper/repeater stages — but there was no composition
    primitive: four hand-rolled spec_from_file_location dances, each of them
    a silent breakage point when a plugin folder is renamed. This is the one
    home.

    `anchor` should be the CALLING plugin's ``__file__``: every layout keeps
    all plugins in one flat dir (dev repo tree, %LOCALAPPDATA% installs, a
    TECHDECK_PLUGINS_DIR override), so the sibling always sits beside the
    caller. Without an anchor the dir is resolved the way the app's
    PluginLoader resolves it.

    Raises UserFacingError when the sibling isn't installed; the sibling's
    own import errors propagate untouched — callers keep their existing
    memoize-and-degrade wrappers and decide what a missing sibling means.
    """
    import importlib.util
    if anchor:
        plugins_dir = Path(anchor).resolve().parents[1]
    else:
        from techdeck.core.plugin_loader import PluginLoader
        plugins_dir = PluginLoader._default_plugins_dir()
    run_py = plugins_dir / plugin_id / "run.py"
    if not run_py.is_file():
        raise UserFacingError(
            problem=f"The '{plugin_id}' app isn't installed, and this app "
                    f"needs it.",
            fix="Update TechDeck so the full app set is installed, then run "
                "again.",
            detail=str(run_py),
        )
    spec = importlib.util.spec_from_file_location(
        f"techdeck_sibling_{plugin_id}", str(run_py))
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def find_922_batch_path(root: Path, batch: str) -> Optional[Path]:
    """Locate 'Batch {n}' under the 922 root, checking the live root first and
    then the '1 - Completed' archive. Returns None if not found."""
    live = root / f"Batch {batch}"
    if live.exists():
        return live
    archived = root / "1 - Completed" / f"Batch {batch}"
    if archived.exists():
        return archived
    return None


def find_911_batch_folder(qtdr_root: Path, batch: str) -> Optional[Path]:
    """Locate a 911 batch folder (named exactly like the batch, e.g. 'V060')
    directly under the given root. Case-insensitive. Returns None if absent."""
    exact = qtdr_root / batch
    if exact.exists():
        return exact
    batch_upper = batch.upper()
    if qtdr_root.exists():
        for entry in qtdr_root.iterdir():
            if entry.is_dir() and entry.name.upper() == batch_upper:
                return entry
    return None


def request_922_batch_folder(
    params: dict, base_override: str = ""
) -> Optional[tuple]:
    """Resolve the 922 batch for this run by FOLDER PICK, not typed number.

    The one batch-entry flow for 922 plugins (promoted from 922 LST Organizer
    v3.2.0, which pioneered it; before this, six plugins hand-copied a typed
    request_batch_number + parse + find dance that had drifted three ways).

    Order:
      1. Family cache hit — an earlier 922 plugin in this queued run already
         picked/typed the batch (``shared_state["922"]["batch_number"]``);
         reuse it without prompting.
      2. Folder pick via request_directory, starting at the 922 root — Sentry
         Drone capable (no ``style=`` passed), native dialog otherwise. The
         batch number comes from the picked folder's NAME ("Batch 483").
      3. Seed the family cache so every later 922 plugin (including the
         still-typed Batch Repeater) reuses this answer.

    Returns ``(batch_no, batch_path)``, or ``None`` when the user cancelled
    the pick (the run's cancel flag is already set by then — just return).
    Raises UserFacingError for a missing root, an unrecognizable folder, or a
    cached batch whose folder has vanished.
    """
    log = params.get("log", print)

    root = resolve_922_root((base_override or "").strip())
    if root is None or not root.exists():
        raise UserFacingError(
            "Couldn't find the '922 QTDR Production Packages' folder.",
            "Make sure OneDrive is synced, or set the Base Directory in this "
            "plugin's Settings, then run again.")

    # 1. Family cache — never re-prompt inside one queued run. The bucket is
    # the literal "922" (matching run_session's fixed buckets), NOT
    # params['plugin_family'], so the answer is shared exactly like
    # request_batch_number's.
    shared_state = params.get("shared_state")
    cached = (shared_state or {}).get("922", {}).get("batch_number")
    if cached:
        batch_no = parse_922_batch(str(cached)) or str(cached)
        batch_path = find_922_batch_path(root, batch_no)
        if batch_path is None:
            raise UserFacingError(
                f"Couldn't find 'Batch {batch_no}' under the 922 folder.",
                "Check the batch folder exists (or run the batch's folder "
                "setup first), then run again.")
        log(f"Batch {batch_no} (shared from an earlier plugin)")
        return batch_no, batch_path

    # 2. Folder pick (drone-capable: style defaults to sentry_style(params)).
    raw = request_directory(params, "Select the 922 batch folder", str(root))
    if not raw:
        # request_directory already flagged the run cancelled.
        log("Folder selection cancelled - nothing was run.")
        return None

    batch_path = Path(raw)
    batch_no = parse_922_batch(batch_path.name)
    if not batch_path.is_dir() or not batch_no:
        raise UserFacingError(
            f"'{batch_path.name}' doesn't look like a 922 batch folder.",
            "Run it again and pick the batch's own folder (the one named "
            "like 'Batch 483').")

    # 3. Seed the family cache for the rest of the queued run.
    if shared_state is not None:
        shared_state.setdefault("922", {})["batch_number"] = batch_no
    log(f"Batch {batch_no}: {batch_path}")
    return batch_no, batch_path


# ─────────────────────────────────────────────────────────────────────────────
# Console input
# ─────────────────────────────────────────────────────────────────────────────

def request_text(params: dict, prompt: str) -> str:
    """Prompt the user for a free-form text answer through the TechDeck
    console, falling back to stdin input() when run standalone. Returns the
    raw string.

    Unlike :func:`request_batch_number` this never reads or writes the
    family-shared cache — every call prompts fresh. Use it for prompts whose
    answers should NOT be reused across same-family plugins (e.g. asking the
    user to pick a line, a mode, a filename).
    """
    console = params.get("console")
    if console is not None and hasattr(console, "request_input"):
        return console.request_input(prompt)
    return input(prompt + " ")


def _user_cancelled(params: dict):
    """Record that the USER backed out of a prompt, then return None.

    The executor decides SUCCESS vs CANCELLED from ONE thing: whether
    ``cancel_event`` is set when ``run()`` returns (plugin_executor ~line 559).
    A plugin that closed a dialog and simply ``return``ed therefore reported a
    clean SUCCESS — success chime, tickets awarded, usage logged — for a run
    the user explicitly abandoned (reported 2026-08-10 against 911 Setup's
    master toggle window; six plugins had the identical hole).

    Setting the flag HERE closes the class instead of the instance: every
    cancellable prompt below funnels through this, so no plugin — present or
    future — can forget. Call sites keep their existing
    ``if not choice: return`` shape and now correctly report CANCELLED.

    Deliberately tolerant: a missing/unset ``cancel_event`` (headless, CLI
    test, an older executor) just means nothing to flag, never a crash.
    """
    event = params.get("cancel_event")
    if event is not None:
        try:
            event.set()
        except Exception:
            pass
    return None


def sentry_style(params: dict) -> Optional[str]:
    """The picker style this plugin should use: ``"chopper_gunner"`` when the
    user OWNS the Sentry Drone gadget (Woogy's Emporium) AND has switched it on
    for this app, else None (a normal Explorer dialog).

    Plugins never call this directly — :func:`request_directory` and
    :func:`request_file` apply it automatically, so a picker plugin gets the
    drone just by declaring the ``sentry_drone`` field in its plugin.json
    (see ``techdeck.core.sentry_mode``). It's exported for the GUI-plugin
    helpers below and for anything that needs to branch on it.
    """
    try:
        from techdeck.core import sentry_mode
        return sentry_mode.style_for(params.get("plugin_id") or "")
    except Exception:
        return None   # headless / standalone: no drone, just the plain dialog


def request_directory(params: dict, title: str = "Select Folder",
                      start_dir: str = "", style: Optional[str] = None) -> Optional[str]:
    """Ask the user to pick a folder via the directory dialog (opened on the
    GUI thread by the console), falling back to a pasted-path stdin prompt
    when run standalone. Returns the chosen path string, or None if the user
    cancelled / entered nothing.

    **Sentry Drone is automatic**: leave ``style`` alone and the MW2 kill-cam
    picker is used whenever the user owns the gadget and has enabled it for
    this app (:func:`sentry_style`) — plugins don't compute that themselves.
    Pass ``style=""`` to force the plain dialog for a prompt the drone should
    never touch. The console independently re-checks ownership, downgrades to
    the native dialog under the professional theme or on any effect failure,
    and an older console without the ``style`` parameter is handled here
    (TypeError → plain call), so a new plugin drop on an old TechDeck works.

    Prefer this over making users paste folder paths through request_text —
    per-run folders (a batch folder, a ROOT of orders) should be clicked,
    not typed.
    """
    if style is None:
        style = sentry_style(params)
    console = params.get("console")
    if console is not None and hasattr(console, "request_directory"):
        if style:
            try:
                picked = console.request_directory(title, start_dir, style)
            except TypeError:
                # Older console signature without style (version tolerance).
                picked = console.request_directory(title, start_dir)
        else:
            picked = console.request_directory(title, start_dir)
        return picked or _user_cancelled(params)
    raw = input(f"{title} (paste path): ").strip().strip('"')
    return raw or _user_cancelled(params)


def request_file(params: dict, title: str = "Select File", start_dir: str = "",
                 name_filter: str = "", style: Optional[str] = None) -> Optional[str]:
    """Ask the user to pick a FILE via the open-file dialog (opened on the GUI
    thread by the console; same blocking contract as request_directory),
    falling back to a pasted-path stdin prompt when run standalone or on an
    older TechDeck without the console method. Returns the chosen path string,
    or None if the user cancelled.

    ``name_filter`` is a QFileDialog filter, e.g. ``"DXF files (*.dxf)"``.
    Sentry Drone applies automatically, exactly as in request_directory —
    pass ``style=""`` to force the plain dialog.
    """
    if style is None:
        style = sentry_style(params)
    console = params.get("console")
    if console is not None and hasattr(console, "request_file"):
        if style:
            try:
                picked = console.request_file(title, start_dir, name_filter, style)
            except TypeError:
                # Older console signature without style (version tolerance).
                picked = console.request_file(title, start_dir, name_filter)
        else:
            picked = console.request_file(title, start_dir, name_filter)
        return picked or _user_cancelled(params)
    raw = input(f"{title} (paste path): ").strip().strip('"')
    return raw or _user_cancelled(params)


def pick_directory_gui(params: dict, title: str = "Select Folder",
                       start_dir: str = "", parent=None) -> Optional[str]:
    """Folder pick for a GUI plugin (``requires_main_thread: true``), where the
    console's blocking worker-thread helpers can't be used.

    Same Sentry Drone contract as :func:`request_directory` — the kill-cam
    picker when the gadget is owned + enabled for this app, otherwise (and on
    the professional theme, or any effect failure) the plain Explorer dialog.
    Call it on the GUI thread; returns the chosen path or None on cancel.
    """
    return _pick_gui(params, title, start_dir, parent, name_filter=None)


def pick_file_gui(params: dict, title: str = "Select File", start_dir: str = "",
                  name_filter: str = "", parent=None) -> Optional[str]:
    """File pick for a GUI plugin — :func:`pick_directory_gui` for files."""
    return _pick_gui(params, title, start_dir, parent,
                     name_filter=name_filter or "All Files (*.*)")


def _pick_gui(params: dict, title: str, start_dir: str, parent,
              name_filter: Optional[str]) -> Optional[str]:
    """Shared body of the two GUI-plugin pickers: try the drone, then fall back
    to the native dialog. ANY failure inside the effect falls back — the toy
    must never block a run."""
    from PySide6.QtWidgets import QFileDialog

    if sentry_style(params):
        try:
            from techdeck.core.settings import SettingsManager
            if not SettingsManager().is_professional():
                from techdeck.ui.widgets import chopper_picker
                if name_filter is None:
                    picked = chopper_picker.pick_folder_chopper(
                        parent, title, start_dir)
                else:
                    picked = chopper_picker.pick_file_chopper(
                        parent, title, start_dir, name_filter)
                return picked or _user_cancelled(params)
        except Exception:
            pass   # fall through to the native dialog
    if name_filter is None:
        return (QFileDialog.getExistingDirectory(parent, title, start_dir)
                or _user_cancelled(params))
    path, _selected = QFileDialog.getOpenFileName(parent, title, start_dir,
                                                  name_filter)
    return path or _user_cancelled(params)


def request_choice(params: dict, title: str, prompt: str,
                   options) -> Optional[str]:
    """Ask the user to pick ONE option via a modal popup with a button per
    option (GUI thread; same blocking contract as request_directory).
    Returns the chosen option string, or None on cancel. Headless, or on an
    older TechDeck without the console method, it falls back to a numbered
    stdin prompt. Use for a small pick-one decision that deserves a dialog
    instead of an in-console prompt (first used by DXF Offset Tool's
    single-file-vs-folder question).
    """
    options = [str(o) for o in options]
    console = params.get("console")
    if console is not None and hasattr(console, "request_choice"):
        return console.request_choice(title, prompt, options) or _user_cancelled(params)
    menu = "  ".join(f"[{i + 1}] {o}" for i, o in enumerate(options))
    raw = input(f"{prompt} {menu} > ").strip()
    if raw.isdigit() and 1 <= int(raw) <= len(options):
        return options[int(raw) - 1]
    for o in options:
        if raw.lower() == o.lower():
            return o
    return _user_cancelled(params)


def request_selection(params: dict, items, done_items=None,
                      **kwargs) -> Optional[list]:
    """Ask the user to tick which of ``items`` to run (SelectionDialog), and
    block until submit. Returns the chosen items, or None if they cancelled.

    ``items`` are display strings; ``done_items`` are flagged "already done".
    ``kwargs`` pass through to the dialog (window_title, header, root_label,
    noun, done_label, prompt_note, subhead_prefix, run_button_text,
    disabled_items, disabled_label, ...).

    Headless — or on an older TechDeck whose console lacks the method — every
    item is returned, so a scripted run selects everything instead of hanging.
    Note the difference from a CANCEL: no dialog means "run it all", an empty
    submit means "run nothing", and only a cancel flags the run.

    Prefer this over calling ``console.request_selection`` directly: three
    plugins did, each re-implementing the hasattr fallback, and each having to
    remember to set ``cancel_event`` by hand — one of them (Baked Beans)
    didn't, so cancelling its process picker scored a ticket-earning success.
    This routes through :func:`_user_cancelled` like every other prompt.
    """
    console = params.get("console")
    if console is not None and hasattr(console, "request_selection"):
        picked = console.request_selection(list(items), done_items, **kwargs)
        if picked is None:
            return _user_cancelled(params)
        return list(picked)
    return list(items)


def request_nest_targets(params: dict, title: str, start_dir: str = "",
                         target_pattern: Optional[str] = None):
    """Two-phase Sentry-Drone folder pick (911 Batch Repeater): the user
    TARGETs a batch folder, the drone optics zoom inside it, then multi-locks
    subfolder targets and EXECUTEs a strike across them. Returns a tuple
    ``(status, batch_path, names)``:

    * ``"ok"`` — picked; names are the locked subfolder names, in lock order.
    * ``"cancelled"`` — the user backed out; treat it as a run cancel.
    * ``"unavailable"`` — headless, an older TechDeck without the console
      method, the professional theme, the Sentry Drone gadget not owned /
      not enabled for this app, or any picker failure. Fall back to a classic
      folder dialog + your own selection window — never block on this.

    ``target_pattern`` (regex, case-insensitive) limits which subfolder names
    can be locked in phase 2 (e.g. the nest-number regex, Hard Rule 3).
    """
    if not sentry_style(params):
        return ("unavailable", "", [])
    console = params.get("console")
    if console is None or not hasattr(console, "request_target_folders"):
        return ("unavailable", "", [])
    try:
        status, path, names = console.request_target_folders(
            title, start_dir, target_pattern)
        return (status, path, list(names))
    except Exception:
        return ("unavailable", "", [])


def show_warning(params: dict, title: str, text: str) -> None:
    """Show a modal warning popup (console GUI thread) and block until the
    user acknowledges it. Use for problems the user must see before moving
    on — a console log line scrolls past unnoticed. Headless (or on an older
    TechDeck without console.show_warning) it just logs the text."""
    console = params.get("console")
    if console is not None and hasattr(console, "show_warning"):
        try:
            console.show_warning(title, text)
            return
        except Exception:
            pass
    log = params.get("log", print)
    log(f"WARNING [{title}]: {text}")


def show_report(params: dict, title: str, subtitle: str, body: str,
                save_path: str = "") -> str:
    """Put a finished report on screen with a Save-as-.txt button. Returns a note
    for the log.

    For a run whose OUTPUT is something a person reads. Writing a .txt into a
    Pilot Program folder and printing the path means the reader has to go and dig
    for it; this puts it in front of them and makes saving their choice.

    Does not block -- the window stays up after the run ends (contrast
    show_warning, which blocks until acknowledged). ``save_path`` is the full file
    path the Save button writes, so pass one INSIDE the folder the run was pointed
    at. Headless, or on an older TechDeck with no console.show_report, it falls
    back to writing the file itself so nothing is ever lost.
    """
    console = params.get("console")
    if console is not None and hasattr(console, "show_report"):
        try:
            console.show_report(title, subtitle, body, save_path)
            return "Report is on screen - press Save as .txt to keep a copy."
        except Exception:
            pass
    if not save_path:
        params.get("log", print)(body)
        return "Report printed above."
    with open(long_path(save_path), "w", encoding="utf-8") as fh:
        fh.write(body)
    return "Saved: %s" % save_path


def load_toggle_memory(memory_key: str) -> dict:
    """The selections last submitted under ``memory_key`` ({} if none/unreadable)."""
    try:
        from techdeck.core.settings import SettingsManager
        return dict(SettingsManager().get_toggle_memory(memory_key) or {})
    except Exception:
        return {}


def save_toggle_memory(memory_key: str, result: dict) -> None:
    """Remember a GroupedToggleDialog result. Never raises — failing to
    remember a preference must not sink the run that just succeeded."""
    try:
        from techdeck.core.settings import SettingsManager
        SettingsManager().set_toggle_memory(memory_key, result)
    except Exception:
        pass


def apply_toggle_memory(groups: list, remembered: dict) -> list:
    """``groups`` with each toggle's ``checked`` replaced by what the user last
    submitted. Pure — returns a new list, mutating nothing.

    The merge is by KEY, in one direction only, and that is what makes the
    memory survive an app update:

      * key in groups AND remembered -> the remembered state wins
      * key in groups, NOT remembered -> its DECLARED default. A toggle added
        by an update ships at whatever default it declares; it never inherits
        a state the user was never shown.
      * key remembered, NOT in groups -> dropped. Removing or renaming a
        toggle retires its memory instead of resurrecting a dead key.

    A child marked ``disabled`` stays unchecked no matter what is remembered —
    disabled means unavailable, not merely off.

    A group declaring ``"remember": False`` is NON-STICKY: it always opens at
    its declared default, no matter what was submitted last time (and
    request_grouped_toggles never saves it). For mode switches that are a
    property of THIS run, not a preference — 911 Setup's PLATE batch toggle is
    the canonical case: a leftover PLATE tick silently applied to the next
    shape batch would fill real paperwork with the wrong template.
    """
    out = []
    for g in groups:
        saved = {} if g.get("remember") is False else \
            (remembered.get(g.get("key")) or {})
        merged = dict(g)
        if "enabled" in saved:
            merged["checked"] = bool(saved["enabled"])
        saved_opts = saved.get("options") or {}
        children = []
        for c in g.get("children") or []:
            child = dict(c)
            if not child.get("disabled") and child.get("key") in saved_opts:
                child["checked"] = bool(saved_opts[child["key"]])
            children.append(child)
        merged["children"] = children
        out.append(merged)
    return out


def request_grouped_toggles(params: dict, groups: list, remember_as: str = "",
                            **kwargs) -> Optional[dict]:
    """Show the two-level master toggle dialog (stages as checkable parents,
    per-stage option toggles as children — GroupedToggleDialog) and block
    until submit. ``groups`` spec and the returned
    ``{group_key: {"enabled": bool, "options": {child_key: bool}}}`` shape are
    documented in techdeck/ui/dialogs/grouped_toggle_dialog.py; ``kwargs``
    pass through (window_title, header, subtext, run_button_text). Returns
    None if the user cancelled.

    ``remember_as`` (normally the plugin id) makes the dialog STICKY: it opens
    with whatever was last submitted under that key and saves the new answer on
    submit, so a user who always runs the same subset sets it up once. Stored
    in settings.json, which an app update leaves alone — see
    SettingsManager.get_toggle_memory and apply_toggle_memory for the merge
    rules that keep it sane across added/removed toggles. A cancel saves
    nothing. Omit it and the dialog is stateless, exactly as before.

    Headless — or on an older TechDeck whose console lacks the method — it
    returns every group/child at its declared default (an all-defaults
    submit), so scripted runs never hang on a dialog. Remembered state is
    applied there too, so a scripted run matches what the GUI would have
    offered.
    """
    if remember_as:
        remembered = load_toggle_memory(remember_as)
        if remembered:
            groups = apply_toggle_memory(groups, remembered)

    console = params.get("console")
    if console is not None and hasattr(console, "request_grouped_toggles"):
        try:
            result = console.request_grouped_toggles(groups, **kwargs)
        except TypeError:
            # Older console signature without these kwargs (version tolerance).
            result = console.request_grouped_toggles(groups)
    else:
        result = {g["key"]: {"enabled": bool(g.get("checked", True)),
                             "options": {c["key"]: (not c.get("disabled")
                                                    and bool(c.get("checked", True)))
                                         for c in g.get("children", [])}}
                  for g in groups}

    if result is None:
        # Closing the master window abandons the whole run -- flag it so the
        # executor reports CANCELLED instead of chiming success and paying out
        # tickets for work nobody asked for (2026-08-10).
        return _user_cancelled(params)

    # Only a real submit is remembered -- a cancel means "not this time",
    # never "make that my default". Groups declaring remember: False are
    # per-run mode switches and are stripped before saving, so their state
    # can never leak into a later run even through a stale settings.json.
    if remember_as:
        no_memory = {g.get("key") for g in groups
                     if g.get("remember") is False}
        save_toggle_memory(remember_as,
                           {k: v for k, v in result.items()
                            if k not in no_memory})
    return result


def plugin_settings(plugin_id: str) -> dict:
    """The saved Settings > Apps values for ANY plugin id (empty dict when
    none are stored or the store is unreadable). Lets an orchestrating plugin
    run a sibling stage with the sibling's own configuration — the same values
    the executor would inject if that plugin ran standalone."""
    try:
        from techdeck.core.settings import SettingsManager
        return dict(SettingsManager().get_plugin_settings(plugin_id) or {})
    except Exception:
        return {}


def set_ticket_units(params: dict, units: int) -> None:
    """Report how many ticket-earning systems this run performed (default 1).

    An orchestrating plugin that runs N sibling stages inside ONE plugin run
    (e.g. 922 Setup running Teams Cards + Batch Repeater + Pallet Stamper)
    calls this with N so the shell awards N x TICKETS_PER_RUN on success
    instead of a single run's worth. Values below 1 clamp to 1; on an older
    TechDeck (executor unaware of the key) the extra units are simply ignored
    and the run pays the normal single-run tickets."""
    try:
        params["ticket_units"] = max(1, int(units))
    except (TypeError, ValueError):
        pass


# Run outcomes a plugin may report via set_run_outcome. Absence of a report
# (the default) is treated as a clean success — so existing plugins that simply
# return are unaffected.
RUN_OUTCOME_WARNING = "warning"   # ran to completion, but with issues to surface
RUN_OUTCOME_PARTIAL = "partial"   # some requested work was skipped or failed


def set_run_outcome(params: dict, status: str, message: str = "") -> None:
    """Report that this run finished WITHOUT raising but not cleanly, so
    TechDeck stops claiming a blank 'completed successfully'.

    Use for the middle ground the old all-or-nothing SUCCESS hid: a Teams
    webhook POST returned non-2xx, a nest had no forecast rows, one of several
    stages was skipped. Pass status = RUN_OUTCOME_WARNING or RUN_OUTCOME_PARTIAL
    and a short plain-English `message`; TechDeck shows it (⚠) instead of the ✅
    line, but the run still counts and still earns tickets — the work happened.

    A hard failure should still RAISE (that path stays ERROR). On an older
    TechDeck unaware of the key the run simply reports the normal success."""
    try:
        params["run_outcome"] = {
            "status": str(status),
            "message": str(message or ""),
        }
    except (TypeError, ValueError):
        pass


def request_batch_number(params: dict, prompt: str) -> str:
    """Prompt for a batch number through the TechDeck console, falling back to
    stdin input() when run standalone (no console). Returns the raw string.

    Family-shared caching: when this plugin runs as part of a same-family
    multi-plugin batch (Home page "Run Selected" with several 911s or several
    922s), the first plugin's answer is cached in
    ``params['shared_state'][family]['batch_number']`` and reused silently by
    subsequent same-family plugins in the same run. Cross-family runs and
    single-plugin runs prompt as before.
    """
    shared_state = params.get("shared_state")
    family = params.get("plugin_family")
    log = params.get("log") if callable(params.get("log")) else None

    if shared_state is not None and family:
        cached = shared_state.get(family, {}).get("batch_number")
        if cached:
            if log:
                log(f"[shared] Using batch number {cached} from prior {family} "
                    f"plugin in this run.")
            return cached

    console = params.get("console")
    if console is not None and hasattr(console, "request_input"):
        answer = console.request_input(prompt)
    else:
        answer = input(prompt + " ")

    if shared_state is not None and family and answer:
        shared_state.setdefault(family, {})["batch_number"] = answer
    return answer


# ─────────────────────────────────────────────────────────────────────────────
# Excel header helpers (openpyxl worksheets)
#
# Columns shift between batches/files, so always look up by header NAME. These
# replace the per-plugin header scanners.
# ─────────────────────────────────────────────────────────────────────────────

def find_header_col(ws, header_name: str, header_row: int) -> Optional[int]:
    """1-based column index in `header_row` whose value matches header_name
    (case-insensitive, stripped), or None."""
    target = header_name.strip().upper()
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is not None and str(val).strip().upper() == target:
            return col
    return None


def header_map(ws, header_row: int) -> dict[str, int]:
    """{UPPERCASE HEADER: 1-based col} for every string cell in `header_row`."""
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if isinstance(val, str) and val.strip():
            out[val.strip().upper()] = col
    return out


def header_col(hdr: dict, name: str) -> Optional[int]:
    """Column for `name` in a header_map, tolerating a header that has GROWN.

    Exact match first, then the leftmost header that STARTS WITH `name`. Use
    this instead of `hdr.get("NAME")` for any header a person maintains and
    might append to.

    Why it exists: the EB 922 Schedule's rating column was renamed `RATING` →
    `RATING/PC COUNT`, and both difficulty readers looked it up exactly. They
    silently found nothing — every Teams card posted with no difficulty label
    and every packet stamped none, for weeks, behind a dismissable warning
    (2026-08-11). Header names are lookup KEYS but they are also human text,
    and humans add words to them.
    """
    target = name.strip().upper()
    if target in hdr:
        return hdr[target]
    matches = [col for key, col in hdr.items() if key.startswith(target)]
    return min(matches) if matches else None


def find_header_row(ws, required: Iterable[str], max_scan: int = 30,
                    prefix_ok: bool = False):
    """Scan the first `max_scan` rows for the row containing all `required`
    header names. Returns (row_index, header_map) or (None, {}).

    Matching is EXACT by default. Pass ``prefix_ok=True`` when a required
    header is one a person maintains and may have appended to — then `RATING`
    also satisfies `RATING/PC COUNT`. It stays opt-in because loosening it
    everywhere would let a required `QTY` latch onto `QTY SHIPPED` on some
    other sheet and silently read the wrong column. Pair it with
    :func:`header_col` to pull those columns back out.
    """
    required_upper = [r.strip().upper() for r in required]
    limit = min(ws.max_row, max_scan)
    for r in range(1, limit + 1):
        row_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                row_map[v.strip().upper()] = c
        if prefix_ok:
            ok = all(header_col(row_map, name) is not None
                     for name in required_upper)
        else:
            ok = all(name in row_map for name in required_upper)
        if ok:
            return r, row_map
    return None, {}


def visible_sheetnames(wb) -> list:
    """Sheet names whose state is 'visible', in workbook order.

    Use this instead of wb.sheetnames whenever SCANNING a workbook for a
    data sheet — by name pattern or by headers. Pricing/PO workbooks often
    carry hidden staging sheets (FORMING, PRICING, template copies) whose
    headers also satisfy the scan, and iterating wb.sheetnames silently
    reads one of those instead of the real sheet (bit 902 DXF Prep: batch
    4130's tab was 'PO- 4130', the name regex missed it, and the fallback
    scan read the hidden 3-row FORMING sheet). Exact-name lookups don't
    need this — sheet names are unique whether hidden or not.
    """
    return [n for n in wb.sheetnames if wb[n].sheet_state == "visible"]


# ─────────────────────────────────────────────────────────────────────────────
# 922 shop-print trees (CAD-AND-SHOP-PRINTS / 7000)
#
# One 922 order folder's shop prints live at
#   {order}\CAD-AND-SHOP-PRINTS\{DYPN}[\{assembly}\{DYPN}]\7000\*.lst
# The depth VARIES (an assembly adds a level), and only TUBE parts get a `7000`
# folder at all — so never assume a fixed depth, and never read "no 7000 folder"
# as an error on its own. Shared by 922_lst_organizer (which pulls the files)
# and 922_batch_repeater (which only audits a pulled repeat for them).
# ─────────────────────────────────────────────────────────────────────────────

CAD_PRINTS_DIR_NAME = "CAD-AND-SHOP-PRINTS"
_CAD_PRINTS_RE = re.compile(r"^cad[-_ ]*and[-_ ]*shop[-_ ]*prints$", re.IGNORECASE)


class ShopPrintScan(NamedTuple):
    """What one order folder's shop-print tree holds.

    Every field describes the SAME tree — either the folder's own, or (only
    when its own yields no .lst at all) its nested `REPEAT\\` fallback. They
    are consistent with each other by construction, so an audit can count
    `lsts` against `seven_k` without inventing a phantom gap.

    lsts        — the .lst files to pull
    cad_dirs    — the CAD-AND-SHOP-PRINTS folder(s) (empty = nothing to pull)
    seven_k     — every 7000 folder under them
    empty_7000  — the 7000 folders that hold NO .lst (a tube part missing its
                  cut file — actionable, unlike having no 7000 folder at all)
    from_nested — True when the answers came from a nested `REPEAT\\` copy
    """
    lsts: list
    cad_dirs: list
    seven_k: list
    empty_7000: list
    from_nested: bool = False

    @property
    def has_cad(self) -> bool:
        return bool(self.cad_dirs)


def scan_shop_print_lsts(order_dir, cancel_event=None) -> ShopPrintScan:
    """Walk ONE order folder's CAD-AND-SHOP-PRINTS tree for its tube `.lst`
    files, plus the structure around them (see ShopPrintScan).

    **A nested `REPEAT\\` subfolder is a FALLBACK, never an addition.** A
    pulled repeat often carries its own prior repeat inside it; those files
    are the batch-before-last's copy, so they are neither pulled nor counted
    while the folder's own tree has anything — and the whole scan switches to
    them only when it doesn't. Counting them alongside is what makes a folder
    read as "1 .lst in 2 7000 folders" and look like it's missing one (real:
    Batch 484's FK345540-H5223262-H48).

    The nested test runs on the path RELATIVE to `order_dir`, so pointing this
    AT a folder inside `REPEAT BATCHES` (the Batch Repeater's audit) scans it
    normally instead of writing off every file as a repeat.

    Cancel is polled inside the walk (Hard Rule 11).
    """
    order_dir = Path(order_dir)
    own: dict = {"cad": [], "seven_k": [], "empty": [], "lsts": []}
    nested: dict = {"cad": [], "seven_k": [], "empty": [], "lsts": []}
    for i, p in enumerate(order_dir.rglob("*")):
        if cancel_event is not None and i % 64 == 0 and cancel_event.is_set():
            break
        if not p.is_dir():
            continue
        is_cad = bool(_CAD_PRINTS_RE.match(p.name.strip()))
        if not is_cad and p.name.strip() != "7000":
            continue
        try:
            rel = [s.lower() for s in p.relative_to(order_dir).parts]
        except ValueError:                      # pragma: no cover - defensive
            rel = [s.lower() for s in p.parts]
        bucket = nested if any("repeat" in s for s in rel) else own
        if is_cad:
            bucket["cad"].append(p)
            continue
        if not any(_CAD_PRINTS_RE.match(s) for s in rel):
            continue                            # a stray '7000' outside the tree
        bucket["seven_k"].append(p)
        found = sorted(f for f in p.glob("*.lst") if f.is_file())
        if found:
            bucket["lsts"].extend(found)
        else:
            bucket["empty"].append(p)
    # One switch for the whole scan, so every field describes the same tree.
    # `cad_dirs` follows it too: a nested CAD folder counts only when the
    # nested copy is the one actually supplying the .lst files — otherwise a
    # folder whose only prints are buried in a stale REPEAT\ still reads as
    # having nothing to give this batch.
    src, from_nested = (nested, True) if not own["lsts"] and nested["lsts"] \
        else (own, False)
    return ShopPrintScan(src["lsts"], src["cad"], src["seven_k"],
                         src["empty"], from_nested)


# ─────────────────────────────────────────────────────────────────────────────
# DYPN part numbers (the shared join key between a PO row and a file on disk)
#
# A DYPN's last dashed segment is a piece number with an OPTIONAL trailing
# revision letter: 'R6455461-H51-4' and 'R6455461-H51-4A' are the SAME piece.
# EB renames it between the PO, the model and the SigmaNest export, so a batch's
# PO can read '-4' while the shop file on disk reads '-4A-STEP' (or the exact
# reverse). Joining on the exact string alone reports ONE physical piece TWICE —
# the file looks foreign ("needs review") and the PO line looks un-pulled
# ("missing") — which is what bit 922 LST Organizer on Batch 485.
#
# So: normalize both sides with normalize_dypn(), try the exact key first, and
# fall back to match_dypn_variant() BEFORE declaring anything missing.
# ─────────────────────────────────────────────────────────────────────────────

# Export tags that make a filename stem differ from its PO DYPN. Order matters
# (tube tag, then STEP, then a duplicated suffix).
_DYPN_TUBE_TAG_RE = re.compile(r"[-_]+P[-_]+Tube\d*$", re.IGNORECASE)  # SigmaNest per-tube export
_DYPN_STEP_RE = re.compile(r"[-_]STEP$", re.IGNORECASE)
_DYPN_DUP_SUFFIX_RE = re.compile(r"([-_])([A-Za-z0-9]+)[-_]\2$", re.IGNORECASE)  # -4A_4A -> -4A
_DYPN_SEG_RE = re.compile(r"^(\d+)([A-Za-z]*)$")


def normalize_dypn(value) -> str:
    """A filename stem (or a raw PO DYPN) reduced to its canonical DYPN,
    upper-cased: the export tags ('-P-Tube2', '-STEP') and a duplicated suffix
    ('-4A_4A') are peeled off. 'R6455461-H51-4A-STEP' -> 'R6455461-H51-4A'."""
    s = str(value or "")
    s = _DYPN_TUBE_TAG_RE.sub("", s)
    s = _DYPN_STEP_RE.sub("", s)
    s = _DYPN_DUP_SUFFIX_RE.sub(r"\1\2", s)
    return s.strip("-_ ").upper()


def split_dypn(part) -> tuple:
    """(head, piece number, revision letter) off a DYPN's last dashed segment:
    'H5222069-H88-4A' -> ('H5222069-H88', '4', 'A'); '...-4' -> (..., '4', '').
    A non-numeric last segment comes back whole as (head, segment, '')."""
    p = str(part or "").strip().upper()
    if "-" not in p:
        return p, "", ""
    head, seg = p.rsplit("-", 1)
    m = _DYPN_SEG_RE.match(seg)
    if not m:
        return head, seg, ""
    return head, m.group(1), m.group(2)


def match_dypn_variant(part, candidates, prefer=None) -> Optional[str]:
    """The one candidate that is `part` under a different trailing revision
    letter ('...-4' vs '...-4A'), in either direction — or None.

    Deliberately conservative, because a wrong hit files a part against the
    wrong PO line: it answers only when the answer is unambiguous.
      * a candidate whose letter matches exactly wins outright;
      * otherwise `prefer` (a predicate, e.g. "same order number") narrows the
        field, and is ignored when it would eliminate everything;
      * more than one survivor -> None, so the caller flags it for a human.

    Candidates are compared after normalize_dypn() but returned verbatim, so
    the caller can use the result as a key back into its own mapping.
    """
    head, num, letter = split_dypn(normalize_dypn(part))
    if not num.isdigit():
        return None                      # no piece number -> nothing to vary
    hits = []
    for c in candidates:
        c_head, c_num, c_letter = split_dypn(normalize_dypn(c))
        if c_head == head and c_num == num:
            hits.append((c, c_letter))
    if not hits:
        return None
    exact = [c for c, c_letter in hits if c_letter == letter]
    if len(exact) == 1:
        return exact[0]
    names = [c for c, _ in hits]
    if len(names) > 1 and prefer is not None:
        narrowed = [c for c in names if prefer(c)]
        if narrowed:
            names = narrowed
    return names[0] if len(names) == 1 else None


# ─────────────────────────────────────────────────────────────────────────────
# 922 tube materials + PO (QF-QU-09) reading
#
# Source-material serials classify a part's stock. STANDARD tubes are
# laser-cuttable (we expect 7000 shop-print folders for orders that contain
# them). OVERSIZED tubes are too thick for the laser (>0.375" NOM) and are cut
# elsewhere, so they do NOT imply a 7000 folder. These sets are the single
# source of truth, shared by 922_lst_organizer and batch_auditor.
# ─────────────────────────────────────────────────────────────────────────────

STANDARD_TUBE_MATERIALS = {
    '218002867', '218003095', '218004492', '218012302', '218019939',
    '218019941', '218021954', '218026206', '218026875', '218026962',
    '218033112', '218136140', '40-00-2020', '40-07-1003',
}

OVERSIZED_TUBE_MATERIALS = {
    '181361440', '218000414', '218001209', '218001975', '218002101',
    '218002191', '218002642', '218002778', '218004273', '218012555',
    '218017859', '218018574', '218019943', '218020020', '218020119',
    '218026338', '218026500', '218032982',
}

ALL_TUBE_MATERIALS = STANDARD_TUBE_MATERIALS | OVERSIZED_TUBE_MATERIALS


def tube_class(serial) -> Optional[str]:
    """Classify a source-material serial: 'standard' (laser tube), 'oversized'
    (too thick for the laser), or None (not a tracked tube material)."""
    if serial is None:
        return None
    s = str(serial).strip()
    if s in STANDARD_TUBE_MATERIALS:
        return "standard"
    if s in OVERSIZED_TUBE_MATERIALS:
        return "oversized"
    return None


def read_qf_qu_09(po_path: Path):
    """Read a 922 'PO H{n} QF-QU-09 REV C' workbook's PO sheet.

    Returns (order_to_serials, dypn_to_order_serial):
      order_to_serials       -> {order: set(source_material_serial, ...)}
      dypn_to_order_serial   -> {dypn: (order, serial)}

    Serials are stripped strings, ready to pass to tube_class(). Header row is
    located by name (ORDER / DYPN / SOURCE MATERIAL), never by position.
    """
    order_to_serials: dict[str, set] = {}
    dypn_to_order_serial: dict[str, tuple] = {}

    wb = load_workbook_resilient(po_path, data_only=True)
    try:
        ws = wb['PO'] if 'PO' in wb.sheetnames else wb.active
        header_row, cols = find_header_row(ws, ['ORDER', 'DYPN', 'SOURCE MATERIAL'])
        if header_row is None:
            return order_to_serials, dypn_to_order_serial

        c_order, c_dypn, c_src = cols['ORDER'], cols['DYPN'], cols['SOURCE MATERIAL']
        for r in range(header_row + 1, ws.max_row + 1):
            def _v(c):
                v = ws.cell(r, c).value
                return str(v).strip() if v not in (None, '') else None
            order, dypn, serial = _v(c_order), _v(c_dypn), _v(c_src)
            if order:
                bucket = order_to_serials.setdefault(order, set())
                if serial:
                    bucket.add(serial)
            if order and dypn:
                dypn_to_order_serial[dypn] = (order, serial)
        return order_to_serials, dypn_to_order_serial
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# PDF helpers (PyMuPDF / fitz)
# ─────────────────────────────────────────────────────────────────────────────

# A part DRAWING carries the standard title-block boilerplate (the tolerance
# block ASA's and Electric Boat's prints share); a 922 work packet never does.
# Measured over EVERY PDF in the live 922 batch folders 2026-08-27: all four
# markers appear on 224 of 224 drawing binders and on 0 of 257 work packets.
# The packets answer instead to their own "Lead Trade" / "Work Instruction"
# text, but the absence of a title block is the cleaner test - it needs no
# guess about which packet layout a batch happens to use.
_TITLE_BLOCK_MARKERS = (
    "DO NOT SCALE DRAWING",
    "UNLESS OTHERWISE SPECIFIED",
    "ENG APPR",
    "MFG APPR",
)
# Two of four. One marker alone could plausibly appear in a packet's notes; all
# four is stricter than a scanned or oddly-encoded print can be relied on to be.
_TITLE_BLOCK_MIN_HITS = 2

# Office lock files (~$...) and stray Python temp files (tmpab12cd34.pdf) are
# not real PDFs of ours. A leftover temp PDF in an order folder would otherwise
# be picked as that order's work packet by "the first PDF in the folder".
_NOT_A_REAL_PDF_RE = re.compile(r'^(~\$|tmp[a-z0-9_]{6,}$)', re.IGNORECASE)

# Drawing binders are exported as Binder1.pdf / Binder7.pdf / BinderH9.pdf.
# This is only a RANKING hint - the title-block read is what actually decides.
_BINDER_NAME_RE = re.compile(r'^binder', re.IGNORECASE)


def is_real_pdf(path) -> bool:
    """False for Office lock files and leftover temp PDFs, which are never ours."""
    return not _NOT_A_REAL_PDF_RE.match(Path(path).stem)


def is_drawing_pdf(path, log=None) -> bool:
    """True when page 1 of `path` is a part DRAWING (a title-block print).

    This is how a drawing binder is told apart from a work packet. Name alone
    cannot do it: the binders sort BEFORE the packet in an order folder
    ('Binder1.pdf' before 'BK394153 NOFORN.pdf' in Windows' case-insensitive
    order), so "the first PDF in the folder" stamps the drawing (reported by a
    922 Pallet Stamper user, 2026-08-20), and a binder is not always named
    'Binder'.

    An UNREADABLE PDF answers False on purpose. Treating a read error as "this
    is a drawing" would silently skip a real work packet and the run would look
    clean; answering False sends the caller on to its own open, which reports
    the failure properly.
    """
    import fitz
    path = Path(path)
    try:
        ensure_local(path)
        doc = fitz.open(long_path(path))
        try:
            text = doc[0].get_text().upper()
        finally:
            doc.close()
    except Exception as exc:
        if log:
            log(f"  NOTE: couldn't read {path.name} to check for a title "
                f"block ({exc}); treating it as a work packet.")
        return False
    return text_has_title_block(text)


def text_has_title_block(page_text: str) -> bool:
    """The title-block test on text you have ALREADY read.

    `is_drawing_pdf` opens the file for you; use this one when you are inside
    an open document anyway and a second open would cost another OneDrive
    round-trip.
    """
    text = (page_text or "").upper()
    hits = sum(1 for marker in _TITLE_BLOCK_MARKERS if marker in text)
    return hits >= _TITLE_BLOCK_MIN_HITS


def find_work_packet(order_dir, log=None) -> Optional[Path]:
    """The work-packet PDF inside a 922 order folder ('{ORDER}-{PPN}').

    Returns None when the folder holds no work packet - INCLUDING the case
    where every PDF in it is a drawing binder. Callers must treat that as
    "nothing to do here", never as "stamp the first PDF anyway".

    Candidates are ranked by NAME (a PDF named for the order first, a
    Binder*.pdf last) and then each is confirmed by READING it: the first one
    whose page 1 is not a title-block drawing wins. The ranking only decides
    which file gets read first, so a binder that isn't named 'Binder' is still
    caught, and the common case costs a single page-1 read.
    """
    order_dir = Path(order_dir)
    pdfs = [p for p in sorted(order_dir.iterdir())
            if is_file(p) and p.suffix.lower() == ".pdf" and is_real_pdf(p)]
    if not pdfs:
        return None

    order_no = order_dir.name.split('-', 1)[0].strip().upper()

    def rank(p: Path) -> tuple:
        named_for_order = p.stem.strip().upper().startswith(order_no)
        looks_like_binder = bool(_BINDER_NAME_RE.match(p.stem))
        return (0 if named_for_order else 1, 1 if looks_like_binder else 0, p.name)

    for pdf in sorted(pdfs, key=rank):
        if not is_drawing_pdf(pdf, log=log):
            return pdf
        if log:
            log(f"  Skipping {pdf.name} - it's a drawing (ASA title block), "
                f"not a work packet.")
    return None


def merge_pdfs(pdfs: list[Path], out_path: Path) -> None:
    """Merge `pdfs` (in order) into a single PDF at out_path."""
    import fitz
    out = fitz.open()
    try:
        for p in pdfs:
            ensure_local(p)
            src = fitz.open(long_path(p))
            try:
                out.insert_pdf(src)
            finally:
                src.close()
        ensure_dir(Path(out_path).parent)
        out.save(long_path(out_path))
    finally:
        out.close()


def save_pdf_atomic(doc, dest_path: Path, close: bool = True) -> None:
    """Save an open fitz document and atomically replace dest_path.

    fitz cannot save in place (especially after redactions), so write to a temp
    file beside the target and os.replace it over the destination.

    ORDER MATTERS ON WINDOWS. An open fitz document keeps a handle on the file
    it was opened from, so replacing that same path while the doc is still open
    fails with ``PermissionError [WinError 5] Access is denied`` — the classic
    in-place-edit case (open a PDF, stamp it, save it back) this helper exists
    to serve. The document is therefore CLOSED before the replace, and `close`
    defaults to True: on return, `doc` is closed and must not be used or closed
    again (fitz raises ``ValueError: document closed`` on a second close).

    Pass ``close=False`` only when `doc` was built from scratch or opened from a
    DIFFERENT file than `dest_path` and you still need it afterwards — if the
    doc's own source path IS `dest_path`, it is closed anyway, because leaving
    it open cannot work.

    A failed save or replace never leaves a stray ``tmp*.pdf`` behind: partial
    temp files are cleaned up on the way out. (They are genuinely dangerous
    here — a leftover temp PDF in a 922 order folder gets picked up as that
    order's work packet by anything that takes "the first PDF in the folder".)
    """
    import fitz
    dest_path = Path(dest_path)
    ensure_dir(dest_path.parent)

    with tempfile.NamedTemporaryFile(
        suffix=".pdf", delete=False, dir=long_path(dest_path.parent)
    ) as tmp:
        tmp_path = Path(tmp.name)

    try:
        doc.save(long_path(tmp_path), incremental=False,
                 encryption=fitz.PDF_ENCRYPT_NONE)

        # Release the source handle before swapping the file underneath it.
        if close or _doc_source_is(doc, dest_path):
            try:
                doc.close()
            except Exception:
                pass

        os.replace(long_path(tmp_path), long_path(dest_path))
    except Exception:
        try:
            os.remove(long_path(tmp_path))
        except OSError:
            pass
        raise


def _doc_source_is(doc, dest_path: Path) -> bool:
    """True when `doc` was opened from `dest_path` (so it holds that handle)."""
    try:
        src = getattr(doc, "name", None)
        if not src:
            return False
        # doc.name carries whatever string opened it - possibly the \?\ form,
        # which Path.resolve() leaves in place. Strip it so a long-path doc
        # still matches its own destination.
        return _strip_long_prefix(src) == _strip_long_prefix(dest_path)
    except Exception:
        return False


# ─────────────────────────────────────────────────────────────────────────────
# Windows long paths (MAX_PATH)
#
# Win32 caps a path at 260 characters unless it is handed over in the
# extended-length form ``\\?\C:\...``. The Pilot Program tree eats most of
# that budget before a plugin adds anything:
#
#   C:\Users\<user>\American Steel & Alum\Communication site - Electric Boat
#   ASA Docs\Pilot Program\1. 1000129724 - Strategic Partnership, Steel
#   Processing Offload\VTDX Award Records\1000129724 SSPO Award 13\...
#
# ...is already ~190 chars, so a descriptive output filename tips it over and
# the write dies with ``FileNotFoundError [Errno 2] No such file or
# directory`` — pointing at a folder that plainly exists. Bit 911 SSPO Award
# Review at 277 chars (DESKTOP-DD35L5F, 2026-08-21); the user "fixed" it by
# renaming the award folder to something shorter, which is not a fix.
#
# Windows 10+ CAN lift the cap process-wide (registry LongPathsEnabled=1 AND a
# longPathAware manifest), but the registry half belongs to corporate IT and is
# off on some machines — so we never rely on it. `long_path()` is the fix that
# works everywhere: prefix the path, and Win32 skips the check entirely.
#
# Gotchas the prefix brings with it, all handled below:
#   * It disables ALL path normalization — the path must be absolute, use
#     backslashes, and contain no "." / ".." components.
#   * UNC shares take a different spelling: \\server\share -> \\?\UNC\server\share
#   * os.makedirs() walks up to the root and chokes on the "\\?" head, so
#     `ensure_dir()` prefixes only the final mkdir of each level.
# ─────────────────────────────────────────────────────────────────────────────

# Prefix well short of 260: the caller may append a suffix (".tmp", a temp
# name beside the target), and a directory at 250 still has to hold filenames.
_LONG_PATH_THRESHOLD = 240
B_SLASH = chr(92)
_UNC_PREFIX = "\\\\?\\UNC\\"
_DOS_PREFIX = "\\\\?\\"


def long_path(path) -> str:
    """Return `path` as a string Windows will accept past the 260-char cap.

    Short paths come back unchanged (the ``\\\\?\\`` form is stricter — no
    relative segments, no forward slashes — so we only reach for it when the
    length actually calls for it). Non-Windows returns the path untouched.

    Wrap the path at the moment of the call, not in the variable:

        wb.save(sdk.long_path(out_path))          # openpyxl / zipfile
        doc.save(sdk.long_path(tmp))              # fitz
        shutil.copy2(sdk.long_path(src), sdk.long_path(dest))

    The SDK helpers (`save_workbook`, `copy_resilient`, `save_pdf_atomic`,
    `merge_pdfs`, the resilient loaders, `ensure_dir`) already do this, so
    prefer them over raw calls — see Hard Rule 14.
    """
    text = str(path)
    if os.name != "nt" or text.startswith(_DOS_PREFIX):
        return text
    if len(text) < _LONG_PATH_THRESHOLD:
        return text
    # The prefix bypasses normalization, so normalize FIRST: absolute,
    # backslashes, no "." / ".." left in it.
    text = os.path.abspath(text)
    if text.startswith("\\\\"):          # UNC share: \\server\share\...
        return _UNC_PREFIX + text[2:]
    return _DOS_PREFIX + text


def _strip_long_prefix(path) -> str:
    """The inverse of `long_path` for comparisons: a resolved, normalized path
    with any extended-length prefix removed, so a string that went through
    `long_path` still compares equal to one that did not."""
    text = str(path)
    if text.startswith(_UNC_PREFIX):
        text = B_SLASH * 2 + text[len(_UNC_PREFIX):]
    elif text.startswith(_DOS_PREFIX):
        text = text[len(_DOS_PREFIX):]
    try:
        text = str(Path(text).resolve())
    except OSError:
        text = os.path.abspath(text)
    text = text.rstrip(B_SLASH + "/")
    return text.lower() if os.name == "nt" else text


def ensure_dir(path) -> Path:
    """makedirs(exist_ok=True) that survives paths past 260 characters.

    ``os.makedirs`` recurses toward the drive root, so handing it a prefixed
    path makes it try to create ``\\\\?`` and raise WinError 123. Instead we
    walk DOWN the missing levels and prefix each individual mkdir. Returns the
    directory as a Path (unprefixed, for normal use afterwards).
    """
    path = Path(path)
    missing = []
    probe = path
    while not os.path.isdir(long_path(probe)):
        missing.append(probe)
        parent = probe.parent
        if parent == probe:              # hit the drive root
            break
        probe = parent
    for level in reversed(missing):
        try:
            os.mkdir(long_path(level))
        except FileExistsError:
            pass
    return path
def exists(path) -> bool:
    """`Path.exists()` that still tells the truth past 260 characters.

    This one fails SILENTLY, which makes it the nastiest member of the
    MAX_PATH family: Win32 refuses the over-length path, Python swallows the
    OSError, and `exists()` answers False for a file that is sitting right
    there. Every `if not dest.exists(): <create it>` guard then fires on a
    file it should have skipped — so the deeper the folder, the more likely a
    plugin quietly overwrites work (922 Setup's "REV C is never overwritten"
    promise rides on exactly that guard).

    Use `sdk.exists` / `sdk.is_file` / `sdk.is_dir` for anything under the
    OneDrive tree (Hard Rule 14)."""
    return os.path.exists(long_path(path))


def is_file(path) -> bool:
    """`Path.is_file()`, long-path safe. See `exists`."""
    return os.path.isfile(long_path(path))


def is_dir(path) -> bool:
    """`Path.is_dir()`, long-path safe. See `exists`."""
    return os.path.isdir(long_path(path))





def save_workbook(wb, path, log=None) -> Path:
    """Save an openpyxl Workbook to `path`, long-path safe and lock-aware.

    Three things every plugin workbook write needs and kept forgetting:
      * the parent folder exists (`ensure_dir`),
      * the path survives MAX_PATH (`long_path`) — otherwise openpyxl's
        zipfile write dies with a baffling "No such file or directory" on a
        folder that is right there,
      * a destination the user has open in Excel raises the clear
        `locked_file_error` instruction instead of a raw Errno 13.

    Use this for EVERY `wb.save(...)` that lands under the OneDrive tree
    (Hard Rule 14). Returns the destination Path.
    """
    path = Path(path)
    ensure_dir(path.parent)
    try:
        wb.save(long_path(path))
    except Exception as exc:
        if _is_share_lock(exc):
            raise locked_file_error(path, exc) from exc
        raise
    if log:
        log(f"Saved: {path.name}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# OneDrive cloud-placeholder resilience
#
# On a Files-On-Demand tree, a file can EXIST (metadata is local — .exists(),
# globs, and size checks all pass) while its CONTENT is cloud-only. If the
# OneDrive client can't download it on demand, reads fail — openpyxl/zipfile
# raise OSError [Errno 22] "Invalid argument" on a workbook that is plainly
# sitting right there. Bit FormingFinder on a freshly-synced machine (the
# Batch 481 PO workbook, v0.8.6.1 debug report NRAPINI-LT).
# ─────────────────────────────────────────────────────────────────────────────

_CLOUD_ATTRS = (
    0x00001000  # FILE_ATTRIBUTE_OFFLINE
    | 0x00040000  # FILE_ATTRIBUTE_RECALL_ON_OPEN
    | 0x00400000  # FILE_ATTRIBUTE_RECALL_ON_DATA_ACCESS
)


def is_cloud_placeholder(path) -> bool:
    """True if `path` is a OneDrive cloud-only / recall-on-access placeholder."""
    try:
        attrs = os.stat(long_path(path)).st_file_attributes
    except (OSError, AttributeError):
        return False
    return bool(attrs & _CLOUD_ATTRS)


def hydrate_cloud_file(path, log=None, attempts: int = 3, delay: float = 1.5,
                       should_stop=None) -> None:
    """Force OneDrive to download a cloud-only file by reading it end to end.

    Raises RuntimeError with user-actionable instructions if the content
    still can't be pulled (OneDrive client stopped, signed out, or offline).

    `should_stop` (optional zero-arg callable) is polled between 1 MiB chunks;
    when it returns True the read stops and the function returns quietly —
    used by the background prefetcher so a cancelled run doesn't keep pulling
    a half-downloaded file. Foreground callers leave it unset.
    """
    path = Path(path)
    last_err = None
    for attempt in range(1, attempts + 1):
        try:
            with open(long_path(path), "rb") as fh:
                while fh.read(1024 * 1024):
                    if should_stop is not None and should_stop():
                        return
            return
        except OSError as exc:
            last_err = exc
            if log:
                log(f"OneDrive download attempt {attempt}/{attempts} failed: {exc}")
            time.sleep(delay)
    raise UserFacingError(
        problem=f"'{path.name}' is stored only in the cloud and Windows "
                f"couldn't download it.",
        fix="Make sure OneDrive is running and signed in, then try again. To "
            "avoid this, right-click the folder in File Explorer and choose "
            "'Always keep on this device'.",
        detail=str(path),
    ) from last_err


def ensure_local(path, log=None) -> None:
    """Make sure `path`'s CONTENT is on disk before a library reads it.

    No-op for normal files; for a OneDrive cloud-only placeholder it forces
    the download (with retries) or raises a user-actionable RuntimeError.
    Call immediately before fitz.open / PdfReader / pd.read_excel / text
    reads of anything under the OneDrive tree (Hard Rule 13)."""
    if is_cloud_placeholder(path):
        if log:
            log(f"'{Path(path).name}' is cloud-only - asking OneDrive to download it...")
        hydrate_cloud_file(path, log=log)


class PrefetchHandle:
    """Handle returned by prefetch_paths(). stop() ends the background work;
    the counters are informational (hydrated = files actually downloaded)."""

    def __init__(self):
        import threading
        self._stop = threading.Event()
        self._threads: list = []
        self.checked = 0
        self.hydrated = 0

    def stop(self) -> None:
        """Ask the background workers to finish up. Returns immediately."""
        self._stop.set()

    def done(self) -> bool:
        return all(not t.is_alive() for t in self._threads)


def prefetch_paths(paths, cancel_event=None, max_workers: int = 3) -> PrefetchHandle:
    """Hydrate OneDrive cloud-only files in the BACKGROUND, best-effort.

    The streaming trick: start this the moment a folder pick tells you which
    files the run will read, and the downloads overlap the prompts the user is
    still answering (and the run's own serial work). By the time the read
    loops arrive, content is local and ensure_local is a single stat call.

    - `paths` may be any iterable of paths; a lazy generator such as
      `folder.rglob('*.pdf')` is ideal — the directory walk itself then also
      happens off-thread.
    - Never raises and never blocks the caller. A file that fails to hydrate
      is simply left for the foreground read to handle exactly as today, with
      the resilient loaders' retries and user-facing messages (Hard Rule 13
      call sites keep their ensure_local either way).
    - Stops early when the plugin's `cancel_event` is set or handle.stop() is
      called — checked between files and between 1 MiB chunks of each file.
    - `max_workers` stays small on purpose: hydration is a full-content
      download through the OneDrive client, and flooding the sync engine just
      moves the queue around.
    """
    import queue
    import threading

    handle = PrefetchHandle()

    def _stopping() -> bool:
        return handle._stop.is_set() or (cancel_event is not None
                                         and cancel_event.is_set())

    q: "queue.Queue" = queue.Queue(maxsize=max_workers * 2)
    _SENTINEL = object()

    def _put(item) -> bool:
        # Bounded put that can't wedge the producer thread forever if the
        # workers have already bailed out after a stop/cancel.
        while True:
            try:
                q.put(item, timeout=0.5)
                return True
            except queue.Full:
                if _stopping():
                    return False

    def _producer():
        try:
            for p in paths:
                if _stopping() or not _put(p):
                    break
        except Exception:
            pass  # a dying generator must not kill the app
        finally:
            for _ in range(max_workers):
                if not _put(_SENTINEL):
                    break

    def _worker():
        while True:
            item = q.get()
            if item is _SENTINEL or _stopping():
                break
            handle.checked += 1
            try:
                if is_cloud_placeholder(item):
                    # Single attempt, no retries: this is opportunistic. The
                    # foreground read keeps the full retry + error messaging.
                    hydrate_cloud_file(item, attempts=1, should_stop=_stopping)
                    handle.hydrated += 1
            except Exception:
                pass

    threads = [threading.Thread(target=_producer, daemon=True,
                                name="techdeck-prefetch-scan")]
    threads += [threading.Thread(target=_worker, daemon=True,
                                 name=f"techdeck-prefetch-{i}")
                for i in range(max_workers)]
    handle._threads = threads
    for t in threads:
        t.start()
    return handle


def locked_file_error(path, cause: Exception | None = None) -> "UserFacingError":
    """A UserFacingError for a file Windows won't let us read because another
    program holds it open — PermissionError / [Errno 13].

    This is a DIFFERENT failure from the cloud-only placeholder (Errno 22 /
    BadZipFile, handled by ensure_local): here the file's content is on disk,
    but Excel/Acrobat has it open, or a OneDrive co-author locked it. No
    amount of retrying or copy-to-temp helps — the first opener chose the
    share mode — so the only fix is for the user to close the file. Bit
    902 DXF Prep when the pricing .xlsm was open in Excel (Errno 13,
    CPENG_TOWERPC, v0.8.6.8)."""
    path = Path(path)
    return UserFacingError(
        problem=f"'{path.name}' is open in another program (usually Excel), so "
                f"TechDeck can't read it.",
        fix=f"Close the file everywhere it's open, then run again. If it stays "
            f"stuck, delete the '~${path.stem}' file in that folder.",
        detail=str(path),
    )


def _is_share_lock(exc: Exception) -> bool:
    """True for a Windows sharing-violation / permission-denied read
    (PermissionError, or any OSError with errno 13). NOT a cloud placeholder
    (that's errno 22)."""
    return isinstance(exc, PermissionError) or getattr(exc, "errno", None) == 13


def _resilient_read(op, path, log=None):
    """The ONE body behind the resilient Excel loaders (they were three
    hand-copied 15-line twins that had already drifted into three spellings
    of the same guard).

    Flow: hydrate a cloud-only placeholder up front (sequential read — the
    recall path OneDrive actually honors; zipfile's backwards seek can stall
    the recall for 60s+ and then fail), then run `op`. On failure:
      * PermissionError / Errno 13 — the file is open in Excel; retrying
        can't help, so raise locked_file_error's clear instruction.
      * The placeholder signature — OSError [Errno 22], or BadZipFile
        (zipfile's end-of-central-directory reader CATCHES that OSError and
        re-raises it as "File is not a zip file", so the failure wears two
        disguises) — while the file still stats as a placeholder: hydrate
        and retry `op` once.
      * Anything else re-raises untouched.
    """
    import zipfile
    ensure_local(path, log=log)
    try:
        return op()
    except Exception as exc:
        if _is_share_lock(exc):
            raise locked_file_error(path, exc) from exc
        if not isinstance(exc, (OSError, zipfile.BadZipFile)) \
                or not is_cloud_placeholder(path):
            raise
        if log:
            log(f"'{Path(path).name}' is cloud-only - asking OneDrive to download it...")
        hydrate_cloud_file(path, log=log)
        return op()


def load_workbook_resilient(path, log=None, **kwargs):
    """openpyxl.load_workbook that survives OneDrive cloud-only placeholders
    AND turns a locked-open workbook into a clear instruction (not a crash).

    The placeholder failure wears two disguises — OSError [Errno 22], and
    zipfile.BadZipFile ("File is not a zip file"; bit 911 Setup that way,
    v0.8.6.5, LAPTOP-1AMLBK7B). A workbook the user still has open in Excel
    fails with PermissionError [Errno 13] instead — surfaced as
    `locked_file_error` so they know to close it (bit 902 DXF Prep,
    v0.8.6.8). Mechanics in `_resilient_read`. Use this for ANY workbook
    under the OneDrive tree (Hard Rule 13)."""
    import openpyxl
    return _resilient_read(
        lambda: openpyxl.load_workbook(long_path(path), **kwargs), path, log=log)


def read_excel_resilient(path, log=None, **kwargs):
    """pandas.read_excel that survives OneDrive placeholders and reports a
    locked-open workbook clearly — the pandas sibling of
    `load_workbook_resilient`. Pass the same args you'd give pd.read_excel
    (sheet_name, header, usecols, ...). Use for ANY user-editable workbook
    read via pandas (Hard Rule 13)."""
    import pandas as pd
    return _resilient_read(
        lambda: pd.read_excel(long_path(path), **kwargs), path, log=log)


def open_excel_resilient(path, log=None):
    """pandas.ExcelFile opened resiliently (placeholder hydrate + locked-open
    message). Returns a pd.ExcelFile whose sheets you then read with
    pd.read_excel(xls, sheet_name=...) — the file lock bites at open time, so
    the returned handle reads sheets without re-touching the lock."""
    import pandas as pd
    return _resilient_read(lambda: pd.ExcelFile(long_path(path)), path, log=log)


def copy_resilient(src, dest, log=None):
    """shutil.copy2 that hydrates a cloud-only source first and reports a
    locked-open file clearly instead of crashing with a raw Errno 13.

    The copy-to-temp pattern (stage a live workbook, then read/edit the copy so
    Excel never sees the synced path) still fails at the COPY if the user has
    the source open in Excel — same PermissionError class as a direct read.
    On a lock we name the actual culprit: usually the source, but on a
    write-back it's the destination. Returns the destination Path."""
    src, dest = Path(src), Path(dest)
    ensure_local(src, log=log)
    try:
        ensure_dir(dest.parent)
        shutil.copy2(long_path(src), long_path(dest))
        return dest
    except Exception as exc:
        if not _is_share_lock(exc):
            raise
        # Which side is locked? If we can't even read the source, it's the
        # source; otherwise the open handle is on the destination.
        culprit = dest
        try:
            with open(long_path(src), "rb") as fh:
                fh.read(1)
        except OSError:
            culprit = src
        raise locked_file_error(culprit, exc) from exc


# ─────────────────────────────────────────────────────────────────────────────
# Power Automate webhook helpers
#
# One JSON POST + one preview-file writer, shared by every plugin that talks
# to a Power Automate flow (922 Setup card creation, 922 Batch Repeater
# REPEAT tagging). Centralized per the Fix Protocol so logging, timeouts and
# error wording stay identical across callers.
# ─────────────────────────────────────────────────────────────────────────────

def post_webhook(url: str, payload: dict, log) -> bool:
    """POST `payload` as JSON to a Power Automate webhook. Returns True on a
    2xx response; logs the failure (and the flow's response body, truncated)
    otherwise. `requests` is bundled (the updater uses it) and respects
    corporate proxies."""
    import requests

    try:
        resp = requests.post(url, json=payload, timeout=60)
    except requests.exceptions.RequestException as exc:
        log(f"ERROR: could not reach the webhook: {exc}")
        log("Check the URL in Settings, your network/VPN, and that the flow is on.")
        return False

    if 200 <= resp.status_code < 300:
        log(f"Webhook accepted the request (HTTP {resp.status_code}).")
        body = (resp.text or "").strip()
        if body:
            log(f"Response: {body[:500]}")
        return True

    log(f"ERROR: webhook returned HTTP {resp.status_code}.")
    log(f"Response: {(resp.text or '')[:500]}")
    return False


def write_payload_preview(payload: dict, filename: str, log) -> None:
    """Save a webhook payload to %LOCALAPPDATA%/TechDeck/<filename> so a dry
    run leaves something inspectable (and replayable into the flow's test
    pane). Failures only log — a preview must never kill the run."""
    import json as _json
    import os
    base = os.environ.get("LOCALAPPDATA") or str(Path.home())
    out = Path(base) / "TechDeck" / filename
    try:
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", encoding="utf-8") as fh:
            _json.dump(payload, fh, indent=2)
        log(f"Preview written to: {out}")
    except OSError as exc:
        log(f"(Could not write preview file: {exc})")
