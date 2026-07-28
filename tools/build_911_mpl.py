r"""
Build the 911 MASTER PARTS LIST (one-time bootstrap).
=====================================================
Compiles `911 MASTER PARTS LIST.xlsx` — the master catalog the 911 Batch
Repeater (v3+) looks repeat parts up in — from the EB 922 Schedule workbook's
snapshot of completed nests.

The schedule is ONLY used here, at compile time; the plugin never reads it.
How the MPL gets refreshed with newly completed nests is TBD (for now, re-run
this tool with a newer schedule export, or append rows by hand).

Sources:
  - CURRENT PIPELINE sheet: dept 911 rows whose STATUS (col F) is exactly
    IN OFFICE / PARTIALLY IN SHOP / IN SHOP / QA (case/space-insensitive;
    'IN OFFICE/HOLD' deliberately does NOT count).
  - SHIPPED sheet: every dept 911 row (status recorded as SHIPPED).

For each completed nest, scans `911 QTDR\{batch}\{nest}\CAD-AND-SHOP-PRINTS\`
and writes one MPL row per part folder:
  DYPN | Batch | Nest | Status | Source Folder (relative to the 911 QTDR
  root, so the list works on every machine's OneDrive layout) | Files

Usage:
  python tools/build_911_mpl.py [--schedule PATH] [--out PATH] [--dry-run]

Defaults: schedule = ~/Downloads/EB 922 Schedule.xlsx, out = the REPEATER
folder under the auto-discovered 911 QTDR root.
"""

import argparse
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from techdeck.core import plugin_sdk as sdk  # noqa: E402

MPL_FILENAME = "911 MASTER PARTS LIST.xlsx"
DONE_STATUSES = {"IN OFFICE", "PARTIALLY IN SHOP", "IN SHOP", "QA"}


def _collect_completed_nests(schedule_path: Path) -> list[tuple[str, str, str]]:
    """[(batch, nest, status)] from the schedule; SHIPPED sheet included."""
    wb = openpyxl.load_workbook(schedule_path, data_only=True)
    out: list[tuple[str, str, str]] = []
    seen: set[tuple[str, str]] = set()

    def add(bn: str, status: str):
        parts = bn.split()
        if len(parts) != 2:
            print(f"  SKIP odd BATCH / NEST value: {bn!r}")
            return
        key = (parts[0].upper(), parts[1].upper())
        if key in seen:
            return
        seen.add(key)
        out.append((parts[0], parts[1], status))

    ws = wb["CURRENT PIPELINE"]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() != "911":
            continue
        bn = str(ws.cell(r, 2).value or "").strip()
        st = str(ws.cell(r, 6).value or "").strip().upper()
        if bn and st in DONE_STATUSES:  # exact match: 'IN OFFICE/HOLD' excluded
            add(bn, st)

    ws = wb["SHIPPED"]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() != "911":
            continue
        bn = str(ws.cell(r, 2).value or "").strip()
        if bn:
            add(bn, "SHIPPED")

    wb.close()
    return out


def build_rows(qtdr_root: Path, nests: list[tuple[str, str, str]]) -> list[tuple]:
    rows = []
    no_folder = []
    no_cad = []
    for batch, nest, status in nests:
        nest_dir = qtdr_root / batch / nest
        if not nest_dir.is_dir():
            no_folder.append(f"{batch} {nest}")
            continue
        cad = nest_dir / "CAD-AND-SHOP-PRINTS"
        if not cad.is_dir():
            no_cad.append(f"{batch} {nest}")
            continue
        for part_dir in sorted(cad.iterdir()):
            if not part_dir.is_dir():
                continue
            files = sorted({p.suffix.lower().lstrip(".")
                            for p in part_dir.iterdir() if p.is_file()})
            rel = part_dir.relative_to(qtdr_root).as_posix()
            rows.append((part_dir.name, batch, nest, status, rel, ", ".join(files)))

    if no_folder:
        print(f"  WARNING: {len(no_folder)} nest(s) with no folder: {', '.join(no_folder)}")
    if no_cad:
        print(f"  WARNING: {len(no_cad)} nest(s) without CAD-AND-SHOP-PRINTS: {', '.join(no_cad)}")

    rows.sort(key=lambda r: (r[0].upper(), r[1].upper(), r[2].upper()))
    return rows


def write_mpl(rows: list[tuple], out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MASTER PARTS"

    headers = ["DYPN", "BATCH", "NEST", "STATUS", "SOURCE FOLDER", "FILES"]
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FFF2CC")  # soft yellow, matches the 911 sheets
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)

    widths = [18, 10, 12, 18, 60, 22]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("--schedule",
                    default=str(Path.home() / "Downloads" / "EB 922 Schedule.xlsx"))
    ap.add_argument("--out", default="",
                    help=f"Output xlsx (default: REPEATER folder \\ {MPL_FILENAME})")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    schedule = Path(args.schedule)
    if not schedule.exists():
        sys.exit(f"Schedule not found: {schedule}")

    qtdr = sdk.resolve_911_qtdr_root()
    if qtdr is None:
        sys.exit("Could not auto-discover the 911 QTDR root (is OneDrive synced?)")
    print(f"911 QTDR root : {qtdr}")

    out_path = Path(args.out) if args.out else (
        qtdr / "04 - Notes - Protocols - Tutorials" / "TECH SERVICES"
        / "REPEATER" / MPL_FILENAME)

    nests = _collect_completed_nests(schedule)
    print(f"Completed nests in schedule: {len(nests)} "
          f"({sum(1 for *_, s in nests if s == 'SHIPPED')} from the SHIPPED sheet)")

    rows = build_rows(qtdr, nests)
    dyp_counts = Counter(r[0].upper() for r in rows)
    dups = {d: n for d, n in dyp_counts.items() if n > 1}
    print(f"Parts catalogued: {len(rows)} ({len(dyp_counts)} unique DYPNs"
          f"{f', {len(dups)} made in more than one nest' if dups else ''})")
    for d, n in sorted(dups.items()):
        print(f"  {d}: {n} nests")

    if args.dry_run:
        print("Dry run - nothing written.")
        return

    write_mpl(rows, out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
