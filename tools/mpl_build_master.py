"""Phase-1 build: merge the per-batch extraction JSONs into the MASTER PARTS
catalog (repeat counts + alternate-suffix alignment + part typing) and write
the sheet.

Modes:
  --preview <xlsx>  write the sheet into a fresh standalone workbook (review)
  --write           write into the live 922 MPL.xlsx (backup first, then a
                    full value-diff of every pre-existing sheet)

Usage:
  python tools/mpl_build_master.py --extract-dir <dir> [--report <json>]
         [--preview <xlsx>] [--write] [--root <922 root>]
"""

import argparse
import datetime as _dt
import json
import sys
from collections import Counter
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "plugins" / "922_batch_repeater"))

import master_parts as mp  # noqa: E402
from techdeck.core import plugin_sdk as sdk  # noqa: E402

QUOTE_REL = Path("2 - Planning") / "EB 922 H# Quote.xlsx"
MPL_NAME = "922 MPL.xlsx"


def load_catalog(extract_dir: Path, today: str):
    cat = mp.Catalog()
    tallies = Counter()
    batches_ok = []
    problems = []
    files = sorted(extract_dir.glob("batch_*.json"),
                   key=lambda p: int(p.stem.split("_")[1]))
    for f in files:
        d = json.loads(f.read_text(encoding="utf-8"))
        tallies[d["status"]] += 1
        if d["status"] != "ok":
            problems.append({"batch": d["batch"], "status": d["status"],
                             "error": d.get("error")})
            continue
        batches_ok.append(d["batch"])
        # occurrences: one per (WO, PPN), WOs in first-appearance (sheet) order
        occ = {}
        for r in d["rows"]:
            occ.setdefault((r["order"], r["ppn"]), []).append(r)
        for (wo, ppn), rows in occ.items():
            cat.merge_occurrence(ppn, d["batch"], wo, rows, today)
    return cat, tallies, batches_ok, problems


def verify_untouched_sheets(before_path: Path, after_path: Path, sheets):
    """Every pre-existing sheet must survive an openpyxl re-save value-
    identical. Returns list of difference strings (empty = clean)."""
    import pandas as pd
    diffs = []
    for s in sheets:
        a = pd.read_excel(before_path, sheet_name=s, header=None)
        b = pd.read_excel(after_path, sheet_name=s, header=None)
        if a.shape != b.shape:
            diffs.append(f"{s}: shape {a.shape} -> {b.shape}")
            continue
        neq = (a.fillna("~na~") != b.fillna("~na~"))
        n = int(neq.values.sum())
        if n:
            diffs.append(f"{s}: {n} cell(s) differ")
    return diffs


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--extract-dir", required=True)
    ap.add_argument("--report")
    ap.add_argument("--preview")
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--root", default="")
    args = ap.parse_args()

    root = Path(sdk.resolve_922_root(args.root) or "")
    if not root.is_dir():
        raise SystemExit("could not resolve the 922 QTDR Production Packages root")
    quote_path = root / QUOTE_REL
    mpl_path = root / MPL_NAME
    today = _dt.date.today().isoformat()

    cat, tallies, batches_ok, problems = load_catalog(
        Path(args.extract_dir), today)
    cat.finalize_times_made()
    pricing = mp.load_pricing_map(quote_path)
    cat.apply_part_types(pricing)
    rows = cat.sorted_rows()

    type_counts = Counter(r["part_type"] for r in rows)
    type_sources = Counter(r["part_type_source"] for r in rows)
    report = {
        "generated": today,
        "batch_tallies": dict(tallies),
        "batches_ok": batches_ok,
        "problems": problems,
        "catalog_rows": len(rows),
        "distinct_ppns": len({r["ppn"].upper() for r in rows}),
        "rows_with_alt_suffixes": sum(1 for r in rows if r["alt_suffixes"]),
        "repeat_ppns": sum(1 for r in {r["ppn"].upper(): r["times_made"]
                                       for r in rows}.values() if r > 1),
        "total_pieces": sum(r["total_pieces"] for r in rows),
        "part_type_counts": dict(type_counts.most_common()),
        "part_type_sources": dict(type_sources.most_common()),
        "merge_warnings": cat.warnings[:200],
        "merge_warning_count": len(cat.warnings),
    }

    if args.preview:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.worksheets[0])
        mp.write_master_sheet(wb, rows, after_sheet="")
        wb.save(args.preview)
        print(f"preview -> {args.preview}")

    if args.write:
        lock = mpl_path.parent / f"~${mpl_path.name}"
        if lock.exists():
            raise SystemExit(f"'{MPL_NAME}' appears open in Excel "
                             f"({lock.name} present) - close it and rerun.")
        import os
        backup_dir = Path(os.path.expandvars(
            r"%LOCALAPPDATA%\TechDeck\mpl_master_build\backups"))
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = backup_dir / f"922 MPL.pre-master.{stamp}.xlsx"
        sdk.copy_resilient(mpl_path, backup)
        print(f"backup -> {backup}")

        wb = sdk.load_workbook_resilient(mpl_path)
        pre_sheets = [s for s in wb.sheetnames
                      if s not in (mp.MASTER_SHEET_NAME, mp.ANALYSIS_SHEET_NAME)]
        mp.write_master_sheet(wb, rows, after_sheet="PO 321+")
        mp.write_analysis_sheet(wb)
        wb.save(mpl_path)
        print(f"wrote '{mp.MASTER_SHEET_NAME}' ({len(rows)} rows) + "
              f"'{mp.ANALYSIS_SHEET_NAME}' -> {mpl_path}")

        diffs = verify_untouched_sheets(backup, mpl_path, pre_sheets)
        report["post_write_sheet_diffs"] = diffs
        if diffs:
            print("!! PRE-EXISTING SHEETS CHANGED - restore from backup:")
            for d in diffs:
                print("   " + d)
        else:
            print(f"verified: {len(pre_sheets)} pre-existing sheet(s) "
                  f"value-identical after save")

    if args.report:
        Path(args.report).write_text(json.dumps(report, indent=1),
                                     encoding="utf-8")
    print(json.dumps({k: report[k] for k in (
        "batch_tallies", "catalog_rows", "distinct_ppns", "repeat_ppns",
        "rows_with_alt_suffixes", "total_pieces", "part_type_counts",
        "part_type_sources", "merge_warning_count")}, indent=1))


if __name__ == "__main__":
    main()
