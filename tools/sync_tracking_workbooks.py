"""Bring the two tracking workbooks up to date.

    python tools/sync_tracking_workbooks.py           # report the gap only
    python tools/sync_tracking_workbooks.py --write   # apply it

Both live in ..\\Other Documents next to the repo:
    TECH_PROCESS_IMPROVEMENT.xlsm     internal build log (TASK | STATE | DESC)
    TechDeck Version Controller.xlsx  the presented record (6 sheets)

Why this is a script and not a hand edit: the workbooks kept drifting because
updating them was a manual step at the end of a long session, and manual steps
at the end of long sessions get skipped. The content lives in
`sync_workbook_content.py`, so keeping the record current is an edit to one
list, and running this is idempotent -- rows are matched on their key column,
so existing entries are updated in place and only new ones are appended.

Read the TONE RULE at the top of the content module before adding anything:
engagement features ARE logged, but the Version Controller is read outside the
team and never uses in-house names.
"""
from __future__ import annotations

import shutil
import sys
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))
import sync_workbook_content as C  # noqa: E402
import sync_workbook_tools as T  # noqa: E402

DOCS = Path(__file__).resolve().parents[2] / "Other Documents"
VC = DOCS / "TechDeck Version Controller.xlsx"
PI = DOCS / "TECH_PROCESS_IMPROVEMENT.xlsm"
TODAY = date.today().strftime("%b %d, %Y").replace(" 0", " ")
CURRENT_VERSION = "0.8.6.10"


# --------------------------------------------------------------------- helpers
def style_from(ws, src_row, dst_row, ncols):
    """Copy a template row's formatting to a new row so appends match."""
    for c in range(1, ncols + 1):
        s, d = ws.cell(row=src_row, column=c), ws.cell(row=dst_row, column=c)
        d.font, d.fill = copy(s.font), copy(s.fill)
        d.border, d.alignment = copy(s.border), copy(s.alignment)
        d.number_format = s.number_format
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def find_row(ws, key, col=1):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip().lower() == key.strip().lower():
            return r
    return None


def upsert(ws, rows, template_row, ncols, key_col=1, log=None):
    """Update rows whose key exists, append the rest. Returns (updated, added)."""
    updated = added = 0
    for values in rows:
        r = find_row(ws, str(values[key_col - 1]), key_col)
        if r is None:
            r = ws.max_row + 1
            style_from(ws, template_row, r, ncols)
            added += 1
            if log is not None:
                log.append(f"    + {values[key_col - 1]}")
        else:
            updated += 1
        for i, v in enumerate(values, start=1):
            if v is not None:
                ws.cell(row=r, column=i).value = v
    return updated, added


def test_count():
    """How many tests the suite actually has, asked of pytest rather than
    remembered. Returns 0 if collection fails, and the claim is left alone."""
    import re
    import subprocess
    try:
        out = subprocess.run(
            [sys.executable, "-m", "pytest", "--collect-only", "-q"],
            cwd=Path(__file__).resolve().parents[1],
            capture_output=True, text=True, timeout=180).stdout
    except (OSError, subprocess.SubprocessError):
        return 0
    return sum(int(m.group(1))
               for m in re.finditer(r"^\S+:\s*(\d+)\s*$", out, re.MULTILINE))


def open_workbook(path, **kw):
    """Load, or explain the two ways that fails on a OneDrive-synced file.

    Excel holds a deny-read lock, so even the dry run dies on PermissionError
    with a traceback that says nothing about the actual cause. (Same class as
    Hard Rule 13's locked-file case -- the content is on disk, another process
    simply will not share it.)
    """
    try:
        return openpyxl.load_workbook(path, **kw)
    except PermissionError:
        raise SystemExit(
            f"\n  {path.name} is open in Excel.\n"
            f"  Close it and run this again -- Excel will not share the file,\n"
            f"  so nothing can be read from it or written to it meanwhile.\n")
    except OSError as exc:
        raise SystemExit(
            f"\n  {path.name} could not be read ({exc}).\n"
            f"  If OneDrive shows it as cloud-only, open the folder in\n"
            f"  Explorer to download it, then run this again.\n")


def set_pair(ws, label, value):
    r = find_row(ws, label)
    if r:
        ws.cell(row=r, column=2).value = value
    return bool(r)


# ------------------------------------------------------------------ styling
# One visual vocabulary shared by every sheet. The sheets had drifted into
# looking like five separate documents -- different header colours, some sheets
# banded and some not, columns too narrow to show a wrapped description.
TITLE_FILL = PatternFill("solid", fgColor="FF1F3864")
HEAD_FILL = PatternFill("solid", fgColor="FFBDD7EE")
SECT_FILL = PatternFill("solid", fgColor="FFD9E1F2")
BAND_FILL = PatternFill("solid", fgColor="FFF2F6FB")
EDGE = Side(style="thin", color="FF9DB2CE")
BOX = Border(left=EDGE, right=EDGE, top=EDGE, bottom=EDGE)
WRAP = Alignment(wrap_text=True, vertical="top")
MID = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _set(cell, value):
    """Assign unless this is the tail of a merged range (read-only there)."""
    if type(cell).__name__ != "MergedCell":
        cell.value = value


def _title(ws, row, ncols, text):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        _set(cell, text if c == 1 else None)
        cell.fill = TITLE_FILL
        cell.font = Font(bold=True, size=12, color="FFFFFFFF")
        cell.border = BOX
        cell.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 26


def _header(ws, row, values):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c)
        _set(cell, v)
        cell.fill = HEAD_FILL
        cell.font = Font(bold=True, size=10)
        cell.border = BOX
        cell.alignment = MID
    ws.row_dimensions[row].height = 24


def _section(ws, row, ncols, text):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        _set(cell, text if c == 1 else None)
        cell.fill = SECT_FILL
        cell.font = Font(bold=True, size=10, color="FF1F3864")
        cell.border = BOX
        cell.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def _data(ws, row, values, band=False):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c)
        _set(cell, v)
        cell.font = Font(size=10)
        cell.border = BOX
        cell.alignment = WRAP
        cell.fill = BAND_FILL if band else PatternFill()
    ws.row_dimensions[row].height = None      # let Excel autofit the wrap


def _widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _view(ws, freeze=None):
    """Set the freeze and CLEAR the saved scroll position.

    A sheet stores where it was last scrolled to as `topLeftCell` on the view,
    and openpyxl carries that through a rewrite untouched. Combined with a
    frozen pane the two disagree -- the view says 'open at row 17', the pane
    says 'the scrollable region starts at row 3' -- and Excel obeys the view,
    so the sheet opens stranded at the bottom with the rows above it
    unreachable. Three sheets shipped like that.

    Setting it to None means 'no opinion', and Excel opens at the top.
    """
    ws.freeze_panes = freeze
    ws.sheet_view.topLeftCell = None
    # Gridlines off everywhere: every cell now carries its own border, and the
    # sheets disagreed about this before (one had them off, five had them on).
    ws.sheet_view.showGridLines = False
    for sel in ws.sheet_view.selection or []:
        sel.activeCell = sel.sqref = freeze or "A1"


def rebuild_tools(wb, notes):
    """AUTOMATION TOOLS is REBUILT, not upserted.

    The problem here was never missing rows -- it was wrong ones. Five tools sat
    under the wrong workflow heading (both 911 extractors and the QR generator
    filed under 922; the 922 runtime tool filed under 911), which is why the 911
    section read as thinner than the work behind it, and four names were stale
    after renames. An upsert can neither move a row between sections nor rename
    one in place; it just adds a second copy alongside the wrong one.
    """
    ws = wb["AUTOMATION TOOLS"]
    ws.delete_rows(1, ws.max_row + 1)
    _widths(ws, {"A": 26, "B": 52, "C": 68, "D": 10, "E": 11})
    _title(ws, 1, 5, "TECHDECK  --  DEPLOYED AUTOMATION TOOLS")
    r, n = 2, 0
    for section, tools in T.TOOL_SECTIONS:
        _section(ws, r, 5, section)
        r += 1
        _header(ws, r, ["TOOL NAME", "WHAT IT DOES", "KEY AUTOMATION",
                        "STATUS", "SINCE VER."])
        r += 1
        for i, row in enumerate(tools):
            _data(ws, r, list(row), band=(i % 2 == 1))
            r += 1
            n += 1
        r += 1                                  # breathing room between sections
    _view(ws, "A2")
    notes.append(f"  AUTOMATION TOOLS           REBUILT -- {n} tools across "
                 f"{len(T.TOOL_SECTIONS)} workflow sections")
    return n


def rebuild_roadmap(wb, notes):
    """ROADMAP is rebuilt too, so a shipped item cannot sit at Planned forever.

    Delivered items STAY on the sheet rather than being deleted -- the sheet's
    job is to show the roadmap being worked through, and a sheet that only ever
    shows what is outstanding hides all of that.
    """
    ws = wb["ROADMAP"]
    ws.delete_rows(1, ws.max_row + 1)
    _widths(ws, {"A": 11, "B": 34, "C": 78, "D": 12, "E": 12})
    _title(ws, 1, 5, "TECHDECK  --  DEVELOPMENT ROADMAP")
    _header(ws, 2, ["PRIORITY", "ITEM", "DESCRIPTION", "WORKFLOW", "PHASE"])
    phase = {"Delivered": 0, "Planned": 1, "Research": 2, "Backlog": 3}
    flow = {"911": 0, "922": 1, "902": 2, "Quality": 3, "Platform": 4}
    pri = {"High": 0, "Medium": 1, "Low": 2}
    rows = sorted(T.ROADMAP_ROWS,
                  key=lambda x: (phase.get(x[4], 9), flow.get(x[3], 9),
                                 pri.get(x[0], 9), x[1]))
    for i, row in enumerate(rows):
        _data(ws, 3 + i, list(row), band=(i % 2 == 1))
    _view(ws, "A3")
    done = sum(1 for x in rows if x[4] == "Delivered")
    notes.append(f"  ROADMAP                    REBUILT -- {len(rows)} items, "
                 f"{done} Delivered")


def restyle(wb, notes):
    """Give the upserted sheets the same look as the rebuilt ones."""
    specs = {
        "VERSION HISTORY": ({"A": 15, "B": 14, "C": 14, "D": 96, "E": 52}, 5, 2),
        "SYSTEM FEATURES": ({"A": 32, "B": 62, "C": 66, "D": 10}, 4, 2),
        "ENGINEERING & RELIABILITY": ({"A": 12, "B": 30, "C": 15, "D": 70,
                                       "E": 62, "F": 40}, 6, 3),
    }
    for sheet, (widths, ncols, head_row) in specs.items():
        ws = wb[sheet]
        _widths(ws, widths)
        _title(ws, 1, ncols, ws.cell(row=1, column=1).value)
        if head_row == 3:                       # sheet carries a subtitle line
            for c in range(1, ncols + 1):
                cell = ws.cell(row=2, column=c)
                cell.fill = SECT_FILL
                cell.font = Font(size=9, italic=True, color="FF1F3864")
                cell.border = BOX
                cell.alignment = Alignment(vertical="center", indent=1)
        _header(ws, head_row, [ws.cell(row=head_row, column=c).value
                               for c in range(1, ncols + 1)])
        band = False
        for r in range(head_row + 1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value in (None, ""):
                continue
            _data(ws, r, [ws.cell(row=r, column=c).value
                          for c in range(1, ncols + 1)], band=band)
            band = not band
        _view(ws, f"A{head_row + 1}")
        notes.append(f"  {sheet:<26} restyled")

    ov = wb["OVERVIEW"]
    _widths(ov, {"A": 24, "B": 80})
    _view(ov)
    _title(ov, 1, 2, ov.cell(row=1, column=1).value)
    for r in range(2, ov.max_row + 1):
        label = ov.cell(row=r, column=1).value
        if label in (None, ""):
            continue
        # A label with nothing beside it is a section heading, not a field.
        if ov.cell(row=r, column=2).value in (None, "") and r != ov.max_row:
            _section(ov, r, 2, label)
            continue
        ov.cell(row=r, column=1).font = Font(bold=True, size=10)
        ov.cell(row=r, column=2).font = Font(size=10)
        for c in (1, 2):
            ov.cell(row=r, column=c).border = BOX
            ov.cell(row=r, column=c).alignment = WRAP
    notes.append("  OVERVIEW                   restyled")


# ------------------------------------------------------------ the two updates
def sync_version_controller(write):
    wb = open_workbook(VC)
    notes = []

    ov = wb["OVERVIEW"]

    for sheet, rows, ncols in (
            ("VERSION HISTORY", C.VERSION_ROWS, 5),
            ("SYSTEM FEATURES", C.SYSTEM_FEATURES, 4),
            ("ENGINEERING & RELIABILITY", C.ENGINEERING, 6)):
        ws = wb[sheet]
        key = 2 if sheet == "ENGINEERING & RELIABILITY" else 1
        log = []
        u, a = upsert(ws, rows, ws.max_row, ncols, key_col=key, log=log)
        notes.append(f"  {sheet:<26} {u} updated, {a} added")
        notes += log

    # Prose written before the tone rule.
    for sheet, old, new in C.PROSE_FIXES:
        ws = wb[sheet]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str) and old in cell.value:
                    cell.value = cell.value.replace(old, new)
                    notes.append(f"  {sheet:<26} reworded: {old[:40]}...")

    # The test-count claim goes stale every time the suite changes, so COUNT
    # rather than hardcode -- the previous literal was already wrong by nine.
    n = test_count()
    r = find_row(wb["ENGINEERING & RELIABILITY"], "Automated Test Suite", 2)
    if r and n:
        wb["ENGINEERING & RELIABILITY"].cell(row=r, column=6).value = \
            f"{n} automated tests, run on every change."
        notes.append(f"  ENGINEERING                test count -> {n}")

    n_tools = rebuild_tools(wb, notes)
    rebuild_roadmap(wb, notes)
    restyle(wb, notes)

    set_pair(ov, "Current Version", CURRENT_VERSION)
    set_pair(ov, "Last Updated", TODAY)
    set_pair(ov, "Automation Tools",
             f"{n_tools} deployed across the 911, 922 and 902 production "
             f"workflows plus shared quality, estimating and shop tools")
    # The summary still claimed two workflows after 902 and quality shipped.
    set_pair(ov, "Supported Workflows",
             "911 QTDR Production Packages  |  922 QTDR Pallet Packages  |  "
             "902 QTDR Production Packages  |  Quality, estimating and shop-"
             "floor tooling")
    notes.append(f"  OVERVIEW                   version -> {CURRENT_VERSION}, "
                 f"tools -> {n_tools}, updated -> {TODAY}")

    if write:
        shutil.copy2(VC, VC.with_suffix(".xlsx.bak"))
        wb.save(VC)
    return notes


def sync_process_improvement(write):
    wb = open_workbook(PI, keep_vba=True)
    ws = wb["Sheet1"]
    notes, renamed = [], 0

    for old, (new, desc) in C.PI_RENAMES.items():
        r = find_row(ws, old)
        if r:
            ws.cell(row=r, column=1).value = new
            ws.cell(row=r, column=3).value = desc
            renamed += 1
            notes.append(f"    ~ {old}  ->  {new}")
    notes.insert(0, f"  Sheet1  {renamed} entries reframed")

    tests = test_count()
    rows = [tuple(v.replace("{tests}", str(tests)) if tests else v for v in row)
            for row in C.PI_NEW]
    log = []
    u, a = upsert(ws, rows, ws.max_row, 3, log=log)
    notes.append(f"  Sheet1  {u} updated, {a} added")
    notes += log

    if write:
        shutil.copy2(PI, PI.with_suffix(".xlsm.bak"))
        wb.save(PI)
    return notes


def main(write=False):
    print("VERSION CONTROLLER")
    for n in sync_version_controller(write):
        print(n)
    print("\nPROCESS IMPROVEMENT LOG")
    for n in sync_process_improvement(write):
        print(n)
    print("\n" + ("WRITTEN" if write else "DRY RUN - pass --write to apply"))


if __name__ == "__main__":
    main("--write" in sys.argv)
